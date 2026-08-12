#!/usr/bin/env python3
"""SEO/GEO brainstorm board -> Excel, post-its gegroepeerd op onderwerp (fases losgelaten)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

PURPLE = "5E4A90"
PURPLE_LT = "EFEBF6"
INK = "1F1B2E"

# post-it kleuren zoals op het bord
C = {
    "roze":   "FCCCF0",
    "rood":   "FC9C9C",
    "geel":   "FCE46C",
    "oranje": "FCB478",
    "blauw":  "8AB4F0",
    "groen":  "60D884",
}

# (thema, tekst, kleur, auteur, opmerking, board-kolom)
ROWS = [
    ("Contentformats & paginatypes", "Nieuwsartikel/Diepte-artikel", "roze", "", "", "Awareness"),
    ("Contentformats & paginatypes", "How-to (Uitleg + werking)", "roze", "", "", "tussen Awareness/Consideration"),
    ("Contentformats & paginatypes", "Listicle (Overzicht + keuzehulp)", "roze", "", "", "tussen Awareness/Consideration"),
    ("Contentformats & paginatypes", "FAQ - ook met oog op wie we zijn wat we doen", "rood", "", "", "tussen Awareness/Consideration"),
    ("Contentformats & paginatypes", "Thema paginas - back to school incl listicle", "rood", "", "", "tussen Awareness/Consideration"),
    ("Contentformats & paginatypes", "Situatie schets - PLP", "rood", "", "", "tussen Awareness/Consideration"),
    ("Contentformats & paginatypes", "Alles voor je vakantie naar ......", "rood", "", "", "tussen Awareness/Consideration"),
    ("Contentformats & paginatypes", "Vergelijking (X vs Y)", "roze", "", "", "Consideration"),
    ("Contentformats & paginatypes", "Top X pagina's; goedkoopste/beste/etc.", "oranje", "", "", "Consideration"),
    ("Contentformats & paginatypes", "Wat is de beste... pagina's op basis van data", "geel", "Peter Neef", "", "Consideration"),
    ("Contentformats & paginatypes", "Longread guides per maincat als kennisbron", "blauw", "", "los blok links op het bord", "los blok"),
    ("Contentformats & paginatypes", "Infographics", "geel", "Peter Neef", "", "Awareness"),
    ("Contentformats & paginatypes", "Doelgroep-guides, denk aan beste laptop voor studenten", "blauw", "", "", "Consideration"),
    ("Contentformats & paginatypes", "Sizing / fit / compatibility guides", "blauw", "", "", "Consideration"),
    ("Contentformats & paginatypes", "Alternatieven hulp", "rood", "", "", "tussen Consideration/Decision"),
    ("Contentformats & paginatypes", "Productreviews", "groen", "", "", "Consideration"),

    ("Technisch & LLM-vindbaarheid", "Schema markup toevoegen", "roze", "", "", "Awareness"),
    ("Technisch & LLM-vindbaarheid", "Javascript rendering afbeeldingen", "roze", "", "", "Awareness"),
    ("Technisch & LLM-vindbaarheid", "LLM txt extra - wie zijn we wat doen we", "rood", "", "los blok links op het bord", "los blok"),
    ("Technisch & LLM-vindbaarheid", "Begrippenlijst die geciteerd kunnen worden, parfum termen etc", "blauw", "", "los blok links op het bord", "los blok"),
    ("Technisch & LLM-vindbaarheid", "Prijsactualeit in LLM????", "rood", "", "spelling zoals op bord", "Decision"),

    ("Eigen data als contentmotor", 'Datagedreven trendrapporten — unieke statistieken uit eigen data ("gemiddelde airfryer-prijs daalde 12% in 2026)', "groen", "", "TOP!-sticker", "Awareness"),
    ("Eigen data als contentmotor", "Proberen het nieuws te halen met onze data", "geel", "Peter Neef", "", "Awareness"),
    ("Eigen data als contentmotor", "Contentkalender met producten met hoge intentie nieuws/interessante feiten", "geel", "Peter Neef", "", "Awareness"),
    ("Eigen data als contentmotor", "Contentkalender gebruiken voor onderzoeken o.b.v. eigen data", "oranje", "", "", "Awareness"),

    ("Keuzehulp & onderbouwing op de site", "Vergelijkmodule?", "rood", "", "", "Consideration"),
    ("Keuzehulp & onderbouwing op de site", "Meer relevante content op PLP", "rood", "", "", "Consideration"),
    ("Keuzehulp & onderbouwing op de site", "Beslist keuze inclusief onderbouwing", "geel", "Peter Neef", "", "Consideration"),
    ("Keuzehulp & onderbouwing op de site", "Eigen score voor shops", "rood", "", "", "Consideration"),
    ("Keuzehulp & onderbouwing op de site", "Shops uitlichter - waarom hier kopen", "rood", "", "", "Consideration"),
    ("Keuzehulp & onderbouwing op de site", "Via beslist nog goedkoper dan direct via het eerste offer", "blauw", "", "", "Consideration"),
    ("Keuzehulp & onderbouwing op de site", "beste deal badge + uitleg (prijs, garantie, levertijd, retourbeleid) wordt geciteerd", "blauw", "", "", "Decision"),
    ("Keuzehulp & onderbouwing op de site", "Gebruikers baseren hun keuze op.....", "rood", "", "", "Decision"),
    ("Keuzehulp & onderbouwing op de site", "Heel duidelijk uitdragen wat beslist doet en wat je allemaal kan vergelijken", "geel", "Peter Neef", "", "Decision"),
    ("Keuzehulp & onderbouwing op de site", "'Keuze van beslist' banner bij beste keuze op browse", "oranje", "", "", "Trust Validation"),

    ("Autoriteit, auteurs & reviews", "Auteurpagina's", "roze", "", "", "Trust Validation"),
    ("Autoriteit, auteurs & reviews", "Beslist medewerkers als auteurs/reviewers", "oranje", "", "", "Trust Validation"),
    ("Autoriteit, auteurs & reviews", "Trustpilot-reputatie", "roze", "", "", "Trust Validation"),
    ("Autoriteit, auteurs & reviews", "Trustpilot reviews", "rood", "", "", "tussen Decision/Trust Validation"),
    ("Autoriteit, auteurs & reviews", "Mogelijkheid tot schrijven review op PLP", "oranje", "", "", "Trust Validation"),
    ("Autoriteit, auteurs & reviews", "(Nog meer) Review uitnodigingen", "oranje", "", "", "Trust Validation"),

    ("Off-site, links & social", 'linkbuilding dmv "beslist topshop" badge, shops weer activeren. Bron voor onafhankelijke shop beoordelingen', "groen", "", "", "Trust Validation"),
    ("Off-site, links & social", "Klasseshops (link building)", "oranje", "", "", "Trust Validation"),
    ("Off-site, links & social", "Reddit Community management", "roze", "", "", "Trust Validation"),
    ("Off-site, links & social", "Reddit comments", "geel", "Peter Neef", "", "Trust Validation"),
    ("Off-site, links & social", "Veel vermeldingen op social media (linkedIn) --> Ziet Google/LLm allemaal als linkjes", "geel", "Peter Neef", "", "Trust Validation"),
    ("Off-site, links & social", "Influencer marketing", "rood", "", "", "Trust Validation"),

    ("Loyaliteit & accounts", "Loyaliteitsfeatures -> beloningen, verlanglijstjes, etc. etc.", "groen", "", "", "Trust Validation"),
    ("Loyaliteit & accounts", "Accounts voor gebruikers", "rood", "", "", "Trust Validation"),
    ("Loyaliteit & accounts", "Deel en win", "rood", "", "", "Trust Validation"),
]

HEADERS = ["Thema", "Post-it", "Kleur", "Auteur", "Opmerking", "Board-kolom"]
WIDTHS = [30, 74, 11, 13, 26, 31]

thin = Side(style="thin", color="D8D2E6")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ---------------------------------------------------------------- Onderwerpen
ws = wb.active
ws.title = "Onderwerpen"

ws["A1"] = "SEO/GEO content brainstorm — onderwerpen"
ws["A1"].font = Font(name="Calibri", size=15, bold=True, color=PURPLE)
ws["A2"] = ("Bron: SEO_GEO brainstorm board. Gegroepeerd op onderwerp — de fases "
            "(Awareness / Consideration / Decision / Trust Validation) zijn losgelaten "
            "en staan alleen nog als referentie in de laatste kolom.")
ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=INK)
ws.merge_cells("A2:F2")
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[1].height = 24
ws.row_dimensions[2].height = 30

HDR = 4
for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(row=HDR, column=i, value=h)
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=PURPLE)
    c.alignment = Alignment(vertical="center", horizontal="left")
    c.border = box
ws.row_dimensions[HDR].height = 22

prev_theme, band = None, False
for r, (thema, tekst, kleur, auteur, opm, kolom) in enumerate(ROWS, start=HDR + 1):
    if thema != prev_theme:
        band = not band
        prev_theme = thema
    tint = PURPLE_LT if band else "FFFFFF"

    vals = [thema, tekst, kleur, auteur, opm, kolom]
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.border = box
        c.alignment = Alignment(wrap_text=(i == 2), vertical="center")
        c.fill = PatternFill("solid", fgColor=tint)
        c.font = Font(name="Calibri", size=11, color=INK)

    ws.cell(row=r, column=1).font = Font(name="Calibri", size=11, bold=True, color=PURPLE)
    # kleurcel krijgt de echte post-it kleur
    kc = ws.cell(row=r, column=3)
    kc.fill = PatternFill("solid", fgColor=C[kleur])
    kc.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=6).font = Font(name="Calibri", size=10, color="6B6480")
    ws.row_dimensions[r].height = 30 if len(tekst) > 60 else 20

last = HDR + len(ROWS)
for i, w in enumerate(WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = f"A{HDR + 1}"
ws.auto_filter.ref = f"A{HDR}:F{last}"

# ---------------------------------------------------------------- Samenvatting
s = wb.create_sheet("Samenvatting")
s["A1"] = "Aantal post-its per thema"
s["A1"].font = Font(name="Calibri", size=13, bold=True, color=PURPLE)

order, seen = [], set()
for t, *_ in ROWS:
    if t not in seen:
        order.append(t); seen.add(t)

for i, h in enumerate(["Thema", "Post-its"], start=1):
    c = s.cell(row=3, column=i, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=PURPLE)
    c.border = box
for r, t in enumerate(order, start=4):
    n = sum(1 for row in ROWS if row[0] == t)
    s.cell(row=r, column=1, value=t).border = box
    c = s.cell(row=r, column=2, value=n)
    c.border = box
    c.alignment = Alignment(horizontal="center")
tot = s.cell(row=4 + len(order), column=1, value="Totaal")
tot.font = Font(bold=True, color=PURPLE)
tot.border = box
c = s.cell(row=4 + len(order), column=2, value=len(ROWS))
c.font = Font(bold=True, color=PURPLE)
c.border = box
c.alignment = Alignment(horizontal="center")

base = 4 + len(order) + 2
s.cell(row=base, column=1, value="Aantal post-its per board-kolom").font = Font(
    name="Calibri", size=13, bold=True, color=PURPLE)
for i, h in enumerate(["Board-kolom", "Post-its"], start=1):
    c = s.cell(row=base + 2, column=i, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=PURPLE)
    c.border = box
kols, seenk = [], set()
for row in ROWS:
    if row[5] not in seenk:
        kols.append(row[5]); seenk.add(row[5])
for r, k in enumerate(kols, start=base + 3):
    s.cell(row=r, column=1, value=k).border = box
    c = s.cell(row=r, column=2, value=sum(1 for row in ROWS if row[5] == k))
    c.border = box
    c.alignment = Alignment(horizontal="center")

s.column_dimensions["A"].width = 36
s.column_dimensions["B"].width = 11

OUT = "/mnt/c/Users/JoepvanSchagen/Downloads/claude/SEO_GEO_brainstorm_onderwerpen.xlsx"
wb.save(OUT)
print("saved:", OUT)
print("post-its:", len(ROWS), "| themas:", len(order), "| board-kolommen:", len(kols))
