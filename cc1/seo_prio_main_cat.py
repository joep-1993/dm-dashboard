#!/usr/bin/env python3
"""
SEO-priority analysis for a single Beslist main category.

Reusable counterpart of the dashboard's SEO Priority tool, scoped to one
main category, with judgment-based suggestions (no fixed % thresholds).
Reads from Redshift + taxv2 only — never writes to the API.

Usage:
    python3 seo_prio_main_cat.py <main_cat_taxv2_id> <main_cat_name> [out.xlsx]

    # Examples (taxv2 IDs from `GET /api/Categories?locale=nl-NL`):
    python3 seo_prio_main_cat.py 30000 "Horloges"
    python3 seo_prio_main_cat.py 32000 "Schoenen"
    python3 seo_prio_main_cat.py 700   "Films & Series"

Output: an xlsx with one row per (category, facet) combo currently effective-
ON, plus any currently-off combos that pull material traffic. Rows are colour-
coded by suggestion (turn_off / review / keep_on / turn_on). The Reason column
explains the call for every row.

Pipeline:
  1. Walk the main-category sub-tree via /api/Categories/{id}.
  2. For each category, pull its facets + effective seoPriority via
     /api/Categories/{id}?includeFacets=true (inheritance already resolved).
  3. Pull 2y of /c/ visits + revenue from Redshift, scoped to main_cat_name.
  4. Build a legacy_id → taxv2_id lookup from each category's urlSlug suffix
     (URLs use legacy big IDs, taxv2 uses small IDs — they only join via the
     numeric suffix on urlSlug).
  5. Parse each URL into (taxv2_cat_id, facet_slug) and fan visits/revenue out
     across the facets the URL touches.
  6. Apply judge() per row + write Excel.

Notes:
  - Date range: last 2 years (matches the dashboard's SEO Priority tool default).
  - .env loaded from dm-tools/.env (REDSHIFT_HOST/USER/PASSWORD/etc.).
  - Audit header X-User-Name: SEO_JOEP on every taxv2 call (read-only here, but
    routine).
"""
import os
import re
import sys
from collections import defaultdict
from datetime import datetime, timedelta

import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv

load_dotenv('/home/joepvanschagen/projects/dm-tools/.env')
import psycopg2
import psycopg2.extras

TAXV2 = "http://producttaxonomyunifiedapi-prod.azure.api.beslist.nl"
H = {"X-User-Name": "SEO_JOEP", "Accept": "application/json"}

if len(sys.argv) < 3:
    print(__doc__)
    sys.exit(1)

MAIN_CAT_ID = int(sys.argv[1])
MAIN_CAT_NAME = sys.argv[2]
default_out = f"/mnt/c/Users/JoepvanSchagen/Downloads/claude/{MAIN_CAT_NAME.lower().replace(' ', '_').replace('&','and')}_seo_prio.xlsx"
OUT_XLSX = sys.argv[3] if len(sys.argv) > 3 else default_out

session = requests.Session()
session.headers.update(H)


def nm_nl(node):
    for l in (node.get("labels") or []):
        if l.get("locale") == "nl-NL":
            return l.get("name")
    return node.get("name", "") or ""


def url_slug_nl(node):
    for l in (node.get("labels") or []):
        if l.get("locale") == "nl-NL":
            return l.get("urlSlug")
    return None


# Legacy IDs are encoded as the trailing numeric suffix on urlSlug
# (e.g. "horloge_649387" → legacy id 649387). The root category's urlSlug
# has no numeric suffix (e.g. "horloge"); URLs without a subcat segment
# (`/products/horloge/c/...`) are matched via slug_to_v2 instead.
_LEGACY_RE = re.compile(r"_(\d+)$")


def walk_subtree(cid, out, depth=0, max_depth=10):
    r = session.get(
        f"{TAXV2}/api/Categories/{cid}?includeFacets=false&includeSubCategories=true",
        timeout=20,
    )
    if r.status_code != 200:
        return
    node = r.json()
    slug = url_slug_nl(node) or ""
    m = _LEGACY_RE.search(slug)
    out.append({
        "id":        cid,
        "name":      nm_nl(node),
        "slug":      slug,
        "legacy_id": m.group(1) if m else None,
        "depth":     depth,
        "parent":    node.get("parentId"),
    })
    if depth >= max_depth:
        return
    for sub in (node.get("subCategories") or []):
        walk_subtree(sub["id"], out, depth + 1, max_depth)


# ── 1. Walk the main-category sub-tree ───────────────────────────────────
print(f"=== 1/4 walk sub-tree for {MAIN_CAT_NAME} (root id={MAIN_CAT_ID}) ===", flush=True)
cats = []
walk_subtree(MAIN_CAT_ID, cats)
print(f"  {len(cats)} categories", flush=True)

legacy_to_v2 = {c["legacy_id"]: c["id"] for c in cats if c["legacy_id"]}
slug_to_v2 = {c["slug"]: c["id"] for c in cats if c["slug"] and not c["legacy_id"]}
cat_name_by_id = {c["id"]: c["name"] for c in cats}
print(f"  legacy_id mappings: {len(legacy_to_v2)} ; root-slug mappings: {len(slug_to_v2)}", flush=True)


# ── 2. Per-category facets + effective seoPriority ───────────────────────
print(f"=== 2/4 fetch facets + effective seoPriority ===", flush=True)
combos = []
for i, c in enumerate(cats):
    r = session.get(
        f"{TAXV2}/api/Categories/{c['id']}?includeFacets=true&includeSubCategories=false",
        timeout=20,
    )
    if r.status_code != 200:
        continue
    for f in (r.json().get("facets") or []):
        nl = next((l for l in (f.get("labels") or []) if l.get("locale") == "nl-NL"), {})
        combos.append({
            "cat_id":      c["id"],
            "cat_name":    c["name"],
            "depth":       c["depth"],
            "facet_id":    f.get("facetId"),
            "facet_slug":  (nl.get("urlSlug") or "").lower(),
            "facet_name":  nl.get("name") or "",
            "seoPriority": f.get("seoPriority"),
            "inh":         f.get("inheritanceStatus"),
        })

on_count = sum(1 for c in combos if c["seoPriority"] is True)
print(f"  total (cat,facet) combos: {len(combos)}   currently effective-true: {on_count}", flush=True)


# ── 3. Redshift /c/ traffic for this main_cat ────────────────────────────
print(f"=== 3/4 pull 2y Redshift /c/ visits+revenue for main_cat='{MAIN_CAT_NAME}' ===", flush=True)
end = datetime.utcnow().date() - timedelta(days=1)
start = end - timedelta(days=365 * 2)
start_key = int(start.strftime("%Y%m%d"))
end_key = int(end.strftime("%Y%m%d"))
print(f"  date range: {start_key} .. {end_key}", flush=True)

conn = psycopg2.connect(
    host=os.getenv("REDSHIFT_HOST"),
    port=os.getenv("REDSHIFT_PORT", "5439"),
    dbname=os.getenv("REDSHIFT_DB"),
    user=os.getenv("REDSHIFT_USER"),
    password=os.getenv("REDSHIFT_PASSWORD"),
)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cur.execute(
    """
    SELECT
        dv.main_cat_name,
        dv.deepest_subcat_name,
        SPLIT_PART(dv.url, '?', 1) AS url,
        COUNT(*) AS visits,
        COALESCE(SUM(fcv.cpc_revenue), 0) + COALESCE(SUM(fcv.ww_revenue), 0) AS revenue
    FROM datamart.fct_visits fcv
    JOIN datamart.dim_visit dv ON fcv.dim_visit_key = dv.dim_visit_key
    JOIN datamart.dim_date dat ON fcv.dim_date_key = dat.dim_date_key
    JOIN chan_deriv.ref_channel_derivation_stats chan
      ON dv.aff_id = chan.aff_id AND dv.channel_id = chan.channel_id
    WHERE dv.is_real_visit = 1
      AND fcv.dim_date_key BETWEEN %s AND %s
      AND dv.url LIKE '%%beslist.nl%%'
      AND dv.url NOT LIKE '%%/r/%%'
      AND dv.url NOT LIKE '%%/p/%%'
      AND dv.url     LIKE '%%/c/%%'
      AND dv.url NOT LIKE '%%/l/%%'
      AND dv.url NOT LIKE '%%/page_%%'
      AND dv.url NOT LIKE '%%#%%'
      AND dv.main_cat_name = %s
      AND dv.deepest_subcat_name IS NOT NULL
    GROUP BY 1, 2, 3
    """,
    (start_key, end_key, MAIN_CAT_NAME),
)
rs_rows = cur.fetchall()
cur.close()
conn.close()
print(f"  {len(rs_rows)} URL rows", flush=True)


# ── 4. Parse URLs + fan visits/revenue across facets ─────────────────────
_SUBCAT_ID_RE = re.compile(r"_(\d+)(?=_|$)")


def parse_url(url):
    """Returns (taxv2_cat_id:int, [(facet_slug, facet_value_id), ...]) or None."""
    try:
        path = url.split("beslist.nl", 1)[1] if "beslist.nl" in url else url
    except Exception:
        return None
    if "/c/" not in path:
        return None
    head, _, facet_part = path.partition("/c/")
    if not facet_part:
        return None
    parts = [p for p in head.split("/") if p]
    if len(parts) < 2:
        return None
    v2_id = None
    if len(parts) >= 3:
        ids = _SUBCAT_ID_RE.findall(parts[-1])
        if ids:
            v2_id = legacy_to_v2.get(ids[-1])
    if v2_id is None:
        # Root-level URL like /products/horloge/c/...
        root_slug = parts[1]
        v2_id = slug_to_v2.get(root_slug)
    if v2_id is None:
        return None
    facets = []
    for chunk in facet_part.split("~~"):
        chunk = chunk.strip("/")
        if not chunk:
            continue
        slug, _, val = chunk.partition("~")
        if slug and val:
            facets.append((slug, val.split("/")[0]))
    if not facets:
        return None
    return v2_id, facets


agg = {}          # (cat_id, slug) -> {visits, revenue, urls, example}
cat_totals = {}   # cat_id        -> {visits, revenue, urls, name}
for r in rs_rows:
    parsed = parse_url(r["url"])
    if not parsed:
        continue
    cat_id, facets = parsed
    visits, revenue = int(r["visits"] or 0), float(r["revenue"] or 0)
    ct = cat_totals.setdefault(cat_id, {"visits": 0, "revenue": 0.0, "urls": 0,
                                        "name": r["deepest_subcat_name"]})
    ct["visits"] += visits
    ct["revenue"] += revenue
    ct["urls"] += 1
    seen = set()
    for slug, _vid in facets:
        slug_l = slug.lower()
        if slug_l in seen:
            continue
        seen.add(slug_l)
        a = agg.setdefault((cat_id, slug_l),
                           {"visits": 0, "revenue": 0.0, "urls": 0, "example": r["url"]})
        a["visits"] += visits
        a["revenue"] += revenue
        a["urls"] += 1

print(f"  parsed {len(agg)} distinct (cat,facet) traffic rows", flush=True)


# ── 5. Judgment + assemble rows ──────────────────────────────────────────
def judge(visits, revenue, pct_v, pct_r, url_count, currently_on):
    """
    Common-sense suggestion. Returns (suggestion, confidence, reason).

    Anchors (over 2 years of traffic):
      - "near zero" = <50 visits AND <€1 revenue
      - "tiny"      = <0.3% visits AND <0.3% revenue AND <500 visits AND <€20
      - "material"  = ≥1% visits OR ≥1% revenue OR ≥1000 visits OR ≥€100
    """
    near_zero = visits < 50 and revenue < 1
    tiny = pct_v < 0.3 and pct_r < 0.3 and visits < 500 and revenue < 20
    material = pct_v >= 1.0 or pct_r >= 1.0 or visits >= 1000 or revenue >= 100

    if currently_on:
        if near_zero:
            return ("turn_off", "high",
                    f"near-zero traffic ({visits} visits, €{revenue:.2f}) over 2y — crawl budget without payoff.")
        if tiny:
            return ("turn_off", "medium",
                    f"only {pct_v:.2f}% visits / {pct_r:.2f}% revenue and {visits} absolute visits — likely not worth crawling.")
        if material:
            return ("keep_on", "high",
                    f"{pct_v:.2f}% visits / {pct_r:.2f}% revenue, {visits:,} visits across {url_count} URLs.")
        return ("review", "manual",
                f"borderline: {pct_v:.2f}% visits / {pct_r:.2f}% revenue, {visits} absolute visits over 2y.")
    # currently OFF / inherit / null
    if material and visits >= 500:
        return ("turn_on", "high",
                f"currently NOT prioritized but pulls {pct_v:.2f}% / {pct_r:.2f}% revenue ({visits:,} visits) — worth indexing.")
    if visits >= 200 and pct_v >= 0.5:
        return ("turn_on", "medium",
                f"currently off but {pct_v:.2f}% visits / {pct_r:.2f}% revenue, {visits} visits — consider on.")
    return ("keep_off", "low", "no meaningful traffic.")


rows_out = []
seen_keys = set()

# (a) Every currently-effective-true combo, regardless of traffic
for c in (x for x in combos if x["seoPriority"] is True):
    key = (c["cat_id"], c["facet_slug"])
    seen_keys.add(key)
    a = agg.get(key, {"visits": 0, "revenue": 0.0, "urls": 0, "example": ""})
    ct = cat_totals.get(c["cat_id"], {"visits": 0, "revenue": 0.0})
    v_tot, r_tot = ct["visits"], ct["revenue"]
    pct_v = (a["visits"] / v_tot * 100) if v_tot else 0.0
    pct_r = (a["revenue"] / r_tot * 100) if r_tot else 0.0
    sug, conf, reason = judge(a["visits"], a["revenue"], pct_v, pct_r, a["urls"], True)
    rows_out.append({
        "main_cat":          MAIN_CAT_NAME,
        "category":          c["cat_name"],
        "cat_id":            c["cat_id"],
        "facet_name":        c["facet_name"],
        "facet_slug":        c["facet_slug"],
        "facet_id":          c["facet_id"],
        "inheritance":       c["inh"],
        "current":           "TRUE",
        "visits_2y":         a["visits"],
        "revenue_2y":        round(a["revenue"], 2),
        "url_count":         a["urls"],
        "pct_visits_in_cat": round(pct_v, 3),
        "pct_revenue_in_cat":round(pct_r, 3),
        "suggestion":        sug,
        "confidence":        conf,
        "reason":            reason,
        "example_url":       a.get("example", "")[:200],
    })

# (b) Currently NOT-true combos with material traffic → turn_on candidates
combo_lookup = {(c["cat_id"], c["facet_slug"]): c for c in combos}
TURN_ON_MIN_VISITS = 200
for (cat_id, slug), a in agg.items():
    if (cat_id, slug) in seen_keys:
        continue
    if a["visits"] < TURN_ON_MIN_VISITS:
        continue
    cmeta = combo_lookup.get((cat_id, slug))
    if cmeta:
        cat_name = cmeta["cat_name"]
        facet_name = cmeta["facet_name"]
        facet_id = cmeta["facet_id"]
        current_seo = cmeta["seoPriority"]
        inh = cmeta["inh"]
    else:
        cat_name = cat_name_by_id.get(cat_id) or cat_totals.get(cat_id, {}).get("name") or f"(cat {cat_id})"
        facet_name = ""
        facet_id = None
        current_seo = None
        inh = ""
    cur_label = ("TRUE" if current_seo is True else
                 "FALSE" if current_seo is False else
                 "inherit/none")
    ct = cat_totals.get(cat_id, {"visits": 0, "revenue": 0.0})
    pct_v = (a["visits"] / ct["visits"] * 100) if ct["visits"] else 0.0
    pct_r = (a["revenue"] / ct["revenue"] * 100) if ct["revenue"] else 0.0
    sug, conf, reason = judge(a["visits"], a["revenue"], pct_v, pct_r, a["urls"], False)
    if sug == "turn_on":
        rows_out.append({
            "main_cat":          MAIN_CAT_NAME,
            "category":          cat_name,
            "cat_id":            cat_id,
            "facet_name":        facet_name,
            "facet_slug":        slug,
            "facet_id":          facet_id,
            "inheritance":       inh,
            "current":           cur_label,
            "visits_2y":         a["visits"],
            "revenue_2y":        round(a["revenue"], 2),
            "url_count":         a["urls"],
            "pct_visits_in_cat": round(pct_v, 3),
            "pct_revenue_in_cat":round(pct_r, 3),
            "suggestion":        sug,
            "confidence":        conf,
            "reason":            reason,
            "example_url":       a.get("example", "")[:200],
        })

order = {"turn_off": 0, "review": 1, "turn_on": 2, "keep_on": 3, "keep_off": 4}
conf_order = {"high": 0, "medium": 1, "low": 2, "manual": 1}
rows_out.sort(key=lambda r: (order.get(r["suggestion"], 9),
                             conf_order.get(r["confidence"], 9),
                             -r["visits_2y"]))


# ── 6. Excel output ──────────────────────────────────────────────────────
print(f"=== 4/4 write {OUT_XLSX} ===", flush=True)
wb = openpyxl.Workbook()
ws = wb.active
ws.title = (MAIN_CAT_NAME[:25] + "_seo_prio").lower().replace(" ", "_")
columns = [
    ("main_cat",           "Main category"),
    ("category",           "Category"),
    ("cat_id",             "Cat ID"),
    ("facet_name",         "Facet"),
    ("facet_slug",         "Facet slug"),
    ("facet_id",           "Facet ID"),
    ("inheritance",        "Inheritance"),
    ("current",            "Current seoPriority"),
    ("suggestion",         "Suggestion"),
    ("confidence",         "Confidence"),
    ("visits_2y",          "Visits (2y)"),
    ("revenue_2y",         "Revenue (2y, €)"),
    ("url_count",          "URLs"),
    ("pct_visits_in_cat",  "% visits in cat"),
    ("pct_revenue_in_cat", "% revenue in cat"),
    ("reason",             "Reason"),
    ("example_url",        "Example URL"),
]
ws.append([label for _, label in columns])
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")

fill_by_sug = {
    "turn_off": PatternFill("solid", fgColor="F8D7DA"),
    "review":   PatternFill("solid", fgColor="FFF3CD"),
    "turn_on":  PatternFill("solid", fgColor="D1ECF1"),
    "keep_on":  PatternFill("solid", fgColor="D4EDDA"),
}
for r in rows_out:
    ws.append([r[c] for c, _ in columns])
    fill = fill_by_sug.get(r["suggestion"])
    if fill:
        for cell in ws[ws.max_row]:
            cell.fill = fill

widths = [14, 28, 8, 22, 22, 8, 12, 18, 12, 12, 12, 14, 8, 14, 16, 60, 60]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A2"
wb.save(OUT_XLSX)

summary = defaultdict(int)
for r in rows_out:
    summary[r["suggestion"]] += 1
print(f"\nwrote {len(rows_out)} rows to {OUT_XLSX}", flush=True)
for k in ("turn_off", "review", "keep_on", "turn_on", "keep_off"):
    print(f"  {k}: {summary.get(k, 0)}", flush=True)
