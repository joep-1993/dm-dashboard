#!/usr/bin/env python3
"""
Companion to pagetitles_blueprint_from_urls.py.

Takes the SEO-traffic faceted URLs from the Redshift query (query.txt; cached in
/tmp/seo_traffic_rows.pkl as rows of (main_cat, deepest_subcat, url, visits, revenue)),
finds the (cat_id, key) combos that have NO blueprint anywhere yet, builds the same
clean deterministic blueprints, and writes them to a NEW sheet `seo_traffic_new`
(blueprint cols + visits + revenue, sorted by revenue desc) inside the
tblPageTitles_blueprint_from_urls workbook.

"No blueprint yet" = NOT in live MySQL beslist.tblPageTitles, NOT in the prior
tblPageTitles_new_from_unique.xlsx, and NOT among the combos already generated in
the blueprint_from_urls deliverable.

Run under the mysql venv (needs pymysql):
  ~/.mysql-venv/bin/python scripts/pagetitles_blueprint_from_seo_traffic.py
"""
import os, sys, json, pickle

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import pagetitles_blueprint_from_urls as bp   # reuse parse/blueprint logic

ROWS_PKL  = '/tmp/seo_traffic_rows.pkl'
OUT_BASE  = bp.OUT_BASE        # .../tblPageTitles_blueprint_from_urls
SHEET     = 'seo_traffic_new'


def load_generated_combos():
    """(cat_id, canon_key) already proposed in our blueprint_from_urls output(s)."""
    import openpyxl
    combos = set()
    for suffix in ('_v2.xlsx', '.xlsx'):
        path = OUT_BASE + suffix
        if not os.path.exists(path):
            continue
        wb = openpyxl.load_workbook(path, read_only=True)
        if 'new_pagetitles' in wb.sheetnames:
            ws = wb['new_pagetitles']
            it = ws.iter_rows(values_only=True); hdr = next(it)
            ci = {h: i for i, h in enumerate(hdr)}
            for r in it:
                combos.add((r[ci['cat_id']], bp.canon_key(r[ci['key']])))
        wb.close()
    return combos


def main():
    import psycopg2
    rows = pickle.load(open(ROWS_PKL, 'rb'))
    print(f"[load] redshift rows={len(rows)}", file=sys.stderr)

    slug2id = json.load(open('/tmp/slug2id.json'))
    id2name = {int(k): v for k, v in json.load(open('/tmp/id2name.json')).items()}
    slug2id_l = {k.lower(): v for k, v in slug2id.items()}

    pg = psycopg2.connect(bp.PG_DSN)
    rules = bp.load_rules(pg.cursor())
    pg.close()

    skip = bp.load_existing_combos()
    n_xlsx = len(skip)
    tbl = bp.load_tblpagetitles_combos()
    gen = load_generated_combos()
    skip |= tbl; skip |= gen
    print(f"[load] skip: xlsx={n_xlsx} tblPageTitles={len(tbl)} generated={len(gen)} "
          f"union={len(skip)}", file=sys.stderr)

    # aggregate visits/revenue per (cat_id, canon_key); keep a representative type set
    agg = {}   # (cat, key) -> [visits, revenue, types]
    no_cat = no_facets = parsed = 0
    for _mc, _sc, url, visits, revenue in rows:
        url = (url or '').lower()
        p = bp.parse_url(url)
        if not p:
            continue
        leaf, types = p
        cat = slug2id_l.get(leaf)
        if cat is None:
            no_cat += 1; continue
        if not types:
            no_facets += 1; continue
        parsed += 1
        key = '~'.join(sorted(types))
        a = agg.get((cat, key))
        if a is None:
            agg[(cat, key)] = [int(visits or 0), float(revenue or 0), types]
        else:
            a[0] += int(visits or 0); a[1] += float(revenue or 0)

    new = {ck: v for ck, v in agg.items() if ck not in skip}
    print(f"[scan] parsed={parsed} no_cat={no_cat} no_facets={no_facets}", file=sys.stderr)
    print(f"[scan] distinct trafficked combos={len(agg)} | NO-blueprint (new)={len(new)}",
          file=sys.stderr)

    # build rows, sort by revenue desc then visits desc
    out_rows = []
    unknown = {}
    for (cat, key), (visits, revenue, types) in new.items():
        cat_id, cat_name, k, title, h1, desc, cc = bp.build_row(
            cat, id2name.get(cat, ''), types, rules, unknown)
        out_rows.append([cat_id, cat_name, k, title, h1, desc, cc,
                         visits, round(revenue, 2)])
    out_rows.sort(key=lambda r: (-r[8], -r[7]))

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    # add the sheet to the current deliverable (_v2 if present, else base)
    target = OUT_BASE + ('_v2.xlsx' if os.path.exists(OUT_BASE + '_v2.xlsx')
                         else '.xlsx')
    wb = openpyxl.load_workbook(target)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)
    hdr = ['cat_id', 'cat_name', 'key', 'title', 'h1_title', 'description',
           'country_code', 'visits', 'revenue']
    ws.append(hdr)
    for cell in ws[1]:
        cell.font = Font(bold=True); cell.fill = PatternFill('solid', fgColor='DDDDDD')
    for r in out_rows:
        ws.append(r)
    try:
        wb.save(target)
        out = target
    except PermissionError:
        out = OUT_BASE + '_v3.xlsx'
        wb.save(out)
    tot_rev = sum(r[8] for r in out_rows)
    print(f"wrote {len(out_rows)} rows to sheet '{SHEET}' (total revenue "
          f"€{tot_rev:,.0f}) -> {out}")


if __name__ == '__main__':
    main()
