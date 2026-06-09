#!/usr/bin/env python3
"""
For each category, rank its facets by SEO visits (Redshift traffic, cached in
/tmp/seo_traffic_rows.pkl), take the top N (default 5, pass N as argv[1]), and
generate the title/h1/description blueprint for EVERY non-empty subset of those
N facets (the power set: 2**N-1 combos for a category with >=N facets).

Writes the COMPLETE set to a new sheet `top{N}_combinations`, and appends only
the combos not already present to `all_combined` (source='top{N}_combinations').

Mind Excel's 1,048,576-row-per-sheet limit: top-8 ~405k fits; top-10 ~1.11M
does not.

Reuses the construction logic from pagetitles_blueprint_from_urls.

Run under the mysql venv:
  ~/.mysql-venv/bin/python scripts/pagetitles_topn_combinations.py [N]
"""
import os, sys, json, pickle, itertools
from collections import defaultdict, Counter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import pagetitles_blueprint_from_urls as bp

ROWS_PKL = '/tmp/seo_traffic_rows.pkl'
WORKBOOK = bp.OUT_BASE + '.xlsx'
TOP_N    = int(sys.argv[1]) if len(sys.argv) > 1 else 5   # facets per category
SHEET    = f'top{TOP_N}_combinations'


def rank_facets_by_visits(slug2id_l):
    """cat_id -> [facet_type, ...] top-N by summed SEO visits of urls using it."""
    rows = pickle.load(open(ROWS_PKL, 'rb'))
    cat_facet = defaultdict(Counter)   # cat -> Counter(facet_type -> visits)
    for _mc, _sc, url, visits, _rev in rows:
        p = bp.parse_url((url or '').lower())
        if not p:
            continue
        leaf, types = p
        cat = slug2id_l.get(leaf)
        if cat is None or not types:
            continue
        v = int(visits or 0)
        for t in types:
            cat_facet[cat][t] += v
    top = {}
    for cat, cnt in cat_facet.items():
        # deterministic: visits desc, then facet name
        ranked = sorted(cnt.items(), key=lambda kv: (-kv[1], kv[0]))
        top[cat] = [t for t, _ in ranked[:TOP_N]]
    return top


def main():
    import psycopg2, openpyxl
    from openpyxl.styles import Font, PatternFill

    slug2id = json.load(open('/tmp/slug2id.json'))
    id2name = {int(k): v for k, v in json.load(open('/tmp/id2name.json')).items()}
    slug2id_l = {k.lower(): v for k, v in slug2id.items()}

    pg = psycopg2.connect(bp.PG_DSN)
    rules = bp.load_rules(pg.cursor())
    pg.close()

    top = rank_facets_by_visits(slug2id_l)
    print(f"[rank] categories with SEO-traffic facets={len(top)}", file=sys.stderr)

    # generate every non-empty subset of each category's top-5
    unknown = {}
    rows = []        # full set for the dedicated sheet
    seen = set()     # (cat, canon_key) generated
    for cat, facets in top.items():
        cat_name = id2name.get(cat, '')
        for r in range(1, len(facets) + 1):
            for combo in itertools.combinations(facets, r):
                types = set(combo)
                key = '~'.join(sorted(types))
                ck = (cat, key)
                if ck in seen:
                    continue
                seen.add(ck)
                row = bp.build_row(cat, cat_name, types, rules, unknown)
                rows.append(row)   # [cat_id, cat_name, key, title, h1, desc, cc]
    print(f"[gen] total combinations={len(rows)}", file=sys.stderr)

    wb = openpyxl.load_workbook(WORKBOOK)

    # --- dedicated sheet: full set ---
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)
    hdr = ['cat_id', 'cat_name', 'key', 'title', 'h1_title', 'description', 'country_code']
    ws.append(hdr)
    for cell in ws[1]:
        cell.font = Font(bold=True); cell.fill = PatternFill('solid', fgColor='DDDDDD')
    for row in sorted(rows, key=lambda r: (r[0], r[2])):
        ws.append(row)

    # --- all_combined: append only net-new combos ---
    ac = wb['all_combined']
    it = ac.iter_rows(values_only=True); achdr = next(it)
    aci = {h: i for i, h in enumerate(achdr)}
    existing = set()
    for r in it:
        existing.add((r[aci['cat_id']], bp.canon_key(r[aci['key']])))
    added = 0
    for cat, cat_name, key, title, h1, desc, cc in rows:
        ck = (cat, bp.canon_key(key))
        if ck in existing:
            continue
        existing.add(ck)
        ac.append([cat, cat_name, key, title, h1, desc, cc, SHEET])
        added += 1

    try:
        wb.save(WORKBOOK); out = WORKBOOK
    except PermissionError:
        out = bp.OUT_BASE + '_v2.xlsx'; wb.save(out)
    print(f"wrote sheet '{SHEET}' = {len(rows)} rows; appended {added} net-new to "
          f"all_combined -> {out}")


if __name__ == '__main__':
    main()
