#!/usr/bin/env python3
"""
Build CLEAN, deterministic tblPageTitles title/h1/description blueprints from the
faceted URLs themselves (NOT from any rendered/curated copy).

Per faceted URL (`.../<leaf-slug>/c/<type>~<vid>~~<type>~<vid>...`):
  leaf-slug                 -> tax v2 cat_id (9xxxxxx)        [/tmp/slug2id.json]
  cat_id                    -> cat_name                       [/tmp/id2name.json]
  facet types (deduped set) -> ordered placeholders           [pa.facet_position_rules]

Facet phrase ("facet combination"):
  - placeholders `!!<facet_type>!!` ordered by facet_position_rules.order_index (asc)
  - the NOUN is a type-facet (is_type_facet) when the set has one; otherwise
    `!!sub_category!!` is inserted at the canonical type-facet slot (order 1700)
  - so every blueprint contains a category or a type-facet.

Templates:
  title       = !!current_query!! <facet phrase> kopen? ✔️ Tot !!DISCOUNT!! korting! | beslist.nl
  h1_title    = <facet phrase>
  description = Zoek je <facet phrase>? &#10062; Vergelijk !!NR!! aanbiedingen en bespaar
                op je aankoop &#10062; Shop <facet phrase> met !!DISCOUNT!! korting online!
                &#10062; beslist.nl

One row per distinct (cat_id, key) where key = '~'.join(sorted(facet types)).
Skips combos already present in the existing
  Downloads/claude/tblPageTitles_new_from_unique.xlsx  (sheet 'new_pagetitles').

Usage:
  pagetitles_blueprint_from_urls.py sample [N]   # print N example blueprints + stats
  pagetitles_blueprint_from_urls.py build        # full run -> new Excel
"""
import os, re, sys, json
from urllib.parse import unquote
import psycopg2, pymysql

CREDS = {}
for _line in open(os.path.expanduser('~/.mysql-creds')):
    if '=' in _line:
        _k, _v = _line.strip().split('=', 1); CREDS[_k] = _v


def my():
    return pymysql.connect(host=CREDS['MYSQL_HOST'], user=CREDS['MYSQL_USER'],
                           password=CREDS['MYSQL_PASSWORD'], db='beslist', charset='utf8mb4')

PG_DSN = ("postgresql://dbadmin:Q9fGRKtUdvdtxsiCM12HeFe0Nki0PvmjZRFLZ9ArmlWdMnDQXX8SdxKnPniqGmq6"
          "@10.1.32.9:5432/n8n-vector-db")

EXISTING_XLSX = '/mnt/c/Users/JoepvanSchagen/Downloads/claude/tblPageTitles_new_from_unique.xlsx'
OUT_BASE      = '/mnt/c/Users/JoepvanSchagen/Downloads/claude/tblPageTitles_blueprint_from_urls'
COUNTRY_CODE  = 'NL'

# canonical slot for the category noun when no type-facet is present:
# after brand/colour/material (merk=3, kleur=22, ...) but before size (maat=2300).
SUBCATEGORY_ORDER = 1700
SUBCATEGORY_PH    = '!!sub_category!!'
# fallback order for any facet type missing from facet_position_rules (rare)
UNKNOWN_ORDER     = 1500
# url params that are not real facets (price slider) -> never blueprint these
IGNORE_FACETS     = {'pricemin', 'pricemax'}

TAIL_TITLE = 'kopen? ✔️ Tot !!DISCOUNT!! korting! | beslist.nl'


def canon_key(s):
    """Canonical comparable form of a stored '~'-joined facet key: lowercase each
    type and re-sort, so skip-set keys match the (lowercased) generated keys."""
    return '~'.join(sorted(t for t in (s or '').lower().split('~') if t))


def parse_url(url):
    """url -> (leaf_slug, set_of_facet_types) or None when not a faceted /c/ url.
    The url is expected already lowercased by the caller."""
    if '/c/' not in url:
        return None
    path, fstr = url.split('/c/', 1)
    segs = [s for s in path.split('/') if s]
    leaf = segs[-1] if segs else ''
    types = set()
    for pair in fstr.split('~~'):
        bits = pair.split('~')
        if len(bits) >= 2 and bits[0]:
            # decode %28(%29 / %20 etc. so the type matches its facet_slug in the rules
            t = unquote(bits[0])
            if t not in IGNORE_FACETS:
                types.add(t)
    return leaf, types


def load_rules(cur):
    """facet_slug -> (order_index, is_type_facet)."""
    cur.execute("SELECT facet_slug, order_index, is_type_facet FROM pa.facet_position_rules")
    rules = {}
    for slug, order, is_type in cur.fetchall():
        rules[slug] = (order if order is not None else UNKNOWN_ORDER, bool(is_type))
    return rules


def load_existing_combos():
    """(cat_id, canon_key) set already present in the prior xlsx deliverable."""
    import openpyxl
    wb = openpyxl.load_workbook(EXISTING_XLSX, read_only=True)
    ws = wb['new_pagetitles']
    it = ws.iter_rows(values_only=True)
    hdr = next(it)
    ci = {h: i for i, h in enumerate(hdr)}
    existing = set()
    for r in it:
        existing.add((r[ci['cat_id']], canon_key(r[ci['key']])))
    wb.close()
    return existing


def load_tblpagetitles_combos():
    """(cat_id, canon_key) set already live in MySQL beslist.tblPageTitles (NL)."""
    conn = my(); c = conn.cursor()
    c.execute("SELECT cat_id, `key` FROM tblPageTitles WHERE country_code='NL'")
    combos = set((r[0], canon_key(r[1])) for r in c.fetchall())
    conn.close()
    return combos


def facet_phrase(types, rules, unknown_counter):
    """Ordered placeholder phrase for a set of facet types.
    Returns the phrase string. Inserts !!sub_category!! at SUBCATEGORY_ORDER when
    the set has no type-facet."""
    items = []   # (order, slug, placeholder)
    has_type = False
    for t in types:
        order, is_type = rules.get(t, (UNKNOWN_ORDER, False))
        if t not in rules:
            unknown_counter[t] = unknown_counter.get(t, 0) + 1
        if is_type:
            has_type = True
        items.append((order, t, f'!!{t}!!'))
    if not has_type:
        items.append((SUBCATEGORY_ORDER, '', SUBCATEGORY_PH))
    items.sort(key=lambda x: (x[0], x[1]))
    return ' '.join(ph for _, _, ph in items)


def build_row(cat, cat_name, types, rules, unknown_counter):
    key = '~'.join(sorted(types))
    phrase = facet_phrase(types, rules, unknown_counter)
    title = f'!!current_query!! {phrase} {TAIL_TITLE}'
    h1 = phrase
    desc = (f'Zoek je {phrase}? &#10062; Vergelijk !!NR!! aanbiedingen en bespaar op je '
            f'aankoop &#10062; Shop {phrase} met !!DISCOUNT!! korting online! &#10062; beslist.nl')
    return [cat, cat_name, key, title, h1, desc, COUNTRY_CODE]


def main():
    mode = sys.argv[1] if len(sys.argv) > 1 else 'sample'
    N = int(sys.argv[2]) if len(sys.argv) > 2 else 25

    slug2id = json.load(open('/tmp/slug2id.json'))
    id2name = {int(k): v for k, v in json.load(open('/tmp/id2name.json')).items()}
    # case-insensitive slug lookup (urls are lowercased before parsing)
    slug2id_l = {k.lower(): v for k, v in slug2id.items()}
    print(f"[load] slug2id={len(slug2id)} id2name={len(id2name)}", file=sys.stderr)

    pg = psycopg2.connect(PG_DSN)
    rules = load_rules(pg.cursor())
    print(f"[load] facet_position_rules={len(rules)} "
          f"(type-facets={sum(1 for _, t in rules.values() if t)})", file=sys.stderr)

    # skip-set = prior xlsx combos UNION live MySQL tblPageTitles combos (canon keys)
    existing = load_existing_combos()
    nx = len(existing)
    tbl = load_tblpagetitles_combos()
    existing |= tbl
    print(f"[load] skip combos: xlsx={nx} tblPageTitles={len(tbl)} union={len(existing)}",
          file=sys.stderr)

    cur = pg.cursor(name='url_stream'); cur.itersize = 50000
    cur.execute("SELECT url FROM pa.urls WHERE url LIKE '%/c/%'")

    seen = set()           # (cat, key) emitted this run (dedup)
    unknown_counter = {}
    rows = []
    scanned = no_cat = no_facets = skipped_existing = dup = 0
    examples = []
    for (url,) in cur:
        scanned += 1
        url = url.lower()                       # lowercase the whole url, then parse
        p = parse_url(url)
        if not p:
            continue
        leaf, types = p
        cat = slug2id_l.get(leaf)
        if cat is None:
            no_cat += 1; continue
        if not types:
            no_facets += 1; continue
        key = '~'.join(sorted(types))           # types already lowercase -> canon form
        ck = (cat, key)
        if ck in seen:
            dup += 1; continue
        seen.add(ck)
        if ck in existing:
            skipped_existing += 1; continue
        row = build_row(cat, id2name.get(cat, ''), types, rules, unknown_counter)
        if mode == 'sample':
            if len(examples) < N:
                examples.append((url, row))
        else:
            rows.append(row)
    pg.close()

    print(f"\n[scan] urls={scanned} no_cat={no_cat} no_facets={no_facets} "
          f"dup={dup} skipped_existing={skipped_existing}", file=sys.stderr)
    print(f"[scan] NEW blueprint rows={len(seen) - skipped_existing - 0}", file=sys.stderr)
    if unknown_counter:
        top = sorted(unknown_counter.items(), key=lambda x: -x[1])[:20]
        print(f"[warn] facet types not in rules: {len(unknown_counter)} "
              f"(placed at order {UNKNOWN_ORDER}); top: {top}", file=sys.stderr)

    if mode == 'sample':
        for url, row in examples:
            cat, cat_name, key, title, h1, desc, cc = row
            print(f"\ncat={cat} ({cat_name})  key={key}")
            print(f"  url  : {url}")
            print(f"  title: {title}")
            print(f"  h1   : {h1}")
            print(f"  desc : {desc}")
        return

    # ---- build mode: write Excel ----
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'new_pagetitles'
    hdr = ['cat_id', 'cat_name', 'key', 'title', 'h1_title', 'description', 'country_code']
    ws.append(hdr)
    for cell in ws[1]:
        cell.font = Font(bold=True); cell.fill = PatternFill('solid', fgColor='DDDDDD')
    # stable order: cat_id then key
    for row in sorted(rows, key=lambda r: (r[0], r[2])):
        ws.append(row)
    os.makedirs(os.path.dirname(OUT_BASE), exist_ok=True)
    out = OUT_BASE + '.xlsx'
    try:
        wb.save(out)
    except PermissionError:
        out = OUT_BASE + '_v2.xlsx'
        wb.save(out)
    print(f"wrote {len(rows)} rows -> {out}")


if __name__ == '__main__':
    main()
