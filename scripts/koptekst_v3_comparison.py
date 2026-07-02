#!/usr/bin/env python3
"""
Benchmark: v1 productie-koptekst vs v3 per-maincat informationele koopgids-prompt.

Voor een representatieve, visit-gewogen set URLs per maincat (uit
backend/data/koptekst_v3_benchmark_urls.json) worden dezelfde producten via de
Beslist Product Search API opgehaald en worden BEIDE kopteksten vers gegenereerd
met identieke productcontext, zodat puur de system message verschilt:
  v1 = backend/gpt_service.generate_product_content  (huidige productie)
  v3 = backend/gpt_service_v3.generate_product_content_v3 (per-maincat)

Beide worden gescoord op compliance-dimensies (zoals v2-benchmark) plus nieuwe
informationele dimensies (meetbare specs, aantal alinea's, koopvraag-signalen).
Output: side-by-side Excel (gegroepeerd per maincat) naar Downloads.

Gebruik:
  cd /home/joepvanschagen/projects/dm-tools
  ./venv/bin/python scripts/koptekst_v3_comparison.py [--per-maincat 2] [--only Maincat] [--out PATH]
"""
import argparse
import json
import os
import re
import sys
import time

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.scraper_service import scrape_product_page_api
from backend.gpt_service import generate_product_content          # v1 (productie)
from backend.gpt_service_v3 import generate_product_content_v3, build_system_message_v3

SAMPLE_PATH = os.path.join(os.path.dirname(__file__), "..", "backend", "data", "koptekst_v3_benchmark_urls.json")

# ---------- Scoring ----------
GENERIC_WORDS = ["ideaal", "perfect", "uitstekend", "een goede keuze", "een heerlijke keuze"]
# meetbare spec = getal + eenheid (kern-signaal van informationele content)
SPEC_RE = re.compile(
    r"\b\d[\d.,]*\s?(mm|cm|m²|m2|meter|kg|gram|g|liter|l|ml|db|watt|w|kwh|volt|v|ampère|ampere|a|mah|inch|\"|dpi|ip\d{2}|°c|graden|km|km/u|kelvin|k|pk|bar|mbar|karaat|tesla|teraflops|spm)\b",
    re.I,
)

def score(text: str) -> dict:
    if not text or text.startswith("[") :
        return {"valid": False}
    t = text.strip()
    plain = re.sub(r"<[^>]+>", " ", t)
    low = plain.lower()
    paras = [p for p in re.split(r"\n\s*\n|<br\s*/?>|</p>|<h[1-6][^>]*>", t) if p.strip()]
    return {
        "valid": True,
        "char_count": len(t),
        "word_count": len(plain.split()),
        "n_paragraphs": max(len(paras), 1),
        "n_links": len(re.findall(r"<a\s+href=", t, re.I)),
        "has_euro": "€" in t or bool(re.search(r"\beuro\b", low)),
        "n_exclamations": plain.count("!"),
        "uses_wij_ons": bool(re.search(r"\b(wij|ons|onze|we)\b", low)),
        "contains_generic": [w for w in GENERIC_WORDS if w in low],
        "has_h3": bool(re.search(r"<h[23]\b", t, re.I)),
        # informationele signalen
        "n_measurable_specs": len(SPEC_RE.findall(t)),
        "has_measurable_spec": bool(SPEC_RE.search(t)),
        "n_questions": plain.count("?"),
    }

def summarize(scores, label):
    v = [s for s in scores if s.get("valid")]
    n = len(v)
    if not n:
        return {"label": label, "n": 0}
    avg = lambda k: sum(s.get(k, 0) for s in v) / n
    pct = lambda k: sum(1 for s in v if s.get(k)) / n * 100.0
    return {
        "label": label, "n": n,
        "avg_words": round(avg("word_count")),
        "avg_chars": round(avg("char_count")),
        "avg_paragraphs": round(avg("n_paragraphs"), 2),
        "pct_multi_para": sum(1 for s in v if s["n_paragraphs"] >= 2) / n * 100.0,
        "pct_has_measurable_spec": pct("has_measurable_spec"),
        "avg_measurable_specs": round(avg("n_measurable_specs"), 2),
        "pct_has_question": sum(1 for s in v if s["n_questions"] > 0) / n * 100.0,
        "avg_links": round(avg("n_links"), 2),
        "pct_has_euro": pct("has_euro"),
        "pct_uses_wij_ons": pct("uses_wij_ons"),
        "avg_exclamations": round(avg("n_exclamations"), 2),
        "pct_contains_generic": sum(1 for s in v if s["contains_generic"]) / n * 100.0,
    }

def load_samples(per_maincat, only):
    rows = json.load(open(SAMPLE_PATH, encoding="utf-8"))
    if only:
        rows = [r for r in rows if r["maincat"].lower() == only.lower()]
    out, counts = [], {}
    for r in rows:
        c = counts.get(r["maincat"], 0)
        if c >= per_maincat:
            continue
        counts[r["maincat"]] = c + 1
        out.append(r)
    return out

def write_excel(out_path, results, s1, s3):
    wb = Workbook()
    an = wb.active; an.title = "Analyse"
    for col, w in {"A": 52, "B": 16, "C": 16, "D": 26}.items():
        an.column_dimensions[col].width = w
    an["A1"] = f"v1 (productie) vs v3 (per-maincat koopgids) — {len(results)} URLs over {len({r['maincat'] for r in results})} maincats"
    an["A1"].font = Font(bold=True, size=14)
    an["A2"] = "Beide vers gegenereerd met IDENTIEKE producten; alleen de system message verschilt."
    an["A2"].font = Font(italic=True, size=10)
    rows = [
        ("METRIEK", "v1", "v3", "doel v3"),
        ("__S__", "Informationele diepte"),
        ("Gem. woorden", s1.get("avg_words"), s3.get("avg_words"), "160-320"),
        ("Gem. tekens", s1.get("avg_chars"), s3.get("avg_chars"), None),
        ("Gem. alinea's", s1.get("avg_paragraphs"), s3.get("avg_paragraphs"), ">=2"),
        ("% meerdere alinea's", f"{s1.get('pct_multi_para',0):.0f}%", f"{s3.get('pct_multi_para',0):.0f}%", "hoog"),
        ("% met meetbare spec (getal+eenheid)", f"{s1.get('pct_has_measurable_spec',0):.0f}%", f"{s3.get('pct_has_measurable_spec',0):.0f}%", "hoog"),
        ("Gem. aantal meetbare specs", s1.get("avg_measurable_specs"), s3.get("avg_measurable_specs"), "hoger"),
        ("% beantwoordt/stelt koopvraag (?)", f"{s1.get('pct_has_question',0):.0f}%", f"{s3.get('pct_has_question',0):.0f}%", "hoger"),
        ("__S__", "Compliance (moet gelijk/beter blijven)"),
        ("Gem. product-links", s1.get("avg_links"), s3.get("avg_links"), None),
        ("% met prijs/euro", f"{s1.get('pct_has_euro',0):.0f}%", f"{s3.get('pct_has_euro',0):.0f}%", "0%"),
        ("% gebruikt wij/ons/onze/we", f"{s1.get('pct_uses_wij_ons',0):.0f}%", f"{s3.get('pct_uses_wij_ons',0):.0f}%", "0%"),
        ("Gem. uitroeptekens", s1.get("avg_exclamations"), s3.get("avg_exclamations"), "0"),
        ("% generieke kwalificatie (ideaal/perfect)", f"{s1.get('pct_contains_generic',0):.0f}%", f"{s3.get('pct_contains_generic',0):.0f}%", "laag"),
    ]
    r = 4
    for e in rows:
        if e[0] == "__S__":
            c = an.cell(row=r, column=1, value=e[1]); c.font = Font(bold=True, size=12)
            c.fill = PatternFill("solid", fgColor="D9E1F2"); r += 1; continue
        for ci, val in enumerate(e, 1):
            an.cell(row=r, column=ci, value=val)
        if r == 4:
            for ci in range(1, 5):
                an.cell(row=r, column=ci).font = Font(bold=True)
        r += 1

    ws = wb.create_sheet("v1 vs v3")
    heads = ["#", "Maincat", "URL", "H1 (scrape)", "#prod", "v1 wrd", "v3 wrd",
             "Koptekst v1 (productie)", "Koptekst v3 (per-maincat)"]
    for c, h in enumerate(heads, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True); cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for i, r_ in enumerate(sorted(results, key=lambda x: x["maincat"]), 2):
        vals = [i-1, r_["maincat"], r_["url"], r_["h1"], r_["n_products"],
                r_["s1"].get("word_count", 0) if r_["s1"].get("valid") else 0,
                r_["s3"].get("word_count", 0) if r_["s3"].get("valid") else 0,
                r_["v1"], r_["v3"]]
        for c, v in enumerate(vals, 1):
            ws.cell(row=i, column=c, value=v).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 300
    for col, w in {1:5,2:18,3:50,4:28,5:7,6:7,7:7,8:78,9:88}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 42; ws.freeze_panes = "D2"

    ps = wb.create_sheet("Prompts")
    import inspect
    from backend import gpt_service
    v1_sys = inspect.getsource(gpt_service.generate_product_content).split('system_message = """')[1].split('"""')[0]
    ps["A1"] = "v1 system message (productie)"; ps["A1"].font = Font(bold=True)
    ps["A2"] = v1_sys
    ps["A4"] = "v3 = per-maincat prompt uit backend/data/kopteksten_maincat_prompts_v3.json + genormaliseerd lengte-/structuurbeleid."; ps["A4"].font = Font(bold=True)
    ps["A5"] = "Voorbeeld v3 system message (Woonaccessoires):"; ps["A5"].font = Font(bold=True)
    ps["A6"] = build_system_message_v3("Woonaccessoires")
    for row in (2, 6):
        ps.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ps.row_dimensions[row].height = 620
    ps.column_dimensions["A"].width = 140

    for attempt in range(3):
        try:
            wb.save(out_path); print(f"\nWrote: {out_path}"); return
        except PermissionError:
            print(f"  save locked, retry {attempt+1}"); time.sleep(1)
    alt = out_path.replace(".xlsx", f"_{int(time.time())}.xlsx")
    wb.save(alt); print(f"\nWrote: {alt}")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--per-maincat", type=int, default=2)
    ap.add_argument("--only", default=None, help="beperk tot één maincat (naam)")
    ap.add_argument("--out", default="/mnt/c/Users/JoepvanSchagen/Downloads/claude/koptekst_v1_vs_v3_2026-07-01.xlsx")
    args = ap.parse_args()

    samples = load_samples(args.per_maincat, args.only)
    print(f"{len(samples)} URLs ({args.per_maincat}/maincat)")
    results = []
    for i, s in enumerate(samples, 1):
        url = s["url"]; mc = s["maincat"]
        print(f"[{i}/{len(samples)}] {mc} :: {url[:70]}")
        try:
            scraped = scrape_product_page_api(url)
        except Exception as e:
            print(f"  scrape error: {e}"); scraped = None
        h1 = (scraped or {}).get("h1_title") or s.get("h1_hint") or "Onbekend"
        products = (scraped or {}).get("products", [])
        if not products:
            print("  geen producten -> skip"); continue
        try:
            v1 = generate_product_content(h1, products)
        except Exception as e:
            v1 = f"[v1 error: {e}]"; print(f"  v1 error: {e}")
        try:
            v3 = generate_product_content_v3(h1, products, mc)
        except Exception as e:
            v3 = f"[v3 error: {e}]"; print(f"  v3 error: {e}")
        results.append({"maincat": mc, "url": url, "h1": h1, "n_products": len(products),
                        "v1": v1, "v3": v3, "s1": score(v1), "s3": score(v3)})
    print(f"\n{len(results)} met producten. Scoren + Excel...")
    s1 = summarize([r["s1"] for r in results], "v1")
    s3 = summarize([r["s3"] for r in results], "v3")
    print("v1:", s1); print("v3:", s3)
    write_excel(args.out, results, s1, s3)

if __name__ == "__main__":
    main()
