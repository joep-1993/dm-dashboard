#!/usr/bin/env python3
"""
Benchmark runner voor v2 koptekst-prompt vs v1 productie.

Trekt N willekeurige URLs met bestaande v1-kopteksten uit pa.kopteksten_content (Postgres),
genereert v2 met dezelfde producten via de Beslist Product Search API,
scoort beide op compliance-dimensies, en schrijft side-by-side Excel naar Downloads.

Gebruik:
  cd /home/joepvanschagen/projects/dm-tools
  source venv/bin/activate
  python3 scripts/koptekst_v2_comparison.py [--n 50] [--seed mijnseed]

V2-prompt zelf staat in backend/gpt_service_v2.py — bewerk die als je het prompt wilt itereren.
"""
import argparse
import os
import re
import sys

# Make backend imports work when run from repo root
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.scraper_service import scrape_product_page_api
from backend.gpt_service_v2 import (
    SYSTEM_MESSAGE_V2,
    generate_product_content_v2,
)


# ---------- Scoring ----------

FORBIDDEN_OPENINGS = [
    "als je op zoek bent naar", "op zoek naar", "ben je op zoek naar", "zoek je",
    "welkom op de", "bij het kiezen van", "bij het selecteren van",
    "bij het uitkiezen van", "bij het overwegen van", "bij de keuze",
    "bij de aanschaf", "bij de zoektocht", "het kiezen van",
    "wanneer je op zoek", "wanneer je overweegt", "wanneer je kiest",
]
FORBIDDEN_GENERIC_WORDS = ["ideaal", "perfect", "uitstekend", "een goede keuze", "een heerlijke keuze"]
FORBIDDEN_AI_ISMS = [
    "laten we eens kijken", "in deze blog", "in deze gids", "hopelijk helpt dit",
    "welkom op de pagina", "tot slot,", "kortom,",
]


def score_koptekst(text: str) -> dict:
    if not text or text.startswith("[no products") or text.startswith("[v2 error"):
        return {"valid": False}
    t = text.strip()
    plain = re.sub(r"<[^>]+>", "", t)
    plain_lower = plain.lower()
    return {
        "valid": True,
        "char_count": len(t),
        "has_h_tags": bool(re.search(r"<h[1-6]\b", t, re.I)),
        "n_links": len(re.findall(r"<a\s+href=", t, re.I)),
        "n_relative_links": len(re.findall(r'href="/p/', t)),
        "n_absolute_links": len(re.findall(r'href="https?://www\.beslist\.nl', t)),
        "has_euro_sign": "€" in t,
        "has_concrete_shop_count": bool(re.search(r"\b\d{1,3}\s*(aanbieders|winkels|shops|verkopers)\b", plain_lower)),
        "has_vague_quantifier_shops": bool(re.search(r"(alle|veel|diverse|meerdere|vele)\s+(aanbieders|winkels|shops|verkopers)", plain_lower)),
        "mentions_beslist": "beslist" in plain_lower,
        "starts_with_forbidden": any(plain_lower.startswith(p) for p in FORBIDDEN_OPENINGS),
        "starts_with_bij_het_kiezen": plain_lower.startswith("bij het kiezen van"),
        "contains_generic_words": [w for w in FORBIDDEN_GENERIC_WORDS if w in plain_lower],
        "contains_ai_isms": [w for w in FORBIDDEN_AI_ISMS if w in plain_lower],
        "n_exclamations": plain.count("!"),
        "uses_wij_ons": bool(re.search(r"\b(wij|ons|onze|we)\b", plain_lower)),
    }


def summarize_scores(scores: list, label: str) -> dict:
    valid = [s for s in scores if s.get("valid")]
    n = len(valid)
    if n == 0:
        return {"label": label, "n": 0}

    def pct(key, val=True):
        return sum(1 for s in valid if s.get(key) == val) / n * 100.0

    def avg(key):
        return sum(s.get(key, 0) for s in valid) / n

    return {
        "label": label, "n": n,
        "avg_char_count": round(avg("char_count")),
        "min_char_count": min(s["char_count"] for s in valid),
        "max_char_count": max(s["char_count"] for s in valid),
        "pct_has_h_tags": pct("has_h_tags"),
        "avg_n_links": round(avg("n_links"), 2),
        "pct_has_relative_links": sum(1 for s in valid if s["n_relative_links"] > 0) / n * 100.0,
        "pct_has_absolute_links": sum(1 for s in valid if s["n_absolute_links"] > 0) / n * 100.0,
        "pct_has_euro": pct("has_euro_sign"),
        "pct_concrete_shop_count": pct("has_concrete_shop_count"),
        "pct_vague_quantifier": pct("has_vague_quantifier_shops"),
        "pct_mentions_beslist": pct("mentions_beslist"),
        "pct_starts_forbidden": pct("starts_with_forbidden"),
        "pct_starts_bij_het_kiezen": pct("starts_with_bij_het_kiezen"),
        "pct_contains_ideaal": sum(1 for s in valid if "ideaal" in (s["contains_generic_words"] or [])) / n * 100.0,
        "pct_contains_perfect": sum(1 for s in valid if "perfect" in (s["contains_generic_words"] or [])) / n * 100.0,
        "pct_contains_any_generic": sum(1 for s in valid if s["contains_generic_words"]) / n * 100.0,
        "pct_contains_ai_ism": sum(1 for s in valid if s["contains_ai_isms"]) / n * 100.0,
        "pct_uses_wij_ons": pct("uses_wij_ons"),
        "avg_exclamations": round(avg("n_exclamations"), 2),
    }


def sample_urls(n: int, seed: str) -> list:
    conn = psycopg2.connect(os.getenv("DATABASE_URL"))
    cur = conn.cursor()
    cur.execute(
        """
        WITH sample AS (
          SELECT k.url_id, k.content, u.url
          FROM pa.kopteksten_content k
          JOIN pa.urls u ON u.url_id = k.url_id
          WHERE k.content IS NOT NULL
            AND LENGTH(k.content) BETWEEN 600 AND 1500
            AND u.url LIKE '/products/%%'
            AND u.is_active = TRUE
          LIMIT 5000
        )
        SELECT url, content FROM sample
        ORDER BY MD5(url || %s)
        LIMIT %s
        """,
        (seed, n),
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def write_excel(out_path: str, results: list, summary_v1: dict, summary_v2: dict):
    wb = Workbook()

    # 1) Analyse sheet first
    an = wb.active
    an.title = "Analyse"
    an.column_dimensions["A"].width = 60
    an.column_dimensions["B"].width = 14
    an.column_dimensions["C"].width = 14
    an.column_dimensions["D"].width = 22

    an["A1"] = f"Aggregate analyse: v1 vs v2 op {len(results)} URLs"
    an["A1"].font = Font(bold=True, size=14)
    an["A2"] = "Steekproef: random URLs uit pa.kopteksten_content (Postgres)"
    an["A3"] = "v1 = bestaande productie-koptekst, v2 = nieuw gegenereerd met backend/gpt_service_v2.py"
    for r in (2, 3):
        an.cell(row=r, column=1).font = Font(italic=True, size=10)

    metric_rows = [
        ("METRIEK", "v1", "v2", "Δ / interpretatie"),
        ("__SECTION__", "Lengte"),
        ("Gemiddelde lengte (chars)", summary_v1.get("avg_char_count"), summary_v2.get("avg_char_count"), None),
        ("Min lengte", summary_v1.get("min_char_count"), summary_v2.get("min_char_count"), None),
        ("Max lengte", summary_v1.get("max_char_count"), summary_v2.get("max_char_count"), None),
        ("__SECTION__", "Format-compliance"),
        ("% met H-tags (verboden in v2)", f"{summary_v1.get('pct_has_h_tags', 0):.0f}%", f"{summary_v2.get('pct_has_h_tags', 0):.0f}%", "v2 doel: 0%"),
        ("Gem. aantal product-links", summary_v1.get("avg_n_links"), summary_v2.get("avg_n_links"), None),
        ("% met relatieve links (/p/...)", f"{summary_v1.get('pct_has_relative_links', 0):.0f}%", f"{summary_v2.get('pct_has_relative_links', 0):.0f}%", "v2 doel: hoog"),
        ("% met absolute links (https://...)", f"{summary_v1.get('pct_has_absolute_links', 0):.0f}%", f"{summary_v2.get('pct_has_absolute_links', 0):.0f}%", "v2 doel: 0%"),
        ("__SECTION__", "Verboden inhoud"),
        ("% met euro-teken (€)", f"{summary_v1.get('pct_has_euro', 0):.0f}%", f"{summary_v2.get('pct_has_euro', 0):.0f}%", "v2 doel: 0%"),
        ("% met concreet aantal aanbieders", f"{summary_v1.get('pct_concrete_shop_count', 0):.0f}%", f"{summary_v2.get('pct_concrete_shop_count', 0):.0f}%", "v2 doel: 0%"),
        ("% gebruikt wij/ons/onze/we", f"{summary_v1.get('pct_uses_wij_ons', 0):.0f}%", f"{summary_v2.get('pct_uses_wij_ons', 0):.0f}%", "doel: 0%"),
        ("Gem. aantal uitroeptekens", summary_v1.get("avg_exclamations"), summary_v2.get("avg_exclamations"), "doel: 0"),
        ("__SECTION__", "Comparison-authority (uniek voor v2)"),
        ("% met vage kwantificeerder (alle/veel/diverse aanbieders)", f"{summary_v1.get('pct_vague_quantifier', 0):.0f}%", f"{summary_v2.get('pct_vague_quantifier', 0):.0f}%", "v2 doel: hoog"),
        ("% noemt 'Beslist'", f"{summary_v1.get('pct_mentions_beslist', 0):.0f}%", f"{summary_v2.get('pct_mentions_beslist', 0):.0f}%", "v2 doel: 100%"),
        ("__SECTION__", "Verboden openingen / generieke woorden"),
        ("% start met verboden zin", f"{summary_v1.get('pct_starts_forbidden', 0):.0f}%", f"{summary_v2.get('pct_starts_forbidden', 0):.0f}%", "doel: 0%"),
        ("% start specifiek met 'Bij het kiezen van'", f"{summary_v1.get('pct_starts_bij_het_kiezen', 0):.0f}%", f"{summary_v2.get('pct_starts_bij_het_kiezen', 0):.0f}%", "doel: 0%"),
        ("% bevat 'ideaal'", f"{summary_v1.get('pct_contains_ideaal', 0):.0f}%", f"{summary_v2.get('pct_contains_ideaal', 0):.0f}%", "doel: 0%"),
        ("% bevat 'perfect'", f"{summary_v1.get('pct_contains_perfect', 0):.0f}%", f"{summary_v2.get('pct_contains_perfect', 0):.0f}%", "doel: 0%"),
        ("% bevat enige generieke kwalificatie", f"{summary_v1.get('pct_contains_any_generic', 0):.0f}%", f"{summary_v2.get('pct_contains_any_generic', 0):.0f}%", "doel: laag"),
        ("% bevat AI-isme", f"{summary_v1.get('pct_contains_ai_ism', 0):.0f}%", f"{summary_v2.get('pct_contains_ai_ism', 0):.0f}%", "doel: 0%"),
    ]

    row = 5
    for entry in metric_rows:
        if entry[0] == "__SECTION__":
            an.cell(row=row, column=1).value = entry[1]
            an.cell(row=row, column=1).font = Font(bold=True, size=12)
            an.cell(row=row, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row += 1
            continue
        for c, val in enumerate(entry, start=1):
            an.cell(row=row, column=c).value = val
        if row == 5:
            for c in range(1, 5):
                an.cell(row=row, column=c).font = Font(bold=True)
        row += 1

    # 2) Side-by-side sheet
    ws = wb.create_sheet("Koptekst v1 vs v2")
    headers = ["#", "URL", "H1 (uit scrape)", "Products scraped",
               "v1 length (chars)", "v2 length (chars)",
               "Koptekst v1 (productie — pa.kopteksten_content)",
               "Koptekst v2 (backend/gpt_service_v2.py)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.value = h
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for i, r in enumerate(results, 2):
        ws.cell(row=i, column=1).value = i - 1
        ws.cell(row=i, column=2).value = r["url"]
        ws.cell(row=i, column=3).value = r["h1"]
        ws.cell(row=i, column=4).value = r["n_products"]
        ws.cell(row=i, column=5).value = r.get("v1_len", 0)
        ws.cell(row=i, column=6).value = r.get("v2_len", 0)
        ws.cell(row=i, column=7).value = r["v1"]
        ws.cell(row=i, column=8).value = r["v2"]
        for c in range(2, 9):
            ws.cell(row=i, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 280

    widths = {1: 5, 2: 55, 3: 32, 4: 10, 5: 11, 6: 11, 7: 80, 8: 90}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 45
    ws.freeze_panes = "C2"

    # 3) Prompts reference sheet
    ps = wb.create_sheet("Prompts")
    ps["A1"] = "Versie"
    ps["A1"].font = Font(bold=True)
    ps["B1"] = "System-prompt"
    ps["B1"].font = Font(bold=True)
    import inspect
    from backend import gpt_service
    v1_sys = inspect.getsource(gpt_service.generate_product_content).split('system_message = """')[1].split('"""')[0]
    ps["A2"] = "v1 (productie)"
    ps["B2"] = v1_sys
    ps["A3"] = "v2 (gpt_service_v2.py)"
    ps["B3"] = SYSTEM_MESSAGE_V2
    for r in (2, 3):
        ps.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ps.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ps.row_dimensions[r].height = 700
    ps.column_dimensions["A"].width = 22
    ps.column_dimensions["B"].width = 130

    # Save (retry on locked file)
    import time as _time
    for attempt in range(3):
        try:
            wb.save(out_path)
            print(f"\nWrote: {out_path}")
            return
        except PermissionError:
            print(f"  save attempt {attempt+1} failed (file locked); retrying...")
            _time.sleep(1)
    ts_out = out_path.replace(".xlsx", f"_{int(_time.time())}.xlsx")
    wb.save(ts_out)
    print(f"\nWrote: {ts_out}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--n", type=int, default=50, help="aantal URLs (default 50)")
    parser.add_argument("--seed", default="seed20260512v50", help="seed voor reproducerebare sample")
    parser.add_argument("--out", default="/mnt/c/Users/JoepvanSchagen/Downloads/koptekst_v1_vs_v2_n50_v4.xlsx")
    args = parser.parse_args()

    samples = sample_urls(args.n, args.seed)
    print(f"Loaded {len(samples)} URLs from pa.kopteksten_content")

    results = []
    for i, (rel_url, v1_content) in enumerate(samples, 1):
        full_url = f"https://www.beslist.nl{rel_url}"
        print(f"\n[{i}/{len(samples)}] {full_url[:80]}")
        try:
            scraped = scrape_product_page_api(full_url)
        except Exception as e:
            scraped = None
            print(f"  scrape error: {e}")

        h1 = (scraped or {}).get("h1_title") or "Onbekend"
        products = (scraped or {}).get("products", [])
        print(f"  h1='{h1}'  scraped {len(products)} products")

        if not products:
            results.append({
                "url": full_url, "h1": h1, "n_products": 0,
                "v1": v1_content, "v2": "[no products — kan v2 niet genereren]",
                "v1_len": len(v1_content or ""), "v2_len": 0,
            })
            continue

        try:
            v2 = generate_product_content_v2(h1, products)
        except Exception as e:
            v2 = f"[v2 error: {e}]"
            print(f"  v2 error: {e}")

        results.append({
            "url": full_url, "h1": h1, "n_products": len(products),
            "v1": v1_content, "v2": v2,
            "v1_len": len(v1_content or ""), "v2_len": len(v2 or ""),
        })

    print("\nScoring all kopteksten...")
    for r in results:
        r["score_v1"] = score_koptekst(r["v1"])
        r["score_v2"] = score_koptekst(r["v2"])
    summary_v1 = summarize_scores([r["score_v1"] for r in results], "v1 (productie)")
    summary_v2 = summarize_scores([r["score_v2"] for r in results], "v2 (nieuw)")
    print(f"v1 valid: {summary_v1.get('n', 0)}, v2 valid: {summary_v2.get('n', 0)}")

    write_excel(args.out, results, summary_v1, summary_v2)


if __name__ == "__main__":
    main()
