"""Pre-classify every (facet_label, category_name) pair against the LLM so
title generation no longer has to classify on demand.

For each category in backend/data/cat_urls.csv:
  - fetch its facets via taxv2 /api/CategoryFacets
  - for each facet, fetch up to N values via /api/Facets/{id}/values
  - classify each value individually with _classify_with_llm
  - if values disagree → store False (tie-break per user spec); else store the
    unanimous verdict

Skips pairs already in pa.facet_type_classifications. Run from repo root:
    PYTHONPATH=. python3 scripts/preclassify_facets.py [--workers 20] [--limit-cats N] [--values-per-facet 2]
"""
from __future__ import annotations

import argparse
import csv
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from backend.database import get_db_connection, return_db_connection  # noqa: E402
from backend.facet_classifier import _classify_with_llm, _persist  # noqa: E402

import openpyxl  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402

TAX_BASE = os.getenv("TAXV2_BASE", "http://producttaxonomyunifiedapi-prod.azure.api.beslist.nl")
TAX_HEADERS = {"X-User-Name": "SEO_JOEP"}
CAT_URLS_CSV = ROOT / "backend" / "data" / "cat_urls.csv"


def load_categories() -> List[Dict]:
    rows = []
    with open(CAT_URLS_CSV, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f, delimiter=";"):
            cat_id = (r.get("cat_id") or "").strip()
            name = (r.get("deepest_cat") or "").strip()
            maincat = (r.get("maincat") or "").strip()
            if cat_id and name:
                rows.append({"cat_id": cat_id, "name": name, "maincat": maincat})
    return rows


def write_excel(rows: List[Dict], path: Path, values_per_facet: int) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pre-classifications"
    headers = ["maincat", "category", "facet_label", "facet_slug", "is_type_facet", "summary"]
    for i in range(values_per_facet):
        headers.extend([f"value_{i+1}", f"verdict_{i+1}", f"reason_{i+1}"])
    ws.append(headers)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="F0F0F0")
    for c in ws[1]:
        c.font = bold
        c.fill = fill
    for r in sorted(rows, key=lambda x: (x["maincat"], x["category"], x["facet_label"])):
        row = [r["maincat"], r["category"], r["facet_label"], r["facet_slug"],
               "TRUE" if r["is_type_facet"] else "FALSE", r["summary"]]
        for i in range(values_per_facet):
            v = r["values"][i] if i < len(r["values"]) else ""
            verdict = "TRUE" if (i < len(r["verdicts"]) and r["verdicts"][i]) else (
                "FALSE" if i < len(r["verdicts"]) else "")
            reason = r["reasons"][i] if i < len(r["reasons"]) else ""
            row.extend([v, verdict, reason])
        ws.append(row)
    widths = [16, 32, 22, 22, 12, 28] + [22, 10, 60] * values_per_facet
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(path)


def fetch_facets(session: requests.Session, cat_id: str) -> List[Dict]:
    """Return [{id, slug, label_nl}] for a category's linked facets."""
    try:
        r = session.get(
            f"{TAX_BASE}/api/CategoryFacets",
            params={"categoryId": cat_id, "locale": "nl-NL"},
            headers=TAX_HEADERS, timeout=20,
        )
        if r.status_code != 200:
            return []
        data = r.json()
        items = data if isinstance(data, list) else data.get("items", [])
        out = []
        for cf in items:
            facet = cf.get("facet") or cf
            slug = (facet.get("urlSlug") or "").lower()
            fid = facet.get("id")
            labels = facet.get("labels") or []
            nl = next((l for l in labels if l.get("locale") == "nl-NL"), {})
            label = nl.get("name") or facet.get("name") or slug
            if fid is not None and label:
                out.append({"id": str(fid), "slug": slug, "label": label})
        return out
    except Exception as e:
        print(f"  [skip cat {cat_id}] CategoryFacets error: {e}")
        return []


def fetch_values(session: requests.Session, facet_id: str, n: int) -> List[str]:
    """Return up to n value names for a facet."""
    try:
        r = session.get(
            f"{TAX_BASE}/api/Facets/{facet_id}/values",
            params={"locale": "nl-NL", "pageSize": max(n, 5), "page": 1},
            headers=TAX_HEADERS, timeout=20,
        )
        if r.status_code != 200:
            return []
        data = r.json()
        items = data.get("items", data) if isinstance(data, dict) else data
        names: List[str] = []
        for v in items[:max(n, 10)]:
            labels = v.get("labels") or []
            name = ""
            if labels:
                name = (labels[0].get("nameInColumn") or labels[0].get("nameOnDetail") or "").strip()
            if name:
                names.append(name)
            if len(names) >= n:
                break
        return names
    except Exception:
        return []


def already_classified() -> set:
    """Return the set of (facet_name_lower, category_lower) already in DB."""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT facet_name, sample_category FROM pa.facet_type_classifications")
        return {(r["facet_name"], r["sample_category"]) for r in cur.fetchall()}
    finally:
        cur.close()
        return_db_connection(conn)


def classify_pair(label: str, category: str, values: List[str]) -> Tuple[bool, str, List[bool], List[str]]:
    """Run the LLM on each value. Tie → False. Returns (final, summary_reason, per_value_verdicts, per_value_reasons)."""
    verdicts: List[bool] = []
    reasons: List[str] = []
    for v in values:
        is_t, reason = _classify_with_llm(label, v, category)
        verdicts.append(is_t)
        reasons.append(reason)
    if not verdicts:
        return False, "no values available", [], []
    if all(verdicts):
        return True, "all values agree → True", verdicts, reasons
    if not any(verdicts):
        return False, "all values agree → False", verdicts, reasons
    return False, "tie-break: mixed verdicts → False", verdicts, reasons


def process_category(cat: Dict, existing: set, values_per_facet: int) -> List[Dict]:
    """Return list of dicts ready for Excel export and DB persistence."""
    session = requests.Session()
    out: List[Dict] = []
    facets = fetch_facets(session, cat["cat_id"])
    for f in facets:
        label_l = f["label"].lower().strip()
        cat_l = cat["name"].lower().strip()
        if (label_l, cat_l) in existing:
            continue
        values = fetch_values(session, f["id"], values_per_facet)
        if len(values) < values_per_facet:
            # Not enough values to apply tie-break rule properly; skip and let
            # on-demand classification handle it later.
            continue
        is_t, summary, verdicts, reasons = classify_pair(f["label"], cat["name"], values)
        out.append({
            "maincat": cat.get("maincat", ""),
            "category": cat["name"],
            "category_l": cat_l,
            "facet_label": f["label"],
            "facet_label_l": label_l,
            "facet_slug": f["slug"],
            "values": values,
            "verdicts": verdicts,
            "reasons": reasons,
            "is_type_facet": is_t,
            "summary": summary,
        })
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workers", type=int, default=20)
    ap.add_argument("--limit-cats", type=int, default=0, help="0 = all categories")
    ap.add_argument("--values-per-facet", type=int, default=2)
    ap.add_argument("--output", type=str,
                    default=str(ROOT / f"facet_preclassification_{int(time.time())}.xlsx"),
                    help="Excel output path")
    ap.add_argument("--persist", action="store_true",
                    help="Also write rows to pa.facet_type_classifications "
                         "(ON CONFLICT DO NOTHING — manual overrides preserved)")
    args = ap.parse_args()

    cats = load_categories()
    if args.limit_cats:
        cats = cats[: args.limit_cats]
    print(f"Categories to process: {len(cats)}")
    existing = already_classified()
    print(f"Already classified: {len(existing)} (label, category) pairs")

    t0 = time.time()
    all_rows: List[Dict] = []
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = {ex.submit(process_category, c, existing, args.values_per_facet): c for c in cats}
        for i, fut in enumerate(as_completed(futs), 1):
            cat = futs[fut]
            try:
                rows = fut.result()
            except Exception as e:
                print(f"[{i}/{len(cats)}] {cat['name']!r}: error {e}")
                continue
            all_rows.extend(rows)
            if args.persist:
                for r in rows:
                    _persist(r["facet_label_l"], r["is_type_facet"],
                             r["values"][0] if r["values"] else "",
                             r["category_l"],
                             (r["summary"] + " | " + " ; ".join(r["reasons"]))[:500])
            if i % 25 == 0 or i == len(cats):
                rate = i / max(1, time.time() - t0)
                print(f"[{i}/{len(cats)}] new_pairs={len(all_rows)} rate={rate:.1f} cat/s")

    out_path = Path(args.output)
    write_excel(all_rows, out_path, args.values_per_facet)
    n_true = sum(1 for r in all_rows if r["is_type_facet"])
    print(f"\nDone in {time.time() - t0:.1f}s. New pairs: {len(all_rows)} ({n_true} True / {len(all_rows)-n_true} False)")
    print(f"Excel: {out_path}")
    if not args.persist:
        print("(--persist not set; rows NOT written to pa.facet_type_classifications)")


if __name__ == "__main__":
    main()
