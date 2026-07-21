#!/usr/bin/env python3
"""Bouw een Excel met de belangrijkste prompt-verschillen per hoofdcategorie (v3).

v3 gebruikt per maincat een eigen, individueel geschreven prompt. Het skelet
(mini-koopgids, 'je', geen prijzen, geen uitroeptekens, alleen producten uit de
lijst linken) is gedeeld; de kern van het verschil zit in: welke subcategorieen
genoemd worden, welke koopgids-onderwerpen/keuzecriteria behandeld moeten worden,
het vakjargon, de structuurvorm en de richtlengte.
"""
import json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/home/joepvanschagen/projects/dm-dashboard/backend/data/kopteksten_maincat_prompts_v3.json"
OUT = "/mnt/c/Users/JoepvanSchagen/Downloads/claude/koptekst_v3_prompt_verschillen_per_maincat_2026-07-02.xlsx"

d = json.load(open(SRC))

INHOUD_START = re.compile(r'^(INHOUD|INHOUDELIJKE|SCHRIJF ZO)\b', re.I)
NEXT_HEAD = re.compile(r'^(STRUCTUUR|STIJL|LENGTE|VORM|TERMINOLOGIE|KADER|BESLIST-KADER|'
                       r'PRIJSVERGELIJK|LINKREGELS|ALGEMEEN|TOON|TONE|HARDE REGELS|VASTE REGELS)\b', re.I)
TERM_START = re.compile(r'^TERMINOLOGIE\b', re.I)


def get_block(lines, start_re, stop_re):
    out, on = [], False
    for l in lines:
        s = l.strip()
        if on and stop_re.match(s):
            break
        if start_re.match(s):
            on = True
        if on:
            out.append(l.rstrip())
    return out


def subcats(intro):
    kw = ('subcategorie', 'zoals', 'bijvoorbeeld', 'denk aan', 'onderwerpen', ' als ', 'o.a.')
    for m in re.finditer(r'\(([^()]+)\)', intro):
        g = m.group(1).strip()
        gl = g.lower()
        if gl in ('koptekst', '"koptekst"', 'een prijsvergelijker', 'koptekst)'):
            continue
        if len(g) > 12 and (any(k in gl for k in kw) or ',' in g):
            return re.sub(r'^(subcategorie[eën]*\s+(zoals|als)\s+|zoals\s+|bijvoorbeeld\s+|o\.a\.\s+'
                          r'|denk aan\s+(onderwerpen als\s+|subcategorie[eën]*\s+als\s+)?|onderwerpen als\s+)',
                          '', g, flags=re.I).strip()
    # fallback: 'denk aan ... .' buiten haakjes
    m = re.search(r'denk aan\s+(?:onderwerpen als\s+|subcategorie[eën]*\s+als\s+)?(.+?)(?:\.\s|\. beslist|$)', intro, re.I)
    if m:
        return m.group(1).strip()
    return "— (generiek geschreven; subcategorieën blijken uit de koopgids-onderwerpen)"


def lengte(p):
    found = []
    for m in re.finditer(r'(\d[\d.]*)\s*(?:tot|-|–|—)\s*(\d[\d.]*)\s*woord', p):
        a = int(m.group(1).replace('.', '')); b = int(m.group(2).replace('.', ''))
        if a < 40:  # sla de ankertekst-regel '3 tot 5 woorden' over
            continue
        v = f"{a}–{b} woorden"
        if v not in found:
            found.append(v)
    return " / ".join(found) if found else "—"


def structuur(p):
    pl = p.lower()
    if any(s in pl for s in ['één samenhangende', 'éénalinea', 'behoud de bestaande éénalinea',
                             'behoud de éénalinea', 'schrijf één ']):
        return "Eén doorlopende alinea (uitzondering)"
    return "2–4 korte alinea’s (koopgids-structuur)"


def themes(block):
    labs = []
    for l in block:
        s = l.strip()
        m = re.match(r'^(?:[-*•]|\d+\.)\s*(.+)$', s)
        if not m:
            continue
        t = m.group(1)
        t = re.split(r'\s[—–:]\s|:\s|\s\(', t)[0].strip(' .—–:')
        if 2 < len(t) < 60:
            labs.append(t)
    seen, out = set(), []
    for t in labs:
        if t.lower() not in seen:
            seen.add(t.lower()); out.append(t)
    return out


def block_text(block, drop_header=True):
    lines = block[1:] if (drop_header and block) else block
    return "\n".join(l for l in lines).strip()


rows = []
for k, v in d.items():
    p = v["prompt"]; lines = p.split("\n"); intro = lines[0]
    ib = get_block(lines, INHOUD_START, NEXT_HEAD)
    tb = get_block(lines, TERM_START, NEXT_HEAD)
    rows.append({
        "maincat": k,
        "slug": v["slug"],
        "subcats": subcats(intro),
        "struct": structuur(p),
        "lengte": lengte(p),
        "themes": "\n".join(f"• {t}" for t in themes(ib)),
        "inhoud": block_text(ib) or "(inhoud verweven in de algemene regels)",
        "termino": block_text(tb),
        "prompt": p,
    })

# ---------- Excel ----------
wb = Workbook()

# kleuren
HDR = PatternFill("solid", fgColor="1F4E78")
ALT = PatternFill("solid", fgColor="EAF1F8")
TITLE = Font(bold=True, color="FFFFFF", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# --- Sheet 1: Toelichting ---
ws0 = wb.active
ws0.title = "Toelichting"
intro_txt = [
    ("Kopteksten v3 – prompt-verschillen per hoofdcategorie", 14, True, "1F4E78"),
    ("", 11, False, None),
    ("Wat is v3?", 12, True, "1F4E78"),
    ("In v3 krijgt elke hoofdcategorie een eigen, apart geschreven systeemprompt. "
     "Waar v1 één generieke promo-blurb-prompt gebruikte, is v3 per maincat opgebouwd "
     "als informatieve mini-koopgids, gebaseerd op analyse van de content die in Google "
     "op dat type zoekopdracht rankt.", 11, False, None),
    ("", 11, False, None),
    ("Wat is GEDEELD tussen alle 31 prompts (het skelet)?", 12, True, "1F4E78"),
    ("• Rol: SEO-introtekst / koptekst boven een categoriepagina van beslist.nl (prijsvergelijker).", 11, False, None),
    ("• Doel: oprecht behulpzame mini-koopgids, geen reclamepraat.", 11, False, None),
    ("• Aanspreekvorm 'je'; nooit 'wij/onze/ons/we'.", 11, False, None),
    ("• Nooit prijzen/bedragen/kortingen (evergreen); geen uitroeptekens.", 11, False, None),
    ("• Alleen linken naar producten uit de meegeleverde lijst, ankertekst = productnaam (3–5 woorden), "
     "geen vage ankers ('klik hier'), nooit URL's verzinnen.", 11, False, None),
    ("• Bewuste afwijking van de 'één-alinea'-regel richting een scanbare koopgids.", 11, False, None),
    ("", 11, False, None),
    ("Waarin VERSCHILLEN de prompts per maincat (zie tabblad 'Verschillen per maincat')?", 12, True, "1F4E78"),
    ("1. Genoemde subcategorieën – de voorbeelden waarop de prompt is afgestemd.", 11, False, None),
    ("2. Koopgids-onderwerpen / keuzecriteria – het hart van het verschil: welke koopvragen "
     "en meetbare criteria de tekst moet behandelen (per subcategorie).", 11, False, None),
    ("3. Vakjargon / terminologie – de kopers­woorden die de prompt expliciet oplegt.", 11, False, None),
    ("4. Structuurvorm – vrijwel overal 2–4 korte alinea’s; Films & Series houdt bewust één alinea aan.", 11, False, None),
    ("5. Richtlengte – varieert sterk: van ~120 woorden (Boeken, accessoires) tot 700–1400 woorden "
     "(Horloges, Klussen, Huishoudelijk) afhankelijk van de complexiteit van de koopbeslissing.", 11, False, None),
    ("", 11, False, None),
    (f"Bron: {SRC}", 9, False, "808080"),
    ("Gegenereerd: 2026-07-02. De volledige v3-prompt per maincat staat in de laatste kolom van tabblad 2.",
     9, False, "808080"),
]
for i, (txt, sz, bold, col) in enumerate(intro_txt, 1):
    c = ws0.cell(row=i, column=1, value=txt)
    c.font = Font(size=sz, bold=bold, color=col or "000000")
    c.alignment = Alignment(wrap_text=True, vertical="top")
ws0.column_dimensions["A"].width = 120

# --- Sheet 2: Verschillen per maincat ---
ws = wb.create_sheet("Verschillen per maincat")
cols = [
    ("Hoofdcategorie", 20),
    ("Slug", 18),
    ("Genoemde subcategorieën", 40),
    ("Structuurvorm", 26),
    ("Richtlengte", 22),
    ("Koopgids-onderwerpen (kern van het verschil)", 42),
    ("Volledige INHOUD-sectie uit de prompt", 70),
    ("Opgelegd vakjargon / terminologie", 55),
    ("Volledige v3-prompt", 100),
]
for j, (name, w) in enumerate(cols, 1):
    c = ws.cell(row=1, column=j, value=name)
    c.fill = HDR; c.font = TITLE; c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    c.border = BORDER
    ws.column_dimensions[get_column_letter(j)].width = w

for i, r in enumerate(rows, start=2):
    vals = [r["maincat"], r["slug"], r["subcats"], r["struct"], r["lengte"],
            r["themes"], r["inhoud"], r["termino"], r["prompt"]]
    for j, val in enumerate(vals, 1):
        c = ws.cell(row=i, column=j, value=val)
        c.alignment = WRAP
        c.border = BORDER
        if i % 2 == 0:
            c.fill = ALT
    ws.cell(row=i, column=1).font = Font(bold=True)

ws.freeze_panes = "C2"
ws.row_dimensions[1].height = 34
for i in range(2, len(rows) + 2):
    ws.row_dimensions[i].height = 150

wb.save(OUT)
print("Saved:", OUT)
print("Maincats:", len(rows))
EOF_MARKER = None
