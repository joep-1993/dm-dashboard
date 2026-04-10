"""
Export all unique titles with scores to Excel.
Includes: URL, title, h1_title, title_score, title_score_issue
"""
import os
import sys
from datetime import datetime

from dotenv import load_dotenv
load_dotenv()
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from backend.database import get_db_connection, return_db_connection


def export():
    conn = get_db_connection()
    cur = conn.cursor()

    print("Fetching all titles with scores...")
    cur.execute("""
        SELECT url, title, h1_title, title_score, title_score_issue
        FROM pa.unique_titles
        WHERE title IS NOT NULL AND title != ''
        ORDER BY title_score ASC NULLS LAST, url
    """)
    rows = cur.fetchall()
    cur.close()
    return_db_connection(conn)

    print(f"Fetched {len(rows):,} titles. Building Excel...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Title Scores"

    # Headers
    headers = ["URL", "Title", "H1 Title", "Score", "Issue"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        bottom=Side(style='thin', color='D9D9D9')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Color fills for scores
    score_fills = {
        1: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red
        2: PatternFill(start_color="FF3300", end_color="FF3300", fill_type="solid"),
        3: PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"),
        4: PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid"),
        5: PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid"),  # Yellow
        6: PatternFill(start_color="CCDD00", end_color="CCDD00", fill_type="solid"),
        7: PatternFill(start_color="99CC00", end_color="99CC00", fill_type="solid"),
        8: PatternFill(start_color="66BB00", end_color="66BB00", fill_type="solid"),
        9: PatternFill(start_color="33AA00", end_color="33AA00", fill_type="solid"),
        10: PatternFill(start_color="00AA00", end_color="00AA00", fill_type="solid"),  # Green
    }

    # Data rows
    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row["url"])
        ws.cell(row=i, column=2, value=row["title"])
        ws.cell(row=i, column=3, value=row["h1_title"])

        score = row["title_score"]
        score_cell = ws.cell(row=i, column=4, value=score)
        if score and score in score_fills:
            score_cell.fill = score_fills[score]
            score_cell.font = Font(bold=True, color="FFFFFF" if score <= 4 else "000000")
        score_cell.alignment = Alignment(horizontal='center')

        ws.cell(row=i, column=5, value=row["title_score_issue"])

        # Light border for readability
        for col in range(1, 6):
            ws.cell(row=i, column=col).border = thin_border

        if i % 100000 == 0:
            print(f"  Written {i:,} rows...")

    # Column widths
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 50

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f"A1:E{len(rows) + 1}"

    # Summary sheet
    ws2 = wb.create_sheet("Score Distribution")
    ws2.cell(row=1, column=1, value="Score").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Count").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=3, value="Percentage").font = header_font
    ws2.cell(row=1, column=3).fill = header_fill

    # Count per score
    score_counts = {}
    unscored = 0
    for row in rows:
        s = row["title_score"]
        if s is None:
            unscored += 1
        else:
            score_counts[s] = score_counts.get(s, 0) + 1

    total = len(rows)
    r = 2
    for score in sorted(score_counts.keys()):
        ws2.cell(row=r, column=1, value=score)
        ws2.cell(row=r, column=2, value=score_counts[score])
        ws2.cell(row=r, column=3, value=f"{score_counts[score]/total*100:.1f}%")
        if score in score_fills:
            ws2.cell(row=r, column=1).fill = score_fills[score]
        r += 1

    if unscored:
        ws2.cell(row=r, column=1, value="Not scored")
        ws2.cell(row=r, column=2, value=unscored)
        ws2.cell(row=r, column=3, value=f"{unscored/total*100:.1f}%")
        r += 1

    ws2.cell(row=r + 1, column=1, value="Total").font = Font(bold=True)
    ws2.cell(row=r + 1, column=2, value=total).font = Font(bold=True)

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12

    output_path = os.path.expanduser("~/unique_titles_scored.xlsx")
    print(f"Saving to {output_path}...")
    wb.save(output_path)
    print(f"Done! {len(rows):,} titles exported to {output_path}")


if __name__ == "__main__":
    export()
