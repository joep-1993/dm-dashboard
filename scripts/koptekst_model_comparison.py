#!/usr/bin/env python3
"""
Benchmark: dezelfde koptekst-prompt (v3, productie) op verschillende OpenAI-modellen.

Doel: kwaliteit vs. kosten vergelijken zodat je kunt kiezen welk model het meest
geschikt is voor de kopteksten. Per URL wordt de productcontext één keer opgehaald
en daarna aan ELK model identiek aangeboden — zo verschilt alleen het model.

Per call worden tokens (incl. reasoning- en cached-tokens), latency en kosten
vastgelegd. Output: Excel met samenvatting + grafieken, side-by-side kopteksten,
per-call kostendetail en de gehanteerde prijslijst.

Gebruik:
  cd /home/joepvanschagen/projects/dm-dashboard
  ./venv/bin/python scripts/koptekst_model_comparison.py [--per-maincat 1] [--limit 8]
                                                         [--models a,b,c] [--only Maincat]
                                                         [--reasoning-effort low] [--out PATH]
"""
import argparse
import json
import os
import re
import sys
import time
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))

from openai import OpenAI
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.scraper_service import scrape_product_page_api
from backend.gpt_service_v3 import (
    build_system_message_v3,
    create_product_recommendation_prompt_v3,
)
from backend.gpt_service import fix_truncated_urls

SAMPLE_PATH = os.path.join(os.path.dirname(__file__), "..", "backend", "data",
                           "koptekst_v3_benchmark_urls.json")

# ---------------------------------------------------------------- pricing ----
# USD per 1M tokens. Bron: developers.openai.com/api/docs/pricing, opgehaald 2026-08-18.
# Actualiseer deze tabel als OpenAI de prijzen wijzigt — alle kosten hieronder
# volgen hieruit en zijn verder niet hardcoded.
PRICING_DATE = "2026-08-18"
PRICING = {
    # model            input   cached-in  output
    "gpt-4o-mini":    {"in": 0.15, "cached_in": 0.075, "out": 0.60},
    "gpt-5.6-terra":  {"in": 2.00, "cached_in": 0.20,  "out": 12.00},
    "gpt-5.6-luna":   {"in": 0.20, "cached_in": 0.02,  "out": 1.20},
    "gpt-5.6-sol":    {"in": 5.00, "cached_in": 0.50,  "out": 30.00},
}

# Modellen die geen temperature accepteren en max_completion_tokens vereisen
# (reasoning-modellen: reasoning-tokens tellen mee als output-tokens).
REASONING_PREFIXES = ("gpt-5", "o1", "o3", "o4")

DEFAULT_MODELS = ["gpt-4o-mini", "gpt-5.6-terra", "gpt-5.6-luna"]

_client = None


def client() -> OpenAI:
    global _client
    if _client is None:
        _client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    return _client


def is_reasoning(model: str) -> bool:
    return model.startswith(REASONING_PREFIXES)


def usd(model: str, prompt_tokens: int, cached_tokens: int, completion_tokens: int) -> float:
    """Kosten in USD. Cached input wordt tegen het cached-tarief gerekend."""
    p = PRICING.get(model)
    if not p:
        return 0.0
    fresh_in = max(prompt_tokens - cached_tokens, 0)
    return (fresh_in * p["in"] + cached_tokens * p["cached_in"] + completion_tokens * p["out"]) / 1_000_000


def generate(model: str, h1: str, products: list, maincat: str, effort: str | None) -> dict:
    """Genereer één koptekst met het opgegeven model. Prompt = identiek aan productie-v3."""
    system_message = build_system_message_v3(maincat)
    user_prompt = create_product_recommendation_prompt_v3(h1, products)
    kwargs = {"model": model,
              "messages": [{"role": "system", "content": system_message},
                           {"role": "user", "content": user_prompt}]}
    if is_reasoning(model):
        # reasoning-tokens vallen binnen dit budget -> ruimer zetten dan de 2000 van v1/v3
        kwargs["max_completion_tokens"] = 4000
        if effort:
            kwargs["reasoning_effort"] = effort
    else:
        kwargs["max_tokens"] = 2000
        kwargs["temperature"] = 0.7

    t0 = time.time()
    resp = client().chat.completions.create(**kwargs)
    latency = time.time() - t0

    u = resp.usage
    ctd = getattr(u, "completion_tokens_details", None)
    ptd = getattr(u, "prompt_tokens_details", None)
    reasoning_tokens = getattr(ctd, "reasoning_tokens", 0) or 0
    cached_tokens = getattr(ptd, "cached_tokens", 0) or 0
    text = resp.choices[0].message.content or ""
    truncated = resp.choices[0].finish_reason == "length"
    if truncated:
        print(f"    ! afgekapt (finish_reason=length) op {model}")
    return {
        "text": fix_truncated_urls(text, products) if text else "",
        "prompt_tokens": u.prompt_tokens,
        "cached_tokens": cached_tokens,
        "completion_tokens": u.completion_tokens,
        "reasoning_tokens": reasoning_tokens,
        "latency": latency,
        "truncated": truncated,
        "cost": usd(model, u.prompt_tokens, cached_tokens, u.completion_tokens),
    }


# --------------------------------------------------------------- scoring ----
# Zelfde compliance-dimensies als scripts/koptekst_v3_comparison.py, zodat de
# uitkomsten vergelijkbaar zijn met de eerdere v1-vs-v3 benchmark.
GENERIC_WORDS = ["ideaal", "perfect", "uitstekend", "een goede keuze", "een heerlijke keuze"]
SPEC_RE = re.compile(
    r"\b\d[\d.,]*\s?(mm|cm|m²|m2|meter|kg|gram|g|liter|l|ml|db|watt|w|kwh|volt|v|ampère|ampere|a|mah|inch|\"|dpi|ip\d{2}|°c|graden|km|km/u|kelvin|k|pk|bar|mbar|karaat|tesla|teraflops|spm)\b",
    re.I,
)
BANNED_OPENINGS = [
    "bij het kiezen", "bij het selecteren", "bij het uitkiezen", "bij het overwegen",
    "bij de keuze", "bij de aanschaf", "bij de zoektocht", "het kiezen van",
    "als je op zoek", "op zoek naar", "ben je op zoek", "zoek je een",
    "wanneer je op zoek", "welkom op de",
]


def score(text: str) -> dict:
    if not text or text.startswith("["):
        return {"valid": False}
    t = text.strip()
    plain = re.sub(r"<[^>]+>", " ", t)
    low = plain.lower()
    opening = low.lstrip()[:40]
    paras = [p for p in re.split(r"\n\s*\n|<br\s*/?>|</p>|<h[1-6][^>]*>", t) if p.strip()]
    return {
        "valid": True,
        "char_count": len(t),
        "word_count": len(plain.split()),
        "n_paragraphs": max(len(paras), 1),
        "n_links": len(re.findall(r"<a\s+href=", t, re.I)),
        "has_euro": "€" in t or bool(re.search(r"\beuro\b", low)),
        "n_exclamations": plain.count("!"),
        "uses_wij_ons": bool(re.search(r"\b(wij|ons|onze|we)\b", low)),
        "contains_generic": bool([w for w in GENERIC_WORDS if w in low]),
        "banned_opening": any(opening.startswith(b) for b in BANNED_OPENINGS),
        "mentions_beslist": "beslist" in low,
        "n_measurable_specs": len(SPEC_RE.findall(t)),
        "has_measurable_spec": bool(SPEC_RE.search(t)),
        "n_questions": plain.count("?"),
    }


def summarize(rows: list, model: str) -> dict:
    """rows = [{'gen': {...}, 'score': {...}}, ...] voor één model."""
    valid = [r for r in rows if r["score"].get("valid")]
    n = len(valid)
    out = {"model": model, "n_calls": len(rows), "n_valid": n}
    if not rows:
        return out
    avg_g = lambda k: sum(r["gen"].get(k, 0) for r in rows) / len(rows)
    out.update({
        "cost_total": sum(r["gen"]["cost"] for r in rows),
        "cost_avg": avg_g("cost"),
        "cost_per_1000": avg_g("cost") * 1000,
        "in_tokens_avg": avg_g("prompt_tokens"),
        "out_tokens_avg": avg_g("completion_tokens"),
        "reasoning_tokens_avg": avg_g("reasoning_tokens"),
        "latency_avg": avg_g("latency"),
        "n_truncated": sum(1 for r in rows if r["gen"].get("truncated")),
    })
    if n:
        avg_s = lambda k: sum(r["score"].get(k, 0) for r in valid) / n
        pct = lambda k: sum(1 for r in valid if r["score"].get(k)) / n * 100.0
        out.update({
            "word_count_avg": avg_s("word_count"),
            "char_count_avg": avg_s("char_count"),
            "n_paragraphs_avg": avg_s("n_paragraphs"),
            "n_links_avg": avg_s("n_links"),
            "specs_avg": avg_s("n_measurable_specs"),
            "pct_has_spec": pct("has_measurable_spec"),
            "pct_mentions_beslist": pct("mentions_beslist"),
            "pct_banned_opening": pct("banned_opening"),
            "pct_has_euro": pct("has_euro"),
            "pct_wij_ons": pct("uses_wij_ons"),
            "pct_generic_words": pct("contains_generic"),
            "excl_avg": avg_s("n_exclamations"),
        })
    return out


# ---------------------------------------------------------------- samples ----
def load_samples(per_maincat: int, only: str | None, limit: int | None) -> list:
    with open(SAMPLE_PATH, encoding="utf-8") as f:
        data = json.load(f)
    by_mc = {}
    for s in data:
        if only and s["maincat"] != only:
            continue
        by_mc.setdefault(s["maincat"], []).append(s)
    picked = []
    for mc in sorted(by_mc):
        rows = sorted(by_mc[mc], key=lambda x: -x.get("visits", 0))[:per_maincat]
        picked.extend(rows)
    picked.sort(key=lambda x: -x.get("visits", 0))
    return picked[:limit] if limit else picked


# ------------------------------------------------------------------ excel ----
HEAD_FILL = PatternFill("solid", fgColor="5E4A90")


def _head(ws, row, headers, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w


def write_excel(out_path, models, results, summaries, effort, volume):
    wb = Workbook()

    # --- Samenvatting -------------------------------------------------------
    ws = wb.active
    ws.title = "Samenvatting"
    ws["A1"] = "Koptekst-modellen: kwaliteit vs. kosten"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (f"Prompt: v3 productie (per-maincat) — identiek voor alle modellen. "
                f"URLs: {len(results)}. reasoning_effort: {effort or 'default'}. "
                f"Prijzen per {PRICING_DATE}.")
    ws["A2"].alignment = Alignment(wrap_text=True)

    rows = [
        ("Kosten per koptekst (USD)", "cost_avg", "0.000000"),
        (f"Kosten per {volume:,} kopteksten (USD)".replace(",", "."), None, "0.00"),
        ("Totale kosten deze test (USD)", "cost_total", "0.0000"),
        ("Input-tokens (gem.)", "in_tokens_avg", "0"),
        ("Output-tokens (gem.)", "out_tokens_avg", "0"),
        ("  waarvan reasoning-tokens (gem.)", "reasoning_tokens_avg", "0"),
        ("Latency in sec (gem.)", "latency_avg", "0.0"),
        ("Afgekapte responses", "n_truncated", "0"),
        ("", None, None),
        ("Woorden (gem.)", "word_count_avg", "0"),
        ("Tekens (gem.)", "char_count_avg", "0"),
        ("Alinea's (gem.)", "n_paragraphs_avg", "0.0"),
        ("Productlinks (gem.)", "n_links_avg", "0.0"),
        ("Meetbare specs (gem.)", "specs_avg", "0.0"),
        ("% met meetbare spec", "pct_has_spec", "0"),
        ("% noemt Beslist", "pct_mentions_beslist", "0"),
        ("", None, None),
        ("% verboden opening (lager=beter)", "pct_banned_opening", "0"),
        ("% met prijs/euro (lager=beter)", "pct_has_euro", "0"),
        ("% wij/ons/onze (lager=beter)", "pct_wij_ons", "0"),
        ("% generieke woorden (lager=beter)", "pct_generic_words", "0"),
        ("Uitroeptekens (gem., lager=beter)", "excl_avg", "0.0"),
    ]
    _head(ws, 4, ["Metric"] + models, [34] + [16] * len(models))
    r = 5
    for label, key, fmt in rows:
        if label == "":
            r += 1
            continue
        ws.cell(row=r, column=1, value=label).font = Font(bold=label.startswith("Kosten"))
        for c, m in enumerate(models, 2):
            s = summaries[m]
            val = s.get("cost_avg", 0) * volume if key is None else s.get(key)
            cell = ws.cell(row=r, column=c, value=val)
            if fmt:
                cell.number_format = fmt
        r += 1

    # grafiek: kosten per volume
    chart = BarChart()
    chart.title = f"Kosten per {volume} kopteksten (USD)"
    chart.y_axis.title = "USD"
    chart.height, chart.width = 7, 12
    data = Reference(ws, min_col=2, max_col=1 + len(models), min_row=6, max_row=6)
    cats = Reference(ws, min_col=2, max_col=1 + len(models), min_row=4, max_row=4)
    chart.add_data(data, from_rows=True, titles_from_data=False)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, f"A{r + 2}")

    # grafiek: latency
    lat_row = 5 + [i for i, (l, _, _) in enumerate(rows) if l.startswith("Latency")][0]
    chart2 = BarChart()
    chart2.title = "Gemiddelde latency (sec)"
    chart2.y_axis.title = "seconden"
    chart2.height, chart2.width = 7, 12
    chart2.add_data(Reference(ws, min_col=2, max_col=1 + len(models), min_row=lat_row, max_row=lat_row),
                    from_rows=True, titles_from_data=False)
    chart2.set_categories(cats)
    chart2.legend = None
    ws.add_chart(chart2, f"J{r + 2}")

    # grafiek: kwaliteitssignalen
    chart3 = BarChart()
    chart3.title = "Kwaliteitssignalen (%)"
    chart3.height, chart3.width = 8, 20
    q_labels = ["% met meetbare spec", "% noemt Beslist", "% verboden opening",
                "% met prijs/euro", "% wij/ons/onze", "% generieke woorden"]
    q_rows = [5 + i for i, (l, _, _) in enumerate(rows)
              if any(l.startswith(q) for q in q_labels)]
    if q_rows:
        chart3.add_data(Reference(ws, min_col=1, max_col=1 + len(models),
                                  min_row=min(q_rows), max_row=max(q_rows)),
                        from_rows=True, titles_from_data=True)
        chart3.set_categories(cats)
        ws.add_chart(chart3, f"A{r + 20}")

    # --- Kopteksten side-by-side -------------------------------------------
    ws2 = wb.create_sheet("Kopteksten")
    _head(ws2, 1, ["#", "Maincat", "H1", "URL"] + models,
          [5, 16, 26, 46] + [80] * len(models))
    for i, res in enumerate(results, 2):
        ws2.cell(row=i, column=1, value=i - 1)
        ws2.cell(row=i, column=2, value=res["maincat"])
        ws2.cell(row=i, column=3, value=res["h1"])
        ws2.cell(row=i, column=4, value=res["url"])
        for c, m in enumerate(models, 5):
            ws2.cell(row=i, column=c, value=res["by_model"][m]["gen"]["text"] or "[leeg]")
        for c in range(1, 5 + len(models)):
            ws2.cell(row=i, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[i].height = 300
    ws2.row_dimensions[1].height = 30
    ws2.freeze_panes = "E2"

    # --- Kostendetail per call ---------------------------------------------
    ws3 = wb.create_sheet("Kosten detail")
    _head(ws3, 1, ["#", "Maincat", "H1", "Model", "Input-tokens", "Cached", "Output-tokens",
                   "Reasoning-tokens", "Woorden", "Latency (s)", "Kosten (USD)"],
          [5, 16, 26, 16, 12, 10, 13, 15, 10, 11, 13])
    r = 2
    for i, res in enumerate(results, 1):
        for m in models:
            g = res["by_model"][m]["gen"]
            s = res["by_model"][m]["score"]
            vals = [i, res["maincat"], res["h1"], m, g["prompt_tokens"], g["cached_tokens"],
                    g["completion_tokens"], g["reasoning_tokens"], s.get("word_count", 0),
                    round(g["latency"], 2), g["cost"]]
            for c, v in enumerate(vals, 1):
                cell = ws3.cell(row=r, column=c, value=v)
                if c == 11:
                    cell.number_format = "0.000000"
            r += 1
    ws3.freeze_panes = "A2"

    # --- Prijzen ------------------------------------------------------------
    ws4 = wb.create_sheet("Prijzen")
    _head(ws4, 1, ["Model", "Input $/1M", "Cached input $/1M", "Output $/1M"], [20, 14, 18, 14])
    for r, m in enumerate(models, 2):
        p = PRICING.get(m, {})
        ws4.cell(row=r, column=1, value=m)
        ws4.cell(row=r, column=2, value=p.get("in"))
        ws4.cell(row=r, column=3, value=p.get("cached_in"))
        ws4.cell(row=r, column=4, value=p.get("out"))
    ws4.cell(row=len(models) + 3, column=1,
             value=f"Bron: developers.openai.com/api/docs/pricing, opgehaald {PRICING_DATE}. "
                   f"Standaardtarief (geen batch/flex).")

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    for attempt in range(3):
        try:
            wb.save(out_path)
            print(f"\nWrote: {out_path}")
            return out_path
        except PermissionError:
            print(f"  save locked, retry {attempt + 1}")
            time.sleep(1)
    alt = out_path.replace(".xlsx", f"_{int(time.time())}.xlsx")
    wb.save(alt)
    print(f"\nWrote: {alt}")
    return alt


# ------------------------------------------------------------------- main ----
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--models", default=",".join(DEFAULT_MODELS),
                    help="komma-gescheiden model-ids")
    ap.add_argument("--per-maincat", type=int, default=1)
    ap.add_argument("--limit", type=int, default=8, help="max aantal URLs (0 = alles)")
    ap.add_argument("--only", default=None, help="beperk tot één maincat")
    ap.add_argument("--reasoning-effort", default=None,
                    choices=["none", "low", "medium", "high"],
                    help="alleen voor gpt-5.x/o-modellen; default = model-default")
    ap.add_argument("--volume", type=int, default=10000,
                    help="volume voor de kosten-extrapolatie (default 10.000 kopteksten)")
    ap.add_argument("--out", default=f"/mnt/c/Users/JoepvanSchagen/Downloads/claude/"
                                     f"koptekst_model_vergelijking_{date.today()}.xlsx")
    args = ap.parse_args()

    models = [m.strip() for m in args.models.split(",") if m.strip()]
    unknown = [m for m in models if m not in PRICING]
    if unknown:
        print(f"! Geen prijs bekend voor: {', '.join(unknown)} — kosten worden 0. "
              f"Vul PRICING aan bovenin dit script.")

    samples = load_samples(args.per_maincat, args.only, args.limit or None)
    print(f"{len(samples)} URLs x {len(models)} modellen = {len(samples) * len(models)} calls")
    print(f"Modellen: {', '.join(models)}")

    results = []
    for i, s in enumerate(samples, 1):
        url, mc = s["url"], s["maincat"]
        print(f"[{i}/{len(samples)}] {mc} :: {url[:70]}")
        try:
            scraped = scrape_product_page_api(url)
        except Exception as e:
            print(f"  scrape error: {e}")
            scraped = None
        h1 = (scraped or {}).get("h1_title") or s.get("h1_hint") or "Onbekend"
        products = (scraped or {}).get("products", [])
        if not products:
            print("  geen producten -> skip")
            continue
        by_model = {}
        for m in models:
            try:
                gen = generate(m, h1, products, mc, args.reasoning_effort)
                print(f"    {m:16s} {gen['completion_tokens']:5d} out "
                      f"({gen['reasoning_tokens']} reasoning)  "
                      f"{gen['latency']:5.1f}s  ${gen['cost']:.6f}")
            except Exception as e:
                print(f"    {m:16s} ERROR: {e}")
                gen = {"text": f"[{m} error: {e}]", "prompt_tokens": 0, "cached_tokens": 0,
                       "completion_tokens": 0, "reasoning_tokens": 0, "latency": 0.0,
                       "truncated": False, "cost": 0.0}
            by_model[m] = {"gen": gen, "score": score(gen["text"])}
        results.append({"maincat": mc, "url": url, "h1": h1,
                        "n_products": len(products), "by_model": by_model})

    if not results:
        print("Geen resultaten (geen producten gevonden).")
        return

    summaries = {m: summarize([r["by_model"][m] for r in results], m) for m in models}

    print(f"\n{'':34s}" + "".join(f"{m:>18s}" for m in models))
    for label, key, fmt in [
        ("Kosten/koptekst (USD)", "cost_avg", "{:.6f}"),
        (f"Kosten/{args.volume} kopteksten", "_vol", "${:.2f}"),
        ("Output-tokens (gem.)", "out_tokens_avg", "{:.0f}"),
        ("  waarvan reasoning", "reasoning_tokens_avg", "{:.0f}"),
        ("Latency (s)", "latency_avg", "{:.1f}"),
        ("Woorden (gem.)", "word_count_avg", "{:.0f}"),
        ("Meetbare specs (gem.)", "specs_avg", "{:.1f}"),
        ("% noemt Beslist", "pct_mentions_beslist", "{:.0f}%"),
        ("% verboden opening", "pct_banned_opening", "{:.0f}%"),
        ("% prijs/euro genoemd", "pct_has_euro", "{:.0f}%"),
    ]:
        line = f"{label:34s}"
        for m in models:
            v = (summaries[m].get("cost_avg", 0) * args.volume) if key == "_vol" \
                else summaries[m].get(key, 0)
            line += f"{fmt.format(v):>18s}"
        print(line)

    write_excel(args.out, models, results, summaries, args.reasoning_effort, args.volume)


if __name__ == "__main__":
    main()
