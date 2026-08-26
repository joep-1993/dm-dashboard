# -*- coding: utf-8 -*-
"""
Stap 3 van /findfacetvalues: maandelijks zoekvolume ophalen en naar Excel schrijven.

Leest de geschifte kandidaten uit keep.txt (één per regel; alles na een '#' is commentaar)
en schrijft <workdir>/<bestandsnaam>.xlsx.

    python volumes.py --workdir DIR [--min-volume 50] [--out PAD]

Waarom batches van 500 en retries: GenerateKeywordHistoricalMetrics geeft bij grote
batches niet voor elk keyword een rij terug, zonder error. Een ontbrekende rij is dan
niet te onderscheiden van "echt 0". Daarom wordt per batch bijgehouden wat wél terugkwam
en worden ontbrekende keywords herhaald opgevraagd tot alles een waarde heeft.
"""
import argparse
import json
import os
import sys

REPO = "/home/joepvanschagen/projects/dm-dashboard"
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(REPO, "backend"))

from dotenv import load_dotenv
load_dotenv(os.path.join(REPO, ".env"))

from common import kp_key  # noqa: E402
import keyword_planner_service as kp  # noqa: E402

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

PURPLE = "5E4A90"


def get_volumes(keywords, log=print):
    """{keyword -> avg monthly searches}. Kleine batches + retries op ontbrekende rijen."""
    keys = sorted({kp_key(k) for k in keywords if kp.validate_keyword(kp_key(k))})
    vol, absent = {}, list(keys)
    client = kp._get_client()
    ci = 0
    for rnd, batch_size in enumerate([500, 40, 40, 40, 20]):
        if not absent:
            break
        todo, still = absent, []
        for i in range(0, len(todo), batch_size):
            batch = todo[i:i + batch_size]
            while ci < len(kp.CUSTOMER_IDS):
                res = kp._query_search_volumes(client, batch, kp.CUSTOMER_IDS[ci])
                if res is None:          # quota op deze customer_id -> volgende
                    ci += 1
                    continue
                for k in batch:
                    (vol.__setitem__(k, res[k]) if k in res else still.append(k))
                break
            else:
                log("  ALLE customer_ids op — resterende keywords zonder waarde.")
                still.extend(batch)
                return vol, still
        log(f"  ronde {rnd + 1}: {len(todo) - len(still)} waarden binnen, {len(still)} nog te doen")
        absent = still
    return vol, absent


def write_excel(path, rows, meta, min_volume):
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor=PURPLE)
    thin = Side(style="thin", color="D6CFE8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = [
        ("Facetwaarde", lambda r: r["kandidaat"], 34),
        ("Zoekvolume p/m (NL)", lambda r: int(r["zoekvolume"]), 15),
        ("Producten met term in titel", lambda r: r["producten_titel"], 13),
        ("Producten met term in beschrijving", lambda r: r["producten_desc"], 14),
        ("Producten totaal", lambda r: r["producten_totaal"], 11),
        ("Bestaat al in deze categorie", lambda r: "ja" if r["bestaat_al_in_categorie"] else "", 13),
        ("Bekend als facetwaarde elders", lambda r: "ja" if r["bekend_als_facetwaarde_elders"] else "", 13),
        ("Voorbeeldproduct", lambda r: (r.get("voorbeeld") or "")[:120], 60),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{meta['facet_gevraagd']}-waarden"[:31]
    ws.append([c[0] for c in cols])
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i)
        cell.font, cell.fill, cell.border = hdr_font, hdr_fill, border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = c[2]
    for r in rows:
        ws.append([c[1](r) for c in cols])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(cols)):
        for cell in row:
            cell.border = border
            if isinstance(cell.value, int):
                cell.alignment = Alignment(horizontal="right")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    ws.row_dimensions[1].height = 32

    cat = meta["categorie"]
    ws2 = wb.create_sheet("Toelichting")
    lines = [
        (f"Kandidaat-facetwaarden voor '{meta['facet_gevraagd']}' — {cat['category_name']}", True),
        ("", False),
        (f"Categorie: {cat['category_name']} — categorie-ID {cat['category_id']} (urlName {cat['url_name']})", False),
        (f"Hoofdcategorie: {cat['main_category_name']} ({cat['main_category_id']})", False),
        (f"URL: {cat['url']}", False),
        (f"Peildatum: {meta['datum']}", False),
        ("", False),
        ("Facet", True),
        ((f"'{meta['facet_gevraagd']}' zit al op deze categorie (facet-ID {meta['facet_op_categorie']['facet_id']}) "
          f"met {len(meta['facet_bestaande_waarden'])} waarden."
          if meta.get("facet_op_categorie") else
          f"'{meta['facet_gevraagd']}' zit nog NIET op deze categorie."), False),
        (("Bestaande waarden: " + ", ".join(meta["facet_bestaande_waarden"][:40]))
         if meta["facet_bestaande_waarden"] else "Bestaande waarden: (geen)", False),
        (f"Elders in de taxonomie heeft dit facet {len(meta['facet_referentiewaarden'])} bestaande waarden; "
         "die zijn als vormreferentie gebruikt.", False),
        ("", False),
        ("Bron van de kandidaten", True),
        (f"{meta['producten_opgehaald']} producten opgehaald via de Search API, "
         f"waarvan {meta['producten_met_beschrijving']} met productbeschrijving.", False),
        ((f"De categorie heeft er {meta['producten_in_categorie']}; er is dus een steekproef gebruikt."
          if meta.get("producten_in_categorie") and meta["producten_in_categorie"] > meta["producten_opgehaald"]
          else "Alle producten in de categorie zijn meegenomen."), False),
        ("Uit titel + beschrijving zijn woordcombinaties geëxtraheerd (stopwoorden, maten en generieke "
         "marketingtaal eruit), waarna de lijst handmatig is geschift op het gevraagde facetsoort.", False),
        (f"Drempel: term komt voor in >={meta['drempels']['min_title']} producttitels of "
         f">={meta['drempels']['min_desc']} beschrijvingen.", False),
        (f"{meta['n_beoordeeld']} kandidaten beoordeeld, {meta['n_gehouden']} gehouden als "
         f"'{meta['facet_gevraagd']}'-waarde, {len(rows)} daarvan met >{min_volume} zoekvolume per maand.", False),
        ("", False),
        ("Zoekvolume", True),
        ("Google Ads Keyword Planner (GenerateKeywordHistoricalMetrics) — gemiddeld maandelijks zoekvolume,", False),
        ("geo Nederland (2528), taal Nederlands (1010), netwerk Google Search.", False),
        ("Opgevraagd in batches van 500 met retries: de API laat in grote batches willekeurig rijen weg,", False),
        ("en zo'n ontbrekende rij is niet te onderscheiden van een echte 0.", False),
        ("Volumes zijn Keyword Planner-buckets (10, 20, 40, 70, 90, 110, 140, ...), geen exacte aantallen.", False),
    ]
    if meta.get("zonder_waarde"):
        lines.append((f"Zonder waarde gebleven: {', '.join(meta['zonder_waarde'][:20])}", False))
    for i, (txt, bold) in enumerate(lines, 1):
        c = ws2.cell(row=i, column=1, value=txt)
        c.font = Font(bold=bold, size=12 if bold and i == 1 else 11,
                      color=PURPLE if bold else "000000")
    ws2.column_dimensions["A"].width = 125
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workdir", required=True)
    ap.add_argument("--min-volume", type=int, default=50)
    ap.add_argument("--out", default=None)
    ap.add_argument("--datum", default=None)
    a = ap.parse_args()

    scan = json.load(open(os.path.join(a.workdir, "scan.json"), encoding="utf-8"))
    keep_path = os.path.join(a.workdir, "keep.txt")
    if not os.path.exists(keep_path):
        raise SystemExit(f"{keep_path} ontbreekt — schift eerst de kandidaten.")
    keep = []
    for line in open(keep_path, encoding="utf-8"):
        line = line.split("#")[0].strip()
        if line:
            keep.append(line)
    if not keep:
        raise SystemExit("keep.txt is leeg.")

    by_term = {r["kandidaat"]: r for r in scan["kandidaten"]}
    rows, unknown = [], []
    for k in keep:
        if k not in by_term:
            unknown.append(k)
        r = dict(by_term.get(k) or {"kandidaat": k, "producten_titel": 0, "producten_desc": 0,
                                    "producten_totaal": 0, "bestaat_al_in_categorie": False,
                                    "bekend_als_facetwaarde_elders": False})
        r["kandidaat"] = k
        rows.append(r)
    if unknown:
        print(f"LET OP: {len(unknown)} termen uit keep.txt komen niet in de scan voor en krijgen "
              f"producttelling 0: {unknown[:10]}")

    print(f"{len(rows)} geschifte kandidaten; zoekvolume ophalen ...")
    vol, absent = get_volumes([r["kandidaat"] for r in rows])
    for r in rows:
        r["zoekvolume"] = vol.get(kp_key(r["kandidaat"]), 0) or 0

    hits = [r for r in rows if r["zoekvolume"] > a.min_volume]
    hits.sort(key=lambda r: (-r["zoekvolume"], -r["producten_titel"]))

    from datetime import date
    meta = dict(scan)
    meta["datum"] = a.datum or date.today().strftime("%d-%m-%Y")
    meta["n_beoordeeld"] = len(scan["kandidaten"])
    meta["n_gehouden"] = len(rows)
    meta["zonder_waarde"] = absent

    cat = scan["categorie"]
    stamp = date.today().strftime("%Y%m%d")
    default = os.path.join(
        "/mnt/c/Users/JoepvanSchagen/Downloads/claude",
        f"{cat['category_name'].replace(' ', '_')}_{scan['facet_gevraagd'].replace(' ', '_')}"
        f"_kandidaten_{stamp}.xlsx")
    out = a.out or default
    write_excel(out, hits, meta, a.min_volume)

    print(f"\n{len(hits)} van de {len(rows)} kandidaten hebben >{a.min_volume} zoekvolume.")
    if absent:
        print(f"LET OP: {len(absent)} keywords kregen geen waarde terug: {absent[:10]}")
    print(f"Excel: {out}")
    for r in hits[:15]:
        print(f"  {r['zoekvolume']:>7}  {r['kandidaat']}")


if __name__ == "__main__":
    main()
