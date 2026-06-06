#!/usr/bin/env python3
"""
Reverse-templatize Unique Titles (pa.unique_titles_content) into
beslist.tblPageTitles blueprint rows.

Per page:
  URL  -> leaf category slug + ordered [(facet_type, value_id), ...]
  slug -> tax v2 cat_id (9xxxxxx)  [/tmp/slug2id.json]
  value_id -> label forms (name_on_detail / name_in_column) from
              taxonomy.`nl-nl_FacetValue`
  Replace each literal facet value (longest form first) in the rendered
  title/h1/description with !!facet_type!!, and the category name with
  !!sub_category!! / !!sub_category_lower!!.

Aggregates one representative template per (cat_id, key) and reports a
QA flag when a facet on the page could not be located in the copy.

Usage:
  pagetitles_from_unique.py sample [N]   # print before/after examples + stats
  pagetitles_from_unique.py build        # full run -> Excel
"""
import os, re, sys, json, html
from collections import defaultdict, Counter
import pymysql, psycopg2

CREDS = {}
for line in open(os.path.expanduser('~/.mysql-creds')):
    if '=' in line:
        k, v = line.strip().split('=', 1); CREDS[k] = v
PG_DSN = "postgresql://dbadmin:Q9fGRKtUdvdtxsiCM12HeFe0Nki0PvmjZRFLZ9ArmlWdMnDQXX8SdxKnPniqGmq6@10.1.32.9:5432/n8n-vector-db"

SIC = re.compile(r'<!--\s*SIC:?\s*(.*?)\s*-->', re.I)
SOD = re.compile(r'<!--\s*SOD:?\s*(.*?)\s*-->', re.I)

def my():
    return pymysql.connect(host=CREDS['MYSQL_HOST'], user=CREDS['MYSQL_USER'],
                           password=CREDS['MYSQL_PASSWORD'], db='beslist', charset='utf8mb4')

def clean(s):
    if not s: return ''
    s = html.unescape(s)
    # if SIC/SOD markup present, drop it (we read the dedicated columns instead)
    s = SIC.sub('', s); s = SOD.sub('', s)
    s = re.sub(r'<!--.*?-->', '', s)   # strip any remaining HTML comments (e.g. <!--3-->)
    return s.strip()

def load_value_labels():
    """value_id -> ordered list of distinct candidate label strings (longest first)."""
    conn = my(); c = conn.cursor()
    c.execute("SELECT facet_value_id, name_on_detail, name_in_column, name, caption "
              "FROM taxonomy.`nl-nl_FacetValue`")
    labels = {}
    for vid, sod, sic, name, cap in c.fetchall():
        cands = []
        for raw in (sod, sic, clean(name), clean(cap)):
            r = (raw or '').strip()
            if r and r not in cands:
                cands.append(r)
        cands.sort(key=len, reverse=True)
        labels[vid] = cands
    conn.close()
    return labels

import unicodedata
STOP = {'met','en','de','het','een','van','voor','op','in','te','set','online','kopen',
        'nieuw','nieuwe','goedkoop','beste','aanbieding','aanbiedingen','full','los'}

from functools import lru_cache

_DBL = re.compile(r'(aa|ee|oo|uu)([bcdfgklmnprstz])$', re.I)

def dutch_inflect(w):
    """Attributive -e inflection forms of a Dutch adjective base.
    Verstelbaar->verstelbare, Groot->grote, Sportief->sportieve,
    Lichtgrijs->lichtgrijze, Stijlvol->stijlvolle, Groen->groene."""
    low = w.lower(); forms = set()
    # double vowel + final consonant -> single vowel + (voiced) consonant + e
    m = _DBL.search(low)
    if m:
        cons = {'s': 'z', 'f': 'v'}.get(m.group(2), m.group(2))
        forms.add(w[:m.start()] + m.group(1)[0] + cons + 'e')
    if low.endswith('f'): forms.add(w[:-1] + 've')      # sportief->sportieve
    if low.endswith('s'): forms.add(w[:-1] + 'ze')      # grijs->grijze
    # consonant doubling after a single short vowel (vol->volle, dun->dunne)
    m2 = re.search(r'[^aeiou]([aeiou])([bdfgklmnprst])$', low)
    if m2 and not m:
        forms.add(w + m2.group(2) + 'e')
    forms.add(w + 'e')                                   # plain (groen->groene)
    return forms

@lru_cache(maxsize=200000)
def variants(label):
    """Dutch-tolerant surface variants of a label for matching (longest first).
    Plural/singular + adjectival inflection. Cached: labels recur across pages."""
    l = label.strip()
    low = l.lower()
    if   low.endswith('en'): stem = l[:-2]
    elif low.endswith('e'):  stem = l[:-1]
    elif low.endswith('s'):  stem = l[:-1]
    else:                    stem = l
    out = set()
    for base in {l, stem}:
        if base:
            out |= {base, base + 'e', base + 'en', base + 's'}
            out |= dutch_inflect(base)
    return tuple(sorted({v for v in out if v}, key=len, reverse=True))

def flex_src(s):
    """Regex source for a label, flexible on internal whitespace and hyphens.
    '128 GB' also matches '128GB';  '4-deurs' also matches '4 deurs'."""
    esc = re.escape(s.strip())
    esc = re.sub(r'(?:\\ )+', r'\\s*', esc)   # escaped spaces -> \s*
    esc = esc.replace('\\-', r'[\s\-]*')       # hyphen -> optional space/hyphen
    return esc

@lru_cache(maxsize=400000)
def word_re(s):
    # word-ish boundaries that tolerate punctuation in values like "35,5", "Van Lier"
    return re.compile(r'(?<![\w])' + flex_src(s) + r'(?![\w])', re.IGNORECASE)

@lru_cache(maxsize=20000)
def category_sources(cat_name):
    """Ordered regex sources to locate the category name in a title, tolerant of
    how it is written in curated copy:
      - 'Eten & drinken'  -> 'Eten en Drinken'  (& as en)
      - 'Cadeaus & gadgets' -> 'Cadeaus gadgets' (& dropped)
      - 'Cases & hoesjes' -> 'cases'             (single constituent word)
    Most specific (full phrase) first, then single words."""
    name = cat_name.strip()
    parts = [p for p in re.split(r'\s*&\s*|\s+', name) if p and p.lower() != 'en']
    srcs = []
    if len(parts) >= 2:
        joiner = r'\s*(?:&|en)?\s*'                      # space / 'en' / '&' interchangeable
        def alt(p):                                     # per-word variant alternation
            return '(?:' + '|'.join(flex_src(v) for v in variants(p)) + ')'
        srcs.append(joiner.join(alt(p) for p in parts)) # full phrase, plural/singular tolerant
        for p in sorted(parts, key=len, reverse=True):  # then any significant word
            if len(p) >= 4:
                for v in variants(p):
                    srcs.append(flex_src(v))
    else:
        # single-word category: keep inflection-aware variants
        for v in variants(name):
            srcs.append(flex_src(v))
    return tuple(srcs)

_PLACE   = re.compile(r'!![^!]+!!')
_SENT    = re.compile('[\ue000-\uf8ff]')
_NOTSENT = re.compile('[^\ue000-\uf8ff]+')   # runs without a placeholder sentinel

def templatize(text, facet_labels, cat_name, positional=False, satisfied=frozenset()):
    """facet_labels: ordered dict {placeholder: [candidate label strings]}.
    Returns (templated_text, matched_phs, fuzzy_phs).
    Existing !!...!! placeholders are masked to digit-free sentinels so that
    no label/category match can corrupt the inside of a placeholder
    (e.g. facet type 'maat_(helmen)' must not have 'helmen' re-wrapped)."""
    if not text:
        return text, set(), set()
    store = []
    def stash(ph):
        store.append(ph)
        return chr(0xe000 + len(store) - 1)   # opaque single-char sentinel
    # mask placeholders already present in the source (e.g. !!DISCOUNT!!)
    work = _PLACE.sub(lambda m: stash(m.group(0)), text)
    matched = set(); fuzzy = set()

    def place_first(rx, ph):
        nonlocal work
        m = rx.search(work)
        if not m:
            return False
        work = work[:m.start()] + stash(ph) + work[m.end():]
        return True

    # ---- pass 1: exact/flexible label matching, longest label first ----
    flat = [(lab, ph) for ph, labs in facet_labels.items() for lab in labs]
    for label, ph in sorted(flat, key=lambda x: len(x[0]), reverse=True):
        if ph in matched:
            continue
        for var in variants(label):
            if place_first(word_re(var), ph):
                matched.add(ph); break

    # ---- category name -> !!sub_category!! / !!sub_category_lower!! ----
    # case-insensitive; the matched text's case decides which placeholder
    if cat_name:
        for src in category_sources(cat_name):
            rx = re.compile(r'(?<![\w])' + src + r'(?![\w])', re.IGNORECASE)
            m = rx.search(work)
            if m:
                ph = '!!sub_category!!' if m.group(0)[:1].isupper() else '!!sub_category_lower!!'
                work = work[:m.start()] + stash(ph) + work[m.end():]
                matched.add(ph); break

    # ---- pass 2: elimination — target only the facets still unaddressed ----
    for ph, labs in facet_labels.items():
        if ph in matched:
            continue
        hit = False
        # 2a. numeric-core: a number in the label, possibly glued to a unit in the copy
        for lab in labs:
            m = re.search(r'\d+(?:[.,]\d+)?', lab)
            if not m:
                continue
            num = re.escape(m.group(0))
            unit = re.search(r'\d[.,\d]*\s*([A-Za-z]{1,5})', lab)
            unit_src = r'\s*' + re.escape(unit.group(1)) if unit else r'\s*[A-Za-z]{0,4}'
            rx = re.compile(r'(?<![\w.,])' + num + unit_src + r'(?![\w])', re.IGNORECASE)
            if place_first(rx, ph):
                matched.add(ph); fuzzy.add(ph); hit = True; break
        if hit:
            continue
        # 2b. distinctive-word: a salient word from the label present verbatim in the copy
        words = sorted({w for lab in labs for w in re.findall(r"[A-Za-zÀ-ſ]{4,}", lab)
                        if w.lower() not in STOP}, key=len, reverse=True)
        for w in words:
            if place_first(re.compile(r'(?<![\w])' + flex_src(w) + r'(?![\w])', re.IGNORECASE), ph):
                matched.add(ph); fuzzy.add(ph); break

    # ---- pass 3: positional elimination (title only, last resort) ----
    # House rule: everything before 'kopen' is a facet/category/connector. So if
    # the title subject has exactly ONE contiguous literal chunk and exactly ONE
    # facet is still unaddressed, that chunk must be that facet (covers values
    # with no/mismatched label, e.g. grote_maten_mode 'XXXL').
    if positional:
        # Candidate targets for the leftover = facets not matched in the title,
        # plus the category slot ('CAT') if the category name wasn't matched
        # either (covers head-noun/hypernym cases: 'stoelen' for Eetkamerstoelen).
        # Disambiguate via facets satisfied elsewhere (h1/description). Bind only
        # when the target is unique — one candidate total, or one not-elsewhere.
        SUBCATS = {'!!sub_category!!', '!!sub_category_lower!!'}
        candidates = [ph for ph in facet_labels if ph not in matched]
        if cat_name and not (matched & SUBCATS):
            candidates = candidates + ['CAT']
        not_elsewhere = [x for x in candidates
                         if (x not in satisfied) and not (x == 'CAT' and (satisfied & SUBCATS))]
        target = (candidates[0] if len(candidates) == 1
                  else not_elsewhere[0] if len(not_elsewhere) == 1 else None)
        if target:
            cut = work.lower().find('kopen')
            subj_end = cut if cut != -1 else len(work)
            runs = []   # (start, end) of contiguous literal content within the subject
            for seg in _NOTSENT.finditer(work[:subj_end]):
                toks = list(re.finditer(r'\S+', seg.group(0)))
                content = [t for t in toks
                           if re.sub(r'[^0-9A-Za-zÀ-ſ]', '', t.group(0))
                           and re.sub(r'[^0-9A-Za-zÀ-ſ]', '', t.group(0)).lower() not in CONNECT]
                if content:
                    runs.append((seg.start() + content[0].start(),
                                 seg.start() + content[-1].end()))
            if len(runs) == 1:
                s, e = runs[0]
                if target == 'CAT':
                    chunk = work[s:e].lstrip()
                    ph = '!!sub_category!!' if chunk[:1].isupper() else '!!sub_category_lower!!'
                else:
                    ph = target
                work = work[:s] + stash(ph) + work[e:]
                matched.add(ph); fuzzy.add(ph)

    # unmask
    out = _SENT.sub(lambda m: store[ord(m.group(0)) - 0xe000], work)
    return out, matched, fuzzy

def parse_url(url):
    if '/c/' not in url:
        return None
    path, fstr = url.split('/c/', 1)
    segs = [s for s in path.split('/') if s]
    leaf = segs[-1] if segs else ''
    pairs = []
    for pair in fstr.split('~~'):
        bits = pair.split('~')
        if len(bits) >= 2 and bits[0]:
            try: pairs.append((bits[0], int(bits[1])))
            except ValueError: pass
    return leaf, pairs

def main():
    mode = sys.argv[1] if len(sys.argv) > 1 else 'sample'
    N = int(sys.argv[2]) if len(sys.argv) > 2 else 25

    slug2id = json.load(open('/tmp/slug2id.json'))
    id2name = {int(k): v for k, v in json.load(open('/tmp/id2name.json')).items()}
    print(f"[load] slug2id={len(slug2id)} id2name={len(id2name)}", file=sys.stderr)
    labels = load_value_labels()
    print(f"[load] value labels={len(labels)}", file=sys.stderr)

    # existing combos to detect NEW vs existing
    conn = my(); c = conn.cursor()
    c.execute("SELECT cat_id,`key` FROM tblPageTitles WHERE country_code='NL'")
    existing = set((r[0], r[1]) for r in c.fetchall())
    conn.close()

    pg = psycopg2.connect(PG_DSN)
    pc = pg.cursor(name='stream'); pc.itersize = 20000
    pc.execute("SELECT u.url, c.title, c.h1_title, c.description "
               "FROM pa.unique_titles_content c JOIN pa.urls u ON u.url_id=c.url_id "
               "WHERE c.title IS NOT NULL AND c.title<>''")

    # aggregate per (cat_id,key): representative template + counts + QA
    agg = defaultdict(lambda: {'titles': Counter(), 'h1s': Counter(), 'descs': Counter(),
                               'n': 0, 'unmatched_pages': 0, 'fuzzy_pages': 0})
    examples = []
    scanned = 0
    for url, title, h1, desc in pc:
        scanned += 1
        p = parse_url(url)
        if not p: continue
        leaf, pairs = p
        cat = slug2id.get(leaf)
        if cat is None or not pairs: continue
        cat_name = id2name.get(cat, '')
        key = '~'.join(sorted({t for t, _ in pairs}))
        facet_labels = {}            # placeholder -> candidate labels (ordered)
        page_value_phs = set()
        for ftype, vid in pairs:
            ph = f'!!{ftype}!!'
            page_value_phs.add(ph)
            facet_labels.setdefault(ph, [])
            for lab in labels.get(vid, []):
                if lab not in facet_labels[ph]:
                    facet_labels[ph].append(lab)
        # h1/description first, so the title's positional pass knows which facets
        # are already satisfied elsewhere (and thus not the title's leftover)
        h_out, m2, f2 = templatize(h1, facet_labels, cat_name)
        d_out, m3, f3 = templatize(desc, facet_labels, cat_name)
        t_out, m1, f1 = templatize(title, facet_labels, cat_name,
                                   positional=True, satisfied=(m2 | m3))
        matched = m1 | m2 | m3
        # QA: did every facet on the page get represented somewhere?
        unmatched = page_value_phs - matched
        a = agg[(cat, key)]
        a['titles'][t_out] += 1; a['h1s'][h_out] += 1; a['descs'][d_out] += 1
        a['n'] += 1
        if unmatched: a['unmatched_pages'] += 1
        if (f1 | f2 | f3): a['fuzzy_pages'] += 1
        if mode == 'sample' and len(examples) < N and unmatched != page_value_phs:
            examples.append((url, cat, cat_name, key, title, t_out, h1, h_out, sorted(unmatched)))
    pg.close()

    if mode == 'sample':
        print(f"\n=== scanned {scanned} pages, {len(agg)} combos ===\n")
        for url, cat, cn, key, t0, t1, h0, h1v, un in examples[:N]:
            print(f"cat={cat} ({cn})  key={key}")
            print(f"  URL   : {url}")
            print(f"  title0: {t0}")
            print(f"  title→: {t1}")
            print(f"  h1   0: {h0}")
            print(f"  h1   →: {h1v}")
            if un: print(f"  ⚠ unmatched facets: {un}")
            print()
        # stats
        tot = sum(a['n'] for a in agg.values())
        new = {k: a for k, a in agg.items() if k not in existing}
        clean_combos = sum(1 for a in new.values() if a['unmatched_pages'] == 0)
        print(f"NEW combos: {len(new)} | fully-matched combos: {clean_combos} "
              f"({100*clean_combos/max(len(new),1):.1f}%)")
    elif mode == 'excel':
        import pickle
        agg2 = pickle.load(open('/tmp/agg.pkl', 'rb'))
        build_excel(agg2, existing, id2name)
    else:
        import pickle
        pickle.dump(dict(agg), open('/tmp/agg.pkl', 'wb'))   # cache for instant re-export
        build_excel(agg, existing, id2name)

# connector words that legitimately sit between facets in the subject
CONNECT = {'met','en','voor','de','het','een','van','op','in','te','online','of','tot',
           '&','+','je','uw','der','den','ter','aan','bij',
           # fixed dimensional descriptor words in curated copy (not facets)
           'maat','hoog','breed','diep','lang','lange','hoge','brede','diepe',
           'lage','laag','hoogte','breedte','diepte','lengte','cm','mm','x',
           # fixed marketing adjectives in curated copy (not facets)
           'goedkope','goedkoop','goedkoopste','voordelige','voordelig','voordeligste',
           'mooie','mooiste','beste','artikelen','exclusieve','exclusief','luxe'}

def residual_before_kopen(title):
    """Per the house rule, all facets sit before 'kopen' in the title.
    Return the literal (non-placeholder, non-connector) words left in that
    subject segment — each is an un-templatized facet worth review."""
    t = title
    low = t.lower()
    cut = low.find('kopen')
    if cut == -1:
        cut = low.find('?')
    if cut == -1:
        cut = low.find('|')
    subj = t[:cut] if cut != -1 else t
    subj = re.sub(r'!![^!]+!!', ' ', subj)          # drop placeholders
    words = re.findall(r"[A-Za-zÀ-ſ0-9]+(?:[.,][0-9]+)?", subj)
    # length-1 tokens are apostrophe/model fragments ("BH's"->'s'), never facets
    return [w for w in words if len(w) > 1 and w.lower() not in CONNECT]

def build_excel(agg, existing, id2name):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'new_pagetitles'
    hdr = ['cat_id', 'cat_name', 'key', 'title', 'h1_title', 'description',
           'country_code', 'support_pages', 'qa_unmatched_pct', 'fuzzy_pct',
           'missing_key_facets', 'covers_all_key_facets', 'residual_before_kopen', 'review']
    ws.append(hdr)
    for cell in ws[1]:
        cell.font = Font(bold=True); cell.fill = PatternFill('solid', fgColor='DDDDDD')
    rows = 0
    for (cat, key), a in sorted(agg.items(), key=lambda x: -x[1]['n']):
        if (cat, key) in existing:      # only fill gaps
            continue
        title = a['titles'].most_common(1)[0][0] or ''
        h1 = a['h1s'].most_common(1)[0][0] or ''
        desc = a['descs'].most_common(1)[0][0] or ''
        # house style: titles begin with the !!current_query!! token
        if not title.startswith('!!current_query!!'):
            title = '!!current_query!! ' + title
        qa = 100 * a['unmatched_pages'] / max(a['n'], 1)
        fz = 100 * a['fuzzy_pages'] / max(a['n'], 1)
        # which key facets are absent across ALL fields (title + h1 + description)?
        blob = title + ' ' + h1 + ' ' + desc
        missing = [t for t in key.split('~') if t and f'!!{t}!!' not in blob]
        covers = 'no' if missing else 'yes'
        residual = residual_before_kopen(title)
        # authoritative correctness signal: no un-templatized literal terms in the
        # title subject (before 'kopen?'). missing/qa are informational only.
        review = 'CHECK' if residual else ''
        ws.append([cat, id2name.get(cat, ''), key, title, h1, desc, 'NL',
                   a['n'], round(qa, 1), round(fz, 1),
                   ','.join(missing), covers, ' '.join(residual), review])
        rows += 1
    base = '/mnt/c/Users/JoepvanSchagen/Downloads/claude/tblPageTitles_new_from_unique'
    os.makedirs(os.path.dirname(base), exist_ok=True)
    out = base + '.xlsx'
    try:
        wb.save(out)
    except PermissionError:
        out = base + '_v2.xlsx'          # original is open/locked in Excel
        wb.save(out)
    print(f"wrote {rows} rows -> {out}")

if __name__ == '__main__':
    main()
