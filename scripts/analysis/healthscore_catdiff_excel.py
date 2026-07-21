#!/usr/bin/env python3
"""Build a comparison Excel (old Healthscore HS1.0 vs new HS2.0) for the 10
validation categories, from the per-URL diff CSV produced by
healthscore_catdiff.py.

CSV columns: cat_id,cat_name,npath,sample_url,in_hs1,in_hs2,status,june_visits,june_revenue
  status in {kept, added, dropped, uncovered}
    kept      = in both HS1 and HS2
    added     = HS2 only  (coverage HS2.0 gains)
    dropped   = HS1 only  (URLs HS2.0 removes)
    uncovered = in neither (remaining opportunity)

HS2.0 config for this run: all-channel knee + climatology, 1-month forward-max
look-ahead, seasonal per-category caps, score = 0.89*pct(log SEO visits) +
0.11*pct(log SEO revenue) within category. Holdout = complete June 2026 (real
SEO visits/revenue), predictor = 90d before June (leakage-free).
"""
import csv
import sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV_IN = "/mnt/c/Users/JoepvanSchagen/Downloads/claude/hs2_catdiff_seasonal_v2.csv"
XLSX_OUT = "/mnt/c/Users/JoepvanSchagen/Downloads/claude/HS2.0_vs_HS1.0_vergelijking.xlsx"

CAT_ORDER = [
    ("9000047", "Stoelen"), ("9000066", "Eetkamerstoelen"),
    ("9000608", "Sneakers"), ("9000953", "Voer"),
    ("9002072", "Douchewanden"), ("9005282", "Mobiele telefoons"),
    ("9005317", "Airconditionings"), ("9001646", "Dekbedovertrekken"),
    ("9003581", "Grasmaaiers"), ("9000668", "Shirts"),
]

# ---- palette (avoid grey labels per house style) ----
NAVY = "1F3864"
BLUE = "2E5AAC"
GREEN = "C6EFCE"
GREEN_TX = "1E6B34"
RED = "FFC7CE"
RED_TX = "9C0006"
YELLOW = "FFF2CC"
HEAD_TX = "FFFFFF"
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def tf(x):
    return str(x).strip().lower() in ("true", "t", "1")


def load(fn):
    rows = []
    with open(fn, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            r["june_visits"] = int(r.get("june_visits") or 0)
            r["june_revenue"] = float(r.get("june_revenue") or 0)
            r["_in1"] = tf(r.get("in_hs1"))
            r["_in2"] = tf(r.get("in_hs2"))
            rows.append(r)
    return rows


def style_header(ws, row, ncols, fill=NAVY):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=HEAD_TX, size=11)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autowidth(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build():
    rows = load(CSV_IN)
    by_cat = defaultdict(list)
    for r in rows:
        by_cat[r["cat_id"]].append(r)

    wb = Workbook()

    # ================= Sheet 1: Samenvatting =================
    ws = wb.active
    ws.title = "Samenvatting"
    ws["A1"] = "HS2.0 vs oude Healthscore (HS1.0) — 10 testcategorieën"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws["A2"] = ("Holdout: juni 2026 (echte SEO-bezoeken/omzet). HS2.0 = all-channel caps + "
                "1-maand look-ahead, seasonal caps, score 0.89·bezoeken + 0.11·omzet.")
    ws["A2"].font = Font(italic=True, size=9, color=BLUE)

    hdr = ["cat_id", "Categorie", "HS1 URLs", "HS2 URLs", "Δ URLs", "Δ URLs %",
           "behouden", "toegevoegd", "verwijderd",
           "vis-dekking HS1 %", "vis-dekking HS2 %", "Δ vis pp",
           "omzet-dekking HS1 %", "omzet-dekking HS2 %", "Δ omzet pp",
           "juni bezoeken (tot)", "juni omzet € (tot)"]
    hrow = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(row=hrow, column=i, value=h)
    style_header(ws, hrow, len(hdr))
    ws.freeze_panes = f"A{hrow+1}"

    tot = defaultdict(float)
    r0 = hrow + 1
    rownum = r0
    for cid, nm in CAT_ORDER:
        crows = by_cat.get(cid, [])
        tv = sum(x["june_visits"] for x in crows)
        tr = sum(x["june_revenue"] for x in crows)
        n1 = sum(1 for x in crows if x["_in1"])
        n2 = sum(1 for x in crows if x["_in2"])
        kept = sum(1 for x in crows if x["_in1"] and x["_in2"])
        added = sum(1 for x in crows if x["_in2"] and not x["_in1"])
        dropped = sum(1 for x in crows if x["_in1"] and not x["_in2"])
        v1 = sum(x["june_visits"] for x in crows if x["_in1"])
        v2 = sum(x["june_visits"] for x in crows if x["_in2"])
        rv1 = sum(x["june_revenue"] for x in crows if x["_in1"])
        rv2 = sum(x["june_revenue"] for x in crows if x["_in2"])
        vc1 = 100 * v1 / tv if tv else 0
        vc2 = 100 * v2 / tv if tv else 0
        rc1 = 100 * rv1 / tr if tr else 0
        rc2 = 100 * rv2 / tr if tr else 0
        vals = [int(cid), nm, n1, n2, n2 - n1, (100 * (n2 - n1) / n1 if n1 else 0),
                kept, added, dropped,
                vc1, vc2, vc2 - vc1, rc1, rc2, rc2 - rc1, tv, tr]
        for i, v in enumerate(vals, 1):
            ws.cell(row=rownum, column=i, value=v)
        # accumulate weighted totals
        tot["n1"] += n1; tot["n2"] += n2; tot["kept"] += kept
        tot["added"] += added; tot["dropped"] += dropped
        tot["tv"] += tv; tot["tr"] += tr
        tot["v1"] += v1; tot["v2"] += v2; tot["rv1"] += rv1; tot["rv2"] += rv2
        rownum += 1

    # total row
    tvc1 = 100 * tot["v1"] / tot["tv"] if tot["tv"] else 0
    tvc2 = 100 * tot["v2"] / tot["tv"] if tot["tv"] else 0
    trc1 = 100 * tot["rv1"] / tot["tr"] if tot["tr"] else 0
    trc2 = 100 * tot["rv2"] / tot["tr"] if tot["tr"] else 0
    tvals = ["", "TOTAAL (10 cats)", int(tot["n1"]), int(tot["n2"]),
             int(tot["n2"] - tot["n1"]),
             (100 * (tot["n2"] - tot["n1"]) / tot["n1"] if tot["n1"] else 0),
             int(tot["kept"]), int(tot["added"]), int(tot["dropped"]),
             tvc1, tvc2, tvc2 - tvc1, trc1, trc2, trc2 - trc1,
             int(tot["tv"]), tot["tr"]]
    for i, v in enumerate(tvals, 1):
        cell = ws.cell(row=rownum, column=i, value=v)
        cell.font = Font(bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=YELLOW)

    # number formats + delta coloring
    last = rownum
    for rr in range(r0, last + 1):
        for col in (6, 10, 11, 12, 13, 14, 15):  # percents / pp
            ws.cell(row=rr, column=col).number_format = "0.0"
        ws.cell(row=rr, column=16).number_format = "#,##0"
        ws.cell(row=rr, column=17).number_format = "€ #,##0"
        for col in (3, 4, 5, 7, 8, 9):
            ws.cell(row=rr, column=col).number_format = "#,##0"
        # color the pp-delta columns
        for col in (12, 15):
            c = ws.cell(row=rr, column=col)
            if c.value is None:
                continue
            if c.value > 0.05:
                c.fill = PatternFill("solid", fgColor=GREEN); c.font = Font(color=GREEN_TX, bold=(rr == last))
            elif c.value < -0.05:
                c.fill = PatternFill("solid", fgColor=RED); c.font = Font(color=RED_TX, bold=(rr == last))
        for rr2 in [rr]:
            for col in range(1, len(hdr) + 1):
                ws.cell(row=rr2, column=col).border = BORDER
    autowidth(ws, [9, 20, 9, 9, 8, 9, 10, 11, 11, 12, 12, 9, 13, 13, 10, 14, 14])

    # ================= Sheet 2: Analyse =================
    wa = wb.create_sheet("Analyse")
    added_all = [x for x in rows if x["_in2"] and not x["_in1"]]
    dropped_all = [x for x in rows if x["_in1"] and not x["_in2"]]
    dropped_dead = [x for x in dropped_all if x["june_visits"] == 0]
    dropped_traf = [x for x in dropped_all if x["june_visits"] > 0]
    added_traf = [x for x in added_all if x["june_visits"] > 0]
    add_vis = sum(x["june_visits"] for x in added_all)
    add_rev = sum(x["june_revenue"] for x in added_all)
    drop_vis = sum(x["june_visits"] for x in dropped_all)
    drop_rev = sum(x["june_revenue"] for x in dropped_all)

    lines = [
        ("HS2.0 vs oude Healthscore — analyse (10 testcategorieën, juni 2026 holdout)", 14, NAVY, True),
        ("", 10, "000000", False),
        ("Kernconclusie", 12, BLUE, True),
        (f"HS2.0 dekt MEER SEO-verkeer met MINDER URLs. Van de {int(tot['n1']):,} URLs in de oude set "
         f"verwijdert HS2.0 er {len(dropped_all):,} en voegt er {len(added_all):,} nieuwe toe → "
         f"{int(tot['n2']):,} URLs totaal.", 10, "000000", False),
        (f"Bezoekdekking stijgt van {tvc1:.1f}% naar {tvc2:.1f}% ({tvc2-tvc1:+.1f}pp); "
         f"omzetdekking van {trc1:.1f}% naar {trc2:.1f}% ({trc2-trc1:+.1f}pp).", 10, "000000", False),
        ("", 10, "000000", False),
        ("Waarom de verwijderingen géén verlies zijn", 12, BLUE, True),
        (f"Van de {len(dropped_all):,} verwijderde URLs hadden er {len(dropped_dead):,} "
         f"({100*len(dropped_dead)/len(dropped_all):.0f}%) NUL bezoeken in juni — pure dode ballast.", 10, "000000", False),
        (f"Slechts {len(dropped_traf):,} verwijderde URLs hadden nog verkeer, samen {drop_vis:,.0f} bezoeken "
         f"(€{drop_rev:,.0f}). Dit is de eerlijke 'churn' — deels teruggevangen door de nieuwe-URL-bak in de volledige set.", 10, "000000", False),
        ("", 10, "000000", False),
        ("Wat HS2.0 wint", 12, BLUE, True),
        (f"De {len(added_all):,} toegevoegde URLs brengen {add_vis:,.0f} juni-bezoeken (€{add_rev:,.0f}) binnen "
         f"die de oude set liet liggen; {len(added_traf):,} ervan hebben aantoonbaar verkeer.", 10, "000000", False),
        (f"Netto verkeersruil: +{add_vis:,.0f} toegevoegd vs −{drop_vis:,.0f} verwijderd = "
         f"{add_vis-drop_vis:+,.0f} bezoeken netto op deze 10 categorieën.", 10, "000000", False),
        ("", 10, "000000", False),
        ("Bekende aandachtspunten (per categorie)", 12, BLUE, True),
        ("• Sneakers −6pp: GEEN cap-probleem maar holdout-artefact. Mrt–mei (voorspelvenster) is Sneakers' "
         "seizoensdal; er zijn simpelweg minder kandidaat-URLs dan de cap toelaat (availability-bound). "
         "Echte zomerruns zullen dit dichten.", 10, "000000", False),
        ("• Douchewanden omzet −33pp: kleine-categorie-ruis. De visit-gedreven score laat 1–2 losse, hoog-€ "
         "pagina's vallen; bezoekdekking is daar juist +31pp.", 10, "000000", False),
        ("• Eetkamerstoelen omzet −7pp: zelfde visit-led effect, bezoekdekking +16pp.", 10, "000000", False),
        ("", 10, "000000", False),
        ("Effect van de twee nieuwe cap-wijzigingen", 12, BLUE, True),
        ("• All-channel caps (i.p.v. SEO-only): grotere base-caps voor Stoelen/Eetkamerstoelen/Dekbed/Voer → "
         "+0,7pp bezoek op deze set, maar +13% URL-footprint.", 10, "000000", False),
        ("• 1-maand look-ahead: in juni ~geen effect omdat deze seizoenscats in juni al op/over hun piek zitten. "
         "Het effect zit in de aanloopmaand (bv. mei voor Airco); een mei-holdout is nodig om dit te meten.", 10, "000000", False),
    ]
    rr = 1
    for text, size, color, bold in lines:
        c = wa.cell(row=rr, column=1, value=text)
        c.font = Font(size=size, color=color, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1
    wa.column_dimensions["A"].width = 120

    # ================= Sheets 3 & 4: Top toegevoegd / verwijderd =================
    def url_sheet(title, data, note):
        wsx = wb.create_sheet(title)
        wsx.cell(row=1, column=1, value=note).font = Font(italic=True, color=BLUE, size=9)
        hdr2 = ["cat_id", "Categorie", "juni bezoeken", "juni omzet €", "voorbeeld-URL"]
        for i, h in enumerate(hdr2, 1):
            wsx.cell(row=2, column=i, value=h)
        style_header(wsx, 2, len(hdr2))
        wsx.freeze_panes = "A3"
        data = sorted(data, key=lambda x: (-x["june_visits"], -x["june_revenue"]))
        for j, x in enumerate(data, 3):
            wsx.cell(row=j, column=1, value=int(x["cat_id"]))
            wsx.cell(row=j, column=2, value=x["cat_name"])
            wsx.cell(row=j, column=3, value=x["june_visits"]).number_format = "#,##0"
            wsx.cell(row=j, column=4, value=round(x["june_revenue"], 2)).number_format = "€ #,##0.00"
            wsx.cell(row=j, column=5, value=x.get("sample_url") or x.get("npath"))
        autowidth(wsx, [9, 20, 14, 14, 110])
        return wsx

    url_sheet("Toegevoegd (HS2.0 wint)",
              added_all,
              "URLs die HS2.0 toevoegt en de oude set miste — gesorteerd op juni-bezoeken (de dekkingswinst).")
    url_sheet("Verwijderd (met verkeer)",
              dropped_traf,
              "Verwijderde URLs DIE NOG VERKEER HADDEN — de eerlijke churn. (De dode ballast met 0 bezoeken staat niet hier.)")

    # ================= Sheet 5: Volledige detail =================
    wsd = wb.create_sheet("Volledig detail")
    hdr3 = ["cat_id", "Categorie", "status", "in HS1", "in HS2", "juni bezoeken", "juni omzet €", "URL"]
    for i, h in enumerate(hdr3, 1):
        wsd.cell(row=1, column=i, value=h)
    style_header(wsd, 1, len(hdr3))
    wsd.freeze_panes = "A2"
    order_idx = {cid: i for i, (cid, _) in enumerate(CAT_ORDER)}
    st_rank = {"added": 0, "kept": 1, "dropped": 2, "uncovered": 3}
    srt = sorted(rows, key=lambda x: (order_idx.get(x["cat_id"], 99),
                                      st_rank.get(x["status"], 9),
                                      -x["june_visits"]))
    for j, x in enumerate(srt, 2):
        wsd.cell(row=j, column=1, value=int(x["cat_id"]))
        wsd.cell(row=j, column=2, value=x["cat_name"])
        sc = wsd.cell(row=j, column=3, value=x["status"])
        if x["status"] == "added":
            sc.fill = PatternFill("solid", fgColor=GREEN); sc.font = Font(color=GREEN_TX)
        elif x["status"] == "dropped":
            sc.fill = PatternFill("solid", fgColor=RED); sc.font = Font(color=RED_TX)
        wsd.cell(row=j, column=4, value="ja" if x["_in1"] else "")
        wsd.cell(row=j, column=5, value="ja" if x["_in2"] else "")
        wsd.cell(row=j, column=6, value=x["june_visits"]).number_format = "#,##0"
        wsd.cell(row=j, column=7, value=round(x["june_revenue"], 2)).number_format = "€ #,##0.00"
        wsd.cell(row=j, column=8, value=x.get("sample_url") or x.get("npath"))
    autowidth(wsd, [9, 20, 11, 8, 8, 14, 14, 110])
    wsd.auto_filter.ref = f"A1:H{len(srt)+1}"

    wb.save(XLSX_OUT)
    print(f"Wrote {XLSX_OUT}")
    print(f"  categories: {len(CAT_ORDER)}  | HS1 {int(tot['n1']):,} URLs -> HS2 {int(tot['n2']):,} URLs")
    print(f"  vis coverage {tvc1:.1f}% -> {tvc2:.1f}% ({tvc2-tvc1:+.1f}pp)")
    print(f"  rev coverage {trc1:.1f}% -> {trc2:.1f}% ({trc2-trc1:+.1f}pp)")
    print(f"  dropped {len(dropped_all):,} ({len(dropped_dead):,} dead / {len(dropped_traf):,} w-traffic) | added {len(added_all):,}")


if __name__ == "__main__":
    build()
