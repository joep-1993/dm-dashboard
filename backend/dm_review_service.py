"""
DM Review data writer (slide 2 — Omzet SEO / Cijfers).

Pulls fresh data from Redshift and appends it to the SharePoint-synced workbook
"review_dm_seo.xlsx" (in joep's personal OneDrive). Four feeds:

  - `visits_omzet`       monthly  visits + omzet  SEO + DMA organic
  - `visits_omzet_dag`   daily    visits + omzet  SEO + DMA organic
  - `serp`               monthly  avg position    per URL type (Cat/C/PLP/R)
  - `serp_device`        daily    avg position    DESKTOP / MOBILE (NL)

The behavior is incremental: for each tab the writer finds the latest date
already in the column, then asks Redshift for everything strictly after it.
"""

import logging
import os
from copy import copy
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell import Cell

from backend.database import get_redshift_connection, return_redshift_connection

logger = logging.getLogger(__name__)

EXCEL_PATH = "/mnt/c/Users/JoepvanSchagen/OneDrive - Beslist.nl BV/review_dm_seo.xlsx"
PPTX_PATH = "/mnt/c/Users/JoepvanSchagen/OneDrive - Beslist.nl BV/DM review_NEW.pptx"
PPTX_SLIDE_INDEX = 1  # 0-based, so slide 2

SHEET_MONTHLY = "visits_omzet"
SHEET_DAILY = "visits_omzet_dag"
SHEET_SERP_TYPE = "serp"
SHEET_SERP_DEVICE = "serp_device"

CHANNELS = ("SEO", "DMA organic")
DEVICES = ("DESKTOP", "MOBILE")
URL_TYPES_SQL = ("Browse-url zonder /r/ en /c/", "C-url", "PLP", "R-url")
URL_TYPE_DISPLAY = {
    "Browse-url zonder /r/ en /c/": "Cat-url",
    "C-url": "C-url",
    "PLP": "PLP",
    "R-url": "R-url",
}

DUTCH_MONTHS = [
    "Januari", "Februari", "Maart", "April", "Mei", "Juni",
    "Juli", "Augustus", "September", "Oktober", "November", "December",
]

# Lookback windows: how far back to refresh on each run. Covers data settling
# (affiliate_revenue lands within ~30d; GSC backfills ~3d) so re-runs pick up
# corrections. Rows older than the lookback are left untouched.
LOOKBACK_DAYS_DAILY = 60        # daily tabs
LOOKBACK_MONTHS_MONTHLY = 3     # monthly visits_omzet tab


# ---------------------------------------------------------------------------
# Excel helpers (copy-style writer matches performance_standup pattern)
# ---------------------------------------------------------------------------

def _check_file_lock() -> Optional[str]:
    d, base = os.path.split(EXCEL_PATH)
    lock = os.path.join(d, "~$" + base)
    if os.path.exists(lock):
        return (f"The Excel file is currently open. Please close "
                f"'{base}' in Excel and try again.")
    if not os.path.exists(EXCEL_PATH):
        return f"Excel file not found at {EXCEL_PATH}"
    return None


def _write_cell(sheet, row: int, col: int, value) -> Cell:
    """Write value and inherit formatting from the cell above (so appended rows
    look like the prior rows: same fill / font / number format)."""
    cell = sheet.cell(row=row, column=col, value=value)
    if row > 1:
        src = sheet.cell(row=row - 1, column=col)
        if src.has_style:
            cell.font = copy(src.font)
            cell.fill = copy(src.fill)
            cell.border = copy(src.border)
            cell.alignment = copy(src.alignment)
            cell.protection = copy(src.protection)
            cell.number_format = src.number_format
    return cell


def _latest_date_in_col(sheet, date_col: int = 1, header_rows: int = 1) -> Optional[date]:
    latest: Optional[date] = None
    for r in range(header_rows + 1, sheet.max_row + 1):
        v = sheet.cell(row=r, column=date_col).value
        d = v.date() if isinstance(v, datetime) else (v if isinstance(v, date) else None)
        if d and (latest is None or d > latest):
            latest = d
    return latest


def _first_empty_row(sheet, date_col: int = 1, header_rows: int = 1) -> int:
    for r in range(header_rows + 1, sheet.max_row + 2):
        if sheet.cell(row=r, column=date_col).value in (None, ""):
            return r
    return sheet.max_row + 1


def _build_row_index(sheet, date_col: int, key_col: int,
                     header_rows: int = 1) -> Dict[Tuple[date, str], int]:
    """Index existing rows by (date, key_value) so we can UPSERT in place.
    key_col is the column whose value distinguishes rows that share a date
    (e.g. kanaal or device)."""
    idx: Dict[Tuple[date, str], int] = {}
    for r in range(header_rows + 1, sheet.max_row + 1):
        v = sheet.cell(row=r, column=date_col).value
        k = sheet.cell(row=r, column=key_col).value
        d = v.date() if isinstance(v, datetime) else (v if isinstance(v, date) else None)
        if d and isinstance(k, str):
            idx[(d, k)] = r
    return idx


def _months_back(today: date, n: int) -> int:
    """Return yyyymm that is `n` whole months before `today`."""
    m = today.month - n
    y = today.year
    while m <= 0:
        m += 12
        y -= 1
    return y * 100 + m


def _yyyymm_anchor_date(yyyymm: int, today: date) -> date:
    """Anchor date for the lookback windows when processing a given month.

    Uses the last day of that month, but never later than today — so processing
    the current (most-recent) month behaves exactly like an un-parameterized run,
    while processing an older month (e.g. May while it's June) anchors the
    windows to that month instead of "now"."""
    y, m = divmod(yyyymm, 100)
    first_next = date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)
    last_day = first_next - timedelta(days=1)
    return min(last_day, today)


# ---------------------------------------------------------------------------
# Redshift queries
# ---------------------------------------------------------------------------

_OMZET_EXPR = ("SUM(fv.cpc_revenue + fv.ww_revenue "
               "+ COALESCE(fv.affiliate_revenue, 0))")


def _fetch_monthly_channel(conn, start_yyyymm: int) -> List[Dict]:
    """Monthly visits + omzet for SEO + DMA organic, from start_yyyymm inclusive."""
    sql = f"""
        SELECT  (fv.dim_date_key / 100)::int        AS yyyymm,
                c.marketing_channel                 AS kanaal,
                COUNT(*)                            AS visits,
                {_OMZET_EXPR}                       AS omzet
        FROM    datamart.fct_visits fv
        JOIN    datamart.dim_visit  dv ON fv.dim_visit_key = dv.dim_visit_key
        JOIN    chan_deriv.ref_channel_derivation_stats c
                  ON dv.aff_id = c.aff_id AND dv.channel_id = c.channel_id
        WHERE   fv.dim_date_key >= %s
          AND   c.marketing_channel IN ('SEO','DMA organic')
          AND   dv.is_real_visit = 1
        GROUP BY 1, 2
        ORDER BY 1, 2
    """
    start_key = start_yyyymm * 100 + 1
    with conn.cursor() as cur:
        cur.execute(sql, (start_key,))
        return [dict(r) for r in cur.fetchall()]


def _fetch_daily_channel(conn, after_dk: int) -> List[Dict]:
    """Daily visits + omzet for SEO + DMA organic, dim_date_key > after_dk."""
    sql = f"""
        SELECT  fv.dim_date_key                     AS dk,
                c.marketing_channel                 AS kanaal,
                COUNT(*)                            AS visits,
                {_OMZET_EXPR}                       AS omzet
        FROM    datamart.fct_visits fv
        JOIN    datamart.dim_visit  dv ON fv.dim_visit_key = dv.dim_visit_key
        JOIN    chan_deriv.ref_channel_derivation_stats c
                  ON dv.aff_id = c.aff_id AND dv.channel_id = c.channel_id
        WHERE   fv.dim_date_key > %s
          AND   c.marketing_channel IN ('SEO','DMA organic')
          AND   dv.is_real_visit = 1
        GROUP BY 1, 2
        ORDER BY 1, 2
    """
    with conn.cursor() as cur:
        cur.execute(sql, (after_dk,))
        return [dict(r) for r in cur.fetchall()]


def _fetch_monthly_serp_by_url_type(conn, yyyymm: int) -> Dict[str, float]:
    """Impression-weighted avg_position per URL type for a single yyyymm."""
    sql = """
        SELECT  type_url,
                SUM(avg_position * impressions)::float
                  / NULLIF(SUM(impressions), 0)     AS weighted_pos
        FROM    bt.search_console
        WHERE   country = 'nld'
          AND   deleted_ind = 0
          AND   dim_date_key / 100 = %s
          AND   type_url IN ('Browse-url zonder /r/ en /c/','C-url','PLP','R-url')
        GROUP BY type_url
    """
    with conn.cursor() as cur:
        cur.execute(sql, (yyyymm,))
        return {r["type_url"]: float(r["weighted_pos"]) for r in cur.fetchall()
                if r["weighted_pos"] is not None}


def _fetch_daily_serp_by_device(conn, after_dk: int) -> List[Dict]:
    """Per-day impression-weighted avg_position for DESKTOP + MOBILE, NL only."""
    sql = """
        SELECT  dim_date_key                        AS dk,
                device,
                SUM(avg_position * impressions)::float
                  / NULLIF(SUM(impressions), 0)     AS weighted_pos
        FROM    bt.search_console
        WHERE   country = 'nld'
          AND   deleted_ind = 0
          AND   device IN ('DESKTOP','MOBILE')
          AND   dim_date_key > %s
        GROUP BY 1, 2
        ORDER BY 1, 2
    """
    with conn.cursor() as cur:
        cur.execute(sql, (after_dk,))
        return [dict(r) for r in cur.fetchall()]


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _yyyymm_to_first_of_month(yyyymm: int) -> datetime:
    return datetime(yyyymm // 100, yyyymm % 100, 1)


def _dk_to_date(dk: int) -> datetime:
    return datetime.strptime(str(dk), "%Y%m%d")


def _write_monthly_channel(sheet, rows: List[Dict]) -> Tuple[int, int]:
    """UPSERT monthly rows. Returns (updated, appended)."""
    if not rows:
        return 0, 0

    # Existing rows are keyed by (first-of-month-date, kanaal)
    idx = _build_row_index(sheet, date_col=1, key_col=4)
    next_row = _first_empty_row(sheet)
    updated = appended = 0
    by_month: Dict[int, Dict[str, Dict]] = {}
    for r in rows:
        by_month.setdefault(r["yyyymm"], {})[r["kanaal"]] = r

    for yyyymm in sorted(by_month):
        month_dt = _yyyymm_to_first_of_month(yyyymm).date()
        for kanaal in CHANNELS:
            data = by_month[yyyymm].get(kanaal)
            if not data:
                continue
            existing_row = idx.get((month_dt, kanaal))
            row = existing_row if existing_row else next_row
            _write_cell(sheet, row, 1, _yyyymm_to_first_of_month(yyyymm))
            _write_cell(sheet, row, 2, int(data["visits"]))
            _write_cell(sheet, row, 3, float(data["omzet"] or 0))
            _write_cell(sheet, row, 4, kanaal)
            if existing_row:
                updated += 1
            else:
                idx[(month_dt, kanaal)] = next_row
                next_row += 1
                appended += 1
    return updated, appended


def _write_daily_channel(sheet, rows: List[Dict]) -> Tuple[int, int]:
    """UPSERT daily rows. Returns (updated, appended)."""
    if not rows:
        return 0, 0

    idx = _build_row_index(sheet, date_col=1, key_col=4)
    next_row = _first_empty_row(sheet)
    updated = appended = 0
    by_day: Dict[int, Dict[str, Dict]] = {}
    for r in rows:
        by_day.setdefault(r["dk"], {})[r["kanaal"]] = r

    for dk in sorted(by_day):
        d_dt = _dk_to_date(dk)
        key_date = d_dt.date()
        for kanaal in CHANNELS:
            data = by_day[dk].get(kanaal)
            if not data:
                continue
            existing_row = idx.get((key_date, kanaal))
            row = existing_row if existing_row else next_row
            _write_cell(sheet, row, 1, d_dt)
            _write_cell(sheet, row, 2, int(data["visits"]))
            _write_cell(sheet, row, 3, float(data["omzet"] or 0))
            _write_cell(sheet, row, 4, kanaal)
            if existing_row:
                updated += 1
            else:
                idx[(key_date, kanaal)] = next_row
                next_row += 1
                appended += 1
    return updated, appended


def _last_data_row(sheet, date_col: int = 1) -> int:
    """Row of the last non-empty date cell."""
    last = 1
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=date_col).value is not None:
            last = r
    return last


# Pivot tables on these tabs source from columns A:D (A:C for serp_device).
# Other tabs (top_3_10_*, etc.) are out of scope for the dm_review refresh.
_PIVOT_SOURCE_COLS = {
    "visits_omzet":     "A:D",
    "visits_omzet_dag": "A:D",
    "serp_device":      "A:C",
}


def _extend_pivot_sources(wb) -> List[str]:
    """For each pivot whose source is one of the dm_review feed sheets, extend
    its `cacheSource.worksheetSource.ref` to cover the current data range and
    flip `refreshOnLoad = True`. Returns a list of human-readable change notes."""
    notes: List[str] = []
    for sn, cols in _PIVOT_SOURCE_COLS.items():
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        last_row = _last_data_row(ws)
        if last_row < 2:
            continue
        col_first, col_last = cols.split(":")
        new_ref = f"{col_first}1:{col_last}{last_row}"
        for pt in (getattr(ws, "_pivots", None) or []):
            cd = pt.cache
            src = getattr(cd, "cacheSource", None)
            wsr = getattr(src, "worksheetSource", None) if src else None
            if wsr is None:
                continue
            old_ref = wsr.ref
            wsr.ref = new_ref
            cd.refreshOnLoad = True
            notes.append(f"{sn}: {pt.name} {old_ref} → {new_ref}")
    return notes


def _roll_pivot_filter_window(wb) -> List[str]:
    """For each pivot on the dm_review feed sheets, if its first pivotField has
    a rolling filter (some items hidden), preserve the window size and slide
    it forward to include the most-recent N items. No-op for pivots that show
    all items.

    Without this, newly-cached date items come in with `h=True` (hidden) and
    the rolling-12-months chart never sees the new month.
    """
    from openpyxl.xml.functions import tostring
    import re

    notes: List[str] = []
    for sn in _PIVOT_SOURCE_COLS:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for pt in (getattr(ws, "_pivots", None) or []):
            cd = pt.cache
            if not cd.cacheFields:
                continue
            si = cd.cacheFields[0].sharedItems
            xml = tostring(si.to_tree()).decode()
            dates = re.findall(r'<d v="([^"]+)"/>', xml)
            if not dates:
                continue  # not a date-keyed pivot
            pf = pt.pivotFields[0]
            data_items = [it for it in pf.items if it.t in (None, "data")]
            window = sum(1 for it in data_items if it.h is not True)
            if window == 0 or window >= len(data_items):
                continue  # no rolling filter configured
            # Pair each data item with its cached date and pick the window most-recent
            items_dated = [(it, dates[it.x] if it.x is not None and it.x < len(dates) else "")
                           for it in data_items]
            items_dated.sort(key=lambda p: p[1])  # ascending by date string (ISO)
            keep_ids = {id(p[0]) for p in items_dated[-window:]}
            changed = 0
            for it in data_items:
                desired_hidden = id(it) not in keep_ids
                current_hidden = it.h is True
                if desired_hidden != current_hidden:
                    it.h = True if desired_hidden else None
                    changed += 1
            if changed:
                first_visible = items_dated[-window][1][:10]
                last_visible  = items_dated[-1][1][:10]
                notes.append(
                    f"{sn}: {pt.name} window={window} → {first_visible} … {last_visible} ({changed} item flag(s) flipped)"
                )
    return notes


def _write_daily_serp_device(sheet, rows: List[Dict]) -> Tuple[int, int]:
    """UPSERT serp_device rows. Returns (updated, appended)."""
    if not rows:
        return 0, 0
    valid = [r for r in rows if r["weighted_pos"] is not None]
    if not valid:
        return 0, 0

    idx = _build_row_index(sheet, date_col=1, key_col=3)
    next_row = _first_empty_row(sheet)
    updated = appended = 0
    by_day: Dict[int, Dict[str, float]] = {}
    for r in valid:
        by_day.setdefault(r["dk"], {})[r["device"]] = float(r["weighted_pos"])

    for dk in sorted(by_day):
        d_dt = _dk_to_date(dk)
        key_date = d_dt.date()
        for device in DEVICES:
            pos = by_day[dk].get(device)
            if pos is None:
                continue
            existing_row = idx.get((key_date, device))
            row = existing_row if existing_row else next_row
            _write_cell(sheet, row, 1, d_dt)
            _write_cell(sheet, row, 2, pos)
            _write_cell(sheet, row, 3, device)
            if existing_row:
                updated += 1
            else:
                idx[(key_date, device)] = next_row
                next_row += 1
                appended += 1
    return updated, appended


def _serp_header_months(sheet) -> Tuple[Dict[int, int], Optional[int]]:
    """Scan the header row for month-name columns.

    Returns (col_by_offset_from_latest, delta_col).
    Because the header uses bare Dutch month names that recycle annually, we
    walk left-to-right assuming chronological order from an unknown start year.
    Anchor to the existing data: 'April' at col 18 corresponds to the most
    recent month with data (which the caller passes in). Returns the column of
    the Delta header if present.
    """
    delta_col: Optional[int] = None
    month_cols: List[int] = []
    for c in range(2, sheet.max_column + 1):
        v = sheet.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip().lower() == "delta":
            delta_col = c
            break
        if isinstance(v, str) and v.strip().capitalize() in DUTCH_MONTHS:
            month_cols.append(c)
    return {i: c for i, c in enumerate(month_cols)}, delta_col


def _insert_serp_month_column(sheet, new_month_name: str, delta_col: Optional[int]) -> int:
    """Insert a new column just before Delta (or at the right end if no Delta).
    Returns the column index of the newly inserted column.
    Header row gets the Dutch month name; styles inherit from the column to the left.
    """
    insert_at = delta_col if delta_col else sheet.max_column + 1
    sheet.insert_cols(insert_at)
    # openpyxl insert_cols does NOT copy formatting; clone from the column to the left.
    src_col = max(2, insert_at - 1)
    for r in range(1, max(sheet.max_row, 5) + 1):
        src = sheet.cell(row=r, column=src_col)
        new = sheet.cell(row=r, column=insert_at)
        if src.has_style:
            new.font = copy(src.font)
            new.fill = copy(src.fill)
            new.border = copy(src.border)
            new.alignment = copy(src.alignment)
            new.protection = copy(src.protection)
            new.number_format = src.number_format
    sheet.cell(row=1, column=insert_at, value=new_month_name)
    return insert_at


def _write_monthly_serp_type(sheet, new_yyyymm: int, positions: Dict[str, float]) -> int:
    """Insert a new month column with the latest avg-position data and refresh Delta.

    Idempotent: if the rightmost month-name column already equals the target
    month (i.e. a previous run for the same month), reuse it instead of
    inserting a duplicate.

    Assumes rows 2..5 are the four URL types in this order:
        Cat-url, C-url, PLP, R-url
    """
    month_cols, delta_col = _serp_header_months(sheet)
    if not month_cols:
        return 0

    last_month_col = month_cols[max(month_cols)]
    new_month_name = DUTCH_MONTHS[(new_yyyymm % 100) - 1]
    last_month_name = sheet.cell(row=1, column=last_month_col).value
    last_month_name = (last_month_name or "").strip() if isinstance(last_month_name, str) else ""

    if last_month_name == new_month_name:
        # Already have a column for this month — overwrite in place, don't insert.
        new_col = last_month_col
        # prev_col is the second-to-last month column
        sorted_month_cols = [month_cols[i] for i in sorted(month_cols)]
        prev_col = sorted_month_cols[-2] if len(sorted_month_cols) >= 2 else last_month_col
    else:
        new_col = _insert_serp_month_column(sheet, new_month_name, delta_col)
        if delta_col and new_col <= delta_col:
            delta_col += 1  # shifted right by the insert
        prev_col = last_month_col if last_month_col < new_col else last_month_col + 1

    written = 0
    for r in range(2, sheet.max_row + 1):
        label = sheet.cell(row=r, column=1).value
        if not isinstance(label, str):
            continue
        # match either display name or its raw search_console value
        display_label = label.strip()
        sql_label = next(
            (k for k, v in URL_TYPE_DISPLAY.items() if v == display_label),
            None,
        )
        if sql_label is None:
            continue
        pos = positions.get(sql_label)
        if pos is None:
            continue
        _write_cell(sheet, r, new_col, pos)
        written += 1
        if delta_col is not None:
            prev_val = sheet.cell(row=r, column=prev_col).value
            try:
                prev_f = float(prev_val) if prev_val is not None else None
            except (TypeError, ValueError):
                prev_f = None
            if prev_f is not None and prev_f != 0:
                _write_cell(sheet, r, delta_col, (pos - prev_f) / prev_f)
    return written


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def run_dm_review(target_yyyymm: Optional[int] = None) -> Dict:
    """Refresh slide-2 feeds in review_dm_seo.xlsx.

    Behavior: re-fetches a lookback window from Redshift each run and UPSERTs
    into the workbook. Existing rows within the window get updated in place;
    rows for new dates get appended. Rows older than the lookback are left
    untouched.

    `target_yyyymm` selects which month slide 2 is built for (the serp column +
    the target/behaald cards). Defaults to the current month. Choosing an older
    month anchors the refresh windows to that month and makes the target cards
    read that month's row in visits_omzet instead of the latest.
    """
    lock_err = _check_file_lock()
    if lock_err:
        return {"status": "error", "error": lock_err}

    logger.info("Opening workbook")
    wb = load_workbook(EXCEL_PATH, data_only=False, keep_vba=False)

    ws_monthly = wb[SHEET_MONTHLY]
    ws_daily = wb[SHEET_DAILY]
    ws_serp = wb[SHEET_SERP_TYPE]
    ws_serp_dev = wb[SHEET_SERP_DEVICE]

    today = datetime.now().date()

    # The serp tab is anchored to the month being processed (defaults to the
    # current month). We refresh that one column (insert or overwrite in place).
    serp_target_yyyymm = target_yyyymm or (today.year * 100 + today.month)
    # Anchor the lookback windows to the processed month, capped at today.
    anchor = _yyyymm_anchor_date(serp_target_yyyymm, today)

    # Lookback windows (intersected with the tab's existing range to handle
    # initial backfill — if a tab is empty we still pull from the lookback).
    daily_start = anchor - timedelta(days=LOOKBACK_DAYS_DAILY)
    daily_after_dk = int((daily_start - timedelta(days=1)).strftime("%Y%m%d"))

    monthly_start_yyyymm = _months_back(anchor, LOOKBACK_MONTHS_MONTHLY)

    logger.info(
        "Refresh windows — monthly:>=%s daily:>%s serp_device:>%s serp_target:%s",
        monthly_start_yyyymm, daily_after_dk, daily_after_dk, serp_target_yyyymm,
    )

    conn = get_redshift_connection()
    try:
        monthly_rows = _fetch_monthly_channel(conn, monthly_start_yyyymm)
        daily_rows = _fetch_daily_channel(conn, daily_after_dk)
        serp_dev_rows = _fetch_daily_serp_by_device(conn, daily_after_dk)
        serp_type_positions = _fetch_monthly_serp_by_url_type(conn, serp_target_yyyymm)
    finally:
        return_redshift_connection(conn)

    # Re-check lock right before write
    lock_err = _check_file_lock()
    if lock_err:
        return {"status": "error", "error": lock_err}

    monthly_up, monthly_app = _write_monthly_channel(ws_monthly, monthly_rows)
    daily_up, daily_app = _write_daily_channel(ws_daily, daily_rows)
    serp_dev_up, serp_dev_app = _write_daily_serp_device(ws_serp_dev, serp_dev_rows)

    # Extend pivot-table source ranges to cover the newly appended rows and
    # mark them refresh-on-load so the chart pivots in the file recompute when
    # Excel (or the linked pptx) next opens it.
    pivot_updates = _extend_pivot_sources(wb)
    # Slide each rolling-window pivot forward so the newest item is visible
    # and the oldest in the previous window drops off.
    pivot_window_updates = _roll_pivot_filter_window(wb)
    serp_cells = (
        _write_monthly_serp_type(ws_serp, serp_target_yyyymm, serp_type_positions)
        if serp_type_positions else 0
    )

    logger.info("Saving workbook")
    try:
        wb.save(EXCEL_PATH)
    except PermissionError as e:
        return {"status": "error",
                "error": f"Cannot save — file is locked. Close Excel and retry. ({e})"}

    # Update the SERP rankings table + the target/behaald cards on slide 2.
    # Pass the in-memory workbook (already saved above) so the pptx updaters read
    # slide-2 values from RAM instead of re-reading the 14MB file over the slow
    # OneDrive/WSL mount — a re-read there can hang while OneDrive syncs the save.
    from backend.dm_review_pptx_tables import update_serp_table, update_target_cards
    pptx_table_result = update_serp_table(PPTX_PATH, wb, PPTX_SLIDE_INDEX,
                                          target_yyyymm=serp_target_yyyymm)
    if pptx_table_result.get("status") != "ok":
        logger.warning("serp table update skipped: %s", pptx_table_result.get("error"))
    pptx_targets_result = update_target_cards(PPTX_PATH, wb, PPTX_SLIDE_INDEX,
                                              target_yyyymm=serp_target_yyyymm)
    if pptx_targets_result.get("status") != "ok":
        logger.warning("target cards update skipped: %s", pptx_targets_result.get("error"))

    return {
        "status": "ok",
        "refresh_window": {
            "monthly_start_yyyymm": monthly_start_yyyymm,
            "daily_start_date": daily_start.isoformat(),
            "serp_target_yyyymm": serp_target_yyyymm,
        },
        "rows_written": {
            "visits_omzet":     {"updated": monthly_up,  "appended": monthly_app},
            "visits_omzet_dag": {"updated": daily_up,    "appended": daily_app},
            "serp_device":      {"updated": serp_dev_up, "appended": serp_dev_app},
            "serp":             {"cells_written": serp_cells},
        },
        "pivot_sources_updated": pivot_updates,
        "pivot_windows_rolled": pivot_window_updates,
        "pptx_serp_table": pptx_table_result,
        "pptx_target_cards": pptx_targets_result,
    }
