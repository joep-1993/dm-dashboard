"""
Regression tests for DMA+ shop exclusions across MULTIPLE Google Ads accounts
(2026-08-26).

Since 2026-08-20 the NL market also runs on 4089798584 ("DMA NL 2"), whose
category campaigns carry a suffix: PLA/<cat>_<tier>_limit (DMA Level 1/2/3
ladder) and PLA/<cat>_<tier>_label (the "DMA: Label A/B/C" portfolios). DMA+
used to look in the primary account only, so shops dropped from a category kept
serving there.

What these tests pin down:
  * exclude AND reverse-exclude visit every account in EXCLUSION_ACCOUNTS;
  * each mutate goes to the account the campaign actually lives in;
  * DMA NL 2 is probed with BOTH its name variants (_limit and _label);
  * a category slot only counts as "missing" when NO account has it — DMA NL 2
    holds a subset of the categories, so the suffixed names must never inflate
    the missing-campaigns list;
  * markets with one account (BE) behave exactly as before;
  * dry_run still calls nothing;
  * the campaign header line stays parseable by the dashboard's log parser.

Everything is monkeypatched — no Google Ads calls are made.

Run:  ./venv/bin/python -m pytest backend/test_dma_plus_multi_account.py -q
   or ./venv/bin/python backend/test_dma_plus_multi_account.py
"""
import io
import os
import sys
import contextlib

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from backend import campaign_processor as cp  # noqa: E402
from backend.dma_plus_service import _parse_affected_entities  # noqa: E402

NL1 = "3800751597"   # DMA NL      — plain "PLA/<cat>_<tier>"
NL2 = "4089798584"   # DMA NL 2    — "PLA/<cat>_<tier>_limit" / "_label"
BE = "9920951707"    # DMA BE      — plain, single account

# Fixture: 2 categories under maincat 77, cl1 = a.
# Afzuigkappen exists in both accounts, Koffiemolens only in DMA NL — which is
# the normal case, DMA NL 2 covers ~350 of the ~3200 categories.
FAKE_ACCOUNTS = {
    NL1: {"PLA/Afzuigkappen_a": "ag-nl1-afz", "PLA/Koffiemolens_a": "ag-nl1-kof"},
    NL2: {"PLA/Afzuigkappen_a_limit": "ag-nl2-limit",
          "PLA/Afzuigkappen_a_label": "ag-nl2-label"},
    BE: {"PLA/Afzuigkappen_a": "ag-be-afz"},
}

CALLS = []  # (op, customer_id, ad_group_id, shops)


def _fake_prefetch(client, customer_id, campaign_prefix="PLA/"):
    return {
        name: {
            "resource_name": f"customers/{customer_id}/campaigns/{name}",
            # ad groups keep the UNSUFFIXED name in DMA NL 2, as in production
            "ad_groups": [{"id": ag, "name": "PLA/Afzuigkappen_a",
                           "resource_name": f"customers/{customer_id}/adGroups/{ag}"}],
        }
        for name, ag in FAKE_ACCOUNTS[customer_id].items()
    }


def _fake_add(client, customer_id, ad_group_id, ad_group_name, shop_names):
    CALLS.append(("add", customer_id, ad_group_id, tuple(sorted(shop_names))))
    return {"success": list(shop_names), "already_excluded": [], "errors": []}


def _fake_reverse(client, customer_id, ad_group_id, ad_group_name, shop_names):
    CALLS.append(("remove", customer_id, ad_group_id, tuple(sorted(shop_names))))
    return {"success": list(shop_names), "not_found": [], "errors": []}


def _patch():
    cp.prefetch_pla_campaigns_and_ad_groups = _fake_prefetch
    cp.add_shop_exclusions_batch = _fake_add
    cp.reverse_exclusion_batch = _fake_reverse
    cp.time.sleep = lambda *a, **k: None
    CALLS.clear()


def _workbook(sheet_name, shop="proshop.nl"):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    cat = wb.create_sheet(cp.SHEET_CAT_IDS)
    cat.append(["maincat", "maincat_id", "deepest_cat", "cat_id"])
    cat.append(["Witgoed", "77", "Afzuigkappen", "1001"])
    cat.append(["Witgoed", "77", "Koffiemolens", "1002"])
    sheet = wb.create_sheet(sheet_name)
    sheet.append(["shop", "shop_id", "maincat", "maincat_id", "cl1", "status", "error"])
    sheet.append([shop, "", "Witgoed", "77", "a", None, None])
    return wb, sheet


def _run(fn, sheet_name, customer_id=NL1, **kwargs):
    _patch()
    wb, sheet = _workbook(sheet_name)
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        fn(client=None, workbook=wb, customer_id=customer_id, **kwargs)
    return list(CALLS), sheet, buf.getvalue()


def test_exclusion_reaches_both_nl_accounts():
    calls, sheet, log = _run(cp.process_exclusion_sheet_v2, cp.SHEET_EXCLUSION)

    # 2 campaigns in DMA NL + the _limit/_label pair in DMA NL 2
    assert len(calls) == 4, calls
    assert ("add", NL1, "ag-nl1-afz", ("proshop.nl",)) in calls
    assert ("add", NL1, "ag-nl1-kof", ("proshop.nl",)) in calls
    assert ("add", NL2, "ag-nl2-limit", ("proshop.nl",)) in calls
    assert ("add", NL2, "ag-nl2-label", ("proshop.nl",)) in calls

    # Every mutate carries the account its campaign lives in
    for _, cid, ag_id, _shops in calls:
        assert cid == (NL2 if ag_id.startswith("ag-nl2") else NL1), (cid, ag_id)

    assert sheet.cell(row=2, column=cp.COL_EX_STATUS + 1).value is True
    assert "DMA NL 2 (4089798584): 2" in log


def test_missing_campaigns_not_inflated_by_suffixed_names():
    """Koffiemolens has no DMA NL 2 counterpart — that is not a missing campaign."""
    _calls, _sheet, log = _run(cp.process_exclusion_sheet_v2, cp.SHEET_EXCLUSION)
    assert "PLA/Koffiemolens_a_limit" not in log
    assert "PLA/Koffiemolens_a_label" not in log
    assert "Campaign not found in Google Ads cache" not in log

    parsed = _parse_affected_entities(log)
    assert parsed["missing_campaigns"] == []
    # Suffixed campaigns pair up with their ad group, so the export shows them
    pairs = {(c, a) for c, a, _t in parsed["campaign_ad_group_pairs"]}
    assert ("PLA/Afzuigkappen_a_limit", "PLA/Afzuigkappen_a") in pairs
    assert ("PLA/Afzuigkappen_a_label", "PLA/Afzuigkappen_a") in pairs


def test_slot_missing_only_when_absent_everywhere():
    _patch()
    wb, sheet = _workbook(cp.SHEET_EXCLUSION)
    wb[cp.SHEET_CAT_IDS].append(["Witgoed", "77", "Vaatwassers", "1003"])  # in neither account
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        cp.process_exclusion_sheet_v2(client=None, workbook=wb, customer_id=NL1)
    log = buf.getvalue()
    assert "Campaign not found in Google Ads cache: PLA/Vaatwassers_a" in log
    # reported once, under the primary account's naming convention
    assert log.count("Campaign not found in Google Ads cache") == 1
    assert _parse_affected_entities(log)["missing_campaigns"] == ["PLA/Vaatwassers_a"]


def test_maincat_coverage_stays_per_slot():
    """A slot served from two accounts is still one category — never >100%."""
    _calls, _sheet, log = _run(cp.process_exclusion_sheet_v2, cp.SHEET_EXCLUSION)
    assert "Campaigns found in Witgoed: 2/2 (100%)" in log, log[-1500:]
    assert "Campaigns matched in Witgoed: 4 (+2 in secondary accounts)" in log, log[-1500:]


def test_reverse_exclusion_reaches_both_nl_accounts():
    """Un-excluding must cover DMA NL 2, else a returning shop stays blocked there."""
    calls, sheet, _log = _run(cp.process_reverse_exclusion_sheet, "verwijderen")
    assert len(calls) == 4, calls
    assert all(op == "remove" for op, *_ in calls)
    assert {cid for _op, cid, *_ in calls} == {NL1, NL2}
    assert sheet.cell(row=2, column=6).value is True


def test_combined_exclusion_covers_both_ops_in_both_accounts():
    _patch()
    exc_wb, exc_sheet = _workbook(cp.SHEET_EXCLUSION, shop="proshop.nl")
    rex_wb, rex_sheet = _workbook("verwijderen", shop="bobplaza.com")
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        result = cp.process_combined_exclusion_v2(
            client=None, exc_workbook=exc_wb, rex_workbook=rex_wb, customer_id=NL1)

    adds = [c for c in CALLS if c[0] == "add"]
    removes = [c for c in CALLS if c[0] == "remove"]
    assert len(adds) == 4 and len(removes) == 4, (adds, removes)
    assert {c[1] for c in adds} == {NL1, NL2}
    assert {c[1] for c in removes} == {NL1, NL2}
    assert all(c[3] == ("proshop.nl",) for c in adds)
    assert all(c[3] == ("bobplaza.com",) for c in removes)
    assert exc_sheet.cell(row=2, column=6).value is True
    assert rex_sheet.cell(row=2, column=6).value is True
    assert result["exclude"]["errors"] == 0
    assert result["reverse_exclude"]["errors"] == 0


def test_single_account_market_unchanged():
    calls, _sheet, log = _run(cp.process_exclusion_sheet_v2, cp.SHEET_EXCLUSION,
                              customer_id=BE)
    assert {cid for _op, cid, *_ in calls} == {BE}
    # no per-account block for a market with one account
    assert "Campaigns matched per account" not in log


def test_dry_run_touches_nothing():
    calls, _sheet, _log = _run(cp.process_exclusion_sheet_v2, cp.SHEET_EXCLUSION,
                               dry_run=True)
    assert calls == []


def test_missing_campaign_names_with_spaces_survive_the_parser():
    """deepest_cats containing a space must not be cut in half in the export."""
    log = (
        "    ⚠️  Campaign not found in Google Ads cache: PLA/Philips 1000 series_a\n"
        "    ⚠️  Campaign not found in cache: PLA/Tefal Easy Fry Dual XXL_a\n"  # combined-run wording
        "Missing campaigns in Huishoudelijk (1): PLA/Philips 2000 series_a\n"
    )
    missing = _parse_affected_entities(log)["missing_campaigns"]
    assert missing == ["PLA/Philips 1000 series_a", "PLA/Philips 2000 series_a",
                       "PLA/Tefal Easy Fry Dual XXL_a"], missing


def test_account_helpers():
    assert cp.exclusion_account_ids(NL1) == [NL1, NL2]
    assert cp.exclusion_account_ids(BE) == [BE]
    assert cp.exclusion_account_ids("9999999999") == ["9999999999"]
    # primary account never gets a tag; the second one does
    assert cp._account_tag(NL1, [NL1, NL2]) == ""
    assert cp._account_tag(NL2, [NL1, NL2]) == " [DMA NL 2]"

    caches = {NL1: _fake_prefetch(None, NL1), NL2: _fake_prefetch(None, NL2)}
    names = [m["campaign_name"] for m in cp.find_category_campaigns(caches, "Afzuigkappen", "a")]
    assert names == ["PLA/Afzuigkappen_a", "PLA/Afzuigkappen_a_limit",
                     "PLA/Afzuigkappen_a_label"], names
    assert cp.find_category_campaigns(caches, "Koffiemolens", "a")[0]["customer_id"] == NL1
    assert cp.find_category_campaigns(caches, "Vaatwassers", "a") == []


if __name__ == "__main__":
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    failed = 0
    for t in tests:
        try:
            t()
            print(f"PASS  {t.__name__}")
        except AssertionError as e:
            failed += 1
            print(f"FAIL  {t.__name__}: {e}")
        except Exception as e:
            failed += 1
            print(f"ERROR {t.__name__}: {type(e).__name__}: {e}")
    print(f"\n{len(tests) - failed}/{len(tests)} passed")
    sys.exit(1 if failed else 0)
