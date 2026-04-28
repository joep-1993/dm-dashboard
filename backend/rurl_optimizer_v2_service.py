"""
Background-task wrapper for the bundled R-URL Optimizer.

The optimizer (backend/rurl_optimizer/) ships as a self-contained CLI. We run
it as a subprocess per task and tail stdout into an in-memory task dict so the
dashboard can poll progress, stream logs, and download the resulting CSV.
"""
from __future__ import annotations

import logging
import os
import re
import shlex
import subprocess
import sys
import threading
import uuid
from collections import deque
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional

logger = logging.getLogger(__name__)

ENGINE_VERSION = 2
PKG_DIR = Path(__file__).parent / "rurl_optimizer_v2"
# Persisted under backend/data/ (same dir as the history JSON) so output
# files survive uvicorn restarts and /tmp wipes — required for the Export
# button on old history rows.
OUTPUT_DIR = Path(__file__).parent / "data" / "rurl-optimizer-v2-output"
INPUT_DIR = Path(__file__).parent / "data" / "rurl-optimizer-v2-input"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
INPUT_DIR.mkdir(parents=True, exist_ok=True)

_TASKS: Dict[str, Dict[str, Any]] = {}
_TASKS_LOCK = threading.Lock()
# Tracks the runner thread per task so the stale-task sweep can tell
# "subprocess died" apart from "still in Redshift fetch / Python phase".
_THREADS: Dict[str, threading.Thread] = {}

# History is persisted to disk (same pattern as DMA+) so uvicorn --reload or
# a machine reboot doesn't wipe the Recent runs table.
_HISTORY_FILE: Path = Path(__file__).parent / "data" / "rurl_optimizer_v2_history.json"
_HISTORY_LOCK = threading.Lock()


def _load_history_from_disk() -> deque:
    if _HISTORY_FILE.exists():
        try:
            import json as _json
            data = _json.loads(_HISTORY_FILE.read_text(encoding="utf-8"))
            if isinstance(data, list):
                return deque(data, maxlen=200)
        except Exception as e:
            logger.warning(f"Failed to load rurl history from {_HISTORY_FILE}: {e}")
    return deque(maxlen=200)


def _save_history_to_disk():
    try:
        import json as _json
        _HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
        _HISTORY_FILE.write_text(
            _json.dumps(list(_HISTORY), default=str, ensure_ascii=False),
            encoding="utf-8",
        )
    except Exception as e:
        logger.warning(f"Failed to save rurl history to {_HISTORY_FILE}: {e}")


_HISTORY: deque = _load_history_from_disk()

# tqdm writes lines like "Processing:  42%|████▏     | 4200/10000 [00:15<00:20, ...]"
_TQDM_RE = re.compile(r"(?P<pct>\d+)%\|.*?\|\s*(?P<cur>\d+)/(?P<tot>\d+)")


def _set(task_id: str, patch: Dict[str, Any]) -> None:
    with _TASKS_LOCK:
        t = _TASKS.setdefault(task_id, {})
        t.update(patch)


def _get(task_id: str) -> Optional[Dict[str, Any]]:
    with _TASKS_LOCK:
        return dict(_TASKS[task_id]) if task_id in _TASKS else None


def _append_log(task_id: str, line: str) -> None:
    with _TASKS_LOCK:
        t = _TASKS.setdefault(task_id, {})
        log = t.setdefault("log", [])
        log.append(line)
        # cap log to last 500 lines
        if len(log) > 500:
            del log[: len(log) - 500]


def _run_subprocess(task_id: str, argv: list[str], output_path: Path, script: str, optional_output: bool = False) -> None:
    # Capture prior state so an optional auxiliary stage can restore it if it
    # exits cleanly without producing output (e.g. global pass, no globals).
    prior = _get(task_id) or {}
    prior_state = {
        "status": prior.get("status"),
        "progress": prior.get("progress"),
        "message": prior.get("message"),
        "output_path": prior.get("output_path"),
        "script": prior.get("script"),
    } if optional_output else None
    _set(task_id, {
        "status": "running",
        "progress": 0,
        "message": "Starting...",
        "started_at": datetime.now().isoformat(),
        "output_path": str(output_path),
        "script": script,
    })
    _append_log(task_id, f"--- Running {script} ---")

    env = os.environ.copy()
    # Ensure the bundled package can import its sibling modules (src/, config).
    env["PYTHONPATH"] = str(PKG_DIR) + os.pathsep + env.get("PYTHONPATH", "")
    # Silence noisy DeprecationWarning / FutureWarning from pandas / pyarrow.
    env["PYTHONWARNINGS"] = "ignore::DeprecationWarning,ignore::FutureWarning"
    # Force UTF-8 for stdio so non-ASCII glyphs in print() don't crash on
    # Windows (default cp1252) — see UnicodeEncodeError on '✓' etc.
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUTF8"] = "1"

    try:
        proc = subprocess.Popen(
            argv,
            cwd=str(PKG_DIR),
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
            # Force UTF-8 decode on the parent side too (subprocess already
            # writes UTF-8 via PYTHONIOENCODING). Without this, the parent
            # uses the system default — cp1252 on Windows — and a single
            # non-ASCII byte in the engine's stdout (a product name, a
            # Dutch log line, a URL with ü/é) raises UnicodeDecodeError
            # mid-loop and kills the runner thread silently.
            encoding="utf-8",
            errors="replace",
        )
    except FileNotFoundError as e:
        _set(task_id, {"status": "failed", "error": f"Failed to start: {e}"})
        return

    _set(task_id, {"pid": proc.pid})

    assert proc.stdout is not None
    for raw_line in proc.stdout:
        line = raw_line.rstrip()
        if not line:
            continue
        _append_log(task_id, line)

        m = _TQDM_RE.search(line)
        if m:
            pct = int(m.group("pct"))
            cur = int(m.group("cur"))
            tot = int(m.group("tot"))
            _set(task_id, {
                "progress": pct,
                "message": f"Processing {cur:,}/{tot:,} URLs",
            })
        elif line.startswith("Pre-loading") or "Loading " in line:
            _set(task_id, {"progress": 5, "message": "Loading data..."})
        elif "Data loaded" in line or "Data cached" in line:
            _set(task_id, {"progress": 15, "message": "Data ready"})
        elif line.startswith("Saving to"):
            _set(task_id, {"progress": 98, "message": "Saving output..."})

        # check if task was cancelled
        if _get(task_id) and _get(task_id).get("cancel_requested"):
            proc.terminate()
            try:
                proc.wait(timeout=5)
            except subprocess.TimeoutExpired:
                proc.kill()
            _set(task_id, {"status": "cancelled", "message": "Cancelled by user"})
            _history_append(task_id)
            return

    rc = proc.wait()
    if rc == 0 and output_path.exists():
        _set(task_id, {
            "status": "completed",
            "progress": 100,
            "message": f"Done. Output: {output_path.name}",
            "finished_at": datetime.now().isoformat(),
        })
    elif rc == 0 and optional_output:
        # Auxiliary stage exited cleanly without writing output (e.g. the
        # global R-URL pass when no global URLs are present). Restore the
        # prior status / output_path so the main stage's success still
        # stands and downstream steps (cache merge, xlsx, DB save) run.
        if prior_state:
            _set(task_id, {k: v for k, v in prior_state.items() if v is not None})
        _append_log(task_id, f"[info] {script}: nothing to write (no output file)")
        return
    else:
        _set(task_id, {
            "status": "failed",
            "error": f"Exit code {rc}",
            "finished_at": datetime.now().isoformat(),
        })
    _history_append(task_id)


def _history_update_latest(task_id: str, patch: Dict[str, Any]) -> None:
    """Patch the most recent history entry for this task_id (if any)."""
    with _HISTORY_LOCK:
        if _HISTORY and _HISTORY[0].get("task_id") == task_id:
            _HISTORY[0].update(patch)
            _save_history_to_disk()


def _history_append(task_id: str) -> None:
    t = _get(task_id)
    if not t:
        return
    new_entry = {
        "task_id": task_id,
        "script": t.get("script"),
        "status": t.get("status"),
        "started_at": t.get("started_at"),
        "finished_at": t.get("finished_at"),
        "message": t.get("message"),
        "error": t.get("error"),
        "output_path": t.get("output_path"),
        "params": t.get("params"),
    }
    with _HISTORY_LOCK:
        # Dedupe by task_id: a single user-initiated run can hit this
        # function multiple times (main pass, global pass, cache_only,
        # xlsx step). Update the existing entry in place so the UI shows
        # one row per run, while preserving the original start time and
        # the user-facing script name from the first stage.
        for i, h in enumerate(_HISTORY):
            if h.get("task_id") == task_id:
                merged = {**h, **new_entry}
                if h.get("started_at"):
                    merged["started_at"] = h["started_at"]
                if h.get("script") and h["script"] != "process_global_rurls":
                    merged["script"] = h["script"]
                if h.get("params"):
                    merged["params"] = h["params"]
                _HISTORY[i] = merged
                _save_history_to_disk()
                return
        _HISTORY.appendleft(new_entry)
        _save_history_to_disk()


TAXV2_BASE = "http://producttaxonomyunifiedapi-prod.azure.api.beslist.nl"
_SLUG_TO_MAINCAT: Dict[str, str] = {}
_CAT_ID_TO_DEEPEST: Dict[str, str] = {}


def _ensure_cat_id_lookup() -> Dict[str, str]:
    """Build a deepest cat_id -> readable deepest_cat name (lazy, cached)."""
    global _CAT_ID_TO_DEEPEST
    if _CAT_ID_TO_DEEPEST:
        return _CAT_ID_TO_DEEPEST
    try:
        from backend.category_keyword_service import PRELOADED_CATEGORIES
        _CAT_ID_TO_DEEPEST = {
            str(c["cat_id"]): c["deepest_cat"]
            for c in PRELOADED_CATEGORIES
            if c.get("cat_id") and c.get("deepest_cat")
        }
    except Exception as e:
        logger.warning(f"could not build cat_id->deepest lookup: {e}")
    return _CAT_ID_TO_DEEPEST


def _deepest_category_from_redirect(redirect_url) -> str:
    """Resolve the readable deepest-category name from a redirect URL.

    Extracts the trailing _<digits> from the path before any /c/ facet
    segment and looks it up in PRELOADED_CATEGORIES (cat_id -> deepest_cat).
    Used by the cross-engine 'Export all' endpoint where the per-row
    redirect_category isn't available (rurl_processed doesn't cache it).
    """
    if not redirect_url or not isinstance(redirect_url, str):
        return ""
    path = redirect_url.split("/c/", 1)[0]
    parts = path.rstrip("/").split("/")
    cat_id = ""
    if parts:
        for tok in reversed(parts[-1].split("_")):
            if tok.isdigit():
                cat_id = tok
                break
    if not cat_id:
        return ""
    return _ensure_cat_id_lookup().get(cat_id, "")


def _ensure_slug_lookup() -> Dict[str, str]:
    """Build a urlSlug -> readable name map for root categories via taxv2.

    Lazy + cached for the lifetime of the worker. Falls back to an empty
    dict on any API failure; callers handle the miss with slug cleanup.
    """
    global _SLUG_TO_MAINCAT
    if _SLUG_TO_MAINCAT:
        return _SLUG_TO_MAINCAT
    try:
        import requests
        r = requests.get(
            f"{TAXV2_BASE}/api/Categories",
            params={"rootCategoriesOnly": "true", "locale": "nl-NL"},
            timeout=10,
        )
        r.raise_for_status()
        for cat in r.json() or []:
            for lbl in (cat.get("labels") or []):
                slug = (lbl.get("urlSlug") or "").strip()
                name = (lbl.get("name") or "").strip()
                if slug and name:
                    _SLUG_TO_MAINCAT[slug] = name
    except Exception as e:
        logger.warning(f"taxv2 main-category lookup failed: {e}")
    return _SLUG_TO_MAINCAT


def _main_category_from_redirect(redirect_url) -> str:
    """Resolve the readable main-category name for a redirect URL via taxv2.

    Extracts the main-category slug from `/products/<slug>/...` and looks
    it up against taxv2's root-category urlSlug list. Falls back to a
    cleaned slug when taxv2 doesn't have a match.
    """
    import re as _re
    if not redirect_url or not isinstance(redirect_url, str):
        return ""
    m = _re.search(r"/products/([^/]+)", redirect_url)
    if not m:
        return ""
    slug = m.group(1).strip()
    name = _ensure_slug_lookup().get(slug)
    if name:
        return name
    return slug.replace("_", " ").strip()


def _write_xlsx_output(df, csv_path: Path) -> Path:
    """Project to the user-facing columns and write as .xlsx.

    Output schema (per 2.0 changes):
      old url | new url | score | main_category | deepest_category | visits | revenue | reason
    """
    import pandas as pd

    out = pd.DataFrame()
    out["old url"] = df.get("original_url", pd.Series(dtype=object))
    out["new url"] = df.get("redirect_url", pd.Series(dtype=object))
    out["score"] = df.get("reliability_score", pd.Series(dtype="Int64"))
    if "redirect_url" in df.columns:
        out["main_category"] = df["redirect_url"].apply(_main_category_from_redirect)
    else:
        out["main_category"] = ""
    out["deepest_category"] = df.get("redirect_category", pd.Series(dtype=object))
    out["visits"] = df.get("visits", pd.Series(dtype=object))
    out["revenue"] = df.get("visit_rev", pd.Series(dtype=object))
    out["reason"] = df.get("reason", pd.Series(dtype=object))

    # Sort by score (descending). Coerce to numeric so non-numeric values
    # (NaN, strings) sort to the bottom rather than crashing the comparison.
    out["__score_sort"] = pd.to_numeric(out["score"], errors="coerce")
    out = out.sort_values("__score_sort", ascending=False, na_position="last").drop(
        columns="__score_sort"
    )

    xlsx_path = csv_path.with_suffix(".xlsx")
    out.to_excel(xlsx_path, index=False)

    # Post-process: center-align score / visits / revenue columns.
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
        wb = load_workbook(xlsx_path)
        ws = wb.active
        center = Alignment(horizontal="center", vertical="center")
        col_idx = {name: i + 1 for i, name in enumerate(out.columns)}
        for name in ("score", "visits", "revenue"):
            ci = col_idx.get(name)
            if not ci:
                continue
            for row in range(1, ws.max_row + 1):
                ws.cell(row=row, column=ci).alignment = center
        wb.save(xlsx_path)
    except Exception as e:
        logger.warning(f"xlsx alignment post-process failed: {e}")

    try:
        csv_path.unlink()
    except FileNotFoundError:
        pass
    return xlsx_path


def _normalize_upload(csv_bytes: bytes) -> bytes:
    """
    Ensure the uploaded CSV has a column named 'r_url'. If the file has a
    single column with any other name, rename it. If it has multiple columns,
    leave it untouched (the optimizer expects an 'r_url' column then, which
    the Redshift path already produces).
    """
    import io
    import pandas as pd
    try:
        df = pd.read_csv(io.BytesIO(csv_bytes))
    except Exception:
        return csv_bytes
    if "r_url" in df.columns:
        return csv_bytes
    if len(df.columns) == 1:
        df.columns = ["r_url"]
        buf = io.StringIO()
        df.to_csv(buf, index=False)
        return buf.getvalue().encode("utf-8")
    return csv_bytes


def _read_url_column(csv_path: Path, url_column: str) -> list[str]:
    """Read only the URL column from a CSV. Returns empty list on any failure."""
    try:
        import pandas as pd
        df = pd.read_csv(csv_path, usecols=[url_column])
        return df[url_column].dropna().astype(str).tolist()
    except Exception:
        return []


def _filter_input_csv(csv_path: Path, url_column: str, skip_urls: set[str]) -> None:
    """Rewrite CSV in place keeping only rows whose URL is NOT in skip_urls."""
    if not skip_urls:
        return
    import pandas as pd
    df = pd.read_csv(csv_path)
    if url_column not in df.columns:
        return
    mask = ~df[url_column].astype(str).isin(skip_urls)
    df[mask].to_csv(csv_path, index=False)


def _url_contains_shopname(url: str) -> bool:
    """True if the R-URL keyword segment contains any SHOP_NAME word.

    Mirrors the engine's V30 short-circuit: extract /r/<keyword>, replace
    underscores with spaces, run detect_shops_in_keyword. Used as a
    pre-filter so shop-name URLs don't burn through the row_limit.
    """
    if not url:
        return False
    import re as _re
    m = _re.search(r"/r/([^/?#]+)", url)
    if not m:
        return False
    keyword = m.group(1).replace("_", " ").strip()
    if not keyword:
        return False
    try:
        from backend.rurl_optimizer_v2.src.validation_rules import detect_shops_in_keyword
        return bool(detect_shops_in_keyword(keyword))
    except Exception as e:
        logger.warning(f"shopname pre-filter failed: {e}")
        return False


def _fetch_redshift_rurls(lookback_days: int = 365, row_limit: Optional[int] = None) -> bytes:
    """Pull R-URLs with visits + revenue from Redshift for the last N days."""
    from backend.database import get_redshift_connection
    import csv
    import io

    today = datetime.now().strftime("%Y%m%d")
    from datetime import timedelta
    start = (datetime.now() - timedelta(days=lookback_days)).strftime("%Y%m%d")

    sql = f"""
    SELECT
        dv.main_cat_name AS maincat,
        dv.deepest_subcat_name AS deepest_cat,
        SPLIT_PART(dv.url, '?', 1) AS r_url,
        count(*) AS visits,
        sum(fcv.cpc_revenue) + sum(fcv.ww_revenue) AS visit_rev
    FROM datamart.fct_visits fcv
    JOIN datamart.dim_visit dv ON fcv.dim_visit_key = dv.dim_visit_key
    JOIN datamart.dim_date dat ON fcv.dim_date_key = dat.dim_date_key
    JOIN chan_deriv.ref_channel_derivation_stats chan
      ON dv.aff_id = chan.aff_id AND dv.channel_id = chan.channel_id
    WHERE dv.is_real_visit = 1
      AND fcv.dim_date_key BETWEEN {start} AND {today}
      AND dv.url LIKE '%beslist.nl%'
      AND dv.url LIKE '%/r/%'
    GROUP BY 1, 2, 3
    ORDER BY visits DESC
    """
    if row_limit and row_limit > 0:
        sql += f"\n    LIMIT {int(row_limit)}"

    conn = get_redshift_connection()
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["maincat", "deepest_cat", "r_url", "visits", "visit_rev"])
    for r in rows:
        # cursor returns RealDictRow — access by key
        w.writerow([r["maincat"], r["deepest_cat"], r["r_url"], r["visits"], r["visit_rev"]])
    return buf.getvalue().encode("utf-8")


def start_optimize(
    csv_bytes: Optional[bytes],
    filename: Optional[str],
    workers: Optional[int],
    threshold: int,
    multi_facet: bool,
    url_column: str,
    also_global: bool,
    source: str = "upload",
    lookback_days: int = 365,
    row_limit: Optional[int] = None,
    force_reprocess: bool = False,
    exclude_shopnames: bool = False,
) -> str:
    task_id = uuid.uuid4().hex[:8]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    if source == "redshift":
        filename = f"redshift_{lookback_days}d.csv"
        # Defer the fetch to the worker thread — it can take 30-60s.
        input_path = INPUT_DIR / f"input_{task_id}_{filename}"
    else:
        if not csv_bytes:
            raise ValueError("csv_bytes required when source=upload")
        input_path = INPUT_DIR / f"input_{task_id}_{filename}"
        input_path.write_bytes(_normalize_upload(csv_bytes))

    output_path = OUTPUT_DIR / f"redirects_{task_id}_{ts}.csv"

    argv = [
        sys.executable,
        "main_parallel_v2.py",
        str(input_path),
        "-o", str(output_path),
        "-c", url_column,
        "--threshold", str(threshold),
    ]
    if workers:
        argv += ["-w", str(workers)]
    if multi_facet:
        argv.append("--multi-facet")

    _set(task_id, {
        "status": "queued",
        "progress": 0,
        "message": "Queued",
        "log": [],
        "params": {
            "filename": filename,
            "workers": workers,
            "threshold": threshold,
            "multi_facet": multi_facet,
            "url_column": url_column,
            "also_global": also_global,
            "source": source,
            "lookback_days": lookback_days if source == "redshift" else None,
            "row_limit": row_limit if source == "redshift" else None,
            "force_reprocess": force_reprocess,
        },
    })

    def _runner():
        if source == "redshift":
            _set(task_id, {"status": "running", "progress": 1,
                           "message": f"Querying Redshift (last {lookback_days} days)..."})
            _append_log(task_id,
                        f"Fetching R-URLs from Redshift (last {lookback_days} days"
                        + (f", limit {row_limit}" if row_limit else "") + ")...")
            try:
                # Iterative fetch: start with a 10x oversample, then double
                # the LIMIT each round if we don't have enough fresh URLs
                # after dropping cached + (optionally) shopname rows. Stops
                # when we have row_limit fresh URLs, Redshift is exhausted
                # (returned fewer rows than asked), or we hit MAX_FETCH.
                needs_oversample = (not force_reprocess) or exclude_shopnames
                if needs_oversample and row_limit:
                    import io as _io
                    import pandas as _pd
                    from backend import rurl_optimizer_persistence as pers
                    MAX_FETCH = 1_000_000
                    fetch_limit = max(row_limit * 10, 1000)
                    df_rs = None
                    raw_n = 0
                    cached_n = 0
                    shop_n = 0
                    while True:
                        data = _fetch_redshift_rurls(lookback_days, fetch_limit)
                        df_full = _pd.read_csv(_io.BytesIO(data))
                        raw_n = len(df_full)
                        df_filtered = df_full
                        if not force_reprocess:
                            cached = pers.already_processed(df_filtered["r_url"].tolist())
                            df_filtered = df_filtered[~df_filtered["r_url"].isin(cached)]
                            cached_n = len(cached)
                        if exclude_shopnames:
                            shop_mask = df_filtered["r_url"].apply(_url_contains_shopname)
                            shop_n = int(shop_mask.sum())
                            df_filtered = df_filtered[~shop_mask]
                        if len(df_filtered) >= row_limit:
                            df_rs = df_filtered.head(row_limit)
                            break
                        if raw_n < fetch_limit:
                            # Redshift returned everything it has — can't go further.
                            df_rs = df_filtered
                            _append_log(task_id,
                                        f"Redshift exhausted at {raw_n:,} rows; "
                                        f"only {len(df_filtered):,} fresh URLs available.")
                            break
                        if fetch_limit >= MAX_FETCH:
                            df_rs = df_filtered
                            _append_log(task_id,
                                        f"[warn] hit fetch cap of {MAX_FETCH:,}; "
                                        f"only {len(df_filtered):,} fresh URLs available.")
                            break
                        fetch_limit = min(fetch_limit * 2, MAX_FETCH)
                        _append_log(task_id,
                                    f"Only {len(df_filtered):,} fresh URLs after filtering "
                                    f"{raw_n:,} rows; expanding fetch to {fetch_limit:,}...")
                    buf = _io.StringIO()
                    df_rs.to_csv(buf, index=False)
                    data = buf.getvalue().encode("utf-8")
                    parts = [f"Redshift returned {raw_n:,} rows"]
                    if cached_n:
                        parts.append(f"{cached_n:,} already processed")
                    if shop_n:
                        parts.append(f"{shop_n:,} contained shopnames")
                    parts.append(f"keeping {len(df_rs):,} fresh URLs")
                    _append_log(task_id, "; ".join(parts) + ".")
                else:
                    data = _fetch_redshift_rurls(lookback_days, row_limit)
                input_path.write_bytes(data)
                nrows = data.count(b"\n") - 1
                _append_log(task_id, f"Redshift returned {nrows:,} rows -> {input_path.name}")
            except Exception as e:
                _set(task_id, {"status": "failed", "error": f"Redshift fetch failed: {e}",
                               "finished_at": datetime.now().isoformat()})
                _history_append(task_id)
                return

        # Persistence: filter out URLs already processed (unless forced).
        all_input_urls = _read_url_column(input_path, url_column)
        cached_urls: set[str] = set()
        if not force_reprocess and all_input_urls:
            try:
                from backend import rurl_optimizer_persistence as pers
                cached_urls = pers.already_processed(all_input_urls)
            except Exception as e:
                _append_log(task_id, f"[warn] persistence lookup failed: {e} — processing all URLs")
                cached_urls = set()
        if cached_urls:
            _append_log(task_id,
                        f"Persistence: {len(cached_urls):,} of {len(all_input_urls):,} URLs "
                        f"already processed — skipping them.")
            _filter_input_csv(input_path, url_column, cached_urls)

        # Optional: drop URLs whose keyword contains a SHOP_NAME so they
        # don't burn through the row_limit. For the redshift path the
        # input file was already filtered upstream; this is mostly for
        # the upload path (and as a safety net if the upstream filter
        # missed anything).
        if exclude_shopnames:
            remaining_urls = _read_url_column(input_path, url_column)
            shop_urls = {u for u in remaining_urls if _url_contains_shopname(u)}
            if shop_urls:
                _append_log(task_id,
                            f"Shopname filter: dropping {len(shop_urls):,} of "
                            f"{len(remaining_urls):,} remaining URLs.")
                _filter_input_csv(input_path, url_column, shop_urls)

        # Short-circuit: every URL is cached — write output directly from the cache.
        if cached_urls and len(cached_urls) >= len(all_input_urls):
            try:
                from backend import rurl_optimizer_persistence as pers
                prev_df = pers.load_previous(list(cached_urls))
                xlsx_path = _write_xlsx_output(prev_df, output_path)
                _set(task_id, {
                    "status": "completed", "progress": 100,
                    "message": f"All {len(cached_urls):,} URLs served from cache",
                    "started_at": datetime.now().isoformat(),
                    "finished_at": datetime.now().isoformat(),
                    "output_path": str(xlsx_path), "script": "cache_only",
                })
                _append_log(task_id, f"Wrote {len(prev_df):,} cached rows to {xlsx_path.name}")
                _history_append(task_id)
            except Exception as e:
                _set(task_id, {"status": "failed", "error": f"cache-only write failed: {e}",
                               "finished_at": datetime.now().isoformat()})
                _history_append(task_id)
            return

        _run_subprocess(task_id, argv, output_path, script="main_parallel_v2")

        # 1. Upsert the fresh rows to rurl_processed BEFORE the global chain
        #    (so cached-row merging doesn't pollute global-pass input).
        if (_get(task_id) or {}).get("status") == "completed" and output_path.exists():
            try:
                from backend import rurl_optimizer_persistence as pers
                import pandas as pd
                fresh_df = pd.read_csv(output_path)
                n_up = pers.upsert_results(fresh_df)
                _append_log(task_id, f"Cached {n_up:,} fresh rows to rurl_processed.")
            except Exception as e:
                _append_log(task_id, f"[warn] persistence upsert failed: {e}")

        # 2. Optional: global R-URL chain (runs on fresh rows only — cached rows
        #    have no main_category column, which would confuse its keyword filter).
        if also_global and (_get(task_id) or {}).get("status") == "completed":
            global_out = OUTPUT_DIR / f"redirects_global_{task_id}_{ts}.csv"
            argv2 = [
                sys.executable, "-c",
                (
                    "import sys; sys.path.insert(0, '.'); "
                    "import process_global_rurls as g; "
                    f"g.INPUT_FILE = {str(output_path)!r}; "
                    f"g.OUTPUT_FILE = {str(global_out)!r}; "
                    "g.main()"
                ),
            ]
            _append_log(task_id, "--- Running global R-URL pass on fresh rows ---")
            _run_subprocess(task_id, argv2, global_out, script="process_global_rurls", optional_output=True)

        # 3. Merge cached rows into whatever the current output CSV is (main or global).
        if cached_urls and (_get(task_id) or {}).get("status") == "completed":
            try:
                from backend import rurl_optimizer_persistence as pers
                import pandas as pd
                current_csv = Path((_get(task_id) or {}).get("output_path", ""))
                if current_csv.suffix == ".csv" and current_csv.exists():
                    fresh_df = pd.read_csv(current_csv)
                    prev_df = pers.load_previous(list(cached_urls))
                    for col in fresh_df.columns:
                        if col not in prev_df.columns:
                            prev_df[col] = pd.NA
                    prev_df = prev_df[fresh_df.columns.tolist()]
                    final_df = pd.concat([fresh_df, prev_df], ignore_index=True)
                    final_df.to_csv(current_csv, index=False)
                    _append_log(task_id,
                                f"Merged {len(prev_df):,} cached rows with "
                                f"{len(fresh_df):,} freshly-processed -> "
                                f"{len(final_df):,} total.")
            except Exception as e:
                _append_log(task_id, f"[warn] cache merge failed: {e}")

        # 4. Emit a clear final summary + convert to xlsx.
        if (_get(task_id) or {}).get("status") == "completed":
            try:
                import pandas as pd
                final_csv = Path((_get(task_id) or {}).get("output_path", ""))
                if final_csv.suffix == ".csv" and final_csv.exists():
                    df = pd.read_csv(final_csv)
                    succ = int(df.get("success", pd.Series(dtype=bool)).sum()) \
                        if "success" in df.columns else len(df)
                    fresh_n = len(df) - len(cached_urls)
                    summary = (
                        f"==> Final: {len(df):,} total URLs "
                        f"({fresh_n:,} processed + {len(cached_urls):,} from cache), "
                        f"{succ:,} successful."
                    )
                    _append_log(task_id, summary)
                    xlsx_path = _write_xlsx_output(df, final_csv)
                    _set(task_id, {
                        "output_path": str(xlsx_path),
                        "message": f"Done. {len(df):,} URLs, {succ:,} successful -> {xlsx_path.name}",
                    })
                    _history_update_latest(task_id, {
                        "output_path": str(xlsx_path),
                        "message": f"Done. {len(df):,} URLs, {succ:,} successful",
                    })
            except Exception as e:
                _append_log(task_id, f"[warn] xlsx conversion failed: {e}")

        # Persist the final output bytes to Postgres so /download survives
        # any /tmp wipe / restart. Mark history with output_in_db=True so
        # the frontend only renders Export when bytes are durable.
        try:
            final_path = Path((_get(task_id) or {}).get("output_path", ""))
            if (_get(task_id) or {}).get("status") == "completed" and final_path.exists():
                from backend import rurl_optimizer_persistence as pers
                mime = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        if final_path.suffix.lower() == ".xlsx" else "text/csv")
                pers.save_run_output(task_id, ENGINE_VERSION, final_path.name, mime, final_path.read_bytes())
                _set(task_id, {"output_in_db": True})
                _history_update_latest(task_id, {"output_in_db": True})
        except Exception as e:
            _append_log(task_id, f"[warn] save_run_output failed: {e}")

    th = threading.Thread(target=_runner, daemon=True)
    _THREADS[task_id] = th
    th.start()
    return task_id


def _is_pid_alive(pid) -> bool:
    if not pid:
        return False
    try:
        os.kill(int(pid), 0)
        return True
    except (OSError, ValueError):
        return False


def _sweep_stale_tasks() -> None:
    """Finalize 'running' tasks whose runner thread is no longer alive.

    Only triggers when the runner thread has exited (so a task still in the
    Redshift fetch / Python orchestration phase isn't prematurely failed).
    If a final output file exists on disk we mark the task completed;
    otherwise failed.
    """
    with _TASKS_LOCK:
        ids = list(_TASKS.keys())
    for tid in ids:
        t = _get(tid) or {}
        if t.get("status") != "running":
            continue
        th = _THREADS.get(tid)
        if th is not None and th.is_alive():
            continue
        out_path = t.get("output_path") or ""
        if out_path and Path(out_path).exists():
            _set(tid, {
                "status": "completed",
                "progress": 100,
                "message": t.get("message") or f"Recovered. Output: {Path(out_path).name}",
                "finished_at": t.get("finished_at") or datetime.now().isoformat(),
            })
        else:
            _set(tid, {
                "status": "failed",
                "error": t.get("error") or "Process exited without finalizing status",
                "finished_at": t.get("finished_at") or datetime.now().isoformat(),
            })
        _history_append(tid)


def get_status(task_id: str) -> Optional[Dict[str, Any]]:
    _sweep_stale_tasks()
    return _get(task_id)


def cancel(task_id: str) -> bool:
    t = _get(task_id)
    if not t or t.get("status") not in ("queued", "running"):
        return False
    _set(task_id, {"cancel_requested": True})
    return True


def get_output_bytes(task_id: str):
    """Return (filename, mime, content_bytes) from the DB, or None."""
    try:
        from backend import rurl_optimizer_persistence as pers
        return pers.get_run_output(task_id)
    except Exception as e:
        logger.warning(f"get_run_output failed: {e}")
        return None


def get_history() -> list:
    """Return history with output_in_db flagged for the Export button."""
    _sweep_stale_tasks()
    try:
        from backend import rurl_optimizer_persistence as pers
        db_ids = pers.list_run_output_task_ids()
    except Exception as e:
        logger.warning(f"history db check failed: {e}")
        return list(_HISTORY)

    with _HISTORY_LOCK:
        updated = False
        for i, h in enumerate(_HISTORY):
            tid = h.get("task_id")
            in_db = tid in db_ids
            if in_db and not h.get("output_in_db"):
                _HISTORY[i] = {**h, "output_in_db": True}
                updated = True
        if updated:
            _save_history_to_disk()
        return list(_HISTORY)


def delete_history_entry(task_id: str) -> bool:
    """Remove a history entry and its output file. Returns True if removed."""
    removed = False
    with _HISTORY_LOCK:
        kept = [h for h in _HISTORY if h.get("task_id") != task_id]
        if len(kept) != len(_HISTORY):
            _HISTORY.clear()
            _HISTORY.extend(kept)
            _save_history_to_disk()
            removed = True

    out_path: Optional[str] = None
    with _TASKS_LOCK:
        t = _TASKS.pop(task_id, None)
        if t:
            out_path = t.get("output_path")
    if out_path:
        try:
            p = Path(out_path)
            if p.exists():
                p.unlink()
        except Exception as e:
            logger.warning(f"Failed to remove output file {out_path}: {e}")
    try:
        from backend import rurl_optimizer_persistence as pers
        pers.delete_run_output(task_id)
    except Exception as e:
        logger.warning(f"delete_run_output failed: {e}")
    return removed
