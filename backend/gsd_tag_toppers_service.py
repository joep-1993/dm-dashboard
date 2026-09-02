"""
GSD Tag Toppers Service
=======================

Bulk "add-only" onderhoud van tag_toppers-campagnes vanuit een Excel-lijst.

Per regel (shop_id / shop_name / country / productids):

  1. Zoek de `[label:tag_toppers]`-campagne van die shop. Bestaat die niet, maak
     hem aan (PAUSED, zoals GSD_tagtoppers.py) en neem meteen de negatieve
     zoekwoorden over van een zustercampagne van dezelfde shop.
  2. Voeg de product ids toe aan de listing-tree van die campagne. Dit is
     STRIKT ADD-ONLY: bestaande ids blijven staan. Er wordt nooit een boom
     afgebroken en opnieuw opgebouwd — dat is precies waar de losse
     GSD_tagtoppers.py-flow (`rebuild_tree_with_specific_item_ids`) wél toe
     overgaat, en waarom die hier niet gebruikt wordt.
  3. Sluit dezelfde ids uit in ALLE niet-REMOVED zustercampagnes van die shop
     (alles zonder `[label:tag_toppers]`), ook add-only: bestaande uitsluitingen
     blijven bestaan.

Identiteit van een shop = shopnaam ÉN shop_id. `shop_id` alleen is geen sleutel
(652237 is zowel Bruna.nl als Hubfootwear.com), dus daarop matchen zou items van
de ene shop in de campagnes van een andere zetten.

De dry-run leest alleen; er gaat geen enkele mutatie naar Google Ads tenzij
`start_run(dry_run=False)` wordt aangeroepen.
"""
import io
import json
import logging
import re
import threading
import time
from collections import OrderedDict, defaultdict
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from google.ads.googleads.client import GoogleAdsClient
from google.ads.googleads.errors import GoogleAdsException

from backend.database import get_db_connection, return_db_connection
from backend.gsd_campaigns_service import (
    GEO_TARGETS,
    TRACKING_TEMPLATES,
    _get_client,
    create_location_op,
    ensure_campaign_label_exists,
    get_negatives,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Direct Shopping accounts (CPR). Zelfde ids als GSD_tagtoppers.py.
COUNTRY_ACCOUNTS = {
    "NL": {"customer_id": "7938980174", "mc_id": "5592708765"},
    "BE": {"customer_id": "2454295509", "mc_id": "5588879919"},
    "DE": {"customer_id": "4192567576", "mc_id": "5342886105"},
}

TAG_TOPPERS_TOKEN = "[label:tag_toppers]"
CHANNEL_TOKEN = "[channel:directshopping]"
TAG_TOPPERS_AD_GROUP = "tag_toppers"
TAG_TOPPERS_SCRIPT_LABEL = "TAGTOPPERS_SCRIPT"

DEFAULT_BID_MICROS = 200_000      # €0,20 — gelijk aan GSD_tagtoppers.py
BUDGET_MICROS = 30_000_000        # €30/dag — gelijk aan GSD_tagtoppers.py

# Google Ads weigert erg grote mutates; ruim onder de limiet blijven.
MUTATE_CHUNK = 1000

# Pogingen voor één unit->subdivision conversie. CONCURRENT_MODIFICATION komt bij
# partial_failure als regel in de respons terug in plaats van als exception, dus
# _mutate_with_retry vangt hem daar niet; vooral de eerste convert in een ad group
# botst met de mutate die eraan voorafging.
CONVERT_RETRIES = 3

# Pogingen voor een hele rij. Een 503 uit de transportlaag slaat toe vóór er iets
# gepland is, dus zo'n rij levert een "Fout" met een lege uitklap op — er is dan
# letterlijk niets gebeurd. Opnieuw proberen mag: de tool is add-only en wat er al
# staat komt als "bestaat al" terug, dus een herhaling schrijft niets dubbel.
ROW_RETRIES = 3

# Rijen worden parallel verwerkt: het leeswerk is ~7s per rij en dat is bijna
# helemaal wachten op de API. Elke rij is een andere shop en dus een andere
# campagne, en gelijktijdige rijen op dezelfde shop worden alsnog geserialiseerd
# door _shop_lock.
RUN_WORKERS = 6

# Voorkeur bij meerdere zusters voor het overnemen van negatives.
SIBLING_LABEL_PREF = ["[label:a]", "[label:b]", "[label:c]", "[label:no_ean]", "[label:no_data]"]
NEG_MATCH_ORDER = {"EXACT": 0, "PHRASE": 1, "BROAD": 2}

# Een product id is meestal 26-28 alfanumerieke tekens, maar niet altijd: in de
# accounts staan ook ids van 7 (w2tjgr6, wqwgabp). Een ondergrens van 15 gooide die
# stilletjes weg. De telcel `number_of_productids` wordt nu op twee andere manieren
# uitgesloten: numerieke cellen slaan we over (Excel bewaart dat getal als int) en
# een token dat volledig uit cijfers bestaat telt niet als id.
_ID_RE = re.compile(r"^[A-Za-z0-9]{5,60}$")
_SPLIT_RE = re.compile(r"[;,|\s]+")

SHOP_RE = re.compile(r"\[shop:([^\]]+)\]")
SHOP_ID_RE = re.compile(r"\[shop_id:([^\]]+)\]")


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def _shop_key(name: Optional[str]) -> str:
    """Normaliseert een shopnaam voor vergelijking tussen campagnes van dezelfde shop.

    Case verschilt tussen campagnes van één shop ([shop:All4fysio.nl] vs
    [shop:all4fysio.nl]) en locale-suffixen zijn dezelfde shop (e5.be|NL ==
    e5.be/nl), dus knippen op de eerste | of /. Bewust conservatief: Decantalo.com
    en Decantalo.de blijven verschillend, net als Ubisoft.com en store.ubisoft.com.
    """
    return re.sub(r"\s*[|/].*$", "", (name or "").strip().lower())


def _clean_shop_name(name: Optional[str]) -> str:
    """Shopnaam zoals hij in een campagnenaam terechtkomt."""
    return (name or "").split("|")[0].strip()


def parse_workbook(data: bytes) -> Dict[str, Any]:
    """Leest de kandidaten-Excel.

    Verwacht kolommen A shop_id, B shop_name, C country, D productids,
    E number_of_productids. Product ids worden uit ALLE cellen vanaf kolom D
    gehaald en gesplitst op komma/puntkomma/pipe/whitespace, zodat een rij die
    over honderden cellen is uitgesmeerd (Excel's celgrens van 32.767 tekens)
    vanzelf goed gaat.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows: List[Dict[str, Any]] = []
    warnings: List[str] = []

    for excel_row, raw in enumerate(ws.iter_rows(values_only=True), start=1):
        if excel_row == 1:
            continue  # header
        if not raw or not any(v not in (None, "") for v in raw):
            continue

        shop_id = raw[0] if len(raw) > 0 else None
        shop_name = raw[1] if len(raw) > 1 else None
        country = raw[2] if len(raw) > 2 else None

        ids: "OrderedDict[str, None]" = OrderedDict()
        numeriek_overgeslagen = []
        for cell in raw[3:]:
            if cell in (None, ""):
                continue
            # Een int/float-cel is de telling, nooit een id.
            if isinstance(cell, (int, float)) and not isinstance(cell, bool):
                continue
            for token in _SPLIT_RE.split(str(cell)):
                token = token.strip().strip('"').strip("'")
                if not token or not _ID_RE.match(token):
                    continue
                if token.isdigit():
                    # Als tekst opgeslagen telling. Wordt gemeld, zodat een echt
                    # numeriek product id niet ongemerkt verdwijnt.
                    numeriek_overgeslagen.append(token)
                    continue
                # Google normaliseert item-ids naar kleine letters: de bomen bevatten
                # uitsluitend lowercase, ook waar de bron hoofdletters heeft. Hier
                # meteen normaliseren, anders vergelijkt de tool "4RjLg6oD…" met
                # "4rjlg6od…", ziet elk id als ontbrekend, en stuurt ops die Google
                # als duplicaat weigert. Zie de LEARNINGS bij Makro.nl.
                ids[token.lower()] = None

        shop_id_s = str(shop_id).strip() if shop_id is not None else ""
        shop_name_s = str(shop_name).strip() if shop_name else ""
        country_s = str(country).strip().upper() if country else ""

        if not shop_id_s or not shop_name_s:
            warnings.append(f"Rij {excel_row}: shop_id of shop_name ontbreekt — overgeslagen")
            continue
        if country_s not in COUNTRY_ACCOUNTS:
            warnings.append(f"Rij {excel_row}: onbekend land {country_s!r} — overgeslagen")
            continue
        if not ids:
            warnings.append(f"Rij {excel_row}: geen product ids gevonden — overgeslagen")
            continue

        if numeriek_overgeslagen:
            warnings.append(
                f"Rij {excel_row}: {len(numeriek_overgeslagen)} volledig numeriek(e) "
                f"token(s) overgeslagen als telling ({', '.join(numeriek_overgeslagen[:3])}) "
                f"— als dit product ids zijn, meld het"
            )

        stated = raw[4] if len(raw) > 4 else None
        stated_n = stated if isinstance(stated, int) else None
        if stated_n is not None and stated_n != len(ids):
            warnings.append(
                f"Rij {excel_row} ({shop_name_s} {country_s}): telcel zegt {stated_n} "
                f"maar er zijn {len(ids)} unieke ids gevonden"
            )

        rows.append({
            "excel_row": excel_row,
            "shop_id": shop_id_s,
            "shop_name": shop_name_s,
            "country": country_s,
            "item_ids": list(ids),
        })

    wb.close()
    return {
        "rows": rows,
        "warnings": warnings,
        "total_rows": len(rows),
        "total_ids": sum(len(r["item_ids"]) for r in rows),
    }


# ---------------------------------------------------------------------------
# Google Ads reads
# ---------------------------------------------------------------------------

def _customer_id(country: str) -> str:
    acct = COUNTRY_ACCOUNTS.get((country or "").upper())
    if not acct:
        raise ValueError(f"Onbekend land {country!r}")
    return acct["customer_id"]


def _fetch_shop_campaigns(client, customer_id: str, shop_id: str, shop_name: str) -> Dict[str, List[dict]]:
    """Alle niet-REMOVED campagnes van deze shop, gesplitst in tag_toppers en zusters.

    Matcht op shop_id (GAQL) en daarna op genormaliseerde shopnaam in Python,
    omdat GAQL's LIKE case-sensitive is en de naam per campagne in case verschilt.
    """
    ga = client.get_service("GoogleAdsService")
    q = f"""
        SELECT campaign.id, campaign.name, campaign.resource_name, campaign.status,
               campaign.shopping_setting.merchant_id
        FROM campaign
        WHERE campaign.name LIKE '%shop_id:{shop_id}]%'
          AND campaign.status != 'REMOVED'
    """
    want = _shop_key(shop_name)
    tag_toppers: List[dict] = []
    siblings: List[dict] = []
    for row in ga.search(customer_id=customer_id, query=q):
        name = row.campaign.name
        m = SHOP_RE.search(name)
        if not m or _shop_key(m.group(1)) != want:
            continue
        entry = {
            "id": row.campaign.id,
            "name": name,
            "resource": row.campaign.resource_name,
            "status": row.campaign.status.name,
            "merchant_id": int(row.campaign.shopping_setting.merchant_id or 0) or None,
        }
        if TAG_TOPPERS_TOKEN in name.lower():
            tag_toppers.append(entry)
        else:
            siblings.append(entry)
    tag_toppers.sort(key=lambda c: (c["status"] != "ENABLED", c["id"]))
    siblings.sort(key=lambda c: c["id"])
    return {"tag_toppers": tag_toppers, "siblings": siblings}


def _read_campaign_tree(client, customer_id: str, campaign_id: int) -> Dict[str, Dict[str, dict]]:
    """{ad_group_id: {resource_name: node}} voor alle listing-groups in een campagne.

    Eén query per campagne in plaats van één per ad group: een GSD-campagne heeft
    14 prijsbucket-ad-groups, dus dat scheelt bij 622 rijen duizenden calls.
    """
    ga = client.get_service("GoogleAdsService")
    q = f"""
        SELECT
          ad_group.id,
          ad_group.name,
          ad_group.status,
          ad_group_criterion.resource_name,
          ad_group_criterion.listing_group.type,
          ad_group_criterion.listing_group.parent_ad_group_criterion,
          ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
          ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
          ad_group_criterion.listing_group.case_value.product_item_id.value,
          ad_group_criterion.listing_group.case_value.product_brand.value,
          ad_group_criterion.listing_group.case_value.product_type.level,
          ad_group_criterion.listing_group.case_value.product_type.value,
          ad_group_criterion.listing_group.case_value.product_category.level,
          ad_group_criterion.listing_group.case_value.product_category.category_id,
          ad_group_criterion.listing_group.case_value.product_condition.condition,
          ad_group_criterion.listing_group.case_value.product_channel.channel,
          ad_group_criterion.negative,
          ad_group_criterion.cpc_bid_micros
        FROM ad_group_criterion
        WHERE campaign.id = {campaign_id}
          AND ad_group_criterion.type = 'LISTING_GROUP'
          AND ad_group_criterion.status != 'REMOVED'
          AND ad_group.status != 'REMOVED'
    """
    out: Dict[str, Dict[str, dict]] = {}
    for row in ga.search(customer_id=customer_id, query=q):
        agc = row.ad_group_criterion
        lg = agc.listing_group
        if lg.type_.name not in ("SUBDIVISION", "UNIT"):
            continue
        cv = lg.case_value
        which = cv._pb.WhichOneof("dimension")
        # which is None bij een OTHERS-node: die heeft geen eigen case value en erft
        # de dimensie van zijn siblings. dim blijft dan None — dat is NIET hetzelfde
        # als "item-id niveau", zie _level_dim().
        dim = index = value = item_id = level = None
        if which == "product_item_id":
            dim = "item_id"
            item_id = cv.product_item_id.value  # "" => OTHERS
        elif which == "product_custom_attribute":
            dim = "custom_attr"
            index = cv.product_custom_attribute.index.name
            value = cv.product_custom_attribute.value
        elif which == "product_brand":
            dim = "brand"
            value = cv.product_brand.value
        elif which == "product_type":
            dim = "product_type"
            level = cv.product_type.level.name
            value = cv.product_type.value
        elif which == "product_category":
            dim = "category"
            level = cv.product_category.level.name
            value = str(cv.product_category.category_id or "")
        elif which == "product_condition":
            dim = "condition"
            value = cv.product_condition.condition.name
        elif which == "product_channel":
            dim = "channel"
            value = cv.product_channel.channel.name
        ag_id = str(row.ad_group.id)
        out.setdefault(ag_id, {})[agc.resource_name] = {
            "resource": agc.resource_name,
            "ad_group_id": ag_id,
            "ad_group_name": row.ad_group.name,
            "type": lg.type_.name,
            "parent": lg.parent_ad_group_criterion or None,
            "dim": dim,
            "index": index,   # custom_attr: INDEX0..INDEX4
            "level": level,   # product_type / category: LEVEL1..LEVEL5
            "value": value,
            "item_id": item_id,
            "negative": bool(agc.negative),
            "bid": int(agc.cpc_bid_micros or 0),
        }
    return out


_AG_CPC_CACHE: Dict[Tuple[str, str], int] = {}
_AG_CPC_LOCK = threading.Lock()


def _ad_group_cpc(client, customer_id: str, ad_group_id: str) -> int:
    """Het default cpc_bid_micros van de ad group, gecachet per (klant, ad group).

    Nodig omdat _read_campaign_tree een erfelijk bod als 0 opslaat
    (`int(agc.cpc_bid_micros or 0)`). Zonder deze terugval kreeg een leaf die zijn
    bod van de ad group erft bij conversie hardgecodeerd DEFAULT_BID_MICROS (EUR 0,20)
    opgelegd — op een mogelijk ENABLED niet-tag-toppers-campagne. Zelfde aanpak als
    dma_exclusions_service._ad_group_cpc.
    """
    key = (str(customer_id), str(ad_group_id))
    with _AG_CPC_LOCK:
        if key in _AG_CPC_CACHE:
            return _AG_CPC_CACHE[key]
    val = 0
    try:
        ga = client.get_service("GoogleAdsService")
        q = (f"SELECT ad_group.cpc_bid_micros FROM ad_group "
             f"WHERE ad_group.id = {int(ad_group_id)}")
        for row in ga.search(customer_id=str(customer_id), query=q):
            val = int(row.ad_group.cpc_bid_micros or 0)
            break
    except GoogleAdsException as ex:
        logger.warning("cpc-lookup mislukt voor ad group %s: %s", ad_group_id, _err(ex))
    with _AG_CPC_LOCK:
        _AG_CPC_CACHE[key] = val
    return val


def _children_index(nodes: Dict[str, dict]) -> Dict[Optional[str], List[dict]]:
    """parent_resource -> kinderen, één keer opgebouwd per boom.

    _children() was een volledige scan over nodes.values(). _item_id_containers()
    riep daar via _level_dim/_level_spec voor ELKE subdivision op, en
    _convertible_leaves() voor elke UNIT — dus O(nodes^2) per boom, en dat maal
    per zuster maal per Excel-rij. Op een ad group met een attribuutniveau van 200
    waarden over een paar duizend item-ids loopt dat in de miljoenen iteraties.
    """
    idx: Dict[Optional[str], List[dict]] = defaultdict(list)
    for n in nodes.values():
        idx[n["parent"]].append(n)
    return idx


def _children(nodes: Dict[str, dict], parent_resource: Optional[str],
              idx: Optional[Dict[Optional[str], List[dict]]] = None) -> List[dict]:
    if idx is not None:
        return idx.get(parent_resource, [])
    return [n for n in nodes.values() if n["parent"] == parent_resource]


def _root(nodes: Dict[str, dict]) -> Optional[dict]:
    for n in nodes.values():
        if not n["parent"]:
            return n
    return None


# ---------------------------------------------------------------------------
# Planning: wat zou er gebeuren
# ---------------------------------------------------------------------------

def _plan_tag_toppers_adds(nodes: Dict[str, dict], item_ids: List[str]) -> Dict[str, Any]:
    """Welke ids ontbreken nog als POSITIEVE unit, en ónder welk knooppunt ze horen.

    In een boom die de tool zelf bouwt zit item-id direct onder de root. Maar niet elke
    tag_toppers-campagne is zo gebouwd: Notino's [label_test]-campagne heeft onder de
    root eerst een custom-attribute niveau en pas daaronder de item-ids. Blind onder de
    root hangen gaf daar "Dimension type of listing group must be the same as that of
    its siblings" voor alle 1105 ids. Het item-id-niveau wordt daarom opgezocht in
    plaats van aangenomen.
    """
    root = _root(nodes)
    if root is None:
        return {"root": None, "parent": None, "existing": set(),
                "missing": list(item_ids), "note": ""}

    parent, note = root, ""
    if _level_dim(nodes, root["resource"]) != "item_id":
        containers = _item_id_containers(nodes)
        if len(containers) == 1:
            parent = containers[0]
            note = "item-ids zitten een niveau dieper"
        elif len(containers) > 1:
            parent, note = None, (f"{len(containers)} item-id-niveaus in de boom — "
                                  "niet te bepalen waar de ids horen")
        elif _children(nodes, root["resource"]):
            parent, note = None, "geen item-id-niveau in de tag_toppers-boom"
        # geen kinderen onder de root: item-id is dan het niveau dat we zelf maken

    # Positief én negatief: Google staat maar één node per case value toe, dus een
    # id dat al als negatief onder deze root hangt kan er niet positief bij.
    existing = {
        n["item_id"] for n in nodes.values()
        if n["dim"] == "item_id" and n["item_id"]
    }
    missing = [i for i in item_ids if i not in existing]
    return {"root": root, "parent": parent, "existing": existing,
            "missing": missing, "note": note}


def _level_spec(nodes: Dict[str, dict], parent_resource: Optional[str],
                idx: Optional[Dict[Optional[str], List[dict]]] = None) -> Optional[dict]:
    """Hoe is het niveau ónder `parent_resource` opgedeeld: {'dim','index','level'}.

    Alle kinderen van één subdivision delen per definitie dezelfde dimensie, maar de
    OTHERS-node draagt hem niet: die komt terug zonder case value. Dimensie én de
    bijbehorende index/level zijn dus alleen af te lezen van de siblings die er wél
    een hebben.

    Eerder werd "UNIT zonder case value" als item-id niveau gelezen. Dat klopt alleen
    als de siblings item-ids zijn; bij een niveau op merk of producttype leverde het
    item-id-units onder brand-siblings op, en dus "Dimension type of listing group
    must be the same as that of its siblings".
    """
    spec = None
    for k in _children(nodes, parent_resource, idx):
        if not k["dim"]:
            continue
        if spec is None:
            spec = {"dim": k["dim"], "index": k["index"], "level": k["level"]}
        elif spec["dim"] != k["dim"]:
            return None   # gemengd niveau: bestaat niet volgens Google, dus niet raden
    return spec


def _level_dim(nodes: Dict[str, dict], parent_resource: Optional[str],
               idx: Optional[Dict[Optional[str], List[dict]]] = None) -> Optional[str]:
    spec = _level_spec(nodes, parent_resource, idx)
    return spec["dim"] if spec else None


def _item_id_containers(nodes: Dict[str, dict],
                        idx: Optional[Dict[Optional[str], List[dict]]] = None) -> List[dict]:
    """SUBDIVISIONs waarvan het niveau eronder écht op item-id zit."""
    idx = idx if idx is not None else _children_index(nodes)
    return [n for n in nodes.values()
            if n["type"] == "SUBDIVISION" and _level_dim(nodes, n["resource"], idx) == "item_id"]


def _convertible_leaves(nodes: Dict[str, dict],
                        idx: Optional[Dict[Optional[str], List[dict]]] = None) -> Tuple[List[dict], set]:
    """(leaves die we mogen omzetten, dimensies die we niet aankunnen).

    Een positieve biddable UNIT wordt een SUBDIVISION met item-id OTHERS (positief,
    originele bid) plus de negatieve ids. NEGATIEVE units zijn uitsluitingen en
    blijven met rust — die omzetten zou bestaande uitsluitingen wissen.

    De leaf wordt exact teruggebouwd uit het niveau-spec van zijn ouders plus zijn
    eigen waarde; een lege waarde is de OTHERS-node van dat niveau. Dimensies buiten
    WRITABLE_DIMS (categorie, staat, kanaal) kunnen we niet schrijven en worden
    gemeld in plaats van benaderd met een op die Google toch afkeurt.
    """
    leaves: List[dict] = []
    unsupported: set = set()
    idx = idx if idx is not None else _children_index(nodes)
    for n in nodes.values():
        if n["type"] != "UNIT" or n["negative"]:
            continue
        lvl = _level_spec(nodes, n["parent"], idx)
        if lvl is None:
            continue  # niveau niet te bepalen: niets doen is hier veiliger dan raden
        if lvl["dim"] == "item_id":
            continue  # zit al op item-id niveau
        if lvl["dim"] not in WRITABLE_DIMS:
            unsupported.add(lvl["dim"])
            continue
        # Eigen waarde als die er is, anders "" => dit ís de OTHERS-node.
        spec = dict(lvl)
        spec["value"] = n["value"] or ""
        leaves.append({**n, "spec": spec})
    return leaves, unsupported


def _plan_sibling_exclusions(nodes: Dict[str, dict], item_ids: List[str]) -> Dict[str, Any]:
    """Plan per ad group: waar komen de negatieve item-ids terecht.

    Twee vormen, gelijk aan wat de listing-tree van GSD kent:
      * er is al een item-id niveau  -> negatieve unit erbij hangen (goedkoop)
      * de leaf is een biddable UNIT -> omzetten naar SUBDIVISION met item-id
                                        OTHERS (positief, originele bid) + negatieven
    """
    idx = _children_index(nodes)
    containers = _item_id_containers(nodes, idx)
    appends: List[Dict[str, Any]] = []
    converts: List[Dict[str, Any]] = []
    unsupported: set = set()

    # GEEN if/else. Een boom kan BEIDE vormen tegelijk hebben — een tak die al op
    # item-id niveau zit én een tak die nog een biddable leaf is. Met een else bleef
    # `converts` in dat geval leeg, rapporteerde de ad group "niets te doen", en
    # bleef die tweede tak gewoon op de tag-topper-producten bieden. Dat is bereikbaar
    # in normaal gebruik: een half geconverteerde boom is precies wat de volgende run
    # inleest na een cancel of een mislukte convert.
    # Beide helften draaien is in de zuivere gevallen gedragsidentiek, want
    # _convertible_leaves() slaat een leaf die al op item_id-niveau zit zelf al over.
    for c in containers:
        kids = _children(nodes, c["resource"], idx)
        # Ook de POSITIEVE item-ids meetellen: één node per case value, dus een
        # id dat er positief hangt kan er niet negatief bij en levert anders
        # LISTING_GROUP_ALREADY_EXISTS op.
        already = {k["item_id"] for k in kids if k["dim"] == "item_id" and k["item_id"]}
        missing = [i for i in item_ids if i not in already]
        if missing:
            appends.append({"parent": c["resource"], "missing": missing})

    leaves, unsupported = _convertible_leaves(nodes, idx)
    for leaf in leaves:
        converts.append({"leaf": leaf, "missing": list(item_ids)})

    return {
        "appends": appends,
        "converts": converts,
        "unsupported": sorted(unsupported),
        "n_new": sum(len(a["missing"]) for a in appends) + sum(len(c["missing"]) for c in converts),
    }


# ---------------------------------------------------------------------------
# Negatives overnemen van een zustercampagne
# ---------------------------------------------------------------------------

def _neg_key(text: str, match_type: str) -> Tuple[str, str]:
    """Google matcht negatives case-insensitief maar bewaart 'emma' en 'Emma' wél
    als aparte criteria; dedupliceer daarom op (lowercase tekst, match type)."""
    return (str(text).strip().lower(), match_type)


def _fetch_campaign_negatives(client, customer_id: str, campaign_id: int) -> "OrderedDict":
    ga = client.get_service("GoogleAdsService")
    q = f"""
        SELECT campaign_criterion.keyword.text,
               campaign_criterion.keyword.match_type,
               campaign_criterion.status
        FROM campaign_criterion
        WHERE campaign.id = {campaign_id}
          AND campaign_criterion.type = 'KEYWORD'
          AND campaign_criterion.negative = TRUE
    """
    out: "OrderedDict" = OrderedDict()
    for row in ga.search(customer_id=customer_id, query=q):
        cc = row.campaign_criterion
        if cc.status.name == "REMOVED":
            continue
        out.setdefault(_neg_key(cc.keyword.text, cc.keyword.match_type.name),
                       (cc.keyword.text, cc.keyword.match_type.name))
    return out


def _pick_negatives_source(siblings: List[dict]) -> Optional[dict]:
    """ENABLED zuster heeft voorkeur, PAUSED is de fallback; daarbinnen label:a, b, c, ..."""
    for status in ("ENABLED", "PAUSED"):
        pool = [s for s in siblings if s["status"] == status and CHANNEL_TOKEN in s["name"].lower()]
        if not pool:
            continue
        for lbl in SIBLING_LABEL_PREF:
            hits = sorted([s for s in pool if lbl in s["name"].lower()], key=lambda s: s["id"])
            if hits:
                return hits[0]
        return sorted(pool, key=lambda s: s["id"])[0]
    return None


def _copy_negatives(client, customer_id: str, target_resource: str, target_id: int,
                    source: dict) -> Tuple[int, int, List[str]]:
    """Neemt de negatives van `source` over in de doelcampagne. Idempotent.

    Geeft (gepland, toegevoegd, fouten) terug. Gepland is expres apart: hiervoor
    werd alleen `added` teruggegeven en gaf de aanroeper dat door als ZOWEL planned
    als applied, waardoor planned per definitie gelijk was aan applied en 40 van de
    100 gelande negatives als "40 gepland / 40 toegepast / ok" op het scherm kwam.
    """
    src = _fetch_campaign_negatives(client, customer_id, source["id"])
    dst = _fetch_campaign_negatives(client, customer_id, target_id)
    missing = [v for k, v in src.items() if k not in dst]
    missing.sort(key=lambda x: (NEG_MATCH_ORDER.get(x[1], 9), x[0].lower()))
    if not missing:
        return 0, 0, []

    svc = client.get_service("CampaignCriterionService")
    ops = []
    for text, mt in missing:
        op = client.get_type("CampaignCriterionOperation")
        c = op.create
        c.campaign = target_resource
        c.negative = True
        c.keyword.text = text
        c.keyword.match_type = client.enums.KeywordMatchTypeEnum[mt]
        ops.append(op)

    added, errors = 0, []
    for i in range(0, len(ops), 200):
        chunk = ops[i:i + 200]
        # partial_failure MOET via een request-object; als kwarg wordt het geweigerd
        # door google-ads v28/v29.
        req = client.get_type("MutateCampaignCriteriaRequest")
        req.customer_id = str(customer_id)
        req.operations.extend(chunk)
        req.partial_failure = True
        try:
            resp = svc.mutate_campaign_criteria(request=req)
        except GoogleAdsException as ex:
            errors.append(_err(ex))
            continue
        # partial_failure staat aan, dus per-operatie-afwijzingen komen NIET als
        # exception binnen maar in resp.partial_failure_error. Die werd hier nooit
        # gelezen, zodat een afgekeurd zoekwoord spoorloos verdween.
        skipped, chunk_errors, failed_idx, retry_idx = _read_partial_failure(client, resp)
        errors.extend(chunk_errors)
        if retry_idx:
            errors.append(f"{len(retry_idx)} negative(s) tijdelijk geweigerd "
                          f"(concurrent modification) — niet opnieuw geprobeerd")
        added += sum(1 for r in resp.results if r.resource_name)
    return len(missing), added, errors


# ---------------------------------------------------------------------------
# Google Ads writes
# ---------------------------------------------------------------------------

class _Temp:
    """Per-mutate generator van tijdelijke (negatieve) criterion-ids."""

    def __init__(self):
        self.n = 0

    def path(self, client, customer_id, ad_group_id) -> str:
        self.n -= 1
        return client.get_service("AdGroupCriterionService").ad_group_criterion_path(
            customer_id, str(ad_group_id), str(self.n))


# Dimensies die we terug kunnen schrijven. Alles daarbuiten (categorie, staat,
# kanaal) wordt herkend bij het lezen maar nooit zelf aangemaakt.
WRITABLE_DIMS = ("item_id", "custom_attr", "brand", "product_type")


def _set_case_value(client, lg, spec: Dict[str, Any]) -> None:
    """Schrijft de case value van één node.

    Een lege `value` is de OTHERS-node van dat niveau: de dimensie moet dan wél
    gezet zijn maar zonder waarde, anders leest Google hem als een node zonder
    dimensie. Voor custom_attr en product_type hoort index respectievelijk level
    er ook bij een OTHERS-node bij — die komen van een sibling, want de OTHERS-node
    draagt ze zelf niet.
    """
    dim = spec["dim"]
    value = spec.get("value") or ""
    cv = lg.case_value
    if dim == "item_id":
        if value:
            cv.product_item_id.value = value
        else:
            client.copy_from(cv.product_item_id, client.get_type("ProductItemIdInfo"))
    elif dim == "custom_attr":
        cv.product_custom_attribute.index = \
            client.enums.ProductCustomAttributeIndexEnum[spec["index"]]
        if value:
            cv.product_custom_attribute.value = value
    elif dim == "brand":
        if value:
            cv.product_brand.value = value
        else:
            client.copy_from(cv.product_brand, client.get_type("ProductBrandInfo"))
    elif dim == "product_type":
        cv.product_type.level = client.enums.ProductTypeLevelEnum[spec["level"]]
        if value:
            cv.product_type.value = value
    else:
        raise ValueError(f"dimensie {dim!r} kan niet geschreven worden")


def _spec_from_legacy(item_id_value, custom_attr):
    if item_id_value is not None:
        return {"dim": "item_id", "value": item_id_value}
    if custom_attr is not None:
        return {"dim": "custom_attr", "index": custom_attr["index"],
                "value": custom_attr["value"]}
    return None


def _unit_op(client, customer_id, ad_group_id, temp, parent_resource, *,
             item_id_value=None, custom_attr=None, spec=None, negative=False, bid=None):
    op = client.get_type("AdGroupCriterionOperation")
    cr = op.create
    cr.resource_name = temp.path(client, customer_id, ad_group_id)
    cr.status = client.enums.AdGroupCriterionStatusEnum.ENABLED
    if bid and not negative:
        cr.cpc_bid_micros = bid
    lg = cr.listing_group
    lg.type_ = client.enums.ListingGroupTypeEnum.UNIT
    if parent_resource:
        lg.parent_ad_group_criterion = parent_resource
    spec = spec or _spec_from_legacy(item_id_value, custom_attr)
    if spec is not None:
        _set_case_value(client, lg, spec)
    if negative:
        cr.negative = True
    return op, cr.resource_name


def _subdiv_op(client, customer_id, ad_group_id, temp, parent_resource, *,
               custom_attr=None, spec=None):
    op = client.get_type("AdGroupCriterionOperation")
    cr = op.create
    cr.resource_name = temp.path(client, customer_id, ad_group_id)
    cr.status = client.enums.AdGroupCriterionStatusEnum.ENABLED
    lg = cr.listing_group
    lg.type_ = client.enums.ListingGroupTypeEnum.SUBDIVISION
    if parent_resource:
        lg.parent_ad_group_criterion = parent_resource
    spec = spec or _spec_from_legacy(None, custom_attr)
    if spec is not None:
        _set_case_value(client, lg, spec)
    return op, cr.resource_name


def _remove_op(client, resource_name):
    op = client.get_type("AdGroupCriterionOperation")
    op.remove = resource_name
    return op


def _dedupe_msgs(msgs: List[str]) -> List[str]:
    """Identieke meldingen samenvouwen tot 'melding (3x)'.

    Eén mislukte mutate levert per operatie dezelfde regel op; ongefilterd werd dat
    een muur van drie keer dezelfde zin, drie keer herhaald.
    """
    counts: "OrderedDict[str, int]" = OrderedDict()
    for m in msgs:
        key = str(m).strip()
        counts[key] = counts.get(key, 0) + 1
    return [f"{k} ({v}x)" if v > 1 else k for k, v in counts.items()]


def _is_transient(ex: Exception) -> bool:
    """Transportfouten waarbij niet de rij stuk is maar de verbinding.

    Google geeft onder belasting 503/UNAVAILABLE terug; dat komt langs als een
    google.api_core-exception, dus buiten GoogleAdsException om, en gooit in `work()`
    de hele rij eruit vóór er iets gepland is.
    """
    if type(ex).__name__ in ("ServiceUnavailable", "DeadlineExceeded",
                             "InternalServerError", "TooManyRequests", "Aborted"):
        return True
    tekst = str(ex).lower()
    return ("service is currently unavailable" in tekst
            or "deadline exceeded" in tekst
            or "try again later" in tekst)


def _err(ex: Exception) -> str:
    if isinstance(ex, GoogleAdsException):
        try:
            return "; ".join(e.message for e in ex.failure.errors)[:400]
        except Exception:
            return str(ex)[:400]
    return f"{type(ex).__name__}: {ex}"[:400]


# Transiënte Google-Ads-foutfamilies. Deze lijst is GELIJK aan die van
# gsd_campaigns_service._is_retryable_gads. Hij liep uiteen: daar werden ook
# internal_error en quota_error opnieuw geprobeerd, hier alleen
# CONCURRENT_MODIFICATION — zodat identieke Google-storingen door de ene module
# werden opgevangen en door de andere als rijfout gerapporteerd. Houd ze samen.
_RETRYABLE_GADS_FAMILIES = ("database_error", "internal_error", "quota_error")
_RETRYABLE_GADS_CODES = (
    "CONCURRENT_MODIFICATION", "INTERNAL_ERROR", "TRANSIENT_ERROR",
    "RESOURCE_EXHAUSTED", "RESOURCE_TEMPORARILY_EXHAUSTED",
)


def _is_concurrent_modification(ex: GoogleAdsException) -> bool:
    """True bij een transiënte Google-fout die veilig opnieuw te proberen is."""
    try:
        for e in ex.failure.errors:
            code = e.error_code
            for family in _RETRYABLE_GADS_FAMILIES:
                val = getattr(code, family, 0)
                name = getattr(val, "name", str(val))
                if name in _RETRYABLE_GADS_CODES:
                    return True
            if "same resource at once" in (e.message or ""):
                return True
    except Exception:
        pass
    msg = str(ex)
    return "CONCURRENT_MODIFICATION" in msg or "modify the same resource" in msg


def _is_already_exists(ex: GoogleAdsException) -> bool:
    """LISTING_GROUP_ALREADY_EXISTS is geen fout maar een no-op: de node staat er al.

    Dat is precies wat add-only hoort te doen, dus zo'n operatie telt als
    overgeslagen. Kan alsnog voorkomen als de boom een id al POSITIEF bevat waar wij
    een negatief willen zetten — Google staat maar één node per case value toe,
    ongeacht de negative-vlag.
    """
    try:
        for e in ex.failure.errors:
            code = e.error_code
            if getattr(code, "criterion_error", None) and \
                    code.criterion_error.name == "LISTING_GROUP_ALREADY_EXISTS":
                return True
            if "already exists" in (e.message or "").lower():
                return True
    except Exception:
        pass
    return False


def _mutate_with_retry(client, customer_id: str, ops: List[Any], retries: int = 4):
    """Voert één mutate uit en herhaalt bij CONCURRENT_MODIFICATION met backoff.

    partial_failure staat aan zodat één afgekeurde operatie niet het hele blok van
    maximaal MUTATE_CHUNK sloopt — zonder dat kostte één 'bestaat al' tot 1000
    geldige mutaties.
    """
    svc = client.get_service("AdGroupCriterionService")
    delay = 2.0
    last = None
    for attempt in range(retries):
        req = client.get_type("MutateAdGroupCriteriaRequest")
        req.customer_id = str(customer_id)
        req.operations.extend(ops)
        req.partial_failure = True
        try:
            return svc.mutate_ad_group_criteria(request=req), None
        except GoogleAdsException as ex:
            last = ex
            if not _is_concurrent_modification(ex) or attempt == retries - 1:
                break
            logger.warning("CONCURRENT_MODIFICATION, opnieuw over %.0fs (poging %d/%d)",
                           delay, attempt + 1, retries)
            time.sleep(delay)
            delay *= 2
    return None, last


def _read_partial_failure(client, resp) -> Tuple[int, List[str], set, set]:
    """(overgeslagen, echte fouten, mislukte indexen, opnieuw-te-proberen indexen).

    CONCURRENT_MODIFICATION komt met partial_failure niet meer als exception binnen
    maar als regel in de respons. Die apart teruggeven, want alleen daar hoort een
    retry op — de rest is een echte fout.
    """
    skipped, errors, failed, retryable = 0, [], set(), set()
    pfe = getattr(resp, "partial_failure_error", None)
    if not pfe or not pfe.details:
        return skipped, errors, failed, retryable
    failure_type = client.get_type("GoogleAdsFailure")
    for det in pfe.details:
        try:
            f = type(failure_type).deserialize(det.value)
        except Exception:
            continue
        for err in f.errors:
            idx = None
            for el in err.location.field_path_elements:
                if el.field_name == "operations":
                    idx = el.index
            if idx is not None:
                failed.add(idx)
            code = err.error_code
            msg = err.message or ""
            is_dup = (getattr(code, "criterion_error", None)
                      and code.criterion_error.name == "LISTING_GROUP_ALREADY_EXISTS") \
                or "already exists" in msg.lower()
            is_busy = (getattr(code, "database_error", None)
                       and code.database_error.name == "CONCURRENT_MODIFICATION") \
                or "same resource at once" in msg
            # Product-group-ops zijn atomair per ad group: één ongeldige op laat álle
            # andere ops voor diezelfde ad group falen met deze algemene melding. Dat
            # is geen eigen fout maar bijvangst, dus opnieuw indienen zónder de echt
            # ongeldige ops. Bij Makro.nl sloopten 28 duplicaten zo 392 goede ops.
            is_bijvangst = ("atomic within the same ad group" in msg
                            or "another operation targeting the same ad group" in msg)
            if is_dup:
                skipped += 1
            elif (is_busy or is_bijvangst) and idx is not None:
                retryable.add(idx)
            else:
                errors.append(msg)
    return skipped, errors, failed, retryable


def _submit_chunk(client, customer_id: str, ops: List[Any],
                  retries: int = 4) -> Tuple[int, int, List[str]]:
    """Eén blok uitvoeren, met retry op ALLEEN de operaties die botsten.

    De retry in _mutate_with_retry vangt CONCURRENT_MODIFICATION op het
    exception-pad. Met partial_failure aan komt diezelfde fout echter per operatie
    in de respons terug en liep hij die retry mis — vandaar dat we hier de
    afzonderlijke botsers opnieuw indienen in plaats van het hele blok.
    """
    pending = list(ops)
    done, skipped, errors = 0, 0, []
    delay = 2.0
    for attempt in range(retries):
        resp, ex = _mutate_with_retry(client, customer_id, pending)
        if resp is None:
            if _is_already_exists(ex):
                skipped += len(pending)
            else:
                errors.append(_err(ex))
            return done, skipped, errors

        s, errs, failed, retryable = _read_partial_failure(client, resp)
        skipped += s
        errors.extend(errs)
        done += sum(1 for j, r in enumerate(resp.results)
                    if r.resource_name and j not in failed)

        if not retryable:
            return done, skipped, errors
        if attempt == retries - 1:
            errors.append(f"na {retries} pogingen niet geland: "
                          f"{len(retryable)} operatie(s)")
            return done, skipped, errors

        # Alleen de te herhalen ops: de duplicaten en de echt ongeldige vallen eruit,
        # en juist dat maakt de volgende poging kansrijk bij de atomaire bijvangst.
        pending = [pending[j] for j in sorted(retryable) if j < len(pending)]
        logger.warning("%d operatie(s) opnieuw indienen over %.0fs (poging %d/%d)",
                       len(pending), delay, attempt + 1, retries)
        time.sleep(delay)
        delay *= 2
    return done, skipped, errors


def _mutate_criteria(client, customer_id: str, ops: List[Any]) -> Tuple[int, int, List[str]]:
    """Voert criterion-operaties uit in blokken.

    Geeft (geland, overgeslagen, fouten) terug. 'Overgeslagen' is uitsluitend
    LISTING_GROUP_ALREADY_EXISTS — de node stond er al, wat bij een add-only tool
    het gewenste eindresultaat is en dus geen fout.
    """
    done, skipped, errors = 0, 0, []
    for i in range(0, len(ops), MUTATE_CHUNK):
        # Cancel grijpt tussen de blokken. Alles is add-only, dus halverwege stoppen
        # laat een consistente boom achter en een volgende run vult de rest aan.
        if _cancelled():
            errors.append(f"afgebroken: {len(ops) - i} operatie(s) niet uitgevoerd")
            break
        d, s, e = _submit_chunk(client, customer_id, ops[i:i + MUTATE_CHUNK])
        done += d
        skipped += s
        errors.extend(e)
        _count_mutations(d + s)
    return done, skipped, errors


def _apply_tag_toppers_adds(client, customer_id: str, ad_group_id: str,
                            root_resource: str, missing: List[str]) -> Tuple[int, int, List[str]]:
    """Hangt ontbrekende ids als POSITIEVE units onder de bestaande root. Add-only."""
    if not missing:
        return 0, 0, []
    temp = _Temp()
    ops = []
    for item_id in missing:
        op, _ = _unit_op(client, customer_id, ad_group_id, temp, root_resource,
                         item_id_value=item_id, negative=False, bid=DEFAULT_BID_MICROS)
        ops.append(op)
    return _mutate_criteria(client, customer_id, ops)


def _apply_sibling_exclusions(client, customer_id: str, ad_group_id: str,
                              plan: Dict[str, Any]) -> Tuple[int, int, List[str]]:
    """Voegt negatieve item-ids toe. Bestaande uitsluitingen blijven staan."""
    done, skipped, errors = 0, 0, []

    for app in plan["appends"]:
        if _cancelled():
            break
        temp = _Temp()
        ops = []
        for item_id in app["missing"]:
            op, _ = _unit_op(client, customer_id, ad_group_id, temp, app["parent"],
                             item_id_value=item_id, negative=True)
            ops.append(op)
        d, s, e = _mutate_criteria(client, customer_id, ops)
        done += d
        skipped += s
        errors.extend(e)

    for conv in plan["converts"]:
        if _cancelled():
            break
        leaf = conv["leaf"]
        # Stap 1 is atomisch: de biddable leaf verdwijnt en wordt in dezelfde
        # mutate vervangen door een subdivision met item-id OTHERS, zodat de ad
        # group nooit even zonder dat targeting-pad zit. De subdivision krijgt exact
        # de case value van de leaf terug — of die nu op custom attribute, merk of
        # producttype zit — anders verschuift het targeting-pad.
        temp = _Temp()
        # Erft de leaf zijn bod van de ad group, dan leest hij hier als 0 en zette
        # `or DEFAULT_BID_MICROS` er EUR 0,20 op. Val terug op het echte ad-groepbod;
        # `or None` erachter is belangrijk voor auto-bidding groepen, die helemaal
        # geen cpc_bid_micros op de unit willen.
        bid = leaf["bid"] or _ad_group_cpc(client, customer_id, ad_group_id) or None
        ops = [_remove_op(client, leaf["resource"])]
        sub_op, sub_res = _subdiv_op(client, customer_id, ad_group_id, temp,
                                     leaf["parent"], spec=leaf["spec"])
        ops.append(sub_op)
        others_op, _ = _unit_op(client, customer_id, ad_group_id, temp, sub_res,
                                item_id_value="", negative=False, bid=bid)
        ops.append(others_op)
        # partial_failure staat aan, dus een afgekeurde subdivision-op komt niet als
        # exception binnen maar als lege resource name. Daar geen kinderen onder
        # hangen: die krijgen dan een lege parent en falen op REQUIRED_FIELD_MISSING,
        # wat de echte oorzaak juist verbergt.
        #
        # CONCURRENT_MODIFICATION komt langs dezelfde weg terug — als regel in de
        # respons, niet als exception — dus _mutate_with_retry ziet hem hier niet.
        # Zonder eigen lus sneuvelt vooral de eerste convert in een ad group, terwijl
        # Google nog bezig is met de vorige mutate. Opnieuw indienen is veilig: landde
        # de remove wel en de subdivision niet, dan slaagt de tweede poging juist
        # doordat de botsende node al weg is.
        real_sub, reden, als_overgeslagen = "", "", False
        for poging in range(CONVERT_RETRIES):
            resp, ex = _mutate_with_retry(client, customer_id, ops)
            if resp is None:
                if _is_already_exists(ex):
                    als_overgeslagen = True
                else:
                    reden = _err(ex)
                break
            real_sub = resp.results[1].resource_name if len(resp.results) > 1 else ""
            if real_sub:
                break
            s, perrs, _f, retryable = _read_partial_failure(client, resp)
            if s and not perrs and not retryable:
                als_overgeslagen = True   # stond er al: gewenste eindtoestand
                break
            reden = perrs[0] if perrs else "botsende mutatie op dezelfde boom"
            if not retryable or poging == CONVERT_RETRIES - 1:
                break
            time.sleep(2.0 * (poging + 1))

        if not real_sub:
            if als_overgeslagen:
                skipped += 1
            else:
                errors.append(f"convert {leaf['resource']}: "
                              + (reden or "subdivision niet aangemaakt"))
            continue

        temp2 = _Temp()
        neg_ops = []
        for item_id in conv["missing"]:
            op, _ = _unit_op(client, customer_id, ad_group_id, temp2, real_sub,
                             item_id_value=item_id, negative=True)
            neg_ops.append(op)
        d, s, e = _mutate_criteria(client, customer_id, neg_ops)
        done += d
        skipped += s
        errors.extend(e)

    return done, skipped, errors


def _merchant_id_from_campaigns(campaigns: List[dict]) -> Optional[int]:
    """Het Merchant Center id uit de AL OP SHOPNAAM GEFILTERDE campagnes van deze shop.

    Vervangt de losse GAQL-lookup, die alleen op `shop_id` matchte. De docstring
    bovenaan dit bestand zegt waarom dat niet mag: identiteit van een shop is
    shopnaam EN shop_id — 652237 is zowel Bruna.nl als Hubfootwear.com. Een
    shop_id-only lookup kon dus het MC-account van een ándere shop teruggeven en
    daarmee een nieuwe campagne aan het verkeerde account hangen. Deterministisch
    kiezen: het meest voorkomende id, bij gelijkspel het laagste.
    """
    counts: Dict[int, int] = {}
    for c in campaigns:
        mid = c.get("merchant_id")
        if mid:
            counts[mid] = counts.get(mid, 0) + 1
    if not counts:
        return None
    return sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))[0][0]


def _create_tag_toppers_campaign(client, customer_id: str, country: str,
                                 shop_id: str, shop_name: str,
                                 mc_id: Optional[int] = None) -> Tuple[Optional[dict], List[str]]:
    """Maakt een tag_toppers-campagne aan (PAUSED) met ad group, boomwortel en ad.

    Bewust NIET via gsd_campaigns_service.add_standard_shopping_campaign: die zet
    `feed_label` en een ander budget, terwijl de bestaande tag_toppers-campagnes
    dat niet hebben. Hier wordt de conventie van GSD_tagtoppers.py aangehouden.
    """
    errors: List[str] = []
    base_shop = _clean_shop_name(shop_name)
    campaign_name = (f"[shop:{base_shop}] [shop_id:{shop_id}] "
                     f"[channel:directshopping] [label:tag_toppers]")
    if mc_id is None:
        return None, ["geen Merchant Center id gevonden bij de bestaande campagnes "
                      "van deze shop — campagne niet aangemaakt"]

    budget_service = client.get_service("CampaignBudgetService")
    budget_op = client.get_type("CampaignBudgetOperation")
    budget = budget_op.create
    budget.name = f"budget_{base_shop}_{shop_id}_directshopping_tag_toppers_{int(time.time())}"
    budget.delivery_method = client.enums.BudgetDeliveryMethodEnum.STANDARD
    budget.amount_micros = BUDGET_MICROS
    budget.explicitly_shared = False
    try:
        budget_res = budget_service.mutate_campaign_budgets(
            customer_id=customer_id, operations=[budget_op]).results[0].resource_name
    except GoogleAdsException as ex:
        return None, [f"budget: {_err(ex)}"]

    campaign_service = client.get_service("CampaignService")
    camp_op = client.get_type("CampaignOperation")
    camp = camp_op.create
    camp.name = campaign_name
    camp.campaign_budget = budget_res
    camp.advertising_channel_type = client.enums.AdvertisingChannelTypeEnum.SHOPPING
    camp.shopping_setting.merchant_id = int(mc_id)
    camp.shopping_setting.campaign_priority = 0
    camp.shopping_setting.enable_local = True
    camp.tracking_url_template = TRACKING_TEMPLATES[country]
    camp.contains_eu_political_advertising = (
        client.enums.EuPoliticalAdvertisingStatusEnum.DOES_NOT_CONTAIN_EU_POLITICAL_ADVERTISING
    )
    camp.status = client.enums.CampaignStatusEnum.PAUSED
    camp.manual_cpc.enhanced_cpc_enabled = False
    try:
        camp_res = campaign_service.mutate_campaigns(
            customer_id=customer_id, operations=[camp_op]).results[0].resource_name
    except GoogleAdsException as ex:
        return None, [f"campagne: {_err(ex)}"]

    campaign_id = int(camp_res.split("/")[-1])
    time.sleep(2)  # laten propageren voor de child-objecten

    try:
        crit_service = client.get_service("CampaignCriterionService")
        crit_service.mutate_campaign_criteria(
            customer_id=customer_id, operations=[create_location_op(client, camp_res, country)])
    except Exception as ex:
        errors.append(f"geo: {_err(ex)}")

    try:
        label_res = ensure_campaign_label_exists(client, customer_id, TAG_TOPPERS_SCRIPT_LABEL)
        if label_res:
            lbl_op = client.get_type("CampaignLabelOperation")
            lbl_op.create.campaign = camp_res
            lbl_op.create.label = label_res
            client.get_service("CampaignLabelService").mutate_campaign_labels(
                customer_id=customer_id, operations=[lbl_op])
    except Exception as ex:
        errors.append(f"label: {_err(ex)}")

    ad_group_service = client.get_service("AdGroupService")
    ag_op = client.get_type("AdGroupOperation")
    ag = ag_op.create
    ag.campaign = camp_res
    ag.name = TAG_TOPPERS_AD_GROUP
    ag.cpc_bid_micros = DEFAULT_BID_MICROS
    ag.status = client.enums.AdGroupStatusEnum.ENABLED
    try:
        ag_res = ad_group_service.mutate_ad_groups(
            customer_id=customer_id, operations=[ag_op]).results[0].resource_name
    except GoogleAdsException as ex:
        return None, errors + [f"ad group: {_err(ex)}"]

    ad_group_id = ag_res.split("/")[-1]
    time.sleep(1)

    # Boomwortel: root SUBDIVISION + item-id OTHERS NEGATIEF. Daarmee toont de
    # campagne niets tot de ids eronder gehangen worden (inclusieve logica).
    temp = _Temp()
    root_op, root_tmp = _subdiv_op(client, customer_id, ad_group_id, temp, None, custom_attr=None)
    others_op, _ = _unit_op(client, customer_id, ad_group_id, temp, root_tmp,
                            item_id_value="", negative=True)
    # Via _mutate_with_retry: op een ad group die net is aangemaakt is Google intern
    # nog bezig, en dan komt CONCURRENT_MODIFICATION terug ("same resource at once").
    # Die is transient — retryen met backoff in plaats van de halve campagne laten
    # stranden en een lege campagne + budget achterlaten.
    resp, ex = _mutate_with_retry(client, customer_id, [root_op, others_op])
    if resp is None:
        return None, errors + [f"boomwortel: {_err(ex)}"]
    # Beide ops moeten geland zijn. _mutate_with_retry zet partial_failure aan, dus
    # een mislukte OTHERS-op geeft geen exception — en zonder die negatieve OTHERS
    # toont de campagne álle producten in plaats van alleen de tag toppers.
    landed = [r.resource_name for r in resp.results] if resp.results else []
    if len(landed) < 2 or not all(landed[:2]):
        _s, perrs, _f, _r = _read_partial_failure(client, resp)
        return None, errors + [
            "boomwortel: " + (perrs[0] if perrs else "root of item-id OTHERS niet aangemaakt")]
    root_resource = landed[0]

    time.sleep(1)
    try:
        ad_op = client.get_type("AdGroupAdOperation")
        ad = ad_op.create
        ad.ad_group = ag_res
        ad.status = client.enums.AdGroupAdStatusEnum.ENABLED
        client.copy_from(ad.ad.shopping_product_ad, client.get_type("ShoppingProductAdInfo"))
        client.get_service("AdGroupAdService").mutate_ad_group_ads(
            customer_id=customer_id, operations=[ad_op])
    except GoogleAdsException as ex:
        errors.append(f"shopping ad: {_err(ex)}")

    return {
        "id": campaign_id,
        "name": campaign_name,
        "resource": camp_res,
        "status": "PAUSED",
        "ad_group_id": ad_group_id,
        "root_resource": root_resource,
    }, errors


def _ensure_tag_toppers_tree(client, customer_id: str, campaign: dict):
    """Herstelt een tag_toppers-campagne die bestaat maar geen listing-tree heeft.

    Ontstaat als _create_tag_toppers_campaign strandde ná campagne en ad group maar
    vóór de boomwortel: wat overblijft is een campagne die niets kan vertonen, en die
    elke volgende run als "bestaat" leest en afserveert met "geen listing-tree
    gevonden". Zonder herstel blijft dat zo, want de aanmaak-tak wordt nooit meer
    bereikt. Kalenderwinkel.nl (24121408709) was er zo een: 1 ad group, 0 criteria,
    0 advertenties.

    Geeft (ad_group_id, root_resource, errors); root_resource is None als het misging.
    """
    ga = client.get_service("GoogleAdsService")
    errors: List[str] = []

    q = f"""
        SELECT ad_group.id, ad_group.name FROM ad_group
        WHERE campaign.id = {campaign['id']} AND ad_group.status != 'REMOVED'
    """
    ags = [(str(r.ad_group.id), r.ad_group.name) for r in ga.search(customer_id=customer_id, query=q)]
    ag_id = next((i for i, name in ags if name == TAG_TOPPERS_AD_GROUP), None)
    if ag_id is None and ags:
        ag_id = ags[0][0]
    if ag_id is None:
        ag_op = client.get_type("AdGroupOperation")
        ag = ag_op.create
        ag.campaign = campaign["resource"]
        ag.name = TAG_TOPPERS_AD_GROUP
        ag.cpc_bid_micros = DEFAULT_BID_MICROS
        ag.status = client.enums.AdGroupStatusEnum.ENABLED
        try:
            ag_res = client.get_service("AdGroupService").mutate_ad_groups(
                customer_id=customer_id, operations=[ag_op]).results[0].resource_name
            ag_id = ag_res.split("/")[-1]
            time.sleep(1)
        except GoogleAdsException as ex:
            return None, None, errors + [f"ad group herstellen: {_err(ex)}"]

    temp = _Temp()
    root_op, root_tmp = _subdiv_op(client, customer_id, ag_id, temp, None, spec=None)
    others_op, _ = _unit_op(client, customer_id, ag_id, temp, root_tmp,
                            item_id_value="", negative=True)
    resp, ex = _mutate_with_retry(client, customer_id, [root_op, others_op])
    if resp is None:
        return None, None, errors + [f"boomwortel herstellen: {_err(ex)}"]
    landed = [r.resource_name for r in resp.results] if resp.results else []
    if len(landed) < 2 or not all(landed[:2]):
        _s, perrs, _f, _r = _read_partial_failure(client, resp)
        return None, None, errors + [
            "boomwortel herstellen: "
            + (perrs[0] if perrs else "root of item-id OTHERS niet aangemaakt")]

    # Zonder shopping ad vertoont de campagne nog steeds niets, en die ontbreekt bij
    # precies dezelfde afgebroken aanmaak.
    try:
        q2 = (f"SELECT ad_group_ad.resource_name FROM ad_group_ad "
              f"WHERE ad_group.id = {ag_id} AND ad_group_ad.status != 'REMOVED'")
        if not list(ga.search(customer_id=customer_id, query=q2)):
            time.sleep(1)
            ad_op = client.get_type("AdGroupAdOperation")
            ad = ad_op.create
            ad.ad_group = f"customers/{customer_id}/adGroups/{ag_id}"
            ad.status = client.enums.AdGroupAdStatusEnum.ENABLED
            client.copy_from(ad.ad.shopping_product_ad, client.get_type("ShoppingProductAdInfo"))
            client.get_service("AdGroupAdService").mutate_ad_group_ads(
                customer_id=customer_id, operations=[ad_op])
    except GoogleAdsException as ex:
        errors.append(f"shopping ad herstellen: {_err(ex)}")

    return ag_id, landed[0], errors


# ---------------------------------------------------------------------------
# Run orchestration
# ---------------------------------------------------------------------------

_state_lock = threading.Lock()
_state: Dict[str, Any] = {
    "running": False,
    "dry_run": True,
    "current": 0,
    "total": 0,
    "cancel": False,
    "results": [],
    "started_at": None,
    "finished_at": None,
    "summary": {},
    # Aantal geschreven criteria. De voortgangsbalk telt rijen, en bij weinig rijen
    # met veel werk (Toolmax: 2 rijen, 33.536 mutaties) beweegt die minutenlang niet.
    # Deze teller loopt wél door en laat zien dat er iets gebeurt.
    "mutations": 0,
}


def _cancelled() -> bool:
    with _state_lock:
        return bool(_state["cancel"])


def _count_mutations(n: int) -> None:
    if not n:
        return
    with _state_lock:
        _state["mutations"] += n


def get_progress() -> Dict[str, Any]:
    with _state_lock:
        return {
            "running": _state["running"],
            "dry_run": _state["dry_run"],
            "current": _state["current"],
            "total": _state["total"],
            "started_at": _state["started_at"].isoformat() if _state["started_at"] else None,
            "finished_at": _state["finished_at"].isoformat() if _state["finished_at"] else None,
            "mutations": _state["mutations"],
            "cancelling": bool(_state["cancel"]) and _state["running"],
            # Meegestuurd zodat de tegels elke poll kunnen meelopen. De tabel ophalen
            # is duur (honderden rijen), deze dict is een handvol getallen.
            "summary": dict(_state["summary"]),
        }


def get_results() -> Dict[str, Any]:
    with _state_lock:
        return {
            "running": _state["running"],
            "dry_run": _state["dry_run"],
            "results": list(_state["results"]),
            "summary": dict(_state["summary"]),
        }


def cancel_run() -> None:
    with _state_lock:
        _state["cancel"] = True


def _process_row(client, row: Dict[str, Any], dry_run: bool) -> Dict[str, Any]:
    """Eén Excel-regel: plannen en (als dry_run False) uitvoeren."""
    country = row["country"]
    customer_id = _customer_id(country)
    # Ook hier normaliseren: een rij kan behalve uit de Excel ook uit
    # gsd_tag_toppers_items komen, en daar staan de ids uit oudere imports nog met
    # hoofdletters. Dedupe met behoud van volgorde, want twee ids die alleen in case
    # verschillen zijn voor Google hetzelfde id.
    item_ids = list(OrderedDict((i.lower(), None) for i in row["item_ids"]))

    res: Dict[str, Any] = {
        "excel_row": row["excel_row"],
        "shop_id": row["shop_id"],
        "shop_name": row["shop_name"],
        "country": country,
        "n_ids": len(item_ids),
        "campaign_action": "",
        "campaign_name": "",
        # 1 zodra de campagne er echt staat. `campaign_action == "aanmaken"` zegt
        # alleen dát er een campagne nodig was: bij een mislukte create blijft die
        # op "aanmaken" staan, dus tellen op dat veld overschat wat er is aangemaakt.
        "campaign_created": 0,
        "ids_already_present": 0,
        "ids_to_add": 0,
        "ids_added": 0,
        "siblings": 0,
        "sibling_ad_groups": 0,
        "exclusions_to_add": 0,
        "exclusions_added": 0,
        "negatives_source": "",
        "negatives_copied": 0,
        "status": "ok",
        "errors": [],
        # Per campagne/ad group wat er gepland en wat er echt geland is. De tabel
        # klapt hierop uit, zodat een run niet alleen een totaal maar ook een
        # aanwijsbare plek van mislukking geeft.
        "targets": [],
    }

    def target(kind, campaign, planned, applied=0, ad_group_id=None, errors=None,
               note="", skipped=0):
        errs = _dedupe_msgs([str(e)[:300] for e in (errors or [])])
        if errs:
            status = "fout" if applied == 0 else "deels"
        elif dry_run:
            status = "gepland"
        elif skipped and applied == 0:
            # Niets geschreven omdat alles er al stond: dat is het gewenste
            # eindresultaat van een add-only tool, geen mislukking.
            status = "overgeslagen"
        elif applied + skipped >= planned:
            status = "ok"
        else:
            status = "deels" if applied else "fout"
        res["targets"].append({
            "kind": kind,
            "campaign": campaign,
            "ad_group_id": ad_group_id,
            "planned": planned,
            "applied": applied,
            "skipped": skipped,
            "status": status,
            "note": note,
            "errors": errs[:5],
        })

    camps = _fetch_shop_campaigns(client, customer_id, row["shop_id"], row["shop_name"])
    siblings = camps["siblings"]
    res["siblings"] = len(siblings)

    # ---- 1/2: tag_toppers-campagne -------------------------------------
    tt = camps["tag_toppers"][0] if camps["tag_toppers"] else None

    if tt is None:
        res["campaign_action"] = "aanmaken"
        res["campaign_name"] = (f"[shop:{_clean_shop_name(row['shop_name'])}] "
                                f"[shop_id:{row['shop_id']}] [channel:directshopping] "
                                f"[label:tag_toppers]")
        res["ids_to_add"] = len(item_ids)
        source = _pick_negatives_source(siblings)
        res["negatives_source"] = source["name"] if source else "geen zuster gevonden"

        if not dry_run:
            # MC-id uit de zusters van DEZE shop (naam + shop_id), niet uit een
            # shop_id-only lookup die een andere shop kan aanwijzen.
            created, errs = _create_tag_toppers_campaign(
                client, customer_id, country, row["shop_id"], row["shop_name"],
                mc_id=_merchant_id_from_campaigns(siblings + camps["tag_toppers"]))
            res["errors"].extend(errs)
            if created is None:
                target("aanmaken", res["campaign_name"], 1, 0, errors=errs)
                res["status"] = "fout"
                return res
            res["campaign_name"] = created["name"]
            res["campaign_created"] = 1
            target("aanmaken", created["name"], 1, 1, errors=errs, note="PAUSED")
            added, skipped, errs2 = _apply_tag_toppers_adds(
                client, customer_id, created["ad_group_id"], created["root_resource"], item_ids)
            res["ids_added"] = added
            res["errors"].extend(errs2)
            target("toevoegen", created["name"], len(item_ids), added,
                   ad_group_id=created["ad_group_id"], errors=errs2, skipped=skipped)
            if source:
                planned3, n, errs3 = _copy_negatives(client, customer_id, created["resource"],
                                                     created["id"], source)
                res["negatives_copied"] = n
                res["errors"].extend(errs3)
                # Geen "bron: <campagnenaam>" in de note: die naam is bijna even lang
                # als de rij zelf en de zuster is al af te leiden uit de shop.
                target("negatives", created["name"], planned3, n, errors=errs3)
        else:
            target("aanmaken", res["campaign_name"], 1, 0, note="PAUSED")
            target("toevoegen", res["campaign_name"], len(item_ids), 0)
            if source:
                src_negs = _fetch_campaign_negatives(client, customer_id, source["id"])
                res["negatives_copied"] = len(src_negs)
                target("negatives", res["campaign_name"], len(src_negs), 0)
    else:
        res["campaign_action"] = "bestaat"
        res["campaign_name"] = tt["name"]
        trees = _read_campaign_tree(client, customer_id, tt["id"])
        # de tag_toppers-campagne heeft één ad group; pak de grootste boom als er meer zijn
        ag_id, nodes = (max(trees.items(), key=lambda kv: len(kv[1]))
                        if trees else (None, {}))
        plan = _plan_tag_toppers_adds(nodes, item_ids)
        res["ids_already_present"] = len(item_ids) - len(plan["missing"])
        res["ids_to_add"] = len(plan["missing"])

        if not dry_run and plan["missing"]:
            if plan["root"] is None:
                # Campagne zonder boom: herstellen in plaats van elke run opnieuw
                # dezelfde fout melden. Zo'n campagne kan tot dat moment niets vertonen.
                heal_ag, heal_root, heal_errs = _ensure_tag_toppers_tree(
                    client, customer_id, tt)
                res["errors"].extend(heal_errs)
                if heal_root is None:
                    msg = "geen listing-tree gevonden en herstel mislukt"
                    res["errors"].append(msg)
                    res["status"] = "fout"
                    target("toevoegen", tt["name"], len(plan["missing"]), 0,
                           ad_group_id=ag_id, errors=heal_errs or [msg])
                else:
                    ag_id = heal_ag
                    added, skipped, errs = _apply_tag_toppers_adds(
                        client, customer_id, ag_id, heal_root, plan["missing"])
                    res["ids_added"] = added
                    res["errors"].extend(errs)
                    target("toevoegen", tt["name"], len(plan["missing"]), added,
                           ad_group_id=ag_id, errors=errs, skipped=skipped,
                           note="boom hersteld")
                    # De afgebroken aanmaak kwam nooit tot de negatives, dus die
                    # ontbreken hier per definitie — geverifieerd: alle 12 achtergebleven
                    # campagnes hadden er 0. _copy_negatives vergelijkt met wat er al
                    # staat, dus dit blijft idempotent bij een volgende run.
                    source = _pick_negatives_source(siblings)
                    res["negatives_source"] = source["name"] if source else "geen zuster gevonden"
                    if source:
                        planned_neg, n_neg, errs3 = _copy_negatives(
                            client, customer_id, tt["resource"], tt["id"], source)
                        res["negatives_copied"] = n_neg
                        res["errors"].extend(errs3)
                        target("negatives", tt["name"], planned_neg, n_neg, errors=errs3)
            elif plan["parent"] is None:
                # Boomvorm die we niet veilig kunnen aanvullen: melden in plaats van de
                # ids naast een sibling van een andere dimensie hangen.
                target("toevoegen", tt["name"], len(plan["missing"]), 0,
                       ad_group_id=ag_id, note=plan["note"])
            else:
                added, skipped, errs = _apply_tag_toppers_adds(
                    client, customer_id, ag_id, plan["parent"]["resource"], plan["missing"])
                res["ids_added"] = added
                res["errors"].extend(errs)
                target("toevoegen", tt["name"], len(plan["missing"]), added,
                       ad_group_id=ag_id, errors=errs, skipped=skipped, note=plan["note"])
        elif plan["missing"]:
            target("toevoegen", tt["name"], len(plan["missing"]), 0, ad_group_id=ag_id,
                   note=("campagne heeft geen boom — wordt hersteld"
                         if plan["root"] is None else plan["note"]))
        else:
            target("toevoegen", tt["name"], 0, 0, ad_group_id=ag_id,
                   note="alle ids stonden er al")

    # ---- 3: uitsluiten in de zustercampagnes ---------------------------
    for sib in siblings:
        if _cancelled():
            break
        try:
            trees = _read_campaign_tree(client, customer_id, sib["id"])
        except GoogleAdsException as ex:
            msg = _err(ex)
            res["errors"].append(f"{sib['name']}: {msg}")
            target("uitsluiten", sib["name"], 0, 0, errors=[msg])
            continue
        for ag_id, nodes in trees.items():
            if not nodes:
                continue
            res["sibling_ad_groups"] += 1
            plan = _plan_sibling_exclusions(nodes, item_ids)
            res["exclusions_to_add"] += plan["n_new"]
            if not dry_run and plan["n_new"]:
                done, skipped, errs = _apply_sibling_exclusions(client, customer_id, ag_id, plan)
                res["exclusions_added"] += done
                res["errors"].extend(f"{sib['name']}/{ag_id}: {e}" for e in errs)
                target("uitsluiten", sib["name"], plan["n_new"], done,
                       ad_group_id=ag_id, errors=errs, skipped=skipped)
            else:
                # Een niveau op merk/producttype/categorie kunnen we niet terugbouwen.
                # Dat expliciet melden, anders leest het als "niets te doen" terwijl de
                # uitsluitingen daar wél nodig zijn.
                onbekend = plan.get("unsupported") or []
                if onbekend:
                    note = ("niveau op " + "/".join(onbekend)
                            + " — uitsluiten hier niet ondersteund")
                elif not plan["n_new"]:
                    note = "niets te doen"
                else:
                    note = ""
                target("uitsluiten", sib["name"], plan["n_new"], 0, ad_group_id=ag_id,
                       note=note)

    if res["errors"]:
        res["status"] = "fout" if res["status"] == "fout" else "deels"
    res["errors"] = _dedupe_msgs(res["errors"])[:10]
    return res


def _failed_row(row: Dict[str, Any], message: str) -> Dict[str, Any]:
    return {
        "excel_row": row.get("excel_row"),
        "shop_id": row.get("shop_id"),
        "shop_name": row.get("shop_name"),
        "country": row.get("country"),
        "n_ids": len(row.get("item_ids") or []),
        "campaign_action": "", "campaign_name": "",
        "ids_already_present": 0, "ids_to_add": 0, "ids_added": 0,
        "siblings": 0, "sibling_ad_groups": 0,
        "exclusions_to_add": 0, "exclusions_added": 0,
        "negatives_source": "", "negatives_copied": 0,
        "status": "fout", "errors": [message], "targets": [],
        "campaign_created": 0,
    }


# Per (account, shop) serialisatie. In de huidige lijst is elke (land, shop_id)
# uniek, dus twee workers raken nooit dezelfde campagne — maar een volgend bestand
# kan dat wel bevatten, en twee gelijktijdige mutates op dezelfde boom lezen een
# stale tree en racen (verloren nodes / CONCURRENT_MODIFICATION).
_shop_locks: Dict[Tuple[str, str], threading.Lock] = {}
_shop_locks_guard = threading.Lock()


def _shop_lock(customer_id: str, shop_id: str) -> threading.Lock:
    key = (str(customer_id), str(shop_id))
    lk = _shop_locks.get(key)
    if lk is None:
        with _shop_locks_guard:
            lk = _shop_locks.get(key)
            if lk is None:
                lk = threading.Lock()
                _shop_locks[key] = lk
    return lk


def _summarize(results: List[Dict[str, Any]]) -> Dict[str, Any]:
    """De tegels boven de resultatentabel. Draait ook tijdens de run over de tot dan
    toe afgeronde rijen, zodat de tegels meelopen in plaats van op 0 te blijven tot
    het eind. O(n) per afgeronde rij is verwaarloosbaar bij een paar honderd rijen."""
    return {
        "rows": len(results),
        # to_create = wat er nodig was (de preview-kant), created = wat er echt staat.
        "campaigns_to_create": sum(1 for r in results if r["campaign_action"] == "aanmaken"),
        "campaigns_created": sum(r.get("campaign_created", 0) for r in results),
        "ids_to_add": sum(r["ids_to_add"] for r in results),
        "ids_added": sum(r["ids_added"] for r in results),
        "exclusions_to_add": sum(r["exclusions_to_add"] for r in results),
        "exclusions_added": sum(r["exclusions_added"] for r in results),
        "negatives_copied": sum(r["negatives_copied"] for r in results),
        "errors": sum(1 for r in results if r["status"] != "ok"),
    }


def _run(rows: List[Dict[str, Any]], dry_run: bool) -> None:
    from concurrent.futures import ThreadPoolExecutor, as_completed

    # _get_client() MOET binnen de try staan die _state["running"] weer vrijgeeft.
    # Stond hij ervoor, dan liet een ontbrekende of ongeldige credential de thread
    # sterven vóór de finally: de vlag bleef staan, elke volgende run kreeg 409
    # "Er loopt al een run", en omdat uvicorn zonder --reload draait haalde alleen
    # een handmatige herstart dat weg. `client` is een closure-variabele van work(),
    # dat pas binnen de try wordt aangeroepen, dus dit is verder gedragsneutraal.
    client = None
    results: List[Dict[str, Any]] = []
    done_count = 0

    def work(row):
        laatste = None
        for poging in range(ROW_RETRIES):
            with _state_lock:
                if _state["cancel"]:
                    return None
            try:
                with _shop_lock(_customer_id(row["country"]), row["shop_id"]):
                    return _process_row(client, row, dry_run)
            except Exception as ex:
                laatste = ex
                if not _is_transient(ex) or poging == ROW_RETRIES - 1:
                    logger.exception("Tag Toppers rij %s mislukt", row.get("excel_row"))
                    break
                wacht = 3.0 * (poging + 1)
                logger.warning("Tag Toppers rij %s: %s — opnieuw over %.0fs",
                               row.get("excel_row"), _err(ex), wacht)
                time.sleep(wacht)
        return _failed_row(row, _err(laatste))

    fatal: Optional[str] = None
    try:
        client = _get_client()
        with ThreadPoolExecutor(max_workers=RUN_WORKERS) as pool:
            futures = {pool.submit(work, r): r for r in rows}
            for fut in as_completed(futures):
                res = fut.result()
                done_count += 1
                if res is not None:
                    results.append(res)
                with _state_lock:
                    _state["current"] = done_count
                    _state["results"] = sorted(results, key=lambda r: r["excel_row"] or 0)
                    _state["summary"] = _summarize(results)
        results.sort(key=lambda r: r["excel_row"] or 0)
    except Exception as ex:  # noqa: BLE001
        # De thread is een daemon, dus een exceptie hier verdween spoorloos en de
        # UI zag alleen een run die nul rijen opleverde. Vastleggen in de samenvatting.
        fatal = _err(ex) if isinstance(ex, GoogleAdsException) else f"{type(ex).__name__}: {ex}"
        logger.exception("tag-toppers run afgebroken")
    finally:
        summary = _summarize(results)
        if fatal:
            summary["fatal_error"] = fatal
        with _state_lock:
            _state["running"] = False
            _state["results"] = list(results)
            _state["summary"] = summary
            _state["finished_at"] = datetime.now(timezone.utc).replace(tzinfo=None)
            started = _state["started_at"]
            cancelled = _state["cancel"]
            finished = _state["finished_at"]
            filename = _uploaded.get("filename")

        _save_run(started_at=started, finished_at=finished, dry_run=dry_run,
                  cancelled=cancelled, filename=filename,
                  results=results, summary=summary)


# ---------------------------------------------------------------------------
# Run-historie (overleeft een herstart)
# ---------------------------------------------------------------------------

_RUNS_TABLE_READY = False


def _ensure_runs_table() -> None:
    global _RUNS_TABLE_READY
    if _RUNS_TABLE_READY:
        return
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS gsd_tag_toppers_runs (
                    id                   SERIAL PRIMARY KEY,
                    started_at           TIMESTAMP NOT NULL,
                    finished_at          TIMESTAMP,
                    dry_run              BOOLEAN NOT NULL,
                    cancelled            BOOLEAN NOT NULL DEFAULT FALSE,
                    filename             TEXT,
                    n_rows               INTEGER NOT NULL DEFAULT 0,
                    total_ids            INTEGER NOT NULL DEFAULT 0,
                    campaigns_to_create  INTEGER NOT NULL DEFAULT 0,
                    ids_planned          INTEGER NOT NULL DEFAULT 0,
                    ids_added            INTEGER NOT NULL DEFAULT 0,
                    exclusions_planned   INTEGER NOT NULL DEFAULT 0,
                    exclusions_added     INTEGER NOT NULL DEFAULT 0,
                    negatives_copied     INTEGER NOT NULL DEFAULT 0,
                    rows_with_errors     INTEGER NOT NULL DEFAULT 0,
                    summary              JSONB
                )
            """)
            # De rijen zelf, zodat een oude run geëxporteerd kan worden met
            # dezelfde kolommen als de resultatentabel. Als losse ALTER omdat de
            # tabel al bestond voordat de export er was.
            cur.execute("ALTER TABLE gsd_tag_toppers_runs "
                        "ADD COLUMN IF NOT EXISTS results JSONB")
            cur.execute("""
                CREATE INDEX IF NOT EXISTS gsd_tag_toppers_runs_started_idx
                ON gsd_tag_toppers_runs (started_at DESC)
            """)
        conn.commit()
        _RUNS_TABLE_READY = True
    except Exception:
        conn.rollback()
        raise
    finally:
        return_db_connection(conn)


def _save_run(*, started_at, finished_at, dry_run, cancelled, filename,
              results: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    """Legt één afgeronde run vast. Best-effort: een run mag hier niet op stuklopen."""
    try:
        _ensure_runs_table()
        conn = get_db_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO gsd_tag_toppers_runs
                        (started_at, finished_at, dry_run, cancelled, filename,
                         n_rows, total_ids, campaigns_to_create,
                         ids_planned, ids_added, exclusions_planned, exclusions_added,
                         negatives_copied, rows_with_errors, summary, results)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    started_at, finished_at, dry_run, cancelled, filename,
                    summary.get("rows", 0),
                    sum(r.get("n_ids", 0) for r in results),
                    summary.get("campaigns_to_create", 0),
                    summary.get("ids_to_add", 0),
                    summary.get("ids_added", 0),
                    summary.get("exclusions_to_add", 0),
                    summary.get("exclusions_added", 0),
                    summary.get("negatives_copied", 0),
                    summary.get("errors", 0),
                    json.dumps(summary),
                    json.dumps(_export_rows(results)),
                ))
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            return_db_connection(conn)
    except Exception as ex:
        logger.error("Kon de run niet vastleggen: %s", ex)


# De volledige rijen, inclusief `targets` — die zijn de inhoud van de uitklap, dus
# zonder hen kan een oude run niet teruggezet worden in het resultatenscherm. Kost
# ruwweg 1-2 MB JSONB per run van 620 rijen; JSONB comprimeert dat verder.
def _export_rows(results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return results


def get_run_detail(run_id: int) -> Optional[Dict[str, Any]]:
    """Rijen + samenvatting van één run; None als de run niet bestaat.

    Genoeg om het resultatenscherm precies terug te zetten zoals het na die run was.
    """
    _ensure_runs_table()
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT results, summary, dry_run, started_at
                FROM gsd_tag_toppers_runs WHERE id = %s
            """, (run_id,))
            row = cur.fetchone()
            if row is None:
                return None
            rec = dict(row)
            started = rec.get("started_at")
            return {
                "results": rec.get("results") or [],
                "summary": rec.get("summary") or {},
                "dry_run": bool(rec.get("dry_run")),
                "started_at": started.isoformat() if hasattr(started, "isoformat") else started,
            }
    finally:
        return_db_connection(conn)


def get_runs(limit: int = 100) -> List[Dict[str, Any]]:
    """Recente runs, nieuwste eerst. Tijden zijn UTC (de shared Postgres draait Etc/UTC)."""
    try:
        _ensure_runs_table()
    except Exception as ex:
        logger.error("Kon de run-tabel niet aanmaken: %s", ex)
        return []
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, started_at, finished_at, dry_run, cancelled, filename,
                       n_rows, total_ids, campaigns_to_create,
                       ids_planned, ids_added, exclusions_planned, exclusions_added,
                       negatives_copied, rows_with_errors
                FROM gsd_tag_toppers_runs
                ORDER BY started_at DESC
                LIMIT %s
            """, (limit,))
            # De pool levert een RealDictCursor, dus rijen zijn al dicts.
            out = []
            for row in cur.fetchall():
                rec = dict(row)
                for k in ("started_at", "finished_at"):
                    v = rec.get(k)
                    rec[k] = v.isoformat() if hasattr(v, "isoformat") else v
                out.append(rec)
            return out
    finally:
        return_db_connection(conn)


# ---------------------------------------------------------------------------
# Beheerde ids — de gewenste staat per shop/land, één regel per item id
# ---------------------------------------------------------------------------
#
# Google Ads blijft de WERKELIJKE staat; deze tabel is wat er zou moeten staan.
# Dat onderscheid is de hele reden dat hij bestaat: de audit van 2026-08-07 vond
# 257 campagnes met gaten omdat "wat hoort hier te staan" nergens was vastgelegd
# behalve in Google Ads zelf — precies de plek die je wilt controleren.

_ITEMS_TABLE_READY = False


def _ensure_items_table() -> None:
    global _ITEMS_TABLE_READY
    if _ITEMS_TABLE_READY:
        return
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS gsd_tag_toppers_items (
                    id          SERIAL PRIMARY KEY,
                    country     TEXT NOT NULL,
                    shop_id     TEXT NOT NULL,
                    shop_name   TEXT NOT NULL,
                    shop_key    TEXT NOT NULL,
                    item_id     TEXT NOT NULL,
                    active      BOOLEAN NOT NULL DEFAULT TRUE,
                    added_at    TIMESTAMP NOT NULL DEFAULT now(),
                    removed_at  TIMESTAMP,
                    source      TEXT
                )
            """)
            # De sleutel is shop_id ÉN shop_key samen, precies de identiteit die de
            # campagne-matcher gebruikt. Geen van beide alleen volstaat: shop_id is
            # niet uniek per shop (652237 = Bruna.nl én Hubfootwear.com), en shop_key
            # knipt op de eerste | waardoor Vente-unique.be (358561) en
            # Vente-unique.be|Marketplace (665200) — twee echt verschillende shops —
            # op één hoop zouden belanden.
            cur.execute("""
                CREATE UNIQUE INDEX IF NOT EXISTS gsd_tag_toppers_items_uniq
                ON gsd_tag_toppers_items (country, shop_id, shop_key, item_id)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS gsd_tag_toppers_items_shop_idx
                ON gsd_tag_toppers_items (country, shop_id, shop_key) WHERE active
            """)
        conn.commit()
        _ITEMS_TABLE_READY = True
    except Exception:
        conn.rollback()
        raise
    finally:
        return_db_connection(conn)


def import_items(rows: List[Dict[str, Any]], source: str) -> Dict[str, Any]:
    """Zet rijen (shop_id/shop_name/country/item_ids) in de beheerde staat.

    Idempotent: bestaande actieve ids blijven ongemoeid, eerder verwijderde ids
    worden weer actief. Er wordt hier niets naar Google Ads geschreven.
    """
    from psycopg2.extras import execute_values

    _ensure_items_table()

    # Eerst platslaan naar (country, shop_key) -> tuples, zodat we per shop kunnen
    # opzoeken wat er al staat. Één INSERT per id zou ~96k round trips zijn naar
    # 10.1.32.9; dit gaat in blokken.
    per_shop: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for row in rows:
        country = (row.get("country") or "").upper()
        shop_id = str(row.get("shop_id") or "").strip()
        shop_name = (row.get("shop_name") or "").strip()
        key = _shop_key(shop_name)
        if not country or not shop_id or not key:
            continue
        bucket = per_shop.setdefault((country, shop_id, key), {
            "shop_name": shop_name, "ids": set()})
        bucket["ids"].update(str(i) for i in (row.get("item_ids") or []))

    nieuw = heractiveerd = ongewijzigd = 0
    payload: List[Tuple] = []

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            for (country, shop_id, key), b in per_shop.items():
                cur.execute("""
                    SELECT item_id, active FROM gsd_tag_toppers_items
                    WHERE country = %s AND shop_id = %s AND shop_key = %s
                """, (country, shop_id, key))
                bestaand = {dict(r)["item_id"]: dict(r)["active"] for r in cur.fetchall()}
                for item_id in b["ids"]:
                    was = bestaand.get(item_id)
                    if was is None:
                        nieuw += 1
                    elif was is False:
                        heractiveerd += 1
                    else:
                        ongewijzigd += 1
                    payload.append((country, shop_id, b["shop_name"], key,
                                    item_id, source))

            for i in range(0, len(payload), 5000):
                execute_values(cur, """
                    INSERT INTO gsd_tag_toppers_items
                        (country, shop_id, shop_name, shop_key, item_id, source)
                    VALUES %s
                    ON CONFLICT (country, shop_id, shop_key, item_id) DO UPDATE SET
                        shop_id    = EXCLUDED.shop_id,
                        shop_name  = EXCLUDED.shop_name,
                        active     = TRUE,
                        removed_at = NULL
                """, payload[i:i + 5000])
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        return_db_connection(conn)

    return {"nieuw": nieuw, "heractiveerd": heractiveerd,
            "ongewijzigd": ongewijzigd, "shops": len(per_shop),
            "totaal_verwerkt": len(payload), "bron": source}


# --- vullen vanuit Google Ads ----------------------------------------------
# Zonder dit start de tabel leeg en weerspiegelt hij de werkelijkheid niet. Leest
# per tag_toppers-campagne welke ids POSITIEF getarget zijn; dat is per definitie
# de huidige gewenste staat zoals die live staat.

_seed_state: Dict[str, Any] = {"running": False, "current": 0, "total": 0,
                               "result": None, "error": None}


def get_seed_progress() -> Dict[str, Any]:
    with _state_lock:
        return dict(_seed_state)


def collect_live_targets(progress: bool = True) -> List[Dict[str, Any]]:
    """Alle tag_toppers-campagnes met de item ids die ze positief targeten."""
    from concurrent.futures import ThreadPoolExecutor, as_completed

    client = _get_client()
    out: List[Dict[str, Any]] = []

    for land in ("NL", "BE", "DE"):
        cid = _customer_id(land)
        ga = client.get_service("GoogleAdsService")
        camps = [
            {"id": r.campaign.id, "name": r.campaign.name}
            for r in ga.search(customer_id=cid, query="""
                SELECT campaign.id, campaign.name FROM campaign
                WHERE campaign.status != 'REMOVED'
            """)
            if TAG_TOPPERS_TOKEN in r.campaign.name.lower()
        ]
        if progress:
            with _state_lock:
                _seed_state["total"] += len(camps)

        def work(c):
            shop = SHOP_RE.search(c["name"])
            sid = SHOP_ID_RE.search(c["name"])
            if not shop or not sid:
                return None
            trees = _read_campaign_tree(client, cid, c["id"])
            if not trees:
                return None
            _, nodes = max(trees.items(), key=lambda kv: len(kv[1]))
            ids = sorted({n["item_id"] for n in nodes.values()
                          if n["dim"] == "item_id" and n["item_id"] and not n["negative"]})
            if not ids:
                return None
            return {"shop_id": sid.group(1).strip(), "shop_name": shop.group(1),
                    "country": land, "item_ids": ids}

        with ThreadPoolExecutor(max_workers=RUN_WORKERS) as pool:
            for fut in as_completed([pool.submit(work, c) for c in camps]):
                try:
                    res = fut.result()
                except Exception as ex:
                    logger.warning("seed: campagne overgeslagen: %s", _err(ex))
                    res = None
                if res:
                    out.append(res)
                if progress:
                    with _state_lock:
                        _seed_state["current"] += 1
    return out


def start_seed_from_ads() -> Dict[str, Any]:
    """Vult de beheerde staat met wat er nu live getarget wordt. Draait in de
    achtergrond: 881 campagnes uitlezen kost een paar minuten."""
    with _state_lock:
        if _seed_state["running"]:
            raise RuntimeError("Er loopt al een vulling")
        _seed_state.update({"running": True, "current": 0, "total": 0,
                            "result": None, "error": None})

    def run():
        try:
            rows = collect_live_targets()
            res = import_items(rows, "google-ads")
            with _state_lock:
                _seed_state["result"] = res
        except Exception as ex:
            logger.exception("Vullen vanuit Google Ads mislukt")
            with _state_lock:
                _seed_state["error"] = _err(ex)
        finally:
            with _state_lock:
                _seed_state["running"] = False

    threading.Thread(target=run, daemon=True).start()
    return {"started": True}


def items_summary() -> Dict[str, Any]:
    """Per land/shop hoeveel ids beheerd worden, plus het totaal."""
    try:
        _ensure_items_table()
    except Exception as ex:
        logger.error("items-tabel niet beschikbaar: %s", ex)
        return {"totaal": 0, "shops": 0, "per_land": [], "rijen": []}
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT country, shop_id, MAX(shop_name) AS shop_name, shop_key,
                       COUNT(*) FILTER (WHERE active)     AS actief,
                       COUNT(*) FILTER (WHERE NOT active) AS verwijderd,
                       MAX(added_at) AS laatst_gewijzigd
                FROM gsd_tag_toppers_items
                GROUP BY country, shop_id, shop_key
                ORDER BY country, MAX(shop_name)
            """)
            rijen = []
            for r in cur.fetchall():
                rec = dict(r)
                lg = rec.get("laatst_gewijzigd")
                rec["laatst_gewijzigd"] = lg.isoformat() if hasattr(lg, "isoformat") else lg
                rijen.append(rec)
            cur.execute("""
                SELECT country, COUNT(*) FILTER (WHERE active) AS actief
                FROM gsd_tag_toppers_items GROUP BY country ORDER BY country
            """)
            per_land = [dict(r) for r in cur.fetchall()]
    finally:
        return_db_connection(conn)
    return {
        "totaal": sum(r["actief"] for r in rijen),
        "shops": len(rijen),
        "per_land": per_land,
        "rijen": rijen,
    }


def items_for_run(country: Optional[str] = None) -> List[Dict[str, Any]]:
    """De beheerde staat in het rijformaat dat een run verwacht."""
    _ensure_items_table()
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT country, shop_id, MAX(shop_name) AS shop_name,
                       shop_key, ARRAY_AGG(item_id ORDER BY item_id) AS ids
                FROM gsd_tag_toppers_items
                WHERE active AND (%s IS NULL OR country = %s)
                GROUP BY country, shop_id, shop_key
                ORDER BY country, shop_key, shop_id
            """, (country, country))
            out = []
            for i, r in enumerate(cur.fetchall(), start=2):
                rec = dict(r)
                out.append({
                    "excel_row": i,
                    "shop_id": rec["shop_id"],
                    "shop_name": rec["shop_name"],
                    "country": rec["country"],
                    "item_ids": list(rec["ids"]),
                })
            return out
    finally:
        return_db_connection(conn)


_uploaded: Dict[str, Any] = {"rows": [], "warnings": [], "filename": None}


def set_uploaded(parsed: Dict[str, Any], filename: Optional[str] = None) -> None:
    """Bewaart de geüploade lijst zodat dry-run en echte run dezelfde rijen draaien."""
    with _state_lock:
        _uploaded["rows"] = parsed["rows"]
        _uploaded["warnings"] = parsed["warnings"]
        _uploaded["filename"] = filename


def get_uploaded() -> Dict[str, Any]:
    with _state_lock:
        return {
            "filename": _uploaded["filename"],
            "rows": len(_uploaded["rows"]),
            "warnings": list(_uploaded["warnings"]),
            "total_ids": sum(len(r["item_ids"]) for r in _uploaded["rows"]),
        }


def uploaded_rows() -> List[Dict[str, Any]]:
    with _state_lock:
        return list(_uploaded["rows"])


def start_run(rows: List[Dict[str, Any]], dry_run: bool = True) -> Dict[str, Any]:
    with _state_lock:
        if _state["running"]:
            raise RuntimeError("Er loopt al een run")
        _state.update({
            "running": True,
            "dry_run": dry_run,
            "current": 0,
            "total": len(rows),
            "cancel": False,
            "results": [],
            "summary": {},
            "mutations": 0,
            # Naive UTC: de shared Postgres draait Etc/UTC en het frontend plakt er
            # een "Z" achter voordat het naar Europe/Amsterdam omrekent.
            "started_at": datetime.now(timezone.utc).replace(tzinfo=None),
            "finished_at": None,
        })
    threading.Thread(target=_run, args=(rows, dry_run), daemon=True).start()
    return {"started": True, "total": len(rows), "dry_run": dry_run}
