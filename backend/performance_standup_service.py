"""
Performance Standup data writer.

Pulls SEO + DMA-organic + GSAAS performance for a date range, plus
maincat/bidcat comparison data (latest-in-range vs. that day a week earlier),
and writes everything into the shared SharePoint-synced Excel file
"2025 Dagstats SEO-Overig-GSaaS omzet stand up.xlsx".

Channels:
  - SEO only / DMA organic / GSAAS rows  -> single chan_deriv.marketing_channel
  - Comparison sheets                     -> SEO + Overig Kanaal, is_real_visit=1
"""

import os
import logging
from copy import copy
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from backend.database import get_redshift_connection, return_redshift_connection

logger = logging.getLogger(__name__)

EXCEL_PATH = "/mnt/c/Users/JoepvanSchagen/Beslist.nl BV/SEO - Documenten/2025 Dagstats SEO-Overig-GSaaS omzet stand up.xlsx"

SHEET_SEO_ONLY = "SEO only"
SHEET_GSAAS = "GSaaS"
SHEET_BIDCAT_VISITS = "BIDCAT visit vergelijking"
SHEET_BIDCAT_OMZET = "BIDCAT omzet vergelijking"
SHEET_MAINCAT_VISITS = "MAINCAT visit vergelijking"

# Column letters (1-indexed for openpyxl)
SEO_DATE_COL = 3        # C
SEO_WRITE_COLS = {       # column number -> field in daily_seo row
    4: "visits",         # D
    5: "cpc",            # E
    6: "ww",             # F
    7: "omzet",          # G
    8: "omzet_excl_aff", # H
    9: "opb",            # I
    10: "opb_excl_aff",  # J
    11: "bounce",        # K
    12: "ctr",           # L
}
SEO_DMA_ORG_COL = 13     # M

GSAAS_DATE_COL = 2       # B
GSAAS_VISITS_COL = 3     # C


# ---------------------------------------------------------------------------
# Redshift queries
# ---------------------------------------------------------------------------

def _date_key(d: date) -> int:
    return int(d.strftime("%Y%m%d"))


def _fetch_daily_channel(conn, dates: List[date], channel: str) -> Dict[date, Dict]:
    """Per-date metrics for a single chan_deriv.marketing_channel value.
    Returns dict keyed by date with visits/cpc/ww/omzet/etc.
    """
    if not dates:
        return {}
    keys = [_date_key(d) for d in dates]
    placeholders = ",".join(["%s"] * len(keys))
    sql = f"""
        SELECT fv.dim_date_key AS d,
               COUNT(*) AS visits,
               SUM(fv.cpc_revenue) AS cpc,
               SUM(fv.ww_revenue) AS ww,
               SUM(fv.cpc_revenue + fv.ww_revenue + COALESCE(fv.affiliate_revenue,0)) AS omzet,
               SUM(fv.cpc_revenue + fv.ww_revenue) AS omzet_excl_aff,
               SUM(fv.cpc_revenue + fv.ww_revenue + COALESCE(fv.affiliate_revenue,0))::float
                 / NULLIF(COUNT(*),0) AS opb,
               SUM(fv.cpc_revenue + fv.ww_revenue)::float
                 / NULLIF(COUNT(*),0) AS opb_excl_aff,
               SUM(CASE WHEN COALESCE(fv.number_of_cpc_productclicks,0)=0
                          AND COALESCE(fv.number_of_ww_productclicks,0)=0
                        THEN 1 ELSE 0 END)::float
                 / NULLIF(COUNT(*),0) AS bounce,
               (SUM(COALESCE(fv.number_of_bvb_clicks,0))
                  + SUM(COALESCE(fv.number_of_outclicks,0)))::float
                 / NULLIF(COUNT(*),0) AS ctr
        FROM datamart.fct_visits fv
        JOIN datamart.dim_visit dv  ON fv.dim_visit_key = dv.dim_visit_key
        JOIN chan_deriv.ref_channel_derivation_stats c
          ON dv.aff_id = c.aff_id AND dv.channel_id = c.channel_id
        WHERE fv.dim_date_key IN ({placeholders})
          AND c.marketing_channel = %s
          AND dv.is_real_visit = 1
        GROUP BY fv.dim_date_key
    """
    with conn.cursor() as cur:
        cur.execute(sql, keys + [channel])
        rows = cur.fetchall()
    out: Dict[date, Dict] = {}
    for r in rows:
        d = datetime.strptime(str(r["d"]), "%Y%m%d").date()
        out[d] = {k: (float(v) if v is not None and k not in ("visits",) else v)
                  for k, v in dict(r).items() if k != "d"}
    return out


def _fetch_comparison_visits(conn, p1: date, p2: date) -> Tuple[List[Dict], List[Dict]]:
    """Bidcat + maincat visits comparison for SEO+Overig with is_real_visit=1.
    Returns (bidcat_rows, maincat_rows) — each row dict with maincat/(bidcat)/p1/p2.
    """
    p1k, p2k = _date_key(p1), _date_key(p2)
    sql_bidcat = """
        SELECT COALESCE(cat.main_category_name,'-') AS maincat,
               COALESCE(cat.bid_category_name,'-')  AS bidcat,
               SUM(CASE WHEN fv.dim_date_key = %s THEN 1 ELSE 0 END) AS p1,
               SUM(CASE WHEN fv.dim_date_key = %s THEN 1 ELSE 0 END) AS p2
        FROM datamart.fct_visits fv
        JOIN datamart.dim_visit dv ON fv.dim_visit_key = dv.dim_visit_key
        JOIN chan_deriv.ref_channel_derivation_stats c
          ON dv.aff_id = c.aff_id AND dv.channel_id = c.channel_id
        JOIN datamart.dim_category cat ON fv.dim_category_key = cat.dim_category_key
        WHERE fv.dim_date_key IN (%s, %s)
          AND c.marketing_channel IN ('SEO','Overig Kanaal')
          AND dv.is_real_visit = 1
        GROUP BY 1, 2
        HAVING SUM(CASE WHEN fv.dim_date_key=%s THEN 1 ELSE 0 END) > 0
            OR SUM(CASE WHEN fv.dim_date_key=%s THEN 1 ELSE 0 END) > 0
        ORDER BY p1 DESC NULLS LAST, p2 DESC NULLS LAST
    """
    sql_maincat = """
        SELECT cat.main_category_name AS maincat,
               SUM(CASE WHEN fv.dim_date_key = %s THEN 1 ELSE 0 END) AS p1,
               SUM(CASE WHEN fv.dim_date_key = %s THEN 1 ELSE 0 END) AS p2
        FROM datamart.fct_visits fv
        JOIN datamart.dim_visit dv ON fv.dim_visit_key = dv.dim_visit_key
        JOIN chan_deriv.ref_channel_derivation_stats c
          ON dv.aff_id = c.aff_id AND dv.channel_id = c.channel_id
        JOIN datamart.dim_category cat ON fv.dim_category_key = cat.dim_category_key
        WHERE fv.dim_date_key IN (%s, %s)
          AND c.marketing_channel IN ('SEO','Overig Kanaal')
          AND dv.is_real_visit = 1
        GROUP BY cat.main_category_name
        HAVING SUM(CASE WHEN fv.dim_date_key=%s THEN 1 ELSE 0 END) > 0
            OR SUM(CASE WHEN fv.dim_date_key=%s THEN 1 ELSE 0 END) > 0
        ORDER BY p1 DESC NULLS LAST, p2 DESC NULLS LAST
    """
    with conn.cursor() as cur:
        cur.execute(sql_bidcat, (p1k, p2k, p1k, p2k, p1k, p2k))
        bidcat_rows = [dict(r) for r in cur.fetchall()]
        cur.execute(sql_maincat, (p1k, p2k, p1k, p2k, p1k, p2k))
        maincat_rows = [dict(r) for r in cur.fetchall()]
    return bidcat_rows, maincat_rows


def _fetch_comparison_omzet(conn, p1: date, p2: date) -> List[Dict]:
    """Bidcat CPC+WW omzet comparison for SEO+Overig with is_real_visit=1."""
    p1k, p2k = _date_key(p1), _date_key(p2)
    sql = """
        SELECT COALESCE(cat.main_category_name,'-') AS maincat,
               COALESCE(cat.bid_category_name,'-')  AS bidcat,
               SUM(CASE WHEN fv.dim_date_key = %s THEN fv.cpc_revenue + fv.ww_revenue ELSE 0 END) AS p1,
               SUM(CASE WHEN fv.dim_date_key = %s THEN fv.cpc_revenue + fv.ww_revenue ELSE 0 END) AS p2
        FROM datamart.fct_visits fv
        JOIN datamart.dim_visit dv ON fv.dim_visit_key = dv.dim_visit_key
        JOIN chan_deriv.ref_channel_derivation_stats c
          ON dv.aff_id = c.aff_id AND dv.channel_id = c.channel_id
        JOIN datamart.dim_category cat ON fv.dim_category_key = cat.dim_category_key
        WHERE fv.dim_date_key IN (%s, %s)
          AND c.marketing_channel IN ('SEO','Overig Kanaal')
          AND dv.is_real_visit = 1
        GROUP BY 1, 2
        HAVING SUM(CASE WHEN fv.dim_date_key=%s THEN fv.cpc_revenue + fv.ww_revenue ELSE 0 END) > 0
            OR SUM(CASE WHEN fv.dim_date_key=%s THEN fv.cpc_revenue + fv.ww_revenue ELSE 0 END) > 0
        ORDER BY p1 DESC, p2 DESC
    """
    with conn.cursor() as cur:
        cur.execute(sql, (p1k, p2k, p1k, p2k, p1k, p2k))
        return [dict(r) for r in cur.fetchall()]


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _check_file_lock() -> Optional[str]:
    """Excel creates an ~$file.xlsx lock when the file is open. Returns error string or None."""
    d, base = os.path.split(EXCEL_PATH)
    lock = os.path.join(d, "~$" + base)
    if os.path.exists(lock):
        return ("The Excel file is currently open. Please close "
                "'2025 Dagstats SEO-Overig-GSaaS omzet stand up.xlsx' in Excel and try again.")
    if not os.path.exists(EXCEL_PATH):
        return f"Excel file not found at {EXCEL_PATH}"
    return None


def _write_cell(sheet, row: int, col: int, value):
    """Write a value and copy formatting from the cell directly above (so new rows
    inherit number format / font / fill / border / alignment from older rows)."""
    cell = sheet.cell(row=row, column=col, value=value)
    if row > 1:
        src = sheet.cell(row=row - 1, column=col)
        if src.has_style:
            cell.font = copy(src.font)
            cell.fill = copy(src.fill)
            cell.border = copy(src.border)
            cell.alignment = copy(src.alignment)
            cell.protection = copy(src.protection)
            cell.number_format = src.number_format


def _build_date_row_map(sheet, date_col: int) -> Dict[date, int]:
    """Scan a column for datetime values and return date->row map."""
    m: Dict[date, int] = {}
    for r in range(2, sheet.max_row + 1):
        v = sheet.cell(row=r, column=date_col).value
        if isinstance(v, datetime):
            m[v.date()] = r
        elif isinstance(v, date):
            m[v] = r
    return m


def _verify_and_write_seo_only(sheet, daily_seo: Dict[date, Dict], daily_dma_org: Dict[date, Dict],
                               dates: List[date]) -> List[str]:
    """Write D..L (SEO metrics) and M (DMA organic visits) per date. Verify date in col C matches."""
    date_map = _build_date_row_map(sheet, SEO_DATE_COL)
    errors: List[str] = []
    for d in dates:
        if d not in date_map:
            errors.append(f"SEO only: no row found for {d.isoformat()}")
            continue
        row = date_map[d]
        cell_date = sheet.cell(row=row, column=SEO_DATE_COL).value
        cell_date = cell_date.date() if isinstance(cell_date, datetime) else cell_date
        if cell_date != d:
            errors.append(f"SEO only row {row}: date mismatch ({cell_date} vs {d})")
            continue
        seo_data = daily_seo.get(d)
        if seo_data:
            for col, field in SEO_WRITE_COLS.items():
                _write_cell(sheet, row, col, seo_data.get(field))
        dma_data = daily_dma_org.get(d)
        if dma_data:
            _write_cell(sheet, row, SEO_DMA_ORG_COL, dma_data.get("visits"))
    return errors


def _verify_and_write_gsaas(sheet, daily_gsaas: Dict[date, Dict], dates: List[date]) -> List[str]:
    """Write GSAAS visits (col C) per date. Verify date in col B matches."""
    date_map = _build_date_row_map(sheet, GSAAS_DATE_COL)
    errors: List[str] = []
    for d in dates:
        if d not in date_map:
            errors.append(f"GSaaS: no row found for {d.isoformat()}")
            continue
        row = date_map[d]
        cell_date = sheet.cell(row=row, column=GSAAS_DATE_COL).value
        cell_date = cell_date.date() if isinstance(cell_date, datetime) else cell_date
        if cell_date != d:
            errors.append(f"GSaaS row {row}: date mismatch ({cell_date} vs {d})")
            continue
        gsaas_data = daily_gsaas.get(d)
        if gsaas_data:
            _write_cell(sheet, row, GSAAS_VISITS_COL, gsaas_data.get("visits"))
    return errors


def _clear_data_rows(sheet, first_data_row: int, ncols: int):
    """Clear cells (preserves header). Doesn't delete rows — just sets values to None."""
    for r in range(first_data_row, sheet.max_row + 1):
        for c in range(1, ncols + 1):
            sheet.cell(row=r, column=c, value=None)


def _write_bidcat_visits(sheet, rows: List[Dict]):
    """Overwrite data rows (start row 2). Col A = `=C{row}`, B/C maincat/bidcat, D/E p1/p2, F/G delta abs/rel."""
    first = 2
    _clear_data_rows(sheet, first, 7)
    for i, r in enumerate(rows):
        rn = first + i
        p1, p2 = int(r["p1"] or 0), int(r["p2"] or 0)
        delta_abs = p2 - p1
        delta_rel = (delta_abs / p1) if p1 else None
        _write_cell(sheet, rn, 1, f"=C{rn}")
        _write_cell(sheet, rn, 2, r["maincat"])
        _write_cell(sheet, rn, 3, r["bidcat"])
        _write_cell(sheet, rn, 4, p1)
        _write_cell(sheet, rn, 5, p2)
        _write_cell(sheet, rn, 6, delta_abs)
        _write_cell(sheet, rn, 7, delta_rel)


def _write_bidcat_omzet(sheet, rows: List[Dict]):
    """BIDCAT omzet vergelijking has row 1 = 'Bidcat filter' and row 2 = header; data starts row 3."""
    first = 3
    _clear_data_rows(sheet, first, 7)
    for i, r in enumerate(rows):
        rn = first + i
        p1, p2 = float(r["p1"] or 0), float(r["p2"] or 0)
        delta_abs = p2 - p1
        delta_rel = (delta_abs / p1) if p1 else None
        _write_cell(sheet, rn, 1, f"=C{rn}")
        _write_cell(sheet, rn, 2, r["maincat"])
        _write_cell(sheet, rn, 3, r["bidcat"])
        _write_cell(sheet, rn, 4, p1)
        _write_cell(sheet, rn, 5, p2)
        _write_cell(sheet, rn, 6, delta_abs)
        _write_cell(sheet, rn, 7, delta_rel)


def _write_maincat_visits(sheet, rows: List[Dict]):
    """MAINCAT visit vergelijking has header row 1; data starts row 2 (no col A formula)."""
    first = 2
    _clear_data_rows(sheet, first, 5)
    for i, r in enumerate(rows):
        rn = first + i
        p1, p2 = int(r["p1"] or 0), int(r["p2"] or 0)
        delta_abs = p2 - p1
        delta_rel = (delta_abs / p1) if p1 else None
        _write_cell(sheet, rn, 1, r["maincat"])
        _write_cell(sheet, rn, 2, p1)
        _write_cell(sheet, rn, 3, p2)
        _write_cell(sheet, rn, 4, delta_abs)
        _write_cell(sheet, rn, 5, delta_rel)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def run_standup(start_date: date, end_date: date) -> Dict:
    """Run the full standup: fetch from Redshift and write to the Excel file.

    Comparison sheets always compare end_date vs end_date-7d (visits)
    and end_date-1 vs end_date-8 (omzet).
    """
    if start_date > end_date:
        return {"status": "error", "error": "start_date must be <= end_date"}

    lock_err = _check_file_lock()
    if lock_err:
        return {"status": "error", "error": lock_err}

    dates = []
    d = start_date
    while d <= end_date:
        dates.append(d)
        d += timedelta(days=1)

    # Comparison dates
    cmp_visits_p2 = end_date
    cmp_visits_p1 = end_date - timedelta(days=7)
    cmp_omzet_p2 = end_date - timedelta(days=1)
    cmp_omzet_p1 = end_date - timedelta(days=8)

    conn = get_redshift_connection()
    try:
        logger.info("Fetching SEO daily for %s..%s", start_date, end_date)
        seo_daily = _fetch_daily_channel(conn, dates, "SEO")
        logger.info("Fetching DMA organic daily")
        dma_daily = _fetch_daily_channel(conn, dates, "DMA organic")
        logger.info("Fetching GSAAS daily")
        gsaas_daily = _fetch_daily_channel(conn, dates, "GSAAS")

        logger.info("Fetching visits comparison %s vs %s", cmp_visits_p1, cmp_visits_p2)
        bidcat_visits, maincat_visits = _fetch_comparison_visits(conn, cmp_visits_p1, cmp_visits_p2)
        logger.info("Fetching omzet comparison %s vs %s", cmp_omzet_p1, cmp_omzet_p2)
        bidcat_omzet = _fetch_comparison_omzet(conn, cmp_omzet_p1, cmp_omzet_p2)
    finally:
        return_redshift_connection(conn)

    # Re-check lock right before opening (someone might've opened Excel in the meantime).
    lock_err = _check_file_lock()
    if lock_err:
        return {"status": "error", "error": lock_err}

    logger.info("Opening workbook")
    wb = load_workbook(EXCEL_PATH, data_only=False, keep_vba=False)

    seo_errors  = _verify_and_write_seo_only(wb[SHEET_SEO_ONLY], seo_daily, dma_daily, dates)
    gsaas_errors = _verify_and_write_gsaas(wb[SHEET_GSAAS], gsaas_daily, dates)

    _write_bidcat_visits(wb[SHEET_BIDCAT_VISITS], bidcat_visits)
    _write_bidcat_omzet(wb[SHEET_BIDCAT_OMZET], bidcat_omzet)
    _write_maincat_visits(wb[SHEET_MAINCAT_VISITS], maincat_visits)

    logger.info("Saving workbook")
    try:
        wb.save(EXCEL_PATH)
    except PermissionError as e:
        return {"status": "error",
                "error": f"Cannot save — file is locked. Close Excel and retry. ({e})"}

    return {
        "status": "ok",
        "date_range": {"start": start_date.isoformat(), "end": end_date.isoformat()},
        "comparison": {
            "visits_p1": cmp_visits_p1.isoformat(),
            "visits_p2": cmp_visits_p2.isoformat(),
            "omzet_p1":  cmp_omzet_p1.isoformat(),
            "omzet_p2":  cmp_omzet_p2.isoformat(),
        },
        "rows_written": {
            "seo_dates_in_range": len(dates),
            "seo_dates_found_in_data": sum(1 for d in dates if d in seo_daily),
            "dma_dates_found_in_data": sum(1 for d in dates if d in dma_daily),
            "gsaas_dates_found_in_data": sum(1 for d in dates if d in gsaas_daily),
            "bidcat_visits": len(bidcat_visits),
            "bidcat_omzet":  len(bidcat_omzet),
            "maincat_visits": len(maincat_visits),
        },
        "warnings": seo_errors + gsaas_errors,
    }
