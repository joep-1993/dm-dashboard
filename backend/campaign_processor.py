"""
DMA Shop Campaigns Processor

This script processes Excel files with shop campaign data and updates Google Ads
listing trees with custom label 3 targeting (shop name).

Usage:
    python campaign_processor.py

Configuration:
    - Excel file path: EXCEL_FILE_PATH constant below
    - Customer ID: CUSTOMER_ID constant below
    - Google Ads credentials: google-ads.yaml in the same directory or set via environment
"""

import sys
import os
import time
import uuid
import platform
import threading
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional, Dict, Any
from google.ads.googleads.client import GoogleAdsClient
from google.ads.googleads.errors import GoogleAdsException
import openpyxl
from openpyxl import load_workbook
from dotenv import load_dotenv
import shutil
from datetime import datetime

# Load environment variables
load_dotenv()

# Add script directory to Python path for imports
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

# Import helper functions (add your existing helper functions to google_ads_helpers.py)
try:
    from google_ads_helpers import (
        safe_remove_entire_listing_tree,
        create_listing_group_subdivision,
        create_listing_group_unit_biddable,
        add_standard_shopping_campaign,
        add_shopping_ad_group,
        add_shopping_product_ad,
        enable_negative_list_for_campaign,
    )
except ImportError as e:
    print(f"⚠️  Warning: Could not import helper functions from google_ads_helpers.py")
    print(f"   Error: {e}")
    print(f"   Script directory: {SCRIPT_DIR}")
    print(f"   Please ensure google_ads_helpers.py is in the same directory as this script")

# ============================================================================
# CONFIGURATION
# ============================================================================

# ---- Switch country here ----
COUNTRY = "NL"  # Set to "NL" or "BE"
# ------------------------------

CUSTOMER_ID_NL = "3800751597"
CUSTOMER_ID_BE = "9920951707"
MERCHANT_CENTER_ID_NL = 140784594
MERCHANT_CENTER_ID_BE = 140784810  # TODO: update if BE uses a different Merchant Center ID

COUNTRY_CONFIG = {
    "NL": {"customer_id": CUSTOMER_ID_NL, "merchant_center_id": MERCHANT_CENTER_ID_NL, "exclude_dataedis": True},
    "BE": {"customer_id": CUSTOMER_ID_BE, "merchant_center_id": MERCHANT_CENTER_ID_BE, "exclude_dataedis": False},
}

CUSTOMER_ID = COUNTRY_CONFIG[COUNTRY]["customer_id"]
MERCHANT_CENTER_ID = COUNTRY_CONFIG[COUNTRY]["merchant_center_id"]
EXCLUDE_DATAEDIS = COUNTRY_CONFIG[COUNTRY]["exclude_dataedis"]

MCC_ACCOUNT_ID = "3011145605"  # MCC account where bid strategies are stored
DEFAULT_BID_MICROS = 200_000  # €0.20

# Negative keyword list to add to all created campaigns
NEGATIVE_LIST_NAME = "DMA negatives"

# Label applied to every campaign + ad group created by the DMA+ dashboard,
# so operators can later see at a glance which entities were generated here.
DM_DASHBOARD_LABEL_NAME = "DM_DASHBOARD"

# Cache the label resource_name per customer_id so we hit LabelService once
# per run rather than once per campaign.
_dm_dashboard_label_cache: Dict[str, str] = {}


def ensure_dm_dashboard_label(client: "GoogleAdsClient", customer_id: str) -> Optional[str]:
    """Look up DM_DASHBOARD; create it if missing. Returns label.resource_name, or None on failure."""
    if customer_id in _dm_dashboard_label_cache:
        return _dm_dashboard_label_cache[customer_id]

    try:
        ga_service = client.get_service("GoogleAdsService")
        query = (
            f"SELECT label.resource_name FROM label "
            f"WHERE label.name = '{DM_DASHBOARD_LABEL_NAME}' LIMIT 1"
        )
        for row in ga_service.search(customer_id=customer_id, query=query):
            rn = row.label.resource_name
            _dm_dashboard_label_cache[customer_id] = rn
            return rn
    except Exception as e:
        print(f"   ⚠️  Failed to look up '{DM_DASHBOARD_LABEL_NAME}' label: {str(e)[:200]}")

    # Not found → create it. The description (and background color) live on
    # the label's text_label sub-message, not directly on Label — setting
    # `op.create.description` raises "Unknown field for Label: description".
    try:
        label_service = client.get_service("LabelService")
        op = client.get_type("LabelOperation")
        op.create.name = DM_DASHBOARD_LABEL_NAME
        op.create.text_label.description = "Created automatically by the DMA+ dashboard (dm-tools)."
        resp = label_service.mutate_labels(customer_id=customer_id, operations=[op])
        rn = resp.results[0].resource_name
        _dm_dashboard_label_cache[customer_id] = rn
        print(f"   🏷️  Created label '{DM_DASHBOARD_LABEL_NAME}': {rn}")
        return rn
    except Exception as e:
        print(f"   ⚠️  Failed to create '{DM_DASHBOARD_LABEL_NAME}' label: {str(e)[:200]}")
        return None


def apply_dm_dashboard_label_to_campaign(
    client: "GoogleAdsClient", customer_id: str,
    campaign_resource_name: str, label_resource_name: str,
) -> bool:
    """Attach the DM_DASHBOARD label to a campaign. Returns True on success or if already attached."""
    try:
        svc = client.get_service("CampaignLabelService")
        op = client.get_type("CampaignLabelOperation")
        op.create.campaign = campaign_resource_name
        op.create.label = label_resource_name
        svc.mutate_campaign_labels(customer_id=customer_id, operations=[op])
        return True
    except Exception as e:
        msg = str(e)
        # Google Ads returns an error if the label is already on the campaign;
        # treat that as success so re-runs don't spam failures.
        if "already" in msg.lower() or "duplicate" in msg.lower():
            return True
        print(f"   ⚠️  Failed to label campaign with DM_DASHBOARD: {msg[:200]}")
        return False


def apply_dm_dashboard_label_to_ad_group(
    client: "GoogleAdsClient", customer_id: str,
    ad_group_resource_name: str, label_resource_name: str,
) -> bool:
    """Attach the DM_DASHBOARD label to an ad group. Returns True on success or if already attached."""
    try:
        svc = client.get_service("AdGroupLabelService")
        op = client.get_type("AdGroupLabelOperation")
        op.create.ad_group = ad_group_resource_name
        op.create.label = label_resource_name
        svc.mutate_ad_group_labels(customer_id=customer_id, operations=[op])
        return True
    except Exception as e:
        msg = str(e)
        if "already" in msg.lower() or "duplicate" in msg.lower():
            return True
        print(f"   ⚠️  Failed to label ad group with DM_DASHBOARD: {msg[:200]}")
        return False

# Bid strategy mapping based on custom label 1
BID_STRATEGY_MAPPING = {
    'a': 'DMA: DMA+ shops A - 0,25',
    'b': 'DMA: DMA+ shops B - 0,21',
    'c': 'DMA: DMA+ shops C - 0,17'
}

# Auto-detect Excel file path based on operating system
def get_excel_path():
    """
    Automatically detect the correct Excel file path based on OS.

    Returns:
        str: Path to Excel file (WSL format for Linux, Windows format for Windows)

    Override with env var DMA_EXCEL_PATH. Dashboard callers always pass
    `excel_path=` explicitly, so these defaults only matter for standalone
    CLI runs from this module's __main__.
    """
    env_override = os.getenv("DMA_EXCEL_PATH")
    if env_override:
        return env_override
    # Legacy defaults (kept for CLI-run backwards compat on the original author's box).
    windows_path = "c:/Users/JoepvanSchagen/Downloads/claude/dma_script_uitbreiding.xlsx"
    wsl_path = "/mnt/c/Users/JoepvanSchagen/Downloads/claude/dma_script_uitbreiding.xlsx"

    system = platform.system().lower()

    if system == "windows":
        # Running on native Windows (PyCharm on Windows)
        return windows_path
    elif system == "linux":
        # Check if running on WSL
        if os.path.exists("/proc/version"):
            with open("/proc/version", "r") as f:
                if "microsoft" in f.read().lower():
                    # Running on WSL
                    return wsl_path
        # Running on native Linux - try WSL path first, fall back to Windows path
        if os.path.exists(wsl_path):
            return wsl_path
        return windows_path
    else:
        # Default to Windows path for other systems (macOS, etc.)
        return windows_path

def get_reverse_exclusion_path():
    """
    Get the path to the reverse exclusion Excel file.

    Override with env var DMA_REVERSE_EXCEL_PATH. Dashboard callers always
    pass the path explicitly; defaults only matter for standalone CLI runs.
    """
    env_override = os.getenv("DMA_REVERSE_EXCEL_PATH")
    if env_override:
        return env_override
    windows_path = "C:/Users/JoepvanSchagen/Downloads/claude/dma_script_uitbreiding_reverse.xlsx"
    wsl_path = "/mnt/c/Users/JoepvanSchagen/Downloads/claude/dma_script_uitbreiding_reverse.xlsx"

    system = platform.system().lower()

    if system == "windows":
        return windows_path
    elif system == "linux":
        if os.path.exists("/proc/version"):
            with open("/proc/version", "r") as f:
                if "microsoft" in f.read().lower():
                    return wsl_path
        if os.path.exists(wsl_path):
            return wsl_path
        return windows_path
    else:
        return windows_path

REVERSE_EXCLUSION_FILE_PATH = get_reverse_exclusion_path()

#EXCEL_FILE_PATH = get_excel_path()
EXCEL_FILE_PATH = get_reverse_exclusion_path()

# Sheet names
SHEET_INCLUSION = "toevoegen"  # Inclusion sheet
SHEET_EXCLUSION = "uitsluiten"  # Exclusion sheet
SHEET_REVERSE_INCLUSION = "toevoegen"  # Reverse inclusion sheet (remove ad groups)
SHEET_ENABLE_INCLUSION = "adgroups_heractiveren"  # Enable inclusion sheet (enable ad groups)

# Column indices (0-based) - INCLUSION SHEET (toevoegen) - NEW STRUCTURE (v2)
COL_CAMPAIGN_NAME = 0  # Column A: campaign_name
COL_AD_GROUP_NAME = 1  # Column B: ad group_name (also used as shop_name for CL3)
COL_SHOP_ID = 2        # Column C: Shop ID
COL_MAINCAT = 3        # Column D: maincat
COL_MAINCAT_ID = 4     # Column E: maincat_id (used as CL4)
COL_CUSTOM_LABEL_1 = 5 # Column F: custom label 1
COL_BUDGET = 6         # Column G: budget
COL_STATUS = 7         # Column H: result (TRUE/FALSE)
COL_ERROR = 8          # Column I: Error message (when status is FALSE)

# Column indices (0-based) - INCLUSION SHEET (toevoegen) - LEGACY STRUCTURE
COL_LEGACY_SHOP_NAME = 0      # Column A: Shop name
COL_LEGACY_SHOP_ID = 1        # Column B: Shop ID
COL_LEGACY_MAINCAT = 2        # Column C: maincat
COL_LEGACY_MAINCAT_ID = 3     # Column D: maincat_id
COL_LEGACY_CUSTOM_LABEL_1 = 4 # Column E: custom label 1
COL_LEGACY_BUDGET = 5         # Column F: budget
COL_LEGACY_STATUS = 6         # Column G: Status (TRUE/FALSE)
COL_LEGACY_ERROR = 7          # Column H: Error message (when status is FALSE)

# Column indices (0-based) - EXCLUSION SHEET (uitsluiten) - NEW STRUCTURE
COL_EX_SHOP_NAME = 0      # Column A: Shop name
COL_EX_SHOP_ID = 1        # Column B: Shop ID
COL_EX_MAINCAT = 2        # Column C: maincat
COL_EX_MAINCAT_ID = 3     # Column D: maincat_id
COL_EX_CUSTOM_LABEL_1 = 4 # Column E: custom label 1
COL_EX_STATUS = 5         # Column F: result (TRUE/FALSE)
COL_EX_ERROR = 6          # Column G: Error message (when status is FALSE)

# Column indices (0-based) - CAT_IDS SHEET (category mappings)
COL_CAT_MAINCAT = 0       # Column A: maincat
COL_CAT_MAINCAT_ID = 1    # Column B: maincat_id
COL_CAT_DEEPEST_CAT = 2   # Column C: deepest_cat
COL_CAT_CAT_ID = 3        # Column D: cat_id

# Sheet names
SHEET_CAT_IDS = "cat_ids"

# Column indices (0-based) - UITBREIDING SHEET (extension/expansion)
COL_UIT_SHOP_NAME = 0      # Column A: Shop name
COL_UIT_SHOP_ID = 1        # Column B: Shop ID (not used)
COL_UIT_MAINCAT = 2        # Column C: maincat (category name)
COL_UIT_MAINCAT_ID = 3     # Column D: maincat_id (used as CL4)
COL_UIT_CUSTOM_LABEL_1 = 4 # Column E: custom label 1 (a/b/c)
COL_UIT_BUDGET = 5         # Column F: budget
COL_UIT_STATUS = 6         # Column G: result (TRUE/FALSE)
COL_UIT_ERROR = 7          # Column H: Error message (when status is FALSE)

# Sheet name for uitbreiding
SHEET_UITBREIDING = "toevoegen"  # Using same sheet name as inclusion

# Sheet name for check sheet (replace pipe-version exclusions)
SHEET_CHECK = "check"

# Column indices (0-based) - CHECK SHEET - same structure as exclusion sheet
COL_CHK_SHOP_NAME = 0      # Column A: Shop name (with |)
COL_CHK_SHOP_ID = 1        # Column B: Shop ID
COL_CHK_MAINCAT = 2        # Column C: maincat
COL_CHK_MAINCAT_ID = 3     # Column D: maincat_id
COL_CHK_CUSTOM_LABEL_1 = 4 # Column E: custom label 1
COL_CHK_STATUS = 5         # Column F: result (TRUE/FALSE)
COL_CHK_ERROR = 6          # Column G: Error message

# Sheet name for check_new sheet (replace CL3 pipe-version targeting via direct ad group reference)
SHEET_CHECK_NEW = "check_new"

# Column indices (0-based) - CHECK_NEW SHEET
COL_CHNEW_SHOP_NAME = 0       # Column A: shop_name (with |)
COL_CHNEW_AD_GROUP_NAME = 1   # Column B: ad_group_name
COL_CHNEW_CAMPAIGN_NAME = 2   # Column C: campaign_name
COL_CHNEW_STATUS = 3          # Column D: result (TRUE/FALSE)
COL_CHNEW_ERROR = 4           # Column E: Error message


# ============================================================================
# GOOGLE ADS CLIENT INITIALIZATION
# ============================================================================

def load_google_oauth_from_env():
    """
    Load Google OAuth credentials from environment variables.
    Uses the same env vars as the working GSD-campaigns script.
    """
    client_id = os.getenv("GOOGLE_CLIENT_ID")
    client_secret = os.getenv("GOOGLE_CLIENT_SECRET")
    missing = []
    if not client_id:
        missing.append("GOOGLE_CLIENT_ID")
    if not client_secret:
        missing.append("GOOGLE_CLIENT_SECRET")
    if missing:
        raise RuntimeError(
            f"Environment variables missing: {', '.join(missing)}.\n"
            "Set them in Windows with: setx GOOGLE_CLIENT_ID \"...\" and setx GOOGLE_CLIENT_SECRET \"...\"\n"
            "Or temporarily with: set GOOGLE_CLIENT_ID=... & set GOOGLE_CLIENT_SECRET=..."
        )
    return client_id, client_secret


def initialize_google_ads_client():
    """
    Initialize Google Ads API client.

    Uses the same authentication approach as 'create GSD-campaigns WB.py':
    - Loads all credentials from environment variables

    Returns:
        GoogleAdsClient: Initialized client

    Required environment variables:
        GOOGLE_ADS_REFRESH_TOKEN: OAuth refresh token
        GOOGLE_ADS_DEVELOPER_TOKEN: Developer token
        GOOGLE_ADS_LOGIN_CUSTOMER_ID: MCC account ID (default: 3011145605)
        GOOGLE_CLIENT_ID: OAuth client ID
        GOOGLE_CLIENT_SECRET: OAuth client secret
    """
    try:
        # Load all credentials from environment variables
        print("Loading Google Ads credentials from environment variables...")
        client_id, client_secret = load_google_oauth_from_env()

        refresh_token = os.getenv("GOOGLE_ADS_REFRESH_TOKEN")
        developer_token = os.getenv("GOOGLE_ADS_DEVELOPER_TOKEN")
        login_customer_id = os.getenv("GOOGLE_ADS_LOGIN_CUSTOMER_ID", "3011145605")

        if not refresh_token:
            raise RuntimeError("GOOGLE_ADS_REFRESH_TOKEN environment variable not set")
        if not developer_token:
            raise RuntimeError("GOOGLE_ADS_DEVELOPER_TOKEN environment variable not set")

        credentials = {
            "developer_token": developer_token,
            "refresh_token": refresh_token,
            "client_id": client_id,
            "client_secret": client_secret,
            "login_customer_id": login_customer_id,
            "use_proto_plus": True
        }

        client = GoogleAdsClient.load_from_dict(credentials)
        print("✅ Google Ads client initialized successfully")

        return client
    except RuntimeError as e:
        print(f"❌ Error: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error initializing Google Ads client: {e}")
        print("   Make sure your environment variables are set:")
        print("   - GOOGLE_CLIENT_ID")
        print("   - GOOGLE_CLIENT_SECRET")
        sys.exit(1)


# ============================================================================
# BID STRATEGY RETRIEVAL
# ============================================================================

def get_bid_strategy_by_name(
    client: GoogleAdsClient,
    customer_id: str,
    strategy_name: str
) -> Optional[str]:
    """
    Retrieve portfolio bid strategy by name.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        strategy_name: Bid strategy name to search for

    Returns:
        Bid strategy resource name or None if not found
    """
    ga_service = client.get_service("GoogleAdsService")

    # Escape single quotes in strategy name for GAQL (replace ' with \')
    escaped_strategy_name = strategy_name.replace("'", "\\'")

    query = f"""
        SELECT
            bidding_strategy.id,
            bidding_strategy.name,
            bidding_strategy.resource_name
        FROM bidding_strategy
        WHERE bidding_strategy.name = '{escaped_strategy_name}'
        LIMIT 1
    """

    try:
        response = ga_service.search(customer_id=customer_id, query=query)

        for row in response:
            print(f"   📊 Found bid strategy: {row.bidding_strategy.name} (ID: {row.bidding_strategy.id})")
            return row.bidding_strategy.resource_name

        print(f"   ⚠️  Bid strategy '{strategy_name}' not found")
        return None

    except Exception as e:
        print(f"   ❌ Error searching for bid strategy '{strategy_name}': {e}")
        return None


# ============================================================================
# CAMPAIGN AND AD GROUP RETRIEVAL
# ============================================================================

def get_campaign_by_name_pattern(
    client: GoogleAdsClient,
    customer_id: str,
    name_pattern: str
) -> Optional[Dict[str, Any]]:
    """
    Retrieve campaign by name pattern.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        name_pattern: Campaign name pattern (e.g., "PLA/Electronics_A")

    Returns:
        Dict with campaign info (id, name, resource_name) or None if not found
    """
    ga_service = client.get_service("GoogleAdsService")

    # Escape single quotes in name pattern for GAQL (replace ' with \')
    escaped_name_pattern = name_pattern.replace("'", "\\'")

    query = f"""
        SELECT
            campaign.id,
            campaign.name,
            campaign.resource_name,
            campaign.status
        FROM campaign
        WHERE campaign.name LIKE '%{escaped_name_pattern}%'
            AND campaign.status != 'REMOVED'
        LIMIT 1
    """

    try:
        response = ga_service.search(customer_id=customer_id, query=query)

        for row in response:
            campaign = row.campaign
            return {
                'id': campaign.id,
                'name': campaign.name,
                'resource_name': campaign.resource_name,
                'status': campaign.status.name
            }

        return None

    except GoogleAdsException as e:
        print(f"❌ Error searching for campaign '{name_pattern}': {e}")
        return None


def get_ad_group_from_campaign(
    client: GoogleAdsClient,
    customer_id: str,
    campaign_id: int
) -> Optional[Dict[str, Any]]:
    """
    Retrieve the first active ad group from a campaign.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        campaign_id: Campaign ID

    Returns:
        Dict with ad group info (id, name, resource_name) or None if not found
    """
    ga_service = client.get_service("GoogleAdsService")

    query = f"""
        SELECT
            ad_group.id,
            ad_group.name,
            ad_group.resource_name,
            ad_group.status
        FROM ad_group
        WHERE ad_group.campaign = 'customers/{customer_id}/campaigns/{campaign_id}'
            AND ad_group.status != 'REMOVED'
        LIMIT 1
    """

    try:
        response = ga_service.search(customer_id=customer_id, query=query)

        for row in response:
            ad_group = row.ad_group
            return {
                'id': ad_group.id,
                'name': ad_group.name,
                'resource_name': ad_group.resource_name,
                'status': ad_group.status.name
            }

        return None

    except GoogleAdsException as e:
        print(f"❌ Error retrieving ad group for campaign {campaign_id}: {e}")
        return None


def get_campaign_and_ad_group_by_pattern(
    client: GoogleAdsClient,
    customer_id: str,
    name_pattern: str
) -> Optional[Dict[str, Any]]:
    """
    Retrieve campaign AND ad group by campaign name pattern in a single query.
    This is more efficient than making two separate API calls.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        name_pattern: Campaign name pattern (e.g., "PLA/Electronics_A")

    Returns:
        Dict with campaign and ad_group info:
        {
            'campaign': {'id': ..., 'name': ..., 'resource_name': ..., 'status': ...},
            'ad_group': {'id': ..., 'name': ..., 'resource_name': ..., 'status': ...}
        }
        or None if not found
    """
    ga_service = client.get_service("GoogleAdsService")

    # Escape single quotes in name pattern for GAQL (replace ' with \')
    escaped_name_pattern = name_pattern.replace("'", "\\'")

    query = f"""
        SELECT
            campaign.id,
            campaign.name,
            campaign.resource_name,
            campaign.status,
            ad_group.id,
            ad_group.name,
            ad_group.resource_name,
            ad_group.status
        FROM ad_group
        WHERE campaign.name LIKE '%{escaped_name_pattern}%'
            AND campaign.status != 'REMOVED'
            AND ad_group.status != 'REMOVED'
        LIMIT 1
    """

    try:
        response = ga_service.search(customer_id=customer_id, query=query)

        for row in response:
            return {
                'campaign': {
                    'id': row.campaign.id,
                    'name': row.campaign.name,
                    'resource_name': row.campaign.resource_name,
                    'status': row.campaign.status.name
                },
                'ad_group': {
                    'id': row.ad_group.id,
                    'name': row.ad_group.name,
                    'resource_name': row.ad_group.resource_name,
                    'status': row.ad_group.status.name
                }
            }

        return None

    except GoogleAdsException as e:
        print(f"❌ Error searching for campaign+ad group '{name_pattern}': {e}")
        return None


# ============================================================================
# LISTING TREE REBUILD FUNCTIONS (Custom Label 3 Targeting)
# ============================================================================

def rebuild_tree_with_custom_label_3_inclusion(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: int,
    shop_name: str,
    default_bid_micros: int = DEFAULT_BID_MICROS
):
    """
    Rebuild listing tree to TARGET (include) a specific shop name via custom label 3.

    Structure:
    Root SUBDIVISION
    ├─ Custom Label 3 = shop_name [POSITIVE, biddable] → Target this shop
    └─ Custom Label 3 OTHERS [NEGATIVE] → Exclude all other shops

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_name: Shop name to target (custom label 3 value)
        default_bid_micros: Bid amount in micros
    """
    print(f"   Rebuilding tree to TARGET shop '{shop_name}' (custom label 3)")

    # Remove existing tree
    safe_remove_entire_listing_tree(client, customer_id, str(ad_group_id))
    time.sleep(0.5)

    agc_service = client.get_service("AdGroupCriterionService")

    # MUTATE 1: Create root SUBDIVISION + Custom Label 3 OTHERS (negative)
    ops1 = []

    # 1. ROOT SUBDIVISION
    root_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=str(ad_group_id),
        parent_ad_group_criterion_resource_name=None,
        listing_dimension_info=None
    )
    root_tmp = root_op.create.resource_name
    ops1.append(root_op)

    # 2. Custom Label 3 OTHERS (negative - blocks all other shops)
    dim_cl3_others = client.get_type("ListingDimensionInfo")
    dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3  # INDEX3 = Custom Label 3
    # Don't set value - OTHERS case

    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=root_tmp,
            listing_dimension_info=dim_cl3_others,
            targeting_negative=True,  # NEGATIVE - blocks everything else
            cpc_bid_micros=None
        )
    )

    # Execute first mutate
    resp1 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops1)
    root_actual = resp1.results[0].resource_name
    time.sleep(0.5)

    # MUTATE 2: Add specific shop name as POSITIVE unit
    ops2 = []

    dim_shop = client.get_type("ListingDimensionInfo")
    dim_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3  # INDEX3 = Custom Label 3
    dim_shop.product_custom_attribute.value = shop_name

    ops2.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=root_actual,
            listing_dimension_info=dim_shop,
            targeting_negative=False,  # POSITIVE targeting
            cpc_bid_micros=default_bid_micros
        )
    )

    agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops2)
    print(f"   ✅ Tree rebuilt: ONLY targeting shop '{shop_name}'")


def rebuild_tree_with_custom_label_3_exclusion(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: int,
    shop_name: str,
    default_bid_micros: int = DEFAULT_BID_MICROS
):
    """
    Rebuild listing tree to EXCLUDE a specific shop name via custom label 3.

    Following the pattern from rebuild_tree_with_label_and_item_ids in example_functions.txt:
    1. Read existing tree structure
    2. Collect ALL custom label structures (CL0, CL1, etc.) EXCEPT CL3
    3. Rebuild tree preserving those structures
    4. Add CL3 exclusion

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_name: Shop name to exclude (custom label 3 value)
        default_bid_micros: Bid amount in micros
    """
    print(f"   Rebuilding tree to EXCLUDE shop '{shop_name}' (custom label 3)")

    # Step 1: Read existing tree structure
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    query = f"""
        SELECT
            ad_group_criterion.resource_name,
            ad_group_criterion.listing_group.type,
            ad_group_criterion.listing_group.parent_ad_group_criterion,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
            ad_group_criterion.negative,
            ad_group_criterion.cpc_bid_micros
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
    """

    try:
        results = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        print(f"   ❌ Error reading existing tree: {e}")
        raise  # Re-raise exception so calling code can handle it properly

    # Step 2: Collect ALL custom label structures to preserve (EXCEPT CL2/INDEX2 and CL3/INDEX3)
    custom_label_structures = []
    custom_label_subdivisions = []

    if results:
        for row in results:
            criterion = row.ad_group_criterion
            lg = criterion.listing_group
            case_val = lg.case_value

            if (case_val and
                case_val._pb.WhichOneof("dimension") == "product_custom_attribute"):
                index_name = case_val.product_custom_attribute.index.name
                value = case_val.product_custom_attribute.value

                # Skip Custom Label 2 (INDEX2) and Custom Label 3 (INDEX3) - we're replacing them
                # INDEX2 is the old (incorrect) shop name targeting, INDEX3 is the new (correct) one
                if index_name == 'INDEX2' or index_name == 'INDEX3':
                    continue

                # Skip OTHERS cases (empty value)
                if not value or value == '':
                    continue

                # Collect SUBDIVISION nodes separately
                if lg.type_.name == 'SUBDIVISION':
                    custom_label_subdivisions.append({
                        'index': index_name,
                        'value': value,
                        'parent': lg.parent_ad_group_criterion if lg.parent_ad_group_criterion else None
                    })

                # Preserve all other custom label UNIT nodes (both negative and positive)
                if lg.type_.name == 'UNIT':
                    custom_label_structures.append({
                        'index': index_name,
                        'value': value,
                        'negative': criterion.negative,
                        'bid_micros': criterion.cpc_bid_micros
                    })

    if custom_label_subdivisions:
        print(f"      ℹ️ Found {len(custom_label_subdivisions)} existing subdivision(s):")
        for struct in custom_label_subdivisions:
            print(f"         - {struct['index']}: '{struct['value']}' (SUBDIVISION)")

    if custom_label_structures:
        print(f"      ℹ️ Preserving {len(custom_label_structures)} existing UNIT structure(s):")
        for struct in custom_label_structures:
            neg_str = "[NEGATIVE]" if struct['negative'] else "[POSITIVE]"
            print(f"         - {struct['index']}: '{struct['value']}' {neg_str}")

    # Step 3: Remove old tree
    safe_remove_entire_listing_tree(client, customer_id, str(ad_group_id))
    # No sleep needed - API operations are synchronous

    agc_service = client.get_service("AdGroupCriterionService")

    # Step 4: Rebuild tree hierarchically with preserved structures + CL3 exclusion
    # Use SUBDIVISIONS to determine hierarchy, not UNIT nodes

    # Group subdivisions by INDEX (dimension type)
    cl0_subdivisions = [s for s in custom_label_subdivisions if s['index'] == 'INDEX0']
    cl1_subdivisions = [s for s in custom_label_subdivisions if s['index'] == 'INDEX1']

    # Group UNIT structures by INDEX
    cl0_units = [s for s in custom_label_structures if s['index'] == 'INDEX0']
    cl1_units = [s for s in custom_label_structures if s['index'] == 'INDEX1']

    ops1 = []

    # 1. ROOT SUBDIVISION
    root_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=str(ad_group_id),
        parent_ad_group_criterion_resource_name=None,
        listing_dimension_info=None
    )
    root_tmp = root_op.create.resource_name
    ops1.append(root_op)

    # Determine hierarchy based on SUBDIVISIONS (not units)
    current_parent_tmp = root_tmp
    deepest_subdivision_tmp = root_tmp
    result_index_map = [0]  # Track which result index is which subdivision

    # If CL0 or CL1 subdivisions exist, rebuild them
    if cl0_subdivisions:
        # Build CL0 level
        cl0_subdiv = cl0_subdivisions[0]

        # Create CL0 subdivision
        dim_cl0 = client.get_type("ListingDimensionInfo")
        dim_cl0.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
        dim_cl0.product_custom_attribute.value = cl0_subdiv['value']

        cl0_subdivision_op = create_listing_group_subdivision(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=current_parent_tmp,
            listing_dimension_info=dim_cl0
        )
        cl0_subdivision_tmp = cl0_subdivision_op.create.resource_name
        ops1.append(cl0_subdivision_op)
        result_index_map.append(len(ops1) - 1)  # Track CL0 subdivision index

        # Add CL0 OTHERS (negative)
        dim_cl0_others = client.get_type("ListingDimensionInfo")
        dim_cl0_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
        ops1.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=current_parent_tmp,
                listing_dimension_info=dim_cl0_others,
                targeting_negative=True,
                cpc_bid_micros=None
            )
        )

        current_parent_tmp = cl0_subdivision_tmp
        deepest_subdivision_tmp = cl0_subdivision_tmp

    if cl1_subdivisions:
        # Build CL1 level under current parent
        cl1_subdiv = cl1_subdivisions[0]

        # Create CL1 subdivision
        dim_cl1 = client.get_type("ListingDimensionInfo")
        dim_cl1.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
        dim_cl1.product_custom_attribute.value = cl1_subdiv['value']

        cl1_subdivision_op = create_listing_group_subdivision(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=current_parent_tmp,
            listing_dimension_info=dim_cl1
        )
        cl1_subdivision_tmp = cl1_subdivision_op.create.resource_name
        ops1.append(cl1_subdivision_op)
        result_index_map.append(len(ops1) - 1)  # Track CL1 subdivision index

        # Add CL1 OTHERS (negative)
        dim_cl1_others = client.get_type("ListingDimensionInfo")
        dim_cl1_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
        ops1.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=current_parent_tmp,
                listing_dimension_info=dim_cl1_others,
                targeting_negative=True,
                cpc_bid_micros=None
            )
        )

        current_parent_tmp = cl1_subdivision_tmp
        deepest_subdivision_tmp = cl1_subdivision_tmp

    # If there are CL0 units under the deepest subdivision, we need to convert them to subdivisions
    # and nest CL3 under them (following pattern from rebuild_tree_with_label_and_item_ids)
    if cl0_units:
        # For each CL0 unit, create as subdivision and add CL3 under it
        for unit in cl0_units:
            # Create CL0 subdivision (instead of unit)
            dim_cl0_subdiv = client.get_type("ListingDimensionInfo")
            dim_cl0_subdiv.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
            dim_cl0_subdiv.product_custom_attribute.value = unit['value']

            cl0_unit_subdivision_op = create_listing_group_subdivision(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=deepest_subdivision_tmp,
                listing_dimension_info=dim_cl0_subdiv
            )
            cl0_unit_subdivision_tmp = cl0_unit_subdivision_op.create.resource_name
            ops1.append(cl0_unit_subdivision_op)

            # Add CL3 OTHERS under this CL0 subdivision
            dim_cl3_others = client.get_type("ListingDimensionInfo")
            dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
            ops1.append(
                create_listing_group_unit_biddable(
                    client=client,
                    customer_id=customer_id,
                    ad_group_id=str(ad_group_id),
                    parent_ad_group_criterion_resource_name=cl0_unit_subdivision_tmp,
                    listing_dimension_info=dim_cl3_others,
                    targeting_negative=False,
                    cpc_bid_micros=unit['bid_micros']  # Use the original bid from CL0 unit
                )
            )

        # Add CL0 OTHERS (negative) under deepest subdivision
        dim_cl0_others = client.get_type("ListingDimensionInfo")
        dim_cl0_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
        ops1.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=deepest_subdivision_tmp,
                listing_dimension_info=dim_cl0_others,
                targeting_negative=True,
                cpc_bid_micros=None
            )
        )
    else:
        # No CL0 units - just add CL3 directly under deepest subdivision
        dim_cl3_others = client.get_type("ListingDimensionInfo")
        dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
        ops1.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=deepest_subdivision_tmp,
                listing_dimension_info=dim_cl3_others,
                targeting_negative=False,
                cpc_bid_micros=default_bid_micros
            )
        )

    # Execute first mutate
    try:
        resp1 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops1)
    except Exception as e:
        print(f"   ❌ Error rebuilding tree: {e}")
        raise  # Re-raise exception so calling code can handle it properly

    # No sleep needed - API operations are synchronous

    # MUTATE 2: Add shop exclusion under each CL0 subdivision (if they exist)
    ops2 = []

    if cl0_units:
        # We created CL0 subdivisions - need to find their actual resource names and add exclusion to each
        # Calculate the index of the first CL0 subdivision in results
        base_index = 1  # Start after ROOT
        if cl0_subdivisions:
            base_index += 2  # CL0 subdivision + CL0 OTHERS
        if cl1_subdivisions:
            base_index += 2  # CL1 subdivision + CL1 OTHERS

        # Each CL0 unit became: CL0 subdivision + CL3 OTHERS
        # So CL0 subdivisions are at: base_index, base_index+2, base_index+4, ...
        for i, unit in enumerate(cl0_units):
            cl0_subdivision_actual = resp1.results[base_index + (i * 2)].resource_name

            dim_shop = client.get_type("ListingDimensionInfo")
            dim_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
            dim_shop.product_custom_attribute.value = shop_name
            ops2.append(
                create_listing_group_unit_biddable(
                    client=client,
                    customer_id=customer_id,
                    ad_group_id=str(ad_group_id),
                    parent_ad_group_criterion_resource_name=cl0_subdivision_actual,
                    listing_dimension_info=dim_shop,
                    targeting_negative=True,
                    cpc_bid_micros=None
                )
            )
    else:
        # No CL0 units - add exclusion under the deepest subdivision (CL1 or ROOT)
        if cl1_subdivisions:
            if cl0_subdivisions:
                deepest_subdivision_actual = resp1.results[3].resource_name  # ROOT, CL0 subdivision, CL0 OTHERS, CL1 subdivision
            else:
                deepest_subdivision_actual = resp1.results[1].resource_name  # ROOT, CL1 subdivision
        elif cl0_subdivisions:
            deepest_subdivision_actual = resp1.results[1].resource_name  # ROOT, CL0 subdivision
        else:
            deepest_subdivision_actual = resp1.results[0].resource_name  # ROOT

        dim_shop = client.get_type("ListingDimensionInfo")
        dim_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
        dim_shop.product_custom_attribute.value = shop_name
        ops2.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=deepest_subdivision_actual,
                listing_dimension_info=dim_shop,
                targeting_negative=True,
                cpc_bid_micros=None
            )
        )

    try:
        agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops2)
    except Exception as e:
        print(f"   ❌ Error adding shop exclusion: {e}")
        raise  # Re-raise exception so calling code can handle it properly

    preserved_count = len(custom_label_structures)
    if preserved_count > 0:
        print(f"   ✅ Tree rebuilt: EXCLUDING shop '{shop_name}', preserved {preserved_count} existing structure(s)")
    else:
        print(f"   ✅ Tree rebuilt: EXCLUDING shop '{shop_name}', showing all others.")


def rebuild_tree_with_shop_exclusions(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: int,
    shop_names: list,
    required_cl0_value: str = None,
    default_bid_micros: int = DEFAULT_BID_MICROS
):
    """
    Rebuild listing tree with CL3 shop exclusions while preserving item ID exclusions.
    Validates and enforces CL0 and CL1 targeting based on Excel data and ad group name.

    Tree structure (with item IDs):
    ROOT (subdivision)
    ├─ CL0 = diepste_cat_id (subdivision) - from Excel column D
    │  ├─ CL1 = custom_label_1 (subdivision) - from ad group name suffix
    │  │  ├─ CL3 = shop1 (unit, negative) - exclude shop 1
    │  │  ├─ CL3 = shop2 (unit, negative) - exclude shop 2
    │  │  └─ CL3 OTHERS (subdivision) - for all other shops:
    │  │     ├─ ITEM_ID = xxx (unit, negative) - preserved exclusions
    │  │     ├─ ITEM_ID = yyy (unit, negative) - preserved exclusions
    │  │     └─ ITEM_ID OTHERS (unit, positive with bid)
    │  └─ CL1 OTHERS (unit, negative)
    └─ CL0 OTHERS (unit, negative)

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_names: List of shop names to exclude (CL3 values)
        required_cl0_value: Required CL0 value from Excel (diepste_cat_id)
        default_bid_micros: Bid amount in micros
    """
    print(f"   Rebuilding tree to EXCLUDE {len(shop_names)} shop(s): {', '.join(shop_names)}")

    # Step 1: Get ad group name to check for CL1 suffix requirement
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    # Query ad group name
    ag_name_query = f"""
        SELECT ad_group.name
        FROM ad_group
        WHERE ad_group.id = {ad_group_id}
    """

    try:
        ag_results = list(ga_service.search(customer_id=customer_id, query=ag_name_query))
        ad_group_name = ag_results[0].ad_group.name if ag_results else None
    except Exception as e:
        print(f"   ⚠️  Warning: Could not read ad group name: {e}")
        ad_group_name = None

    # Check if ad group name ends with _a, _b, or _c
    required_cl1 = None
    if ad_group_name:
        for suffix in ['_a', '_b', '_c']:
            if ad_group_name.endswith(suffix):
                required_cl1 = suffix[1:]  # Remove underscore: "_a" → "a"
                print(f"   📌 Ad group name ends with '{suffix}' → CL1 must be '{required_cl1}'")
                break

    # Step 2: Read existing tree to find CL0, CL1, and item ID exclusions
    query = f"""
        SELECT
            ad_group_criterion.listing_group.type,
            ad_group_criterion.listing_group.case_value.product_item_id.value,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
            ad_group_criterion.cpc_bid_micros,
            ad_group_criterion.negative
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
    """

    try:
        results = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        print(f"   ❌ Error reading existing tree: {e}")
        raise

    # Extract CL0, CL1, item IDs, existing shop exclusions, and bid from existing tree
    cl0_value = None
    cl1_value = None
    existing_bid = default_bid_micros
    item_id_exclusions = []  # List of item IDs to preserve
    existing_shop_exclusions = []  # List of existing CL3 shop exclusions to preserve

    for row in results:
        case_value = row.ad_group_criterion.listing_group.case_value

        # Check for item ID
        if case_value.product_item_id.value:
            # Only preserve NEGATIVE item IDs (exclusions)
            if row.ad_group_criterion.negative:
                item_id_exclusions.append(case_value.product_item_id.value)

        # Check for custom attributes (CL0-CL4)
        if case_value.product_custom_attribute:
            index = case_value.product_custom_attribute.index.name
            value = case_value.product_custom_attribute.value

            # Get CL0 and CL1 from any node (subdivision or unit)
            if index == 'INDEX0' and value:
                cl0_value = value
            elif index == 'INDEX1' and value:
                cl1_value = value
            # Capture existing CL3 shop exclusions (NEGATIVE units with value, not OTHERS)
            elif index == 'INDEX3' and value:
                if (row.ad_group_criterion.listing_group.type.name == 'UNIT' and
                    row.ad_group_criterion.negative):
                    existing_shop_exclusions.append(value)

            # Capture existing bid from positive units only
            if (row.ad_group_criterion.listing_group.type.name == 'UNIT' and
                not row.ad_group_criterion.negative and
                row.ad_group_criterion.cpc_bid_micros):
                existing_bid = row.ad_group_criterion.cpc_bid_micros

    # Override CL0 if required value is specified from Excel
    if required_cl0_value:
        if cl0_value and cl0_value != required_cl0_value:
            print(f"   ⚠️  Overriding existing CL0='{cl0_value}' with required CL0='{required_cl0_value}' (from Excel diepste_cat_id)")
        cl0_value = required_cl0_value

    # Override CL1 if ad group name requires specific value
    if required_cl1:
        if cl1_value and cl1_value != required_cl1:
            print(f"   ⚠️  Overriding existing CL1='{cl1_value}' with required CL1='{required_cl1}' (from ad group name)")
        cl1_value = required_cl1

    # Validate we have required values
    if not cl0_value:
        if required_cl0_value:
            cl0_value = required_cl0_value
        else:
            raise Exception(f"Could not find CL0 value in existing tree and Excel doesn't specify one")
    if not cl1_value:
        raise Exception(f"Could not find CL1 value in existing tree and ad group name doesn't specify one")

    # Log what we found
    print(f"   Found existing structure: CL0={cl0_value}, CL1={cl1_value}, bid={existing_bid/10000:.2f}€")
    if existing_shop_exclusions:
        print(f"   Found {len(existing_shop_exclusions)} existing shop exclusion(s): {', '.join(existing_shop_exclusions)}")
    if item_id_exclusions:
        print(f"   Found {len(item_id_exclusions)} item ID exclusion(s)")

    # Merge new shop exclusions with existing ones (preserve all existing)
    # IMPORTANT: Use lowercase for comparison to avoid duplicates due to case differences
    existing_lower = {shop.lower(): shop for shop in existing_shop_exclusions}  # Map lowercase to original
    all_shop_exclusions = set(existing_shop_exclusions)  # Start with existing (preserve original case)
    new_shops_added = []

    for shop in shop_names:
        shop_lower = shop.lower()
        if shop_lower not in existing_lower:
            all_shop_exclusions.add(shop)
            existing_lower[shop_lower] = shop  # Track this one too
            new_shops_added.append(shop)

    if new_shops_added:
        print(f"   Adding {len(new_shops_added)} new shop exclusion(s): {', '.join(new_shops_added)}")
    else:
        print(f"   No new shop exclusions to add (all {len(shop_names)} already exist)")

    # Convert back to sorted list for consistent ordering (case-insensitive sort)
    shop_names = sorted(all_shop_exclusions, key=str.lower)
    print(f"   Total shop exclusions after merge: {len(shop_names)}")

    # Step 3: Remove entire tree
    safe_remove_entire_listing_tree(client, customer_id, str(ad_group_id))
    print(f"   Removed existing tree")

    # Step 4: Rebuild tree with shop exclusions and preserved item IDs
    has_item_ids = len(item_id_exclusions) > 0

    # Rebuild tree with multiple shop exclusions
    agc_service = client.get_service("AdGroupCriterionService")

    # MUTATE 1: Create ROOT + CL0 subdivision + CL0 OTHERS (satisfies CL0) + ROOT OTHERS
    ops1 = []

    # ROOT subdivision
    root_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=str(ad_group_id),
        parent_ad_group_criterion_resource_name=None,
        listing_dimension_info=None
    )
    root_tmp = root_op.create.resource_name
    ops1.append(root_op)

    # CL0 subdivision (under ROOT)
    dim_cl0 = client.get_type("ListingDimensionInfo")
    dim_cl0.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
    dim_cl0.product_custom_attribute.value = str(cl0_value)

    cl0_subdivision_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=str(ad_group_id),
        parent_ad_group_criterion_resource_name=root_tmp,
        listing_dimension_info=dim_cl0
    )
    cl0_subdivision_tmp = cl0_subdivision_op.create.resource_name
    ops1.append(cl0_subdivision_op)

    # CL1 OTHERS (negative - under CL0) - This satisfies CL0 subdivision requirement
    dim_cl1_others_temp = client.get_type("ListingDimensionInfo")
    dim_cl1_others_temp.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=cl0_subdivision_tmp,  # Under CL0!
            listing_dimension_info=dim_cl1_others_temp,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    # CL0 OTHERS (negative - under ROOT) - This satisfies ROOT subdivision requirement
    dim_cl0_others = client.get_type("ListingDimensionInfo")
    dim_cl0_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=root_tmp,  # Under ROOT
            listing_dimension_info=dim_cl0_others,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    try:
        response1 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops1)
        cl0_actual = response1.results[1].resource_name
    except Exception as e:
        raise Exception(f"Error creating ROOT and CL0: {e}")

    # MUTATE 2: Create CL1 subdivision + CL3 OTHERS (subdivision if item IDs, else unit)
    ops2 = []

    # CL1 subdivision (specific value, e.g., "b")
    dim_cl1 = client.get_type("ListingDimensionInfo")
    dim_cl1.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
    dim_cl1.product_custom_attribute.value = str(cl1_value)

    cl1_subdivision_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=str(ad_group_id),
        parent_ad_group_criterion_resource_name=cl0_actual,
        listing_dimension_info=dim_cl1
    )
    cl1_subdivision_tmp = cl1_subdivision_op.create.resource_name
    ops2.append(cl1_subdivision_op)

    # CL3 OTHERS - subdivision if item IDs exist, else unit
    dim_cl3_others = client.get_type("ListingDimensionInfo")
    dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3

    if has_item_ids:
        # Create as SUBDIVISION to hold item ID exclusions underneath
        cl3_others_op = create_listing_group_subdivision(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=cl1_subdivision_tmp,
            listing_dimension_info=dim_cl3_others
        )
        cl3_others_tmp = cl3_others_op.create.resource_name
        ops2.append(cl3_others_op)

        # Add ITEM_ID OTHERS under CL3 OTHERS to satisfy subdivision requirement
        dim_item_others = client.get_type("ListingDimensionInfo")
        dim_item_others.product_item_id = client.get_type("ProductItemIdInfo")
        # Don't set value - this makes it OTHERS
        ops2.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=cl3_others_tmp,
                listing_dimension_info=dim_item_others,
                targeting_negative=False,  # Positive
                cpc_bid_micros=existing_bid
            )
        )
    else:
        # Create as UNIT with bid (no item IDs to preserve)
        ops2.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=cl1_subdivision_tmp,
                listing_dimension_info=dim_cl3_others,
                targeting_negative=False,  # Positive
                cpc_bid_micros=existing_bid
            )
        )

    try:
        response2 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops2)
        cl1_actual = response2.results[0].resource_name
        if has_item_ids:
            cl3_others_actual = response2.results[1].resource_name  # Get actual CL3 OTHERS resource name
    except Exception as e:
        raise Exception(f"Error creating CL1 and CL3 OTHERS: {e}")

    # MUTATE 3: Add individual shop exclusions (CL3 OTHERS already exists from MUTATE 2)
    ops3 = []

    # Add each shop as a negative CL3 unit
    for shop in shop_names:
        dim_cl3_shop = client.get_type("ListingDimensionInfo")
        dim_cl3_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
        dim_cl3_shop.product_custom_attribute.value = str(shop)

        ops3.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=str(ad_group_id),
                parent_ad_group_criterion_resource_name=cl1_actual,
                listing_dimension_info=dim_cl3_shop,
                targeting_negative=True,  # NEGATIVE = exclude this shop
                cpc_bid_micros=None
            )
        )

    # Execute shop exclusions
    try:
        agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops3)
    except Exception as e:
        raise Exception(f"Error adding shop exclusions: {e}")

    # MUTATE 4: Add item ID exclusions under CL3 OTHERS (if any exist)
    if has_item_ids:
        ops4 = []

        # Add each item ID as a negative unit under CL3 OTHERS
        for item_id in item_id_exclusions:
            dim_item_id = client.get_type("ListingDimensionInfo")
            dim_item_id.product_item_id = client.get_type("ProductItemIdInfo")
            dim_item_id.product_item_id.value = item_id

            ops4.append(
                create_listing_group_unit_biddable(
                    client=client,
                    customer_id=customer_id,
                    ad_group_id=str(ad_group_id),
                    parent_ad_group_criterion_resource_name=cl3_others_actual,
                    listing_dimension_info=dim_item_id,
                    targeting_negative=True,  # NEGATIVE = exclude this item ID
                    cpc_bid_micros=None
                )
            )

        # Execute item ID exclusions
        try:
            agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops4)
            print(f"   ✅ Tree rebuilt with {len(shop_names)} shop exclusion(s) and {len(item_id_exclusions)} item ID exclusion(s) preserved")
        except Exception as e:
            raise Exception(f"Error adding item ID exclusions: {e}")
    else:
        print(f"   ✅ Tree rebuilt with {len(shop_names)} shop exclusion(s)")


def build_listing_tree_for_inclusion(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    custom_label_1: str,
    maincat_id: str,
    shop_name: str,
    default_bid_micros: int = DEFAULT_BID_MICROS
):
    """
    Build listing tree for inclusion logic (NEW STRUCTURE):

    Tree structure:
    ROOT (subdivision)
    ├─ Custom Label 3 = shop_name (subdivision)
    │  ├─ Custom Label 3 OTHERS (unit, negative)
    │  └─ Custom Label 4 = maincat_id (subdivision)
    │     ├─ Custom Label 4 OTHERS (unit, negative)
    │     ├─ Custom Label 1 = custom_label_1 (unit, biddable, positive) ← Added in MUTATE 2
    │     └─ Custom Label 1 OTHERS (unit, negative) ← Created in MUTATE 1 with temp name
    └─ Custom Label 3 OTHERS (unit, negative)

    CRITICAL: Google Ads requires that when you create a SUBDIVISION, you must
    provide its OTHERS case in the SAME mutate operation using temporary resource names.

    MUTATE 1: Create root + CL3 subdivision + CL4 subdivision + all OTHERS cases
    MUTATE 2: Add positive custom_label_1 target under maincat subdivision

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        custom_label_1: Custom label 1 value (a/b/c)
        maincat_id: Main category ID to target (custom label 4)
        shop_name: Shop name to target (custom label 3)
        default_bid_micros: Default bid in micros
    """
    print(f"      Building tree: Shop={shop_name}, Maincat ID={maincat_id}, CL1={custom_label_1}")

    # Check if listing tree already exists - if so, skip to preserve exclusions
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    check_query = f"""
        SELECT ad_group_criterion.resource_name
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
        LIMIT 1
    """

    try:
        existing_tree = list(ga_service.search(customer_id=customer_id, query=check_query))
        if existing_tree:
            print(f"      ℹ️  Listing tree already exists - skipping to preserve exclusions")
            return
    except Exception:
        pass  # No existing tree, proceed to create

    agc_service = client.get_service("AdGroupCriterionService")

    # MUTATE 1: Create root + CL3 subdivision + CL4 subdivision + all OTHERS cases
    # CRITICAL: When creating a subdivision, you MUST provide its OTHERS case in the SAME mutate
    ops1 = []

    # 1. ROOT SUBDIVISION
    root_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=None,
        listing_dimension_info=None
    )
    root_tmp = root_op.create.resource_name
    ops1.append(root_op)

    # 2. Custom Label 3 subdivision (Custom Label 3 = shop_name)
    dim_cl3 = client.get_type("ListingDimensionInfo")
    dim_cl3.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3  # INDEX3 = Custom Label 3
    dim_cl3.product_custom_attribute.value = str(shop_name)

    cl3_subdivision_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=root_tmp,
        listing_dimension_info=dim_cl3
    )
    cl3_subdivision_tmp = cl3_subdivision_op.create.resource_name
    ops1.append(cl3_subdivision_op)

    # 3. Custom Label 3 OTHERS (negative - blocks other shops)
    # This is a child of ROOT and satisfies the OTHERS requirement for root
    dim_cl3_others = client.get_type("ListingDimensionInfo")
    dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
    # Don't set value - OTHERS case

    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=root_tmp,
            listing_dimension_info=dim_cl3_others,
            targeting_negative=True,  # NEGATIVE
            cpc_bid_micros=None
        )
    )

    # 4. Maincat ID subdivision (Custom Label 4 = maincat_id)
    # This is a child of CL3 subdivision (using TEMP name)
    dim_maincat = client.get_type("ListingDimensionInfo")
    dim_maincat.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX4  # INDEX4 = Custom Label 4
    dim_maincat.product_custom_attribute.value = str(maincat_id)

    maincat_subdivision_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=cl3_subdivision_tmp,  # Under CL3, not ROOT!
        listing_dimension_info=dim_maincat
    )
    maincat_subdivision_tmp = maincat_subdivision_op.create.resource_name
    ops1.append(maincat_subdivision_op)

    # 5. Custom Label 4 OTHERS (negative - blocks other categories)
    # This is a child of CL3 subdivision and satisfies the OTHERS requirement for CL3
    dim_cl4_others = client.get_type("ListingDimensionInfo")
    dim_cl4_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX4
    # Don't set value - OTHERS case

    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=cl3_subdivision_tmp,  # Child of CL3
            listing_dimension_info=dim_cl4_others,
            targeting_negative=True,  # NEGATIVE
            cpc_bid_micros=None
        )
    )

    # 6. Custom Label 1 OTHERS (negative - blocks other CL1 values)
    # This is a child of maincat_id subdivision (using TEMP name) and satisfies its OTHERS requirement
    dim_cl1_others = client.get_type("ListingDimensionInfo")
    dim_cl1_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
    # Don't set value - OTHERS case

    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=maincat_subdivision_tmp,  # Using TEMP name!
            listing_dimension_info=dim_cl1_others,
            targeting_negative=True,  # NEGATIVE - block other CL1 values
            cpc_bid_micros=None
        )
    )

    # Execute first mutate
    resp1 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops1)
    maincat_subdivision_actual = resp1.results[3].resource_name  # Fourth result is maincat subdivision (0=root, 1=cl3, 2=cl3_others, 3=cl4)

    # Small delay to prevent concurrent modification errors
    time.sleep(0.5)

    # MUTATE 2: Under maincat_id, add the positive custom_label_1 target
    # Note: CL1 OTHERS was already created in MUTATE 1
    ops2 = []

    # Custom Label 1 (Custom Label 1 = custom_label_1) - POSITIVE target
    dim_cl1 = client.get_type("ListingDimensionInfo")
    dim_cl1.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1  # INDEX1 = Custom Label 1
    dim_cl1.product_custom_attribute.value = str(custom_label_1)

    ops2.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=maincat_subdivision_actual,
            listing_dimension_info=dim_cl1,
            targeting_negative=False,  # POSITIVE - target this CL1 value
            cpc_bid_micros=10_000  # 1 cent = €0.01 = 10,000 micros
        )
    )

    # Execute second mutate
    agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops2)
    print(f"      ✅ Tree created: Shop '{shop_name}' → Maincat '{maincat_id}' → CL1 '{custom_label_1}'")


def build_listing_tree_for_inclusion_v2(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    shop_name: str,
    maincat_ids: list,
    custom_label_1: str,
    default_bid_micros: int = DEFAULT_BID_MICROS
):
    """
    Build listing tree for inclusion logic (V2 - with CL1 targeting + dataedis exclusion):

    Tree structure:
    ROOT (subdivision)
    ├─ CL3 = shop_name (subdivision)
    │  ├─ CL4 = maincat_id_1 (subdivision)
    │  │  ├─ CL1 = custom_label_1 (subdivision)
    │  │  │  ├─ CL0 OTHERS (unit, positive, biddable)
    │  │  │  └─ CL0 = dataedis (unit, negative)
    │  │  └─ CL1 OTHERS (unit, negative)
    │  ├─ CL4 = maincat_id_2 (subdivision)  [if multiple maincat_ids]
    │  │  ├─ CL1 = custom_label_1 (subdivision)
    │  │  │  ├─ CL0 OTHERS (unit, positive, biddable)
    │  │  │  └─ CL0 = dataedis (unit, negative)
    │  │  └─ CL1 OTHERS (unit, negative)
    │  └─ CL4 OTHERS (unit, negative)
    └─ CL3 OTHERS (unit, negative)

    IMPORTANT: This function will NOT rebuild the tree if one already exists,
    to preserve any existing exclusions that may have been added.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_name: Shop name to target (custom label 3) - same as ad_group_name
        maincat_ids: List of maincat IDs to target (custom label 4)
        custom_label_1: CL1 value (a/b/c)
        default_bid_micros: Default bid in micros
    """
    print(f"      Building tree: Shop={shop_name}, Maincat IDs={maincat_ids}, CL1={custom_label_1}")

    # Check if listing tree already exists - if so, skip to preserve exclusions
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    check_query = f"""
        SELECT ad_group_criterion.resource_name
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
        LIMIT 1
    """

    try:
        existing_tree = list(ga_service.search(customer_id=customer_id, query=check_query))
        if existing_tree:
            print(f"      ℹ️  Listing tree already exists - skipping to preserve exclusions")
            return
    except Exception:
        pass  # No existing tree, proceed to create

    agc_service = client.get_service("AdGroupCriterionService")

    # =========================================================================
    # MUTATE 1: Create all subdivisions + their OTHERS cases
    # ROOT + CL3 subdiv + CL3 OTHERS + CL4 OTHERS + [CL4 subdiv + CL1 OTHERS] per maincat
    # =========================================================================
    ops1 = []

    # [0] ROOT SUBDIVISION
    root_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=None,
        listing_dimension_info=None
    )
    root_tmp = root_op.create.resource_name
    ops1.append(root_op)

    # [1] CL3 = shop_name subdivision (under root)
    dim_cl3 = client.get_type("ListingDimensionInfo")
    dim_cl3.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
    dim_cl3.product_custom_attribute.value = str(shop_name)

    cl3_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=root_tmp,
        listing_dimension_info=dim_cl3
    )
    cl3_tmp = cl3_op.create.resource_name
    ops1.append(cl3_op)

    # [2] CL3 OTHERS (unit, negative, under root)
    dim_cl3_others = client.get_type("ListingDimensionInfo")
    dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=root_tmp,
            listing_dimension_info=dim_cl3_others,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    # [3] CL4 OTHERS (unit, negative, under CL3)
    dim_cl4_others = client.get_type("ListingDimensionInfo")
    dim_cl4_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX4
    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=cl3_tmp,
            listing_dimension_info=dim_cl4_others,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    # For each maincat_id: [4+i*2] CL4 subdivision + [5+i*2] CL1 OTHERS
    for maincat_id in maincat_ids:
        # CL4 = maincat_id subdivision (under CL3)
        dim_cl4 = client.get_type("ListingDimensionInfo")
        dim_cl4.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX4
        dim_cl4.product_custom_attribute.value = str(maincat_id)

        cl4_op = create_listing_group_subdivision(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=cl3_tmp,
            listing_dimension_info=dim_cl4
        )
        cl4_tmp = cl4_op.create.resource_name
        ops1.append(cl4_op)

        # CL1 OTHERS (unit, negative, under this CL4 subdivision)
        dim_cl1_others = client.get_type("ListingDimensionInfo")
        dim_cl1_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
        ops1.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=ad_group_id,
                parent_ad_group_criterion_resource_name=cl4_tmp,
                listing_dimension_info=dim_cl1_others,
                targeting_negative=True,
                cpc_bid_micros=None
            )
        )

    # Execute MUTATE 1
    resp1 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops1)

    time.sleep(0.5)

    # =========================================================================
    # MUTATE 2: Add CL1 subdivisions (with CL0 children) under each CL4
    # CL4 subdivision actual names are at response indices: 4 + i*2
    # =========================================================================
    ops2 = []

    # Track CL1 subdivision temp names for MUTATE 3
    cl1_subdiv_temps = []

    for i, maincat_id in enumerate(maincat_ids):
        cl4_actual = resp1.results[4 + i * 2].resource_name

        if EXCLUDE_DATAEDIS:
            # CL1 = custom_label_1 subdivision (under CL4) — subdivides into CL0 for dataedis exclusion
            dim_cl1 = client.get_type("ListingDimensionInfo")
            dim_cl1.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
            dim_cl1.product_custom_attribute.value = str(custom_label_1)

            cl1_op = create_listing_group_subdivision(
                client=client,
                customer_id=customer_id,
                ad_group_id=ad_group_id,
                parent_ad_group_criterion_resource_name=cl4_actual,
                listing_dimension_info=dim_cl1
            )
            cl1_tmp = cl1_op.create.resource_name
            cl1_subdiv_temps.append(cl1_tmp)
            ops2.append(cl1_op)

            # CL0 OTHERS (unit, positive, biddable, under CL1) — catches all non-excluded
            dim_cl0_others = client.get_type("ListingDimensionInfo")
            dim_cl0_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
            ops2.append(
                create_listing_group_unit_biddable(
                    client=client,
                    customer_id=customer_id,
                    ad_group_id=ad_group_id,
                    parent_ad_group_criterion_resource_name=cl1_tmp,
                    listing_dimension_info=dim_cl0_others,
                    targeting_negative=False,
                    cpc_bid_micros=10_000  # 1 cent = 10,000 micros
                )
            )

            # CL0 = dataedis (unit, negative, under CL1)
            dim_cl0_excl = client.get_type("ListingDimensionInfo")
            dim_cl0_excl.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
            dim_cl0_excl.product_custom_attribute.value = "dataedis"
            ops2.append(
                create_listing_group_unit_biddable(
                    client=client,
                    customer_id=customer_id,
                    ad_group_id=ad_group_id,
                    parent_ad_group_criterion_resource_name=cl1_tmp,
                    listing_dimension_info=dim_cl0_excl,
                    targeting_negative=True,
                    cpc_bid_micros=None
                )
            )
        else:
            # CL1 = custom_label_1 unit (under CL4) — no CL0 subdivision needed
            dim_cl1 = client.get_type("ListingDimensionInfo")
            dim_cl1.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
            dim_cl1.product_custom_attribute.value = str(custom_label_1)

            ops2.append(
                create_listing_group_unit_biddable(
                    client=client,
                    customer_id=customer_id,
                    ad_group_id=ad_group_id,
                    parent_ad_group_criterion_resource_name=cl4_actual,
                    listing_dimension_info=dim_cl1,
                    targeting_negative=False,
                    cpc_bid_micros=10_000  # 1 cent = 10,000 micros
                )
            )

    # Execute MUTATE 2
    agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops2)
    dataedis_msg = " → dataedis excluded" if EXCLUDE_DATAEDIS else ""
    print(f"      ✅ Tree created: Shop '{shop_name}' → {len(maincat_ids)} maincat(s) → CL1 '{custom_label_1}'{dataedis_msg}")


def build_listing_tree_with_cl1(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    shop_name: str,
    maincat_ids: list,
    custom_label_1: str,
    default_bid_micros: int = DEFAULT_BID_MICROS
):
    """
    Build listing tree with CL3 (shop), CL4 (maincat) subdivisions, and CL1 targeting.

    Tree structure:
    ROOT (subdivision)
    ├─ CL3 = shop_name (subdivision)
    │  ├─ CL4 = maincat_id_1 (subdivision)
    │  │  ├─ CL1 = custom_label_1 (unit, positive, biddable)
    │  │  └─ CL1 OTHERS (unit, negative)
    │  ├─ CL4 = maincat_id_2 (subdivision)  [if multiple maincat_ids]
    │  │  ├─ CL1 = custom_label_1 (unit, positive, biddable)
    │  │  └─ CL1 OTHERS (unit, negative)
    │  └─ CL4 OTHERS (unit, negative)
    └─ CL3 OTHERS (unit, negative)

    IMPORTANT: This function does NOT check for an existing tree. The caller must
    remove the old tree first (via safe_remove_entire_listing_tree).

    MUTATE 1: root + CL3 subdiv + CL3 OTHERS + CL4 OTHERS + [CL4 subdiv + CL1 OTHERS] per maincat
    MUTATE 2: CL1 positive units under each CL4 subdivision (using actual resource names)

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_name: Shop name for CL3 targeting (already split at |)
        maincat_ids: List of maincat IDs for CL4 targeting
        custom_label_1: CL1 value (a/b/c)
        default_bid_micros: Default bid in micros
    """
    print(f"      Building tree with CL1: Shop={shop_name}, Maincat IDs={maincat_ids}, CL1={custom_label_1}")

    agc_service = client.get_service("AdGroupCriterionService")

    # =========================================================================
    # MUTATE 1: Create all subdivisions + their OTHERS cases
    # =========================================================================
    ops1 = []

    # [0] ROOT subdivision
    root_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=None,
        listing_dimension_info=None
    )
    root_tmp = root_op.create.resource_name
    ops1.append(root_op)

    # [1] CL3 = shop_name subdivision (under root)
    dim_cl3 = client.get_type("ListingDimensionInfo")
    dim_cl3.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
    dim_cl3.product_custom_attribute.value = str(shop_name)

    cl3_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=root_tmp,
        listing_dimension_info=dim_cl3
    )
    cl3_tmp = cl3_op.create.resource_name
    ops1.append(cl3_op)

    # [2] CL3 OTHERS (unit, negative, under root)
    dim_cl3_others = client.get_type("ListingDimensionInfo")
    dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=root_tmp,
            listing_dimension_info=dim_cl3_others,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    # [3] CL4 OTHERS (unit, negative, under CL3)
    dim_cl4_others = client.get_type("ListingDimensionInfo")
    dim_cl4_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX4
    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=cl3_tmp,
            listing_dimension_info=dim_cl4_others,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    # For each maincat_id: [4+i*2] CL4 subdivision + [5+i*2] CL1 OTHERS
    for maincat_id in maincat_ids:
        # CL4 = maincat_id subdivision (under CL3)
        dim_cl4 = client.get_type("ListingDimensionInfo")
        dim_cl4.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX4
        dim_cl4.product_custom_attribute.value = str(maincat_id)

        cl4_op = create_listing_group_subdivision(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=cl3_tmp,
            listing_dimension_info=dim_cl4
        )
        cl4_tmp = cl4_op.create.resource_name
        ops1.append(cl4_op)

        # CL1 OTHERS (unit, negative, under this CL4 subdivision)
        dim_cl1_others = client.get_type("ListingDimensionInfo")
        dim_cl1_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
        ops1.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=ad_group_id,
                parent_ad_group_criterion_resource_name=cl4_tmp,
                listing_dimension_info=dim_cl1_others,
                targeting_negative=True,
                cpc_bid_micros=None
            )
        )

    # Execute MUTATE 1
    resp1 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops1)

    # Small delay to prevent concurrent modification errors
    time.sleep(0.5)

    # =========================================================================
    # MUTATE 2: Add positive CL1 targets under each CL4 subdivision
    # =========================================================================
    # CL4 subdivision actual names are at response indices: 4 + i*2
    ops2 = []

    for i, maincat_id in enumerate(maincat_ids):
        cl4_actual = resp1.results[4 + i * 2].resource_name

        dim_cl1 = client.get_type("ListingDimensionInfo")
        dim_cl1.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
        dim_cl1.product_custom_attribute.value = str(custom_label_1)

        ops2.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=ad_group_id,
                parent_ad_group_criterion_resource_name=cl4_actual,
                listing_dimension_info=dim_cl1,
                targeting_negative=False,
                cpc_bid_micros=10_000  # 1 cent = 10,000 micros
            )
        )

    # Execute MUTATE 2
    agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops2)
    print(f"      ✅ Tree created: Shop '{shop_name}' → {len(maincat_ids)} maincat(s) → CL1 '{custom_label_1}'")


def build_listing_tree_for_uitbreiding(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    shop_name: str,
    maincat_id: str,
    custom_label_1: str,
    default_bid_micros: int = DEFAULT_BID_MICROS
):
    """
    Build listing tree for uitbreiding (extension) logic:

    Tree structure:
    ROOT (subdivision)
    └─ CL1 = custom_label_1 (subdivision)
       ├─ CL3 = shop_name (subdivision)
       │  ├─ CL4 = maincat_id (unit, biddable, positive)
       │  └─ CL4 OTHERS (unit, negative)
       └─ CL3 OTHERS (unit, negative)
    └─ CL1 OTHERS (unit, negative)

    This targets:
    - Custom Label 1 = a/b/c (variant)
    - Custom Label 3 = shop_name
    - Custom Label 4 = maincat_id (category)

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_name: Shop name to target (custom label 3)
        maincat_id: Category ID to target (custom label 4)
        custom_label_1: Label value (a/b/c) for custom label 1
        default_bid_micros: Default bid in micros
    """
    print(f"      Building tree: CL1={custom_label_1}, Shop={shop_name}, Maincat={maincat_id}")

    # Check if listing tree already exists - if so, skip to preserve exclusions
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    check_query = f"""
        SELECT ad_group_criterion.resource_name
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
        LIMIT 1
    """

    try:
        existing_tree = list(ga_service.search(customer_id=customer_id, query=check_query))
        if existing_tree:
            print(f"      ℹ️  Listing tree already exists - skipping to preserve exclusions")
            return
    except Exception:
        pass  # No existing tree, proceed to create

    agc_service = client.get_service("AdGroupCriterionService")

    # MUTATE 1: Create ROOT + CL1 subdivision + CL1 OTHERS
    # Also need to add CL3 OTHERS under CL1 subdivision (required for subdivision)
    ops1 = []

    # 1. ROOT SUBDIVISION
    root_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=None,
        listing_dimension_info=None
    )
    root_tmp = root_op.create.resource_name
    ops1.append(root_op)

    # 2. Custom Label 1 subdivision (CL1 = a/b/c)
    dim_cl1 = client.get_type("ListingDimensionInfo")
    dim_cl1.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
    dim_cl1.product_custom_attribute.value = str(custom_label_1)

    cl1_subdivision_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=root_tmp,
        listing_dimension_info=dim_cl1
    )
    cl1_subdivision_tmp = cl1_subdivision_op.create.resource_name
    ops1.append(cl1_subdivision_op)

    # 3. Custom Label 3 OTHERS under CL1 subdivision (required for CL1 subdivision)
    dim_cl3_others = client.get_type("ListingDimensionInfo")
    dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3

    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=cl1_subdivision_tmp,
            listing_dimension_info=dim_cl3_others,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    # 4. Custom Label 1 OTHERS (negative - blocks other variants)
    dim_cl1_others = client.get_type("ListingDimensionInfo")
    dim_cl1_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1

    ops1.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=root_tmp,
            listing_dimension_info=dim_cl1_others,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    # Execute first mutate
    resp1 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops1)
    cl1_subdivision_actual = resp1.results[1].resource_name  # Second result is CL1 subdivision

    # Wait for API to process before next mutate
    time.sleep(2)

    # MUTATE 2: Create CL3 subdivision under CL1 + CL4 OTHERS under CL3
    ops2 = []

    # 5. Custom Label 3 subdivision (CL3 = shop_name)
    dim_cl3 = client.get_type("ListingDimensionInfo")
    dim_cl3.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
    dim_cl3.product_custom_attribute.value = str(shop_name)

    cl3_subdivision_op = create_listing_group_subdivision(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=cl1_subdivision_actual,
        listing_dimension_info=dim_cl3
    )
    cl3_subdivision_tmp = cl3_subdivision_op.create.resource_name
    ops2.append(cl3_subdivision_op)

    # 6. Custom Label 4 OTHERS (negative - blocks other categories)
    dim_cl4_others = client.get_type("ListingDimensionInfo")
    dim_cl4_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX4

    ops2.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=cl3_subdivision_tmp,
            listing_dimension_info=dim_cl4_others,
            targeting_negative=True,
            cpc_bid_micros=None
        )
    )

    # Execute second mutate
    resp2 = agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops2)
    cl3_subdivision_actual = resp2.results[0].resource_name  # First result is CL3 subdivision

    # Wait for API to process before next mutate
    time.sleep(2)

    # MUTATE 3: Add maincat_id as positive CL4 unit
    ops3 = []

    dim_cl4 = client.get_type("ListingDimensionInfo")
    dim_cl4.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX4
    dim_cl4.product_custom_attribute.value = str(maincat_id)

    ops3.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=cl3_subdivision_actual,
            listing_dimension_info=dim_cl4,
            targeting_negative=False,  # POSITIVE - target this maincat
            cpc_bid_micros=10_000  # 1 cent = €0.01 = 10,000 micros
        )
    )

    # Execute third mutate
    agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops3)
    print(f"      ✅ Tree created: CL1='{custom_label_1}' → CL3='{shop_name}' → CL4='{maincat_id}'")


# ============================================================================
# EXCEL PROCESSING
# ============================================================================

def process_inclusion_sheet_v2(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None,
    dry_run: bool = False,
):
    """
    Process the 'toevoegen' (inclusion) sheet - V2 (NEW STRUCTURE).

    Excel columns:
    A. shop_name - Used to build ad group name: PLA/{shop_name}_{cl1}
    B. Shop ID (not used)
    C. maincat - Used to build campaign name: PLA/{maincat} store_{cl1}
    D. maincat_id - Used as Custom Label 4 (multiple per ad group)
    E. custom label 1 - Used in both campaign and ad group names
    F. budget (daily budget in EUR)
    G. result (TRUE/FALSE) - updated by script
    H. error message

    Naming conventions:
    - Campaign name: PLA/{maincat} store_{cl1}
    - Ad group name: PLA/{shop_name}_{cl1}

    Groups rows by derived campaign name (maincat + cl1), then by shop_name.
    For each campaign:
    1. Create campaign with derived name (status: PAUSED)
       - Uses budget from column F
       - Applies bid strategy from MCC based on custom_label_1
    2. For each unique shop_name within the campaign (status: ENABLED):
       - Create ad group with derived name
       - Collect all maincat_ids for that ad group
       - Build listing tree with shop_name as CL3
       - All maincat_ids as positive CL4 units
    3. Update column G (result) with TRUE/FALSE per row

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving)
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING INCLUSION SHEET (V2): '{SHEET_INCLUSION}'")
    if dry_run:
        print("(DRY RUN: no campaigns, ad groups, listing trees or ads will actually be written to Google Ads)")
    print(f"{'='*70}\n")

    # Resolve (or create) the DM_DASHBOARD label once up-front. Live runs only —
    # in dry_run mode we'd be labeling against fake resource names, so we skip
    # the API call entirely and just note the intent.
    dm_label_resource = None
    if dry_run:
        print(f"   [DRY RUN] Would ensure label '{DM_DASHBOARD_LABEL_NAME}' exists and apply it to each new campaign + ad group")
    else:
        dm_label_resource = ensure_dm_dashboard_label(client, customer_id)
        if dm_label_resource:
            print(f"   🏷️  Using label '{DM_DASHBOARD_LABEL_NAME}': {dm_label_resource}")

    try:
        sheet = workbook[SHEET_INCLUSION]
    except KeyError:
        print(f"❌ Sheet '{SHEET_INCLUSION}' not found in workbook")
        return

    # Load a separate workbook with data_only=True to read calculated values from formulas
    # This is needed because cells may contain VLOOKUP formulas instead of plain values
    data_workbook = None
    data_sheet = None
    if file_path:
        try:
            data_workbook = load_workbook(file_path, data_only=True)
            data_sheet = data_workbook[SHEET_INCLUSION]
            print("   (Using data_only mode to read formula results)")
        except Exception as e:
            print(f"   ⚠️  Could not load data_only workbook: {e}")
            print(f"   (Will read formulas as-is - make sure cells contain values, not formulas)")

    # Column indices for this sheet
    COL_SHOP_NAME = 0      # A: shop_name
    COL_SHOP_ID = 1        # B: Shop ID (not used)
    COL_MAINCAT = 2        # C: maincat
    COL_MAINCAT_ID = 3     # D: maincat_id
    COL_CL1 = 4            # E: custom label 1
    COL_BUDGET = 5         # F: budget
    COL_RESULT = 6         # G: result (TRUE/FALSE)
    COL_ERR = 7            # H: error message

    # Step 1: Read all rows and group by campaign (maincat + cl1), then by shop_name
    campaigns = defaultdict(lambda: {
        'maincat': None,
        'cl1': None,
        'ad_groups': defaultdict(lambda: {'maincat_ids': set(), 'shop_id': None, 'rows': []}),
        'budget': None,
        'rows': []
    })

    print("Step 1: Reading and grouping rows by campaign (maincat + cl1) and ad group (shop_name)...")
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if status column is empty (use original sheet for status check)
        status_value = row[COL_RESULT].value if len(row) > COL_RESULT else None

        # Skip rows that already have a status (TRUE/FALSE)
        if status_value is not None and status_value != '':
            continue

        # Read values from data_only sheet if available (to get formula results)
        # Otherwise fall back to original sheet
        if data_sheet:
            shop_name = data_sheet.cell(row=idx, column=COL_SHOP_NAME + 1).value
            shop_id = data_sheet.cell(row=idx, column=COL_SHOP_ID + 1).value
            maincat = data_sheet.cell(row=idx, column=COL_MAINCAT + 1).value
            maincat_id = data_sheet.cell(row=idx, column=COL_MAINCAT_ID + 1).value
            custom_label_1 = data_sheet.cell(row=idx, column=COL_CL1 + 1).value
            budget = data_sheet.cell(row=idx, column=COL_BUDGET + 1).value
        else:
            shop_name = row[COL_SHOP_NAME].value
            shop_id = row[COL_SHOP_ID].value
            maincat = row[COL_MAINCAT].value
            maincat_id = row[COL_MAINCAT_ID].value
            custom_label_1 = row[COL_CL1].value
            budget = row[COL_BUDGET].value

        # Validate required fields
        if not shop_name or not maincat or not maincat_id or not custom_label_1:
            print(f"   ⚠️  [Row {idx}] Missing required fields (shop_name/maincat/maincat_id/cl1), skipping")
            sheet.cell(row=idx, column=COL_RESULT + 1).value = False
            sheet.cell(row=idx, column=COL_ERR + 1).value = "Missing required fields"
            continue

        # Build campaign name from maincat and cl1
        campaign_name = f"PLA/{maincat} store_{custom_label_1}"

        # Store campaign-level data
        campaigns[campaign_name]['maincat'] = maincat
        campaigns[campaign_name]['cl1'] = custom_label_1
        campaigns[campaign_name]['budget'] = budget
        campaigns[campaign_name]['rows'].append({'idx': idx, 'row': row})

        # Store ad group data - collect all maincat_ids for this shop
        campaigns[campaign_name]['ad_groups'][shop_name]['maincat_ids'].add(maincat_id)
        campaigns[campaign_name]['ad_groups'][shop_name]['shop_id'] = shop_id
        campaigns[campaign_name]['ad_groups'][shop_name]['rows'].append({'idx': idx, 'row': row})

    print(f"   Found {len(campaigns)} campaign(s) to process")
    total_ad_groups = sum(len(c['ad_groups']) for c in campaigns.values())
    print(f"   Total ad groups: {total_ad_groups}\n")

    # Step 2: Process each campaign
    total_campaigns = len(campaigns)
    successful_campaigns = 0

    for campaign_idx, (campaign_name, campaign_data) in enumerate(campaigns.items(), start=1):
        print(f"\n{'─'*70}")
        print(f"CAMPAIGN {campaign_idx}/{total_campaigns}: {campaign_name}")
        print(f"{'─'*70}")

        budget_value = campaign_data['budget']
        custom_label_1 = campaign_data['cl1']
        maincat = campaign_data['maincat']
        ad_groups = campaign_data['ad_groups']

        print(f"   Maincat: {maincat}")
        print(f"   Budget: {budget_value} EUR")
        print(f"   Custom Label 1: {custom_label_1}")
        print(f"   Ad Groups (shops): {len(ad_groups)}")

        try:
            # Campaign configuration
            merchant_center_account_id = MERCHANT_CENTER_ID
            budget_name = f"Budget_{campaign_name}"
            tracking_template = ""
            country = COUNTRY

            # Convert budget from EUR to micros
            try:
                budget_micros = int(float(budget_value) * 1_000_000) if budget_value else 10_000_000
            except (ValueError, TypeError):
                print(f"   ⚠️  Invalid budget value '{budget_value}', using default 10 EUR")
                budget_micros = 10_000_000

            # Get bid strategy based on custom label 1
            bid_strategy_resource_name = None
            if custom_label_1 and custom_label_1 in BID_STRATEGY_MAPPING:
                bid_strategy_name = BID_STRATEGY_MAPPING[custom_label_1]
                print(f"   Looking up bid strategy: {bid_strategy_name}")
                bid_strategy_resource_name = get_bid_strategy_by_name(
                    client=client,
                    customer_id=MCC_ACCOUNT_ID,
                    strategy_name=bid_strategy_name
                )

            # Get first ad group's shop info for campaign metadata
            first_ag_name = list(ad_groups.keys())[0]
            first_ag_data = ad_groups[first_ag_name]

            # Create campaign (status: PAUSED - set in add_standard_shopping_campaign)
            print(f"\n   Creating campaign: {campaign_name}")
            if dry_run:
                # Fake resource name keeps the downstream .split('/')[-1] logic happy
                campaign_resource_name = f"customers/{customer_id}/campaigns/DRY_RUN_{uuid.uuid4().hex[:10]}"
                print(f"   [DRY RUN] Would create campaign → {campaign_resource_name}")
            else:
                campaign_resource_name = add_standard_shopping_campaign(
                    client=client,
                    customer_id=customer_id,
                    merchant_center_account_id=merchant_center_account_id,
                    campaign_name=campaign_name,
                    budget_name=budget_name,
                    tracking_template=tracking_template,
                    country=country,
                    shopid=first_ag_data['shop_id'],
                    shopname=first_ag_name,
                    label=custom_label_1,
                    budget=budget_micros,
                    bidding_strategy_resource_name=bid_strategy_resource_name
                )

            if not campaign_resource_name:
                raise Exception("Failed to create/find campaign")

            print(f"   ✅ Campaign ready: {campaign_resource_name}")

            # Tag with DM_DASHBOARD label so operators can identify entities
            # created via the dashboard. Skipped in dry_run (fake resource name).
            if not dry_run and dm_label_resource:
                if apply_dm_dashboard_label_to_campaign(client, customer_id, campaign_resource_name, dm_label_resource):
                    print(f"   🏷️  Labeled campaign with '{DM_DASHBOARD_LABEL_NAME}'")

            # Add negative keyword list to campaign
            if NEGATIVE_LIST_NAME:
                if dry_run:
                    print(f"   [DRY RUN] Would attach negative keyword list '{NEGATIVE_LIST_NAME}'")
                else:
                    enable_negative_list_for_campaign(
                        client=client,
                        customer_id=customer_id,
                        campaign_resource_name=campaign_resource_name,
                        negative_list_name=NEGATIVE_LIST_NAME
                    )

            # Wait after campaign setup before processing ad groups
            time.sleep(1.0)

            # Process each ad group (shop) within this campaign
            print(f"\n   Processing {len(ad_groups)} ad group(s)...")
            ad_groups_processed = []
            ad_group_errors = {}

            for ag_idx, (shop_name, ag_data) in enumerate(ad_groups.items(), start=1):
                # Build ad group name: PLA/{shop_name}_{cl1}
                ad_group_name = f"PLA/{shop_name}_{custom_label_1}"
                print(f"\n   ──── Ad Group {ag_idx}/{len(ad_groups)}: {ad_group_name} ────")
                print(f"      (Shop: {shop_name})")

                try:
                    maincat_ids = sorted(ag_data['maincat_ids'])
                    print(f"      Maincat IDs (CL4): {maincat_ids}")

                    # Create ad group (status: ENABLED - set in add_shopping_ad_group)
                    if dry_run:
                        ad_group_resource_name = f"customers/{customer_id}/adGroups/DRY_RUN_{uuid.uuid4().hex[:10]}"
                        print(f"      [DRY RUN] Would create ad group → {ad_group_resource_name}")
                    else:
                        ad_group_resource_name, _ = add_shopping_ad_group(
                            client=client,
                            customer_id=customer_id,
                            campaign_resource_name=campaign_resource_name,
                            ad_group_name=ad_group_name,
                            campaign_name=campaign_name
                        )

                    if not ad_group_resource_name:
                        raise Exception(f"Failed to create/find ad group")

                    print(f"      ✅ Ad group ready: {ad_group_resource_name}")

                    # Tag ad group with DM_DASHBOARD label (live runs only)
                    if not dry_run and dm_label_resource:
                        if apply_dm_dashboard_label_to_ad_group(client, customer_id, ad_group_resource_name, dm_label_resource):
                            print(f"      🏷️  Labeled ad group with '{DM_DASHBOARD_LABEL_NAME}'")

                    # Wait after ad group creation before building tree (skip delay in dry_run)
                    if not dry_run:
                        time.sleep(1.0)

                    # Extract ad group ID
                    ad_group_id = ad_group_resource_name.split('/')[-1]

                    # For CL3 targeting, split shop_name at | and use first part
                    # e.g. "Hbm-machines.com|NL" becomes "Hbm-machines.com"
                    shop_name_for_targeting = shop_name.split('|')[0] if '|' in shop_name else shop_name
                    if shop_name_for_targeting != shop_name:
                        print(f"      CL3 targeting: '{shop_name_for_targeting}' (split from '{shop_name}')")

                    # Build listing tree with V2 function (CL1 targeting + dataedis exclusion)
                    if dry_run:
                        dataedis_msg = " → dataedis excluded" if EXCLUDE_DATAEDIS else ""
                        print(f"      [DRY RUN] Would build tree: Shop '{shop_name_for_targeting}' → {len(maincat_ids)} maincat(s) → CL1 '{custom_label_1}'{dataedis_msg}")
                    else:
                        build_listing_tree_for_inclusion_v2(
                            client=client,
                            customer_id=customer_id,
                            ad_group_id=ad_group_id,
                            shop_name=shop_name_for_targeting,  # Use split shop_name for CL3
                            maincat_ids=maincat_ids,
                            custom_label_1=custom_label_1
                        )

                    # Wait after tree creation before creating ad (skip delay in dry_run)
                    if not dry_run:
                        time.sleep(2.0)

                    # Create shopping product ad
                    print(f"      Creating shopping product ad...")
                    if dry_run:
                        print(f"      [DRY RUN] Would create shopping product ad")
                    else:
                        add_shopping_product_ad(
                            client=client,
                            customer_id=customer_id,
                            ad_group_resource_name=ad_group_resource_name
                        )

                    ad_groups_processed.append(shop_name)
                    print(f"      ✅ Ad group completed: {ad_group_name}")

                    # Wait between ad groups to prevent concurrent modification
                    time.sleep(1.0)

                except Exception as e:
                    error_msg = str(e)
                    print(f"      ❌ Failed: {error_msg}")
                    ad_group_errors[shop_name] = error_msg

            # Mark rows as successful/failed
            for shop_name, ag_data in ad_groups.items():
                for row_info in ag_data['rows']:
                    row_num = row_info['idx']
                    if shop_name in ad_groups_processed:
                        sheet.cell(row=row_num, column=COL_RESULT + 1).value = True
                        sheet.cell(row=row_num, column=COL_ERR + 1).value = ""
                    else:
                        sheet.cell(row=row_num, column=COL_RESULT + 1).value = False
                        error_msg = ad_group_errors.get(shop_name, "Failed to process ad group")
                        sheet.cell(row=row_num, column=COL_ERR + 1).value = error_msg[:100]

            if len(ad_groups_processed) > 0:
                successful_campaigns += 1
                print(f"\n   ✅ CAMPAIGN COMPLETED: {len(ad_groups_processed)}/{len(ad_groups)} ad groups processed")

        except Exception as e:
            error_msg = str(e)
            print(f"\n   ❌ CAMPAIGN FAILED: {error_msg}")
            # Mark all rows for this campaign as failed
            for row_info in campaign_data['rows']:
                row_num = row_info['idx']
                sheet.cell(row=row_num, column=COL_RESULT + 1).value = False
                sheet.cell(row=row_num, column=COL_ERR + 1).value = f"Campaign failed: {error_msg[:80]}"

        # Save periodically
        if file_path and campaign_idx % 5 == 0:
            print(f"\n   💾 Saving progress...")
            try:
                workbook.save(file_path)
            except Exception as save_error:
                print(f"   ⚠️  Error saving: {save_error}")

        # Wait between campaigns to prevent concurrent modification
        time.sleep(2.0)

    # Final save
    if file_path:
        print(f"\n💾 Final save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"⚠️  Error on final save: {save_error}")

    # Close data_only workbook if it was opened
    if data_workbook:
        data_workbook.close()

    print(f"\n{'='*70}")
    print(f"INCLUSION SHEET (V2) SUMMARY")
    print(f"{'='*70}")
    print(f"Total campaigns: {total_campaigns}")
    print(f"✅ Successful: {successful_campaigns}")
    print(f"❌ Failed: {total_campaigns - successful_campaigns}")
    print(f"{'='*70}\n")


def pause_ad_group(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_resource_name: str
) -> bool:
    """
    Pause an ad group by setting its status to PAUSED.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_resource_name: Resource name of the ad group to pause

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        from google.protobuf import field_mask_pb2

        ad_group_service = client.get_service("AdGroupService")
        ad_group_operation = client.get_type("AdGroupOperation")

        ad_group = ad_group_operation.update
        ad_group.resource_name = ad_group_resource_name
        ad_group.status = client.enums.AdGroupStatusEnum.PAUSED

        # Set the field mask to only update the status field
        ad_group_operation.update_mask.CopyFrom(
            field_mask_pb2.FieldMask(paths=["status"])
        )

        ad_group_service.mutate_ad_groups(
            customer_id=customer_id,
            operations=[ad_group_operation]
        )

        return True
    except GoogleAdsException as ex:
        print(f"      ❌ Google Ads API error: {ex.error.code().name}")
        for error in ex.failure.errors:
            print(f"         {error.message}")
        return False
    except Exception as e:
        print(f"      ❌ Error pausing ad group: {str(e)}")
        return False


def enable_ad_group(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_resource_name: str
) -> bool:
    """
    Enable an ad group by setting its status to ENABLED.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_resource_name: Resource name of the ad group to enable

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        from google.protobuf import field_mask_pb2

        ad_group_service = client.get_service("AdGroupService")
        ad_group_operation = client.get_type("AdGroupOperation")

        ad_group = ad_group_operation.update
        ad_group.resource_name = ad_group_resource_name
        ad_group.status = client.enums.AdGroupStatusEnum.ENABLED

        # Set the field mask to only update the status field
        ad_group_operation.update_mask.CopyFrom(
            field_mask_pb2.FieldMask(paths=["status"])
        )

        ad_group_service.mutate_ad_groups(
            customer_id=customer_id,
            operations=[ad_group_operation]
        )

        return True
    except GoogleAdsException as ex:
        print(f"      ❌ Google Ads API error: {ex.error.code().name}")
        for error in ex.failure.errors:
            print(f"         {error.message}")
        return False
    except Exception as e:
        print(f"      ❌ Error enabling ad group: {str(e)}")
        return False


def remove_ad_group(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_resource_name: str
) -> bool:
    """
    Remove (delete) an ad group using the remove operation.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_resource_name: Resource name of the ad group to remove

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        ad_group_service = client.get_service("AdGroupService")
        ad_group_operation = client.get_type("AdGroupOperation")

        # Use the remove operation instead of update with REMOVED status
        ad_group_operation.remove = ad_group_resource_name

        ad_group_service.mutate_ad_groups(
            customer_id=customer_id,
            operations=[ad_group_operation]
        )

        return True
    except GoogleAdsException as ex:
        print(f"      ❌ Google Ads API error: {ex.error.code().name}")
        for error in ex.failure.errors:
            print(f"         {error.message}")
        return False
    except Exception as e:
        print(f"      ❌ Error removing ad group: {str(e)}")
        return False


def find_ad_group_in_campaign(
    client: GoogleAdsClient,
    customer_id: str,
    campaign_name: str,
    ad_group_name: str
) -> Optional[Dict[str, Any]]:
    """
    Find an ad group by campaign name and ad group name.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        campaign_name: Name of the campaign
        ad_group_name: Name of the ad group

    Returns:
        dict with ad_group info or None if not found
    """
    google_ads_service = client.get_service("GoogleAdsService")

    # Escape special characters for GAQL (only single quotes need escaping)
    def escape_gaql_string(s):
        s = s.replace("'", "\\'")
        return s

    escaped_campaign_name = escape_gaql_string(campaign_name)
    escaped_ad_group_name = escape_gaql_string(ad_group_name)

    query = f"""
        SELECT
            ad_group.id,
            ad_group.resource_name,
            ad_group.name,
            ad_group.status,
            campaign.id,
            campaign.name,
            campaign.resource_name,
            campaign.status
        FROM ad_group
        WHERE campaign.name = '{escaped_campaign_name}'
        AND ad_group.name = '{escaped_ad_group_name}'
        AND ad_group.status IN ('ENABLED', 'PAUSED')
        AND campaign.status != 'REMOVED'
        LIMIT 1
    """

    try:
        response = google_ads_service.search(customer_id=customer_id, query=query)
        for row in response:
            return {
                'ad_group_id': row.ad_group.id,
                'ad_group_resource_name': row.ad_group.resource_name,
                'ad_group_name': row.ad_group.name,
                'ad_group_status': row.ad_group.status.name,
                'campaign_id': row.campaign.id,
                'campaign_name': row.campaign.name,
                'campaign_resource_name': row.campaign.resource_name
            }
    except GoogleAdsException as ex:
        print(f"      ❌ Error searching for ad group: {ex}")

    return None


def process_reverse_inclusion_sheet_v2(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None,
    dry_run: bool = False,
):
    """
    Process the 'toevoegen' sheet - removes (deletes) ad groups.

    This is the reverse of process_inclusion_sheet_v2. Instead of creating
    ad groups, it finds existing ad groups and removes them.

    Excel columns:
    A. shop_name - Used to build ad group name: PLA/{shop_name}_{cl1}
    B. Shop ID (not used)
    C. maincat - Used to build campaign name: PLA/{maincat} store_{cl1}
    D. maincat_id (not used)
    E. custom label 1 - Used in both campaign and ad group names
    F. budget (ignored)
    G. result (TRUE/FALSE) - updated by script
    H. Error message

    Naming conventions:
    - Campaign name: PLA/{maincat} store_{cl1}
    - Ad group name: PLA/{shop_name}_{cl1}

    Groups rows by derived campaign name (from maincat + cl1), then removes each
    unique ad group (derived from shop_name + cl1) within that campaign.

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving)
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING REVERSE INCLUSION SHEET (V2): '{SHEET_REVERSE_INCLUSION}'")
    print(f"(REMOVING AD GROUPS)")
    if dry_run:
        print("(DRY RUN: no ad groups will actually be removed from Google Ads)")
    print(f"{'='*70}\n")

    try:
        sheet = workbook[SHEET_REVERSE_INCLUSION]
    except KeyError:
        print(f"❌ Sheet '{SHEET_REVERSE_INCLUSION}' not found in workbook")
        return

    # Column indices for toevoegen sheet
    COL_SHOP_NAME = 0      # A: Shop name (= ad group name)
    COL_SHOP_ID = 1        # B: Shop ID (not used)
    COL_MAINCAT = 2        # C: maincat (category name)
    COL_MAINCAT_ID = 3     # D: maincat_id (not used)
    COL_CL1 = 4            # E: custom label 1
    COL_BUDGET = 5         # F: budget (ignored)
    COL_RESULT = 6         # G: result (TRUE/FALSE)
    COL_ERR = 7            # H: Error message

    # Step 1: Read all rows and group by campaign (derived from maincat + cl1)
    # Each shop_name is an ad group within that campaign
    campaigns_to_process = defaultdict(lambda: {
        'maincat': None,
        'cl1': None,
        'ad_groups': defaultdict(lambda: {'rows': []})  # shop_name -> rows
    })

    print("Step 1: Reading and grouping rows by campaign (maincat + cl1) and ad group (shop_name)...")
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if status column is empty
        status_value = row[COL_RESULT].value if len(row) > COL_RESULT else None

        # Skip rows that already have a status (TRUE/FALSE)
        if status_value is not None and status_value != '':
            continue

        shop_name = row[COL_SHOP_NAME].value  # This is the ad group name
        maincat = row[COL_MAINCAT].value
        cl1 = row[COL_CL1].value

        # Validate required fields
        if not shop_name or not maincat or not cl1:
            print(f"   ⚠️  [Row {idx}] Missing required fields (shop_name/maincat/cl1), skipping")
            sheet.cell(row=idx, column=COL_RESULT + 1).value = False
            sheet.cell(row=idx, column=COL_ERR + 1).value = "Missing required fields"
            continue

        # Build campaign name from maincat and cl1
        campaign_name = f"PLA/{maincat} store_{cl1}"

        # Group by campaign, then by ad group (shop_name)
        campaigns_to_process[campaign_name]['maincat'] = maincat
        campaigns_to_process[campaign_name]['cl1'] = cl1
        campaigns_to_process[campaign_name]['ad_groups'][shop_name]['rows'].append({'idx': idx, 'row': row})

    total_campaigns = len(campaigns_to_process)
    total_ad_groups = sum(len(c['ad_groups']) for c in campaigns_to_process.values())
    print(f"   Found {total_campaigns} campaign(s) with {total_ad_groups} unique ad group(s) to remove")

    # Step 2: Process each campaign and its ad groups
    successful_removals = 0
    failed_removals = 0
    processed_ag_count = 0

    for camp_idx, (campaign_name, campaign_data) in enumerate(campaigns_to_process.items(), start=1):
        print(f"\n{'─'*70}")
        print(f"CAMPAIGN {camp_idx}/{total_campaigns}: {campaign_name}")
        print(f"{'─'*70}")
        print(f"   Maincat: {campaign_data['maincat']}")
        print(f"   Custom Label 1: {campaign_data['cl1']}")
        print(f"   Ad Groups to remove: {len(campaign_data['ad_groups'])}")

        # Process each ad group (shop_name) in this campaign
        for shop_name, ag_data in campaign_data['ad_groups'].items():
            processed_ag_count += 1

            # Build ad group name: PLA/{shop_name}_{cl1}
            ad_group_name = f"PLA/{shop_name}_{campaign_data['cl1']}"
            print(f"\n   ──── Ad Group: {ad_group_name} ────")
            print(f"      (Shop: {shop_name})")

            try:
                # Find the ad group
                print(f"      Searching for ad group...")
                ad_group_info = find_ad_group_in_campaign(
                    client=client,
                    customer_id=customer_id,
                    campaign_name=campaign_name,
                    ad_group_name=ad_group_name
                )

                if not ad_group_info:
                    raise Exception(f"Ad group not found in campaign")

                print(f"      ✅ Found ad group (ID: {ad_group_info['ad_group_id']})")
                print(f"         Current status: {ad_group_info['ad_group_status']}")

                # Check if already removed
                if ad_group_info['ad_group_status'] == 'REMOVED':
                    print(f"      ℹ️  Ad group is already REMOVED")
                    # Mark as successful anyway
                    for row_info in ag_data['rows']:
                        row_num = row_info['idx']
                        sheet.cell(row=row_num, column=COL_RESULT + 1).value = True
                        sheet.cell(row=row_num, column=COL_ERR + 1).value = "Already removed"
                    successful_removals += 1
                    continue

                # Remove the ad group (skipped under dry_run)
                if dry_run:
                    print(f"      [DRY RUN] Would remove ad group")
                    success = True
                else:
                    print(f"      Removing ad group...")
                    success = remove_ad_group(
                        client=client,
                        customer_id=customer_id,
                        ad_group_resource_name=ad_group_info['ad_group_resource_name']
                    )

                if success:
                    print(f"      ✅ Ad group removed successfully")
                    successful_removals += 1
                    # Mark all rows for this ad group as successful
                    for row_info in ag_data['rows']:
                        row_num = row_info['idx']
                        sheet.cell(row=row_num, column=COL_RESULT + 1).value = True
                        sheet.cell(row=row_num, column=COL_ERR + 1).value = ""
                else:
                    raise Exception("Failed to remove ad group")

                time.sleep(0.3)  # Small delay between API calls

            except Exception as e:
                error_msg = str(e)
                print(f"      ❌ Failed: {error_msg}")
                failed_removals += 1
                # Mark all rows for this ad group as failed
                for row_info in ag_data['rows']:
                    row_num = row_info['idx']
                    sheet.cell(row=row_num, column=COL_RESULT + 1).value = False
                    sheet.cell(row=row_num, column=COL_ERR + 1).value = error_msg[:100]

            # Save periodically
            if file_path and processed_ag_count % 10 == 0:
                print(f"\n   💾 Saving progress...")
                try:
                    workbook.save(file_path)
                except Exception as save_error:
                    print(f"   ⚠️  Error saving: {save_error}")

    # Final save
    if file_path:
        print(f"\n💾 Final save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"⚠️  Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"REVERSE INCLUSION SHEET (V2) SUMMARY")
    print(f"{'='*70}")
    print(f"Total campaigns: {total_campaigns}")
    print(f"Total ad groups: {total_ad_groups}")
    print(f"✅ Removed: {successful_removals}")
    print(f"❌ Failed: {failed_removals}")
    print(f"{'='*70}\n")


def process_enable_inclusion_sheet_v2(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None,
    sheet_name: str = "adgroups_heractiveren"
):
    """
    Process a sheet to ENABLE ad groups (reverse of pause).

    This enables previously paused ad groups, bringing them back online.

    Excel columns (adgroups_heractiveren):
    A. shop_name - Used to build ad group name: PLA/{shop_name}_{cl1}
    B. Shop ID (not used)
    C. maincat - Used to build campaign name: PLA/{maincat} store_{cl1}
    D. maincat_id (not used)
    E. custom label 1 - Used in both campaign and ad group names
    F. budget (ignored)
    G. result (TRUE/FALSE) - updated by script
    H. Error message

    Naming conventions:
    - Campaign name: PLA/{maincat} store_{cl1}
    - Ad group name: PLA/{shop_name}_{cl1}

    Groups rows by derived campaign name (from maincat + cl1), then enables each
    unique ad group (derived from shop_name + cl1) within that campaign.

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving)
        sheet_name: Name of sheet to process (default: 'hervatten')
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING ENABLE INCLUSION SHEET (V2): '{sheet_name}'")
    print(f"{'='*70}\n")

    try:
        sheet = workbook[sheet_name]
    except KeyError:
        print(f"❌ Sheet '{sheet_name}' not found in workbook")
        return

    # Column indices for this sheet
    COL_SHOP_NAME = 0      # A: Shop name (= ad group name)
    COL_SHOP_ID = 1        # B: Shop ID (not used)
    COL_MAINCAT = 2        # C: maincat (category name)
    COL_MAINCAT_ID = 3     # D: maincat_id (not used)
    COL_CL1 = 4            # E: custom label 1
    COL_BUDGET = 5         # F: budget (ignored)
    COL_RESULT = 6         # G: result (TRUE/FALSE)
    COL_ERR = 7            # H: Error message

    # Step 1: Read all rows and group by campaign (derived from maincat + cl1)
    # Each shop_name is an ad group within that campaign
    campaigns_to_process = defaultdict(lambda: {
        'maincat': None,
        'cl1': None,
        'ad_groups': defaultdict(lambda: {'rows': []})  # shop_name -> rows
    })

    print("Step 1: Reading and grouping rows by campaign (maincat + cl1) and ad group (shop_name)...")
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if status column is empty
        status_value = row[COL_RESULT].value if len(row) > COL_RESULT else None

        # Skip rows that already have a status (TRUE/FALSE)
        if status_value is not None and status_value != '':
            continue

        shop_name = row[COL_SHOP_NAME].value  # This is the ad group name
        maincat = row[COL_MAINCAT].value
        cl1 = row[COL_CL1].value

        # Validate required fields
        if not shop_name or not maincat or not cl1:
            print(f"   ⚠️  [Row {idx}] Missing required fields (shop_name/maincat/cl1), skipping")
            sheet.cell(row=idx, column=COL_RESULT + 1).value = False
            sheet.cell(row=idx, column=COL_ERR + 1).value = "Missing required fields"
            continue

        # Build campaign name from maincat and cl1
        campaign_name = f"PLA/{maincat} store_{cl1}"

        # Group by campaign, then by ad group (shop_name)
        campaigns_to_process[campaign_name]['maincat'] = maincat
        campaigns_to_process[campaign_name]['cl1'] = cl1
        campaigns_to_process[campaign_name]['ad_groups'][shop_name]['rows'].append({'idx': idx, 'row': row})

    total_campaigns = len(campaigns_to_process)
    total_ad_groups = sum(len(c['ad_groups']) for c in campaigns_to_process.values())
    print(f"   Found {total_campaigns} campaign(s) with {total_ad_groups} unique ad group(s) to enable")

    # Step 2: Process each campaign and its ad groups
    successful_enables = 0
    failed_enables = 0
    processed_ag_count = 0

    for camp_idx, (campaign_name, campaign_data) in enumerate(campaigns_to_process.items(), start=1):
        print(f"\n{'─'*70}")
        print(f"CAMPAIGN {camp_idx}/{total_campaigns}: {campaign_name}")
        print(f"{'─'*70}")
        print(f"   Maincat: {campaign_data['maincat']}")
        print(f"   Custom Label 1: {campaign_data['cl1']}")
        print(f"   Ad Groups to enable: {len(campaign_data['ad_groups'])}")

        # Process each ad group (shop_name) in this campaign
        for shop_name, ag_data in campaign_data['ad_groups'].items():
            processed_ag_count += 1

            # Build ad group name: PLA/{shop_name}_{cl1}
            ad_group_name = f"PLA/{shop_name}_{campaign_data['cl1']}"
            print(f"\n   ──── Ad Group: {ad_group_name} ────")
            print(f"      (Shop: {shop_name})")

            try:
                # Find the ad group
                print(f"      Searching for ad group...")
                ad_group_info = find_ad_group_in_campaign(
                    client=client,
                    customer_id=customer_id,
                    campaign_name=campaign_name,
                    ad_group_name=ad_group_name
                )

                if not ad_group_info:
                    raise Exception(f"Ad group not found in campaign")

                print(f"      ✅ Found ad group (ID: {ad_group_info['ad_group_id']})")
                print(f"         Current status: {ad_group_info['ad_group_status']}")

                # Check if already enabled
                if ad_group_info['ad_group_status'] == 'ENABLED':
                    print(f"      ℹ️  Ad group is already ENABLED")
                    # Mark as successful anyway
                    for row_info in ag_data['rows']:
                        row_num = row_info['idx']
                        sheet.cell(row=row_num, column=COL_RESULT + 1).value = True
                        sheet.cell(row=row_num, column=COL_ERR + 1).value = "Already enabled"
                    successful_enables += 1
                    continue

                # Enable the ad group
                print(f"      Enabling ad group...")
                success = enable_ad_group(
                    client=client,
                    customer_id=customer_id,
                    ad_group_resource_name=ad_group_info['ad_group_resource_name']
                )

                if success:
                    print(f"      ✅ Ad group enabled successfully")
                    successful_enables += 1
                    # Mark all rows for this ad group as successful
                    for row_info in ag_data['rows']:
                        row_num = row_info['idx']
                        sheet.cell(row=row_num, column=COL_RESULT + 1).value = True
                        sheet.cell(row=row_num, column=COL_ERR + 1).value = ""
                else:
                    raise Exception("Failed to enable ad group")

                time.sleep(0.3)  # Small delay between API calls

            except Exception as e:
                error_msg = str(e)
                print(f"      ❌ Failed: {error_msg}")
                failed_enables += 1
                # Mark all rows for this ad group as failed
                for row_info in ag_data['rows']:
                    row_num = row_info['idx']
                    sheet.cell(row=row_num, column=COL_RESULT + 1).value = False
                    sheet.cell(row=row_num, column=COL_ERR + 1).value = error_msg[:100]

            # Save periodically
            if file_path and processed_ag_count % 10 == 0:
                print(f"\n   💾 Saving progress...")
                try:
                    workbook.save(file_path)
                except Exception as save_error:
                    print(f"   ⚠️  Error saving: {save_error}")

    # Final save
    if file_path:
        print(f"\n💾 Final save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"⚠️  Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"ENABLE INCLUSION SHEET (V2) SUMMARY")
    print(f"{'='*70}")
    print(f"Total campaigns: {total_campaigns}")
    print(f"Total ad groups: {total_ad_groups}")
    print(f"✅ Enabled: {successful_enables}")
    print(f"❌ Failed: {failed_enables}")
    print(f"{'='*70}\n")


def process_pause_ad_groups_sheet(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None,
    sheet_name: str = "adgroups_pauzeren"
):
    """
    Process the 'adgroups_pauzeren' sheet - pauses ad groups.

    Excel columns (adgroups_pauzeren):
    A. shop_name - Used to build ad group name: PLA/{shop_name}_{custom_value}
    B. Shop ID (not used)
    C. maincat - Used to build campaign name: PLA/{maincat} store_{custom_value}
    D. check (not used)
    E. maincat_id (not used)
    F. custom label 1 - Used in both campaign and ad group names
    G. budget (not used)
    H. result (TRUE/FALSE) - updated by script
    I. Error message

    Naming conventions:
    - Campaign name: PLA/{maincat} store_{custom_value}
    - Ad group name: PLA/{shop_name}_{custom_value}

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving)
        sheet_name: Name of sheet to process (default: 'adgroups_pauzeren')
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING PAUSE AD GROUPS SHEET: '{sheet_name}'")
    print(f"{'='*70}\n")

    try:
        sheet = workbook[sheet_name]
    except KeyError:
        print(f"❌ Sheet '{sheet_name}' not found in workbook")
        return

    # Column indices
    COL_SHOP_NAME = 0      # A: Shop name
    COL_MAINCAT = 2        # C: maincat
    COL_CUSTOM_VALUE = 5   # F: custom label 1
    COL_RESULT = 7         # H: result (TRUE/FALSE)
    COL_ERR = 8            # I: Error message

    # Step 1: Read all rows and group by campaign
    campaigns_to_process = defaultdict(lambda: {
        'maincat': None,
        'custom_value': None,
        'ad_groups': defaultdict(lambda: {'rows': []})
    })

    print("Step 1: Reading and grouping rows by campaign and ad group...")
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Skip rows that already have a status
        status_value = row[COL_RESULT].value if len(row) > COL_RESULT else None
        if status_value is not None and status_value != '':
            continue

        shop_name = row[COL_SHOP_NAME].value
        maincat = row[COL_MAINCAT].value
        custom_value = row[COL_CUSTOM_VALUE].value

        # Validate required fields
        if not shop_name or not maincat or not custom_value:
            print(f"   ⚠️  [Row {idx}] Missing required fields (shop_name/maincat/custom_value), skipping")
            sheet.cell(row=idx, column=COL_RESULT + 1).value = False
            sheet.cell(row=idx, column=COL_ERR + 1).value = "Missing required fields"
            continue

        custom_value = str(custom_value).strip()

        # Build campaign name: PLA/{maincat} store_{custom_value}
        campaign_name = f"PLA/{maincat} store_{custom_value}"

        campaigns_to_process[campaign_name]['maincat'] = maincat
        campaigns_to_process[campaign_name]['custom_value'] = custom_value
        campaigns_to_process[campaign_name]['ad_groups'][shop_name]['rows'].append({'idx': idx, 'row': row})

    total_campaigns = len(campaigns_to_process)
    total_ad_groups = sum(len(c['ad_groups']) for c in campaigns_to_process.values())
    print(f"   Found {total_campaigns} campaign(s) with {total_ad_groups} unique ad group(s) to pause")

    # Step 2: Process each campaign and its ad groups
    successful_pauses = 0
    failed_pauses = 0
    processed_ag_count = 0

    for camp_idx, (campaign_name, campaign_data) in enumerate(campaigns_to_process.items(), start=1):
        print(f"\n{'─'*70}")
        print(f"CAMPAIGN {camp_idx}/{total_campaigns}: {campaign_name}")
        print(f"{'─'*70}")
        print(f"   Maincat: {campaign_data['maincat']}")
        print(f"   Custom Value: {campaign_data['custom_value']}")
        print(f"   Ad Groups to pause: {len(campaign_data['ad_groups'])}")

        for shop_name, ag_data in campaign_data['ad_groups'].items():
            processed_ag_count += 1

            # Build ad group name: PLA/{shop_name}_{custom_value}
            ad_group_name = f"PLA/{shop_name}_{campaign_data['custom_value']}"
            print(f"\n   ──── Ad Group: {ad_group_name} ────")
            print(f"      (Shop: {shop_name})")

            try:
                # Find the ad group
                print(f"      Searching for ad group...")
                ad_group_info = find_ad_group_in_campaign(
                    client=client,
                    customer_id=customer_id,
                    campaign_name=campaign_name,
                    ad_group_name=ad_group_name
                )

                if not ad_group_info:
                    raise Exception(f"Ad group not found in campaign")

                print(f"      ✅ Found ad group (ID: {ad_group_info['ad_group_id']})")
                print(f"         Current status: {ad_group_info['ad_group_status']}")

                # Check if already paused
                if ad_group_info['ad_group_status'] == 'PAUSED':
                    print(f"      ℹ️  Ad group is already PAUSED")
                    for row_info in ag_data['rows']:
                        row_num = row_info['idx']
                        sheet.cell(row=row_num, column=COL_RESULT + 1).value = True
                        sheet.cell(row=row_num, column=COL_ERR + 1).value = "Already paused"
                    successful_pauses += 1
                    continue

                # Pause the ad group
                print(f"      Pausing ad group...")
                success = pause_ad_group(
                    client=client,
                    customer_id=customer_id,
                    ad_group_resource_name=ad_group_info['ad_group_resource_name']
                )

                if success:
                    print(f"      ✅ Ad group paused successfully")
                    successful_pauses += 1
                    for row_info in ag_data['rows']:
                        row_num = row_info['idx']
                        sheet.cell(row=row_num, column=COL_RESULT + 1).value = True
                        sheet.cell(row=row_num, column=COL_ERR + 1).value = ""
                else:
                    raise Exception("Failed to pause ad group")

                time.sleep(0.3)  # Small delay between API calls

            except Exception as e:
                error_msg = str(e)
                print(f"      ❌ Failed: {error_msg}")
                failed_pauses += 1
                for row_info in ag_data['rows']:
                    row_num = row_info['idx']
                    sheet.cell(row=row_num, column=COL_RESULT + 1).value = False
                    sheet.cell(row=row_num, column=COL_ERR + 1).value = error_msg[:100]

            # Save periodically
            if file_path and processed_ag_count % 10 == 0:
                print(f"\n   💾 Saving progress...")
                try:
                    workbook.save(file_path)
                except Exception as save_error:
                    print(f"   ⚠️  Error saving: {save_error}")

    # Final save
    if file_path:
        print(f"\n💾 Final save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"⚠️  Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"PAUSE AD GROUPS SHEET SUMMARY")
    print(f"{'='*70}")
    print(f"Total campaigns: {total_campaigns}")
    print(f"Total ad groups: {total_ad_groups}")
    print(f"✅ Paused: {successful_pauses}")
    print(f"❌ Failed: {failed_pauses}")
    print(f"{'='*70}\n")


def process_inclusion_sheet_legacy(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None
):
    """
    Process the 'toevoegen' (inclusion) sheet - LEGACY VERSION.

    Excel columns (OLD):
    A. Shop name
    B. Shop ID
    C. maincat
    D. maincat_id
    E. custom label 1
    F. budget (daily budget in EUR)
    G. Status (TRUE/FALSE) - updated by script

    Groups rows by unique combination of (maincat, custom_label_1) ONLY.
    For each group:
    1. Create ONE campaign with name: PLA/{maincat} store_{custom_label_1}
       - Uses budget from column F (converted to micros)
       - Applies bid strategy from MCC based on custom_label_1
    2. Create MULTIPLE ad groups (one per unique shop_name in group)
       - Ad group names: PLA/{shop_name}_{custom_label_1}
    3. Build listing tree for EACH ad group:
       - Target maincat_id as custom label 4
       - Subdivide and target shop_name as custom label 3
       - Exclude everything else at both levels
       - Bid: 1 cent (10,000 micros)
    4. Update column G (status) with TRUE/FALSE per row based on shop success

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving progress)
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING INCLUSION SHEET (LEGACY): '{SHEET_INCLUSION}'")
    print(f"{'='*70}\n")

    try:
        sheet = workbook[SHEET_INCLUSION]
    except KeyError:
        print(f"❌ Sheet '{SHEET_INCLUSION}' not found in workbook")
        return

    # Step 1: Read all rows and group by (maincat, custom_label_1) only
    groups = defaultdict(list)  # key: (maincat, custom_label_1), value: list of row data

    print("Step 1: Reading and grouping rows...")
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if status column (G) is empty - if so, this is where we start processing
        status_value = row[COL_LEGACY_STATUS].value

        # Skip rows that already have a status (TRUE/FALSE)
        if status_value is not None and status_value != '':
            continue

        shop_name = row[COL_LEGACY_SHOP_NAME].value
        shop_id = row[COL_LEGACY_SHOP_ID].value
        maincat = row[COL_LEGACY_MAINCAT].value
        maincat_id = row[COL_LEGACY_MAINCAT_ID].value
        custom_label_1 = row[COL_LEGACY_CUSTOM_LABEL_1].value
        budget = row[COL_LEGACY_BUDGET].value

        # Validate required fields
        if not shop_name or not maincat or not maincat_id or not custom_label_1:
            print(f"   ⚠️  [Row {idx}] Missing required fields (shop_name/maincat/maincat_id/custom_label_1), skipping")
            row[COL_LEGACY_STATUS].value = False
            # Only write to error column if it exists
            if len(row) > COL_LEGACY_ERROR:
                row[COL_LEGACY_ERROR].value = "Missing required fields (shop_name/maincat/maincat_id/custom_label_1)"
            continue

        # Group by (maincat, custom_label_1) only - multiple shops per campaign
        group_key = (maincat, custom_label_1)

        # Store row data
        groups[group_key].append({
            'row_idx': idx,
            'row_obj': row,
            'shop_name': shop_name,
            'shop_id': shop_id,
            'maincat': maincat,
            'maincat_id': maincat_id,
            'custom_label_1': custom_label_1,
            'budget': budget
        })

    print(f"   Found {len(groups)} unique group(s) to process\n")

    # Step 2: Process each group
    total_groups = len(groups)
    successful_groups = 0

    for group_idx, (group_key, rows_in_group) in enumerate(groups.items(), start=1):
        # Delay between groups to prevent concurrent modification errors
        if group_idx > 1:
            time.sleep(2)

        maincat, custom_label_1 = group_key

        print(f"\n{'─'*70}")
        print(f"GROUP {group_idx}/{total_groups}: {maincat} | {custom_label_1}")
        print(f"   Rows in group: {len(rows_in_group)}")
        print(f"{'─'*70}")

        # Get metadata from first row (all rows in group share same maincat, maincat_id, budget)
        first_row = rows_in_group[0]
        maincat_id = first_row['maincat_id']
        budget_value = first_row['budget']

        # Get unique shops in this group
        unique_shops = {}  # shop_name -> shop_id mapping
        for row_data in rows_in_group:
            unique_shops[row_data['shop_name']] = row_data['shop_id']

        print(f"   Maincat ID: {maincat_id}")
        print(f"   Budget: {budget_value} EUR")
        print(f"   Unique shops in group: {len(unique_shops)}")

        try:
            # Build campaign name: PLA/{maincat} store_{custom_label_1}
            campaign_name = f"PLA/{maincat} store_{custom_label_1}"
            print(f"\n   Step 1: Checking for existing campaign or creating new: {campaign_name}")

            # Campaign configuration
            merchant_center_account_id = MERCHANT_CENTER_ID
            budget_name = f"Budget_{campaign_name}"
            tracking_template = ""  # Not needed
            country = COUNTRY

            # Convert budget from EUR to micros (EUR * 1,000,000)
            # Default to 10 EUR if budget is missing or invalid
            try:
                budget_micros = int(float(budget_value) * 1_000_000) if budget_value else 10_000_000
            except (ValueError, TypeError):
                print(f"   ⚠️  Invalid budget value '{budget_value}', using default 10 EUR")
                budget_micros = 10_000_000

            # Get bid strategy based on custom label 1 (from MCC account)
            bid_strategy_resource_name = None
            if custom_label_1 in BID_STRATEGY_MAPPING:
                bid_strategy_name = BID_STRATEGY_MAPPING[custom_label_1]
                print(f"   Looking up bid strategy: {bid_strategy_name} (in MCC account)")
                bid_strategy_resource_name = get_bid_strategy_by_name(
                    client=client,
                    customer_id=MCC_ACCOUNT_ID,  # Search in MCC account
                    strategy_name=bid_strategy_name
                )

            # Use first shop's ID for campaign metadata
            first_shop_id = list(unique_shops.values())[0]
            first_shop_name = list(unique_shops.keys())[0]

            campaign_resource_name = add_standard_shopping_campaign(
                client=client,
                customer_id=customer_id,
                merchant_center_account_id=merchant_center_account_id,
                campaign_name=campaign_name,
                budget_name=budget_name,
                tracking_template=tracking_template,
                country=country,
                shopid=first_shop_id,
                shopname=first_shop_name,
                label=custom_label_1,
                budget=budget_micros,
                bidding_strategy_resource_name=bid_strategy_resource_name
            )

            if not campaign_resource_name:
                raise Exception("Failed to create/find campaign")

            print(f"   Campaign resource: {campaign_resource_name}")

            # Check/create multiple ad groups - one for each unique shop
            print(f"\n   Step 2: Processing ad groups for {len(unique_shops)} shop(s)...")
            shops_processed_successfully = []
            shop_errors = {}  # Track errors per shop

            for shop_idx, (shop_name, shop_id) in enumerate(unique_shops.items(), start=1):
                # Delay between shops to prevent concurrent modification errors
                if shop_idx > 1:
                    time.sleep(1.5)

                print(f"\n   ──── Shop {shop_idx}/{len(unique_shops)}: {shop_name} ────")

                try:
                    # Build ad group name: PLA/{shop_name}_{custom_label_1}
                    ad_group_name = f"PLA/{shop_name}_{custom_label_1}"
                    print(f"      Checking/creating ad group: {ad_group_name}")

                    ad_group_resource_name, _ = add_shopping_ad_group(
                        client=client,
                        customer_id=customer_id,
                        campaign_resource_name=campaign_resource_name,
                        ad_group_name=ad_group_name,
                        campaign_name=campaign_name
                    )

                    if not ad_group_resource_name:
                        raise Exception(f"Failed to create/find ad group for {shop_name}")

                    print(f"      ✅ Ad group ready: {ad_group_resource_name}")

                    # Extract ad group ID from resource name
                    ad_group_id = ad_group_resource_name.split('/')[-1]

                    # Build listing tree for this shop
                    print(f"      Building listing tree...")
                    build_listing_tree_for_inclusion(
                        client=client,
                        customer_id=customer_id,
                        ad_group_id=ad_group_id,
                        custom_label_1=custom_label_1,
                        maincat_id=maincat_id,
                        shop_name=shop_name,
                        default_bid_micros=DEFAULT_BID_MICROS
                    )

                    print(f"      ✅ Listing tree created for {shop_name}")

                    # Create shopping product ad in the ad group
                    print(f"      Creating shopping product ad...")
                    ad_resource_name = add_shopping_product_ad(
                        client=client,
                        customer_id=customer_id,
                        ad_group_resource_name=ad_group_resource_name
                    )

                    if not ad_resource_name:
                        print(f"      ⚠️  Warning: Failed to create shopping ad for {shop_name}")

                    shops_processed_successfully.append(shop_name)

                    # Small delay to avoid concurrent modification issues
                    time.sleep(1)

                except Exception as e:
                    error_msg = str(e)
                    print(f"      ❌ Failed to process shop {shop_name}: {error_msg}")
                    shop_errors[shop_name] = error_msg
                    # Continue with next shop instead of failing entire group

            # Mark rows as successful/failed based on their shop
            for row_data in rows_in_group:
                if row_data['shop_name'] in shops_processed_successfully:
                    row_data['row_obj'][COL_LEGACY_STATUS].value = True
                    # Clear error message on success (only if column exists)
                    if len(row_data['row_obj']) > COL_LEGACY_ERROR:
                        row_data['row_obj'][COL_LEGACY_ERROR].value = ""
                else:
                    row_data['row_obj'][COL_LEGACY_STATUS].value = False
                    # Add error message if available (only if column exists)
                    if len(row_data['row_obj']) > COL_LEGACY_ERROR:
                        if row_data['shop_name'] in shop_errors:
                            row_data['row_obj'][COL_LEGACY_ERROR].value = shop_errors[row_data['shop_name']]
                        else:
                            row_data['row_obj'][COL_LEGACY_ERROR].value = "Failed to process shop"

            if len(shops_processed_successfully) > 0:
                successful_groups += 1
                print(f"\n   ✅ GROUP {group_idx} COMPLETED: {len(shops_processed_successfully)}/{len(unique_shops)} shops processed")

            # Save progress periodically
            if file_path and group_idx % 5 == 0:
                print(f"\n   💾 Saving progress...")
                try:
                    workbook.save(file_path)
                except Exception as save_error:
                    print(f"   ⚠️  Failed to save progress: {save_error}")

        except Exception as e:
            error_msg = str(e)
            print(f"\n   ❌ GROUP {group_idx} FAILED: {error_msg}")
            # Mark all rows in this group as failed
            for row_data in rows_in_group:
                row_data['row_obj'][COL_LEGACY_STATUS].value = False
                # Only write error message if column exists
                if len(row_data['row_obj']) > COL_LEGACY_ERROR:
                    row_data['row_obj'][COL_LEGACY_ERROR].value = f"Group failed: {error_msg}"

    # Final save
    if file_path:
        print(f"\n💾 Final save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"⚠️  Failed to save: {save_error}")

    print(f"\n{'='*70}")
    print(f"INCLUSION SHEET (LEGACY) SUMMARY: {successful_groups}/{total_groups} groups processed successfully")
    print(f"{'='*70}\n")


def _process_single_exclusion_row(
    row_data: dict,
    client: GoogleAdsClient,
    customer_id: str,
    rate_limit_seconds: float
) -> dict:
    """
    Process a single exclusion row (worker function for parallel processing).

    Args:
        row_data: Dict containing row information
        client: Google Ads client
        customer_id: Customer ID
        rate_limit_seconds: Rate limit delay

    Returns:
        Dict with results: {'success': bool, 'error': str or None}
    """
    idx = row_data['idx']
    shop_name = row_data['shop_name']
    cat_uitsluiten = row_data['cat_uitsluiten']
    custom_label_1 = row_data['custom_label_1']

    print(f"\n[Row {idx}] Processing exclusion for shop: {shop_name}")
    print(f"         Category: {cat_uitsluiten}, Custom Label 1: {custom_label_1}")

    # Build campaign name pattern
    campaign_pattern = f"PLA/{cat_uitsluiten}_{custom_label_1}"
    print(f"   Searching for campaign+ad group: {campaign_pattern}")

    # Use combined lookup (saves 1 API call)
    result = get_campaign_and_ad_group_by_pattern(client, customer_id, campaign_pattern)
    if not result:
        print(f"   ❌ Campaign or ad group not found")
        return {
            'success': False,
            'error': f"Campaign not found: {campaign_pattern}"
        }

    campaign = result['campaign']
    ad_group = result['ad_group']

    print(f"   ✅ Found campaign: {campaign['name']} (ID: {campaign['id']})")
    print(f"   ✅ Found ad group: {ad_group['name']} (ID: {ad_group['id']})")

    # Rebuild tree with shop name exclusion
    try:
        rebuild_tree_with_custom_label_3_exclusion(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group['id'],
            shop_name=shop_name,
            default_bid_micros=DEFAULT_BID_MICROS
        )
        print(f"   ✅ SUCCESS - Row {idx} completed")

        # Rate limiting ONLY after successful processing
        if rate_limit_seconds > 0:
            time.sleep(rate_limit_seconds)

        return {'success': True, 'error': None}

    except Exception as e:
        error_msg = str(e)
        print(f"   ❌ Error rebuilding tree: {error_msg}")
        return {
            'success': False,
            'error': f"Error rebuilding tree: {error_msg[:500]}"
        }


def process_uitbreiding_sheet(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None,
    save_interval: int = 5
):
    """
    Process the uitbreiding (extension) sheet - adds shops to existing category campaigns.

    OPTIMIZED VERSION:
    - Groups shops by (maincat, cl1) for batch processing
    - Finds/creates campaign ONCE per group instead of once per shop
    - Reduces API calls significantly

    Excel columns:
    A. Shop name
    B. Shop ID (not used)
    C. maincat (category name)
    D. maincat_id (used as CL4)
    E. custom label 1 (a/b/c)
    F. budget
    G. result (TRUE/FALSE) - updated by script
    H. error message (when status is FALSE)

    Logic:
    1. Group rows by (maincat, custom_label_1)
    2. For each group, find/create campaign ONCE: PLA/{maincat} store_{cl1}
    3. For each shop in group, create ad group: PLA/{shop_name}_{cl1}
    4. Build listing tree for each ad group

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving)
        save_interval: Save progress every N groups
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING UITBREIDING SHEET: '{SHEET_UITBREIDING}'")
    print(f"(OPTIMIZED: Grouping shops by maincat + cl1)")
    print(f"{'='*70}")

    try:
        sheet = workbook[SHEET_UITBREIDING]
    except KeyError:
        print(f"❌ Sheet '{SHEET_UITBREIDING}' not found in workbook")
        return

    ga_service = client.get_service("GoogleAdsService")

    # =========================================================================
    # STEP 1: Group all rows by (maincat, cl1) for efficient batch processing
    # =========================================================================
    print("\nGrouping rows by (maincat, cl1)...")

    # Structure: {(maincat, cl1): [row_data_dict, ...]}
    groups = defaultdict(list)
    rows_with_missing_fields = []

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if already processed
        status_value = row[COL_UIT_STATUS].value if len(row) > COL_UIT_STATUS else None
        if status_value is not None and status_value != '':
            continue

        shop_name = row[COL_UIT_SHOP_NAME].value
        maincat = row[COL_UIT_MAINCAT].value
        maincat_id = row[COL_UIT_MAINCAT_ID].value
        custom_label_1 = row[COL_UIT_CUSTOM_LABEL_1].value
        budget = row[COL_UIT_BUDGET].value

        # Skip empty rows
        if not shop_name:
            continue

        # Track rows with missing required fields
        if not maincat or not maincat_id or not custom_label_1:
            rows_with_missing_fields.append(idx)
            continue

        group_key = (str(maincat), str(custom_label_1))
        groups[group_key].append({
            'row_idx': idx,
            'shop_name': shop_name,
            'maincat': maincat,
            'maincat_id': maincat_id,
            'custom_label_1': custom_label_1,
            'budget': budget
        })

    # Mark rows with missing fields as errors
    for idx in rows_with_missing_fields:
        print(f"[Row {idx}] ⚠️  Missing required fields, skipping")
        sheet.cell(row=idx, column=COL_UIT_STATUS + 1).value = False
        sheet.cell(row=idx, column=COL_UIT_ERROR + 1).value = "Missing required fields"

    total_groups = len(groups)
    total_rows = sum(len(rows) for rows in groups.values())
    print(f"Found {total_rows} row(s) in {total_groups} unique (maincat, cl1) group(s)")
    print(f"Rows with missing fields: {len(rows_with_missing_fields)}")

    if total_groups == 0:
        print("No rows to process.")
        return

    # =========================================================================
    # STEP 2: Process each group - find/create campaign ONCE per group
    # =========================================================================
    success_count = 0
    error_count = 0
    groups_processed = 0

    for (maincat, cl1), rows_in_group in groups.items():
        groups_processed += 1

        print(f"\n{'='*60}")
        print(f"[Group {groups_processed}/{total_groups}] {maincat} | {cl1}")
        print(f"  Shops in group: {len(rows_in_group)}")
        print(f"{'='*60}")

        # Get metadata from first row
        first_row = rows_in_group[0]
        maincat_id = first_row['maincat_id']
        budget = first_row['budget']

        # Build campaign name - ONCE for entire group
        campaign_name = f"PLA/{maincat} store_{cl1}"
        print(f"\n  Campaign: {campaign_name}")

        try:
            # Step 1: Find or create campaign ONCE for entire group
            escaped_campaign_name = campaign_name.replace("'", "\\'")
            campaign_query = f"""
                SELECT campaign.id, campaign.resource_name, campaign.status
                FROM campaign
                WHERE campaign.name = '{escaped_campaign_name}'
            """

            campaign_results = list(ga_service.search(customer_id=customer_id, query=campaign_query))
            campaign_resource_name = None

            for result in campaign_results:
                if result.campaign.status != client.enums.CampaignStatusEnum.REMOVED:
                    campaign_resource_name = result.campaign.resource_name
                    print(f"  ✅ Found existing campaign")
                    break

            if not campaign_resource_name:
                # Create new campaign
                print(f"  📦 Creating new campaign...")

                # Convert budget from EUR to micros
                try:
                    budget_micros = int(float(budget) * 1_000_000) if budget else 10_000_000
                except (ValueError, TypeError):
                    print(f"     ⚠️  Invalid budget '{budget}', using default 10 EUR")
                    budget_micros = 10_000_000

                # Get bid strategy based on custom label 1
                bid_strategy_resource_name = None
                if cl1 and cl1 in BID_STRATEGY_MAPPING:
                    bid_strategy_name = BID_STRATEGY_MAPPING[cl1]
                    print(f"     Looking up bid strategy: {bid_strategy_name}")
                    bid_strategy_resource_name = get_bid_strategy_by_name(
                        client=client,
                        customer_id=MCC_ACCOUNT_ID,
                        strategy_name=bid_strategy_name
                    )

                # Create campaign
                merchant_center_account_id = MERCHANT_CENTER_ID
                budget_name = f"Budget_{campaign_name}"
                first_shop = rows_in_group[0]['shop_name']

                campaign_resource_name = add_standard_shopping_campaign(
                    client=client,
                    customer_id=customer_id,
                    merchant_center_account_id=merchant_center_account_id,
                    campaign_name=campaign_name,
                    budget_name=budget_name,
                    tracking_template="",
                    country=COUNTRY,
                    shopid=None,
                    shopname=first_shop,
                    label=cl1,
                    budget=budget_micros,
                    bidding_strategy_resource_name=bid_strategy_resource_name
                )

                if not campaign_resource_name:
                    raise Exception("Failed to create campaign")

                print(f"  ✅ Campaign created")

                # Add negative keyword list to new campaign
                if NEGATIVE_LIST_NAME:
                    print(f"     Adding negative keyword list: {NEGATIVE_LIST_NAME}")
                    enable_negative_list_for_campaign(
                        client=client,
                        customer_id=customer_id,
                        campaign_resource_name=campaign_resource_name,
                        negative_list_name=NEGATIVE_LIST_NAME
                    )

            # Step 2: Process each shop in the group
            print(f"\n  Processing {len(rows_in_group)} shop(s)...")

            for shop_idx, row_data in enumerate(rows_in_group, start=1):
                idx = row_data['row_idx']
                shop_name = row_data['shop_name']
                shop_maincat_id = row_data['maincat_id']

                print(f"\n    [{shop_idx}/{len(rows_in_group)}] {shop_name}")

                try:
                    ad_group_name = f"PLA/{shop_name}_{cl1}"

                    # Look for existing ad group
                    escaped_ad_group_name = ad_group_name.replace("'", "\\'")
                    ad_group_query = f"""
                        SELECT ad_group.id, ad_group.resource_name
                        FROM ad_group
                        WHERE ad_group.campaign = '{campaign_resource_name}'
                        AND ad_group.name = '{escaped_ad_group_name}'
                        AND ad_group.status != 'REMOVED'
                    """

                    ad_group_results = list(ga_service.search(customer_id=customer_id, query=ad_group_query))
                    ad_group_resource_name = None

                    for result in ad_group_results:
                        ad_group_resource_name = result.ad_group.resource_name
                        print(f"      ✅ Found existing ad group")
                        break

                    if not ad_group_resource_name:
                        # Create new ad group
                        print(f"      📦 Creating ad group: {ad_group_name}")
                        ad_group_resource_name, _ = add_shopping_ad_group(
                            client=client,
                            customer_id=customer_id,
                            campaign_resource_name=campaign_resource_name,
                            ad_group_name=ad_group_name,
                            campaign_name=campaign_name
                        )

                        if not ad_group_resource_name:
                            raise Exception("Failed to create ad group")

                        print(f"      ✅ Ad group created")

                    # Build listing tree
                    ad_group_id = ad_group_resource_name.split('/')[-1]

                    build_listing_tree_for_uitbreiding(
                        client=client,
                        customer_id=customer_id,
                        ad_group_id=ad_group_id,
                        shop_name=shop_name,
                        maincat_id=str(shop_maincat_id),
                        custom_label_1=str(cl1)
                    )

                    # Create shopping product ad
                    add_shopping_product_ad(
                        client=client,
                        customer_id=customer_id,
                        ad_group_resource_name=ad_group_resource_name
                    )

                    # Mark success
                    sheet.cell(row=idx, column=COL_UIT_STATUS + 1).value = True
                    sheet.cell(row=idx, column=COL_UIT_ERROR + 1).value = ""
                    success_count += 1
                    print(f"      ✅ Row {idx} completed")

                    # Rate limiting between shops
                    time.sleep(1.5)

                except Exception as shop_e:
                    error_msg = str(shop_e)
                    print(f"      ❌ Error: {error_msg[:60]}")

                    # Categorize errors
                    if "CONCURRENT_MODIFICATION" in error_msg:
                        friendly_error = "Concurrent modification (retry needed)"
                    elif "NOT_FOUND" in error_msg.upper():
                        friendly_error = "Resource not found"
                    elif "SUBDIVISION_REQUIRES_OTHERS_CASE" in error_msg:
                        friendly_error = "Tree structure error"
                    else:
                        friendly_error = error_msg[:80]

                    sheet.cell(row=idx, column=COL_UIT_STATUS + 1).value = False
                    sheet.cell(row=idx, column=COL_UIT_ERROR + 1).value = friendly_error
                    error_count += 1

        except Exception as group_e:
            # Campaign-level error - mark all rows in group as failed
            error_msg = str(group_e)
            print(f"\n  ❌ GROUP FAILED: {error_msg[:80]}")

            for row_data in rows_in_group:
                idx = row_data['row_idx']
                sheet.cell(row=idx, column=COL_UIT_STATUS + 1).value = False
                sheet.cell(row=idx, column=COL_UIT_ERROR + 1).value = f"Campaign error: {error_msg[:60]}"
                error_count += 1

        # Save periodically (every N groups)
        if file_path and groups_processed % save_interval == 0:
            print(f"\n💾 Saving progress ({groups_processed} groups processed)...")
            try:
                workbook.save(file_path)
            except Exception as save_error:
                print(f"⚠️  Error saving: {save_error}")

        # Delay between groups to avoid concurrent modification
        if groups_processed < total_groups:
            time.sleep(2)

    # Final save
    if file_path:
        print(f"\n💾 Final save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"⚠️  Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"UITBREIDING SHEET SUMMARY (OPTIMIZED)")
    print(f"{'='*70}")
    print(f"Total groups processed: {groups_processed}")
    print(f"Total rows processed: {success_count + error_count}")
    print(f"Rows with missing fields: {len(rows_with_missing_fields)}")
    print(f"✅ Successful: {success_count}")
    print(f"❌ Failed: {error_count + len(rows_with_missing_fields)}")
    print(f"{'='*70}\n")


def load_cat_ids_mapping(workbook: openpyxl.Workbook) -> dict:
    """
    Load the cat_ids sheet and create a mapping of maincat_id -> list of deepest_cat values.

    Args:
        workbook: Excel workbook containing cat_ids sheet

    Returns:
        dict: {maincat_id: [deepest_cat1, deepest_cat2, ...]}
    """
    try:
        sheet = workbook[SHEET_CAT_IDS]
    except KeyError:
        print(f"❌ Sheet '{SHEET_CAT_IDS}' not found in workbook")
        return {}

    mapping = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        maincat_id = row[COL_CAT_MAINCAT_ID]
        deepest_cat = row[COL_CAT_DEEPEST_CAT]

        if maincat_id and deepest_cat:
            maincat_id_str = str(maincat_id)
            if maincat_id_str not in mapping:
                mapping[maincat_id_str] = set()
            mapping[maincat_id_str].add(str(deepest_cat))

    # Convert sets to sorted lists
    for key in mapping:
        mapping[key] = sorted(mapping[key])

    print(f"   Loaded {len(mapping)} maincat_id mappings from '{SHEET_CAT_IDS}' sheet")
    return mapping


def add_shop_exclusion_to_ad_group(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    shop_name: str
):
    """
    Add a shop name as CL3 exclusion to an ad group's listing tree.
    Preserves existing tree structure and adds the shop as a negative CL3 unit.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_name: Shop name to exclude (CL3 value)
    """
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    agc_service = client.get_service("AdGroupCriterionService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    # Step 1: Read existing tree structure
    query = f"""
        SELECT
            ad_group_criterion.resource_name,
            ad_group_criterion.listing_group.type,
            ad_group_criterion.listing_group.parent_ad_group_criterion,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
            ad_group_criterion.negative,
            ad_group_criterion.cpc_bid_micros
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
    """

    results = list(ga_service.search(customer_id=customer_id, query=query))

    if not results:
        print(f"      ⚠️  No listing tree found in ad group {ad_group_id}")
        return False

    # Find parent for CL3 nodes (by looking at existing CL3 nodes including CL3 OTHERS)
    parent_for_cl3 = None
    existing_cl3_exclusions = set()

    for row in results:
        criterion = row.ad_group_criterion
        lg = criterion.listing_group

        # Check for CL3 nodes (INDEX3) - get parent from any CL3 node
        if lg.case_value.product_custom_attribute.index.name == 'INDEX3':
            value = lg.case_value.product_custom_attribute.value

            if value:
                # This is a specific CL3 value (shop name)
                if criterion.negative:
                    existing_cl3_exclusions.add(value.lower())

            # Get the parent of any CL3 node - this is where we add new CL3 exclusions
            if lg.parent_ad_group_criterion:
                parent_for_cl3 = lg.parent_ad_group_criterion

    if not parent_for_cl3:
        # No CL3 level exists — CL1 is likely a UNIT instead of SUBDIVISION.
        # Rebuild: remove CL1 UNIT, create CL1 SUBDIVISION, add CL3 catch-all + exclusion.
        print(f"      ℹ️  No CL3 level found, attempting to rebuild CL1 as SUBDIVISION...")

        # Find the CL1 UNIT node and its parent (CL0 SUBDIVISION)
        cl1_unit_resource = None
        cl1_parent_resource = None
        cl1_value = None
        cl1_bid_micros = None
        cl1_negative = False

        for row in results:
            criterion = row.ad_group_criterion
            lg = criterion.listing_group
            if lg.case_value.product_custom_attribute.index.name == 'INDEX1' and lg.type_.name == 'UNIT':
                cl1_unit_resource = criterion.resource_name
                cl1_parent_resource = lg.parent_ad_group_criterion
                cl1_value = lg.case_value.product_custom_attribute.value
                cl1_bid_micros = criterion.cpc_bid_micros if not criterion.negative else None
                cl1_negative = criterion.negative
                break

        if not cl1_unit_resource or not cl1_parent_resource or not cl1_value:
            print(f"      ⚠️  No parent for CL3 found and could not find CL1 UNIT to rebuild in ad group {ad_group_id}")
            return False

        print(f"      Found CL1 UNIT: value='{cl1_value}', rebuilding as SUBDIVISION with CL3 level...")

        try:
            operations = []

            # 1. Remove the existing CL1 UNIT
            remove_op = client.get_type("AdGroupCriterionOperation")
            remove_op.remove = cl1_unit_resource
            operations.append(remove_op)

            # 2. Create CL1 SUBDIVISION (same value, same parent)
            dim_cl1 = client.get_type("ListingDimensionInfo")
            dim_cl1.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
            dim_cl1.product_custom_attribute.value = cl1_value

            cl1_subdiv_op = create_listing_group_subdivision(
                client=client,
                customer_id=customer_id,
                ad_group_id=ad_group_id,
                parent_ad_group_criterion_resource_name=cl1_parent_resource,
                listing_dimension_info=dim_cl1,
            )
            cl1_subdiv_resource = cl1_subdiv_op.create.resource_name
            operations.append(cl1_subdiv_op)

            # 3. Create CL3 catch-all UNIT (biddable, everything else) under the new CL1 SUBDIVISION
            dim_cl3_others = client.get_type("ListingDimensionInfo")
            dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3

            cl3_catchall_op = create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=ad_group_id,
                parent_ad_group_criterion_resource_name=cl1_subdiv_resource,
                listing_dimension_info=dim_cl3_others,
                targeting_negative=False,
                cpc_bid_micros=cl1_bid_micros if not cl1_negative else None,
            )
            operations.append(cl3_catchall_op)

            # 4. Add the shop exclusion as CL3 negative unit
            dim_cl3_shop = client.get_type("ListingDimensionInfo")
            dim_cl3_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
            dim_cl3_shop.product_custom_attribute.value = shop_name

            cl3_excl_op = create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=ad_group_id,
                parent_ad_group_criterion_resource_name=cl1_subdiv_resource,
                listing_dimension_info=dim_cl3_shop,
                targeting_negative=True,
                cpc_bid_micros=None,
            )
            operations.append(cl3_excl_op)

            # Execute all operations atomically
            agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=operations)
            print(f"      ✅ Rebuilt CL1 as SUBDIVISION and added exclusion: CL3='{shop_name}'")
            return True

        except Exception as e:
            print(f"      ❌ Error rebuilding tree: {str(e)[:150]}")
            return False

    # Check if shop is already excluded
    if shop_name.lower() in existing_cl3_exclusions:
        print(f"      ℹ️  Shop '{shop_name}' already excluded")
        return True

    # Step 2: Add the shop exclusion as a new CL3 negative unit
    dim_cl3_shop = client.get_type("ListingDimensionInfo")
    dim_cl3_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
    dim_cl3_shop.product_custom_attribute.value = shop_name

    op = create_listing_group_unit_biddable(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=parent_for_cl3,
        listing_dimension_info=dim_cl3_shop,
        targeting_negative=True,
        cpc_bid_micros=None
    )

    try:
        agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=[op])
        print(f"      ✅ Added exclusion: CL3='{shop_name}'")
        return True
    except Exception as e:
        error_msg = str(e)
        if "LISTING_GROUP_ALREADY_EXISTS" in error_msg:
            print(f"      ℹ️  Shop '{shop_name}' already excluded (duplicate)")
            return True
        else:
            print(f"      ❌ Error adding exclusion: {error_msg[:100]}")
            return False


def reverse_exclusion(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    ad_group_name: str,
    shop_name: str
):
    """
    Remove a shop exclusion from an ad group's listing tree.
    This is the reverse of add_shop_exclusion_to_ad_group - it finds and removes
    the CL3 negative criterion for the specified shop.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_name: Shop name to un-exclude (remove CL3 exclusion)

    Returns:
        bool: True if exclusion was removed or didn't exist, False on error
    """
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    agc_service = client.get_service("AdGroupCriterionService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    # Step 1: Read existing tree structure to find the CL3 exclusion
    query = f"""
        SELECT
            ad_group_criterion.resource_name,
            ad_group_criterion.listing_group.type,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
            ad_group_criterion.negative
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
    """

    try:
        results = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        print(f"      ❌ Error reading tree: {e}")
        return False

    if not results:
        print(f"      ⚠️  No listing tree found in ad group {ad_group_name} {ad_group_id}")
        return False

    # Find the CL3 exclusion criterion for this shop
    criterion_to_remove = None
    shop_name_lower = shop_name.lower()

    for row in results:
        criterion = row.ad_group_criterion
        lg = criterion.listing_group

        # Check for CL3 nodes (INDEX3)
        if lg.case_value.product_custom_attribute.index.name == 'INDEX3':
            value = lg.case_value.product_custom_attribute.value

            # Check if this is the shop we want to un-exclude
            if value and value.lower() == shop_name_lower and criterion.negative:
                criterion_to_remove = criterion.resource_name
                break

    if not criterion_to_remove:
        print(f"      ℹ️  Shop '{shop_name}' is not excluded (no exclusion found)")
        return True  # Not an error - shop wasn't excluded

    # Step 2: Remove the exclusion criterion
    op = client.get_type("AdGroupCriterionOperation")
    op.remove = criterion_to_remove

    try:
        agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=[op])
        #print(f"      ✅ Removed exclusion: CL3='{shop_name}' from {ad_group_name}")
        return True
    except Exception as e:
        error_msg = str(e)
        print(f"      ❌ Error removing exclusion: {error_msg[:100]}")
        return False


def reverse_exclusion_batch(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    ad_group_name: str,
    shop_names: list
) -> dict:
    """
    Remove multiple shop exclusions from an ad group's listing tree in one batch.

    This is an optimized version that reads the listing tree ONCE and removes
    all matching shop exclusions in a single mutate operation.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        ad_group_name: Ad group name (for logging)
        shop_names: List of shop names to un-exclude

    Returns:
        dict: {
            'success': list of shop names that were successfully removed,
            'not_found': list of shop names that weren't excluded,
            'errors': list of (shop_name, error_msg) tuples
        }
    """
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    agc_service = client.get_service("AdGroupCriterionService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    result = {
        'success': [],
        'not_found': [],
        'errors': []
    }

    # Normalize shop names for case-insensitive matching
    shop_names_lower = {name.lower(): name for name in shop_names}

    # Step 1: Read existing tree structure ONCE
    query = f"""
        SELECT
            ad_group_criterion.resource_name,
            ad_group_criterion.listing_group.type,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
            ad_group_criterion.negative
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
    """

    try:
        results = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        # All shops failed due to read error
        for shop_name in shop_names:
            result['errors'].append((shop_name, f"Error reading tree: {str(e)[:50]}"))
        return result

    if not results:
        # No listing tree - all shops count as "not found"
        result['not_found'] = list(shop_names)
        return result

    # Step 2: Find ALL CL3 exclusion criteria matching any of our shops
    criteria_to_remove = []  # List of (resource_name, shop_name)
    found_shops = set()

    for row in results:
        criterion = row.ad_group_criterion
        lg = criterion.listing_group

        # Check for CL3 nodes (INDEX3)
        if lg.case_value.product_custom_attribute.index.name == 'INDEX3':
            value = lg.case_value.product_custom_attribute.value

            if value and criterion.negative:
                value_lower = value.lower()
                if value_lower in shop_names_lower:
                    original_shop_name = shop_names_lower[value_lower]
                    criteria_to_remove.append((criterion.resource_name, original_shop_name))
                    found_shops.add(value_lower)

    # Track shops that weren't excluded
    for shop_lower, shop_name in shop_names_lower.items():
        if shop_lower not in found_shops:
            result['not_found'].append(shop_name)

    if not criteria_to_remove:
        # No exclusions to remove
        return result

    # Step 3: Remove all exclusion criteria in ONE batch operation
    operations = []
    for resource_name, shop_name in criteria_to_remove:
        op = client.get_type("AdGroupCriterionOperation")
        op.remove = resource_name
        operations.append((op, shop_name))

    try:
        # Execute batch removal
        response = agc_service.mutate_ad_group_criteria(
            customer_id=customer_id,
            operations=[op for op, _ in operations]
        )
        # If the API returned fewer results than ops (shouldn't happen with the
        # default partial_failure=False, but defensive), surface that.
        if len(response.results) != len(operations):
            print(f"      ⚠️  reverse_exclusion_batch: ad_group {ad_group_name} — "
                  f"API returned {len(response.results)} results for {len(operations)} ops!")
        for _, shop_name in operations:
            result['success'].append(shop_name)
    except GoogleAdsException as gae:
        # Walk failure.errors so each individual error is surfaced — the previous
        # str(gae)[:100] truncated the most useful diagnostic part.
        details = []
        try:
            for err in gae.failure.errors:
                code = err.error_code.WhichOneof("error_code") if err.error_code else "?"
                details.append(f"[{code}] {err.message}")
        except Exception:
            details.append(str(gae)[:300])
        full_msg = " | ".join(details) if details else str(gae)[:300]
        print(f"      ❌ reverse_exclusion_batch GoogleAdsException for ad_group {ad_group_name}: {full_msg}")
        for _, shop_name in operations:
            result['errors'].append((shop_name, full_msg[:200]))
    except Exception as e:
        msg = f"{type(e).__name__}: {str(e)[:300]}"
        print(f"      ❌ reverse_exclusion_batch unexpected exception for ad_group {ad_group_name}: {msg}")
        for _, shop_name in operations:
            result['errors'].append((shop_name, msg[:200]))

    return result


def prepare_shop_exclusion_operation(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    shop_name: str,
    listing_group_cache: dict = None
) -> tuple:
    """
    Prepare a shop exclusion operation without executing it.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        shop_name: Shop name to exclude (CL3 value)
        listing_group_cache: Optional pre-fetched listing group data

    Returns:
        tuple: (operation, status, message)
        - operation: The mutation operation (or None if skip/error)
        - status: 'ready', 'skip', or 'error'
        - message: Description of result
    """
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    # Use cache if provided, otherwise query
    if listing_group_cache and ad_group_id in listing_group_cache:
        cache_entry = listing_group_cache[ad_group_id]
        parent_for_cl3 = cache_entry.get('parent_for_cl3')
        existing_cl3_exclusions = cache_entry.get('cl3_exclusions', set())
    else:
        # Query listing group structure
        query = f"""
            SELECT
                ad_group_criterion.resource_name,
                ad_group_criterion.listing_group.type,
                ad_group_criterion.listing_group.parent_ad_group_criterion,
                ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
                ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
                ad_group_criterion.negative
            FROM ad_group_criterion
            WHERE ad_group_criterion.ad_group = '{ag_path}'
                AND ad_group_criterion.type = 'LISTING_GROUP'
        """

        try:
            results = list(ga_service.search(customer_id=customer_id, query=query))
        except Exception as e:
            return (None, 'error', f"Query error: {str(e)[:50]}")

        if not results:
            return (None, 'error', "No listing tree found")

        parent_for_cl3 = None
        existing_cl3_exclusions = set()

        for row in results:
            criterion = row.ad_group_criterion
            lg = criterion.listing_group

            # Safely get index name
            try:
                index_name = lg.case_value.product_custom_attribute.index.name
            except (AttributeError, TypeError):
                index_name = None

            # Check for CL3 nodes (INDEX3) - get parent from any CL3 node
            if index_name == 'INDEX3':
                value_str = lg.case_value.product_custom_attribute.value
                if value_str and criterion.negative:
                    existing_cl3_exclusions.add(value_str.lower())
                if lg.parent_ad_group_criterion:
                    parent_for_cl3 = lg.parent_ad_group_criterion

    if not parent_for_cl3:
        return (None, 'error', "No parent for CL3 found")

    # Check if already excluded
    if shop_name.lower() in existing_cl3_exclusions:
        return (None, 'skip', f"Already excluded")

    # Create the operation
    dim_cl3_shop = client.get_type("ListingDimensionInfo")
    dim_cl3_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
    dim_cl3_shop.product_custom_attribute.value = shop_name

    op = create_listing_group_unit_biddable(
        client=client,
        customer_id=customer_id,
        ad_group_id=ad_group_id,
        parent_ad_group_criterion_resource_name=parent_for_cl3,
        listing_dimension_info=dim_cl3_shop,
        targeting_negative=True,
        cpc_bid_micros=None
    )

    return (op, 'ready', "Operation prepared")


def execute_exclusion_batch(
    client: GoogleAdsClient,
    customer_id: str,
    operations: list,
    batch_size: int = 10
) -> tuple:
    """
    Execute a batch of exclusion operations.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        operations: List of (operation, ad_group_name, shop_name) tuples
        batch_size: Max operations per API call

    Returns:
        tuple: (success_count, error_count, errors_list)
    """
    if not operations:
        return (0, 0, [])

    agc_service = client.get_service("AdGroupCriterionService")
    success_count = 0
    error_count = 0
    errors = []

    # Process in batches
    for i in range(0, len(operations), batch_size):
        batch = operations[i:i + batch_size]
        ops = [item[0] for item in batch]

        try:
            agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops)
            success_count += len(batch)
            for _, ag_name, shop_name in batch:
                print(f"      ✅ {ag_name}: excluded '{shop_name}'")
        except Exception as e:
            error_msg = str(e)
            if "LISTING_GROUP_ALREADY_EXISTS" in error_msg:
                # Some might already exist, count as success
                success_count += len(batch)
                print(f"      ℹ️  Batch contained duplicates (already excluded)")
            else:
                # Batch failed - try individual operations
                print(f"      ⚠️  Batch failed, retrying individually...")
                for op, ag_name, shop_name in batch:
                    try:
                        agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=[op])
                        success_count += 1
                        print(f"      ✅ {ag_name}: excluded '{shop_name}'")
                    except Exception as ind_e:
                        ind_error = str(ind_e)
                        if "LISTING_GROUP_ALREADY_EXISTS" in ind_error:
                            success_count += 1
                            print(f"      ℹ️  {ag_name}: already excluded")
                        else:
                            error_count += 1
                            errors.append(f"{ag_name}: {ind_error[:50]}")
                            print(f"      ❌ {ag_name}: {ind_error[:50]}")
                    time.sleep(0.1)

        # Small delay between batches
        if i + batch_size < len(operations):
            time.sleep(0.2)

    return (success_count, error_count, errors)


def add_shop_exclusions_batch(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    ad_group_name: str,
    shop_names: list
) -> dict:
    """
    Add multiple shop exclusions to an ad group's listing tree in one batch.

    This is an optimized version that reads the listing tree ONCE and adds
    all exclusions in a single mutate operation.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        ad_group_name: Ad group name (for logging)
        shop_names: List of shop names to exclude

    Returns:
        dict: {
            'success': list of shop names that were successfully excluded,
            'already_excluded': list of shop names that were already excluded,
            'errors': list of (shop_name, error_msg) tuples
        }
    """
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    agc_service = client.get_service("AdGroupCriterionService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    result = {
        'success': [],
        'already_excluded': [],
        'errors': []
    }

    # Normalize shop names for case-insensitive matching
    shop_names_lower = {name.lower(): name for name in shop_names}

    # Step 1: Read existing tree structure ONCE
    query = f"""
        SELECT
            ad_group_criterion.resource_name,
            ad_group_criterion.listing_group.type,
            ad_group_criterion.listing_group.parent_ad_group_criterion,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
            ad_group_criterion.negative,
            ad_group_criterion.cpc_bid_micros
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
    """

    try:
        results = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        # All shops failed due to read error
        for shop_name in shop_names:
            result['errors'].append((shop_name, f"Error reading tree: {str(e)[:50]}"))
        return result

    if not results:
        for shop_name in shop_names:
            result['errors'].append((shop_name, "No listing tree found"))
        return result

    # Step 2: Find parent for CL3 and existing exclusions
    parent_for_cl3 = None
    existing_cl3_exclusions = set()

    for row in results:
        criterion = row.ad_group_criterion
        lg = criterion.listing_group

        try:
            index_name = lg.case_value.product_custom_attribute.index.name
        except (AttributeError, TypeError):
            index_name = None

        if index_name == 'INDEX3':
            value_str = lg.case_value.product_custom_attribute.value
            if value_str and criterion.negative:
                existing_cl3_exclusions.add(value_str.lower())
            if lg.parent_ad_group_criterion:
                parent_for_cl3 = lg.parent_ad_group_criterion

    if not parent_for_cl3:
        # No CL3 level exists — find the deepest positive UNIT (not INDEX3) and
        # convert it to a SUBDIVISION, then add CL3 children for exclusions.
        # This works regardless of which custom label is the leaf (CL0, CL1, CL2, CL4).
        print(f"      ℹ️  No CL3 level found, searching for deepest positive UNIT to subdivide...")

        leaf_unit = None  # Will hold: {resource, parent, value, bid, index_name, index_enum}

        for row in results:
            criterion = row.ad_group_criterion
            lg = criterion.listing_group
            try:
                index_name = lg.case_value.product_custom_attribute.index.name
                index_enum = lg.case_value.product_custom_attribute.index
            except (AttributeError, TypeError):
                index_name = None
                index_enum = None
            if (index_name and index_name != 'INDEX3'
                    and lg.type_.name == 'UNIT'
                    and not criterion.negative
                    and lg.case_value.product_custom_attribute.value):
                leaf_unit = {
                    'resource': criterion.resource_name,
                    'parent': lg.parent_ad_group_criterion,
                    'value': lg.case_value.product_custom_attribute.value,
                    'bid': criterion.cpc_bid_micros if criterion.cpc_bid_micros > 0 else DEFAULT_BID_MICROS,
                    'index_name': index_name,
                    'index_enum': index_enum,
                }
                break

        if not leaf_unit or not leaf_unit['parent']:
            for shop_name in shop_names:
                result['errors'].append((shop_name, "No parent for CL3 found"))
            return result

        print(f"      Found {leaf_unit['index_name']} UNIT: value='{leaf_unit['value']}', rebuilding as SUBDIVISION with CL3 level...")

        try:
            rebuild_ops = []

            # 1. Remove the existing UNIT
            remove_op = client.get_type("AdGroupCriterionOperation")
            remove_op.remove = leaf_unit['resource']
            rebuild_ops.append(remove_op)

            # 2. Create SUBDIVISION with the same index, value, and parent
            dim_leaf = client.get_type("ListingDimensionInfo")
            dim_leaf.product_custom_attribute.index = leaf_unit['index_enum']
            dim_leaf.product_custom_attribute.value = leaf_unit['value']

            leaf_subdiv_op = create_listing_group_subdivision(
                client=client,
                customer_id=customer_id,
                ad_group_id=ad_group_id,
                parent_ad_group_criterion_resource_name=leaf_unit['parent'],
                listing_dimension_info=dim_leaf,
            )
            leaf_subdiv_resource = leaf_subdiv_op.create.resource_name
            rebuild_ops.append(leaf_subdiv_op)

            # 3. Create CL3 catch-all UNIT (biddable, everything else) under the new SUBDIVISION
            dim_cl3_others = client.get_type("ListingDimensionInfo")
            dim_cl3_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3

            cl3_catchall_op = create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=ad_group_id,
                parent_ad_group_criterion_resource_name=leaf_subdiv_resource,
                listing_dimension_info=dim_cl3_others,
                targeting_negative=False,
                cpc_bid_micros=leaf_unit['bid'],
            )
            rebuild_ops.append(cl3_catchall_op)

            # 4. Add all shop exclusions as CL3 negative units
            for shop_name in shop_names:
                dim_cl3_shop = client.get_type("ListingDimensionInfo")
                dim_cl3_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
                dim_cl3_shop.product_custom_attribute.value = shop_name

                cl3_excl_op = create_listing_group_unit_biddable(
                    client=client,
                    customer_id=customer_id,
                    ad_group_id=ad_group_id,
                    parent_ad_group_criterion_resource_name=leaf_subdiv_resource,
                    listing_dimension_info=dim_cl3_shop,
                    targeting_negative=True,
                    cpc_bid_micros=None,
                )
                rebuild_ops.append(cl3_excl_op)

            # Execute all operations atomically
            agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=rebuild_ops)
            print(f"      ✅ Rebuilt {leaf_unit['index_name']} as SUBDIVISION and added {len(shop_names)} exclusion(s)")
            result['success'] = list(shop_names)
            return result

        except GoogleAdsException as gae:
            print(f"      ❌ Error rebuilding tree (GoogleAdsException):")
            for error in gae.failure.errors:
                print(f"         {error.error_code}: {error.message}")
                if error.location:
                    print(f"         Location: {error.location}")
            first_err = gae.failure.errors[0].message[:80] if gae.failure.errors else "unknown error"
            for shop_name in shop_names:
                result['errors'].append((shop_name, f"Rebuild failed: {first_err}"))
            return result
        except Exception as e:
            print(f"      ❌ Error rebuilding tree: {str(e)[:150]}")
            for shop_name in shop_names:
                result['errors'].append((shop_name, f"Rebuild failed: {str(e)[:50]}"))
            return result

    # Step 3: Determine which shops to add vs skip
    operations = []
    for shop_lower, shop_name in shop_names_lower.items():
        if shop_lower in existing_cl3_exclusions:
            result['already_excluded'].append(shop_name)
        else:
            # Create operation for this shop
            dim_cl3_shop = client.get_type("ListingDimensionInfo")
            dim_cl3_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
            dim_cl3_shop.product_custom_attribute.value = shop_name

            op = create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=ad_group_id,
                parent_ad_group_criterion_resource_name=parent_for_cl3,
                listing_dimension_info=dim_cl3_shop,
                targeting_negative=True,
                cpc_bid_micros=None
            )
            operations.append((op, shop_name))

    if not operations:
        # All shops were already excluded
        return result

    # Step 4: Execute batch
    try:
        ops = [op for op, _ in operations]
        agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops)
        for _, shop_name in operations:
            result['success'].append(shop_name)
    except GoogleAdsException as gae:
        error_str = str(gae)
        if "LISTING_GROUP_ALREADY_EXISTS" in error_str:
            # Some were duplicates - try individually
            for op, shop_name in operations:
                try:
                    agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=[op])
                    result['success'].append(shop_name)
                except Exception as ind_e:
                    ind_error = str(ind_e)
                    if "LISTING_GROUP_ALREADY_EXISTS" in ind_error:
                        result['already_excluded'].append(shop_name)
                    else:
                        result['errors'].append((shop_name, ind_error[:50]))
        else:
            # Batch failed - try individually
            for op, shop_name in operations:
                try:
                    agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=[op])
                    result['success'].append(shop_name)
                except Exception as ind_e:
                    ind_error = str(ind_e)
                    if "LISTING_GROUP_ALREADY_EXISTS" in ind_error:
                        result['already_excluded'].append(shop_name)
                    else:
                        result['errors'].append((shop_name, ind_error[:50]))
    except Exception as e:
        error_msg = str(e)[:300]
        for _, shop_name in operations:
            result['errors'].append((shop_name, error_msg))

    return result


def replace_shop_exclusions_batch(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    ad_group_name: str,
    replacements: dict
) -> dict:
    """
    Replace shop exclusions in an ad group's listing tree (REMOVE old + CREATE new).

    For each replacement, removes the old pipe-version exclusion and creates
    a new clean lowercase exclusion in a single batch mutate call.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        ad_group_name: Ad group name (for logging)
        replacements: Dict of {old_name: new_name} to replace
                      e.g. {"Artandcraft.com|NL": "artandcraft.com"}

    Returns:
        dict: {
            'success': list of (old_name, new_name) tuples that were replaced,
            'not_found': list of old_names that weren't found as exclusions,
            'already_clean': list of old_names where clean version already existed (old removed),
            'errors': list of (old_name, error_msg) tuples
        }
    """
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    agc_service = client.get_service("AdGroupCriterionService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    result = {
        'success': [],
        'not_found': [],
        'already_clean': [],
        'errors': []
    }

    # Normalize old names for case-insensitive matching
    old_names_lower = {old.lower(): (old, new) for old, new in replacements.items()}
    # Collect all new (clean) names for checking if they already exist
    new_names_lower = {new.lower() for new in replacements.values()}

    # Step 1: Read existing tree structure ONCE
    query = f"""
        SELECT
            ad_group_criterion.resource_name,
            ad_group_criterion.listing_group.type,
            ad_group_criterion.listing_group.parent_ad_group_criterion,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
            ad_group_criterion.negative
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
    """

    try:
        results = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        for old_name in replacements:
            result['errors'].append((old_name, f"Error reading tree: {str(e)[:50]}"))
        return result

    if not results:
        result['not_found'] = list(replacements.keys())
        return result

    # Step 2: Analyze tree - find parent, old exclusions, and existing clean versions
    parent_for_cl3 = None
    old_criteria = {}  # {value_lower: (resource_name, original_old_name)}
    existing_clean = set()  # lowercase values of existing negative CL3 nodes

    for row in results:
        criterion = row.ad_group_criterion
        lg = criterion.listing_group

        try:
            index_name = lg.case_value.product_custom_attribute.index.name
        except (AttributeError, TypeError):
            index_name = None

        if index_name == 'INDEX3':
            value = lg.case_value.product_custom_attribute.value
            if lg.parent_ad_group_criterion:
                parent_for_cl3 = lg.parent_ad_group_criterion

            if value and criterion.negative:
                value_lower = value.lower()
                # Check if this is an old (pipe) exclusion we want to replace
                if value_lower in old_names_lower:
                    old_criteria[value_lower] = (criterion.resource_name, old_names_lower[value_lower][0])
                # Track all existing clean exclusions
                if value_lower in new_names_lower:
                    existing_clean.add(value_lower)

    # Step 3: Build operations
    operations = []  # List of (op, old_name, action_type)
    # action_type: 'replace' or 'remove_only' (when clean already exists)

    for old_lower, (old_name, new_name) in old_names_lower.items():
        if old_lower not in old_criteria:
            result['not_found'].append(old_name)
            continue

        resource_name = old_criteria[old_lower][0]
        new_lower = new_name.lower()

        # REMOVE operation for the old pipe-version
        remove_op = client.get_type("AdGroupCriterionOperation")
        remove_op.remove = resource_name
        operations.append((remove_op, old_name, 'remove'))

        if new_lower in existing_clean:
            # Clean version already exists - just remove the old one
            operations.append((None, old_name, 'already_clean_marker'))
        else:
            if not parent_for_cl3:
                result['errors'].append((old_name, "No parent for CL3 found"))
                # Remove the REMOVE op we just added since we can't complete the replacement
                operations.pop()
                continue

            # CREATE operation for the new clean version
            dim_cl3_shop = client.get_type("ListingDimensionInfo")
            dim_cl3_shop.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX3
            dim_cl3_shop.product_custom_attribute.value = new_name

            create_op = create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=ad_group_id,
                parent_ad_group_criterion_resource_name=parent_for_cl3,
                listing_dimension_info=dim_cl3_shop,
                targeting_negative=True,
                cpc_bid_micros=None
            )
            operations.append((create_op, old_name, 'create'))
            # Mark this clean name as "will exist" to avoid duplicate creates
            existing_clean.add(new_lower)

    if not operations:
        return result

    # Step 4: Execute batch - filter out marker entries
    real_ops = [(op, old_name, action) for op, old_name, action in operations if op is not None]
    marker_names = {old_name for op, old_name, action in operations if action == 'already_clean_marker'}

    if real_ops:
        try:
            agc_service.mutate_ad_group_criteria(
                customer_id=customer_id,
                operations=[op for op, _, _ in real_ops]
            )
            # Batch succeeded - categorize results
            replaced_names = set()
            for op, old_name, action in real_ops:
                if action == 'create':
                    new_name = replacements[old_name]
                    result['success'].append((old_name, new_name))
                    replaced_names.add(old_name)

            # Names that had only REMOVE (clean already existed)
            for old_name in marker_names:
                if old_name not in replaced_names:
                    new_name = replacements[old_name]
                    result['already_clean'].append((old_name, new_name))

        except GoogleAdsException as gae:
            # Batch failed - try per-replacement (REMOVE+CREATE pair)
            error_str = str(gae)[:300]
            print(f"      Batch failed ({error_str}), falling back to individual replacements...")
            _replace_individual_fallback(
                client, customer_id, agc_service, operations, replacements, marker_names, result
            )
        except Exception as e:
            error_msg = str(e)[:300]
            for old_name in replacements:
                if old_name not in [n for n in result['not_found']]:
                    result['errors'].append((old_name, error_msg))
    else:
        # Only markers (all clean versions already existed, just needed remove)
        for old_name in marker_names:
            new_name = replacements[old_name]
            result['already_clean'].append((old_name, new_name))

    return result


def _replace_individual_fallback(client, customer_id, agc_service, operations, replacements, marker_names, result):
    """Fallback: execute replacement operations individually per shop."""
    # Group operations by old_name
    ops_by_name = {}
    for op, old_name, action in operations:
        if old_name not in ops_by_name:
            ops_by_name[old_name] = []
        if op is not None:
            ops_by_name[old_name].append(op)

    for old_name, ops in ops_by_name.items():
        if not ops:
            continue
        try:
            agc_service.mutate_ad_group_criteria(
                customer_id=customer_id,
                operations=ops
            )
            new_name = replacements[old_name]
            if old_name in marker_names:
                result['already_clean'].append((old_name, new_name))
            else:
                result['success'].append((old_name, new_name))
        except Exception as ind_e:
            result['errors'].append((old_name, str(ind_e)[:50]))


def prefetch_pla_campaigns_and_ad_groups(
    client: GoogleAdsClient,
    customer_id: str,
    campaign_prefix: str = "PLA/"
) -> dict:
    """
    Pre-fetch all PLA campaigns and their ad groups in a single query.

    Returns a dict structured as:
    {
        'campaign_name': {
            'resource_name': 'customers/xxx/campaigns/yyy',
            'ad_groups': [
                {'id': 123, 'name': 'ag_name', 'resource_name': 'customers/xxx/adGroups/zzz'},
                ...
            ]
        },
        ...
    }
    """
    print(f"\n📥 Pre-fetching campaigns and ad groups (prefix: {campaign_prefix})...")

    ga_service = client.get_service("GoogleAdsService")

    escaped_prefix = campaign_prefix.replace("'", "\\'")
    query = f"""
        SELECT
            campaign.id,
            campaign.name,
            campaign.resource_name,
            ad_group.id,
            ad_group.name,
            ad_group.resource_name
        FROM ad_group
        WHERE campaign.name LIKE '{escaped_prefix}%'
        AND campaign.status != 'REMOVED'
        AND ad_group.status != 'REMOVED'
        ORDER BY campaign.name, ad_group.name
    """

    try:
        results = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        print(f"❌ Error pre-fetching: {e}")
        return {}

    # Build the cache structure
    cache = {}
    for row in results:
        campaign_name = row.campaign.name
        if campaign_name not in cache:
            cache[campaign_name] = {
                'resource_name': row.campaign.resource_name,
                'ad_groups': []
            }
        cache[campaign_name]['ad_groups'].append({
            'id': row.ad_group.id,
            'name': row.ad_group.name,
            'resource_name': row.ad_group.resource_name
        })

    total_campaigns = len(cache)
    total_ad_groups = sum(len(c['ad_groups']) for c in cache.values())
    print(f"✅ Cached {total_campaigns} campaigns with {total_ad_groups} ad groups\n")

    return cache


def process_exclusion_sheet_v2(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None,
    save_interval: int = 10,
    dry_run: bool = False,
):
    """
    Process the 'uitsluiten' (exclusion) sheet - V2 with cat_ids mapping.

    OPTIMIZED VERSION (V2):
    - Groups shops by (maincat_id, cl1) for batch processing
    - Reads each ad group's listing tree ONCE per group instead of once per shop
    - Pre-fetches all PLA campaigns and ad groups upfront
    - Uses batch mutations for faster processing

    Excel columns (uitsluiten):
    A. Shop name - shop to exclude
    B. Shop ID (not used)
    C. maincat - category name
    D. maincat_id - used to look up deepest_cats
    E. custom label 1 (a/b/c)
    F. result (TRUE/FALSE) - updated by script
    G. error message (when status is FALSE)

    Logic:
    1. Group all rows by (maincat_id, custom_label_1)
    2. For each group, look up deepest_cats ONCE
    3. For each deepest_cat, find campaign PLA/{deepest_cat}_{cl1}
    4. For each ad group, add ALL shop exclusions in one batch

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving)
        save_interval: Save progress every N groups
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING EXCLUSION SHEET V2: '{SHEET_EXCLUSION}'")
    print(f"(OPTIMIZED: Grouping shops by maincat_id + cl1)")
    if dry_run:
        print("(DRY RUN: no shop exclusions will actually be added to Google Ads)")
    print(f"{'='*70}")

    # Load cat_ids mapping
    print("\nLoading cat_ids mapping...")
    cat_ids_mapping = load_cat_ids_mapping(workbook)
    if not cat_ids_mapping:
        print("❌ No cat_ids mapping loaded, cannot process exclusions")
        return

    try:
        sheet = workbook[SHEET_EXCLUSION]
    except KeyError:
        print(f"❌ Sheet '{SHEET_EXCLUSION}' not found in workbook")
        return

    # Pre-fetch all PLA campaigns and ad groups
    print("\nPre-fetching PLA campaigns and ad groups...")
    campaign_cache = prefetch_pla_campaigns_and_ad_groups(client, customer_id, "PLA/")

    # =========================================================================
    # STEP 1: Group all rows by (maincat_id, cl1) for efficient batch processing
    # =========================================================================
    print("\nGrouping rows by (maincat_id, cl1)...")

    # Structure: {(maincat_id, cl1): [(row_idx, shop_name), ...]}
    groups = defaultdict(list)
    rows_with_missing_fields = []
    maincat_name_by_id: dict = {}  # maincat_id -> human-readable maincat name (col C)

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if already processed
        status_value = row[COL_EX_STATUS].value if len(row) > COL_EX_STATUS else None
        if status_value is not None and status_value != '':
            continue

        shop_name = row[COL_EX_SHOP_NAME].value
        maincat_name = row[COL_EX_MAINCAT].value if len(row) > COL_EX_MAINCAT else None
        maincat_id = row[COL_EX_MAINCAT_ID].value
        custom_label_1 = row[COL_EX_CUSTOM_LABEL_1].value

        # Skip empty rows
        if not shop_name:
            continue

        # Track rows with missing required fields (maincat_id and cl1 are both required)
        missing = []
        if not maincat_id:
            missing.append("maincat_id")
        if not custom_label_1:
            missing.append("custom_label_1")
        if missing:
            rows_with_missing_fields.append((idx, missing))
            continue

        maincat_id_str = str(maincat_id)
        cl1_str = str(custom_label_1)
        if maincat_name and maincat_id_str not in maincat_name_by_id:
            maincat_name_by_id[maincat_id_str] = str(maincat_name)
        groups[(maincat_id_str, cl1_str)].append((idx, shop_name))

    # Per-maincat counters for the summary:
    #   categories_by_maincat       = distinct deepest_cats under the maincat
    #   slots_by_maincat            = categories × number of cl1 groups
    #   campaigns_found_by_maincat  = how many slots were in the prefetch cache
    #   missing_campaigns_by_maincat = list of PLA/{cat}_{cl1} names NOT in cache
    #                                  (aggregated across all cl1 groups)
    campaigns_found_by_maincat: dict = defaultdict(int)
    categories_by_maincat: dict = {}
    slots_by_maincat: dict = defaultdict(int)
    missing_campaigns_by_maincat: dict = defaultdict(list)

    # Mark rows with missing fields as errors
    for idx, missing in rows_with_missing_fields:
        missing_str = ", ".join(missing)
        print(f"[Row {idx}] ⚠️  Missing required fields: {missing_str}, skipping")
        sheet.cell(row=idx, column=COL_EX_STATUS + 1).value = False
        sheet.cell(row=idx, column=COL_EX_ERROR + 1).value = f"Missing required fields: {missing_str}"

    total_groups = len(groups)
    total_rows = sum(len(rows) for rows in groups.values())
    print(f"Found {total_rows} row(s) in {total_groups} unique (maincat_id, cl1) group(s)")
    print(f"Rows with missing fields: {len(rows_with_missing_fields)}")

    if total_groups == 0:
        print("No rows to process.")
        return

    # =========================================================================
    # STEP 2: Process each group - fetch campaigns/ad groups ONCE per group
    # =========================================================================
    success_count = 0
    error_count = 0
    groups_processed = 0
    # Run-wide action counters — distinct from per-row "successful". A row
    # is TRUE if the shop ends up excluded for any reason (newly added OR
    # already excluded), so "Rows OK" can be 3 while "Exclusions actually
    # added" is 0.
    run_total_added = 0
    run_total_already_excluded = 0
    run_total_mutate_errors = 0
    run_total_batch_calls = 0

    for (maincat_id_str, cl1_str), rows in groups.items():
        groups_processed += 1
        shop_names = [shop_name for _, shop_name in rows]
        row_indices = [idx for idx, _ in rows]

        # For CL3 targeting, split shop_name at | and use first part
        # e.g. "Hbm-machines.com|NL" becomes "Hbm-machines.com"
        shop_names_for_targeting = [name.split('|')[0] if '|' in name else name for name in shop_names]
        # Create mapping from targeting name back to original name(s)
        targeting_to_original = {}
        for orig, tgt in zip(shop_names, shop_names_for_targeting):
            if tgt not in targeting_to_original:
                targeting_to_original[tgt] = []
            targeting_to_original[tgt].append(orig)

        print(f"\n{'='*60}")
        print(f"[Group {groups_processed}/{total_groups}] maincat_id={maincat_id_str}, cl1={cl1_str}")
        print(f"  Shops to process: {len(shop_names)}")
        print(f"  Shop names: {', '.join(shop_names[:5])}{'...' if len(shop_names) > 5 else ''}")
        # Show if any names were split
        split_names = [(orig, tgt) for orig, tgt in zip(shop_names, shop_names_for_targeting) if orig != tgt]
        if split_names:
            print(f"  CL3 targeting (split): {', '.join([f'{tgt} (from {orig})' for orig, tgt in split_names[:3]])}{'...' if len(split_names) > 3 else ''}")

        # Look up deepest_cats for this maincat_id ONCE for the entire group
        deepest_cats = cat_ids_mapping.get(maincat_id_str, [])
        if not deepest_cats:
            print(f"  ⚠️  No deepest_cats found for maincat_id={maincat_id_str}")
            # Mark all rows in this group as failed
            for idx in row_indices:
                sheet.cell(row=idx, column=COL_EX_STATUS + 1).value = False
                sheet.cell(row=idx, column=COL_EX_ERROR + 1).value = f"No deepest_cats for maincat_id={maincat_id_str}"
                error_count += 1
            continue

        print(f"  Found {len(deepest_cats)} deepest_cat(s)")

        # Record category / slot counts for the maincat-level summary at the end
        categories_by_maincat[maincat_id_str] = len(deepest_cats)
        slots_by_maincat[maincat_id_str] += len(deepest_cats)

        # Track results per shop
        shop_results = {shop: {'success': 0, 'already_excluded': 0, 'errors': []} for shop in shop_names}
        campaigns_found = 0
        missing_campaigns: list = []  # PLA/{deepest_cat}_{cl1} names not present in the Google Ads cache
        total_exclusions_added = 0

        # Process each deepest_cat ONCE for all shops in this group
        for deepest_cat in deepest_cats:
            campaign_name = f"PLA/{deepest_cat}_{cl1_str}"
            campaign_data = campaign_cache.get(campaign_name)

            if not campaign_data:
                missing_campaigns.append(campaign_name)
                print(f"    ⚠️  Campaign not found in Google Ads cache: {campaign_name}")

            if campaign_data:
                campaigns_found += 1
                ad_groups = campaign_data['ad_groups']
                print(f"    📁 Campaign: {campaign_name} ({len(ad_groups)} ad group(s))")
                # Emit a tree line per campaign so the Affected Product Trees
                # column in the export has something to show. Format is
                # intentionally parseable by _parse_affected_entities' tree regex.
                mc_name_disp = maincat_name_by_id.get(maincat_id_str, f"maincat_id={maincat_id_str}")
                shops_disp = ", ".join(sorted(set(shop_names))[:5]) + ("..." if len(set(shop_names)) > 5 else "")
                tree_verb = "Tree to modify" if dry_run else "Tree modified"
                print(f"      🌳 {tree_verb}: Campaign '{campaign_name}' → Maincat '{mc_name_disp}' → CL1 '{cl1_str}' → Shops: {shops_disp}")

                for ag in ad_groups:
                    ag_id = str(ag['id'])
                    ag_name = ag['name']

                    # Retry logic for connection errors
                    max_retries = 3
                    retry_delay = 2

                    for attempt in range(max_retries):
                        try:
                            # Call batch function with targeting names (split at |)
                            # Use unique targeting names to avoid duplicates
                            unique_targeting_names = list(set(shop_names_for_targeting))
                            if dry_run:
                                # DRY RUN: simulate — pretend every shop would be added
                                # successfully without actually calling Google Ads.
                                result = {
                                    'success': list(unique_targeting_names),
                                    'already_excluded': [],
                                    'errors': [],
                                }
                            else:
                                result = add_shop_exclusions_batch(
                                    client=client,
                                    customer_id=customer_id,
                                    ad_group_id=ag_id,
                                    ad_group_name=ag_name,
                                    shop_names=unique_targeting_names
                                )

                            # Log results per ad group. NOTE: these locals
                            # (success_count, error_count) SHADOW the outer
                            # row-level counters with the same name. The
                            # run-wide action totals below are the precise
                            # "what Google Ads actually did" numbers.
                            success_count = len(result['success'])
                            already_count = len(result['already_excluded'])
                            error_count = len(result['errors'])
                            run_total_batch_calls += 1
                            run_total_added += success_count
                            run_total_already_excluded += already_count
                            run_total_mutate_errors += error_count
                            if error_count > 0:
                                print(f"      ❌ {ag_name}: {error_count} error(s), {success_count} added, {already_count} already excluded")
                                for shop, err in result['errors'][:3]:  # Show first 3 errors
                                    print(f"         - {shop}: {err[:60]}")
                            elif success_count > 0:
                                print(f"      ✅ {ag_name}: {success_count} added, {already_count} already excluded")
                            else:
                                print(f"      ⏭️  {ag_name}: all {already_count} already excluded")

                            # Aggregate results - map targeting names back to original names
                            for targeting_name in result['success']:
                                # Update all original names that map to this targeting name
                                for orig_name in targeting_to_original.get(targeting_name, [targeting_name]):
                                    if orig_name in shop_results:
                                        shop_results[orig_name]['success'] += 1
                                        total_exclusions_added += 1
                            for targeting_name in result['already_excluded']:
                                for orig_name in targeting_to_original.get(targeting_name, [targeting_name]):
                                    if orig_name in shop_results:
                                        shop_results[orig_name]['already_excluded'] += 1
                            for targeting_name, error in result['errors']:
                                for orig_name in targeting_to_original.get(targeting_name, [targeting_name]):
                                    if orig_name in shop_results:
                                        shop_results[orig_name]['errors'].append(f"{ag_name}: {error}")

                            break  # Success, exit retry loop

                        except Exception as e:
                            error_str = str(e)
                            if "failed to connect" in error_str.lower() or "unavailable" in error_str.lower():
                                if attempt < max_retries - 1:
                                    print(f"    ⚠️  Connection error, retrying in {retry_delay}s...")
                                    time.sleep(retry_delay)
                                    retry_delay *= 2
                                    continue
                            # Non-retryable error or max retries reached
                            error_msg = str(e)[:50]
                            print(f"      ❌ {ag_name}: {error_msg}")
                            for shop in shop_names:
                                shop_results[shop]['errors'].append(f"{ag_name}: {error_msg}")
                            break

                    # Rate limiting delay after each ad group (not each shop!)
                    time.sleep(0.3)

        print(f"  Summary: {campaigns_found} campaign(s), {total_exclusions_added} exclusion(s) added")
        if missing_campaigns:
            preview = ", ".join(missing_campaigns[:10])
            more = f" (+{len(missing_campaigns) - 10} more)" if len(missing_campaigns) > 10 else ""
            print(f"  ⚠️  {len(missing_campaigns)} campaign name(s) missing from Google Ads cache: {preview}{more}")
        campaigns_found_by_maincat[maincat_id_str] += campaigns_found
        missing_campaigns_by_maincat[maincat_id_str].extend(missing_campaigns)

        # =========================================================================
        # STEP 3: Update row statuses based on results
        # =========================================================================
        for idx, shop_name in rows:
            result = shop_results[shop_name]

            # Consider success if: at least one exclusion added OR already excluded
            # AND no errors occurred
            has_errors = len(result['errors']) > 0
            has_activity = result['success'] > 0 or result['already_excluded'] > 0

            if campaigns_found == 0:
                # No campaigns found at all - this is an error
                sheet.cell(row=idx, column=COL_EX_STATUS + 1).value = False
                sheet.cell(row=idx, column=COL_EX_ERROR + 1).value = f"No campaigns found for maincat_id={maincat_id_str}"
                error_count += 1
                print(f"    Row {idx} ({shop_name}): ❌ No campaigns")
            elif has_errors:
                sheet.cell(row=idx, column=COL_EX_STATUS + 1).value = False
                error_summary = "; ".join(result['errors'][:3])
                sheet.cell(row=idx, column=COL_EX_ERROR + 1).value = error_summary[:100]
                error_count += 1
                print(f"    Row {idx} ({shop_name}): ❌ {len(result['errors'])} error(s)")
            else:
                sheet.cell(row=idx, column=COL_EX_STATUS + 1).value = True
                sheet.cell(row=idx, column=COL_EX_ERROR + 1).value = ""
                success_count += 1
                print(f"    Row {idx} ({shop_name}): ✅ added={result['success']}, already={result['already_excluded']}")

        # Save periodically (every N groups)
        if file_path and groups_processed % save_interval == 0:
            print(f"\n💾 Saving progress ({groups_processed} groups processed)...")
            try:
                workbook.save(file_path)
            except Exception as save_error:
                print(f"⚠️  Error saving: {save_error}")

    # Final save
    if file_path:
        print(f"\n💾 Final save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"⚠️  Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"EXCLUSION SHEET V2 SUMMARY (OPTIMIZED)")
    print(f"{'='*70}")
    print(f"Total groups processed: {groups_processed}")
    print(f"Total rows processed: {success_count + error_count}")
    print(f"Rows with missing fields: {len(rows_with_missing_fields)}")
    print(f"✅ Rows OK: {success_count}  (row didn't error — shop is excluded for any reason)")
    print(f"❌ Rows failed: {error_count + len(rows_with_missing_fields)}")
    print()
    print(f"Batch calls made (1 per ad group): {run_total_batch_calls}")
    print(f"  → Exclusions actually added: {run_total_added}")
    print(f"  → Already excluded (no-op): {run_total_already_excluded}")
    print(f"  → Mutate errors: {run_total_mutate_errors}")
    if run_total_batch_calls > 0 and run_total_added == 0:
        print()
        print("⚠️  No exclusions were actually added. Either every shop was")
        print("    already excluded in every ad group, or every mutate was")
        print("    classified as a no-op (check the per-ad-group log lines).")
    print()
    for mc_id, count in campaigns_found_by_maincat.items():
        mc_name = maincat_name_by_id.get(mc_id, f"maincat_id={mc_id}")
        cats = categories_by_maincat.get(mc_id, 0)
        slots = slots_by_maincat.get(mc_id, 0)
        pct = (count / slots * 100) if slots else 0
        missing = missing_campaigns_by_maincat.get(mc_id, [])
        print(f"Categories in {mc_name}: {cats}")
        print(f"Campaigns found in {mc_name}: {count}/{slots} ({pct:.0f}%)")
        if missing:
            # Stays in the log tail because it's part of the final summary block
            # (important: the /api/dma-plus/status "log" field is truncated to the last 5000 chars).
            print(f"Missing campaigns in {mc_name} ({len(missing)}): {', '.join(missing)}")
    print(f"{'='*70}\n")


def process_exclusion_sheet_new(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None,
    save_interval: int = 10
):
    """
    Process the 'uitsluiten_new' exclusion sheet — simplified version.

    Directly uses the campaign name from column C (no cat_ids lookup needed).
    Groups shops by campaign name for batch processing.

    Excel columns (uitsluiten_new):
    A. Shop name - shop to exclude (CL3 targeting)
    B. (not used)
    C. Campaign name - exact campaign to target
    D. result (TRUE/FALSE) - updated by script
    E. error message (when status is FALSE)

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving)
        save_interval: Save progress every N groups
    """
    SHEET_NAME = "uitsluiten_new"
    COL_SHOP = 0       # Column A: shop name
    COL_AG_NAME = 2    # Column C: ad group name
    COL_STATUS = 5     # Column F: result
    COL_ERROR = 6      # Column G: error message

    print(f"\n{'='*70}")
    print(f"PROCESSING EXCLUSION SHEET: '{SHEET_NAME}'")
    print(f"{'='*70}")

    try:
        sheet = workbook[SHEET_NAME]
    except KeyError:
        print(f"❌ Sheet '{SHEET_NAME}' not found in workbook")
        return

    # Pre-fetch all PLA campaigns and ad groups, build ad group name lookup
    print("\nPre-fetching PLA campaigns and ad groups...")
    campaign_cache = prefetch_pla_campaigns_and_ad_groups(client, customer_id, "PLA/")

    # Build a lookup: ad_group_name -> {'id': ..., 'name': ..., 'campaign': ...}
    ag_name_lookup = {}
    for campaign_name, campaign_data in campaign_cache.items():
        for ag in campaign_data['ad_groups']:
            ag_name_lookup[ag['name']] = {
                'id': ag['id'],
                'name': ag['name'],
                'resource_name': ag['resource_name'],
                'campaign': campaign_name
            }
    print(f"   Built lookup for {len(ag_name_lookup)} ad group(s)")

    # =========================================================================
    # STEP 1: Group rows by ad group name for batch processing
    # =========================================================================
    print("\nGrouping rows by ad group name...")

    # Structure: {ad_group_name: [(row_idx, shop_name), ...]}
    groups = defaultdict(list)
    rows_with_missing_fields = []

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if already processed
        status_value = row[COL_STATUS].value if len(row) > COL_STATUS else None
        if status_value is not None and status_value != '':
            continue

        shop_name = row[COL_SHOP].value
        ag_name = row[COL_AG_NAME].value

        # Skip empty rows
        if not shop_name:
            continue

        if not ag_name:
            rows_with_missing_fields.append(idx)
            continue

        shop_name = str(shop_name).strip()
        ag_name = str(ag_name).strip()
        groups[ag_name].append((idx, shop_name))

    # Mark rows with missing fields as errors
    for idx in rows_with_missing_fields:
        print(f"[Row {idx}] ⚠️  Missing ad group name, skipping")
        sheet.cell(row=idx, column=COL_STATUS + 1).value = False
        sheet.cell(row=idx, column=COL_ERROR + 1).value = "Missing ad group name"

    total_groups = len(groups)
    total_rows = sum(len(rows) for rows in groups.values())
    print(f"Found {total_rows} row(s) in {total_groups} unique ad group(s)")
    if rows_with_missing_fields:
        print(f"Rows with missing fields: {len(rows_with_missing_fields)}")

    if total_groups == 0:
        print("No rows to process.")
        return

    # =========================================================================
    # STEP 2: Process each ad group
    # =========================================================================
    success_count = 0
    error_count = 0
    groups_processed = 0

    for ag_name_key, rows in groups.items():
        groups_processed += 1
        shop_names = [shop_name for _, shop_name in rows]
        row_indices = [idx for idx, _ in rows]

        # For CL3 targeting, split shop_name at | and use first part
        shop_names_for_targeting = [name.split('|')[0] if '|' in name else name for name in shop_names]
        targeting_to_original = {}
        for orig, tgt in zip(shop_names, shop_names_for_targeting):
            if tgt not in targeting_to_original:
                targeting_to_original[tgt] = []
            targeting_to_original[tgt].append(orig)

        print(f"\n{'='*60}")
        print(f"[Group {groups_processed}/{total_groups}] Ad group: {ag_name_key}")
        print(f"  Shops to exclude: {len(shop_names)}")
        print(f"  Shop names: {', '.join(shop_names[:5])}{'...' if len(shop_names) > 5 else ''}")

        # Look up ad group in cache
        ag_data = ag_name_lookup.get(ag_name_key)
        if not ag_data:
            print(f"  ❌ Ad group not found: '{ag_name_key}'")
            for idx in row_indices:
                sheet.cell(row=idx, column=COL_STATUS + 1).value = False
                sheet.cell(row=idx, column=COL_ERROR + 1).value = "Ad group not found"
                error_count += 1
            continue

        ag_id = str(ag_data['id'])
        print(f"  Campaign: {ag_data['campaign']}")

        # Track results per shop
        shop_results = {shop: {'success': 0, 'already_excluded': 0, 'errors': []} for shop in shop_names}
        total_exclusions_added = 0

        max_retries = 3
        retry_delay = 2

        for attempt in range(max_retries):
            try:
                unique_targeting_names = list(set(shop_names_for_targeting))
                result = add_shop_exclusions_batch(
                    client=client,
                    customer_id=customer_id,
                    ad_group_id=ag_id,
                    ad_group_name=ag_name_key,
                    shop_names=unique_targeting_names
                )

                s_count = len(result['success'])
                a_count = len(result['already_excluded'])
                e_count = len(result['errors'])
                if e_count > 0:
                    print(f"    ❌ {e_count} error(s), {s_count} added, {a_count} already excluded")
                    for shop, err in result['errors'][:3]:
                        print(f"       - {shop}: {err[:60]}")
                elif s_count > 0:
                    print(f"    ✅ {s_count} added, {a_count} already excluded")
                else:
                    print(f"    ⏭️  all {a_count} already excluded")

                # Aggregate results — map targeting names back to original names
                for targeting_name in result['success']:
                    for orig_name in targeting_to_original.get(targeting_name, [targeting_name]):
                        if orig_name in shop_results:
                            shop_results[orig_name]['success'] += 1
                            total_exclusions_added += 1
                for targeting_name in result['already_excluded']:
                    for orig_name in targeting_to_original.get(targeting_name, [targeting_name]):
                        if orig_name in shop_results:
                            shop_results[orig_name]['already_excluded'] += 1
                for targeting_name, error in result['errors']:
                    for orig_name in targeting_to_original.get(targeting_name, [targeting_name]):
                        if orig_name in shop_results:
                            shop_results[orig_name]['errors'].append(error)

                break  # Success, exit retry loop

            except Exception as e:
                error_str = str(e)
                if ("failed to connect" in error_str.lower() or "unavailable" in error_str.lower()) and attempt < max_retries - 1:
                    print(f"    ⚠️  Connection error, retrying in {retry_delay}s...")
                    time.sleep(retry_delay)
                    retry_delay *= 2
                    continue
                error_msg = str(e)[:50]
                print(f"    ❌ {error_msg}")
                for shop in shop_names:
                    shop_results[shop]['errors'].append(error_msg)
                break

        time.sleep(0.3)

        # =========================================================================
        # STEP 3: Update row statuses
        # =========================================================================
        for idx, shop_name in rows:
            r = shop_results[shop_name]
            has_errors = len(r['errors']) > 0

            if has_errors:
                sheet.cell(row=idx, column=COL_STATUS + 1).value = False
                error_summary = "; ".join(r['errors'][:3])
                sheet.cell(row=idx, column=COL_ERROR + 1).value = error_summary[:100]
                error_count += 1
            else:
                sheet.cell(row=idx, column=COL_STATUS + 1).value = True
                sheet.cell(row=idx, column=COL_ERROR + 1).value = ""
                success_count += 1

        # Save periodically
        if file_path and groups_processed % save_interval == 0:
            print(f"\n💾 Saving progress ({groups_processed} groups processed)...")
            try:
                workbook.save(file_path)
            except Exception as save_error:
                print(f"⚠️  Error saving: {save_error}")

    # Final save
    if file_path:
        print(f"\n💾 Final save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"⚠️  Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"EXCLUSION SHEET '{SHEET_NAME}' SUMMARY")
    print(f"{'='*70}")
    print(f"Total ad groups processed: {groups_processed}")
    print(f"Total rows processed: {success_count + error_count}")
    if rows_with_missing_fields:
        print(f"Rows with missing fields: {len(rows_with_missing_fields)}")
    print(f"✅ Successful: {success_count}")
    print(f"❌ Failed: {error_count + len(rows_with_missing_fields)}")
    print(f"{'='*70}\n")


def process_check_sheet(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None,
    save_interval: int = 10
):
    """
    Process the 'check' sheet - replace pipe-version shop exclusions with clean lowercase versions.

    Reads the 'check' sheet and for each shop_name containing '|', finds the matching
    CL3 exclusion in the listing tree and replaces it with a clean version (lowercase,
    without pipe and country suffix).

    Example: "Artandcraft.com|NL" -> "artandcraft.com"

    Excel columns (check):
    A. Shop name - shop with pipe (e.g. "Artandcraft.com|NL")
    B. Shop ID (not used)
    C. maincat - category name
    D. maincat_id - used to look up deepest_cats
    E. custom label 1 (a/b/c)
    F. result (TRUE/FALSE) - updated by script
    G. error message (when status is FALSE)

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving)
        save_interval: Save progress every N groups
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING CHECK SHEET: '{SHEET_CHECK}'")
    print(f"(Replace pipe-version exclusions with clean lowercase versions)")
    print(f"{'='*70}")

    # Load cat_ids mapping
    print("\nLoading cat_ids mapping...")
    cat_ids_mapping = load_cat_ids_mapping(workbook)
    if not cat_ids_mapping:
        print("No cat_ids mapping loaded, cannot process check sheet")
        return

    try:
        sheet = workbook[SHEET_CHECK]
    except KeyError:
        print(f"Sheet '{SHEET_CHECK}' not found in workbook")
        return

    # Pre-fetch all PLA campaigns and ad groups
    print("\nPre-fetching PLA campaigns and ad groups...")
    campaign_cache = prefetch_pla_campaigns_and_ad_groups(client, customer_id, "PLA/")

    # =========================================================================
    # STEP 1: Group all rows by (maincat_id, cl1)
    # =========================================================================
    print("\nGrouping rows by (maincat_id, cl1)...")

    # Structure: {(maincat_id, cl1): [(row_idx, shop_name), ...]}
    groups = defaultdict(list)
    rows_with_missing_fields = []

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if already processed
        status_value = row[COL_CHK_STATUS].value if len(row) > COL_CHK_STATUS else None
        if status_value is not None and status_value != '':
            continue

        shop_name = row[COL_CHK_SHOP_NAME].value
        maincat_id = row[COL_CHK_MAINCAT_ID].value
        custom_label_1 = row[COL_CHK_CUSTOM_LABEL_1].value

        # Skip empty rows
        if not shop_name:
            continue

        # Validate that shop_name contains '|'
        if '|' not in str(shop_name):
            print(f"[Row {idx}] Skipping '{shop_name}' - no pipe character found")
            sheet.cell(row=idx, column=COL_CHK_STATUS + 1).value = False
            sheet.cell(row=idx, column=COL_CHK_ERROR + 1).value = "No pipe character in shop name"
            continue

        # Track rows with missing required fields
        if not maincat_id or not custom_label_1:
            rows_with_missing_fields.append(idx)
            continue

        maincat_id_str = str(maincat_id)
        cl1_str = str(custom_label_1)
        groups[(maincat_id_str, cl1_str)].append((idx, str(shop_name)))

    # Mark rows with missing fields as errors
    for idx in rows_with_missing_fields:
        print(f"[Row {idx}] Missing required fields, skipping")
        sheet.cell(row=idx, column=COL_CHK_STATUS + 1).value = False
        sheet.cell(row=idx, column=COL_CHK_ERROR + 1).value = "Missing required fields"

    total_groups = len(groups)
    total_rows = sum(len(rows) for rows in groups.values())
    print(f"Found {total_rows} row(s) in {total_groups} unique (maincat_id, cl1) group(s)")
    print(f"Rows with missing fields: {len(rows_with_missing_fields)}")

    if total_groups == 0:
        print("No rows to process.")
        return

    # =========================================================================
    # STEP 2: Process each group
    # =========================================================================
    success_count = 0
    error_count = 0
    groups_processed = 0

    for (maincat_id_str, cl1_str), rows in groups.items():
        groups_processed += 1
        shop_names = [shop_name for _, shop_name in rows]
        row_indices = [idx for idx, _ in rows]

        # Build replacements dict: {old_name: clean_name}
        replacements = {}
        for shop_name in shop_names:
            clean_name = shop_name.split('|')[0].lower()
            replacements[shop_name] = clean_name

        print(f"\n{'='*60}")
        print(f"[Group {groups_processed}/{total_groups}] maincat_id={maincat_id_str}, cl1={cl1_str}")
        print(f"  Shops to replace: {len(shop_names)}")
        for old, new in list(replacements.items())[:5]:
            print(f"    {old} -> {new}")
        if len(replacements) > 5:
            print(f"    ... and {len(replacements) - 5} more")

        # Look up deepest_cats for this maincat_id
        deepest_cats = cat_ids_mapping.get(maincat_id_str, [])
        if not deepest_cats:
            print(f"  No deepest_cats found for maincat_id={maincat_id_str}")
            for idx in row_indices:
                sheet.cell(row=idx, column=COL_CHK_STATUS + 1).value = False
                sheet.cell(row=idx, column=COL_CHK_ERROR + 1).value = f"No deepest_cats for maincat_id={maincat_id_str}"
                error_count += 1
            continue

        print(f"  Found {len(deepest_cats)} deepest_cat(s)")

        # Track results per shop
        shop_results = {shop: {'success': 0, 'already_clean': 0, 'not_found': 0, 'errors': []} for shop in shop_names}
        campaigns_found = 0
        total_replacements = 0

        # Process each deepest_cat
        for deepest_cat in deepest_cats:
            campaign_name = f"PLA/{deepest_cat}_{cl1_str}"
            campaign_data = campaign_cache.get(campaign_name)

            if campaign_data:
                campaigns_found += 1
                ad_groups = campaign_data['ad_groups']
                print(f"    Campaign: {campaign_name} ({len(ad_groups)} ad group(s))")

                for ag in ad_groups:
                    ag_id = str(ag['id'])
                    ag_name = ag['name']

                    # Retry logic for connection errors
                    max_retries = 3
                    retry_delay = 2

                    for attempt in range(max_retries):
                        try:
                            batch_result = replace_shop_exclusions_batch(
                                client=client,
                                customer_id=customer_id,
                                ad_group_id=ag_id,
                                ad_group_name=ag_name,
                                replacements=replacements
                            )

                            # Log results per ad group
                            n_success = len(batch_result['success'])
                            n_clean = len(batch_result['already_clean'])
                            n_notfound = len(batch_result['not_found'])
                            n_errors = len(batch_result['errors'])

                            if n_success > 0 or n_clean > 0:
                                print(f"      {ag_name}: {n_success} replaced, {n_clean} already clean, {n_notfound} not found")
                                for old_name, new_name in batch_result['success']:
                                    print(f"        REPLACED: {old_name} -> {new_name}")
                                for old_name, new_name in batch_result['already_clean']:
                                    print(f"        CLEANED (already existed): {old_name} -> {new_name}")
                            elif n_errors > 0:
                                print(f"      {ag_name}: {n_errors} error(s)")
                                for old_name, err in batch_result['errors'][:3]:
                                    print(f"        - {old_name}: {err[:60]}")
                            else:
                                print(f"      {ag_name}: no matching exclusions found")

                            # Aggregate results
                            for old_name, new_name in batch_result['success']:
                                if old_name in shop_results:
                                    shop_results[old_name]['success'] += 1
                                    total_replacements += 1
                            for old_name, new_name in batch_result['already_clean']:
                                if old_name in shop_results:
                                    shop_results[old_name]['already_clean'] += 1
                                    total_replacements += 1
                            for old_name in batch_result['not_found']:
                                if old_name in shop_results:
                                    shop_results[old_name]['not_found'] += 1
                            for old_name, error in batch_result['errors']:
                                if old_name in shop_results:
                                    shop_results[old_name]['errors'].append(f"{ag_name}: {error}")

                            break  # Success, exit retry loop

                        except Exception as e:
                            error_str = str(e)
                            if "failed to connect" in error_str.lower() or "unavailable" in error_str.lower():
                                if attempt < max_retries - 1:
                                    print(f"    Connection error, retrying in {retry_delay}s...")
                                    time.sleep(retry_delay)
                                    retry_delay *= 2
                                    continue
                            # Non-retryable error or max retries reached
                            error_msg = str(e)[:50]
                            print(f"      {ag_name}: {error_msg}")
                            for shop in shop_names:
                                shop_results[shop]['errors'].append(f"{ag_name}: {error_msg}")
                            break

                    # Rate limiting delay
                    time.sleep(0.3)

        print(f"  Summary: {campaigns_found} campaign(s), {total_replacements} replacement(s)")

        # =========================================================================
        # STEP 3: Update row statuses
        # =========================================================================
        for idx, shop_name in rows:
            res = shop_results[shop_name]

            has_errors = len(res['errors']) > 0
            has_activity = res['success'] > 0 or res['already_clean'] > 0

            if campaigns_found == 0:
                sheet.cell(row=idx, column=COL_CHK_STATUS + 1).value = False
                sheet.cell(row=idx, column=COL_CHK_ERROR + 1).value = f"No campaigns found for maincat_id={maincat_id_str}"
                error_count += 1
                print(f"    Row {idx} ({shop_name}): No campaigns")
            elif has_errors:
                sheet.cell(row=idx, column=COL_CHK_STATUS + 1).value = False
                error_summary = "; ".join(res['errors'][:3])
                sheet.cell(row=idx, column=COL_CHK_ERROR + 1).value = error_summary[:100]
                error_count += 1
                print(f"    Row {idx} ({shop_name}): {len(res['errors'])} error(s)")
            elif has_activity:
                sheet.cell(row=idx, column=COL_CHK_STATUS + 1).value = True
                sheet.cell(row=idx, column=COL_CHK_ERROR + 1).value = ""
                success_count += 1
                print(f"    Row {idx} ({shop_name}): replaced={res['success']}, already_clean={res['already_clean']}")
            else:
                # Not found in any ad group - mark as success (nothing to replace)
                sheet.cell(row=idx, column=COL_CHK_STATUS + 1).value = True
                sheet.cell(row=idx, column=COL_CHK_ERROR + 1).value = "Not found in any ad group (no action needed)"
                success_count += 1
                print(f"    Row {idx} ({shop_name}): not found in any ad group")

        # Save periodically
        if file_path and groups_processed % save_interval == 0:
            print(f"\nSaving progress ({groups_processed} groups processed)...")
            try:
                workbook.save(file_path)
            except Exception as save_error:
                print(f"Error saving: {save_error}")

    # Final save
    if file_path:
        print(f"\nFinal save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"CHECK SHEET SUMMARY")
    print(f"{'='*70}")
    print(f"Total groups processed: {groups_processed}")
    print(f"Total rows processed: {success_count + error_count}")
    print(f"Rows with missing fields: {len(rows_with_missing_fields)}")
    print(f"Successful: {success_count}")
    print(f"Failed: {error_count + len(rows_with_missing_fields)}")
    print(f"{'='*70}\n")


def process_check_cl1_sheet(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None,
    save_interval: int = 10
):
    """
    Check and fix CL1 targeting for ad groups created by process_inclusion_sheet_v2.

    Reads the 'toevoegen' sheet and for each ad group checks whether the listing
    tree has CL1 (custom_label_1) targeting matching the cl1 from the ad group name.
    If CL1 is missing, the tree is rebuilt with CL1 + maincat_id targeting while
    preserving any existing CL3 shop exclusions.

    Excel columns (toevoegen) - same as process_inclusion_sheet_v2:
    A. shop_name
    B. Shop ID (not used)
    C. maincat
    D. maincat_id
    E. custom label 1 (a/b/c)
    F. budget (not used here)
    G. result (TRUE/FALSE) - updated by script
    H. error message

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving)
        save_interval: Save progress every N campaigns
    """
    print(f"\n{'='*70}")
    print(f"CHECKING CL1 TARGETING: '{SHEET_INCLUSION}'")
    print(f"(Check and rebuild trees missing CL1 targeting)")
    print(f"{'='*70}")

    try:
        sheet = workbook[SHEET_INCLUSION]
    except KeyError:
        print(f"Sheet '{SHEET_INCLUSION}' not found in workbook")
        return

    # Load data_only workbook for formula results (same pattern as process_inclusion_sheet_v2)
    data_workbook = None
    data_sheet = None
    if file_path:
        try:
            data_workbook = load_workbook(file_path, data_only=True)
            data_sheet = data_workbook[SHEET_INCLUSION]
            print("   (Using data_only mode to read formula results)")
        except Exception as e:
            print(f"   Could not load data_only workbook: {e}")

    # Local column constants (same as process_inclusion_sheet_v2)
    COL_SHOP_NAME = 0      # A: shop_name
    COL_SHOP_ID = 1        # B: Shop ID (not used)
    COL_MAINCAT = 2        # C: maincat
    COL_MAINCAT_ID = 3     # D: maincat_id
    COL_CL1 = 4            # E: custom label 1
    COL_BUDGET = 5         # F: budget (not used here)
    COL_RESULT = 6         # G: result (TRUE/FALSE)
    COL_ERR = 7            # H: error message

    # =========================================================================
    # STEP 1: Read and group rows by campaign (maincat + cl1), then by shop_name
    # =========================================================================
    print("\nStep 1: Reading and grouping rows...")

    campaigns = defaultdict(lambda: {
        'maincat': None,
        'cl1': None,
        'ad_groups': defaultdict(lambda: {'maincat_ids': set(), 'rows': []}),
        'rows': []
    })

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if already processed
        status_value = row[COL_RESULT].value if len(row) > COL_RESULT else None
        if status_value is not None and status_value != '':
            continue

        # Read values from data_only sheet if available
        if data_sheet:
            shop_name = data_sheet.cell(row=idx, column=COL_SHOP_NAME + 1).value
            maincat = data_sheet.cell(row=idx, column=COL_MAINCAT + 1).value
            maincat_id = data_sheet.cell(row=idx, column=COL_MAINCAT_ID + 1).value
            custom_label_1 = data_sheet.cell(row=idx, column=COL_CL1 + 1).value
        else:
            shop_name = row[COL_SHOP_NAME].value
            maincat = row[COL_MAINCAT].value
            maincat_id = row[COL_MAINCAT_ID].value
            custom_label_1 = row[COL_CL1].value

        # Validate required fields
        if not shop_name or not maincat or not maincat_id or not custom_label_1:
            if shop_name:  # Only log if there's a shop_name (skip truly empty rows)
                print(f"   [Row {idx}] Missing required fields, skipping")
                sheet.cell(row=idx, column=COL_RESULT + 1).value = False
                sheet.cell(row=idx, column=COL_ERR + 1).value = "Missing required fields"
            continue

        campaign_name = f"PLA/{maincat} store_{custom_label_1}"

        campaigns[campaign_name]['maincat'] = maincat
        campaigns[campaign_name]['cl1'] = custom_label_1
        campaigns[campaign_name]['rows'].append({'idx': idx, 'row': row})

        campaigns[campaign_name]['ad_groups'][shop_name]['maincat_ids'].add(maincat_id)
        campaigns[campaign_name]['ad_groups'][shop_name]['rows'].append({'idx': idx, 'row': row})

    total_campaigns = len(campaigns)
    total_ad_groups = sum(len(c['ad_groups']) for c in campaigns.values())
    print(f"   Found {total_campaigns} campaign(s), {total_ad_groups} ad group(s) to check\n")

    if total_campaigns == 0:
        print("No rows to process.")
        return

    # =========================================================================
    # STEP 2: Pre-fetch PLA campaigns and ad groups
    # =========================================================================
    print("Step 2: Pre-fetching PLA campaigns and ad groups...")
    campaign_cache = prefetch_pla_campaigns_and_ad_groups(client, customer_id, "PLA/")

    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")

    # =========================================================================
    # STEP 3: Process each campaign
    # =========================================================================
    campaigns_processed = 0
    ag_checked = 0
    ag_rebuilt = 0
    ag_already_ok = 0
    ag_errors = 0

    for campaign_name, campaign_data in campaigns.items():
        campaigns_processed += 1
        cl1 = campaign_data['cl1']
        ad_groups = campaign_data['ad_groups']

        print(f"\n{'='*60}")
        print(f"[Campaign {campaigns_processed}/{total_campaigns}] {campaign_name}")
        print(f"  Expected CL1: {cl1}")
        print(f"  Ad groups to check: {len(ad_groups)}")

        # Find campaign in cache
        cached_campaign = campaign_cache.get(campaign_name)
        if not cached_campaign:
            print(f"  Campaign not found in Google Ads, skipping")
            for shop_name, ag_data in ad_groups.items():
                for row_info in ag_data['rows']:
                    sheet.cell(row=row_info['idx'], column=COL_RESULT + 1).value = False
                    sheet.cell(row=row_info['idx'], column=COL_ERR + 1).value = f"Campaign '{campaign_name}' not found"
                    ag_errors += 1
            continue

        # Build lookup of ad groups by name
        cached_ag_by_name = {ag['name']: ag for ag in cached_campaign['ad_groups']}

        for shop_name, ag_data in ad_groups.items():
            ag_checked += 1
            ad_group_name = f"PLA/{shop_name}_{cl1}"
            maincat_ids = sorted(ag_data['maincat_ids'])

            print(f"\n    Checking: {ad_group_name}")

            # Find ad group in cache
            cached_ag = cached_ag_by_name.get(ad_group_name)
            if not cached_ag:
                print(f"      Ad group not found in Google Ads")
                for row_info in ag_data['rows']:
                    sheet.cell(row=row_info['idx'], column=COL_RESULT + 1).value = False
                    sheet.cell(row=row_info['idx'], column=COL_ERR + 1).value = f"Ad group '{ad_group_name}' not found"
                    ag_errors += 1
                continue

            ad_group_id = str(cached_ag['id'])
            ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

            # Read existing tree
            query = f"""
                SELECT
                    ad_group_criterion.resource_name,
                    ad_group_criterion.listing_group.type,
                    ad_group_criterion.listing_group.parent_ad_group_criterion,
                    ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
                    ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
                    ad_group_criterion.negative
                FROM ad_group_criterion
                WHERE ad_group_criterion.ad_group = '{ag_path}'
                    AND ad_group_criterion.type = 'LISTING_GROUP'
            """

            try:
                tree_rows = list(ga_service.search(customer_id=customer_id, query=query))
            except Exception as e:
                error_msg = f"Error reading tree: {str(e)[:50]}"
                print(f"      {error_msg}")
                for row_info in ag_data['rows']:
                    sheet.cell(row=row_info['idx'], column=COL_RESULT + 1).value = False
                    sheet.cell(row=row_info['idx'], column=COL_ERR + 1).value = error_msg
                    ag_errors += 1
                continue

            if not tree_rows:
                print(f"      No listing tree found, skipping")
                for row_info in ag_data['rows']:
                    sheet.cell(row=row_info['idx'], column=COL_RESULT + 1).value = False
                    sheet.cell(row=row_info['idx'], column=COL_ERR + 1).value = "No listing tree found"
                    ag_errors += 1
                continue

            # Check if CL1 targeting exists and matches
            has_correct_cl1 = False
            existing_cl3_exclusions = []
            shop_name_for_targeting = shop_name.split('|')[0] if '|' in shop_name else shop_name

            for tree_row in tree_rows:
                criterion = tree_row.ad_group_criterion
                lg = criterion.listing_group

                try:
                    index_name = lg.case_value.product_custom_attribute.index.name
                except (AttributeError, TypeError):
                    index_name = None

                if index_name == 'INDEX1':
                    value = lg.case_value.product_custom_attribute.value
                    if value and not criterion.negative and value.lower() == cl1.lower():
                        has_correct_cl1 = True

                # Collect CL3 negative exclusions (shop exclusions to preserve)
                if index_name == 'INDEX3' and criterion.negative:
                    value = lg.case_value.product_custom_attribute.value
                    if value and value.lower() != shop_name_for_targeting.lower():
                        existing_cl3_exclusions.append(value)

            if has_correct_cl1:
                print(f"      CL1='{cl1}' already present, OK")
                ag_already_ok += 1
                for row_info in ag_data['rows']:
                    sheet.cell(row=row_info['idx'], column=COL_RESULT + 1).value = True
                    sheet.cell(row=row_info['idx'], column=COL_ERR + 1).value = ""
                continue

            # CL1 is missing - rebuild the tree
            print(f"      CL1='{cl1}' MISSING - rebuilding tree")
            if existing_cl3_exclusions:
                print(f"      Preserving {len(existing_cl3_exclusions)} CL3 exclusion(s): {', '.join(existing_cl3_exclusions[:5])}")

            # Retry logic for connection errors
            max_retries = 3
            retry_delay = 2
            rebuild_success = False

            for attempt in range(max_retries):
                try:
                    # Step A: Remove old tree
                    safe_remove_entire_listing_tree(client, customer_id, ad_group_id)
                    time.sleep(1.0)

                    # Step B: Build new tree with CL1
                    build_listing_tree_with_cl1(
                        client=client,
                        customer_id=customer_id,
                        ad_group_id=ad_group_id,
                        shop_name=shop_name_for_targeting,
                        maincat_ids=maincat_ids,
                        custom_label_1=cl1
                    )
                    time.sleep(1.0)

                    # Step C: Re-add CL3 exclusions if any
                    if existing_cl3_exclusions:
                        excl_result = add_shop_exclusions_batch(
                            client=client,
                            customer_id=customer_id,
                            ad_group_id=ad_group_id,
                            ad_group_name=ad_group_name,
                            shop_names=existing_cl3_exclusions
                        )
                        print(f"      Re-added exclusions: {len(excl_result['success'])} ok, {len(excl_result['errors'])} errors")

                    rebuild_success = True
                    ag_rebuilt += 1
                    print(f"      REBUILT: {campaign_name} / {ad_group_name} (CL1='{cl1}', {len(maincat_ids)} maincat(s))")
                    break  # Success, exit retry loop

                except Exception as e:
                    error_str = str(e)
                    if "failed to connect" in error_str.lower() or "unavailable" in error_str.lower():
                        if attempt < max_retries - 1:
                            print(f"      Connection error, retrying in {retry_delay}s...")
                            time.sleep(retry_delay)
                            retry_delay *= 2
                            continue
                    # Non-retryable or max retries
                    error_msg = str(e)[:80]
                    print(f"      Error rebuilding: {error_msg}")
                    for row_info in ag_data['rows']:
                        sheet.cell(row=row_info['idx'], column=COL_RESULT + 1).value = False
                        sheet.cell(row=row_info['idx'], column=COL_ERR + 1).value = error_msg[:100]
                        ag_errors += 1
                    break

            if rebuild_success:
                for row_info in ag_data['rows']:
                    sheet.cell(row=row_info['idx'], column=COL_RESULT + 1).value = True
                    sheet.cell(row=row_info['idx'], column=COL_ERR + 1).value = ""

            # Rate limiting
            time.sleep(0.3)

        # Save periodically
        if file_path and campaigns_processed % save_interval == 0:
            print(f"\nSaving progress ({campaigns_processed} campaigns processed)...")
            try:
                workbook.save(file_path)
            except Exception as save_error:
                print(f"Error saving: {save_error}")

    # Final save
    if file_path:
        print(f"\nFinal save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"CL1 CHECK SUMMARY")
    print(f"{'='*70}")
    print(f"Campaigns processed: {campaigns_processed}")
    print(f"Ad groups checked: {ag_checked}")
    print(f"  Already OK: {ag_already_ok}")
    print(f"  Rebuilt: {ag_rebuilt}")
    print(f"  Errors: {ag_errors}")
    print(f"{'='*70}\n")


def process_check_new_sheet(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None,
    save_interval: int = 10
):
    """
    Replace CL3 shop_name subdivision values containing '|' with clean lowercase versions.

    Reads the 'check_new' sheet where each row directly specifies the campaign name
    and ad group name. For each row, reads the listing tree, removes it, and rebuilds
    with the clean shop name while preserving CL4/CL1 targeting and CL3 exclusions.

    Example: CL3 = "Artandcraft.com|NL" -> "artandcraft.com"

    Excel columns (check_new):
    A. shop_name (with |)
    B. ad_group_name
    C. campaign_name
    D. result (TRUE/FALSE) - updated by script
    E. error message

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file (for saving)
        save_interval: Save progress every N rows
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING CHECK_NEW SHEET: '{SHEET_CHECK_NEW}'")
    print(f"(Replace CL3 pipe-version targeting with clean lowercase versions)")
    print(f"{'='*70}")

    try:
        sheet = workbook[SHEET_CHECK_NEW]
    except KeyError:
        print(f"Sheet '{SHEET_CHECK_NEW}' not found in workbook")
        return

    # Pre-fetch all PLA campaigns and ad groups
    print("\nPre-fetching PLA campaigns and ad groups...")
    campaign_cache = prefetch_pla_campaigns_and_ad_groups(client, customer_id, "PLA/")

    # Build lookup: campaign_name -> {ag_name -> ag_data}
    campaign_ag_lookup = {}
    for camp_name, camp_data in campaign_cache.items():
        campaign_ag_lookup[camp_name] = {ag['name']: ag for ag in camp_data['ad_groups']}

    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")

    # =========================================================================
    # Process each row
    # =========================================================================
    rows_processed = 0
    success_count = 0
    skip_count = 0
    error_count = 0

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if already processed
        status_value = row[COL_CHNEW_STATUS].value if len(row) > COL_CHNEW_STATUS else None
        if status_value is not None and status_value != '':
            continue

        shop_name = row[COL_CHNEW_SHOP_NAME].value
        ad_group_name = row[COL_CHNEW_AD_GROUP_NAME].value
        campaign_name = row[COL_CHNEW_CAMPAIGN_NAME].value

        # Skip empty rows
        if not shop_name:
            continue

        shop_name = str(shop_name).strip()
        ad_group_name = str(ad_group_name).strip() if ad_group_name else None
        campaign_name = str(campaign_name).strip() if campaign_name else None

        rows_processed += 1

        # Validate required fields
        if not ad_group_name or not campaign_name:
            print(f"[Row {idx}] Missing ad_group_name or campaign_name, skipping")
            sheet.cell(row=idx, column=COL_CHNEW_STATUS + 1).value = False
            sheet.cell(row=idx, column=COL_CHNEW_ERROR + 1).value = "Missing required fields"
            error_count += 1
            continue

        # Check if shop_name contains '|'
        if '|' not in shop_name:
            print(f"[Row {idx}] '{shop_name}' has no pipe, skipping")
            sheet.cell(row=idx, column=COL_CHNEW_STATUS + 1).value = True
            sheet.cell(row=idx, column=COL_CHNEW_ERROR + 1).value = "No pipe in shop_name (already clean)"
            skip_count += 1
            continue

        clean_name = shop_name.split('|')[0].lower()
        print(f"\n[Row {idx}] {shop_name} -> {clean_name}")
        print(f"  Campaign: {campaign_name}")
        print(f"  Ad group: {ad_group_name}")

        # Find campaign in cache
        ag_lookup = campaign_ag_lookup.get(campaign_name)
        if not ag_lookup:
            print(f"  Campaign not found in Google Ads")
            sheet.cell(row=idx, column=COL_CHNEW_STATUS + 1).value = False
            sheet.cell(row=idx, column=COL_CHNEW_ERROR + 1).value = f"Campaign '{campaign_name}' not found"
            error_count += 1
            continue

        # Find ad group in campaign
        cached_ag = ag_lookup.get(ad_group_name)
        if not cached_ag:
            print(f"  Ad group not found in campaign")
            sheet.cell(row=idx, column=COL_CHNEW_STATUS + 1).value = False
            sheet.cell(row=idx, column=COL_CHNEW_ERROR + 1).value = f"Ad group '{ad_group_name}' not found"
            error_count += 1
            continue

        ad_group_id = str(cached_ag['id'])
        ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

        # Read existing tree
        query = f"""
            SELECT
                ad_group_criterion.resource_name,
                ad_group_criterion.listing_group.type,
                ad_group_criterion.listing_group.parent_ad_group_criterion,
                ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
                ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
                ad_group_criterion.negative
            FROM ad_group_criterion
            WHERE ad_group_criterion.ad_group = '{ag_path}'
                AND ad_group_criterion.type = 'LISTING_GROUP'
        """

        try:
            tree_rows = list(ga_service.search(customer_id=customer_id, query=query))
        except Exception as e:
            error_msg = f"Error reading tree: {str(e)[:50]}"
            print(f"  {error_msg}")
            sheet.cell(row=idx, column=COL_CHNEW_STATUS + 1).value = False
            sheet.cell(row=idx, column=COL_CHNEW_ERROR + 1).value = error_msg
            error_count += 1
            continue

        if not tree_rows:
            print(f"  No listing tree found")
            sheet.cell(row=idx, column=COL_CHNEW_STATUS + 1).value = False
            sheet.cell(row=idx, column=COL_CHNEW_ERROR + 1).value = "No listing tree found"
            error_count += 1
            continue

        # Analyze tree structure
        cl3_subdivision_value = None
        maincat_ids = []
        cl1_value = None
        cl3_exclusions = []

        for tree_row in tree_rows:
            criterion = tree_row.ad_group_criterion
            lg = criterion.listing_group

            try:
                index_name = lg.case_value.product_custom_attribute.index.name
            except (AttributeError, TypeError):
                index_name = None

            if index_name == 'INDEX3':
                value = lg.case_value.product_custom_attribute.value
                if value and lg.type_.name == 'SUBDIVISION':
                    cl3_subdivision_value = value
                elif value and criterion.negative:
                    cl3_exclusions.append(value)

            elif index_name == 'INDEX4':
                value = lg.case_value.product_custom_attribute.value
                if value and not criterion.negative:
                    maincat_ids.append(value)

            elif index_name == 'INDEX1':
                value = lg.case_value.product_custom_attribute.value
                if value and not criterion.negative:
                    cl1_value = value

        # Check if CL3 actually needs replacing
        if not cl3_subdivision_value:
            print(f"  No CL3 subdivision found in tree")
            sheet.cell(row=idx, column=COL_CHNEW_STATUS + 1).value = False
            sheet.cell(row=idx, column=COL_CHNEW_ERROR + 1).value = "No CL3 subdivision found"
            error_count += 1
            continue

        if '|' not in cl3_subdivision_value:
            print(f"  CL3='{cl3_subdivision_value}' already clean")
            sheet.cell(row=idx, column=COL_CHNEW_STATUS + 1).value = True
            sheet.cell(row=idx, column=COL_CHNEW_ERROR + 1).value = "CL3 already clean"
            skip_count += 1
            continue

        if not maincat_ids:
            print(f"  No CL4 (maincat_id) targeting found in tree")
            sheet.cell(row=idx, column=COL_CHNEW_STATUS + 1).value = False
            sheet.cell(row=idx, column=COL_CHNEW_ERROR + 1).value = "No CL4 targeting found"
            error_count += 1
            continue

        maincat_ids = sorted(set(maincat_ids))

        print(f"  CL3: '{cl3_subdivision_value}' -> '{clean_name}'")
        print(f"  CL4 maincat_ids: {maincat_ids}")
        if cl1_value:
            print(f"  CL1: {cl1_value}")
        if cl3_exclusions:
            print(f"  Preserving {len(cl3_exclusions)} CL3 exclusion(s)")

        # Rebuild the tree
        max_retries = 3
        retry_delay = 2
        rebuild_success = False

        for attempt in range(max_retries):
            try:
                # Step 1: Remove old tree
                safe_remove_entire_listing_tree(client, customer_id, ad_group_id)
                time.sleep(1.0)

                # Step 2: Rebuild with clean name
                if cl1_value:
                    build_listing_tree_with_cl1(
                        client=client,
                        customer_id=customer_id,
                        ad_group_id=ad_group_id,
                        shop_name=clean_name,
                        maincat_ids=maincat_ids,
                        custom_label_1=cl1_value
                    )
                else:
                    build_listing_tree_for_inclusion_v2(
                        client=client,
                        customer_id=customer_id,
                        ad_group_id=ad_group_id,
                        shop_name=clean_name,
                        maincat_ids=maincat_ids
                    )
                time.sleep(1.0)

                # Step 3: Re-add CL3 exclusions if any
                if cl3_exclusions:
                    excl_result = add_shop_exclusions_batch(
                        client=client,
                        customer_id=customer_id,
                        ad_group_id=ad_group_id,
                        ad_group_name=ad_group_name,
                        shop_names=cl3_exclusions
                    )
                    n_ok = len(excl_result['success'])
                    n_err = len(excl_result['errors'])
                    print(f"  Re-added exclusions: {n_ok} ok, {n_err} errors")

                rebuild_success = True
                print(f"  REPLACED: '{cl3_subdivision_value}' -> '{clean_name}' in {campaign_name} / {ad_group_name}")
                break

            except Exception as e:
                error_str = str(e)
                if "failed to connect" in error_str.lower() or "unavailable" in error_str.lower():
                    if attempt < max_retries - 1:
                        print(f"  Connection error, retrying in {retry_delay}s...")
                        time.sleep(retry_delay)
                        retry_delay *= 2
                        continue
                error_msg = str(e)[:80]
                print(f"  Error: {error_msg}")
                sheet.cell(row=idx, column=COL_CHNEW_STATUS + 1).value = False
                sheet.cell(row=idx, column=COL_CHNEW_ERROR + 1).value = error_msg[:100]
                error_count += 1
                break

        if rebuild_success:
            sheet.cell(row=idx, column=COL_CHNEW_STATUS + 1).value = True
            sheet.cell(row=idx, column=COL_CHNEW_ERROR + 1).value = ""
            success_count += 1

        # Rate limiting
        time.sleep(0.3)

        # Save periodically
        if file_path and rows_processed % save_interval == 0:
            print(f"\nSaving progress ({rows_processed} rows processed)...")
            try:
                workbook.save(file_path)
            except Exception as save_error:
                print(f"Error saving: {save_error}")

    # Final save
    if file_path:
        print(f"\nFinal save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"CHECK_NEW SHEET SUMMARY")
    print(f"{'='*70}")
    print(f"Rows processed: {rows_processed}")
    print(f"Successful replacements: {success_count}")
    print(f"Skipped (already clean): {skip_count}")
    print(f"Errors: {error_count}")
    print(f"{'='*70}\n")


def process_exclusion_sheet(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str,
    save_interval: int = 10
):
    """
    Process the 'uitsluiten' (exclusion) sheet with GROUPED PROCESSING.

    Groups rows by campaign (cat_uitsluiten + custom_label_1) and collects all
    shops to exclude for each campaign. Then rebuilds each campaign's tree once
    with all shop exclusions.

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to Excel file for saving
        save_interval: Save workbook every N campaign groups (default: 10)
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING EXCLUSION SHEET: '{SHEET_EXCLUSION}' (GROUPED MODE)")
    print(f"{'='*70}")
    print(f"  Strategy: Group rows by campaign, apply all shop exclusions at once")
    print(f"  Save interval: Every {save_interval} campaign groups")
    print(f"{'='*70}\n")

    try:
        sheet = workbook[SHEET_EXCLUSION]
    except KeyError:
        print(f"❌ Sheet '{SHEET_EXCLUSION}' not found in workbook")
        return

    # Step 1: Group rows by campaign and collect shops
    print("Step 1: Grouping rows by campaign...")
    campaign_groups = defaultdict(lambda: {'rows': [], 'shops': set()})

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if row has enough columns
        if len(row) <= COL_EX_CUSTOM_LABEL_1:
            print(f"⚠️  Row {idx}: Not enough columns (has {len(row)}, needs at least {COL_EX_CUSTOM_LABEL_1 + 1}). Skipping.")
            continue

        # Skip rows that already have a status
        status_cell = sheet.cell(row=idx, column=COL_EX_STATUS + 1)  # +1 because openpyxl is 1-indexed
        if status_cell.value is not None and status_cell.value != '':
            continue

        # Extract values safely
        try:
            shop_name = row[COL_EX_SHOP_NAME].value
            cat_uitsluiten = row[COL_EX_CAT_UITSLUITEN].value
            diepste_cat_id = row[COL_EX_DIEPSTE_CAT_ID].value
            custom_label_1 = row[COL_EX_CUSTOM_LABEL_1].value
        except IndexError as e:
            print(f"⚠️  Row {idx}: Column access error: {e}. Skipping.")
            continue

        # Validate required fields
        if not shop_name or not cat_uitsluiten or not custom_label_1 or not diepste_cat_id:
            sheet.cell(row=idx, column=COL_EX_STATUS + 1).value = False
            sheet.cell(row=idx, column=COL_EX_ERROR + 1).value = "Missing required fields"
            continue

        # Group key: (cat_uitsluiten, custom_label_1)
        group_key = (cat_uitsluiten, str(custom_label_1))

        # Add row and shop to group - store row number, not row tuple
        campaign_groups[group_key]['rows'].append({
            'idx': idx,
            'row_number': idx
        })
        campaign_groups[group_key]['shops'].add(str(shop_name))
        # Store diepste_cat_id (should be same for all rows in group)
        campaign_groups[group_key]['diepste_cat_id'] = str(diepste_cat_id)

    print(f"Found {len(campaign_groups)} campaign group(s) to process")
    print(f"Total rows: {sum(len(g['rows']) for g in campaign_groups.values())}\n")

    if len(campaign_groups) == 0:
        print("✅ No campaign groups to process")
        return

    # Step 2: Process each campaign group
    print("="*70)
    print("Step 2: Processing campaign groups...")
    print("="*70)

    success_count = 0
    fail_count = 0
    groups_processed = 0

    for i, (group_key, group_data) in enumerate(campaign_groups.items(), 1):
        try:
            cat_uitsluiten, custom_label_1 = group_key
        except (ValueError, TypeError) as e:
            print(f"\n❌ ERROR unpacking group_key: {group_key}")
            print(f"   Error: {e}")
            print(f"   Skipping this group...")
            continue

        rows = group_data['rows']
        shops = sorted(group_data['shops'])
        diepste_cat_id = group_data.get('diepste_cat_id')

        campaign_pattern = f"PLA/{cat_uitsluiten}_{custom_label_1}"

        print(f"\n{'─'*70}")
        print(f"GROUP {i}/{len(campaign_groups)}: {campaign_pattern}")
        print(f"{'─'*70}")
        print(f"   Rows in group: {len(rows)}")
        print(f"   Diepste cat ID (CL0): {diepste_cat_id}")
        print(f"   Shops to exclude: {len(shops)}")
        print(f"   Shop names: {', '.join(shops)}")

        try:
            # Find campaign and ad group
            result = get_campaign_and_ad_group_by_pattern(client, customer_id, campaign_pattern)

            if not result:
                print(f"   ❌ Campaign not found")
                # Mark all rows in group as NOT_FOUND
                for row_info in rows:
                    row_num = row_info['row_number']
                    sheet.cell(row=row_num, column=COL_EX_STATUS + 1).value = False
                    sheet.cell(row=row_num, column=COL_EX_ERROR + 1).value = "Campaign not found"
                    fail_count += 1
                continue

            print(f"   ✅ Found: Campaign ID {result['campaign']['id']}, Ad Group ID {result['ad_group']['id']}")

            # Rebuild tree with all shop exclusions and required CL0 targeting
            rebuild_tree_with_shop_exclusions(
                client,
                customer_id,
                result['ad_group']['id'],
                shop_names=shops,  # Pass all shops for this campaign
                required_cl0_value=diepste_cat_id  # Required CL0 from Excel
            )

            # Mark all rows in group as SUCCESS
            for row_info in rows:
                row_num = row_info['row_number']
                sheet.cell(row=row_num, column=COL_EX_STATUS + 1).value = True
                sheet.cell(row=row_num, column=COL_EX_ERROR + 1).value = ""  # Clear error message
                success_count += 1

            groups_processed += 1
            print(f"   ✅ SUCCESS - Tree rebuilt with {len(shops)} shop exclusion(s)")

        except Exception as e:
            print(f"   ❌ ERROR: {e}")
            # Mark all rows in group as ERROR
            # Create brief, user-friendly error message
            error_str = str(e)

            # Shorten common error types
            if "SUBDIVISION_REQUIRES_OTHERS_CASE" in error_str:
                error_msg = "Tree structure error: missing OTHERS case"
            elif "LISTING_GROUP_SUBDIVISION_REQUIRES_OTHERS_CASE" in error_str:
                error_msg = "Tree structure error: missing OTHERS case"
            elif "CONCURRENT_MODIFICATION" in error_str:
                error_msg = "Concurrent modification (retry needed)"
            elif "NOT_FOUND" in error_str or "not found" in error_str.lower():
                error_msg = "Resource not found"
            elif "INVALID_ARGUMENT" in error_str:
                error_msg = "Invalid argument in API call"
            elif "PERMISSION_DENIED" in error_str:
                error_msg = "Permission denied"
            elif "Could not find CL0" in error_str or "Could not find CL1" in error_str:
                error_msg = error_str[:80]  # Keep this one as-is, it's informative
            else:
                # Generic error - truncate but keep key info
                error_msg = error_str[:80] if len(error_str) > 80 else error_str

            for row_info in rows:
                row_num = row_info['row_number']
                sheet.cell(row=row_num, column=COL_EX_STATUS + 1).value = False
                sheet.cell(row=row_num, column=COL_EX_ERROR + 1).value = error_msg
                fail_count += 1

        # Save every N groups
        if i % save_interval == 0:
            print(f"\n   💾 Saving progress... ({i}/{len(campaign_groups)} groups processed)")
            try:
                workbook.save(file_path)
                print(f"   ✅ Progress saved successfully")
            except Exception as save_error:
                print(f"   ⚠️  Error saving file: {save_error}")

    # Final save
    print(f"\n   💾 Final save...")
    try:
        workbook.save(file_path)
        print(f"   ✅ Final save successful")
    except Exception as save_error:
        print(f"   ⚠️  Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"EXCLUSION SHEET SUMMARY")
    print(f"{'='*70}")
    print(f"Total campaign groups processed: {len(campaign_groups)}")
    print(f"✅ Groups successful: {groups_processed}")
    print(f"❌ Groups failed: {len(campaign_groups) - groups_processed}")
    print(f"✅ Total rows marked success: {success_count}")
    print(f"❌ Total rows marked failed: {fail_count}")
    print(f"{'='*70}\n")


# ============================================================================
# CL1 VALIDATION FUNCTIONS (Legacy-based)
# ============================================================================

def validate_cl1_targeting_for_ad_group(
    client: GoogleAdsClient,
    customer_id: str,
    ad_group_id: str,
    ad_group_name: str,
    dry_run: bool = False
) -> dict:
    """
    Validate and fix CL1 (Custom Label 1) targeting for an ad group.

    Checks if the product listing tree has a product group targeting the custom value
    (a, b, or c) that matches the ad group name suffix (_a, _b, or _c).

    If the correct CL1 targeting is missing, it will be added while preserving
    existing exclusions and tree structure.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        ad_group_name: Ad group name (to extract suffix)
        dry_run: If True, only report issues without making changes

    Returns:
        dict with keys:
            - 'status': 'ok', 'fixed', 'skipped', 'error'
            - 'message': Description of what was done/found
            - 'required_cl1': The required CL1 value from ad group name
            - 'existing_cl1': The existing CL1 value(s) found in tree
    """
    result = {
        'status': 'ok',
        'message': '',
        'required_cl1': None,
        'existing_cl1': []
    }

    # Step 1: Extract required CL1 from ad group name suffix
    required_cl1 = None
    for suffix in ['_a', '_b', '_c']:
        if ad_group_name.endswith(suffix):
            required_cl1 = suffix[1:]  # Remove underscore: "_a" → "a"
            break

    if not required_cl1:
        result['status'] = 'skipped'
        result['message'] = f"Ad group name '{ad_group_name}' does not end with _a, _b, or _c"
        return result

    result['required_cl1'] = required_cl1

    # Step 2: Query existing listing tree
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    agc_service = client.get_service("AdGroupCriterionService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    query = f"""
        SELECT
            ad_group_criterion.resource_name,
            ad_group_criterion.listing_group.type,
            ad_group_criterion.listing_group.parent_ad_group_criterion,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
            ad_group_criterion.negative,
            ad_group_criterion.cpc_bid_micros
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
    """

    try:
        rows = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        result['status'] = 'error'
        result['message'] = f"Error querying listing tree: {str(e)[:300]}"
        return result

    if not rows:
        result['status'] = 'skipped'
        result['message'] = "No listing tree found in ad group"
        return result

    # Step 3: Analyze the tree structure
    # Find CL4 nodes (parent for CL1 nodes) and existing CL1 targeting
    cl4_subdivision_resource = None
    cl4_unit_node = None  # CL4 UNIT that needs conversion to SUBDIVISION
    existing_cl1_values = []  # Positive CL1 targets
    cl1_others_exists = False
    existing_bid = DEFAULT_BID_MICROS

    for row in rows:
        criterion = row.ad_group_criterion
        lg = criterion.listing_group
        case_value = lg.case_value

        # Check for custom attribute nodes
        if case_value.product_custom_attribute:
            index = case_value.product_custom_attribute.index.name
            value = case_value.product_custom_attribute.value

            # Track CL4 nodes
            if index == 'INDEX4':
                if lg.type.name == 'SUBDIVISION':
                    cl4_subdivision_resource = criterion.resource_name
                elif lg.type.name == 'UNIT' and not criterion.negative and value:
                    cl4_unit_node = {
                        'resource_name': criterion.resource_name,
                        'parent': lg.parent_ad_group_criterion or None,
                        'value': value,
                        'bid': criterion.cpc_bid_micros or DEFAULT_BID_MICROS,
                    }

            # Track existing CL1 targeting
            if index == 'INDEX1':
                if value:
                    # Specific CL1 value
                    if not criterion.negative:
                        existing_cl1_values.append(value)
                        # Capture bid from positive CL1 units
                        if criterion.cpc_bid_micros:
                            existing_bid = criterion.cpc_bid_micros
                else:
                    # CL1 OTHERS case
                    cl1_others_exists = True

    result['existing_cl1'] = existing_cl1_values

    # Step 4: Check if required CL1 is already targeted
    if required_cl1 in existing_cl1_values:
        result['status'] = 'ok'
        result['message'] = f"CL1='{required_cl1}' already targeted correctly"
        return result

    # CL1 is missing - need to add it
    # If CL4 is a UNIT, we need to convert it to a SUBDIVISION first
    needs_cl4_conversion = False
    if not cl4_subdivision_resource:
        if cl4_unit_node:
            needs_cl4_conversion = True
            existing_bid = cl4_unit_node['bid']
        else:
            result['status'] = 'error'
            result['message'] = "Could not find CL4 node (maincat) to add CL1 under"
            return result

    if dry_run:
        result['status'] = 'fixed'
        msg = f"[DRY RUN] Would add CL1='{required_cl1}' targeting (existing: {existing_cl1_values or 'none'})"
        if needs_cl4_conversion:
            msg += f" (converting CL4='{cl4_unit_node['value']}' UNIT to SUBDIVISION)"
        result['message'] = msg
        return result

    # Step 5: Add the missing CL1 targeting
    ops = []

    if needs_cl4_conversion:
        # Remove the old CL4 UNIT
        remove_op = client.get_type("AdGroupCriterionOperation")
        remove_op.remove = cl4_unit_node['resource_name']
        ops.append(remove_op)

        # Create CL4 SUBDIVISION with same dimension and parent
        dim_cl4 = client.get_type("ListingDimensionInfo")
        dim_cl4.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX4
        dim_cl4.product_custom_attribute.value = cl4_unit_node['value']

        subdiv_op = create_listing_group_subdivision(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=cl4_unit_node['parent'],
            listing_dimension_info=dim_cl4
        )
        cl4_subdivision_resource = subdiv_op.create.resource_name
        ops.append(subdiv_op)

    # If CL1 OTHERS doesn't exist, we need to add it
    # (Every subdivision must have an OTHERS case)
    if not cl1_others_exists:
        dim_cl1_others = client.get_type("ListingDimensionInfo")
        dim_cl1_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
        # Don't set value - this is the OTHERS case

        ops.append(
            create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=ad_group_id,
                parent_ad_group_criterion_resource_name=cl4_subdivision_resource,
                listing_dimension_info=dim_cl1_others,
                targeting_negative=True,  # OTHERS is always negative
                cpc_bid_micros=None
            )
        )

    # Add the positive CL1 target
    dim_cl1 = client.get_type("ListingDimensionInfo")
    dim_cl1.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
    dim_cl1.product_custom_attribute.value = required_cl1

    ops.append(
        create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=ad_group_id,
            parent_ad_group_criterion_resource_name=cl4_subdivision_resource,
            listing_dimension_info=dim_cl1,
            targeting_negative=False,  # Positive target
            cpc_bid_micros=existing_bid  # Use existing bid or default
        )
    )

    try:
        agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops)
        result['status'] = 'fixed'
        result['message'] = f"Added CL1='{required_cl1}' targeting (bid: {existing_bid/1_000_000:.2f}€)"
        if needs_cl4_conversion:
            result['message'] += f" (converted CL4='{cl4_unit_node['value']}' UNIT→SUBDIVISION)"
        if not cl1_others_exists:
            result['message'] += " + added CL1 OTHERS"
        return result
    except Exception as e:
        error_msg = str(e)
        if "LISTING_GROUP_ALREADY_EXISTS" in error_msg:
            result['status'] = 'ok'
            result['message'] = f"CL1='{required_cl1}' already exists (concurrent update)"
            return result
        result['status'] = 'error'
        result['message'] = f"Error adding CL1 targeting: {error_msg[:100]}"
        return result


def validate_cl1_targeting_for_campaigns(
    client: GoogleAdsClient,
    customer_id: str,
    campaign_name_pattern: str = None,
    dry_run: bool = False
) -> dict:
    """
    Validate and fix CL1 targeting for all ad groups in matching campaigns.

    Iterates through all campaigns (optionally filtered by name pattern) and
    validates that each ad group's product listing tree targets the correct
    CL1 value based on the ad group name suffix (_a, _b, or _c).

    Args:
        client: Google Ads client
        customer_id: Customer ID
        campaign_name_pattern: Optional pattern to filter campaigns (uses LIKE)
                              e.g., "PLA/%" for all PLA campaigns
        dry_run: If True, only report issues without making changes

    Returns:
        dict with summary statistics and details
    """
    print(f"\n{'='*70}")
    print("CL1 TARGETING VALIDATION")
    print(f"{'='*70}")
    print(f"Customer ID: {customer_id}")
    print(f"Campaign filter: {campaign_name_pattern or '(all campaigns)'}")
    print(f"Dry run: {dry_run}")
    print(f"{'='*70}\n")

    ga_service = client.get_service("GoogleAdsService")

    # Step 1: Query campaigns and ad groups
    where_clause = "campaign.status != 'REMOVED' AND ad_group.status != 'REMOVED'"
    if campaign_name_pattern:
        # Escape single quotes in pattern
        escaped_pattern = campaign_name_pattern.replace("'", "\\'")
        where_clause += f" AND campaign.name LIKE '{escaped_pattern}'"

    query = f"""
        SELECT
            campaign.id,
            campaign.name,
            campaign.resource_name,
            ad_group.id,
            ad_group.name,
            ad_group.resource_name
        FROM ad_group
        WHERE {where_clause}
        ORDER BY campaign.name, ad_group.name
    """

    try:
        results = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        print(f"❌ Error querying campaigns/ad groups: {e}")
        return {'error': str(e)}

    print(f"Found {len(results)} ad group(s) to validate\n")

    # Step 2: Process each ad group
    stats = {
        'total': len(results),
        'ok': 0,
        'fixed': 0,
        'skipped': 0,
        'error': 0,
        'details': []
    }

    current_campaign = None

    for row in results:
        campaign_name = row.campaign.name
        ag_id = row.ad_group.id
        ag_name = row.ad_group.name

        # Print campaign header when it changes
        if campaign_name != current_campaign:
            current_campaign = campaign_name
            print(f"\n📁 Campaign: {campaign_name}")

        # Validate this ad group
        result = validate_cl1_targeting_for_ad_group(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ag_id),
            ad_group_name=ag_name,
            dry_run=dry_run
        )

        # Update stats
        stats[result['status']] += 1

        # Log result
        status_icon = {
            'ok': '✅',
            'fixed': '🔧',
            'skipped': '⏭️',
            'error': '❌'
        }.get(result['status'], '❓')

        print(f"   {status_icon} {ag_name}: {result['message']}")

        # Store details for reporting
        stats['details'].append({
            'campaign': campaign_name,
            'ad_group': ag_name,
            'ad_group_id': ag_id,
            **result
        })

        # Rate limiting
        if result['status'] == 'fixed' and not dry_run:
            time.sleep(0.5)

    # Print summary
    print(f"\n{'='*70}")
    print("CL1 VALIDATION SUMMARY")
    print(f"{'='*70}")
    print(f"Total ad groups: {stats['total']}")
    print(f"✅ Already correct: {stats['ok']}")
    print(f"🔧 Fixed: {stats['fixed']}")
    print(f"⏭️  Skipped (no _a/_b/_c suffix): {stats['skipped']}")
    print(f"❌ Errors: {stats['error']}")
    print(f"{'='*70}\n")

    # (Previously wrote an xlsx to a hardcoded Windows path here — removed:
    # the same details are already present in stats['details'] and are
    # exported via the dashboard's Export button.)

    return stats


def load_deepest_cat_to_cat_id_mapping(workbook: openpyxl.Workbook) -> dict:
    """
    Load the cat_ids sheet and create a reverse mapping of deepest_cat -> cat_id.

    This is used by validate_listing_trees_for_campaigns to look up the cat_id
    (INDEX0 value) from the campaign name (which contains the deepest_cat).

    Args:
        workbook: Excel workbook containing cat_ids sheet

    Returns:
        dict: {deepest_cat: cat_id} (both as strings)
    """
    try:
        sheet = workbook[SHEET_CAT_IDS]
    except KeyError:
        print(f"❌ Sheet '{SHEET_CAT_IDS}' not found in workbook")
        return {}

    mapping = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        deepest_cat = row[COL_CAT_DEEPEST_CAT]
        cat_id = row[COL_CAT_CAT_ID]

        if deepest_cat and cat_id:
            deepest_cat_str = str(deepest_cat).strip()
            cat_id_str = str(int(cat_id)) if isinstance(cat_id, float) else str(cat_id).strip()
            if deepest_cat_str not in mapping:
                mapping[deepest_cat_str] = cat_id_str

    print(f"   Loaded {len(mapping)} deepest_cat → cat_id mappings from '{SHEET_CAT_IDS}' sheet")
    return mapping


def validate_listing_trees_for_campaigns(
    client: GoogleAdsClient,
    customer_id: str,
    campaign_name_pattern: str = "PLA/%",
    dry_run: bool = True,
    excel_path: str = None
) -> dict:
    """
    Validate that all ad groups in matching campaigns have a listing tree.
    If a tree is missing, create one based on:
    - CL1 (INDEX1): from the ad group/campaign name suffix (_a, _b, _c)
    - Cat_id (INDEX0): looked up from the cat_ids sheet via deepest_cat in campaign name

    Target tree structure:
        ROOT (SUBDIVISION)
        ├─ INDEX1 = cl1 (SUBDIVISION)
        │  ├─ INDEX0 = cat_id (UNIT, positive, biddable)
        │  └─ INDEX0 = OTHERS (UNIT, negative)
        └─ INDEX1 = OTHERS (UNIT, negative)

    Campaign name format: PLA/{deepest_cat}_{cl1}
    Example: PLA/Auto-interieur_a → deepest_cat="Auto-interieur", cl1="a"

    Args:
        client: Google Ads client
        customer_id: Customer ID
        campaign_name_pattern: LIKE pattern to filter campaigns (default: "PLA/%")
        dry_run: If True, only report issues without making changes
        excel_path: Path to Excel file with cat_ids sheet (default: EXCEL_FILE_PATH)

    Returns:
        dict with summary statistics
    """
    print(f"\n{'='*70}")
    print("LISTING TREE VALIDATION")
    print(f"{'='*70}")
    print(f"Customer ID: {customer_id}")
    print(f"Campaign filter: {campaign_name_pattern}")
    print(f"Dry run: {dry_run}")
    print(f"{'='*70}\n")

    # Step 1: Load cat_ids mapping from Excel
    file_path = excel_path or EXCEL_FILE_PATH
    print(f"📂 Loading cat_ids from: {file_path}")
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        deepest_cat_mapping = load_deepest_cat_to_cat_id_mapping(workbook)
        workbook.close()
    except Exception as e:
        print(f"❌ Error loading Excel: {e}")
        return {'error': str(e)}

    if not deepest_cat_mapping:
        print("❌ No cat_id mappings loaded — aborting")
        return {'error': 'No cat_id mappings'}

    # Step 2: Query campaigns and ad groups
    ga_service = client.get_service("GoogleAdsService")
    agc_service = client.get_service("AdGroupCriterionService")
    ag_service = client.get_service("AdGroupService")

    escaped_pattern = campaign_name_pattern.replace("'", "\\'")
    query = f"""
        SELECT
            campaign.id,
            campaign.name,
            ad_group.id,
            ad_group.name,
            ad_group.resource_name
        FROM ad_group
        WHERE campaign.name LIKE '{escaped_pattern}'
          AND campaign.status != 'REMOVED'
          AND ad_group.status != 'REMOVED'
        ORDER BY campaign.name, ad_group.name
    """

    try:
        results = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        print(f"❌ Error querying campaigns/ad groups: {e}")
        return {'error': str(e)}

    print(f"Found {len(results)} ad group(s) to validate\n")

    # Step 3: For each ad group, check if listing tree exists
    stats = {
        'total': len(results),
        'ok': 0,
        'created': 0,
        'skipped': 0,
        'error': 0,
        'details': []
    }

    current_campaign = None

    for row in results:
        campaign_name = row.campaign.name
        ag_id = str(row.ad_group.id)
        ag_name = row.ad_group.name

        if campaign_name != current_campaign:
            current_campaign = campaign_name
            print(f"\n📁 Campaign: {campaign_name}")

        # Parse campaign name: PLA/{deepest_cat}_{cl1}
        # Strip "PLA/" prefix, then split on last "_" to get deepest_cat and cl1
        if not campaign_name.startswith("PLA/"):
            print(f"   ⏭️  {ag_name}: Campaign name doesn't start with PLA/")
            stats['skipped'] += 1
            stats['details'].append({
                'campaign': campaign_name, 'ad_group': ag_name,
                'status': 'skipped', 'message': "Campaign name doesn't start with PLA/"
            })
            continue

        name_part = campaign_name[4:]  # Remove "PLA/"
        last_underscore = name_part.rfind('_')
        if last_underscore == -1:
            print(f"   ⏭️  {ag_name}: No _a/_b/_c suffix found in campaign name")
            stats['skipped'] += 1
            stats['details'].append({
                'campaign': campaign_name, 'ad_group': ag_name,
                'status': 'skipped', 'message': "No _a/_b/_c suffix in campaign name"
            })
            continue

        deepest_cat = name_part[:last_underscore]
        cl1 = name_part[last_underscore + 1:]

        if cl1 not in ('a', 'b', 'c'):
            print(f"   ⏭️  {ag_name}: Suffix '{cl1}' is not a/b/c")
            stats['skipped'] += 1
            stats['details'].append({
                'campaign': campaign_name, 'ad_group': ag_name,
                'status': 'skipped', 'message': f"Suffix '{cl1}' is not a/b/c"
            })
            continue

        # Look up cat_id
        cat_id = deepest_cat_mapping.get(deepest_cat)
        if not cat_id:
            print(f"   ❌ {ag_name}: No cat_id found for deepest_cat='{deepest_cat}'")
            stats['error'] += 1
            stats['details'].append({
                'campaign': campaign_name, 'ad_group': ag_name,
                'status': 'error', 'message': f"No cat_id for deepest_cat='{deepest_cat}'"
            })
            continue

        # Check if listing tree exists
        ag_path = ag_service.ad_group_path(customer_id, ag_id)
        tree_query = f"""
            SELECT ad_group_criterion.resource_name
            FROM ad_group_criterion
            WHERE ad_group_criterion.ad_group = '{ag_path}'
                AND ad_group_criterion.type = 'LISTING_GROUP'
            LIMIT 1
        """

        try:
            tree_rows = list(ga_service.search(customer_id=customer_id, query=tree_query))
        except Exception as e:
            print(f"   ❌ {ag_name}: Error checking tree: {str(e)[:80]}")
            stats['error'] += 1
            stats['details'].append({
                'campaign': campaign_name, 'ad_group': ag_name,
                'status': 'error', 'message': f"Error checking tree: {str(e)[:80]}"
            })
            continue

        if tree_rows:
            stats['ok'] += 1
            stats['details'].append({
                'campaign': campaign_name, 'ad_group': ag_name,
                'status': 'ok', 'message': 'Listing tree exists'
            })
            continue

        # No tree — create one
        if dry_run:
            print(f"   🔧 {ag_name}: [DRY RUN] Would create tree: CL1='{cl1}', CL0='{cat_id}'")
            stats['created'] += 1
            stats['details'].append({
                'campaign': campaign_name, 'ad_group': ag_name,
                'status': 'created', 'message': f"[DRY RUN] Would create: CL1='{cl1}', CL0='{cat_id}'"
            })
            continue

        # Build the listing tree
        try:
            ops = []

            # 1. ROOT SUBDIVISION
            root_op = create_listing_group_subdivision(
                client=client,
                customer_id=customer_id,
                ad_group_id=ag_id,
                parent_ad_group_criterion_resource_name=None,
                listing_dimension_info=None
            )
            root_tmp = root_op.create.resource_name
            ops.append(root_op)

            # 2. INDEX1 = cl1 (SUBDIVISION)
            dim_cl1 = client.get_type("ListingDimensionInfo")
            dim_cl1.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1
            dim_cl1.product_custom_attribute.value = cl1

            cl1_subdiv_op = create_listing_group_subdivision(
                client=client,
                customer_id=customer_id,
                ad_group_id=ag_id,
                parent_ad_group_criterion_resource_name=root_tmp,
                listing_dimension_info=dim_cl1
            )
            cl1_subdiv_tmp = cl1_subdiv_op.create.resource_name
            ops.append(cl1_subdiv_op)

            # 3. INDEX1 = OTHERS (UNIT, negative)
            dim_cl1_others = client.get_type("ListingDimensionInfo")
            dim_cl1_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX1

            ops.append(create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=ag_id,
                parent_ad_group_criterion_resource_name=root_tmp,
                listing_dimension_info=dim_cl1_others,
                targeting_negative=True,
                cpc_bid_micros=None
            ))

            # 4. INDEX0 = cat_id (UNIT, positive, biddable)
            dim_cl0 = client.get_type("ListingDimensionInfo")
            dim_cl0.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
            dim_cl0.product_custom_attribute.value = cat_id

            ops.append(create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=ag_id,
                parent_ad_group_criterion_resource_name=cl1_subdiv_tmp,
                listing_dimension_info=dim_cl0,
                targeting_negative=False,
                cpc_bid_micros=DEFAULT_BID_MICROS
            ))

            # 5. INDEX0 = OTHERS (UNIT, negative)
            dim_cl0_others = client.get_type("ListingDimensionInfo")
            dim_cl0_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0

            ops.append(create_listing_group_unit_biddable(
                client=client,
                customer_id=customer_id,
                ad_group_id=ag_id,
                parent_ad_group_criterion_resource_name=cl1_subdiv_tmp,
                listing_dimension_info=dim_cl0_others,
                targeting_negative=True,
                cpc_bid_micros=None
            ))

            # Execute all operations atomically
            agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops)
            print(f"   ✅ {ag_name}: Created tree: CL1='{cl1}', CL0='{cat_id}'")
            stats['created'] += 1
            stats['details'].append({
                'campaign': campaign_name, 'ad_group': ag_name,
                'status': 'created', 'message': f"Created tree: CL1='{cl1}', CL0='{cat_id}'"
            })

            time.sleep(0.3)

        except GoogleAdsException as gae:
            error_msgs = [f"{e.error_code}: {e.message}" for e in gae.failure.errors]
            print(f"   ❌ {ag_name}: {'; '.join(error_msgs)}")
            stats['error'] += 1
            stats['details'].append({
                'campaign': campaign_name, 'ad_group': ag_name,
                'status': 'error', 'message': f"GoogleAdsException: {error_msgs[0][:300]}"
            })
        except Exception as e:
            print(f"   ❌ {ag_name}: {str(e)[:300]}")
            stats['error'] += 1
            stats['details'].append({
                'campaign': campaign_name, 'ad_group': ag_name,
                'status': 'error', 'message': str(e)[:300]
            })

    # Print summary
    print(f"\n{'='*70}")
    print("LISTING TREE VALIDATION SUMMARY")
    print(f"{'='*70}")
    print(f"Total ad groups: {stats['total']}")
    print(f"✅ Already has tree: {stats['ok']}")
    print(f"🔧 Created: {stats['created']}")
    print(f"⏭️  Skipped: {stats['skipped']}")
    print(f"❌ Errors: {stats['error']}")
    print(f"{'='*70}\n")

    return stats


def validate_ads_for_campaigns(
    client: GoogleAdsClient,
    customer_id: str,
    campaign_name_pattern: str = None,
    fix: bool = False,
    dry_run: bool = False,
) -> dict:
    """
    Validate that all ad groups in matching campaigns have at least one ad.

    Iterates through all campaigns (optionally filtered by name pattern) and
    checks whether each ad group contains a shopping product ad. Optionally
    creates missing ads.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        campaign_name_pattern: Optional pattern to filter campaigns (uses LIKE)
                              e.g., "% store_%" for all store campaigns
        fix: If True, create missing ads; if False, only report

    Returns:
        dict with summary statistics
    """
    print(f"\n{'='*70}")
    print("AD PRESENCE VALIDATION")
    if fix and dry_run:
        print("(DRY RUN: missing ads will be reported but no ads will be created)")
    print(f"{'='*70}")
    print(f"Customer ID: {customer_id}")
    print(f"Campaign filter: {campaign_name_pattern or '(all campaigns)'}")
    print(f"Fix mode: {fix}")
    print(f"{'='*70}\n")

    ga_service = client.get_service("GoogleAdsService")

    # Query all ad groups with their ad count in a single query
    where_clause = "campaign.status != 'REMOVED' AND ad_group.status != 'REMOVED'"
    if campaign_name_pattern:
        escaped_pattern = campaign_name_pattern.replace("'", "\\'")
        where_clause += f" AND campaign.name LIKE '{escaped_pattern}'"

    query = f"""
        SELECT
            campaign.id,
            campaign.name,
            ad_group.id,
            ad_group.name,
            ad_group.resource_name
        FROM ad_group
        WHERE {where_clause}
        ORDER BY campaign.name, ad_group.name
    """

    try:
        ad_groups = list(ga_service.search(customer_id=customer_id, query=query))
    except Exception as e:
        print(f"❌ Error querying ad groups: {e}")
        return {'error': str(e)}

    print(f"Found {len(ad_groups)} ad group(s) to check\n")

    if not ad_groups:
        return {'total': 0, 'with_ads': 0, 'missing_ads': 0, 'fixed': 0, 'errors': 0}

    # Query all ads for matching campaigns in a single call
    ad_where = "campaign.status != 'REMOVED' AND ad_group.status != 'REMOVED' AND ad_group_ad.status != 'REMOVED'"
    if campaign_name_pattern:
        ad_where += f" AND campaign.name LIKE '{escaped_pattern}'"

    ad_query = f"""
        SELECT
            ad_group.id,
            ad_group_ad.ad.id
        FROM ad_group_ad
        WHERE {ad_where}
    """

    try:
        ad_results = list(ga_service.search(customer_id=customer_id, query=ad_query))
    except Exception as e:
        print(f"❌ Error querying ads: {e}")
        return {'error': str(e)}

    # Build set of ad group IDs that have ads
    ag_ids_with_ads = set()
    for row in ad_results:
        ag_ids_with_ads.add(row.ad_group.id)

    print(f"Found {len(ag_ids_with_ads)} ad group(s) with existing ads\n")

    # Check each ad group
    stats = {
        'total': len(ad_groups),
        'with_ads': 0,
        'missing_ads': 0,
        'fixed': 0,
        'errors': 0,
        'details': []
    }

    current_campaign = None
    missing_count_in_campaign = 0

    for row in ad_groups:
        campaign_name = row.campaign.name
        ag_id = row.ad_group.id
        ag_name = row.ad_group.name
        ag_resource = row.ad_group.resource_name

        if campaign_name != current_campaign:
            if current_campaign and missing_count_in_campaign > 0:
                print(f"   ⚠️  {missing_count_in_campaign} ad group(s) missing ads")
            current_campaign = campaign_name
            missing_count_in_campaign = 0
            print(f"\n📁 Campaign: {campaign_name}")

        if ag_id in ag_ids_with_ads:
            stats['with_ads'] += 1
        else:
            stats['missing_ads'] += 1
            missing_count_in_campaign += 1
            print(f"   ❌ {ag_name} — NO AD")

            stats['details'].append({
                'campaign': campaign_name,
                'ad_group': ag_name,
                'ad_group_id': ag_id,
                'ad_group_resource': ag_resource
            })

            if fix:
                if dry_run:
                    print(f"      [DRY RUN] Would create shopping ad")
                    stats['fixed'] += 1
                else:
                    try:
                        add_shopping_product_ad(client, customer_id, ag_resource)
                        print(f"      🔧 Created shopping ad")
                        stats['fixed'] += 1
                        time.sleep(0.3)
                    except Exception as e:
                        print(f"      ❌ Error creating ad: {e}")
                        stats['errors'] += 1

    # Print last campaign's missing count
    if current_campaign and missing_count_in_campaign > 0:
        print(f"   ⚠️  {missing_count_in_campaign} ad group(s) missing ads")

    # Summary
    print(f"\n{'='*70}")
    print("AD VALIDATION SUMMARY")
    print(f"{'='*70}")
    print(f"Total ad groups: {stats['total']}")
    print(f"✅ With ads: {stats['with_ads']}")
    print(f"❌ Missing ads: {stats['missing_ads']}")
    if fix:
        print(f"🔧 Fixed: {stats['fixed']}")
        print(f"❌ Fix errors: {stats['errors']}")
    print(f"{'='*70}\n")

    # (Previously wrote an xlsx to a hardcoded Windows path here — removed:
    # the same details are already present in stats['details'] and are
    # exported via the dashboard's Export button.)

    return stats


def process_reverse_exclusion_sheet(
    client: GoogleAdsClient,
    workbook: openpyxl.Workbook,
    customer_id: str,
    file_path: str = None,
    save_interval: int = 10,
    sheet_name: str = "verwijderen",
    dry_run: bool = False,
):
    """
    Process a sheet to REMOVE shop exclusions (reverse of exclusion).

    OPTIMIZED VERSION: Groups shops by (maincat_id, cl1) and processes them together,
    reading each ad group's listing tree only ONCE per group instead of once per shop.

    This function reads a sheet with shop/category data and removes the CL3 exclusion
    for each shop, effectively un-excluding them from the campaigns.

    Uses the same cat_ids mapping approach as process_exclusion_sheet_v2:
    1. Get maincat_id from input row
    2. Look up all deepest_cats for that maincat_id in cat_ids sheet
    3. For each deepest_cat, find campaign PLA/{deepest_cat}_{cl1}
    4. Remove shop exclusions from all ad groups in that campaign (batched)

    Excel columns (same structure as uitsluiten):
    A. Shop name - shop to un-exclude
    B. Shop ID - (not used, for reference)
    C. maincat - Category name (for reference)
    D. maincat_id - Maincat ID (used to look up deepest_cats via cat_ids sheet)
    E. custom label 1 - Custom label 1 value (for campaign matching)
    F. Status - Will be updated with TRUE/FALSE
    G. Error - Will be updated with error messages

    Args:
        client: Google Ads client
        workbook: Excel workbook
        customer_id: Customer ID
        file_path: Path to save progress
        save_interval: Save every N groups processed
        sheet_name: Name of the sheet to process (default: "verwijderen")
    """
    print(f"\n{'='*70}")
    print(f"PROCESSING REVERSE EXCLUSION SHEET: '{sheet_name}'")
    print(f"(OPTIMIZED: Grouping shops by maincat_id + cl1)")
    if dry_run:
        print("(DRY RUN: no shop exclusions will actually be removed from Google Ads)")
    print(f"{'='*70}")

    # Load cat_ids mapping (same as process_exclusion_sheet_v2)
    print("\nLoading cat_ids mapping...")
    cat_ids_mapping = load_cat_ids_mapping(workbook)
    if not cat_ids_mapping:
        print("❌ No cat_ids mapping loaded, cannot process reverse exclusions")
        return

    if sheet_name not in workbook.sheetnames:
        print(f"⚠️  Sheet '{sheet_name}' not found in workbook")
        print(f"   Available sheets: {workbook.sheetnames}")
        return

    sheet = workbook[sheet_name]

    # Column indices (0-based) - same structure as uitsluiten sheet
    COL_SHOP_NAME = 0      # A: Shop name
    COL_SHOP_ID = 1        # B: Shop ID (not used)
    COL_MAINCAT = 2        # C: maincat (category name, for reference)
    COL_MAINCAT_ID = 3     # D: maincat_id (used to look up deepest_cats)
    COL_CL1 = 4            # E: custom label 1
    COL_STATUS = 5         # F: Status
    COL_ERROR = 6          # G: Error

    # Pre-fetch campaigns cache
    print("\nPre-fetching PLA campaigns and ad groups...")
    campaign_cache = prefetch_pla_campaigns_and_ad_groups(client, customer_id, "PLA/")

    # =========================================================================
    # STEP 1: Group all rows by (maincat_id, cl1) for efficient batch processing
    # =========================================================================
    print("\nGrouping rows by (maincat_id, cl1)...")

    # Structure: {(maincat_id, cl1): [(row_idx, shop_name), ...]}
    groups = defaultdict(list)
    rows_with_missing_fields = []  # Track rows with missing required fields
    maincat_name_by_id: dict = {}  # for tree log description

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        # Check if already processed
        status_value = row[COL_STATUS].value if len(row) > COL_STATUS else None
        if status_value is not None and status_value != '':
            continue

        shop_name = row[COL_SHOP_NAME].value
        maincat_name = row[COL_MAINCAT].value if len(row) > COL_MAINCAT else None
        maincat_id = row[COL_MAINCAT_ID].value
        custom_label_1 = row[COL_CL1].value

        # Skip empty rows
        if not shop_name:
            continue

        # Track rows with missing required fields
        if not maincat_id or not custom_label_1:
            rows_with_missing_fields.append(idx)
            continue

        maincat_id_str = str(maincat_id)
        cl1_str = str(custom_label_1)
        if maincat_name and maincat_id_str not in maincat_name_by_id:
            maincat_name_by_id[maincat_id_str] = str(maincat_name)
        groups[(maincat_id_str, cl1_str)].append((idx, shop_name))

    # Mark rows with missing fields as errors
    for idx in rows_with_missing_fields:
        print(f"[Row {idx}] ⚠️  Missing required fields, skipping")
        sheet.cell(row=idx, column=COL_STATUS + 1).value = False
        sheet.cell(row=idx, column=COL_ERROR + 1).value = "Missing required fields"

    total_groups = len(groups)
    total_rows = sum(len(rows) for rows in groups.values())
    print(f"Found {total_rows} row(s) in {total_groups} unique (maincat_id, cl1) group(s)")
    print(f"Rows with missing fields: {len(rows_with_missing_fields)}")

    if total_groups == 0:
        print("No rows to process.")
        return

    # =========================================================================
    # STEP 2: Process each group - fetch campaigns/ad groups ONCE per group
    # =========================================================================
    success_count = 0
    error_count = 0
    groups_processed = 0
    # Run-wide counters distinct from per-row "successful" — these tell the
    # operator what the API actually did, not just whether the workbook row
    # ended up flagged TRUE. A row is TRUE if the shop is no longer excluded
    # for ANY reason (removed OR was never excluded), so "Rows OK" can be 3
    # while "Exclusions actually removed" is 0.
    run_total_removed = 0
    run_total_already_not_excluded = 0
    run_total_mutate_errors = 0
    run_total_batch_calls = 0

    for (maincat_id_str, cl1_str), rows in groups.items():
        groups_processed += 1
        shop_names = [shop_name for _, shop_name in rows]
        row_indices = [idx for idx, _ in rows]

        # For CL3 targeting, split shop_name at | and use first part
        # e.g. "Hbm-machines.com|NL" becomes "Hbm-machines.com"
        shop_names_for_targeting = [name.split('|')[0] if '|' in name else name for name in shop_names]
        # Create mapping from targeting name back to original name(s)
        targeting_to_original = {}
        for orig, tgt in zip(shop_names, shop_names_for_targeting):
            if tgt not in targeting_to_original:
                targeting_to_original[tgt] = []
            targeting_to_original[tgt].append(orig)

        print(f"\n{'='*60}")
        print(f"[Group {groups_processed}/{total_groups}] maincat_id={maincat_id_str}, cl1={cl1_str}")
        print(f"  Shops to process: {len(shop_names)}")
        print(f"  Shop names: {', '.join(shop_names[:5])}{'...' if len(shop_names) > 5 else ''}")
        # Show if any names were split
        split_names = [(orig, tgt) for orig, tgt in zip(shop_names, shop_names_for_targeting) if orig != tgt]
        if split_names:
            print(f"  CL3 targeting (split): {', '.join([f'{tgt} (from {orig})' for orig, tgt in split_names[:3]])}{'...' if len(split_names) > 3 else ''}")

        # Look up deepest_cats for this maincat_id ONCE for the entire group
        deepest_cats = cat_ids_mapping.get(maincat_id_str, [])
        if not deepest_cats:
            print(f"  ⚠️  No deepest_cats found for maincat_id={maincat_id_str}")
            # Mark all rows in this group as failed
            for idx in row_indices:
                sheet.cell(row=idx, column=COL_STATUS + 1).value = False
                sheet.cell(row=idx, column=COL_ERROR + 1).value = f"No deepest_cats for maincat_id={maincat_id_str}"
                error_count += 1
            continue

        print(f"  Found {len(deepest_cats)} deepest_cat(s)")

        # Track results per shop
        shop_results = {shop: {'success': 0, 'not_found': 0, 'errors': []} for shop in shop_names}
        campaigns_found = 0
        total_exclusions_removed = 0

        # Process each deepest_cat ONCE for all shops in this group
        missing_campaigns: list = []
        for deepest_cat in deepest_cats:
            campaign_name = f"PLA/{deepest_cat}_{cl1_str}"
            campaign_data = campaign_cache.get(campaign_name)

            if not campaign_data:
                missing_campaigns.append(campaign_name)
                print(f"    ⚠️  Campaign not found in Google Ads cache: {campaign_name}")

            if campaign_data:
                campaigns_found += 1
                ad_groups = campaign_data['ad_groups']
                print(f"    📁 Campaign: {campaign_name} ({len(ad_groups)} ad group(s))")
                # Tree line per campaign — parseable by _parse_affected_entities
                mc_name_disp = maincat_name_by_id.get(maincat_id_str, f"maincat_id={maincat_id_str}")
                shops_disp = ", ".join(sorted(set(shop_names))[:5]) + ("..." if len(set(shop_names)) > 5 else "")
                tree_verb = "Tree to modify" if dry_run else "Tree modified"
                print(f"      🌳 {tree_verb}: Campaign '{campaign_name}' → Maincat '{mc_name_disp}' → CL1 '{cl1_str}' → Shops: {shops_disp}")

                for ag in ad_groups:
                    ag_id = str(ag['id'])
                    ag_name = ag['name']

                    # Retry logic for connection errors
                    max_retries = 3
                    retry_delay = 2

                    for attempt in range(max_retries):
                        try:
                            # Call batch function with targeting names (split at |)
                            # Use unique targeting names to avoid duplicates
                            unique_targeting_names = list(set(shop_names_for_targeting))
                            if dry_run:
                                # DRY RUN: simulate — pretend every shop's exclusion
                                # would be removed successfully.
                                result = {
                                    'success': list(unique_targeting_names),
                                    'not_found': [],
                                    'errors': [],
                                }
                            else:
                                result = reverse_exclusion_batch(
                                    client=client,
                                    customer_id=customer_id,
                                    ad_group_id=ag_id,
                                    ad_group_name=ag_name,
                                    shop_names=unique_targeting_names
                                )

                            # Aggregate results - map targeting names back to original names
                            for targeting_name in result['success']:
                                for orig_name in targeting_to_original.get(targeting_name, [targeting_name]):
                                    if orig_name in shop_results:
                                        shop_results[orig_name]['success'] += 1
                                        total_exclusions_removed += 1
                            for targeting_name in result['not_found']:
                                for orig_name in targeting_to_original.get(targeting_name, [targeting_name]):
                                    if orig_name in shop_results:
                                        shop_results[orig_name]['not_found'] += 1
                            for targeting_name, error in result['errors']:
                                for orig_name in targeting_to_original.get(targeting_name, [targeting_name]):
                                    if orig_name in shop_results:
                                        shop_results[orig_name]['errors'].append(f"{ag_name}: {error}")

                            # Per-ad-group status line (parser keys on "<indent> <icon> PLA/...:")
                            s_count = len(result['success'])
                            nf_count = len(result['not_found'])
                            err_count = len(result['errors'])
                            run_total_batch_calls += 1
                            run_total_removed += s_count
                            run_total_already_not_excluded += nf_count
                            run_total_mutate_errors += err_count
                            if err_count > 0:
                                print(f"      ❌ {ag_name}: {err_count} error(s), {s_count} removed, {nf_count} not found")
                                # Print each error in full so transient API problems are visible
                                for shop_e, msg_e in result['errors'][:5]:
                                    print(f"         · {shop_e}: {msg_e}")
                            elif s_count > 0:
                                print(f"      ✅ {ag_name}: {s_count} removed, {nf_count} not found")
                            else:
                                print(f"      ⏭️  {ag_name}: all {nf_count} not found (nothing to remove)")

                            break  # Success, exit retry loop

                        except Exception as e:
                            error_str = str(e)
                            if "failed to connect" in error_str.lower() or "unavailable" in error_str.lower():
                                if attempt < max_retries - 1:
                                    print(f"    ⚠️  Connection error, retrying in {retry_delay}s...")
                                    time.sleep(retry_delay)
                                    retry_delay *= 2
                                    continue
                            # Non-retryable error or max retries reached
                            error_msg = str(e)[:50]
                            print(f"    ❌ {ag_name}: {error_msg}")
                            for shop in shop_names:
                                shop_results[shop]['errors'].append(f"{ag_name}: {error_msg}")
                            break

                    # Rate limiting delay after each ad group (not each shop!)
                    time.sleep(0.3)

        print(f"  Found {campaigns_found} campaign(s), removed {total_exclusions_removed} exclusion(s) total")

        # =========================================================================
        # STEP 3: Update row statuses based on results
        # =========================================================================
        for idx, shop_name in rows:
            result = shop_results[shop_name]

            # Consider success if: at least one exclusion removed OR not found (wasn't excluded)
            # AND no errors occurred
            has_errors = len(result['errors']) > 0
            has_activity = result['success'] > 0 or result['not_found'] > 0

            if campaigns_found == 0:
                # No campaigns found at all - this is an error
                sheet.cell(row=idx, column=COL_STATUS + 1).value = False
                sheet.cell(row=idx, column=COL_ERROR + 1).value = f"No campaigns found for maincat_id={maincat_id_str}"
                error_count += 1
                print(f"    Row {idx} ({shop_name}): ❌ No campaigns")
            elif has_errors:
                sheet.cell(row=idx, column=COL_STATUS + 1).value = False
                error_summary = "; ".join(result['errors'][:3])
                sheet.cell(row=idx, column=COL_ERROR + 1).value = error_summary[:100]
                error_count += 1
                print(f"    Row {idx} ({shop_name}): ❌ {len(result['errors'])} error(s)")
            else:
                sheet.cell(row=idx, column=COL_STATUS + 1).value = True
                sheet.cell(row=idx, column=COL_ERROR + 1).value = ""
                success_count += 1
                print(f"    Row {idx} ({shop_name}): ✅ removed={result['success']}, not_found={result['not_found']}")

        # Save periodically (every N groups)
        if file_path and groups_processed % save_interval == 0:
            print(f"\n💾 Saving progress ({groups_processed} groups processed)...")
            try:
                workbook.save(file_path)
            except Exception as save_error:
                print(f"⚠️  Error saving: {save_error}")

    # Final save
    if file_path:
        print(f"\n💾 Final save...")
        try:
            workbook.save(file_path)
        except Exception as save_error:
            print(f"⚠️  Error on final save: {save_error}")

    print(f"\n{'='*70}")
    print(f"REVERSE EXCLUSION SHEET SUMMARY (OPTIMIZED)")
    print(f"{'='*70}")
    print(f"Total groups processed: {groups_processed}")
    print(f"Total rows processed: {success_count + error_count}")
    print(f"Rows with missing fields: {len(rows_with_missing_fields)}")
    print(f"✅ Rows OK: {success_count}  (row didn't error — shop is no longer excluded for any reason)")
    print(f"❌ Rows failed: {error_count + len(rows_with_missing_fields)}")
    print()
    print(f"Batch calls made (1 per ad group): {run_total_batch_calls}")
    print(f"  → Exclusions actually removed: {run_total_removed}")
    print(f"  → Already not excluded (no-op): {run_total_already_not_excluded}")
    print(f"  → Mutate errors: {run_total_mutate_errors}")
    if run_total_batch_calls > 0 and run_total_removed == 0:
        print()
        print("⚠️  No exclusions were actually removed. Either the shops were already")
        print("    not excluded in any of these ad groups, or every match attempt")
        print("    fell through (check INDEX/value casing in the listing tree).")
    print(f"{'='*70}\n")

# ============================================================================
# PLA STORE CAMPAIGNS - DATAEDIS CL0 EXCLUSION
# ============================================================================

def add_dataedis_exclusion_to_pla_store_campaigns(client, customer_id, dry_run=True):
    """
    For all campaigns starting with 'PLA/' and containing 'store_':
    - Get all ad groups
    - For each ad group, add CL0 = 'dataedis' as a NEGATIVE exclusion in the listing tree

    The exclusion is added at the deepest level (under the CL1 subdivision).
    If CL1 is currently a UNIT, it is converted to a SUBDIVISION first.

    Args:
        client: Google Ads client
        customer_id: Customer ID
        dry_run: If True, only log what would happen without making changes
    """
    ga_service = client.get_service("GoogleAdsService")

    # Step 1: Find all matching campaigns
    campaign_query = """
        SELECT campaign.id, campaign.name, campaign.resource_name
        FROM campaign
        WHERE campaign.name LIKE 'PLA/%'
        AND campaign.name LIKE '%store_%'
        AND campaign.status != 'REMOVED'
        ORDER BY campaign.name
    """
    campaigns = list(ga_service.search(customer_id=customer_id, query=campaign_query))

    print(f"\nFound {len(campaigns)} PLA/store_ campaigns")
    if dry_run:
        print("⚠️  DRY RUN MODE - no changes will be made\n")

    total_ad_groups = 0
    total_modified = 0
    total_skipped = 0
    total_errors = 0

    for camp_row in campaigns:
        campaign_name = camp_row.campaign.name
        campaign_resource_name = camp_row.campaign.resource_name

        # Step 2: Get all ad groups in this campaign
        escaped_campaign_rn = campaign_resource_name.replace("'", "\\'")
        ag_query = f"""
            SELECT ad_group.id, ad_group.name
            FROM ad_group
            WHERE ad_group.campaign = '{escaped_campaign_rn}'
            AND ad_group.status != 'REMOVED'
        """
        ad_groups = list(ga_service.search(customer_id=customer_id, query=ag_query))

        if not ad_groups:
            continue

        print(f"\n{'='*60}")
        print(f"Campaign: {campaign_name} ({len(ad_groups)} ad groups)")
        print(f"{'='*60}")

        for ag_row in ad_groups:
            ad_group_id = ag_row.ad_group.id
            ad_group_name = ag_row.ad_group.name
            total_ad_groups += 1

            print(f"\n  Ad group: {ad_group_name} (ID: {ad_group_id})")

            try:
                result = _add_cl0_exclusion_to_ad_group(
                    client, customer_id, ad_group_id, "dataedis", dry_run=dry_run
                )
                if result == "modified":
                    total_modified += 1
                elif result == "skipped":
                    total_skipped += 1
            except Exception as e:
                print(f"    ❌ Error: {e}")
                total_errors += 1

    # Summary
    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"{'='*60}")
    print(f"Campaigns processed: {len(campaigns)}")
    print(f"Ad groups processed: {total_ad_groups}")
    print(f"Ad groups modified:  {total_modified}")
    print(f"Ad groups skipped:   {total_skipped}")
    print(f"Errors:              {total_errors}")
    if dry_run:
        print(f"\n⚠️  DRY RUN - no changes were made. Set dry_run=False to execute.")


def _add_cl0_exclusion_to_ad_group(client, customer_id, ad_group_id, exclusion_value, dry_run=True):
    """
    Add a CL0 exclusion to an ad group's listing tree.

    Finds the deepest positive leaf UNIT nodes (the actual targeting nodes with bids)
    and adds CL0 children under each one. Works regardless of tree depth or dimension types.

    For each leaf UNIT found:
    - If already has CL0 children with the exclusion: skip
    - If UNIT: Convert to SUBDIVISION and add CL0=OTHERS (POS, bid) + CL0=exclusion (NEG)

    Args:
        client: Google Ads client
        customer_id: Customer ID
        ad_group_id: Ad group ID
        exclusion_value: The CL0 value to exclude (e.g. "dataedis")
        dry_run: If True, only log without making changes

    Returns:
        "modified" if a change was made (or would be made in dry run)
        "skipped" if no change was needed
    """
    ga_service = client.get_service("GoogleAdsService")
    ag_service = client.get_service("AdGroupService")
    agc_service = client.get_service("AdGroupCriterionService")
    ag_path = ag_service.ad_group_path(customer_id, ad_group_id)

    # Query the full listing tree
    query = f"""
        SELECT
            ad_group_criterion.resource_name,
            ad_group_criterion.criterion_id,
            ad_group_criterion.listing_group.type,
            ad_group_criterion.listing_group.parent_ad_group_criterion,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.index,
            ad_group_criterion.listing_group.case_value.product_custom_attribute.value,
            ad_group_criterion.listing_group.case_value.product_item_id.value,
            ad_group_criterion.negative,
            ad_group_criterion.cpc_bid_micros
        FROM ad_group_criterion
        WHERE ad_group_criterion.ad_group = '{ag_path}'
            AND ad_group_criterion.type = 'LISTING_GROUP'
            AND ad_group_criterion.status != 'REMOVED'
    """

    results = list(ga_service.search(customer_id=customer_id, query=query))

    if not results:
        print(f"    ⚠️  No listing tree found - skipping")
        return "skipped"

    # Parse all nodes
    nodes = []
    for row in results:
        lg = row.ad_group_criterion.listing_group
        case_value = lg.case_value

        dim_type = None
        dim_value = None
        dim_index = None  # Store the enum for recreating the dimension

        if case_value.product_item_id.value:
            dim_type = "ITEM_ID"
            dim_value = case_value.product_item_id.value
        elif case_value.product_custom_attribute.index:
            dim_type = case_value.product_custom_attribute.index.name
            dim_index = case_value.product_custom_attribute.index
            dim_value = case_value.product_custom_attribute.value or None

        nodes.append({
            'resource_name': row.ad_group_criterion.resource_name,
            'type': lg.type_.name,
            'parent': lg.parent_ad_group_criterion or None,
            'dim_type': dim_type,
            'dim_index': dim_index,
            'dim_value': dim_value,
            'negative': row.ad_group_criterion.negative,
            'bid': row.ad_group_criterion.cpc_bid_micros
        })

    # Build set of resource names that are parents (i.e. have children)
    parent_resources = set(n['parent'] for n in nodes if n['parent'])

    # Find the deepest positive leaf UNIT nodes with a bid (the actual targeting nodes)
    # These are UNIT nodes that:
    # - Are not negative (positive targeting)
    # - Have a bid (actual targeting with bid)
    # - Have a dimension value (not OTHERS)
    leaf_targets = [
        n for n in nodes
        if n['type'] == 'UNIT'
        and not n['negative']
        and n['bid']
        and n['dim_value']
    ]

    if not leaf_targets:
        print(f"    ⚠️  No positive leaf UNIT nodes with bid found - skipping")
        return "skipped"

    # Also find SUBDIVISION nodes that already have CL0 children (for the "already exists" check)
    # Check if any parent node already has CL0=exclusion_value as a child
    modified = False

    for leaf in leaf_targets:
        leaf_resource = leaf['resource_name']
        leaf_parent = leaf['parent']
        leaf_dim_type = leaf['dim_type']
        leaf_dim_value = leaf['dim_value']
        leaf_dim_index = leaf['dim_index']
        leaf_bid = leaf['bid'] or DEFAULT_BID_MICROS

        # Check if this leaf's parent already has CL0 children (meaning exclusion might already be at this level)
        # This happens when the leaf is already a SUBDIVISION with CL0 children
        # But since leaf is a UNIT, it can't have children. Check if its siblings are CL0.
        siblings = [n for n in nodes if n['parent'] == leaf_parent and n['resource_name'] != leaf_resource]
        sibling_dim_types = set(s['dim_type'] for s in siblings if s['dim_type'])

        # If siblings are CL0 (INDEX0), this leaf is already at CL0 level - check if exclusion exists
        if 'INDEX0' in sibling_dim_types and leaf_dim_type == 'INDEX0':
            already_excluded = any(
                s['dim_type'] == 'INDEX0' and
                s['dim_value'] == exclusion_value and
                s['negative']
                for s in siblings
            )
            if already_excluded:
                print(f"    ✅ CL0='{exclusion_value}' already excluded at this level - skipping")
                continue

        # Convert this leaf UNIT to SUBDIVISION and add CL0 children
        dim_desc = f"{leaf_dim_type}={leaf_dim_value}"

        if dry_run:
            print(f"    🔍 [DRY RUN] Would convert {dim_desc} UNIT to SUBDIVISION, add CL0='{exclusion_value}' exclusion (bid: €{leaf_bid/1_000_000:.2f})")
            modified = True
            continue

        ops = []

        # 1. REMOVE the old leaf UNIT
        remove_op = client.get_type("AdGroupCriterionOperation")
        remove_op.remove = leaf_resource
        ops.append(remove_op)

        # 2. CREATE new SUBDIVISION with same dimension (same parent, same case value)
        dim_info = client.get_type("ListingDimensionInfo")
        if leaf_dim_type == 'ITEM_ID':
            dim_info.product_item_id.value = leaf_dim_value
        elif leaf_dim_index is not None:
            dim_info.product_custom_attribute.index = leaf_dim_index
            if leaf_dim_value:
                dim_info.product_custom_attribute.value = leaf_dim_value

        subdiv_op = create_listing_group_subdivision(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=leaf_parent,
            listing_dimension_info=dim_info
        )
        subdiv_tmp = subdiv_op.create.resource_name
        ops.append(subdiv_op)

        # 3. CREATE CL0 OTHERS (positive, with original bid) - required for subdivision
        dim_cl0_others = client.get_type("ListingDimensionInfo")
        dim_cl0_others.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0

        ops.append(create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=subdiv_tmp,
            listing_dimension_info=dim_cl0_others,
            targeting_negative=False,
            cpc_bid_micros=leaf_bid
        ))

        # 4. CREATE CL0 = exclusion_value (negative)
        dim_cl0_excl = client.get_type("ListingDimensionInfo")
        dim_cl0_excl.product_custom_attribute.index = client.enums.ProductCustomAttributeIndexEnum.INDEX0
        dim_cl0_excl.product_custom_attribute.value = exclusion_value

        ops.append(create_listing_group_unit_biddable(
            client=client,
            customer_id=customer_id,
            ad_group_id=str(ad_group_id),
            parent_ad_group_criterion_resource_name=subdiv_tmp,
            listing_dimension_info=dim_cl0_excl,
            targeting_negative=True,
            cpc_bid_micros=None
        ))

        # Execute all in one atomic mutate (removes processed before creates)
        agc_service.mutate_ad_group_criteria(customer_id=customer_id, operations=ops)
        print(f"    ✅ Converted {dim_desc} to subdivision, added CL0='{exclusion_value}' exclusion")
        modified = True

    return "modified" if modified else "skipped"


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """
    Main execution function.
    """
    print(f"\n{'='*70}")
    print("DMA SHOP CAMPAIGNS PROCESSOR")
    print(f"{'='*70}")
    print(f"Country: {COUNTRY}")
    print(f"Operating System: {platform.system()}")
    print(f"Customer ID: {CUSTOMER_ID}")
    print(f"Excel File: {EXCEL_FILE_PATH}")
    print(f"{'='*70}\n")

    # Initialize Google Ads client
    client = initialize_google_ads_client()
    #add_dataedis_exclusion_to_pla_store_campaigns(client, CUSTOMER_ID, False)

    # Create a working copy of the Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    working_copy_path = EXCEL_FILE_PATH.replace(".xlsx", f"_working_copy_{timestamp}.xlsx")

    print(f"\n{'='*70}")
    print(f"CREATING WORKING COPY")
    print(f"{'='*70}")
    print(f"Original file: {EXCEL_FILE_PATH}")
    print(f"Working copy:  {working_copy_path}")

    try:
        shutil.copy2(EXCEL_FILE_PATH, working_copy_path)
        print(f"✅ Working copy created successfully")
    except Exception as e:
        print(f"❌ Error creating working copy: {e}")
        sys.exit(1)

    # Load Excel workbook from working copy
    print(f"\n{'='*70}")
    print(f"LOADING WORKING COPY")
    print(f"{'='*70}")
    print(f"Loading: {working_copy_path}")
    try:
        workbook = load_workbook(working_copy_path)
        print(f"✅ Excel file loaded successfully")
        print(f"   Available sheets: {workbook.sheetnames}")
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        sys.exit(1)

    ''' 
    # Process exclusion sheet (V2 - with cat_ids mapping)
    try:
        process_exclusion_sheet_v2(client, workbook, CUSTOMER_ID, working_copy_path)
    except Exception as e:
        print(f"❌ Error processing exclusion sheet: {e}")
            
    # Load reverse exclusion file separately
    print(f"\n{'='*70}")
    print(f"LOADING REVERSE EXCLUSION FILE")
    print(f"{'='*70}")
    print(f"File: {REVERSE_EXCLUSION_FILE_PATH}")
    '''

    reverse_working_copy_path = REVERSE_EXCLUSION_FILE_PATH.replace(".xlsx", f"_working_copy_{timestamp}.xlsx")

    try:
        shutil.copy2(REVERSE_EXCLUSION_FILE_PATH, reverse_working_copy_path)
        print(f"✅ Reverse exclusion working copy created: {reverse_working_copy_path}")
        reverse_workbook = load_workbook(reverse_working_copy_path)
        print(f"✅ Reverse exclusion file loaded successfully")
        print(f"   Available sheets: {reverse_workbook.sheetnames}")

        #process_inclusion_sheet_v2(client, reverse_workbook, CUSTOMER_ID, reverse_working_copy_path)
        #process_exclusion_sheet_new(client, reverse_workbook, CUSTOMER_ID, reverse_working_copy_path)
        process_exclusion_sheet_v2(client, reverse_workbook, CUSTOMER_ID, reverse_working_copy_path)

        #process_check_sheet(client, reverse_workbook, CUSTOMER_ID, reverse_working_copy_path)
        #process_check_cl1_sheet(client, reverse_workbook, CUSTOMER_ID, reverse_working_copy_path)
        #process_check_new_sheet(client, reverse_workbook, CUSTOMER_ID, reverse_working_copy_path)
        #process_reverse_inclusion_sheet_v2(client, reverse_workbook, CUSTOMER_ID, reverse_working_copy_path)
        #process_reverse_exclusion_sheet(client, reverse_workbook, CUSTOMER_ID, reverse_working_copy_path, 10, "verwijderen")
        #process_enable_inclusion_sheet_v2(client, reverse_workbook, CUSTOMER_ID, reverse_working_copy_path, "toevoegen")
        #process_pause_ad_groups_sheet(client, reverse_workbook, CUSTOMER_ID, reverse_working_copy_path)

        # Save reverse exclusion workbook
        reverse_workbook.save(reverse_working_copy_path)
        print(f"✅ Reverse exclusion results saved to: {reverse_working_copy_path}")
    except FileNotFoundError:
        print(f"⚠️  Reverse exclusion file not found: {REVERSE_EXCLUSION_FILE_PATH}")
        print(f"   Skipping reverse exclusion processing")
    except Exception as e:
        print(f"❌ Error processing reverse exclusion file: {e}")


    # Validate cl1 targeting (Dry run)
    #validate_cl1_targeting_for_campaigns(client, CUSTOMER_ID, "% store_%", False)
    #validate_ads_for_campaigns(client, CUSTOMER_ID, "% store_%", True)

    # Validate listing trees exist for reverse campaigns (dry_run=True to preview)
    #validate_listing_trees_for_campaigns(client, CUSTOMER_ID, "PLA/%", dry_run=True)

    '''
    # Process inclusion sheet (V2 - new structure)
    try:
        process_inclusion_sheet_legacy(client, workbook, CUSTOMER_ID, working_copy_path)
    except Exception as e:
        print(f"❌ Error processing inclusion sheet: {e}")
        
    # Process uitbreiding sheet
    try:
       process_uitbreiding_sheet(client, workbook, CUSTOMER_ID, working_copy_path)
    except Exception as e:
       print(f"❌ Error processing uitbreiding sheet: {e}")
    '''

    # Final save to working copy
    print(f"\n{'='*70}")
    print("SAVING FINAL RESULTS")
    print(f"{'='*70}")
    print(f"All results saved to working copy: {working_copy_path}")
    print(f"Original file remains unchanged: {EXCEL_FILE_PATH}")
    print(f"\nTo use the results, rename or copy the working copy to:")
    print(f"  {EXCEL_FILE_PATH}")
    print(f"{'='*70}")

    print(f"\n{'='*70}")
    print("PROCESSING COMPLETE")
    print(f"{'='*70}\n")

if __name__ == "__main__":
    main()
