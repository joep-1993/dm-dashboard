"""
DMA+ Monthly Delta + Category Coverage (ported from dma_script).

Adds two capabilities to the existing DMA+ dashboard:

1. Monthly Delta — process a multi-sheet xlsx (one sheet per country × {Nieuw, Afvallers})
   with 3 columns per row (shop / maincat / maincat_id). Each row fans out to
   cl1 ∈ {a,b,c}. Per country the flow is:
     Nieuw      → Include shops → Exclude shops
     Afvallers  → Reverse exclude → Reverse include

2. Category Coverage — for a given country, write one row per taxv2 category
   with TRUE/FALSE flags (columns A/B/C) for whether PLA/{name}_{cl1} exists
   in Google Ads. Helpful for spotting name mismatches.

This module intentionally does not modify backend/campaign_processor.py — it
builds a fresh workbook with the conventional sheet name (`toevoegen`,
`uitsluiten`, `verwijderen`) for each operation and calls the existing v2
processors unchanged.
"""

from __future__ import annotations

import io
import os
import sys
import time
import uuid
import logging
import threading
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

from backend.dma_plus_service import (
    COUNTRY_CONFIG,
    TaskCancelled,
    _check_cancelled,
    _get_client,
    _get_task,
    _patch_campaign_processor,
    _populate_cat_ids_sheet,
    _set_task,
    _history_append,
    _extract_results,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants — matches the DMA+ delta file layout
# ---------------------------------------------------------------------------
DMA_PLUS_CL1_VALUES = ("a", "b", "c")
DMA_PLUS_DEFAULT_BUDGET = 50.0

DMA_PLUS_SHEETS = {
    "NL": {"nieuw": "NL - Nieuw (aanmaken)", "afvallers": "NL - Afvallers"},
    "BE": {"nieuw": "BE - Nieuw (aanmaken)", "afvallers": "BE - Afvallers"},
}

TAXV2_BASE_URL = "http://producttaxonomyunifiedapi-prod.azure.api.beslist.nl"

# Output files go here so the router can hand them to FileResponse.
OUTPUT_DIR = Path("/tmp/dma-plus-output")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Caches
_TAXV2_TREE_CACHE: dict = {}


# ---------------------------------------------------------------------------
# Taxonomy v2 — full tree via parallel BFS
# ---------------------------------------------------------------------------
def _fetch_taxv2_tree(locale: str = "nl-NL", force_refresh: bool = False) -> dict:
    """
    BFS-crawl the Taxonomy v2 API — the flat /api/Categories endpoint only
    returns root categories, so we traverse /api/Categories/{id}?includeSubCategories=true
    for every node using a thread pool.

    Returns:
      id_to_parent, id_to_name, children_of, leaves_of, id_order
    """
    import urllib.request, json as _json

    if not force_refresh and locale in _TAXV2_TREE_CACHE:
        return _TAXV2_TREE_CACHE[locale]

    def pick_name(labels):
        if not labels:
            return None
        for lab in labels:
            if lab.get("locale") == locale and lab.get("name"):
                return lab["name"]
        return labels[0].get("name")

    def fetch_json(url):
        req = urllib.request.Request(url, headers={"Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            return _json.loads(resp.read().decode("utf-8"))

    def fetch_detail(cid):
        u = f"{TAXV2_BASE_URL}/api/Categories/{cid}?includeSubCategories=true&includeFacets=false"
        try:
            return cid, fetch_json(u)
        except Exception:
            try:
                return cid, fetch_json(u)
            except Exception as e2:
                return cid, {"__error__": str(e2)}

    try:
        roots = fetch_json(f"{TAXV2_BASE_URL}/api/Categories?locale={locale}")
    except Exception as e:
        logger.error(f"Failed to fetch taxonomy roots: {e}")
        return {}

    if not isinstance(roots, list):
        logger.error(f"Unexpected taxonomy response type: {type(roots).__name__}")
        return {}

    id_to_parent: dict = {}
    id_to_name: dict = {}
    children_of = defaultdict(list)
    id_order: list = []

    for cat in roots:
        cid = cat.get("id")
        if cid is None:
            continue
        id_to_parent[cid] = cat.get("parentId")
        id_to_name[cid] = pick_name(cat.get("labels")) or str(cid)
        id_order.append(cid)

    frontier = list(id_to_parent.keys())
    fetch_errors = 0

    with ThreadPoolExecutor(max_workers=12) as ex:
        while frontier:
            results = list(ex.map(fetch_detail, frontier))
            next_frontier = []
            for cid, detail in results:
                if "__error__" in detail:
                    fetch_errors += 1
                    continue
                for sub in (detail.get("subCategories") or []):
                    sid = sub.get("id")
                    if sid is None or sid in id_to_parent:
                        continue
                    id_to_parent[sid] = sub.get("parentId", cid)
                    id_to_name[sid] = pick_name(sub.get("labels")) or str(sid)
                    children_of[cid].append(sid)
                    id_order.append(sid)
                    next_frontier.append(sid)
            frontier = next_frontier

    leaves_of: dict = {}

    def compute_leaves(cid):
        if cid in leaves_of:
            return leaves_of[cid]
        kids = children_of.get(cid, [])
        if not kids:
            leaves_of[cid] = {id_to_name.get(cid, str(cid))}
            return leaves_of[cid]
        acc = set()
        for k in kids:
            acc |= compute_leaves(k)
        leaves_of[cid] = acc
        return acc

    for cid in id_order:
        compute_leaves(cid)

    tree = {
        "id_to_parent": id_to_parent,
        "id_to_name": id_to_name,
        "children_of": dict(children_of),
        "leaves_of": leaves_of,
        "id_order": id_order,
    }
    _TAXV2_TREE_CACHE[locale] = tree
    logger.info(f"Crawled {len(id_order)} taxv2 categories (errors={fetch_errors})")
    return tree


def _fetch_pla_campaign_names(client, customer_id: str) -> set:
    """Query the 'campaign' resource directly so campaigns without ad groups still count."""
    ga = client.get_service("GoogleAdsService")
    q = "SELECT campaign.name FROM campaign WHERE campaign.name LIKE 'PLA/%' AND campaign.status != 'REMOVED'"
    names: set = set()
    try:
        for row in ga.search(customer_id=customer_id, query=q):
            if row.campaign.name:
                names.add(row.campaign.name)
    except Exception as e:
        logger.error(f"Error fetching PLA campaigns: {e}")
    return names


# ---------------------------------------------------------------------------
# Synthetic workbook builders (one per operation)
#
# Each produces a workbook with the *conventional* sheet name that the v2
# processor already reads, so backend/campaign_processor.py doesn't need any
# changes. Source is a 3-col sheet: A=shop, B=maincat, C=maincat_id.
# Every source row fans out to len(cl1_values) rows.
# ---------------------------------------------------------------------------
def _iter_source_rows(src_ws):
    """Yield (src_row_idx, shop, maincat, maincat_id) for non-empty rows."""
    for src_idx, row in enumerate(src_ws.iter_rows(min_row=2, values_only=True), start=2):
        shop = row[0] if len(row) > 0 else None
        maincat = row[1] if len(row) > 1 else None
        maincat_id = row[2] if len(row) > 2 else None
        if not shop or not maincat or maincat_id in (None, ""):
            continue
        yield src_idx, shop, maincat, maincat_id


def _build_inclusion_workbook_from_source(src_ws, budget: float) -> tuple:
    """Build 'toevoegen' sheet. Returns (workbook, list_of_source_row_indices)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "toevoegen"
    ws.append(["shop_name", "Shop ID", "maincat", "maincat_id",
               "custom label 1", "budget", "result", "error message"])
    src_rows: List[int] = []
    for src_idx, shop, maincat, maincat_id in _iter_source_rows(src_ws):
        for cl1 in DMA_PLUS_CL1_VALUES:
            ws.append([shop, "", maincat, maincat_id, cl1, budget, None, None])
            src_rows.append(src_idx)
    return wb, src_rows


def _build_exclusion_workbook_from_source(src_ws) -> tuple:
    """Build 'uitsluiten' + 'cat_ids' sheets. Returns (workbook, src_rows)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "uitsluiten"
    ws.append(["shop_name", "Shop ID", "maincat", "maincat_id",
               "custom label 1", "result", "error message"])
    src_rows: List[int] = []
    for src_idx, shop, maincat, maincat_id in _iter_source_rows(src_ws):
        for cl1 in DMA_PLUS_CL1_VALUES:
            ws.append([shop, "", maincat, maincat_id, cl1, None, None])
            src_rows.append(src_idx)
    _populate_cat_ids_sheet(wb)
    return wb, src_rows


def _build_reverse_exclusion_workbook_from_source(src_ws) -> tuple:
    """Build 'verwijderen' + 'cat_ids' sheets."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "verwijderen"
    ws.append(["shop_name", "Shop ID", "maincat", "maincat_id",
               "custom label 1", "result", "error message"])
    src_rows: List[int] = []
    for src_idx, shop, maincat, maincat_id in _iter_source_rows(src_ws):
        for cl1 in DMA_PLUS_CL1_VALUES:
            ws.append([shop, "", maincat, maincat_id, cl1, None, None])
            src_rows.append(src_idx)
    _populate_cat_ids_sheet(wb)
    return wb, src_rows


def _build_reverse_inclusion_workbook_from_source(src_ws) -> tuple:
    """Build 'toevoegen' sheet — same layout as inclusion, reverse processor removes ad groups."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "toevoegen"
    ws.append(["shop_name", "Shop ID", "maincat", "maincat_id",
               "custom label 1", "budget", "result", "error message"])
    src_rows: List[int] = []
    for src_idx, shop, maincat, maincat_id in _iter_source_rows(src_ws):
        for cl1 in DMA_PLUS_CL1_VALUES:
            ws.append([shop, "", maincat, maincat_id, cl1, DMA_PLUS_DEFAULT_BUDGET, None, None])
            src_rows.append(src_idx)
    return wb, src_rows


# ---------------------------------------------------------------------------
# Task state helpers
# ---------------------------------------------------------------------------
def _append_log(task_id: str, line: str, progress: Optional[int] = None, message: Optional[str] = None):
    t = _get_task(task_id) or {}
    log = t.get("log") or []
    log.append(line)
    # Cap log to last 500 lines to avoid unbounded memory.
    if len(log) > 500:
        log = log[-500:]
    t["log"] = log
    if progress is not None:
        t["progress"] = progress
    if message is not None:
        t["message"] = message
    _set_task(task_id, t)


def _record_errors(task_id: str, errors: list):
    t = _get_task(task_id) or {}
    existing = t.get("errors") or []
    existing.extend(errors)
    t["errors"] = existing
    _set_task(task_id, t)


# ---------------------------------------------------------------------------
# Monthly Delta orchestrator
# ---------------------------------------------------------------------------
def _run_one_operation(task_id, op_label, wb, src_rows, country, source_sheet,
                       processor_call, extract_cols):
    """
    Generic runner:
      - patches campaign_processor to the country
      - captures stdout to log
      - calls processor_call(client, wb, customer_id)
      - extracts results from the conventional sheet and records errors
    extract_cols = (sheet_name, result_col_0based, error_col_0based)
    """
    _check_cancelled(task_id)
    _patch_campaign_processor(country)
    client = _get_client()
    customer_id = COUNTRY_CONFIG[country]["customer_id"]

    sheet_name, result_col, error_col = extract_cols

    old_stdout = sys.stdout
    captured = io.StringIO()
    sys.stdout = captured
    try:
        processor_call(client, wb, customer_id)
    finally:
        sys.stdout = old_stdout

    full_log = captured.getvalue()
    for line in full_log.splitlines():
        _append_log(task_id, line)

    results = _extract_results(wb, sheet_name, result_col, error_col)

    # Map synth row → source row
    errors = []
    rows_with_status = [r for r in results]
    for i, r in enumerate(rows_with_status):
        if r.get("success"):
            continue
        src_row = src_rows[i] if i < len(src_rows) else None
        errors.append({
            "country": country,
            "operation": op_label,
            "source_sheet": source_sheet,
            "source_row": src_row,
            "shop": r.get("shop_name") or r.get("shop") or "",
            "maincat": r.get("maincat") or "",
            "cl1": r.get("cl1") or "",
            "error": r.get("error") or "",
        })
    if errors:
        _record_errors(task_id, errors)

    return len(results), len(errors)


def run_monthly_delta(task_id: str, wb_bytes: bytes, dry_run: bool = False):
    """Background-thread target for the monthly delta flow."""
    try:
        _set_task(task_id, {
            **(_get_task(task_id) or {}),
            "status": "running",
            "progress": 2,
            "message": "Loading workbook..." + (" [DRY RUN]" if dry_run else ""),
            "log": [],
            "errors": [],
            "dry_run": dry_run,
        })

        src_wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)
        _append_log(task_id, f"Source sheets: {src_wb.sheetnames}")
        if dry_run:
            _append_log(task_id, "DRY RUN: no mutations will be sent to Google Ads")

        from backend import campaign_processor as cp

        overall_progress = 5
        summary: dict = {}

        for country in ("NL", "BE"):
            _check_cancelled(task_id)
            _append_log(task_id, f"==== Country: {country} ====", progress=overall_progress,
                        message=f"Processing {country}..." + (" [DRY RUN]" if dry_run else ""))
            summary[country] = {}
            sheets = DMA_PLUS_SHEETS[country]

            nieuw_name = sheets["nieuw"]
            afvallers_name = sheets["afvallers"]

            if nieuw_name in src_wb.sheetnames:
                nieuw_ws = src_wb[nieuw_name]

                # INCLUDE
                _check_cancelled(task_id)
                _append_log(task_id, f"[{country}] Include ({nieuw_name})",
                            message=f"{country} Include...")
                wb_inc, src_inc = _build_inclusion_workbook_from_source(nieuw_ws, DMA_PLUS_DEFAULT_BUDGET)
                if len(src_inc) > 0:
                    n_rows, n_err = _run_one_operation(
                        task_id, "include", wb_inc, src_inc, country, nieuw_name,
                        lambda c, w, cid: cp.process_inclusion_sheet_v2(c, w, cid, dry_run=dry_run),
                        ("toevoegen", 6, 7),
                    )
                    summary[country]["include"] = {"rows": n_rows, "errors": n_err}
                else:
                    _append_log(task_id, f"[{country}] Include: no rows")
                    summary[country]["include"] = {"rows": 0, "errors": 0}

                # EXCLUDE
                _check_cancelled(task_id)
                _append_log(task_id, f"[{country}] Exclude ({nieuw_name})",
                            message=f"{country} Exclude...")
                wb_exc, src_exc = _build_exclusion_workbook_from_source(nieuw_ws)
                if len(src_exc) > 0:
                    n_rows, n_err = _run_one_operation(
                        task_id, "exclude", wb_exc, src_exc, country, nieuw_name,
                        lambda c, w, cid: cp.process_exclusion_sheet_v2(c, w, cid, dry_run=dry_run),
                        ("uitsluiten", 5, 6),
                    )
                    summary[country]["exclude"] = {"rows": n_rows, "errors": n_err}
                else:
                    summary[country]["exclude"] = {"rows": 0, "errors": 0}
            else:
                _append_log(task_id, f"[{country}] '{nieuw_name}' sheet missing — skipping Include/Exclude")

            if afvallers_name in src_wb.sheetnames:
                afv_ws = src_wb[afvallers_name]

                # REVERSE EXCLUDE
                _check_cancelled(task_id)
                _append_log(task_id, f"[{country}] Reverse-exclude ({afvallers_name})",
                            message=f"{country} Reverse exclude...")
                wb_rex, src_rex = _build_reverse_exclusion_workbook_from_source(afv_ws)
                if len(src_rex) > 0:
                    n_rows, n_err = _run_one_operation(
                        task_id, "reverse_exclude", wb_rex, src_rex, country, afvallers_name,
                        lambda c, w, cid: cp.process_reverse_exclusion_sheet(c, w, cid, dry_run=dry_run),
                        ("verwijderen", 5, 6),
                    )
                    summary[country]["reverse_exclude"] = {"rows": n_rows, "errors": n_err}
                else:
                    summary[country]["reverse_exclude"] = {"rows": 0, "errors": 0}

                # REVERSE INCLUDE
                _check_cancelled(task_id)
                _append_log(task_id, f"[{country}] Reverse-include ({afvallers_name})",
                            message=f"{country} Reverse include...")
                wb_rin, src_rin = _build_reverse_inclusion_workbook_from_source(afv_ws)
                if len(src_rin) > 0:
                    n_rows, n_err = _run_one_operation(
                        task_id, "reverse_include", wb_rin, src_rin, country, afvallers_name,
                        lambda c, w, cid: cp.process_reverse_inclusion_sheet_v2(c, w, cid, dry_run=dry_run),
                        ("toevoegen", 6, 7),
                    )
                    summary[country]["reverse_include"] = {"rows": n_rows, "errors": n_err}
                else:
                    summary[country]["reverse_include"] = {"rows": 0, "errors": 0}
            else:
                _append_log(task_id, f"[{country}] '{afvallers_name}' sheet missing — skipping Afvallers")

            overall_progress = min(95, overall_progress + 45)
            _set_task(task_id, {**(_get_task(task_id) or {}), "progress": overall_progress})

        # Write errors report xlsx
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = OUTPUT_DIR / f"dma_plus_monthly_{ts}.xlsx"
        out_wb = openpyxl.Workbook()
        ws = out_wb.active
        ws.title = "summary"
        ws.append(["country", "operation", "rows", "errors"])
        for cc, ops in summary.items():
            for op, v in ops.items():
                ws.append([cc, op, v.get("rows", 0), v.get("errors", 0)])

        err_ws = out_wb.create_sheet("errors")
        err_ws.append(["country", "operation", "source_sheet", "source_row",
                       "shop", "maincat", "cl1", "error"])
        for err in (_get_task(task_id) or {}).get("errors", []):
            err_ws.append([err.get("country"), err.get("operation"), err.get("source_sheet"),
                           err.get("source_row"), err.get("shop"), err.get("maincat"),
                           err.get("cl1"), err.get("error")])
        out_wb.save(out_path)

        total_errors = len((_get_task(task_id) or {}).get("errors", []))
        final = {
            **(_get_task(task_id) or {}),
            "status": "completed",
            "progress": 100,
            "message": f"Done. {total_errors} error(s).",
            "output_path": str(out_path),
            "summary": summary,
        }
        _set_task(task_id, final)
        _history_append({
            "task_id": task_id, "operation": "monthly_delta", "country": "NL+BE",
            "status": "completed", "started_at": final.get("started_at"),
            "finished_at": datetime.now().isoformat(),
            "summary": summary,
            "output_path": str(out_path),
        })

    except TaskCancelled:
        _set_task(task_id, {**(_get_task(task_id) or {}),
                            "status": "cancelled", "message": "Cancelled by user"})
    except BaseException as exc:
        import traceback as _tb
        _append_log(task_id, f"❌ Failed: {type(exc).__name__}: {exc}")
        _append_log(task_id, _tb.format_exc())
        _set_task(task_id, {**(_get_task(task_id) or {}),
                            "status": "failed", "error": f"{type(exc).__name__}: {exc}"})


# ---------------------------------------------------------------------------
# Category Coverage orchestrator
# ---------------------------------------------------------------------------
def run_category_coverage(task_id: str, country: str):
    """Background-thread target for the coverage export."""
    try:
        _set_task(task_id, {
            **(_get_task(task_id) or {}),
            "status": "running",
            "progress": 5,
            "message": "Crawling taxonomy...",
            "log": [],
        })

        tree = _fetch_taxv2_tree()
        if not tree:
            raise RuntimeError("Could not fetch taxonomy tree")
        _append_log(task_id, f"Categories: {len(tree['id_order'])}",
                    progress=40, message="Fetching PLA campaigns...")

        _check_cancelled(task_id)
        _patch_campaign_processor(country)
        client = _get_client()
        customer_id = COUNTRY_CONFIG[country]["customer_id"]
        campaign_names = _fetch_pla_campaign_names(client, customer_id)
        _append_log(task_id, f"PLA campaigns for {country}: {len(campaign_names)}",
                    progress=70, message="Building xlsx...")

        _check_cancelled(task_id)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"coverage_{country}"
        ws.append(["a", "b", "c", "category_id", "category_name", "parent_id", "is_leaf"])

        id_to_parent = tree["id_to_parent"]
        id_to_name = tree["id_to_name"]
        children_of = tree["children_of"]
        counts = {"a": 0, "b": 0, "c": 0}

        for cid in tree["id_order"]:
            name = id_to_name.get(cid, "")
            parent_id = id_to_parent.get(cid)
            is_leaf = not children_of.get(cid)

            flags = []
            for cl1 in DMA_PLUS_CL1_VALUES:
                exists = f"PLA/{name}_{cl1}" in campaign_names
                flags.append(bool(exists))
                if exists:
                    counts[cl1] += 1
            ws.append([flags[0], flags[1], flags[2], cid, name, parent_id, is_leaf])

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = OUTPUT_DIR / f"category_coverage_{country}_{ts}.xlsx"
        wb.save(out_path)
        _append_log(task_id, f"Wrote {len(tree['id_order'])} rows; TRUE counts a={counts['a']} b={counts['b']} c={counts['c']}")

        final = {
            **(_get_task(task_id) or {}),
            "status": "completed",
            "progress": 100,
            "message": f"Done. a={counts['a']} b={counts['b']} c={counts['c']}",
            "output_path": str(out_path),
            "summary": {country: {**counts, "total": len(tree["id_order"])}},
        }
        _set_task(task_id, final)
        _history_append({
            "task_id": task_id, "operation": "coverage", "country": country,
            "status": "completed", "started_at": final.get("started_at"),
            "finished_at": datetime.now().isoformat(),
            "summary": final["summary"],
            "output_path": str(out_path),
        })

    except TaskCancelled:
        _set_task(task_id, {**(_get_task(task_id) or {}),
                            "status": "cancelled", "message": "Cancelled by user"})
    except BaseException as exc:
        import traceback as _tb
        _append_log(task_id, f"❌ Failed: {type(exc).__name__}: {exc}")
        _append_log(task_id, _tb.format_exc())
        _set_task(task_id, {**(_get_task(task_id) or {}),
                            "status": "failed", "error": f"{type(exc).__name__}: {exc}"})


# ---------------------------------------------------------------------------
# Public entry points (called from router)
# ---------------------------------------------------------------------------
def start_monthly(wb_bytes: bytes, dry_run: bool = False) -> str:
    task_id = uuid.uuid4().hex[:8]
    _set_task(task_id, {
        "status": "queued",
        "operation": "monthly_delta",
        "country": "NL+BE",
        "progress": 0,
        "message": "Queued..." + (" [DRY RUN]" if dry_run else ""),
        "started_at": datetime.now().isoformat(),
        "dry_run": dry_run,
    })
    threading.Thread(target=run_monthly_delta, args=(task_id, wb_bytes),
                     kwargs={"dry_run": dry_run}, daemon=True).start()
    return task_id


def start_coverage(country: str) -> str:
    country = (country or "NL").upper()
    if country not in COUNTRY_CONFIG:
        raise ValueError(f"Unknown country {country!r}")
    task_id = uuid.uuid4().hex[:8]
    _set_task(task_id, {
        "status": "queued",
        "operation": "coverage",
        "country": country,
        "progress": 0,
        "message": "Queued...",
        "started_at": datetime.now().isoformat(),
    })
    threading.Thread(target=run_category_coverage, args=(task_id, country), daemon=True).start()
    return task_id


def get_output_path(task_id: str) -> Optional[str]:
    t = _get_task(task_id) or {}
    return t.get("output_path")
