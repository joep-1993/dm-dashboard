from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, HTTPException, UploadFile, File, Request, Response as FastAPIResponse, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse, RedirectResponse, Response, HTMLResponse, JSONResponse
import io
import httpx
from urllib.parse import urljoin
from datetime import datetime
from io import StringIO, BytesIO
import csv
import hashlib
import hmac
import json
import os
import asyncio
import secrets
import tempfile
import time
import re
import threading
import uuid
from functools import partial, wraps
from concurrent.futures import ThreadPoolExecutor

# In-memory store for background validation tasks
_validation_tasks = {}

def _get_validation_task(task_id):
    return _validation_tasks.get(task_id)

def _set_validation_task(task_id, data):
    # Preserve the cancel flag — prevents race condition where a batch
    # completion overwrites a cancel request that arrived mid-batch.
    existing = _validation_tasks.get(task_id)
    if existing and existing.get("cancel") and "cancel" not in data:
        data["cancel"] = True
    _validation_tasks[task_id] = data
from backend.database import get_db_connection, get_output_connection, return_db_connection, return_output_connection
from backend.scraper_service import scrape_product_page, scrape_product_page_api, sanitize_content, is_main_category_url, MAIN_CATEGORY_H1
from backend.gpt_service import generate_product_content, generate_main_category_content, check_content_has_valid_links
from backend.link_validator import validate_content_links, validate_and_fix_content_links
from backend.faq_service import process_single_url_faq
from backend.batch_api_service import start_faq_batch, start_kopteksten_batch, start_titles_batch, get_batch_status
from backend.thema_ads_router import router as thema_ads_router, cleanup_stale_jobs as cleanup_thema_ads_jobs
from backend.gsd_campaigns_router import router as gsd_campaigns_router
from backend.dma_bidding_router import router as dma_bidding_router
from backend.gsd_budgets_router import router as gsd_budgets_router
from backend.mc_id_finder_router import router as mc_id_finder_router
from backend.redshift_upload_router import router as redshift_upload_router
from backend.task_scheduler_router import router as task_scheduler_router
from backend.url_validator_router import router as url_validator_router
from backend.dma_plus_router import router as dma_plus_router
from backend.rurl_optimizer_router import router as rurl_optimizer_router
from backend.rurl_optimizer_v2_router import router as rurl_optimizer_v2_router
from backend.keyword_planner_service import get_search_volumes, test_api_connection as test_keyword_planner_connection
from backend.category_keyword_service import process_category_keywords, PRELOADED_CATEGORIES
from backend.content_publisher import (
    get_total_content_count,
    get_content_batch,
    publish_all_content,
    generate_curl_command,
    start_publish_task,
    get_publish_task_status
)
import psycopg2

app = FastAPI(title="SEO Tools - Unified Platform", version="1.0.0")

# --- Authentication (env-gated; disabled when DASHBOARD_PASSWORD is unset) ---
AUTH_PASSWORD = os.getenv("DASHBOARD_PASSWORD", "")
AUTH_SECRET = os.getenv("DASHBOARD_SECRET", secrets.token_hex(32))
AUTH_COOKIE = "dm_session"
PUBLIC_PATHS = {"/login", "/api/health"}

def _make_token(password: str) -> str:
    return hmac.new(AUTH_SECRET.encode(), password.encode(), hashlib.sha256).hexdigest()

LOGIN_PAGE = """<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>DM Dashboard - Login</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<style>body{background:#f5f5f5;display:flex;align-items:center;justify-content:center;min-height:100vh}
.login-card{max-width:400px;width:100%}.error{color:#dc3545;font-size:.875rem;margin-top:.5rem}</style>
</head><body>
<div class="login-card card shadow-sm"><div class="card-body p-4">
<h4 class="card-title mb-3">DM Dashboard</h4>
<form method="POST" action="/login">
<div class="mb-3"><label class="form-label">Password</label>
<input type="password" name="password" class="form-control" autofocus required></div>
<button type="submit" class="btn btn-primary w-100">Log in</button>
{error}
</form></div></div></body></html>"""

@app.get("/login", response_class=HTMLResponse)
async def login_page():
    return LOGIN_PAGE.replace("{error}", "")

@app.post("/login", response_class=HTMLResponse)
async def login_submit(request: Request):
    form = await request.form()
    password = form.get("password", "")
    if hmac.compare_digest(password, AUTH_PASSWORD):
        response = RedirectResponse(url="/", status_code=303)
        response.set_cookie(AUTH_COOKIE, _make_token(password), httponly=True, samesite="lax", max_age=86400 * 30)
        return response
    return HTMLResponse(
        LOGIN_PAGE.replace("{error}", '<div class="error">Incorrect password</div>'),
        status_code=401,
    )

@app.get("/logout")
async def logout():
    response = RedirectResponse(url="/login")
    response.delete_cookie(AUTH_COOKIE)
    return response

from starlette.middleware.base import BaseHTTPMiddleware

class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        if not AUTH_PASSWORD:
            return await call_next(request)
        path = request.url.path
        if path in PUBLIC_PATHS:
            return await call_next(request)
        token = request.cookies.get(AUTH_COOKIE, "")
        if hmac.compare_digest(token, _make_token(AUTH_PASSWORD)):
            return await call_next(request)
        if path.startswith("/api/"):
            return HTMLResponse('{"detail":"Not authenticated"}', status_code=401, media_type="application/json")
        return RedirectResponse(url="/login")

app.add_middleware(AuthMiddleware)

# Include thema_ads router
app.include_router(thema_ads_router)

# Include gsd_campaigns router
app.include_router(gsd_campaigns_router)

# Include dma_bidding router
app.include_router(dma_bidding_router)

# Include gsd_budgets router
app.include_router(gsd_budgets_router)

# Include mc_id_finder router
app.include_router(mc_id_finder_router)

# Include redshift_upload router
app.include_router(redshift_upload_router)

# Include url_validator router
app.include_router(url_validator_router)

# Include dma_plus router
app.include_router(dma_plus_router)

# Include rurl_optimizer routers (v1 + v2)
app.include_router(rurl_optimizer_router)
app.include_router(rurl_optimizer_v2_router)

# Include task_scheduler router (env-gated — Windows-only, depends on schtasks)
TASK_SCHEDULER_ENABLED = os.getenv("ENABLE_TASK_SCHEDULER", "false").lower() == "true"
if TASK_SCHEDULER_ENABLED:
    app.include_router(task_scheduler_router)

@app.get("/api/config")
async def get_runtime_config():
    """Feature flags the frontend reads on load to reveal gated UI."""
    return {"task_scheduler_enabled": TASK_SCHEDULER_ENABLED}

@app.on_event("startup")
async def startup_event():
    """Run startup tasks for all services."""
    await cleanup_thema_ads_jobs()


@app.on_event("shutdown")
async def shutdown_event():
    """Close long-lived HTTP sessions to prevent CLOSE_WAIT socket buildup."""
    from backend import gpt_service, scraper_service, link_validator, faq_service, ai_titles_service
    from backend.url_validator_service import _taxonomy

    for label, session in [
        ("gpt_service", getattr(gpt_service, "_http_client", None)),
        ("scraper_service", getattr(scraper_service, "_session", None)),
        ("link_validator", getattr(link_validator, "_es_session", None)),
        ("faq_service", getattr(faq_service, "_faq_session", None)),
        ("ai_titles_http", getattr(ai_titles_service, "_http_session", None)),
        ("url_validator", getattr(_taxonomy, "_session", None)),
    ]:
        if session is not None:
            try:
                session.close()
                print(f"[SHUTDOWN] Closed {label} session")
            except Exception as e:
                print(f"[SHUTDOWN] Error closing {label}: {e}")

    # Close OpenAI clients (they wrap an internal httpx client)
    for label, client in [
        ("gpt_service", getattr(gpt_service, "_openai_client", None)),
        ("ai_titles", getattr(ai_titles_service, "_openai_client", None)),
        ("faq_service", getattr(faq_service, "_openai_client", None)),
    ]:
        if client is not None:
            try:
                client.close()
                print(f"[SHUTDOWN] Closed {label} OpenAI client")
            except Exception as e:
                print(f"[SHUTDOWN] Error closing {label}: {e}")

# CORS for frontend. Set CORS_ORIGINS (comma-separated) in env to restrict; default = "*".
_cors_origins_env = os.getenv("CORS_ORIGINS", "*").strip()
_cors_origins = [o.strip() for o in _cors_origins_env.split(",") if o.strip()] or ["*"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=_cors_origins != ["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve frontend static files
app.mount("/static", StaticFiles(directory="frontend"), name="static")

def retry_on_redshift_serialization_error(max_retries=3, initial_delay=0.1):
    """
    Decorator to retry database operations that fail due to Redshift serialization conflicts.
    Error 1023: Serializable isolation violation on table
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            delay = initial_delay
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except psycopg2.Error as e:
                    error_msg = str(e)
                    # Check for Redshift serialization conflict (error code 1023)
                    if "1023" in error_msg or "Serializable isolation violation" in error_msg:
                        if attempt < max_retries - 1:
                            print(f"[RETRY] Redshift serialization conflict detected (attempt {attempt + 1}/{max_retries}), retrying in {delay}s...")
                            time.sleep(delay)
                            delay *= 2  # Exponential backoff
                            continue
                    # Re-raise if not a serialization error or max retries exceeded
                    raise
            return None
        return wrapper
    return decorator

@app.get("/")
def read_root():
    """Redirect to the unified dashboard"""
    return RedirectResponse(url="/static/dashboard.html")

@app.get("/api/health")
def health_check():
    return {"status": "healthy", "service": "dm_tools"}

@app.get("/api/debug/test-scraper")
def debug_test_scraper(url: str):
    """Debug endpoint to test scraper on a single URL"""
    import time
    start = time.time()
    try:
        result = scrape_product_page_api(url)
        elapsed = time.time() - start
        if result is None:
            return {"status": "failed", "reason": "api_returned_none", "elapsed_seconds": elapsed}
        if result.get('error'):
            return {"status": "failed", "reason": result.get('error'), "elapsed_seconds": elapsed}
        return {
            "status": "success",
            "elapsed_seconds": elapsed,
            "product_count": len(result.get('products', [])),
            "h1_title": result.get('h1_title'),
            "product_subject": result.get('product_subject')
        }
    except Exception as e:
        elapsed = time.time() - start
        return {"status": "error", "reason": str(e), "elapsed_seconds": elapsed}

@app.post("/api/generate")
async def generate_text(prompt: str):
    """Example endpoint for AI generation"""
    from backend.gpt_service import simple_completion
    try:
        result = simple_completion(prompt)
        return {"response": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def process_single_url(url: str, conservative_mode: bool = False):
    """Process a single URL - runs in thread pool
    Returns tuple: (result_dict, redshift_operations)

    Args:
        url: URL to process
        conservative_mode: If True, use conservative scraping rate (max 2 URLs/sec)
    """
    result = {"url": url, "status": "pending"}
    conn = None
    final_status = None
    final_reason = None

    try:
        # Use API-based scraper for better product subject extraction
        scraped_data = scrape_product_page_api(url)

        # Check for 503 error (rate limiting) - should stop batch processing
        if scraped_data and scraped_data.get('error') == '503':
            final_status = 'failed'
            final_reason = 'rate_limited_503'
            result["status"] = "failed"
            result["reason"] = "rate_limited_503"
            # DO NOT mark as processed in Redshift - keep in pending for retry
            # 503 errors should stop the batch to avoid further rate limiting
        elif not scraped_data:
            final_status = 'failed'
            final_reason = 'api_failed'
            result["status"] = "failed"
            result["reason"] = "api_failed"
        elif not scraped_data['products'] or len(scraped_data['products']) == 0:
            final_status = 'skipped'
            final_reason = 'no_products_found'
            result["status"] = "skipped"
            result["reason"] = "no_products_found"
        else:
            # Generate AI content using product_subject from selected facets
            try:
                # Check if this is a main category URL - use special prompt + H1 from mapping
                if is_main_category_url(url):
                    from backend.scraper_service import parse_beslist_url
                    main_cat_slug, _, _ = parse_beslist_url(url)
                    content_topic = MAIN_CATEGORY_H1.get(main_cat_slug, scraped_data['h1_title'])
                    print(f"[DEBUG] Main category detected: {url[:80]}... using H1 '{content_topic}' and {len(scraped_data['products'])} products")
                    ai_content = generate_main_category_content(
                        content_topic,
                        scraped_data['products']
                    )
                else:
                    # Use product_subject if available (from API), otherwise fall back to h1_title
                    content_topic = scraped_data.get('product_subject') or scraped_data['h1_title']
                    print(f"[DEBUG] Generating AI content for {url[:80]}... with topic '{content_topic}' and {len(scraped_data['products'])} products")
                    ai_content = generate_product_content(
                        content_topic,
                        scraped_data['products']
                    )
                print(f"[DEBUG] AI content generated, length: {len(ai_content)}")

                # Sanitize content for SQL
                sanitized = sanitize_content(ai_content)

                # Check if content has valid links
                has_valid_links = check_content_has_valid_links(ai_content)
                print(f"[DEBUG] Generated content for {url[:80]}... - Has valid links: {has_valid_links}")
                print(f"[DEBUG] Content preview: {ai_content[:200]}...")

                if not has_valid_links:
                    final_status = 'failed'
                    final_reason = 'no_valid_links'
                    result["status"] = "failed"
                    result["reason"] = "no_valid_links"
                else:
                    # Save content to local PostgreSQL immediately
                    content_conn = None
                    try:
                        content_conn = get_db_connection()
                        content_cur = content_conn.cursor()
                        content_cur.execute("""
                            INSERT INTO pa.content_urls_joep (url, content)
                            VALUES (%s, %s)
                        """, (url, sanitized))
                        content_conn.commit()
                        content_cur.close()
                    finally:
                        if content_conn:
                            return_db_connection(content_conn)

                    final_status = 'success'
                    result["status"] = "success"
                    # Strip HTML tags from preview to avoid broken tags
                    preview_text = re.sub(r'<[^>]+>', '', ai_content)
                    result["content_preview"] = preview_text[:100] + "..."

            except Exception as e:
                final_status = 'failed'
                final_reason = f"ai_generation_error: {str(e)}"
                result["status"] = "failed"
                result["reason"] = f"ai_generation_error: {str(e)}"

        # Single DB transaction at the end with final status
        conn = get_db_connection()
        cur = conn.cursor()

        truncated_reason = final_reason[:255] if final_reason and len(final_reason) > 255 else final_reason

        if final_status == 'skipped' and final_reason and 'no_products_found' in final_reason:
            # Write to shared validation table (applies to both kopteksten and FAQ)
            cur.execute("""
                INSERT INTO pa.url_validation_tracking (url, status, skip_reason)
                VALUES (%s, %s, %s)
                ON CONFLICT (url) DO UPDATE SET status = EXCLUDED.status, skip_reason = EXCLUDED.skip_reason, checked_at = CURRENT_TIMESTAMP
            """, (url, final_status, truncated_reason))
        elif final_reason:
            cur.execute("""
                INSERT INTO pa.jvs_seo_werkvoorraad_kopteksten_check (url, status, skip_reason)
                VALUES (%s, %s, %s)
                ON CONFLICT (url) DO UPDATE SET status = EXCLUDED.status, skip_reason = EXCLUDED.skip_reason
            """, (url, final_status, truncated_reason))
        else:
            cur.execute("""
                INSERT INTO pa.jvs_seo_werkvoorraad_kopteksten_check (url, status)
                VALUES (%s, %s)
                ON CONFLICT (url) DO UPDATE SET status = EXCLUDED.status, skip_reason = NULL
            """, (url, final_status))

        conn.commit()
        print(f"[PROCESSING] {url} - Status: {final_status}" + (f" - Reason: {final_reason}" if final_reason else ""))
        return result

    except Exception as e:
        result["status"] = "failed"
        result["reason"] = f"error: {str(e)}"
        # Try to record error in DB
        try:
            if not conn:
                conn = get_db_connection()
                cur = conn.cursor()
            # Truncate error message to 255 characters to fit VARCHAR(255) column
            error_msg = f"error: {str(e)}"[:255]
            cur.execute("""
                INSERT INTO pa.jvs_seo_werkvoorraad_kopteksten_check (url, status, skip_reason)
                VALUES (%s, 'failed', %s)
                ON CONFLICT (url) DO UPDATE SET status = 'failed', skip_reason = EXCLUDED.skip_reason
            """, (url, error_msg))
            conn.commit()
        except (psycopg2.DatabaseError, psycopg2.Error) as db_err:
            print(f"[process_single_url] Failed to persist error status for {url}: {db_err}")
        return result
    finally:
        if conn:
            cur.close()
            return_db_connection(conn)  # Return connection to pool instead of closing

@app.post("/api/process-urls")
def process_urls(batch_size: int = 2, parallel_workers: int = 1, conservative_mode: bool = False):
    """
    Process batch of URLs for SEO content generation.
    Fetches specified number of URLs, scrapes content, generates AI text, and saves to database.
    Supports parallel processing with configurable workers.

    Args:
        batch_size: Number of URLs to process
        parallel_workers: Number of parallel workers (1-10), ignored if conservative_mode is True
        conservative_mode: If True, use conservative scraping rate (max 2 URLs/sec) with 1 worker. Default: False
    """
    print(f"[ENDPOINT] process_urls called - batch_size={batch_size}, workers={parallel_workers}, conservative={conservative_mode}")

    try:
        # Validate parameters
        if batch_size < 1:
            raise HTTPException(status_code=400, detail="Batch size must be at least 1")

        if parallel_workers < 1 or parallel_workers > 100:
            raise HTTPException(status_code=400, detail="Parallel workers must be between 1 and 100")

        # Conservative mode always uses 1 worker for maximum safety
        if conservative_mode:
            parallel_workers = 1

        print(f"[ENDPOINT] Getting local connection...")
        # Get unprocessed URLs from local PostgreSQL
        local_conn = get_db_connection()
        print(f"[ENDPOINT] Got local connection, creating cursor...")
        local_cur = local_conn.cursor()

        # Fetch unprocessed URLs from local werkvoorraad (URLs not yet in tracking table)
        try:
            print(f"[ENDPOINT] Querying for {batch_size} pending URLs...")
            local_cur.execute("""
                SELECT w.url
                FROM pa.jvs_seo_werkvoorraad w
                WHERE NOT EXISTS (SELECT 1 FROM pa.jvs_seo_werkvoorraad_kopteksten_check t WHERE t.url = w.url)
                  AND NOT EXISTS (SELECT 1 FROM pa.url_validation_tracking v WHERE v.url = w.url)
                LIMIT %s
            """, (batch_size,))

            rows = local_cur.fetchall()
            print(f"[ENDPOINT] Got {len(rows)} URLs from local PostgreSQL")
        finally:
            print(f"[ENDPOINT] Closing cursor and returning connection...")
            local_cur.close()
            return_db_connection(local_conn)
            local_conn = None
            print(f"[ENDPOINT] Connection returned to pool")

        if not rows:
            return {
                "status": "complete",
                "message": "No URLs to process",
                "processed": 0
            }

        urls = [row['url'] for row in rows]

        # Process URLs in parallel using ThreadPoolExecutor
        # Use partial to bind conservative_mode parameter
        process_func = partial(process_single_url, conservative_mode=conservative_mode)
        with ThreadPoolExecutor(max_workers=parallel_workers) as executor:
            results = list(executor.map(process_func, urls))

        # Check for rate limiting
        rate_limited = False
        for result in results:
            if result['status'] == 'failed' and result.get('reason') == 'rate_limited_503':
                rate_limited = True
                print(f"[RATE LIMIT DETECTED] 503 error detected - stopping batch immediately")
                break

        processed_count = sum(1 for r in results if r['status'] == 'success')
        skipped_count = sum(1 for r in results if r['status'] == 'skipped')
        failed_count = sum(1 for r in results if r['status'] == 'failed')

        if rate_limited:
            print(f"[BATCH STOPPED - RATE LIMITED] Processed: {processed_count}/{len(urls)} | Skipped: {skipped_count} | Failed: {failed_count}")
        else:
            print(f"[BATCH COMPLETE] Processed: {processed_count}/{len(urls)} | Skipped: {skipped_count} | Failed: {failed_count}")

        return {
            "status": "rate_limited" if rate_limited else "success",
            "processed": processed_count,
            "total_attempted": len(results),
            "rate_limited": rate_limited,
            "message": "Stopped due to rate limiting - wait before retrying" if rate_limited else None,
            "results": results
        }

    except Exception as e:
        print(f"[ERROR] process_urls failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/status")
def get_status():
    """Get processing status and counts (LOCAL PostgreSQL only)"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Count pending directly (same logic as process_urls query)
        cur.execute("""
            SELECT
                (SELECT COUNT(*) FROM pa.jvs_seo_werkvoorraad) as total,
                (SELECT COUNT(DISTINCT url) FROM pa.content_urls_joep WHERE content IS NOT NULL) as processed,
                (SELECT COUNT(*) FROM pa.url_validation_tracking WHERE status = 'skipped') as skipped,
                (SELECT COUNT(*) FROM pa.jvs_seo_werkvoorraad_kopteksten_check WHERE status = 'failed') as failed,
                (SELECT COUNT(*) FROM pa.jvs_seo_werkvoorraad w
                 WHERE NOT EXISTS (SELECT 1 FROM pa.jvs_seo_werkvoorraad_kopteksten_check t WHERE t.url = w.url)
                   AND NOT EXISTS (SELECT 1 FROM pa.url_validation_tracking v WHERE v.url = w.url)) as pending
        """)
        counts = cur.fetchone()
        total = counts['total']
        processed = counts['processed']
        skipped = counts['skipped']
        failed = counts['failed']
        pending = counts['pending']

        # Get recent results from local PostgreSQL
        try:
            cur.execute("""
                SELECT url, content, created_at
                FROM pa.content_urls_joep
                ORDER BY created_at DESC NULLS LAST
                LIMIT 5
            """)
            recent_rows = cur.fetchall()
            recent = [{'url': r['url'], 'content': r['content'], 'created_at': r['created_at'].isoformat() if r.get('created_at') else None} for r in recent_rows]
        except Exception as e:
            print(f"[DEBUG] Failed to get recent results: {e}")
            recent = []

        cur.close()
        return_db_connection(conn)

        return {
            "total_urls": total,
            "processed": processed,
            "skipped": skipped,
            "failed": failed,
            "pending": pending,
            "recent_results": recent
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/failure-reasons")
def get_failure_reasons():
    """Get breakdown of failure reasons"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT status, skip_reason, COUNT(*) as count
            FROM (
                SELECT status, skip_reason FROM pa.jvs_seo_werkvoorraad_kopteksten_check
                UNION ALL
                SELECT status, skip_reason FROM pa.url_validation_tracking
            ) combined
            GROUP BY status, skip_reason
            ORDER BY count DESC
        """)
        rows = cur.fetchall()

        cur.close()
        return_db_connection(conn)

        return {
            "breakdown": [{"status": r["status"], "reason": r["skip_reason"], "count": r["count"]} for r in rows]
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/export/xlsx")
async def export_xlsx():
    """Export all generated content as Excel XLSX (from local PostgreSQL)"""
    from openpyxl import Workbook

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT url, content
            FROM pa.content_urls_joep
        """)
        rows = cur.fetchall()

        cur.close()
        return_db_connection(conn)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Content Export"

        # Add headers
        ws.append(['url', 'content'])

        # Add data rows
        import re
        # Remove control characters that Excel doesn't allow (except tab, newline, carriage return)
        illegal_chars = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

        for row in rows:
            content = row['content'] if row['content'] else ''
            # Remove illegal control characters
            content = illegal_chars.sub('', content)
            ws.append([row['url'], content])

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=content_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/export/json")
async def export_json():
    """Export all generated content as JSON (from local PostgreSQL)"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT url, content
            FROM pa.content_urls_joep
        """)
        rows = cur.fetchall()

        cur.close()
        return_db_connection(conn)

        # Convert to JSON-serializable format
        data = []
        for row in rows:
            data.append({
                'url': row['url'],
                'content': row['content']
            })

        # Return as downloadable file
        json_str = json.dumps(data, indent=2, ensure_ascii=False)
        return StreamingResponse(
            iter([json_str]),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename=content_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/upload-urls")
async def upload_urls(file: UploadFile = File(...)):
    """Upload a text file with URLs (one per line) to add to the work queue"""
    try:
        # Read file content
        content = await file.read()

        # Try multiple encodings (UTF-16, UTF-8 with BOM, UTF-8, Windows-1252, Latin-1)
        text_content = None
        for encoding in ['utf-16', 'utf-16-le', 'utf-8-sig', 'utf-8', 'windows-1252', 'latin-1']:
            try:
                text_content = content.decode(encoding)
                # Verify no replacement characters
                if '�' not in text_content:
                    break
            except (UnicodeDecodeError, UnicodeError):
                continue

        if text_content is None:
            text_content = content.decode('latin-1')  # Latin-1 never fails

        # Handle both newlines and semicolons as separators (for CSV format)
        lines = text_content.strip().replace('\r\n', '\n').replace('\r', '\n').split('\n')

        # Extract URLs from each line (first column if CSV)
        urls = []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            # If line contains semicolons, it's CSV format - take first column
            if ';' in line:
                url = line.split(';')[0].strip()
            else:
                url = line.strip()

            if url and not url.startswith('url'):  # Skip CSV header
                urls.append(url)

        if not urls:
            raise HTTPException(status_code=400, detail="No URLs found in file")

        # Insert URLs into Redshift work queue
        output_conn = get_output_connection()
        output_cur = output_conn.cursor()

        added_count = 0
        duplicate_count = 0
        base_url = "https://www.beslist.nl"

        # Convert relative URLs to absolute URLs
        full_urls = []
        for url in urls:
            if url.startswith('/'):
                full_url = base_url + url
            else:
                full_url = url
            full_urls.append(full_url)

        # Get existing URLs from database in batches (Redshift performs better with smaller batches)
        existing_urls = set()
        batch_size = 500  # Check in batches of 500

        for i in range(0, len(full_urls), batch_size):
            batch = full_urls[i:i + batch_size]
            placeholders = ','.join(['%s'] * len(batch))
            output_cur.execute(f"""
                SELECT url FROM pa.jvs_seo_werkvoorraad_shopping_season
                WHERE url IN ({placeholders})
            """, batch)
            existing_urls.update(row['url'] for row in output_cur.fetchall())

        # Filter out duplicates
        new_urls = [(url,) for url in full_urls if url not in existing_urls]
        duplicate_count = len(full_urls) - len(new_urls)

        # Batch insert new URLs
        if new_urls:
            # Insert in batches for better Redshift performance
            insert_batch_size = 100
            for i in range(0, len(new_urls), insert_batch_size):
                batch = new_urls[i:i + insert_batch_size]
                output_cur.executemany("""
                    INSERT INTO pa.jvs_seo_werkvoorraad_shopping_season (url, kopteksten)
                    VALUES (%s, 0)
                """, batch)
            added_count = len(new_urls)

        output_conn.commit()
        output_cur.close()
        return_output_connection(output_conn)

        return {
            "status": "success",
            "total_urls": len(urls),
            "added": added_count,
            "duplicates": duplicate_count,
            "message": f"Added {added_count} new URLs, {duplicate_count} duplicates skipped"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/content/lookup")
async def lookup_content(url: str):
    """Look up content for a specific URL."""
    try:
        # Normalize URL - build both relative and absolute variants to match DB
        clean_url = url.strip().lower()
        base_url = "https://www.beslist.nl"

        if clean_url.startswith('http'):
            # Full URL provided - extract path as well
            if 'beslist.nl' in clean_url:
                path_url = '/' + clean_url.split('beslist.nl', 1)[-1].lstrip('/')
                full_url = base_url + path_url
            else:
                path_url = clean_url
                full_url = clean_url
        else:
            # Relative path provided - build full URL
            if not clean_url.startswith('/'):
                clean_url = '/' + clean_url
            path_url = clean_url
            full_url = base_url + path_url

        conn = get_db_connection()
        cur = conn.cursor()

        # Look up in content table - try both full URL and relative path
        cur.execute("""
            SELECT url, content, created_at
            FROM pa.content_urls_joep
            WHERE url = %s OR url = %s
            LIMIT 1
        """, (full_url, path_url))
        row = cur.fetchone()

        cur.close()
        return_db_connection(conn)

        if not row:
            return {
                "found": False,
                "url": clean_url,
                "message": "URL not found in content database"
            }

        return {
            "found": True,
            "url": row['url'],
            "content": row['content'],
            "created_at": row['created_at'].isoformat() if row.get('created_at') else None
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/result/{url:path}")
async def delete_result(url: str):
    """Delete a result and reset the URL back to pending state"""
    try:
        use_redshift = os.getenv("USE_REDSHIFT_OUTPUT", "false").lower() == "true"
        werkvoorraad_table = "pa.jvs_seo_werkvoorraad_shopping_season" if use_redshift else "pa.jvs_seo_werkvoorraad"

        # Delete from output table and update werkvoorraad - with retry on serialization conflicts
        @retry_on_redshift_serialization_error(max_retries=5, initial_delay=0.2)
        def delete_from_output():
            output_conn = get_output_connection()
            output_cur = output_conn.cursor()
            try:
                # Delete content
                output_cur.execute("""
                    DELETE FROM pa.content_urls_joep
                    WHERE url = %s
                """, (url,))

                # Reset kopteksten flag in werkvoorraad
                output_cur.execute(f"""
                    UPDATE {werkvoorraad_table}
                    SET kopteksten = 0
                    WHERE url = %s
                """, (url,))

                output_conn.commit()
            except Exception as e:
                output_conn.rollback()
                raise e
            finally:
                output_cur.close()
                return_output_connection(output_conn)

        delete_from_output()

        # Delete from local tracking table and shared validation table
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            DELETE FROM pa.jvs_seo_werkvoorraad_kopteksten_check
            WHERE url = %s
        """, (url,))
        cur.execute("""
            DELETE FROM pa.url_validation_tracking
            WHERE url = %s
        """, (url,))
        conn.commit()
        cur.close()
        return_db_connection(conn)

        return {
            "status": "success",
            "message": f"Result deleted and URL reset to pending",
            "url": url
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def validate_single_content_es(content_data: tuple) -> dict:
    """Validate and fix links in a single content item using Elasticsearch lookup.

    Args:
        content_data: Tuple of (content_url, content)

    Returns:
        Dict with validation results including:
        - content_url: The URL this content belongs to
        - has_changes: Whether URLs were replaced
        - replaced_urls: List of {old_url, new_url} replacements
        - gone_urls: URLs where product is gone (need reprocessing)
        - valid_urls: URLs that were already correct
        - corrected_content: Content with corrected URLs (if any changes)
    """
    content_url, content = content_data
    return validate_and_fix_content_links(content, content_url)

@app.post("/api/validate-links")
def validate_links(batch_size: int = 10, parallel_workers: int = 3, conservative_mode: bool = False):
    """
    Validate and fix hyperlinks in generated content using Elasticsearch lookup.

    - If a product URL differs from the canonical plpUrl, the content is auto-corrected
    - If a product is GONE from Elasticsearch, the content is deleted and URL queued for reprocessing
    - URLs with only corrected links keep their kopteksten=1 status

    Args:
        batch_size: Number of content items to validate
        parallel_workers: Number of parallel workers (1-10)
        conservative_mode: Legacy parameter, kept for compatibility (ignored)
    """
    try:
        # Validate parameters
        if batch_size < 1:
            raise HTTPException(status_code=400, detail="Batch size must be at least 1")

        if parallel_workers < 1 or parallel_workers > 100:
            raise HTTPException(status_code=400, detail="Parallel workers must be between 1 and 100")

        # Use local PostgreSQL for all operations
        conn = get_db_connection()
        cur = conn.cursor()

        # Fetch unvalidated content URLs using LEFT JOIN for efficiency
        cur.execute("""
            SELECT c.url, c.content
            FROM pa.content_urls_joep c
            LEFT JOIN pa.link_validation_results v ON c.url = v.content_url
            WHERE v.content_url IS NULL
            LIMIT %s
        """, (batch_size,))

        rows = cur.fetchall()

        if not rows:
            cur.close()
            return_db_connection(conn)
            return {
                "status": "complete",
                "message": "No content to validate",
                "validated": 0,
                "urls_corrected": 0,
                "moved_to_pending": 0
            }

        # Prepare content items for parallel validation
        content_items = [(row['url'], row['content']) for row in rows]

        # Process validations in parallel using ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=parallel_workers) as executor:
            validation_results = list(executor.map(validate_single_content_es, content_items))

        results = []
        urls_corrected = 0
        moved_to_pending = 0
        urls_with_gone_products = []  # Only these get kopteksten reset to 0
        urls_to_update_content = []  # URLs where content needs updating (replaced URLs)
        url_to_content = {url: content for url, content in content_items}  # For backup
        url_to_gone_details = {}  # Store gone_urls details for backup

        # Process validation results
        for validation_result in validation_results:
            content_url = validation_result['content_url']
            has_replaced = len(validation_result['replaced_urls']) > 0
            has_gone = len(validation_result['gone_urls']) > 0

            # Calculate totals for tracking
            total_links = len(validation_result['valid_urls']) + len(validation_result['replaced_urls']) + len(validation_result['gone_urls'])

            # Save validation results to local tracking table
            cur.execute("""
                INSERT INTO pa.link_validation_results
                (content_url, total_links, broken_links, valid_links, broken_link_details)
                VALUES (%s, %s, %s, %s, %s)
            """, (
                content_url,
                total_links,
                len(validation_result['gone_urls']),  # Only GONE URLs are truly broken
                len(validation_result['valid_urls']) + len(validation_result['replaced_urls']),  # Replaced URLs are now valid
                json.dumps({
                    'gone_urls': validation_result['gone_urls'],
                    'replaced_urls': validation_result['replaced_urls']
                })
            ))

            # Handle URL replacements - update content in local database
            if has_replaced and not has_gone:
                # Only replaced URLs, no gone URLs - update content, keep kopteksten=1
                # Order matches `UPDATE ... SET content = %s WHERE url = %s` below.
                urls_to_update_content.append((validation_result['corrected_content'], content_url))
                urls_corrected += 1

            # Handle gone products - need to regenerate content
            if has_gone:
                urls_with_gone_products.append(content_url)
                url_to_gone_details[content_url] = validation_result['gone_urls']
                moved_to_pending += 1

            results.append({
                'url': content_url,
                'total_links': total_links,
                'valid_urls': len(validation_result['valid_urls']),
                'replaced_urls': validation_result['replaced_urls'],
                'gone_urls': validation_result['gone_urls'],
                'content_corrected': has_replaced and not has_gone,
                'moved_to_pending': has_gone
            })

        # Update corrected content in local database (batched for performance)
        if urls_to_update_content:
            cur.executemany("""
                UPDATE pa.content_urls_joep
                SET content = %s
                WHERE url = %s
            """, urls_to_update_content)
            print(f"[VALIDATE-LINKS] Updated content for {len(urls_to_update_content)} URLs with corrected links")

        # Delete/reset operations for gone products only
        if urls_with_gone_products:
            placeholders = ','.join(['%s'] * len(urls_with_gone_products))

            # Backup content to history table before deletion
            cur.execute(f"""
                SELECT url, content, created_at FROM pa.content_urls_joep
                WHERE url IN ({placeholders})
            """, urls_with_gone_products)
            content_to_backup = cur.fetchall()

            for row in content_to_backup:
                url = row['url']
                gone_urls = url_to_gone_details.get(url, [])
                cur.execute("""
                    INSERT INTO pa.content_history (url, content, reset_reason, reset_details, original_created_at)
                    VALUES (%s, %s, %s, %s, %s)
                """, (url, row['content'], 'gone_products', json.dumps({'gone_urls': gone_urls}), row['created_at']))

            print(f"[VALIDATE-LINKS] Backed up {len(content_to_backup)} URLs to content_history")

            # Delete from content table (local PostgreSQL)
            cur.execute(f"""
                DELETE FROM pa.content_urls_joep
                WHERE url IN ({placeholders})
            """, urls_with_gone_products)

            # Delete from tracking tables (local PostgreSQL)
            cur.execute(f"""
                DELETE FROM pa.jvs_seo_werkvoorraad_kopteksten_check
                WHERE url IN ({placeholders})
            """, urls_with_gone_products)
            cur.execute(f"""
                DELETE FROM pa.url_validation_tracking
                WHERE url IN ({placeholders})
            """, urls_with_gone_products)

            # Add URLs to werkvoorraad for reprocessing (batched for performance)
            cur.executemany("""
                INSERT INTO pa.jvs_seo_werkvoorraad (url, kopteksten)
                VALUES (%s, 0)
                ON CONFLICT (url) DO UPDATE SET kopteksten = 0
            """, [(url,) for url in urls_with_gone_products])

            print(f"[VALIDATE-LINKS] Deleted content for {len(urls_with_gone_products)} URLs with gone products")

        conn.commit()
        cur.close()
        return_db_connection(conn)

        return {
            "status": "success",
            "validated": len(rows),
            "urls_corrected": urls_corrected,
            "moved_to_pending": moved_to_pending,
            "results": results
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/validate-all-links/status/{task_id}")
def get_validate_all_status(task_id: str):
    """Poll progress of a running validate-all task."""
    task = _get_validation_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task

@app.post("/api/validate-all-links/cancel/{task_id}")
def cancel_validate_all(task_id: str):
    """Cancel a running validate-all task.  Use task_id='all' to cancel every running task."""
    if task_id == "all":
        cancelled = 0
        for tid, task in _validation_tasks.items():
            if task.get("status") == "running":
                task["cancel"] = True
                _validation_tasks[tid] = task
                cancelled += 1
        return {"status": "ok", "cancelled": cancelled}
    task = _get_validation_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    task["cancel"] = True
    _set_validation_task(task_id, task)
    return {"status": "cancelling", "message": "Cancellation requested. Will stop after current batch."}

@app.post("/api/validate-all-links")
def validate_all_links(parallel_workers: int = 3, batch_size: int = 100):
    """Start background validation of ALL content URLs. Returns task_id for polling."""
    if parallel_workers < 1 or parallel_workers > 100:
        raise HTTPException(status_code=400, detail="Parallel workers must be between 1 and 100")
    if batch_size < 1 or batch_size > 500:
        raise HTTPException(status_code=400, detail="Batch size must be between 1 and 500")

    task_id = str(uuid.uuid4())[:8]

    # Count total to validate upfront
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""SELECT COUNT(*) as cnt FROM pa.content_urls_joep c
            LEFT JOIN pa.link_validation_results v ON c.url = v.content_url
            WHERE v.content_url IS NULL""")
        total_to_validate = cur.fetchone()['cnt']
        cur.close()
        return_db_connection(conn)
    except Exception:
        total_to_validate = 0

    _set_validation_task(task_id, {"status": "running", "validated": 0, "total_to_validate": total_to_validate, "urls_corrected": 0, "moved_to_pending": 0})

    def run_validation():
        try:
            total_validated = 0
            total_urls_corrected = 0
            total_moved_to_pending = 0

            while True:
                # Check for cancellation
                task_state = _get_validation_task(task_id)
                if task_state and task_state.get("cancel"):
                    _set_validation_task(task_id, {"status": "cancelled", "total_to_validate": total_to_validate, "validated": total_validated, "urls_corrected": total_urls_corrected, "moved_to_pending": total_moved_to_pending})
                    print(f"[VALIDATE-ALL] Cancelled at {total_validated} URLs.")
                    return

                conn = get_db_connection()
                cur = conn.cursor()

                cur.execute("""
                    SELECT c.url, c.content
                    FROM pa.content_urls_joep c
                    LEFT JOIN pa.link_validation_results v ON c.url = v.content_url
                    WHERE v.content_url IS NULL
                    LIMIT %s
                """, (batch_size,))

                rows = cur.fetchall()

                if not rows:
                    cur.close()
                    return_db_connection(conn)
                    break

                content_items = [(row['url'], row['content']) for row in rows]

                with ThreadPoolExecutor(max_workers=parallel_workers) as executor:
                    validation_results = list(executor.map(validate_single_content_es, content_items))

                urls_corrected = 0
                moved_to_pending = 0
                urls_with_gone_products = []
                urls_to_update_content = []
                url_to_gone_details = {}

                for validation_result in validation_results:
                    content_url = validation_result['content_url']
                    has_replaced = len(validation_result['replaced_urls']) > 0
                    has_gone = len(validation_result['gone_urls']) > 0
                    total_links = len(validation_result['valid_urls']) + len(validation_result['replaced_urls']) + len(validation_result['gone_urls'])

                    cur.execute("""
                        INSERT INTO pa.link_validation_results
                        (content_url, total_links, broken_links, valid_links, broken_link_details)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (
                        content_url, total_links,
                        len(validation_result['gone_urls']),
                        len(validation_result['valid_urls']) + len(validation_result['replaced_urls']),
                        json.dumps({'gone_urls': validation_result['gone_urls'], 'replaced_urls': validation_result['replaced_urls']})
                    ))

                    if has_replaced and not has_gone:
                        # Order matches `UPDATE ... SET content = %s WHERE url = %s` below.
                        urls_to_update_content.append((validation_result['corrected_content'], content_url))
                        urls_corrected += 1

                    if has_gone:
                        urls_with_gone_products.append(content_url)
                        url_to_gone_details[content_url] = validation_result['gone_urls']
                        moved_to_pending += 1

                if urls_to_update_content:
                    cur.executemany("""
                        UPDATE pa.content_urls_joep SET content = %s WHERE url = %s
                    """, urls_to_update_content)

                if urls_with_gone_products:
                    placeholders = ','.join(['%s'] * len(urls_with_gone_products))
                    cur.execute(f"SELECT url, content, created_at FROM pa.content_urls_joep WHERE url IN ({placeholders})", urls_with_gone_products)
                    for row in cur.fetchall():
                        gone_urls = url_to_gone_details.get(row['url'], [])
                        cur.execute("INSERT INTO pa.content_history (url, content, reset_reason, reset_details, original_created_at) VALUES (%s, %s, %s, %s, %s)",
                            (row['url'], row['content'], 'gone_products', json.dumps({'gone_urls': gone_urls}), row['created_at']))
                    cur.execute(f"DELETE FROM pa.content_urls_joep WHERE url IN ({placeholders})", urls_with_gone_products)
                    cur.execute(f"DELETE FROM pa.jvs_seo_werkvoorraad_kopteksten_check WHERE url IN ({placeholders})", urls_with_gone_products)
                    cur.execute(f"DELETE FROM pa.url_validation_tracking WHERE url IN ({placeholders})", urls_with_gone_products)
                    cur.executemany("INSERT INTO pa.jvs_seo_werkvoorraad (url, kopteksten) VALUES (%s, 0) ON CONFLICT (url) DO UPDATE SET kopteksten = 0",
                        [(url,) for url in urls_with_gone_products])

                conn.commit()
                cur.close()
                return_db_connection(conn)

                total_validated += len(rows)
                total_urls_corrected += urls_corrected
                total_moved_to_pending += moved_to_pending
                print(f"[VALIDATE-ALL] Batch complete: {len(rows)} validated, {urls_corrected} corrected, {moved_to_pending} moved to pending. Total so far: {total_validated}")
                _set_validation_task(task_id, {"status": "running", "total_to_validate": total_to_validate, "validated": total_validated, "urls_corrected": total_urls_corrected, "moved_to_pending": total_moved_to_pending})

            _set_validation_task(task_id, {
                "status": "completed",
                "validated": total_validated,
                "urls_corrected": total_urls_corrected,
                "moved_to_pending": total_moved_to_pending
            })
        except Exception as e:
            _set_validation_task(task_id, {"status": "error", "error": str(e)})

    threading.Thread(target=run_validation, daemon=True).start()
    return {"task_id": task_id, "status": "started", "message": "Validation started in background. Poll /api/validate-all-links/status/{task_id} for progress."}

@app.get("/api/validation-history")
async def get_validation_history(limit: int = 20):
    """Get history of link validation results"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT
                content_url,
                total_links,
                broken_links,
                valid_links,
                broken_link_details,
                validated_at
            FROM pa.link_validation_results
            ORDER BY validated_at DESC
            LIMIT %s
        """, (limit,))
        rows = cur.fetchall()

        cur.close()
        return_db_connection(conn)

        return {
            "status": "success",
            "results": rows
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/validation-history/reset")
async def reset_validation_history():
    """Reset all validation history - allows re-validation of all URLs"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Get count before deletion
        cur.execute("SELECT COUNT(*) as count FROM pa.link_validation_results")
        count = cur.fetchone()['count']

        # Delete all validation history
        cur.execute("DELETE FROM pa.link_validation_results")
        conn.commit()

        cur.close()
        return_db_connection(conn)

        return {
            "status": "success",
            "message": f"Reset validation history for {count} URLs",
            "cleared_count": count
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/recheck-skipped-urls/status/{task_id}")
def get_recheck_status(task_id: str):
    """Poll progress of a running recheck task."""
    task = _get_validation_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task

@app.post("/api/recheck-skipped-urls/cancel/{task_id}")
def cancel_recheck(task_id: str):
    """Cancel a running recheck task.  Use task_id='all' to cancel every running task."""
    if task_id == "all":
        cancelled = 0
        for tid, task in _validation_tasks.items():
            if task.get("status") == "running":
                task["cancel"] = True
                _validation_tasks[tid] = task
                cancelled += 1
        return {"status": "ok", "cancelled": cancelled}
    task = _get_validation_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    task["cancel"] = True
    _set_validation_task(task_id, task)
    return {"status": "cancelling"}

@app.post("/api/recheck-skipped-urls")
def recheck_skipped_urls(parallel_workers: int = 3, batch_size: int = 50):
    """Start background re-check of skipped URLs. Returns task_id for polling."""
    from concurrent.futures import ThreadPoolExecutor

    if parallel_workers < 1 or parallel_workers > 100:
        raise HTTPException(status_code=400, detail="Parallel workers must be between 1 and 100")
    if batch_size < 1 or batch_size > 500:
        raise HTTPException(status_code=400, detail="Batch size must be between 1 and 500")

    task_id = str(uuid.uuid4())[:8]

    # Count total to recheck
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""SELECT COUNT(*) as cnt FROM pa.url_validation_tracking
            WHERE status = 'skipped' AND (skip_reason IS NULL OR skip_reason NOT LIKE '%%(rechecked)%%')""")
        total_to_recheck = cur.fetchone()['cnt']
        cur.close()
        return_db_connection(conn)
    except Exception:
        total_to_recheck = 0

    _set_validation_task(task_id, {"status": "running", "rechecked": 0, "total_to_recheck": total_to_recheck, "now_eligible": 0})

    def run_recheck():
      try:
        total_rechecked = 0
        total_now_eligible = 0

        print(f"[RECHECK-SKIPPED] Starting re-check of skipped URLs...")

        while True:
            task_state = _get_validation_task(task_id)
            if task_state and task_state.get("cancel"):
                _set_validation_task(task_id, {"status": "cancelled", "rechecked": total_rechecked, "total_to_recheck": total_to_recheck, "now_eligible": total_now_eligible})
                print(f"[RECHECK-SKIPPED] Cancelled at {total_rechecked}.")
                return
            conn = get_db_connection()
            cur = conn.cursor()

            # Get batch of skipped URLs from shared table (only those not yet rechecked)
            cur.execute("""
                SELECT url FROM pa.url_validation_tracking
                WHERE status = 'skipped'
                  AND (skip_reason IS NULL OR skip_reason NOT LIKE '%%(rechecked)%%')
                LIMIT %s
            """, (batch_size,))
            skipped_rows = cur.fetchall()

            if not skipped_rows:
                cur.close()
                return_db_connection(conn)
                break

            skipped_urls = [row['url'] for row in skipped_rows]

            def check_single_url(url):
                """Check if a URL now has products"""
                try:
                    scraped_data = scrape_product_page_api(url)
                    has_products = scraped_data and scraped_data.get('products') and len(scraped_data['products']) > 0
                    return {'url': url, 'has_products': has_products}
                except Exception as e:
                    print(f"[RECHECK-SKIPPED] Error checking {url}: {e}")
                    return {'url': url, 'has_products': False}

            # Check URLs in parallel
            with ThreadPoolExecutor(max_workers=parallel_workers) as executor:
                results = list(executor.map(check_single_url, skipped_urls))

            now_eligible = [r['url'] for r in results if r['has_products']]
            still_skipped = [r['url'] for r in results if not r['has_products']]

            # Remove eligible URLs from shared table + both feature tables so they can be reprocessed
            if now_eligible:
                placeholders = ','.join(['%s'] * len(now_eligible))
                cur.execute(f"""
                    DELETE FROM pa.url_validation_tracking
                    WHERE url IN ({placeholders})
                """, now_eligible)
                cur.execute(f"""
                    DELETE FROM pa.jvs_seo_werkvoorraad_kopteksten_check
                    WHERE url IN ({placeholders})
                """, now_eligible)
                cur.execute(f"""
                    DELETE FROM pa.faq_tracking
                    WHERE url IN ({placeholders})
                      AND (skip_reason IS NULL OR skip_reason != 'main_category_url')
                """, now_eligible)
                # Make sure they're in werkvoorraad for reprocessing
                cur.executemany("""
                    INSERT INTO pa.jvs_seo_werkvoorraad (url, kopteksten)
                    VALUES (%s, 0)
                    ON CONFLICT (url) DO NOTHING
                """, [(url,) for url in now_eligible])
                total_now_eligible += len(now_eligible)

            # Mark remaining as rechecked in shared table to avoid infinite loop
            if still_skipped:
                placeholders = ','.join(['%s'] * len(still_skipped))
                cur.execute(f"""
                    UPDATE pa.url_validation_tracking
                    SET skip_reason = 'no_products_found (rechecked)'
                    WHERE url IN ({placeholders})
                """, still_skipped)

            conn.commit()
            cur.close()
            return_db_connection(conn)

            total_rechecked += len(skipped_urls)
            print(f"[RECHECK-SKIPPED] Batch: {len(skipped_urls)} checked, {len(now_eligible)} now eligible. Total: {total_rechecked}")
            _set_validation_task(task_id, {"status": "running", "rechecked": total_rechecked, "total_to_recheck": total_to_recheck, "now_eligible": total_now_eligible})

        print(f"[RECHECK-SKIPPED] Complete. Rechecked: {total_rechecked}, Now eligible: {total_now_eligible}")
        _set_validation_task(task_id, {"status": "completed", "rechecked": total_rechecked, "total_to_recheck": total_to_recheck, "now_eligible": total_now_eligible})
      except Exception as e:
        _set_validation_task(task_id, {"status": "error", "error": str(e)})

    threading.Thread(target=run_recheck, daemon=True).start()
    return {"task_id": task_id, "status": "started"}


@app.delete("/api/recheck-skipped-urls/reset")
def reset_skipped_recheck():
    """
    Reset the 'rechecked' marker on skipped URLs, allowing them to be rechecked again.
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            UPDATE pa.url_validation_tracking
            SET skip_reason = 'no_products_found'
            WHERE status = 'skipped' AND skip_reason LIKE '%%(rechecked)%%'
        """)
        count = cur.rowcount

        conn.commit()
        cur.close()
        return_db_connection(conn)

        return {
            "status": "success",
            "message": f"Reset {count} URLs to allow rechecking",
            "reset_count": count
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# FAQ GENERATION ENDPOINTS
# ============================================================================

def process_single_url_faq_wrapper(args: tuple) -> dict:
    """Wrapper for process_single_url_faq to work with ThreadPoolExecutor"""
    url, num_faqs = args
    return process_single_url_faq(url, num_faqs)


@app.get("/api/faq/status")
def get_faq_status():
    """Get FAQ processing status and counts"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Count pending directly (same logic as process_faq_urls query)
        cur.execute("""
            SELECT
                (SELECT COUNT(*) FROM pa.jvs_seo_werkvoorraad) as total,
                (SELECT COUNT(*) FROM pa.faq_content) as processed,
                (SELECT COUNT(*) FROM pa.url_validation_tracking WHERE status = 'skipped') as skipped,
                (SELECT COUNT(*) FROM pa.faq_tracking WHERE status = 'failed') as failed,
                (SELECT COUNT(*) FROM pa.jvs_seo_werkvoorraad w
                 WHERE NOT EXISTS (SELECT 1 FROM pa.url_validation_tracking v WHERE v.url = w.url)
                   AND NOT EXISTS (SELECT 1 FROM pa.faq_tracking t WHERE t.url = w.url AND t.status != 'pending')) as pending
        """)
        counts = cur.fetchone()
        total = counts['total']
        processed = counts['processed']
        skipped = counts['skipped']
        failed = counts['failed']
        pending = counts['pending']

        # Get recent FAQ results
        try:
            cur.execute("""
                SELECT url, page_title, faq_json, schema_org, created_at
                FROM pa.faq_content
                ORDER BY created_at DESC NULLS LAST
                LIMIT 5
            """)
            recent_rows = cur.fetchall()
            recent = [{
                'url': r['url'],
                'page_title': r['page_title'],
                'faq_json': r['faq_json'],
                'schema_org': r['schema_org'],
                'created_at': r['created_at'].isoformat() if r.get('created_at') else None
            } for r in recent_rows]
        except Exception as e:
            print(f"[FAQ] Failed to get recent results: {e}")
            recent = []

        cur.close()
        return_db_connection(conn)

        return {
            "total_urls": total,
            "processed": processed,
            "skipped": skipped,
            "failed": failed,
            "pending": pending,
            "recent_results": recent
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/faq/batch-start")
def start_faq_batch_endpoint(num_faqs: int = 6):
    """Start FAQ batch processing via OpenAI Batch API."""
    return start_faq_batch(num_faqs)

@app.get("/api/faq/batch-status")
def get_faq_batch_status():
    """Get FAQ batch processing status."""
    return get_batch_status("faq")

@app.post("/api/batch-start")
def start_kopteksten_batch_endpoint():
    """Start kopteksten batch processing via OpenAI Batch API."""
    return start_kopteksten_batch()

@app.get("/api/batch-status")
def get_kopteksten_batch_status():
    """Get kopteksten batch processing status."""
    return get_batch_status("kopteksten")

@app.post("/api/faq/process-urls")
def process_faq_urls(batch_size: int = 10, parallel_workers: int = 3, num_faqs: int = 6):
    """
    Process batch of URLs for FAQ generation.

    Args:
        batch_size: Number of URLs to process
        parallel_workers: Number of parallel workers (1-10)
        num_faqs: Number of FAQ items to generate per page (default 5)
    """
    print(f"[FAQ] process_faq_urls called - batch_size={batch_size}, workers={parallel_workers}, num_faqs={num_faqs}")

    try:
        # Validate parameters
        if batch_size < 1:
            raise HTTPException(status_code=400, detail="Batch size must be at least 1")

        if parallel_workers < 1 or parallel_workers > 100:
            raise HTTPException(status_code=400, detail="Parallel workers must be between 1 and 100")

        if num_faqs < 1 or num_faqs > 10:
            raise HTTPException(status_code=400, detail="Number of FAQs must be between 1 and 10")

        # Get unprocessed URLs from local PostgreSQL
        conn = get_db_connection()
        cur = conn.cursor()

        # Fetch unprocessed URLs (not in FAQ tracking, not in shared validation, or pending)
        cur.execute("""
            SELECT w.url
            FROM pa.jvs_seo_werkvoorraad w
            WHERE NOT EXISTS (SELECT 1 FROM pa.url_validation_tracking v WHERE v.url = w.url)
              AND NOT EXISTS (SELECT 1 FROM pa.faq_tracking t WHERE t.url = w.url AND t.status != 'pending')
            LIMIT %s
        """, (batch_size,))

        rows = cur.fetchall()
        cur.close()
        return_db_connection(conn)

        if not rows:
            return {
                "status": "complete",
                "message": "No URLs to process",
                "processed": 0
            }

        urls = [row['url'] for row in rows]

        # Process URLs in parallel using ThreadPoolExecutor
        url_args = [(url, num_faqs) for url in urls]
        with ThreadPoolExecutor(max_workers=parallel_workers) as executor:
            results = list(executor.map(process_single_url_faq_wrapper, url_args))

        # Save results to database using batch inserts for better performance
        conn = get_db_connection()
        cur = conn.cursor()

        # Prepare batch data — skipped (no_products_found) goes to shared table, rest to faq_tracking
        tracking_data = []
        shared_skip_data = []
        content_data = []

        for result in results:
            url = result['url']
            status = result['status']
            reason = result.get('reason')
            truncated_reason = reason[:255] if reason and len(reason) > 255 else reason

            if status == 'skipped' and reason and 'no_products_found' in reason:
                shared_skip_data.append((url, status, truncated_reason))
            else:
                tracking_data.append((url, status, truncated_reason))

            if status == 'success':
                content_data.append((
                    url,
                    result.get('page_title', ''),
                    result.get('faq_json', ''),
                    result.get('schema_org', '')
                ))

            print(f"[FAQ] {url} - Status: {status}" + (f" - Reason: {reason}" if reason else f" - {result.get('faq_count', 0)} FAQs"))

        # Batch insert shared skip data (no_products_found)
        if shared_skip_data:
            cur.executemany("""
                INSERT INTO pa.url_validation_tracking (url, status, skip_reason)
                VALUES (%s, %s, %s)
                ON CONFLICT (url) DO UPDATE SET status = EXCLUDED.status, skip_reason = EXCLUDED.skip_reason, checked_at = CURRENT_TIMESTAMP
            """, shared_skip_data)

        # Batch insert feature-specific tracking data (success, failed)
        if tracking_data:
            cur.executemany("""
                INSERT INTO pa.faq_tracking (url, status, skip_reason)
                VALUES (%s, %s, %s)
                ON CONFLICT (url) DO UPDATE SET status = EXCLUDED.status, skip_reason = EXCLUDED.skip_reason
            """, tracking_data)

        # Batch insert content data
        if content_data:
            cur.executemany("""
                INSERT INTO pa.faq_content (url, page_title, faq_json, schema_org)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (url) DO UPDATE SET page_title = EXCLUDED.page_title, faq_json = EXCLUDED.faq_json, schema_org = EXCLUDED.schema_org
            """, content_data)

        conn.commit()
        cur.close()
        return_db_connection(conn)

        processed_count = sum(1 for r in results if r['status'] == 'success')
        skipped_count = sum(1 for r in results if r['status'] == 'skipped')
        failed_count = sum(1 for r in results if r['status'] == 'failed')

        print(f"[FAQ BATCH COMPLETE] Processed: {processed_count}/{len(urls)} | Skipped: {skipped_count} | Failed: {failed_count}")

        return {
            "status": "success",
            "processed": processed_count,
            "skipped": skipped_count,
            "failed": failed_count,
            "total_attempted": len(results),
            "results": results
        }

    except Exception as e:
        print(f"[FAQ ERROR] process_faq_urls failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


def faq_json_to_html(faq_json_str: str) -> str:
    """Convert FAQ JSON to HTML with bold questions and regular answers."""
    if not faq_json_str:
        return ''
    try:
        faqs = json.loads(faq_json_str)
        html_parts = []
        for faq in faqs:
            question = faq.get('question', '')
            answer = faq.get('answer', '')
            html_parts.append(f'<b>{question}</b><br>{answer}')
        content = '<br><br>'.join(html_parts)
        return f'<br />{content}<br />' if content else ''
    except (json.JSONDecodeError, TypeError):
        return ''


@app.get("/api/faq/export/xlsx")
async def export_faq_xlsx():
    """Export all generated FAQs as Excel XLSX"""
    from openpyxl import Workbook

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT url, page_title, faq_json, schema_org
            FROM pa.faq_content
        """)
        rows = cur.fetchall()

        cur.close()
        return_db_connection(conn)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "FAQ Export"

        # Add headers (matching the format user requested: url, content_faq, content_bottom)
        ws.append(['url', 'content_faq', 'content_bottom'])

        # Add data rows
        import re
        # Remove control characters that Excel doesn't allow
        illegal_chars = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

        for row in rows:
            # JSON-LD for content_faq column (raw JSON, no script tags)
            schema_org = row['schema_org'] if row['schema_org'] else '{}'
            content_faq = illegal_chars.sub('', schema_org)

            # Build HTML for content_bottom column
            content_bottom = faq_json_to_html(row['faq_json'])
            content_bottom = illegal_chars.sub('', content_bottom)

            ws.append([row['url'], content_faq, content_bottom])

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 80
        ws.column_dimensions["B"].width = 100
        ws.column_dimensions["C"].width = 100

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=faq_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/faq/export/json")
async def export_faq_json():
    """Export all generated FAQs as JSON"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT url, page_title, faq_json, schema_org
            FROM pa.faq_content
        """)
        rows = cur.fetchall()

        cur.close()
        return_db_connection(conn)

        # Convert to JSON-serializable format
        data = []
        for row in rows:
            data.append({
                'url': row['url'],
                'page_title': row['page_title'],
                'faqs': json.loads(row['faq_json']) if row['faq_json'] else [],
                'schema_org': json.loads(row['schema_org']) if row['schema_org'] else {}
            })

        # Return as downloadable file
        json_str = json.dumps(data, indent=2, ensure_ascii=False)
        return StreamingResponse(
            iter([json_str]),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename=faq_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/faq/validate-links")
def validate_faq_links_endpoint(batch_size: int = 100, parallel_workers: int = 3):
    """
    Validate hyperlinks in FAQ content using Elasticsearch lookup.

    - Only validates FAQs that haven't been validated yet
    - Checks all product links (/p/) in FAQ answers
    - If product is gone, resets FAQ URL to pending for regeneration
    - Records validation results to avoid re-validating

    Args:
        batch_size: Number of FAQs to validate per batch (default: 100)
        parallel_workers: Number of parallel workers (default: 3)
    """
    from backend.link_validator import validate_faq_links, reset_faq_to_pending
    from concurrent.futures import ThreadPoolExecutor, as_completed

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Get FAQs that haven't been validated yet (LEFT JOIN)
        cur.execute("""
            SELECT c.url, c.faq_json
            FROM pa.faq_content c
            LEFT JOIN pa.faq_validation_results v ON c.url = v.url
            WHERE c.faq_json IS NOT NULL
              AND v.url IS NULL
            LIMIT %s
        """, (batch_size,))
        rows = cur.fetchall()

        cur.close()
        return_db_connection(conn)

        if not rows:
            return {
                "status": "success",
                "message": "No unvalidated FAQs found",
                "validated": 0,
                "reset_to_pending": 0
            }

        # Validate each FAQ and collect results
        all_urls_with_gone = []
        total_links = 0
        total_valid = 0
        total_gone = 0
        validation_records = []

        def validate_single(row):
            url = row['url']
            faq_json = row['faq_json']
            result = validate_faq_links(faq_json)
            return {
                'url': url,
                'total_links': result['total_links'],
                'valid_links': result['valid_links'],
                'gone_links': result['gone_links'],
                'has_gone': result['has_gone_links']
            }

        # Process in parallel
        with ThreadPoolExecutor(max_workers=parallel_workers) as executor:
            futures = {executor.submit(validate_single, row): row for row in rows}

            for future in as_completed(futures):
                result = future.result()
                total_links += result['total_links']
                total_valid += result['valid_links']
                total_gone += len(result['gone_links']) if isinstance(result['gone_links'], list) else result['gone_links']

                validation_records.append(result)

                if result['has_gone']:
                    all_urls_with_gone.append(result['url'])

        # Record validation results (for URLs that passed - no gone products)
        conn = get_db_connection()
        cur = conn.cursor()
        for record in validation_records:
            if not record['has_gone']:  # Only record if no gone products
                cur.execute("""
                    INSERT INTO pa.faq_validation_results (url, total_links, valid_links, gone_links)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (url) DO UPDATE SET
                        total_links = EXCLUDED.total_links,
                        valid_links = EXCLUDED.valid_links,
                        gone_links = EXCLUDED.gone_links,
                        validated_at = CURRENT_TIMESTAMP
                """, (record['url'], record['total_links'], record['valid_links'], 0))
        conn.commit()
        cur.close()
        return_db_connection(conn)

        # Reset FAQs with gone products to pending
        reset_count = 0
        if all_urls_with_gone:
            reset_count = reset_faq_to_pending(all_urls_with_gone)

        return {
            "status": "success",
            "validated": len(rows),
            "total_links_checked": total_links,
            "valid_links": total_valid,
            "gone_links": total_gone,
            "faqs_with_gone_products": len(all_urls_with_gone),
            "reset_to_pending": reset_count,
            "urls_reset": all_urls_with_gone[:20]  # Show first 20 for debugging
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/faq/validate-all-links/status/{task_id}")
def get_faq_validate_all_status(task_id: str):
    """Poll progress of a running FAQ validate-all task."""
    task = _get_validation_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task

@app.post("/api/faq/validate-all-links/cancel/{task_id}")
def cancel_faq_validate_all(task_id: str):
    """Cancel a running FAQ validate-all task.  Use task_id='all' to cancel every running task."""
    if task_id == "all":
        cancelled = 0
        for tid, task in _validation_tasks.items():
            if task.get("status") == "running":
                task["cancel"] = True
                _validation_tasks[tid] = task
                cancelled += 1
        return {"status": "ok", "cancelled": cancelled}
    task = _get_validation_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    task["cancel"] = True
    _set_validation_task(task_id, task)
    return {"status": "cancelling", "message": "Cancellation requested. Will stop after current batch."}

@app.post("/api/faq/validate-all-links")
def validate_all_faq_links(parallel_workers: int = 3, batch_size: int = 500):
    """Start background validation of ALL FAQ links. Returns task_id for polling."""
    from backend.link_validator import validate_faq_links, reset_faq_to_pending
    from concurrent.futures import ThreadPoolExecutor, as_completed

    if parallel_workers < 1 or parallel_workers > 100:
        raise HTTPException(status_code=400, detail="Parallel workers must be between 1 and 100")
    if batch_size < 1 or batch_size > 1000:
        raise HTTPException(status_code=400, detail="Batch size must be between 1 and 1000")

    task_id = str(uuid.uuid4())[:8]

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""SELECT COUNT(*) as cnt FROM pa.faq_content c
            LEFT JOIN pa.faq_validation_results v ON c.url = v.url
            WHERE c.faq_json IS NOT NULL AND v.url IS NULL""")
        total_to_validate = cur.fetchone()['cnt']
        cur.close()
        return_db_connection(conn)
    except Exception:
        total_to_validate = 0

    _set_validation_task(task_id, {"status": "running", "validated": 0, "total_to_validate": total_to_validate, "total_links_checked": 0, "gone_links": 0, "reset_to_pending": 0})

    def run_faq_validation():
      try:
        total_validated = 0
        total_reset = 0
        total_links_checked = 0
        total_gone_links = 0

        while True:
            # Check for cancellation
            task_state = _get_validation_task(task_id)
            if task_state and task_state.get("cancel"):
                _set_validation_task(task_id, {"status": "cancelled", "total_to_validate": total_to_validate, "validated": total_validated, "total_links_checked": total_links_checked, "gone_links": total_gone_links, "reset_to_pending": total_reset})
                print(f"[FAQ-VALIDATE] Cancelled at {total_validated} FAQs.")
                return

            conn = get_db_connection()
            cur = conn.cursor()

            cur.execute("""
                SELECT c.url, c.faq_json
                FROM pa.faq_content c
                LEFT JOIN pa.faq_validation_results v ON c.url = v.url
                WHERE c.faq_json IS NOT NULL
                  AND v.url IS NULL
                LIMIT %s
            """, (batch_size,))
            rows = cur.fetchall()

            cur.close()
            return_db_connection(conn)

            if not rows:
                break

            # Validate each FAQ
            batch_urls_with_gone = []
            validation_records = []

            def validate_single(row):
                url = row['url']
                faq_json = row['faq_json']
                result = validate_faq_links(faq_json)
                return {
                    'url': url,
                    'total_links': result['total_links'],
                    'valid_links': result['valid_links'],
                    'gone_links': result['gone_links'],
                    'has_gone': result['has_gone_links']
                }

            with ThreadPoolExecutor(max_workers=parallel_workers) as executor:
                futures = {executor.submit(validate_single, row): row for row in rows}

                for future in as_completed(futures):
                    result = future.result()
                    total_links_checked += result['total_links']
                    gone_count = len(result['gone_links']) if isinstance(result['gone_links'], list) else result['gone_links']
                    total_gone_links += gone_count

                    validation_records.append(result)

                    if result['has_gone']:
                        batch_urls_with_gone.append(result['url'])

            # Record validation results (for URLs that passed)
            conn = get_db_connection()
            cur = conn.cursor()
            for record in validation_records:
                if not record['has_gone']:
                    cur.execute("""
                        INSERT INTO pa.faq_validation_results (url, total_links, valid_links, gone_links)
                        VALUES (%s, %s, %s, %s)
                        ON CONFLICT (url) DO UPDATE SET
                            total_links = EXCLUDED.total_links,
                            valid_links = EXCLUDED.valid_links,
                            gone_links = EXCLUDED.gone_links,
                            validated_at = CURRENT_TIMESTAMP
                    """, (record['url'], record['total_links'], record['valid_links'], 0))
            conn.commit()
            cur.close()
            return_db_connection(conn)

            # Reset FAQs with gone products
            if batch_urls_with_gone:
                reset_count = reset_faq_to_pending(batch_urls_with_gone)
                total_reset += reset_count

            total_validated += len(rows)
            print(f"[FAQ-VALIDATE] Batch complete: {len(rows)} validated, {len(batch_urls_with_gone)} reset. Total: {total_validated}")
            _set_validation_task(task_id, {"status": "running", "total_to_validate": total_to_validate, "validated": total_validated, "total_links_checked": total_links_checked, "gone_links": total_gone_links, "reset_to_pending": total_reset})

        _set_validation_task(task_id, {
            "status": "completed",
            "validated": total_validated,
            "total_links_checked": total_links_checked,
            "gone_links": total_gone_links,
            "reset_to_pending": total_reset
        })
      except Exception as e:
        _set_validation_task(task_id, {"status": "error", "error": str(e)})

    threading.Thread(target=run_faq_validation, daemon=True).start()
    return {"task_id": task_id, "status": "started", "message": "FAQ validation started in background. Poll /api/faq/validate-all-links/status/{task_id} for progress."}


@app.get("/api/faq/lookup")
async def lookup_faq(url: str):
    """Look up FAQ content for a specific URL."""
    try:
        clean_url = url.strip().lower()
        base_url = "https://www.beslist.nl"
        if clean_url.startswith('http'):
            if 'beslist.nl' in clean_url:
                path_url = '/' + clean_url.split('beslist.nl', 1)[-1].lstrip('/')
            else:
                path_url = clean_url
        else:
            if not clean_url.startswith('/'):
                clean_url = '/' + clean_url
            path_url = clean_url

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT url, page_title, faq_json, schema_org, created_at
            FROM pa.faq_content
            WHERE url = %s OR url = %s
            LIMIT 1
        """, (base_url + path_url, path_url))
        row = cur.fetchone()
        cur.close()
        return_db_connection(conn)

        if not row:
            return {"found": False, "url": clean_url, "message": "URL not found in FAQ database"}

        return {
            "found": True,
            "url": row['url'],
            "page_title": row['page_title'],
            "faq_json": row['faq_json'],
            "schema_org": row['schema_org'],
            "created_at": row['created_at'].isoformat() if row.get('created_at') else None
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/faq/result/{url:path}")
async def delete_faq_result(url: str):
    """Delete FAQ content and reset URL to pending."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM pa.faq_content WHERE url = %s", (url,))
        cur.execute("DELETE FROM pa.faq_tracking WHERE url = %s", (url,))
        cur.execute("DELETE FROM pa.faq_validation_results WHERE url = %s", (url,))
        deleted = cur.rowcount
        conn.commit()
        cur.close()
        return_db_connection(conn)
        return {"status": "success", "message": f"FAQ deleted and URL reset to pending", "url": url}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/faq/validation-history/reset")
def reset_faq_validation_history():
    """
    Reset all FAQ validation history to allow re-validation of all FAQs.
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("DELETE FROM pa.faq_validation_results")
        deleted = cur.rowcount

        conn.commit()
        cur.close()
        return_db_connection(conn)

        return {
            "status": "success",
            "message": f"Reset validation history for {deleted} FAQs",
            "deleted": deleted,
            "cleared_count": deleted
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/faq/recheck-skipped-urls")
def recheck_skipped_faq_urls(parallel_workers: int = 3, batch_size: int = 50):
    """
    Re-check all skipped URLs (shared table) to see if they're now eligible.
    Delegates to the shared recheck endpoint.
    """
    return recheck_skipped_urls(parallel_workers=parallel_workers, batch_size=batch_size)


@app.delete("/api/faq/recheck-skipped-urls/reset")
def reset_faq_skipped_recheck():
    """
    Reset the 'rechecked' marker on skipped URLs. Delegates to shared endpoint.
    """
    return reset_skipped_recheck()


@app.get("/api/export/combined/xlsx")
async def export_combined_xlsx():
    """
    Export combined FAQ and content_top results as Excel XLSX.
    Columns: url, content_faq, content_top, content_bottom, country_language
    Includes ALL URLs from both tables (empty cells where data is missing).
    """
    from openpyxl import Workbook

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Full outer join to get ALL URLs from both tables
        cur.execute("""
            SELECT
                COALESCE(f.url, c.url) as url,
                f.schema_org as content_faq,
                f.faq_json,
                c.content as content_top
            FROM pa.faq_content f
            FULL OUTER JOIN pa.content_urls_joep c ON f.url = c.url
            ORDER BY COALESCE(f.url, c.url)
        """)
        rows = cur.fetchall()

        cur.close()
        return_db_connection(conn)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Combined Export"

        # Add headers
        ws.append(['url', 'content_faq', 'content_top', 'content_bottom', 'country_language'])

        # Add data rows
        import re
        # Remove control characters that Excel doesn't allow
        illegal_chars = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

        for row in rows:
            # JSON-LD for content_faq column (raw JSON, no script tags)
            if row['content_faq']:
                content_faq = illegal_chars.sub('', row['content_faq'])
            else:
                content_faq = ''

            content_top = row['content_top'] if row['content_top'] else ''
            if content_top:
                content_top = illegal_chars.sub('', content_top)

            # Build HTML for content_bottom column (empty if no FAQ data)
            content_bottom = faq_json_to_html(row['faq_json']) if row['faq_json'] else ''
            if content_bottom:
                content_bottom = illegal_chars.sub('', content_bottom)

            ws.append([row['url'], content_faq, content_top, content_bottom, 'nl-nl'])

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 80
        ws.column_dimensions["B"].width = 100
        ws.column_dimensions["C"].width = 100
        ws.column_dimensions["D"].width = 100
        ws.column_dimensions["E"].width = 15

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=combined_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/faq/result/{url:path}")
async def delete_faq_result(url: str):
    """Delete a FAQ result and reset the URL back to pending state"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Delete from FAQ content table
        cur.execute("""
            DELETE FROM pa.faq_content
            WHERE url = %s
        """, (url,))

        # Delete from FAQ tracking table and shared validation table
        cur.execute("""
            DELETE FROM pa.faq_tracking
            WHERE url = %s
        """, (url,))
        cur.execute("""
            DELETE FROM pa.url_validation_tracking
            WHERE url = %s
        """, (url,))

        conn.commit()
        cur.close()
        return_db_connection(conn)

        return {
            "status": "success",
            "message": f"FAQ result deleted and URL reset to pending",
            "url": url
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Content Publishing API
# ============================================================================

@app.get("/api/content-publish/stats")
async def get_content_publish_stats():
    """Get statistics about content available for publishing."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Get counts
        cur.execute("""
            SELECT
                (SELECT COUNT(*) FROM pa.content_urls_joep WHERE content IS NOT NULL) as content_top_count,
                (SELECT COUNT(*) FROM pa.faq_content WHERE faq_json IS NOT NULL) as faq_count,
                (SELECT COUNT(DISTINCT COALESCE(c.url, f.url))
                 FROM pa.content_urls_joep c
                 FULL OUTER JOIN pa.faq_content f ON c.url = f.url
                 WHERE c.content IS NOT NULL OR f.faq_json IS NOT NULL) as total_unique_urls
        """)
        row = cur.fetchone()

        cur.close()
        return_db_connection(conn)

        return {
            "content_top_count": row['content_top_count'],
            "faq_count": row['faq_count'],
            "total_unique_urls": row['total_unique_urls']
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/content-publish/preview")
async def preview_content_publish(offset: int = 0, limit: int = 10):
    """Preview content that will be published."""
    try:
        items = get_content_batch(offset, min(limit, 100))
        total = get_total_content_count()

        return {
            "total": total,
            "offset": offset,
            "limit": limit,
            "items": items
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/content-publish/curl")
async def get_content_publish_curl(limit: int = 10, environment: str = "dev"):
    """
    Generate a curl command for publishing content.

    Args:
        limit: Number of items to include (default: 10, max: 100)
        environment: Target environment (dev, staging, production)
    """
    try:
        if environment not in ("dev", "staging", "production"):
            raise HTTPException(status_code=400, detail="Invalid environment. Use: dev, staging, production")
        curl_cmd = generate_curl_command(limit=min(limit, 100), environment=environment)
        return {
            "environment": environment,
            "curl_command": curl_cmd
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/content-publish")
async def publish_content(environment: str = "dev", content_type: str = "all"):
    """
    Publish content to the website-configuration API in a single call.

    Args:
        environment: Target environment (dev, staging, production)
        content_type: What to publish - "all", "seo_only", or "faq_only"
    """
    try:
        if environment not in ("dev", "staging", "production"):
            raise HTTPException(status_code=400, detail="Invalid environment. Use: dev, staging, production")

        if content_type not in ("all", "seo_only", "faq_only"):
            raise HTTPException(status_code=400, detail="Invalid content_type. Use: all, seo_only, faq_only")

        # Start background task for publishing
        task_id = start_publish_task(environment=environment, content_type=content_type)
        return {
            "status": "started",
            "task_id": task_id,
            "environment": environment,
            "content_type": content_type,
            "message": "Publishing started in background. Use /api/content-publish/status/{task_id} to check progress."
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/content-publish/status/{task_id}")
async def get_publish_status(task_id: str):
    """
    Get the status of a content publishing task.
    """
    try:
        status = get_publish_task_status(task_id)
        if "error" in status and status["error"] == "Task not found":
            raise HTTPException(status_code=404, detail="Task not found")
        return status
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/content-publish/last-push")
async def get_last_publish():
    """Get the timestamp of the last successful publish to production."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT published_at, total_urls, content_type, payload_size_mb, duration_sec
            FROM pa.publish_log
            WHERE environment = 'production' AND status = 'success'
            ORDER BY published_at DESC
            LIMIT 1
        """)
        row = cur.fetchone()
        cur.close()
        return_db_connection(conn)

        if row:
            return {
                "last_push": row['published_at'].isoformat(),
                "total_urls": row['total_urls'],
                "content_type": row['content_type'],
                "payload_size_mb": float(row['payload_size_mb']) if row['payload_size_mb'] else None,
                "duration_sec": float(row['duration_sec']) if row['duration_sec'] else None,
            }
        return {"last_push": None}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Unique Titles Publishing API
# ============================================================================

from backend.unique_titles import (
    init_unique_titles_table,
    bulk_upsert_titles,
    get_all_titles,
    get_titles_count,
    generate_csv_for_upload,
    upload_titles_to_api,
    delete_title,
    search_titles,
    queue_urls_for_generation
)

# Initialize table on startup
try:
    init_unique_titles_table()
except Exception as e:
    print(f"[STARTUP] Could not initialize unique_titles table: {e}")

try:
    from backend.facet_classifier import init_facet_classifications_table
    init_facet_classifications_table()
except Exception as e:
    print(f"[STARTUP] Could not initialize facet_type_classifications table: {e}")

# ============================================================================
# SEO Priority tool
# ============================================================================
from backend import seo_prio_service

try:
    seo_prio_service.init_seo_prio_tables()
except Exception as e:
    print(f"[STARTUP] Could not initialize seo_prio tables: {e}")


@app.get("/api/seo-prio/defaults")
async def seo_prio_defaults():
    start, end = seo_prio_service.default_date_range()
    return {
        "start_date": start,
        "end_date": end,
        "thresholds": seo_prio_service.DEFAULT_THRESHOLDS,
    }


@app.post("/api/seo-prio/start")
async def seo_prio_start(payload: dict):
    start_date = payload.get("start_date")
    end_date = payload.get("end_date")
    if not start_date or not end_date:
        s, e = seo_prio_service.default_date_range()
        start_date = start_date or s
        end_date = end_date or e
    thresholds = payload.get("thresholds") or {}
    run_id = seo_prio_service.start_run({
        "start_date": str(start_date),
        "end_date": str(end_date),
        "thresholds": thresholds,
    })
    return {"run_id": run_id}


@app.get("/api/seo-prio/status/{run_id}")
async def seo_prio_status(run_id: str):
    status = seo_prio_service.get_run_status(run_id)
    if not status:
        return JSONResponse({"error": "not found"}, status_code=404)
    return status


@app.post("/api/seo-prio/stop/{run_id}")
async def seo_prio_stop(run_id: str):
    ok = seo_prio_service.stop_run(run_id)
    return {"stopped": ok}


@app.get("/api/seo-prio/runs")
async def seo_prio_runs():
    return {"runs": seo_prio_service.list_runs()}


@app.get("/api/seo-prio/summary/{run_id}")
async def seo_prio_summary(run_id: str):
    return seo_prio_service.get_run_summary(run_id)


@app.get("/api/seo-prio/results/{run_id}")
async def seo_prio_results(run_id: str, limit: int = 1000, offset: int = 0):
    return seo_prio_service.get_run_results(run_id, limit=limit, offset=offset)


@app.delete("/api/seo-prio/runs/{run_id}")
async def seo_prio_delete(run_id: str):
    ok = seo_prio_service.delete_run(run_id)
    if not ok:
        raise HTTPException(status_code=409, detail="Run is still active; stop it first")
    return {"status": "deleted", "run_id": run_id}


@app.get("/api/seo-prio/export/{run_id}")
async def seo_prio_export(run_id: str):
    from fastapi.responses import StreamingResponse
    blob = seo_prio_service.export_excel(run_id)
    return StreamingResponse(
        io.BytesIO(blob),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="seo_prio_{run_id}.xlsx"'},
    )


@app.get("/api/unique-titles/status")
async def get_unique_titles_status():
    """Get statistics about unique titles in database."""
    try:
        count = get_titles_count()
        return {
            "total_titles": count
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/unique-titles/import")
async def import_unique_titles(file: UploadFile = File(...)):
    """
    Import a CSV file with titles into the database.

    The CSV must have semicolon (;) as separator and include columns:
    url, title, description, h1_title
    (Other columns like id, active, created_at, updated_at are ignored)
    """
    try:
        # Read file content
        content = await file.read()

        # Try to detect encoding
        text_content = None
        for encoding in ['utf-8', 'utf-8-sig', 'utf-16', 'utf-16-le', 'windows-1252', 'latin-1']:
            try:
                text_content = content.decode(encoding)
                if '\x00' not in text_content:  # UTF-16 check
                    break
            except (UnicodeDecodeError, UnicodeError):
                continue

        if text_content is None:
            text_content = content.decode('latin-1')

        # Parse CSV
        lines = text_content.strip().split('\n')
        if not lines:
            raise HTTPException(status_code=400, detail="Empty file")

        # Parse header to find column indices
        header = lines[0].split(';')
        header_lower = [h.strip().lower() for h in header]

        # Find required column indices
        try:
            url_idx = header_lower.index('url')
            title_idx = header_lower.index('title')
            desc_idx = header_lower.index('description')
            h1_idx = header_lower.index('h1_title')
        except ValueError as e:
            raise HTTPException(status_code=400, detail=f"Missing required column: {e}")

        # Parse data rows
        titles = []
        for line in lines[1:]:
            if not line.strip():
                continue

            # Handle quoted fields with semicolons
            parts = []
            in_quotes = False
            current = ''
            for char in line:
                if char == '"':
                    in_quotes = not in_quotes
                elif char == ';' and not in_quotes:
                    parts.append(current.strip().strip('"'))
                    current = ''
                else:
                    current += char
            parts.append(current.strip().strip('"'))

            if len(parts) > max(url_idx, title_idx, desc_idx, h1_idx):
                titles.append({
                    'url': parts[url_idx],
                    'title': parts[title_idx],
                    'description': parts[desc_idx],
                    'h1_title': parts[h1_idx]
                })

        # Bulk upsert to database
        result = bulk_upsert_titles(titles)

        return {
            "status": "success",
            "rows_in_file": len(lines) - 1,
            "imported": result["success_count"],
            "errors": result["error_count"]
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/unique-titles/publish")
async def publish_unique_titles():
    """
    Generate CSV from database and upload to the website-configuration API.
    """
    try:
        result = upload_titles_to_api()
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/unique-titles/preview")
async def preview_unique_titles(offset: int = 0, limit: int = 100):
    """Preview titles from database."""
    try:
        all_titles = get_all_titles()
        total = len(all_titles)
        titles = all_titles[offset:offset + limit]

        return {
            "total": total,
            "offset": offset,
            "limit": limit,
            "titles": titles
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/unique-titles/download-csv")
async def download_unique_titles_csv():
    """Download the generated CSV file."""
    try:
        csv_content = generate_csv_for_upload()

        return StreamingResponse(
            iter([csv_content]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=unique_titles_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/unique-titles/{url:path}")
async def delete_unique_title(url: str):
    """Delete a title by URL."""
    try:
        success = delete_title(url)
        if success:
            return {"status": "success", "message": f"Deleted title for {url}"}
        else:
            raise HTTPException(status_code=404, detail="Title not found")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/unique-titles/queue")
async def queue_unique_titles(payload: dict = Body(...)):
    """Add URLs to the database so they become eligible for AI title generation."""
    raw = payload.get("urls")
    if isinstance(raw, str):
        urls = [line for line in raw.splitlines()]
    elif isinstance(raw, list):
        urls = raw
    else:
        raise HTTPException(status_code=400, detail="Provide 'urls' as a string or list")

    try:
        return queue_urls_for_generation(urls)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/unique-titles/search")
async def search_unique_titles(q: str, limit: int = 100):
    """Search titles by URL or title content."""
    try:
        results = search_titles(q, limit)
        return {
            "query": q,
            "count": len(results),
            "results": results
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# AI Title Generation API
# ============================================================================

from backend.ai_titles_service import (
    init_ai_titles_columns,
    get_processing_status,
    start_processing,
    stop_processing,
    get_ai_titles_stats,
    get_unprocessed_count,
    get_recent_results,
    analyze_and_flag_failures,
)

# Initialize AI titles columns on startup
try:
    init_ai_titles_columns()
except Exception as e:
    print(f"[STARTUP] Could not initialize AI titles columns: {e}")


@app.get("/api/ai-titles/status")
async def get_ai_titles_status():
    """Get AI title processing status."""
    try:
        processing = get_processing_status()
        stats = get_ai_titles_stats()
        return {
            "processing": processing,
            "stats": stats
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ai-titles/batch-start")
def start_ai_titles_batch_endpoint():
    """Start AI titles batch processing via concurrent workers."""
    return start_titles_batch()

@app.get("/api/ai-titles/batch-status")
def get_ai_titles_batch_status():
    """Get AI titles batch processing status."""
    return get_batch_status("titles")

@app.post("/api/ai-titles/start")
async def start_ai_titles_processing(batch_size: int = 100, num_workers: int = 50, use_api: bool = True):
    """Start AI title generation processing.

    Args:
        batch_size: Number of URLs to process. 0 = all pending.
        num_workers: Number of parallel workers.
        use_api: If True, use productsearch API for faceted URLs. If False, use scraping.
    """
    try:
        result = start_processing(batch_size, num_workers, use_api)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ai-titles/stop")
async def stop_ai_titles_processing():
    """Stop AI title generation processing."""
    try:
        result = stop_processing()
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/ai-titles/recent")
async def get_ai_titles_recent(limit: int = 20):
    """Get recently processed AI titles."""
    try:
        results = get_recent_results(limit)
        return {"results": results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ai-titles/flag-predicted-failures")
async def ai_titles_flag_predicted_failures(request: dict):
    """Analyze failed URLs for patterns and flag pending URLs likely to fail."""
    dry_run = request.get("dry_run", True)
    min_fail_rate = request.get("min_fail_rate", 80)
    min_failures = request.get("min_failures", 5)

    try:
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(
            None, analyze_and_flag_failures, dry_run, min_fail_rate, min_failures
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# CANONICAL URL GENERATOR
# =============================================================================

from backend.canonical_service import (
    fetch_urls_from_redshift,
    fetch_urls_for_rules,
    generate_canonicals,
    parse_rules_from_json,
    TransformationRules
)
from pydantic import BaseModel
from typing import List, Optional


class CanonicalRulesRequest(BaseModel):
    """Request model for canonical generation"""
    cat_cat: List[dict] = []
    facet_facet: List[dict] = []
    cat_facet: List[dict] = []
    cat_facet_remove: List[dict] = []
    bucket_bucket: List[dict] = []
    remove_bucket: List[dict] = []
    start_date: str = "20240101"
    end_date: str = "20261231"
    fetch_from_redshift: bool = True
    manual_urls: List[str] = []


@app.post("/api/canonical/generate")
async def generate_canonical_urls(request: CanonicalRulesRequest):
    """
    Generate canonical URLs based on transformation rules.

    Either fetches URLs from Redshift based on the rules, or uses manually provided URLs.
    Then applies all transformation rules to generate canonical versions.
    """
    try:
        # Parse rules
        rules = parse_rules_from_json(request.dict())

        # Get URLs to transform
        if request.fetch_from_redshift:
            urls = fetch_urls_for_rules(rules, request.start_date, request.end_date)
        else:
            urls = request.manual_urls

        if not urls:
            return {
                "status": "success",
                "message": "No URLs found matching the rules",
                "total": 0,
                "results": []
            }

        # Generate canonicals
        results = generate_canonicals(urls, rules)

        # Filter to only include URLs that actually changed
        changed_results = [r for r in results if r["original"] != r["canonical"]]

        return {
            "status": "success",
            "total": len(results),
            "changed": len(changed_results),
            "results": changed_results
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/canonical/preview")
async def preview_canonical_urls(request: CanonicalRulesRequest):
    """
    Preview URLs that would be fetched from Redshift based on the rules.
    Does not apply transformations, just shows what URLs would be affected.
    """
    try:
        rules = parse_rules_from_json(request.dict())
        urls = fetch_urls_for_rules(rules, request.start_date, request.end_date)

        return {
            "status": "success",
            "total": len(urls),
            "urls": urls[:100]  # Limit preview to first 100
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/canonical/fetch-urls")
async def fetch_canonical_urls(
    contains: Optional[str] = None,
    start_date: str = "20240101",
    end_date: str = "20261231",
    limit: int = 1000
):
    """
    Fetch URLs from Redshift that match the given criteria.
    Useful for exploring what URLs exist before defining transformation rules.
    """
    try:
        urls = fetch_urls_from_redshift(
            contains=contains,
            start_date=start_date,
            end_date=end_date,
            limit=limit
        )

        return {
            "status": "success",
            "total": len(urls),
            "urls": urls
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/canonical/transform")
async def transform_single_url(url: str, rules: CanonicalRulesRequest):
    """
    Transform a single URL using the provided rules.
    Useful for testing rules before running on full dataset.
    """
    try:
        from backend.canonical_service import transform_url

        parsed_rules = parse_rules_from_json(rules.dict())
        canonical = transform_url(url, parsed_rules)

        return {
            "original": url,
            "canonical": canonical,
            "changed": url != canonical
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# 301 REDIRECT GENERATOR - Sort facets alphabetically
# =============================================================================

from backend.redirect_301_service import (
    check_facets_sorted,
    fetch_urls_with_facets,
    fetch_urls_with_facets_batched,
    generate_301_redirects,
    parse_facet_rules,
    parse_category_rules,
    extract_patterns_from_rules
)


class Redirect301Request(BaseModel):
    """Request model for 301 redirect generation"""
    contains: Optional[str] = None
    start_date: str = "20240101"
    end_date: str = "20261231"
    fetch_from_redshift: bool = True
    manual_urls: List[str] = []
    limit: int = 10000
    facet_rules: List[dict] = []  # [{"old_facet": "...", "new_facet": "...", "category": "..."}]
    category_rules: List[dict] = []  # [{"old_cat": "...", "new_cat": "...", "new_maincat": "..."}]
    sort_only: bool = False  # If True, only sort facets without applying rules
    auto_filter: bool = True  # If True, auto-extract patterns from rules for filtering


@app.post("/api/301-generator/generate")
async def generate_301_urls(request: Redirect301Request):
    """
    Generate 301 redirects for URLs with unsorted facets or transformations.

    Supports:
    - Sorting facets alphabetically (sort_only=True)
    - Category transformations: {"old_cat": "fietsen_123_456", "new_cat": "fietsen_123"}
    - Facet transformations with full ID: {"old_facet": "merk~123", "new_facet": "materiaal~456"}
    - Facet transformations without ID: {"old_facet": "merk", "new_facet": "materiaal", "category": "/fietsen/"}
    """
    try:
        # Parse rules first (needed for both pattern extraction and transformation)
        facet_rules = parse_facet_rules(request.facet_rules) if request.facet_rules else None
        category_rules = parse_category_rules(request.category_rules) if request.category_rules else None

        if request.fetch_from_redshift:
            # Extract patterns from rules to automatically filter Redshift query (if enabled)
            rule_patterns = []
            if request.auto_filter:
                rule_patterns = extract_patterns_from_rules(facet_rules, category_rules)
                if rule_patterns:
                    print(f"[301-GENERATOR] Auto-filter enabled: {len(rule_patterns)} unique patterns extracted")

            if rule_patterns:
                # Use batched queries - one query per pattern (faster for many patterns)
                url_data = fetch_urls_with_facets_batched(
                    patterns=rule_patterns,
                    start_date=request.start_date,
                    end_date=request.end_date,
                    limit_per_pattern=request.limit // len(rule_patterns) if len(rule_patterns) > 1 else request.limit
                )
            else:
                # No patterns - use single query with optional contains filter
                url_data = fetch_urls_with_facets(
                    contains=request.contains,
                    start_date=request.start_date,
                    end_date=request.end_date,
                    limit=request.limit
                )
            urls = [u["url"] for u in url_data]
        else:
            urls = request.manual_urls

        if not urls:
            return {
                "status": "success",
                "message": "No URLs found",
                "total": 0,
                "needs_redirect": 0,
                "results": []
            }

        results = generate_301_redirects(
            urls,
            facet_rules=facet_rules,
            category_rules=category_rules,
            sort_only=request.sort_only
        )

        # Get patterns that were used for query (for UI feedback)
        search_patterns = []
        if request.fetch_from_redshift and request.auto_filter:
            search_patterns = extract_patterns_from_rules(facet_rules, category_rules)

        return {
            "status": "success",
            "total": len(urls),
            "needs_redirect": len(results),
            "facet_rules_applied": len(request.facet_rules),
            "category_rules_applied": len(request.category_rules),
            "auto_filter_enabled": request.auto_filter,
            "search_patterns": search_patterns,
            "results": results
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/301-generator/fetch-urls")
async def fetch_301_urls(
    contains: Optional[str] = None,
    start_date: str = "20240101",
    end_date: str = "20261231",
    limit: int = 10000
):
    """
    Fetch URLs with facets from Redshift for 301 redirect checking.
    """
    try:
        url_data = fetch_urls_with_facets(
            contains=contains,
            start_date=start_date,
            end_date=end_date,
            limit=limit
        )

        urls = [u["url"] for u in url_data]

        return {
            "status": "success",
            "total": len(urls),
            "urls": urls
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/301-generator/check")
async def check_single_url_facets(url: str):
    """
    Check if a single URL has properly sorted facets.
    """
    try:
        is_sorted, corrected_url = check_facets_sorted(url)

        return {
            "original": url,
            "is_sorted": is_sorted,
            "corrected": corrected_url if not is_sorted else None,
            "needs_redirect": not is_sorted
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# R-FINDER ENDPOINTS - Find /r/ URLs from Redshift
# =============================================================================

from backend.rfinder_service import fetch_r_urls, get_r_url_stats


class RFinderRequest(BaseModel):
    """Request model for R-finder search"""
    filters: Optional[List[str]] = []
    min_visits: Optional[int] = 0
    start_date: Optional[str] = "20210101"
    end_date: Optional[str] = "20261231"
    limit: Optional[int] = 4000


@app.post("/api/rfinder/search")
async def search_r_urls(request: RFinderRequest):
    """
    Search for /r/ URLs from Redshift.

    Applies the same filters as the original GA4-based r-finder script:
    - Must contain /r/
    - Excludes device=, /sitemap/, sortby=, /filters/, /page_, shop_id=, etc.

    Optional filters can be provided to narrow down results (e.g., category segments).
    """
    try:
        # Clean up filters - remove empty strings
        filters = [f for f in (request.filters or []) if f and f.strip()]

        urls = fetch_r_urls(
            filters=filters if filters else None,
            min_visits=request.min_visits or 0,
            start_date=request.start_date or "20210101",
            end_date=request.end_date or "20261231",
            limit=request.limit or 4000
        )

        return {
            "status": "success",
            "total": len(urls),
            "urls": urls
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/rfinder/stats")
async def get_rfinder_stats(
    start_date: str = "20210101",
    end_date: str = "20261231"
):
    """
    Get statistics about /r/ URLs in Redshift.

    Returns total unique URLs and total sessions for the given date range.
    """
    try:
        stats = get_r_url_stats(start_date, end_date)
        return {
            "status": "success",
            **stats
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# Redirect Checker API
# =============================================================================

REDIRECT_CHECKER_USER_AGENT = "Beslist script voor SEO"
REDIRECT_CHECKER_BASE_URL = "https://www.beslist.nl"

def normalize_url(url: str) -> str:
    """Normalize URL by adding base URL for relative paths."""
    url = url.strip()
    if not url:
        return url
    # Already absolute URL
    if url.startswith(('http://', 'https://')):
        return url
    # Relative URL starting with /
    if url.startswith('/'):
        return REDIRECT_CHECKER_BASE_URL + url
    # Relative URL without leading slash
    return REDIRECT_CHECKER_BASE_URL + '/' + url

def extract_canonical_from_html(html: str, base_url: str) -> str:
    """Extract canonical URL from HTML using regex."""
    try:
        patterns = [
            r'<link[^>]+rel=["\']canonical["\'][^>]+href=["\']([^"\']+)["\']',
            r'<link[^>]+href=["\']([^"\']+)["\'][^>]+rel=["\']canonical["\']',
        ]
        for pattern in patterns:
            match = re.search(pattern, html, re.IGNORECASE)
            if match:
                canonical = match.group(1)
                if not canonical.startswith(('http://', 'https://')):
                    canonical = urljoin(base_url, canonical)
                return canonical
    except Exception:
        pass
    return None


class RedirectRateLimiter:
    """Token bucket rate limiter for redirect checker."""
    def __init__(self, rate):
        self.rate = rate
        self.tokens = rate
        self.last_update = time.monotonic()
        self.lock = asyncio.Lock()

    async def __aenter__(self):
        async with self.lock:
            now = time.monotonic()
            elapsed = now - self.last_update
            self.tokens = min(self.rate, self.tokens + elapsed * self.rate)
            self.last_update = now
            if self.tokens < 1:
                wait_time = (1 - self.tokens) / self.rate
                await asyncio.sleep(wait_time)
                self.tokens = 0
            else:
                self.tokens -= 1

    async def __aexit__(self, *args):
        pass


async def check_single_url(client: httpx.AsyncClient, url: str, semaphore, rate_limiter, timeout: int):
    """Check a single URL for status code, redirect, and canonical."""
    async with semaphore:
        async with rate_limiter:
            result = {
                'input_url': url,
                'status_code': None,
                'final_url': None,
                'redirect_url': None,
                'canonical_url': None,
                'error': None
            }
            try:
                # First request without redirects to get initial status
                response = await client.get(url, follow_redirects=False, timeout=timeout)
                initial_status = response.status_code
                result['status_code'] = initial_status

                if initial_status in (301, 302, 303, 307, 308):
                    redirect_location = response.headers.get('Location')
                    if redirect_location:
                        if not redirect_location.startswith(('http://', 'https://')):
                            redirect_location = urljoin(url, redirect_location)
                        result['redirect_url'] = redirect_location

                    # Second request following redirects to get final URL and canonical
                    response = await client.get(url, follow_redirects=True, timeout=timeout)
                    result['final_url'] = str(response.url)
                    if response.status_code == 200:
                        try:
                            html = response.text
                            result['canonical_url'] = extract_canonical_from_html(html, str(response.url))
                        except Exception:
                            pass
                else:
                    # No redirect - final URL is same as input
                    result['final_url'] = str(response.url)
                    if initial_status == 200:
                        try:
                            html = response.text
                            result['canonical_url'] = extract_canonical_from_html(html, str(response.url))
                        except Exception:
                            pass

            except httpx.TimeoutException:
                result['status_code'] = 'TIMEOUT'
                result['error'] = 'Request timed out'
            except httpx.RequestError as e:
                result['status_code'] = 'ERROR'
                result['error'] = str(e)[:100]
            except Exception as e:
                result['status_code'] = 'ERROR'
                result['error'] = str(e)[:100]
            return result


@app.post("/api/redirect-checker/check")
async def redirect_checker_check(request: dict):
    """
    Check URLs for status codes, redirects, and canonical URLs.
    Returns results as a stream of JSON lines.
    """
    urls = request.get('urls', [])
    workers = min(request.get('workers', 20), 50)
    rate = min(request.get('rate', 2), 20)
    timeout = min(request.get('timeout', 15), 60)

    if not urls:
        raise HTTPException(status_code=400, detail="No URLs provided")

    # Normalize URLs (handle relative URLs)
    urls = [normalize_url(url) for url in urls if url.strip()]

    async def generate():
        semaphore = asyncio.Semaphore(workers)
        rate_limiter = RedirectRateLimiter(rate)
        headers = {"User-Agent": REDIRECT_CHECKER_USER_AGENT}
        limits = httpx.Limits(max_connections=workers, max_keepalive_connections=workers)

        async with httpx.AsyncClient(headers=headers, limits=limits) as client:
            tasks = [check_single_url(client, url, semaphore, rate_limiter, timeout) for url in urls]
            for coro in asyncio.as_completed(tasks):
                result = await coro
                yield json.dumps({"type": "result", "data": result}) + "\n"

        yield json.dumps({"type": "done"}) + "\n"

    return StreamingResponse(generate(), media_type="application/x-ndjson")


@app.post("/api/redirect-checker/download")
async def redirect_checker_download(request: dict):
    """Download redirect checker results as Excel file."""
    results = request.get('results', [])
    if not results:
        raise HTTPException(status_code=400, detail="No results provided")

    try:
        import pandas as pd
        df = pd.DataFrame(results)
        cols = ['input_url', 'status_code', 'redirect_url', 'canonical_url', 'final_url', 'error']
        df = df[[c for c in cols if c in df.columns]]

        output = BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)

        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=redirect_check_results.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# URL Checker API
# ============================================================================

def extract_meta_title(html: str) -> str:
    """Extract <title> content from HTML."""
    match = re.search(r'<title[^>]*>(.*?)</title>', html, re.IGNORECASE | re.DOTALL)
    return match.group(1).strip() if match else None

def extract_meta_description(html: str) -> str:
    """Extract meta description from HTML."""
    patterns = [
        r'<meta[^>]+name=["\']description["\'][^>]+content=["\']([^"\']*)["\']',
        r'<meta[^>]+content=["\']([^"\']*)["\'][^>]+name=["\']description["\']',
    ]
    for pattern in patterns:
        match = re.search(pattern, html, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return None

def extract_h1(html: str) -> str:
    """Extract first H1 content from HTML (strip tags inside H1)."""
    match = re.search(r'<h1[^>]*>(.*?)</h1>', html, re.IGNORECASE | re.DOTALL)
    if match:
        h1_content = match.group(1).strip()
        # Strip inner HTML tags
        h1_content = re.sub(r'<[^>]+>', '', h1_content).strip()
        return h1_content
    return None

def extract_product_count(html: str) -> int:
    """Extract product count from Beslist.nl page by finding productCount of a selected facet."""
    match = re.search(r'"productCount":(\d+),"selected":true', html)
    if match:
        return int(match.group(1))
    # Try reversed order
    match = re.search(r'"selected":true[^}]*?"productCount":(\d+)', html)
    if match:
        return int(match.group(1))
    return None


async def check_single_url_metadata(client: httpx.AsyncClient, url: str, semaphore, rate_limiter, timeout: int):
    """Check a single URL and extract metadata: status, redirect, title, description, H1, product count."""
    async with semaphore:
        async with rate_limiter:
            result = {
                'input_url': url,
                'status_code': None,
                'redirect_url': None,
                'final_url': None,
                'meta_title': None,
                'meta_description': None,
                'h1': None,
                'product_count': None,
                'canonical_url': None,
                'error': None
            }
            try:
                # First request without redirects to get initial status
                response = await client.get(url, follow_redirects=False, timeout=timeout)
                initial_status = response.status_code
                result['status_code'] = initial_status

                html = None

                if initial_status in (301, 302, 303, 307, 308):
                    redirect_location = response.headers.get('Location')
                    if redirect_location:
                        if not redirect_location.startswith(('http://', 'https://')):
                            redirect_location = urljoin(url, redirect_location)
                        result['redirect_url'] = redirect_location

                    # Follow redirects to get final URL and parse HTML
                    response = await client.get(url, follow_redirects=True, timeout=timeout)
                    result['final_url'] = str(response.url)
                    if response.status_code == 200:
                        html = response.text
                else:
                    result['final_url'] = str(response.url)
                    if initial_status == 200:
                        html = response.text

                # Extract metadata from HTML
                if html:
                    result['meta_title'] = extract_meta_title(html)
                    result['meta_description'] = extract_meta_description(html)
                    result['h1'] = extract_h1(html)
                    result['product_count'] = extract_product_count(html)
                    result['canonical_url'] = extract_canonical_from_html(html, result['final_url'])

            except httpx.TimeoutException:
                result['status_code'] = 'TIMEOUT'
                result['error'] = 'Request timed out'
            except httpx.RequestError as e:
                result['status_code'] = 'ERROR'
                result['error'] = str(e)[:100]
            except Exception as e:
                result['status_code'] = 'ERROR'
                result['error'] = str(e)[:100]
            return result


@app.post("/api/url-checker/check")
async def url_checker_check(request: dict):
    """
    Check URLs for status codes, metadata (title, description, H1, product count).
    Returns results as a stream of JSON lines.
    """
    urls = request.get('urls', [])
    workers = min(request.get('workers', 10), 10)
    rate = min(request.get('rate', 2), 2)
    timeout = min(request.get('timeout', 15), 60)

    if not urls:
        raise HTTPException(status_code=400, detail="No URLs provided")

    urls = [normalize_url(url) for url in urls if url.strip()]

    async def generate():
        semaphore = asyncio.Semaphore(workers)
        rate_limiter = RedirectRateLimiter(rate)
        headers = {"User-Agent": REDIRECT_CHECKER_USER_AGENT}
        limits = httpx.Limits(max_connections=workers, max_keepalive_connections=workers)

        async with httpx.AsyncClient(headers=headers, limits=limits) as client:
            tasks = [check_single_url_metadata(client, url, semaphore, rate_limiter, timeout) for url in urls]
            for coro in asyncio.as_completed(tasks):
                result = await coro
                yield json.dumps({"type": "result", "data": result}) + "\n"

        yield json.dumps({"type": "done"}) + "\n"

    return StreamingResponse(generate(), media_type="application/x-ndjson")


@app.post("/api/url-checker/upload")
async def url_checker_upload(file: UploadFile = File(...)):
    """Upload a file (.xlsx, .csv, .txt) containing URLs. Returns the list of URLs found."""
    filename = file.filename.lower()
    contents = await file.read()

    try:
        urls = []
        if filename.endswith(('.xlsx', '.xls')):
            import pandas as pd
            df = pd.read_excel(BytesIO(contents))
            if not df.empty:
                urls = df.iloc[:, 0].dropna().astype(str).tolist()
        elif filename.endswith('.csv'):
            import pandas as pd
            df = pd.read_csv(BytesIO(contents))
            if not df.empty:
                urls = df.iloc[:, 0].dropna().astype(str).tolist()
        elif filename.endswith('.txt'):
            text = contents.decode('utf-8', errors='ignore')
            urls = [line.strip() for line in text.splitlines()]
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type. Use .xlsx, .csv, or .txt")

        # Filter to valid URLs
        urls = [u.strip() for u in urls if u.strip() and (u.strip().startswith('http') or u.strip().startswith('/'))]

        if not urls:
            raise HTTPException(status_code=400, detail="No valid URLs found in file")

        return {"urls": urls, "count": len(urls)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/url-checker/download")
async def url_checker_download(request: dict):
    """Download URL checker results as Excel file."""
    results = request.get('results', [])
    if not results:
        raise HTTPException(status_code=400, detail="No results provided")

    try:
        import pandas as pd
        df = pd.DataFrame(results)
        cols = ['input_url', 'status_code', 'redirect_url', 'canonical_url', 'self_canonical', 'meta_title', 'meta_description', 'h1', 'product_count', 'error']
        df = df[[c for c in cols if c in df.columns]]

        output = BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)

        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=url_checker_results.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Keyword Planner API
# ============================================================================

@app.post("/api/keyword-planner/search-volumes")
async def keyword_planner_search_volumes(request: dict):
    """
    Get search volumes for a list of keywords.
    Accepts {"keywords": ["e-bike", "hardloopschoenen", ...]}
    Returns results with original keyword, normalized keyword, and search volume.
    """
    keywords = request.get("keywords", [])
    if not keywords:
        raise HTTPException(status_code=400, detail="No keywords provided")
    if len(keywords) > 50000:
        raise HTTPException(status_code=400, detail="Maximum 50,000 keywords per request")

    try:
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, get_search_volumes, keywords)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/keyword-planner/upload-excel")
async def keyword_planner_upload_excel(file: UploadFile = File(...)):
    """
    Upload an Excel file containing keywords. Reads keywords from the first column.
    Returns search volumes for all keywords found.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")

    try:
        import pandas as pd
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))

        if df.empty:
            raise HTTPException(status_code=400, detail="Excel file is empty")

        # Read keywords from the first column
        keywords = df.iloc[:, 0].dropna().astype(str).tolist()
        keywords = [k.strip() for k in keywords if k.strip()]

        if not keywords:
            raise HTTPException(status_code=400, detail="No keywords found in the first column")
        if len(keywords) > 50000:
            raise HTTPException(status_code=400, detail="Maximum 50,000 keywords per file")

        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, get_search_volumes, keywords)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/keyword-planner/test")
async def keyword_planner_test():
    """Test the Google Ads Keyword Planner API connection."""
    try:
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, test_keyword_planner_connection)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/keyword-planner/download")
async def keyword_planner_download(request: dict):
    """Download keyword planner results as Excel file."""
    results = request.get('results', [])
    if not results:
        raise HTTPException(status_code=400, detail="No results provided")

    try:
        import pandas as pd
        df = pd.DataFrame(results)
        cols = ['original_keyword', 'normalized_keyword', 'search_volume']
        df = df[[c for c in cols if c in df.columns]]
        df = df.sort_values('search_volume', ascending=False)

        output = BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)

        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=keyword_planner_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Category Keyword Volumes API
# ============================================================================

@app.post("/api/keyword-planner/category-volumes")
async def keyword_planner_category_volumes(request: dict):
    """
    Combine a keyword with all preloaded category names.
    Generates singular/plural + keyword/category order combinations,
    looks up search volumes, and aggregates per deepest_cat and maincat.

    Categories are preloaded from backend/categories.xlsx at startup.
    """
    keyword = request.get("keyword", "").strip()
    if not keyword:
        raise HTTPException(status_code=400, detail="No keyword provided")

    try:
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, process_category_keywords, keyword, PRELOADED_CATEGORIES)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/keyword-planner/category-volumes/download")
async def keyword_planner_category_volumes_download(request: dict):
    """
    Download category volume results as Excel file.
    Expects the same format as the input Excel but with added search volume columns.
    """
    import pandas as pd

    deepest_results = request.get('deepest_cat_results', [])
    if not deepest_results:
        raise HTTPException(status_code=400, detail="No results provided")

    try:
        # Build maincat volume lookup
        maincat_volumes = {}
        for r in deepest_results:
            mc = r.get("maincat", "")
            vol = r.get("search_volume", 0)
            if mc not in maincat_volumes:
                maincat_volumes[mc] = 0
            maincat_volumes[mc] += vol

        # Build output DataFrame matching input format + volume columns
        rows = []
        for r in deepest_results:
            rows.append({
                "maincat": r.get("maincat", ""),
                "maincat_id": r.get("maincat_id", ""),
                "deepest_cat": r.get("deepest_cat", ""),
                "cat_id": r.get("cat_id", ""),
                "search_volume_deepest_cat": r.get("search_volume", 0),
                "search_volume_maincat": maincat_volumes.get(r.get("maincat", ""), 0),
            })

        df = pd.DataFrame(rows)
        output = BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)

        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=category_volumes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# IndexNow Endpoints
# ============================================================

@app.post("/api/indexnow/submit")
async def indexnow_submit(request: dict):
    """
    Submit URLs to IndexNow API.
    Accepts {"urls": ["https://...", ...]}
    Deduplicates against previously submitted URLs in Redshift.
    """
    from backend.indexnow_service import submit_urls

    urls = request.get("urls", [])
    if not urls:
        raise HTTPException(status_code=400, detail="No URLs provided")

    urls = [u.strip() for u in urls if u and u.strip()]
    if not urls:
        raise HTTPException(status_code=400, detail="No valid URLs provided")

    try:
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, submit_urls, urls)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/indexnow/upload-excel")
async def indexnow_upload_excel(file: UploadFile = File(...)):
    """
    Upload an Excel file with a URL column and submit to IndexNow.
    """
    from backend.indexnow_service import submit_urls

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")

    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))

        if df.empty:
            raise HTTPException(status_code=400, detail="Excel file is empty")

        # Find URL column (case-insensitive)
        url_col = None
        for col in df.columns:
            if col.strip().upper() == "URL":
                url_col = col
                break

        if url_col is None:
            # Fall back to first column
            url_col = df.columns[0]

        urls = df[url_col].dropna().astype(str).tolist()
        urls = [u.strip() for u in urls if u.strip() and u.strip().startswith("http")]

        if not urls:
            raise HTTPException(status_code=400, detail="No valid URLs found in the file")

        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, submit_urls, urls)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/indexnow/history")
async def indexnow_history(limit: int = 100):
    """Get submission history grouped by date and response code."""
    from backend.indexnow_service import get_submission_history

    try:
        loop = asyncio.get_event_loop()
        history = await loop.run_in_executor(None, get_submission_history, limit)
        return {"status": "success", "history": history}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/indexnow/export/{date}")
async def indexnow_export_by_date(date: str):
    """Export submitted URLs for a specific date as XLSX."""
    from backend.database import get_db_connection, return_db_connection
    from openpyxl import Workbook

    # Validate date format
    try:
        datetime.strptime(date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT url, submitted_date, response_code
            FROM pa.index_now_joep
            WHERE submitted_date = %s
            ORDER BY url
        """, (date,))
        rows = cur.fetchall()
        cur.close()
        return_db_connection(conn)

        wb = Workbook()
        ws = wb.active
        ws.title = "IndexNow"
        ws.append(["url", "submitted_date", "response_code"])
        for row in rows:
            ws.append([row["url"], str(row["submitted_date"]), row["response_code"]])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=indexnow_{date}.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/indexnow/today-count")
async def indexnow_today_count():
    """Get the number of URLs submitted today and the daily limit."""
    from backend.indexnow_service import get_today_count, DAILY_LIMIT

    try:
        loop = asyncio.get_event_loop()
        count = await loop.run_in_executor(None, get_today_count)
        return {"status": "success", "today_count": count, "daily_limit": DAILY_LIMIT}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# SEO Index Checker Endpoints
# ============================================================

@app.post("/api/index-checker/check")
async def index_checker_check(request: dict):
    """
    Check index status for a list of URLs via Google Search Console URL Inspection API.
    Accepts {"urls": ["https://...", ...]}
    """
    from backend.index_checker_service import check_urls

    urls = request.get("urls", [])
    if not urls:
        raise HTTPException(status_code=400, detail="No URLs provided")

    urls = [u.strip() for u in urls if u and u.strip()]
    if not urls:
        raise HTTPException(status_code=400, detail="No valid URLs provided")
    if len(urls) > 8000:
        raise HTTPException(status_code=400, detail="Maximum 8,000 URLs per request (daily quota limit)")

    try:
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, check_urls, urls)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/index-checker/upload-excel")
async def index_checker_upload_excel(file: UploadFile = File(...)):
    """
    Upload an Excel file with URLs and check their index status.
    """
    from backend.index_checker_service import check_urls

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")

    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))

        if df.empty:
            raise HTTPException(status_code=400, detail="Excel file is empty")

        # Find URL column (case-insensitive)
        url_col = None
        for col in df.columns:
            if col.strip().upper() == "URL":
                url_col = col
                break
        if url_col is None:
            url_col = df.columns[0]

        urls = df[url_col].dropna().astype(str).tolist()
        urls = [u.strip() for u in urls if u.strip() and u.strip().startswith("http")]

        if not urls:
            raise HTTPException(status_code=400, detail="No valid URLs found in the file")
        if len(urls) > 8000:
            raise HTTPException(status_code=400, detail="Maximum 8,000 URLs per request (daily quota limit)")

        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, check_urls, urls)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/index-checker/quota")
async def index_checker_quota():
    """Get info about available service accounts and estimated daily quota."""
    from backend.index_checker_service import get_quota_info

    try:
        loop = asyncio.get_event_loop()
        info = await loop.run_in_executor(None, get_quota_info)
        return info
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
