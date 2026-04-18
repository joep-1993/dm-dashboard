"""
DMA+ Service — Web wrapper for campaign_processor.py operations.

Provides background execution, progress tracking, change history, and
NL/BE country switching for the 5 core DMA operations:
  1. process_inclusion_sheet_v2
  2. process_exclusion_sheet_v2
  3. validate_cl1_targeting_for_campaigns
  4. validate_ads_for_campaigns
  5. validate_listing_trees_for_campaigns
"""

import io
import os
import sys
import time
import uuid
import logging
import threading
import traceback
from collections import deque
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional

import openpyxl
from google.ads.googleads.client import GoogleAdsClient

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Google Ads client (shared with dma_bidding_service pattern)
# ---------------------------------------------------------------------------
MCC_CUSTOMER_ID = "3011145605"

COUNTRY_CONFIG = {
    "NL": {"customer_id": "3800751597", "merchant_center_id": 140784594, "exclude_dataedis": True},
    "BE": {"customer_id": "9920951707", "merchant_center_id": 140784810, "exclude_dataedis": False},
}


def _get_client() -> GoogleAdsClient:
    """Initialize Google Ads client from environment, checking both naming conventions."""
    config = {
        "developer_token": os.environ.get("GOOGLE_ADS_DEVELOPER_TOKEN")
                           or os.environ.get("GOOGLE_DEVELOPER_TOKEN", ""),
        "refresh_token": os.environ.get("GOOGLE_ADS_REFRESH_TOKEN")
                         or os.environ.get("GOOGLE_REFRESH_TOKEN", ""),
        "client_id": os.environ.get("GOOGLE_CLIENT_ID", ""),
        "client_secret": os.environ.get("GOOGLE_CLIENT_SECRET", ""),
        "login_customer_id": os.environ.get("GOOGLE_ADS_LOGIN_CUSTOMER_ID")
                             or os.environ.get("GOOGLE_LOGIN_CUSTOMER_ID", MCC_CUSTOMER_ID),
        "use_proto_plus": True,
    }
    return GoogleAdsClient.load_from_dict(config)


# ---------------------------------------------------------------------------
# Task store (in-memory, background tasks)
# ---------------------------------------------------------------------------
_tasks: dict = {}

# Change history is persisted to disk so server restarts (frequent during
# dev) don't wipe it. JSON is fine here: entries are small and capped at 200.
_HISTORY_FILE: Path = Path(__file__).parent / "data" / "dma_plus_history.json"
_history_lock = threading.Lock()


def _load_history_from_disk() -> deque:
    if _HISTORY_FILE.exists():
        try:
            import json as _json
            data = _json.loads(_HISTORY_FILE.read_text(encoding="utf-8"))
            if isinstance(data, list):
                return deque(data, maxlen=200)
        except Exception as e:
            logger.warning(f"Failed to load DMA+ history from {_HISTORY_FILE}: {e}")
    return deque(maxlen=200)


def _save_history_to_disk():
    try:
        import json as _json
        _HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
        _HISTORY_FILE.write_text(
            _json.dumps(list(_history), default=str, ensure_ascii=False),
            encoding="utf-8",
        )
    except Exception as e:
        logger.warning(f"Failed to save DMA+ history to {_HISTORY_FILE}: {e}")


_history: deque = _load_history_from_disk()


def _history_append(entry: dict):
    """Add an entry to the head of the history and persist to disk."""
    with _history_lock:
        _history.appendleft(entry)
        _save_history_to_disk()


def _get_task(task_id: str) -> Optional[dict]:
    return _tasks.get(task_id)


def _set_task(task_id: str, data: dict):
    existing = _tasks.get(task_id)
    if existing and existing.get("cancel") and "cancel" not in data:
        data["cancel"] = True
    _tasks[task_id] = data


class TaskCancelled(Exception):
    """Raised inside a background task when the user hit Cancel."""


def _check_cancelled(task_id: str):
    task = _tasks.get(task_id)
    if task and task.get("cancel"):
        raise TaskCancelled()


# ---------------------------------------------------------------------------
# Patch campaign_processor globals for web use
# ---------------------------------------------------------------------------
def _patch_campaign_processor(country: str):
    """
    Set campaign_processor module globals to match the selected country.
    Must be called BEFORE invoking any processor function.
    """
    from backend import campaign_processor as cp

    cfg = COUNTRY_CONFIG[country]
    cp.COUNTRY = country
    cp.CUSTOMER_ID = cfg["customer_id"]
    cp.MERCHANT_CENTER_ID = cfg["merchant_center_id"]
    cp.EXCLUDE_DATAEDIS = cfg["exclude_dataedis"]


# ---------------------------------------------------------------------------
# Maincat cross-matching (name ↔ id)
# ---------------------------------------------------------------------------
_maincat_name_to_id: Dict[str, str] = {}
_maincat_id_to_name: Dict[str, str] = {}


def _ensure_maincat_mapping():
    """Load maincat_mapping.csv once for name↔id resolution."""
    if _maincat_name_to_id:
        return
    csv_path = Path(__file__).parent / "maincat_mapping.csv"
    if not csv_path.exists():
        return
    import csv
    with open(csv_path, encoding="utf-8") as f:
        for row in csv.DictReader(f, delimiter=";"):
            name = row["maincat"].strip()
            mid = str(row["maincat_id"]).strip()
            _maincat_name_to_id[name.lower()] = mid
            _maincat_id_to_name[mid] = name


def resolve_maincat(maincat: str, maincat_id: str) -> tuple:
    """
    Cross-match maincat name and id. Returns (name, id).
    If only one is provided, resolves the other from maincat_mapping.csv.
    """
    _ensure_maincat_mapping()
    maincat = (maincat or "").strip()
    maincat_id = (maincat_id or "").strip()

    if maincat and not maincat_id:
        maincat_id = _maincat_name_to_id.get(maincat.lower(), "")
    elif maincat_id and not maincat:
        maincat = _maincat_id_to_name.get(maincat_id, "")

    return maincat, maincat_id


# ---------------------------------------------------------------------------
# Build workbook from shop name input (for quick include/exclude)
# ---------------------------------------------------------------------------
def _build_inclusion_workbook(shop_name: str, maincat: str, maincat_id: str,
                              cl1: str, budget: float) -> openpyxl.Workbook:
    """Create a minimal workbook matching the inclusion sheet v2 layout.
    Columns: A=shop_name, B=Shop ID, C=maincat, D=maincat_id, E=cl1, F=budget, G=result, H=error
    cl1 can be comma-separated (e.g. 'a,b,c') to create one row per level."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "toevoegen"
    ws.append(["shop_name", "Shop ID", "maincat", "maincat_id",
               "custom label 1", "budget", "result", "error message"])
    cl1_values = [v.strip() for v in cl1.split(",") if v.strip()]
    for cl in cl1_values:
        ws.append([shop_name, "", maincat, maincat_id, cl, budget, None, None])
    return wb


def _build_exclusion_workbook(shop_name: str, maincat: str, maincat_id: str,
                              cl1: str) -> openpyxl.Workbook:
    """Create a minimal workbook matching the exclusion sheet layout.
    cl1 can be comma-separated (e.g. 'a,b,c') to create one row per level."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "uitsluiten"
    ws.append(["shop_name", "Shop ID", "maincat", "maincat_id", "custom label 1", "result", "error message"])
    cl1_values = [v.strip() for v in cl1.split(",") if v.strip()]
    for cl in cl1_values:
        ws.append([shop_name, "", maincat, maincat_id, cl, None, None])
    # Populate cat_ids sheet from cat_urls.csv + maincat_mapping.csv
    _populate_cat_ids_sheet(wb)
    return wb


def _build_reverse_exclusion_workbook(shop_name: str, maincat: str, maincat_id: str,
                                      cl1: str) -> openpyxl.Workbook:
    """Create a workbook matching process_reverse_exclusion_sheet's 'verwijderen'
    layout — same columns as 'uitsluiten' but a different sheet name."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "verwijderen"
    ws.append(["shop_name", "Shop ID", "maincat", "maincat_id", "custom label 1", "result", "error message"])
    cl1_values = [v.strip() for v in cl1.split(",") if v.strip()]
    for cl in cl1_values:
        ws.append([shop_name, "", maincat, maincat_id, cl, None, None])
    # cat_ids sheet is required (processor calls load_cat_ids_mapping)
    _populate_cat_ids_sheet(wb)
    return wb


_cat_ids_cache: list = []       # [(mc_name, mc_id, deepest_cat, cat_id), ...]
_cat_ids_cache_time: float = 0
_CAT_IDS_TTL = 3600             # 1 hour


def _populate_cat_ids_sheet(wb: openpyxl.Workbook):
    """Add a cat_ids sheet with full maincat_id → deepest_cat mapping.
    Uses cached Taxonomy API v2 data (1h TTL), falls back to cat_urls.csv."""
    global _cat_ids_cache, _cat_ids_cache_time

    ws_cat = wb.create_sheet("cat_ids")
    ws_cat.append(["maincat", "maincat_id", "deepest_cat", "cat_id"])

    _ensure_maincat_mapping()

    # Use cache if fresh
    if _cat_ids_cache and (time.time() - _cat_ids_cache_time) < _CAT_IDS_TTL:
        for row in _cat_ids_cache:
            ws_cat.append(list(row))
        logger.info(f"cat_ids sheet: {len(_cat_ids_cache)} rows from cache")
        return

    # Try live Taxonomy API
    api_rows = _fetch_all_cat_ids_from_taxonomy_api()
    if api_rows:
        _cat_ids_cache = api_rows
        _cat_ids_cache_time = time.time()
        for row in api_rows:
            ws_cat.append(list(row))
        logger.info(f"cat_ids sheet: {len(api_rows)} rows from Taxonomy API v2")
        return

    # Fallback to CSV
    logger.info("Taxonomy API unavailable, falling back to cat_urls.csv")
    _populate_cat_ids_from_csv(ws_cat)


def _fetch_all_cat_ids_from_taxonomy_api() -> list:
    """Fetch all categories from Taxonomy API v2. Returns [(mc_name, mc_id, cat_name, cat_id), ...]."""
    import requests

    TAX_BASE = "http://producttaxonomyunifiedapi-prod.azure.api.beslist.nl"
    TAX_HEADERS = {"X-User-Name": "SEO_JOEP", "Accept": "application/json"}
    result = []

    try:
        for mc_name_lower, mc_id in _maincat_name_to_id.items():
            mc_name = _maincat_id_to_name.get(mc_id, mc_name_lower)
            subcats = _fetch_subcategories_recursive(TAX_BASE, TAX_HEADERS, int(mc_id))
            for cat_name, cat_id in subcats:
                result.append((mc_name, mc_id, cat_name, str(cat_id)))
    except Exception as e:
        logger.warning(f"Taxonomy API fetch failed: {e}")
        return []

    return result


def _fetch_subcategories_recursive(base_url: str, headers: dict, parent_id: int) -> list:
    """Recursively fetch all subcategories under a parent. Returns [(name, id), ...]."""
    import requests

    result = []
    try:
        r = requests.get(
            f"{base_url}/api/Categories/{parent_id}",
            headers=headers, params={"locale": "nl-NL"}, timeout=30,
        )
        if r.status_code != 200:
            return result

        data = r.json()
        for sub in data.get("subCategories", []):
            if not sub.get("isEnabled", True):
                continue
            nl = next((l for l in sub.get("labels", []) if l.get("locale") == "nl-NL"), {})
            name = nl.get("name", "")
            cat_id = sub.get("id")
            if name and cat_id:
                result.append((name, cat_id))
                # Recurse into this subcategory
                result.extend(
                    _fetch_subcategories_recursive(base_url, headers, cat_id)
                )
    except Exception as e:
        logger.warning(f"Failed to fetch subcategories for {parent_id}: {e}")

    return result


def _populate_cat_ids_from_csv(ws_cat):
    """Fallback: populate cat_ids from static cat_urls.csv."""
    cat_urls_csv = Path(__file__).parent / "data" / "cat_urls.csv"
    if not cat_urls_csv.exists():
        return

    import csv
    with open(cat_urls_csv, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, delimiter=";"):
            mc_name = row.get("maincat", "").strip()
            deepest_cat = row.get("deepest_cat", "").strip()
            cat_id = row.get("cat_id", "").strip()
            if not mc_name or not deepest_cat:
                continue
            mc_id = _maincat_name_to_id.get(mc_name.lower(), "")
            if mc_id:
                ws_cat.append([mc_name, mc_id, deepest_cat, cat_id])


# ---------------------------------------------------------------------------
# Extract results from workbook after processing
# ---------------------------------------------------------------------------
def _extract_results(wb: openpyxl.Workbook, sheet_name: str, result_col: int, error_col: int) -> list:
    """Read back results from the processed workbook."""
    results = []
    if sheet_name not in wb.sheetnames:
        return results
    ws = wb[sheet_name]
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if row_idx == 1:
            continue
        cells = list(row)
        if not cells or not cells[0].value:
            continue
        result_val = cells[result_col].value if result_col < len(cells) else None
        error_val = cells[error_col].value if error_col < len(cells) else None
        row_data = [c.value for c in cells]
        results.append({
            "row": row_idx,
            "data": row_data,
            "success": result_val is True or str(result_val).upper() == "TRUE",
            "error": str(error_val) if error_val else None,
        })
    return results


# ---------------------------------------------------------------------------
# Core operation runners (run in background thread)
# ---------------------------------------------------------------------------
def _run_operation(task_id: str, operation: str, country: str,
                   wb_bytes: Optional[bytes] = None,
                   shop_name: Optional[str] = None,
                   maincat: Optional[str] = None,
                   maincat_id: Optional[str] = None,
                   cl1: Optional[str] = None,
                   budget: Optional[float] = None,
                   campaign_pattern: str = None,
                   dry_run: bool = False,
                   fix: bool = False):
    """Background thread target for all 5 operations. Does ALL heavy init
    (maincat resolve, workbook build, Taxonomy API fetch, Google Ads client)
    inside this thread so the /start HTTP handler returns instantly."""
    try:
        existing_started = _get_task(task_id).get("started_at") if _get_task(task_id) else datetime.now().isoformat()
        _set_task(task_id, {
            "status": "initializing",
            "operation": operation,
            "country": country,
            "progress": 2,
            "message": "Resolving category...",
            "started_at": existing_started,
        })
        _check_cancelled(task_id)

        # Resolve maincat name ↔ id (cheap, reads CSV)
        maincat, maincat_id = resolve_maincat(maincat, maincat_id)

        # Build workbook if shop_name provided. _build_exclusion_workbook can
        # fetch the Taxonomy API (slow on cold cache), which is why this is
        # in the background thread.
        wb: Optional[openpyxl.Workbook] = None
        if shop_name and operation in ("inclusion", "reverse_inclusion"):
            # reverse_inclusion reads the same 'toevoegen' sheet as inclusion
            _set_task(task_id, {**_get_task(task_id), "progress": 5, "message": "Building workbook..."})
            _check_cancelled(task_id)
            wb = _build_inclusion_workbook(shop_name, maincat, maincat_id, cl1 or "a", budget or 50.0)
            wb_bytes = None
        elif shop_name and operation == "exclusion":
            _set_task(task_id, {**_get_task(task_id), "progress": 5, "message": "Loading categories (Taxonomy API)..."})
            _check_cancelled(task_id)
            wb = _build_exclusion_workbook(shop_name, maincat, maincat_id, cl1 or "a")
            # Once the workbook is built the shop_name path no longer needs wb_bytes
            wb_bytes = None
        elif shop_name and operation == "reverse_exclusion":
            # reverse_exclusion needs the 'verwijderen' sheet + cat_ids
            _set_task(task_id, {**_get_task(task_id), "progress": 5, "message": "Loading categories (Taxonomy API)..."})
            _check_cancelled(task_id)
            wb = _build_reverse_exclusion_workbook(shop_name, maincat, maincat_id, cl1 or "a")
            wb_bytes = None

        _check_cancelled(task_id)

        # Patch campaign_processor for country
        _set_task(task_id, {**_get_task(task_id), "progress": 8, "message": "Initializing Google Ads client..."})
        _patch_campaign_processor(country)
        from backend import campaign_processor as cp

        # Initialize client
        client = _get_client()
        customer_id = COUNTRY_CONFIG[country]["customer_id"]

        _check_cancelled(task_id)

        _set_task(task_id, {
            **_get_task(task_id),
            "status": "running",
            "progress": 10,
            "message": f"Running {operation}...",
        })

        result_data = None
        full_log = ""  # complete captured stdout — used for affected-entity parsing

        # ---- INCLUSION ----
        if operation == "inclusion":
            if wb_bytes:
                wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)
            if not wb:
                raise ValueError("No workbook provided for inclusion")

            # Redirect stdout to capture progress
            old_stdout = sys.stdout
            captured = io.StringIO()
            sys.stdout = captured
            try:
                cp.process_inclusion_sheet_v2(client, wb, customer_id, dry_run=dry_run)
            finally:
                sys.stdout = old_stdout

            full_log = captured.getvalue()
            results = _extract_results(wb, "toevoegen", 6, 7)  # col G=result, H=error
            result_data = {
                "rows_processed": len(results),
                "successes": sum(1 for r in results if r["success"]),
                "failures": sum(1 for r in results if not r["success"] and r["error"]),
                "details": results,
                "log": full_log[-5000:],  # last 5k chars for display
            }

        # ---- EXCLUSION ----
        elif operation == "exclusion":
            if wb_bytes:
                wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)
            if not wb:
                raise ValueError("No workbook provided for exclusion")

            old_stdout = sys.stdout
            captured = io.StringIO()
            sys.stdout = captured
            try:
                cp.process_exclusion_sheet_v2(client, wb, customer_id, dry_run=dry_run)
            finally:
                sys.stdout = old_stdout

            full_log = captured.getvalue()
            results = _extract_results(wb, "uitsluiten", 5, 6)  # col F=result, G=error
            result_data = {
                "rows_processed": len(results),
                "successes": sum(1 for r in results if r["success"]),
                "failures": sum(1 for r in results if not r["success"] and r["error"]),
                "details": results,
                "log": full_log[-5000:],
            }

        # ---- REVERSE INCLUSION ----
        elif operation == "reverse_inclusion":
            # Same sheet layout as inclusion ('toevoegen'), removes ad groups
            if wb_bytes:
                wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)
            if not wb:
                raise ValueError("No workbook provided for reverse_inclusion")

            old_stdout = sys.stdout
            captured = io.StringIO()
            sys.stdout = captured
            try:
                cp.process_reverse_inclusion_sheet_v2(client, wb, customer_id, dry_run=dry_run)
            finally:
                sys.stdout = old_stdout

            full_log = captured.getvalue()
            results = _extract_results(wb, "toevoegen", 6, 7)  # col G=result, H=error
            result_data = {
                "rows_processed": len(results),
                "successes": sum(1 for r in results if r["success"]),
                "failures": sum(1 for r in results if not r["success"] and r["error"]),
                "details": results,
                "log": full_log[-5000:],
            }

        # ---- REVERSE EXCLUSION ----
        elif operation == "reverse_exclusion":
            # 'verwijderen' sheet layout matches 'uitsluiten'; removes shop exclusions
            if wb_bytes:
                wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)
            if not wb:
                raise ValueError("No workbook provided for reverse_exclusion")

            old_stdout = sys.stdout
            captured = io.StringIO()
            sys.stdout = captured
            try:
                cp.process_reverse_exclusion_sheet(client, wb, customer_id, dry_run=dry_run)
            finally:
                sys.stdout = old_stdout

            full_log = captured.getvalue()
            results = _extract_results(wb, "verwijderen", 5, 6)  # col F=result, G=error
            result_data = {
                "rows_processed": len(results),
                "successes": sum(1 for r in results if r["success"]),
                "failures": sum(1 for r in results if not r["success"] and r["error"]),
                "details": results,
                "log": full_log[-5000:],
            }

        # ---- VALIDATE CL1 ----
        elif operation == "validate_cl1":
            old_stdout = sys.stdout
            captured = io.StringIO()
            sys.stdout = captured
            try:
                result_data = cp.validate_cl1_targeting_for_campaigns(
                    client, customer_id,
                    campaign_name_pattern=campaign_pattern,
                    dry_run=dry_run,
                )
            finally:
                sys.stdout = old_stdout
            full_log = captured.getvalue()
            if result_data:
                result_data["log"] = full_log[-5000:]

        # ---- VALIDATE ADS ----
        elif operation == "validate_ads":
            old_stdout = sys.stdout
            captured = io.StringIO()
            sys.stdout = captured
            try:
                result_data = cp.validate_ads_for_campaigns(
                    client, customer_id,
                    campaign_name_pattern=campaign_pattern,
                    fix=fix,
                )
            finally:
                sys.stdout = old_stdout
            full_log = captured.getvalue()
            if result_data:
                result_data["log"] = full_log[-5000:]

        # ---- VALIDATE LISTING TREES ----
        elif operation == "validate_trees":
            old_stdout = sys.stdout
            captured = io.StringIO()
            sys.stdout = captured
            try:
                excel_path = None
                if wb_bytes:
                    # Save to temp file for cat_ids reading
                    import tempfile
                    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                    tmp.write(wb_bytes)
                    tmp.close()
                    excel_path = tmp.name

                result_data = cp.validate_listing_trees_for_campaigns(
                    client, customer_id,
                    campaign_name_pattern=campaign_pattern or "PLA/%",
                    dry_run=dry_run,
                    excel_path=excel_path,
                )
            finally:
                sys.stdout = old_stdout
            full_log = captured.getvalue()
            if result_data:
                result_data["log"] = full_log[-5000:]

        else:
            raise ValueError(f"Unknown operation: {operation}")

        # Store result
        _set_task(task_id, {
            "status": "completed",
            "operation": operation,
            "country": country,
            "progress": 100,
            "message": "Completed",
            "started_at": _get_task(task_id).get("started_at"),
            "completed_at": datetime.now().isoformat(),
            "result": result_data,
        })

        # Parse affected entities from the FULL log, not the display-truncated
        # copy — otherwise campaigns mentioned before the last 5k chars get
        # dropped from the export when a run has many groups.
        affected = _parse_affected_entities(full_log)

        # Also store affected in result for the status endpoint
        if result_data:
            result_data["affected"] = affected

        # Add to history
        _history_append({
            "task_id": task_id,
            "operation": operation,
            "country": country,
            "started_at": _get_task(task_id).get("started_at"),
            "completed_at": datetime.now().isoformat(),
            "status": "completed",
            "summary": _summarize_result(operation, result_data),
            "affected": affected,
        })

    except TaskCancelled:
        logger.info(f"DMA+ task {task_id} cancelled by user")
        prev = _get_task(task_id) or {}
        _set_task(task_id, {
            "status": "failed",
            "operation": operation,
            "country": country,
            "progress": prev.get("progress", 0),
            "message": "Cancelled by user",
            "started_at": prev.get("started_at", ""),
            "completed_at": datetime.now().isoformat(),
        })
        _history_append({
            "task_id": task_id,
            "operation": operation,
            "country": country,
            "started_at": prev.get("started_at", ""),
            "completed_at": datetime.now().isoformat(),
            "status": "failed",
            "summary": "Cancelled by user",
        })
    except Exception as e:
        logger.error(f"DMA+ task {task_id} failed: {e}", exc_info=True)
        _set_task(task_id, {
            "status": "failed",
            "operation": operation,
            "country": country,
            "progress": 0,
            "message": str(e),
            "error": traceback.format_exc()[-3000:],
            "started_at": _get_task(task_id).get("started_at", ""),
            "completed_at": datetime.now().isoformat(),
        })
        _history_append({
            "task_id": task_id,
            "operation": operation,
            "country": country,
            "started_at": _get_task(task_id).get("started_at", ""),
            "completed_at": datetime.now().isoformat(),
            "status": "failed",
            "summary": f"Error: {str(e)[:200]}",
        })


def _parse_affected_entities(log: str) -> dict:
    """Parse the captured log output to extract affected campaigns, ad groups,
    trees, and missing campaigns (names the processor looked up that weren't
    in the Google Ads cache).

    Also produces `campaign_ad_group_pairs`: a list of (campaign, ad_group)
    tuples preserving the log's campaign→ad-groups structure so the export
    can show each ad group on the same row as its campaign. For exclusion
    the log is:
        📁 Campaign: PLA/X (N ad group(s))
          ⏭️  PLA/X: all 1 already excluded
    — the leading-whitespace ad-group line immediately follows its campaign.
    """
    import re
    campaigns = set()
    ad_groups = set()
    trees = set()
    missing_campaigns = set()
    campaign_ad_group_pairs: list = []  # [(campaign, ad_group_or_empty), ...]
    current_campaign = None
    current_campaign_had_ag = False

    for line in log.splitlines():
        # Campaigns: "Creating campaign: PLA/..." or "Campaign: PLA/..."
        m = re.search(r'(?:Creating|Created|Found|Processing) campaign[:\s]+([^\n(]+)', line, re.IGNORECASE)
        if m:
            campaigns.add(m.group(1).strip().rstrip('.'))

        # Campaign names from "PLA/..." pattern in context. Note: the original
        # regex stopped at whitespace, which truncated names with a space in
        # them (e.g. "PLA/Klussen store_a" → "PLA/Klussen"). We now only
        # apply this broad fallback when the specific parsers above haven't
        # already claimed the line.
        if not m:
            m = re.search(r'campaign.*?(PLA/[^\s,()]+)', line, re.IGNORECASE)
            if m:
                campaigns.add(m.group(1).strip())

        # Ad groups: "Creating ad group: ..." or "Ad group: ..."
        m = re.search(r'(?:Creating|Created|Processing|Found) ad group[:\s]+([^\n(]+)', line, re.IGNORECASE)
        if m:
            ad_groups.add(m.group(1).strip().rstrip('.'))

        # Ad group names from "PLA/" pattern (ad groups also named PLA/...)
        m = re.search(r'ad.group.*?(PLA/[^\s,()]+)', line, re.IGNORECASE)
        if m:
            ad_groups.add(m.group(1).strip())

        # Campaign header line — three formats, all set the current campaign
        # so subsequent ad-group lines pair with it:
        #   exclusion:          "    📁 Campaign: PLA/X (N ad group(s))"
        #   reverse_inclusion:  "CAMPAIGN 1/3: PLA/Klussen store_a"
        #   inclusion:          "   Creating campaign: PLA/Klussen store_a"
        # Names may contain spaces ("PLA/Klussen store_a"), which is why we
        # use ".+?" up to a distinctive terminator rather than "\S+".
        m = re.search(r'Campaign:\s+(PLA/.+?)\s+\(\d+\s+ad\s+group', line)
        if not m:
            m = re.search(r'^\s*CAMPAIGN\s+\d+/\d+:\s+(PLA/.+?)\s*$', line)
        if not m:
            m = re.search(r'Creating\s+campaign:\s+(PLA/.+?)\s*$', line, re.IGNORECASE)
        if m:
            # Flush previous campaign if it had no ad-group line (handles
            # degraded logs and campaigns that errored before any AG loop).
            if current_campaign is not None and not current_campaign_had_ag:
                campaign_ad_group_pairs.append((current_campaign, ""))
            current_campaign = m.group(1).strip()
            current_campaign_had_ag = False
            campaigns.add(current_campaign)

        # Ad-group header — two formats:
        #   reverse_inclusion: "   ──── Ad Group: PLA/wibra.nl_a ────"
        #   inclusion:         "   ──── Ad Group 1/3: PLA/wibra.nl_a ────"
        # The optional "N/M" must sit between "Ad Group" and the colon.
        m_rev_ag = re.search(r'Ad Group(?:\s+\d+/\d+)?:\s*(PLA/\S+)', line)
        if m_rev_ag:
            ag_name = m_rev_ag.group(1).strip()
            ad_groups.add(ag_name)
            if current_campaign is not None:
                campaign_ad_group_pairs.append((current_campaign, ag_name))
                current_campaign_had_ag = True

        # Exclusion per-ad-group status lines like:
        #   "      ⏭️  PLA/Aggregaten_a: all 1 already excluded"
        #   "      ✅ PLA/wibra.nl_a: 3 added, 0 already excluded"
        #   "      ❌ PLA/wibra.nl_a: 2 error(s)"
        #   "      ⏭️  PLA/Accessoires elektrisch gereedschap_a: all 1 already excluded"
        # The capture group terminates on ':' rather than whitespace so names
        # with spaces (deepest_cats like "Accessoires elektrisch gereedschap",
        # "CO2 Meters", "Afvoerbuizen & hulpstukken") are captured whole.
        m = re.search(r'^\s{4,}\S+\s+(PLA/[^:]+):\s', line)
        if m:
            ag_name = m.group(1).strip()
            ad_groups.add(ag_name)
            if current_campaign is not None:
                campaign_ad_group_pairs.append((current_campaign, ag_name))
                current_campaign_had_ag = True

        # Trees: "Tree created: ..." / "Tree rebuilt: ..." (inclusion)
        #        "Tree to modify: ..." / "Tree modified: ..." (exclusion)
        m = re.search(r'Tree (?:created|rebuilt|to modify|modified)[:\s]+(.+)', line)
        if m:
            trees.add(m.group(1).strip())

        # Exclusions added
        m = re.search(r'Adding \d+ new shop exclusion.*?ad group.*?(PLA/[^\s,()]+)', line, re.IGNORECASE)
        if m:
            ad_groups.add(m.group(1).strip())

        # Missing campaigns emitted by process_exclusion_sheet_v2:
        #   "    ⚠️  Campaign not found in Google Ads cache: PLA/Aggregaten_b"
        # and by the final summary block:
        #   "Missing campaigns in Klussen (3): PLA/Aggregaten_a, PLA/Aggregaten_b, ..."
        m = re.search(r'Campaign not found in Google Ads cache:\s+(PLA/\S+)', line)
        if m:
            missing_campaigns.add(m.group(1).strip())
        m = re.search(r'Missing campaigns in [^:]+\(\d+\):\s+(.+)$', line)
        if m:
            for name in m.group(1).split(","):
                name = name.strip()
                if name.startswith("PLA/"):
                    missing_campaigns.add(name)

    # Flush trailing campaign with no ad-group line
    if current_campaign is not None and not current_campaign_had_ag:
        campaign_ad_group_pairs.append((current_campaign, ""))

    return {
        "campaigns": sorted(campaigns),
        "ad_groups": sorted(ad_groups),
        "trees": sorted(trees),
        "missing_campaigns": sorted(missing_campaigns),
        "campaign_ad_group_pairs": campaign_ad_group_pairs,
    }


def _summarize_result(operation: str, result: dict) -> str:
    if not result:
        return "No results"
    if operation in ("inclusion", "exclusion", "reverse_inclusion", "reverse_exclusion"):
        return f"{result.get('successes', 0)} ok, {result.get('failures', 0)} failed of {result.get('rows_processed', 0)} rows"
    elif operation == "validate_cl1":
        return f"{result.get('ok', 0)} ok, {result.get('fixed', 0)} fixed, {result.get('error', 0)} errors of {result.get('total', 0)}"
    elif operation == "validate_ads":
        return f"{result.get('with_ads', 0)} with ads, {result.get('missing_ads', 0)} missing, {result.get('fixed', 0)} fixed"
    elif operation == "validate_trees":
        return f"{result.get('ok', 0)} ok, {result.get('created', 0)} created, {result.get('error', 0)} errors of {result.get('total', 0)}"
    return str(result)[:200]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def start_operation(operation: str, country: str = "NL",
                    wb_bytes: bytes = None,
                    shop_name: str = None,
                    maincat: str = None,
                    maincat_id: str = None,
                    cl1: str = None,
                    budget: float = None,
                    campaign_pattern: str = None,
                    dry_run: bool = False,
                    fix: bool = False) -> str:
    """
    Seed a task record and kick off the background thread. Returns task_id
    instantly — all heavy work (maincat resolve, workbook build, Taxonomy API
    fetch, Google Ads client init) happens inside _run_operation so the HTTP
    handler is not blocked and the Cancel button can flip the flag early.
    """
    task_id = str(uuid.uuid4())[:8]

    _set_task(task_id, {
        "status": "queued",
        "operation": operation,
        "country": country,
        "progress": 0,
        "message": "Queued...",
        "started_at": datetime.now().isoformat(),
    })

    thread = threading.Thread(
        target=_run_operation,
        args=(task_id, operation, country),
        kwargs={
            "wb_bytes": wb_bytes,
            "shop_name": shop_name,
            "maincat": maincat,
            "maincat_id": maincat_id,
            "cl1": cl1,
            "budget": budget,
            "campaign_pattern": campaign_pattern,
            "dry_run": dry_run,
            "fix": fix,
        },
        daemon=True,
    )
    thread.start()

    return task_id


def get_task_status(task_id: str) -> Optional[dict]:
    return _get_task(task_id)


def cancel_task(task_id: str) -> bool:
    task = _get_task(task_id)
    if task and task.get("status") in ("queued", "initializing", "running"):
        task["cancel"] = True
        _set_task(task_id, task)
        return True
    return False


def get_history() -> list:
    return list(_history)


def clear_history():
    with _history_lock:
        _history.clear()
        _save_history_to_disk()


def remove_history_entry(task_id: str) -> bool:
    """Remove a single history entry by task_id. Returns True if found."""
    with _history_lock:
        original_len = len(_history)
        # Rebuild the deque without the matching entry (deque doesn't support
        # del-by-predicate but the history is tiny so a list comprehension is fine).
        remaining = [e for e in _history if e.get("task_id") != task_id]
        if len(remaining) == original_len:
            return False
        _history.clear()
        _history.extend(remaining)
        _save_history_to_disk()
        return True
